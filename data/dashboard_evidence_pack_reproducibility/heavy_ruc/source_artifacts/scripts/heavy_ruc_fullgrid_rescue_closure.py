# -*- coding: utf-8 -*-
"""
Heavy RUC Full-Grid Rescue Closure Workbench
==================================================

Purpose
-------
Run a very small, decision-grade Heavy RUC reconciliation audit.

The previous bespoke solver run reported a strong Heavy RUC candidate:

    HEAVY_RUC__solver_preq_convex_top18
    quarterly MAPE ~ 3.2766%, annual MAPE ~ 2.4446%

but that was measured on a narrower final-validation grid of 180 quarterly pairs /
15 origins. The later finalist-arbitration run tested a larger full-grid validation
and the same named family no longer achieved the same MAPE.

This script focuses only on Heavy RUC and only on the models needed to answer:

    Can the bespoke Heavy RUC family reproduce its 3.277% / 2.445% result on a
    common full rolling-origin grid?

It is deliberately not a broad search. It preserves the exact old dynamic-no-lead
GBM/Ridge family, pure Schiff benchmarks, and a few arbitration static-solver
components, then rebuilds static and prequential ensembles on the same grid.

Expected use
------------
Run from PowerShell, for example:

    conda activate agts312
    cd "C:\\Users\\Adrian Desilvestro\\Downloads"
    python "heavy_ruc_bespoke_reconciliation_fullgrid.py"

Before running, confirm INPUT_XLSX points to the latest Master Copy workbook.

Outputs
-------
A timestamped run folder is created under OUTPUT_ROOT containing:

    recommended_finalists_primary.csv
    recommended_finalists_by_quarterly.csv
    final_summary.csv
    paired_vs_schiff.csv
    paired_vs_pdf_style.csv
    stress_tests.csv
    ensemble_weights.csv
    all_quarterly_predictions.csv
    all_annual_predictions.csv
    pdf_expected_comparison.csv
    top50_by_stream.csv
    stage1_finalist_arbitration_results.xlsx
    stage1_finalist_arbitration_report.md

Notes
-----
Stage 1 actual-driver testing uses realised future explanatory variables. It is
therefore a model-structure / volume-model test, not yet a full production test
with vintage macro or fuel-price forecast-error propagation.
"""

from __future__ import annotations

import json
import math
import re
import shutil
import time
import warnings
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from scipy.optimize import linprog, minimize
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, BayesianRidge, ElasticNet, HuberRegressor
from sklearn.ensemble import GradientBoostingRegressor, ExtraTreesRegressor, RandomForestRegressor

warnings.filterwarnings("ignore")

try:
    from xgboost import XGBRegressor
    HAS_XGBOOST = True
except Exception:
    HAS_XGBOOST = False

# =============================================================================
# 0. User configuration
# =============================================================================

INPUT_XLSX = Path(r"C:\Users\Adrian Desilvestro\OneDrive\Documents\Playground\Revenue Modeling - Strategic Review\04 Models\Inputs\Master Copy revenue modelling workbook.xlsx")
OUTPUT_ROOT = Path(r"C:\Users\Adrian Desilvestro\OneDrive\Documents\Playground\Revenue Modeling - Strategic Review\04 Models\Inputs\heavy_ruc_fullgrid_rescue_closure_outputs")

RUN_PROFILE = "arbitration"  # keep this script narrow and full-grid by design
RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)

MAX_HORIZON = 12
MIN_TRAIN_QUARTERS = 40

# Full-grid validation for curated candidates. This is much smaller than the broad search,
# so origin_stride=1 is intentional.
ORIGIN_STRIDE = 1
MAX_ORIGINS_PER_STREAM: Optional[int] = None

# Feature/timing assumptions.
ALLOW_LEAD_PRICE_FEATURES_FOR_DISCOVERY = True
INCLUDE_XGBOOST_IF_AVAILABLE = False  # keep off by default; sklearn GBM was sufficient and faster/stabler
TREE_N_JOBS = 1

# Candidate generation controls.
ADD_RANDOM_LOCAL_NEIGHBOURS = False
N_RANDOM_LOCAL_NEIGHBOURS_PER_STREAM = 0

# Ensemble controls.
TOP_MODELS_FOR_ENSEMBLES = 18
FIXED_BLEND_STEPS = np.round(np.linspace(0.0, 1.0, 21), 2)

# Explicit fixed-weight ensemble reconstructions.
# These are used to test why the rescue audit found a very strong Heavy RUC blend on
# the narrower common grid.  We rebuild the old/bespoke and new/reconciliation static
# solver components on the full grid, then blend them across the full grid.
OLD_BESPOKE_STATIC_COMPONENT_WEIGHTS = {
    # Earlier Stage 1/PDF-style static solver family.  This came from the old
    # bespoke solver weights file and is the most plausible component behind the
    # strong rescue blend labelled bespoke_solver::HEAVY_RUC__solver_static_convex_top18.
    "HEAVY_RUC__struct__GBR_learning_rate0_08_max_depth1_n_estimators400__ylag__w40": 0.555798,
    "HEAVY_RUC__rich__Elastic_alpha0_05_l1_ratio0_2__noylag__w52": 0.444202,
}

# Extra zero/near-zero old candidates from the historical solver pool.  These are
# included as base candidates because a full-grid re-optimisation may prefer a
# neighbouring candidate even if its old narrow-grid weight was zero.
OLD_BESPOKE_EXTRA_COMPONENTS = [
    "HEAVY_RUC__struct__GBR_learning_rate0_08_max_depth1_n_estimators250__ylag__w40",
    "HEAVY_RUC__struct__GBR_learning_rate0_08_max_depth2_n_estimators150__ylag__w40",
    "HEAVY_RUC__rich__Elastic_alpha0_05_l1_ratio0_2__noylag__wexp",
    "HEAVY_RUC__rich__Elastic_alpha0_1_l1_ratio0_2__noylag__w64",
    "HEAVY_RUC__rich__Elastic_alpha0_1_l1_ratio0_2__noylag__w52",
    "HEAVY_RUC__diff__Elastic_alpha0_05_l1_ratio0_2__noylag__w40",
]

HEAVY_RECON_STATIC_COMPONENT_WEIGHTS = {
    # Positive-weight components from the Heavy RUC full-grid reconciliation static solver.
    "HEAVY_RUC__dynamic_no_leads__Elastic_alpha0_005_l1_ratio0_2__ylag__w64": 0.469332,
    "HEAVY_RUC__schiff__GBR_learning_rate0_06_max_depth1_n_estimators650__noylag__w64": 0.281844,
    "HEAVY_RUC__dynamic_no_leads__GBR_learning_rate0_08_max_depth1_n_estimators400__ylag__w52": 0.144373,
    "HEAVY_RUC__dynamic_no_leads__GBR_learning_rate0_08_max_depth1_n_estimators150__ylag__w40": 0.104451,
}

# Weight grid for the old-vs-reconciliation rescue closure blend.
RESCUE_OLD_RECON_BLEND_STEPS = np.round(np.linspace(0.0, 1.0, 41), 3)

# Expected PDF-finalist metrics from the earlier scientific report. These are not used
# for model fitting; they are reference values in the output comparison table.
PDF_EXPECTED = {
    "PED": {"quarterly_mape": 2.48, "annual_mape": 2.42, "label": "PDF finalist"},
    "LIGHT_RUC": {"quarterly_mape": 9.16, "annual_mape": 6.25, "label": "PDF finalist"},
    "HEAVY_RUC": {"quarterly_mape": 3.80, "annual_mape": 3.07, "label": "PDF finalist"},
}

LATEST_BESPOKE_EXPECTED = {
    "PED": {"quarterly_mape": 2.65701, "annual_mape": 2.60088, "label": "Latest bespoke finalist"},
    "LIGHT_RUC": {"quarterly_mape": 10.3145, "annual_mape": 6.70964, "label": "Latest bespoke finalist"},
    "HEAVY_RUC": {"quarterly_mape": 3.27657, "annual_mape": 2.44456, "label": "Latest bespoke finalist"},
}

POLICY_WINDOWS = {
    "ruc_discount_active": ("2022Q2", "2023Q2"),
    "ruc_discount_reversal": ("2023Q3", "2023Q4"),
    "post_discount_normalisation": ("2024Q1", "2025Q4"),
}

STREAM_ORDER = ["HEAVY_RUC"]
STREAM_LABELS = {
    "PED": "PED VKT per capita",
    "LIGHT_RUC": "Light RUC volume",
    "HEAVY_RUC": "Heavy RUC volume",
}

# =============================================================================
# 1. Utilities
# =============================================================================

def normalise_name(x: Any) -> str:
    s = str(x).strip().lower().replace("\n", " ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def snake_name(x: Any) -> str:
    s = normalise_name(x)
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


def safe_float_series(s: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")
    return pd.to_numeric(
        s.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip(),
        errors="coerce",
    )


def parse_quarter_value(x: Any) -> Optional[pd.Period]:
    if pd.isna(x):
        return None
    if isinstance(x, pd.Period):
        return pd.Period(str(x), freq="Q-DEC")
    if isinstance(x, pd.Timestamp):
        return x.to_period("Q-DEC")
    s = str(x).strip()
    if not s:
        return None
    m = re.search(r"(19\d{2}|20\d{2})\s*[- ]?\s*[Qq]\s*([1-4])", s)
    if m:
        return pd.Period(f"{int(m.group(1))}Q{int(m.group(2))}", freq="Q-DEC")
    m = re.search(r"[Qq]\s*([1-4])\s*[- ]?\s*(19\d{2}|20\d{2})", s)
    if m:
        return pd.Period(f"{int(m.group(2))}Q{int(m.group(1))}", freq="Q-DEC")
    try:
        if re.fullmatch(r"\d+(\.\d+)?", s):
            val = float(s)
            if 20000 < val < 60000:
                return (pd.Timestamp("1899-12-30") + pd.Timedelta(days=val)).to_period("Q-DEC")
    except Exception:
        pass
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.notna(ts):
            return ts.to_period("Q-DEC")
    except Exception:
        pass
    return None


def period_sort_value(p: pd.Period) -> int:
    return int(p.year * 4 + p.quarter)


def june_year_from_period(p: pd.Period) -> int:
    return int(p.year + 1) if p.quarter in (3, 4) else int(p.year)


def period_between(p: pd.Period, start: str, end: str) -> bool:
    return pd.Period(start, freq="Q-DEC") <= p <= pd.Period(end, freq="Q-DEC")


def mape(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    return float(np.mean(np.abs(p[mask] - a[mask]) / np.abs(a[mask])) * 100.0)


def bias_pct(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    return float(np.mean((p[mask] - a[mask]) / np.abs(a[mask])) * 100.0)


def p90_ape(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    ape = np.abs(p[mask] - a[mask]) / np.abs(a[mask]) * 100.0
    return float(np.nanpercentile(ape, 90))


def rmse(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p)
    if mask.sum() == 0:
        return np.nan
    return float(np.sqrt(np.mean((p[mask] - a[mask]) ** 2)))


def safe_exp(x: float) -> float:
    if not np.isfinite(x):
        return np.nan
    return float(np.exp(np.clip(x, -50, 50)))


def dedupe(seq: Iterable[str]) -> List[str]:
    out, seen = [], set()
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def boolish(x: Any) -> bool:
    if isinstance(x, bool):
        return x
    if pd.isna(x):
        return False
    return str(x).strip().lower() in {"true", "1", "yes", "y"}


def ensure_run_dir(root: Path) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    run_dir = root / f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    run_dir.mkdir(parents=True, exist_ok=False)
    (run_dir / "figures").mkdir(exist_ok=True)
    return run_dir

# =============================================================================
# 2. Workbook loading and schema detection
# =============================================================================

TARGET_PATTERNS = {
    "PED": [
        r"light.*petrol.*vkt.*cap",
        r"petrol.*vkt.*cap",
        r"vkt.*per.*cap",
        r"vkt.*pc",
    ],
    "LIGHT_RUC": [
        r"light.*ruc.*volume",
        r"light.*ruc.*net.*km",
        r"light.*ruc.*kilomet",
        r"light.*iuc.*volume",
        r"light.*iuc.*net.*km",
        r"light.*net.*km",
    ],
    "HEAVY_RUC": [
        r"heavy.*ruc.*volume",
        r"heavy.*ruc.*net.*km",
        r"heavy.*ruc.*kilomet",
        r"heavy.*iuc.*volume",
        r"heavy.*iuc.*net.*km",
        r"heavy.*net.*km",
    ],
}

FEATURE_BLACKLIST = [
    "household", "hes", "hils", "survey", "hies", "disposable", "rent", "mortgage",
    "housing cost", "income decile", "brent", "wti", "crude", "barrel", "importer",
    "margin", "spread", "gst", "excise", "levy", "carbon", "ets", "nzu", "emissions",
    "shipping", "refining", "refinery", "forecast", "projection", "error", "residual",
    "mape", "ape", "p50", "p80", "p90", "confidence", "upper", "lower", "oracle", "freeze",
]

FEATURE_ALLOW_TERMS = {
    "PED": ["petrol", "gdp", "gross domestic", "unemployment", "unemploy", "cpi", "population", "per capita", "gdppc", "gdp pc"],
    "LIGHT_RUC": ["diesel", "light ruc", "light iuc", "ruc price", "iuc price", "gdp", "gross domestic", "unemployment", "unemploy", "cpi"],
    "HEAVY_RUC": ["diesel", "heavy ruc", "heavy iuc", "ruc price", "iuc price", "gdp", "gross domestic", "unemployment", "unemploy", "cpi"],
}


def is_log_col(col: str) -> bool:
    n = normalise_name(col)
    return n.startswith("ln ") or " ln " in f" {n} " or "log" in n or "natural log" in n


def load_input_sheet(xlsx_path: Path) -> pd.DataFrame:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")
    xl = pd.ExcelFile(xlsx_path)
    preferred = ["Stage 1 Inputs", "Stage1 Inputs", "Stage 1", "Model Inputs", "Inputs", "Data"]
    norm_to_actual = {normalise_name(s): s for s in xl.sheet_names}
    sheet = None
    for p in preferred:
        if normalise_name(p) in norm_to_actual:
            sheet = norm_to_actual[normalise_name(p)]
            break
    if sheet is None:
        best_score, best_sheet = -1, None
        for s in xl.sheet_names:
            try:
                tmp = pd.read_excel(xlsx_path, sheet_name=s, nrows=8)
            except Exception:
                continue
            joined = " ".join(normalise_name(c) for c in tmp.columns)
            score = sum(kw in joined for kw in ["quarter", "vkt", "ruc", "iuc", "petrol", "diesel", "gdp", "unemployment", "cpi"])
            if score > best_score:
                best_score, best_sheet = score, s
        sheet = best_sheet
    if sheet is None:
        raise RuntimeError("Could not detect a model input sheet.")

    df = pd.read_excel(xlsx_path, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

    norm_cols = {c: normalise_name(c) for c in df.columns}
    quarter_col = None
    for c, n in norm_cols.items():
        if "quarter" in n and df[c].notna().sum() >= 10:
            quarter_col = c
            break
    if quarter_col is None:
        for c, n in norm_cols.items():
            if ("date" in n or "period" in n) and df[c].notna().sum() >= 10:
                quarter_col = c
                break
    if quarter_col is None:
        year_cols = [c for c, n in norm_cols.items() if n in {"year", "financial year", "fiscal year"}]
        q_cols = [c for c, n in norm_cols.items() if n in {"quarter", "qtr", "q"}]
        if not year_cols or not q_cols:
            raise RuntimeError("Could not detect quarter/date column.")
        periods = (df[year_cols[0]].astype(str) + "Q" + df[q_cols[0]].astype(str).str.extract(r"([1-4])", expand=False).fillna(""))
        periods = periods.map(parse_quarter_value)
    else:
        periods = df[quarter_col].map(parse_quarter_value)

    df = df.copy()
    df["__period__"] = periods
    df = df[df["__period__"].notna()].copy()
    df["__sort__"] = df["__period__"].map(period_sort_value)
    df = df.sort_values("__sort__").drop_duplicates("__period__", keep="last")
    df = df.drop(columns=["__sort__"]).reset_index(drop=True)
    df.attrs["input_sheet"] = sheet
    return df


def detect_target_col(df: pd.DataFrame, stream: str) -> Tuple[str, bool]:
    patterns = TARGET_PATTERNS[stream]
    candidates = []
    for c in df.columns:
        if c == "__period__":
            continue
        n = normalise_name(c)
        if "revenue" in n:
            continue
        if any(re.search(p, n) for p in patterns):
            s = safe_float_series(df[c])
            non_na = s.notna().sum()
            if non_na >= 20:
                score = 100 + min(non_na, 200) / 10 - (5 if is_log_col(c) else 0)
                candidates.append((score, c, is_log_col(c), non_na))
    if not candidates:
        raise RuntimeError(f"Could not detect target column for {stream}.")
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1], candidates[0][2]


def detect_feature_cols(df: pd.DataFrame, stream: str, all_targets: Sequence[str]) -> List[str]:
    allow = FEATURE_ALLOW_TERMS[stream]
    cols = []
    for c in df.columns:
        if c == "__period__" or c in all_targets:
            continue
        n = normalise_name(c)
        if any(b in n for b in FEATURE_BLACKLIST):
            continue
        if "revenue" in n:
            continue
        # Avoid pulling in other target-like series unless explicitly a price/rate.
        if any(t in n for t in ["volume", "vkt", "net km", "net kilomet", "kilometres", "kilometers", "kms"]):
            if "price" not in n and "rate" not in n:
                continue
        if not any(term in n for term in allow):
            continue
        s = safe_float_series(df[c])
        if s.notna().sum() >= 20:
            cols.append(c)
    if not cols:
        raise RuntimeError(f"No feature columns detected for {stream}.")
    return cols

# =============================================================================
# 3. Feature engineering
# =============================================================================

@dataclass
class StreamData:
    stream: str
    target_col: str
    target_is_log: bool
    feature_cols: List[str]
    y_raw: pd.Series
    y_log: pd.Series
    exog: pd.DataFrame
    groups: Dict[str, List[str]]
    primary_log: Dict[str, str]


def assign_base_group(stream: str, col: str) -> str:
    n = normalise_name(col)
    if "petrol" in n and ("price" in n or "cents" in n or "cost" in n):
        return "petrol_price"
    if "diesel" in n and ("price" in n or "cents" in n or "cost" in n):
        return "diesel_price"
    if "light" in n and ("ruc" in n or "iuc" in n) and ("price" in n or "rate" in n or "cost" in n):
        return "light_ruc_price"
    if "heavy" in n and ("ruc" in n or "iuc" in n) and ("price" in n or "rate" in n or "cost" in n):
        return "heavy_ruc_price"
    if ("ruc price" in n or "iuc price" in n) and stream == "LIGHT_RUC":
        return "light_ruc_price"
    if ("ruc price" in n or "iuc price" in n) and stream == "HEAVY_RUC":
        return "heavy_ruc_price"
    if "gdp" in n or "gross domestic" in n:
        if "per capita" in n or "pc" in n:
            return "gdp_pc"
        return "gdp"
    if "unemployment" in n or "unemploy" in n:
        return "unemployment"
    if "cpi" in n:
        return "cpi"
    if "population" in n:
        return "population"
    return "other_core"


def add_group(groups: Dict[str, List[str]], group: str, name: str) -> None:
    groups.setdefault(group, [])
    if name not in groups[group]:
        groups[group].append(name)


def build_target_series(df: pd.DataFrame, target_col: str, target_is_log: bool) -> Tuple[pd.Series, pd.Series]:
    s = safe_float_series(df[target_col])
    s.index = df["__period__"]
    if target_is_log:
        y_log = s.copy()
        y_raw = np.exp(y_log)
    else:
        y_raw = s.copy()
        y_log = np.log(y_raw.where(y_raw > 0))
    return y_raw, y_log


def build_exog(df: pd.DataFrame, stream: str, feature_cols: Sequence[str]) -> Tuple[pd.DataFrame, Dict[str, List[str]], Dict[str, str]]:
    idx = df["__period__"]
    exog = pd.DataFrame(index=idx)
    groups: Dict[str, List[str]] = {"time": [], "policy_window": [], "policy_price_signal": [], "policy_interaction": []}
    primary_log: Dict[str, str] = {}

    for col in feature_cols:
        s = safe_float_series(df[col]).copy()
        s.index = idx
        base = snake_name(col)
        group = assign_base_group(stream, col)
        positive_share = (s > 0).mean(skipna=True)
        col_is_log = is_log_col(col)

        level_name = f"{base}__level"
        exog[level_name] = s
        add_group(groups, "core_level", level_name)
        add_group(groups, f"{group}_level", level_name)

        log_s = None
        if col_is_log:
            log_s = s.copy()
            log_name = f"{base}__log"
            exog[log_name] = log_s
            add_group(groups, "core_level", log_name)
            add_group(groups, f"{group}_level", log_name)
            primary_log.setdefault(group, log_name)
        elif positive_share > 0.90:
            log_s = np.log(s.where(s > 0))
            log_name = f"{base}__log"
            exog[log_name] = log_s
            add_group(groups, "core_level", log_name)
            add_group(groups, f"{group}_level", log_name)
            primary_log.setdefault(group, log_name)

        # Lags/diffs for all core variables.
        for lag in [1, 2, 4, 8]:
            nm = f"{base}__lag{lag}"
            exog[nm] = s.shift(lag)
            add_group(groups, "core_lag", nm)
            add_group(groups, f"{group}_lag", nm)
            if log_s is not None:
                nml = f"{base}__log_lag{lag}"
                exog[nml] = log_s.shift(lag)
                add_group(groups, "core_lag", nml)
                add_group(groups, f"{group}_lag", nml)
        for d in [1, 4]:
            nm = f"{base}__diff{d}"
            exog[nm] = s.diff(d)
            add_group(groups, "core_diff", nm)
            add_group(groups, f"{group}_diff", nm)
            if log_s is not None:
                nml = f"{base}__log_diff{d}"
                exog[nml] = log_s.diff(d)
                add_group(groups, "core_diff", nml)
                add_group(groups, f"{group}_diff", nml)

        # Lead price features are kept only for feature sets that explicitly allow leads.
        if ALLOW_LEAD_PRICE_FEATURES_FOR_DISCOVERY and group in {"light_ruc_price", "heavy_ruc_price", "diesel_price", "petrol_price"}:
            for lead in [1, 2, 4]:
                nm = f"{base}__lead{lead}"
                exog[nm] = s.shift(-lead)
                add_group(groups, "core_lead", nm)
                add_group(groups, f"{group}_lead", nm)
                if log_s is not None:
                    nml = f"{base}__log_lead{lead}"
                    exog[nml] = log_s.shift(-lead)
                    add_group(groups, "core_lead", nml)
                    add_group(groups, f"{group}_lead", nml)
                    chg = f"{base}__log_lead{lead}_minus_current"
                    exog[chg] = log_s.shift(-lead) - log_s
                    add_group(groups, "core_lead", chg)
                    add_group(groups, "policy_price_signal", chg)
                    add_group(groups, f"{group}_lead", chg)

    periods = pd.Series(list(exog.index), index=exog.index)
    q = periods.map(lambda p: int(p.quarter))
    exog["time__trend"] = np.arange(len(exog), dtype=float) + 1.0
    exog["time__trend_sq"] = exog["time__trend"] ** 2
    exog["time__post2011"] = periods.map(lambda p: 1.0 if p >= pd.Period("2011Q1", freq="Q-DEC") else 0.0)
    exog["time__post2011_trend"] = exog["time__post2011"] * exog["time__trend"]
    exog["time__post2020"] = periods.map(lambda p: 1.0 if p >= pd.Period("2020Q1", freq="Q-DEC") else 0.0)
    exog["time__covid2020"] = periods.map(lambda p: 1.0 if pd.Period("2020Q1", freq="Q-DEC") <= p <= pd.Period("2020Q4", freq="Q-DEC") else 0.0)
    for qq in [1, 2, 3]:
        nm = f"time__q{qq}"
        exog[nm] = (q == qq).astype(float)
        add_group(groups, "time", nm)
    for nm in ["time__trend", "time__trend_sq", "time__post2011", "time__post2011_trend", "time__post2020", "time__covid2020"]:
        add_group(groups, "time", nm)

    # RUC-specific policy windows and price-shock terms.
    if stream in {"LIGHT_RUC", "HEAVY_RUC"}:
        for key, (start, end) in POLICY_WINDOWS.items():
            nm = f"policy__{key}"
            exog[nm] = periods.map(lambda p, s=start, e=end: 1.0 if period_between(p, s, e) else 0.0)
            add_group(groups, "policy_window", nm)

        ruc_key = "light_ruc_price" if stream == "LIGHT_RUC" else "heavy_ruc_price"
        ruc_log = primary_log.get(ruc_key)
        diesel_log = primary_log.get("diesel_price")
        for label, src in [("ruc", ruc_log), ("diesel", diesel_log)]:
            if src is not None and src in exog.columns:
                z = exog[src]
                exog[f"policy__{label}_log_change_1"] = z.diff(1)
                exog[f"policy__{label}_log_change_4"] = z.diff(4)
                exog[f"policy__{label}_price_jump_up_1"] = z.diff(1).clip(lower=0)
                exog[f"policy__{label}_price_cut_1"] = (-z.diff(1)).clip(lower=0)
                exog[f"policy__abs_{label}_log_change_1"] = z.diff(1).abs()
                for nm in [f"policy__{label}_log_change_1", f"policy__{label}_log_change_4", f"policy__{label}_price_jump_up_1", f"policy__{label}_price_cut_1", f"policy__abs_{label}_log_change_1"]:
                    add_group(groups, "policy_price_signal", nm)
                for lag in [1, 2, 4]:
                    for src2 in [f"policy__{label}_price_jump_up_1", f"policy__{label}_price_cut_1", f"policy__abs_{label}_log_change_1"]:
                        nm = f"{src2}_lag{lag}"
                        exog[nm] = exog[src2].shift(lag)
                        add_group(groups, "policy_price_signal", nm)

        for policy_col in groups.get("policy_window", []):
            for price_col in groups.get("policy_price_signal", [])[:10]:
                nm = f"interact__{policy_col.replace('policy__','')}__{price_col.replace('policy__','')}"
                exog[nm] = exog[policy_col] * exog[price_col]
                add_group(groups, "policy_interaction", nm)

    return exog.sort_index(), groups, primary_log


def pick_group(groups: Dict[str, List[str]], group: str, prefer_log: bool = True, max_n: int = 2) -> List[str]:
    candidates = groups.get(group, [])
    if not candidates:
        return []
    if prefer_log:
        log_cands = [c for c in candidates if "__log" in c and "__diff" not in c and "__lag" not in c and "__lead" not in c]
        if log_cands:
            return log_cands[:max_n]
    clean = [c for c in candidates if "__diff" not in c and "__lag" not in c and "__lead" not in c]
    return clean[:max_n]


def feature_names_for_set(sd: StreamData, feature_set: str, include_target_lags: bool) -> List[str]:
    g = sd.groups
    stream = sd.stream

    time_base = [c for c in g.get("time", []) if c in sd.exog.columns]
    quarter_time = [c for c in time_base if c.startswith("time__q")]

    schiff: List[str] = []
    if stream == "PED":
        schiff += pick_group(g, "petrol_price_level", max_n=1)
        schiff += pick_group(g, "gdp_pc_level", max_n=1) or pick_group(g, "gdp_level", max_n=1)
        schiff += pick_group(g, "unemployment_level", prefer_log=False, max_n=1)
        schiff += ["time__trend", "time__post2011_trend", "time__post2020", "time__covid2020"] + quarter_time
    elif stream == "LIGHT_RUC":
        schiff += pick_group(g, "diesel_price_level", max_n=1)
        schiff += pick_group(g, "light_ruc_price_level", max_n=1)
        # lagged light RUC price is part of the structural RUC intuition.
        schiff += [c for c in g.get("light_ruc_price_lag", []) if "lag1" in c and "__log" in c][:1]
        schiff += pick_group(g, "gdp_level", max_n=1)
        schiff += ["time__post2020"] + quarter_time
    elif stream == "HEAVY_RUC":
        schiff += pick_group(g, "gdp_level", max_n=1)
        schiff += pick_group(g, "heavy_ruc_price_level", max_n=1)
        if feature_set == "schiff" and ALLOW_LEAD_PRICE_FEATURES_FOR_DISCOVERY:
            schiff += [c for c in g.get("heavy_ruc_price_lead", []) if "lead1" in c and "__log" in c][:1]
        schiff += quarter_time

    schiff = [c for c in dedupe(schiff) if c in sd.exog.columns]

    if feature_set in {"schiff", "schiff_no_lead"}:
        cols = schiff
        if feature_set == "schiff_no_lead":
            cols = [c for c in cols if "lead" not in c]
    elif feature_set == "struct":
        cols = [c for c in g.get("core_level", []) if "__diff" not in c and "__lag" not in c and "__lead" not in c] + time_base
    elif feature_set == "diff":
        cols = g.get("core_level", []) + g.get("core_diff", []) + time_base
        cols = [c for c in cols if "__lead" not in c]
    elif feature_set == "rich":
        cols = g.get("core_level", []) + g.get("core_diff", []) + g.get("core_lag", []) + time_base
        cols = [c for c in cols if "__lead" not in c]
    elif feature_set == "dynamic_pruned":
        cols = schiff + g.get("core_diff", []) + g.get("core_lag", []) + g.get("policy_window", []) + g.get("policy_price_signal", [])
        cols = [c for c in cols if "lead" not in c]
    elif feature_set == "dynamic_no_leads":
        cols = g.get("core_level", []) + g.get("core_diff", []) + g.get("core_lag", []) + time_base + g.get("policy_window", []) + g.get("policy_price_signal", []) + g.get("policy_interaction", [])
        cols = [c for c in cols if "lead" not in c]
    elif feature_set == "dynamic_rich":
        cols = g.get("core_level", []) + g.get("core_diff", []) + g.get("core_lag", []) + g.get("core_lead", []) + time_base + g.get("policy_window", []) + g.get("policy_price_signal", []) + g.get("policy_interaction", [])
    elif feature_set == "price_distributed_lags":
        price_groups = []
        for key in ["petrol_price", "diesel_price", "light_ruc_price", "heavy_ruc_price"]:
            price_groups += g.get(f"{key}_level", []) + g.get(f"{key}_diff", []) + g.get(f"{key}_lag", []) + g.get(f"{key}_lead", [])
        cols = price_groups + g.get("gdp_level", []) + g.get("gdp_lag", []) + g.get("gdp_pc_level", []) + g.get("unemployment_level", []) + time_base + g.get("policy_window", []) + g.get("policy_price_signal", [])
        if "no_leads" in feature_set:
            cols = [c for c in cols if "lead" not in c]
    else:
        raise ValueError(f"Unknown feature_set: {feature_set}")

    cols = [c for c in dedupe(cols) if c in sd.exog.columns]
    if include_target_lags:
        cols += ["target__lag1", "target__lag2", "target__lag4", "target__diff1", "target__diff4", "target__roll4_mean", "target__roll8_mean"]
    return dedupe(cols)


def target_lag_features(period: pd.Period, y_hist: Dict[pd.Period, float]) -> Dict[str, float]:
    def get(p: pd.Period) -> float:
        v = y_hist.get(p, np.nan)
        return float(v) if np.isfinite(v) else np.nan
    y1, y2, y4, y5 = get(period - 1), get(period - 2), get(period - 4), get(period - 5)
    last4 = np.array([get(period - i) for i in range(1, 5)], dtype=float)
    last8 = np.array([get(period - i) for i in range(1, 9)], dtype=float)
    return {
        "target__lag1": y1,
        "target__lag2": y2,
        "target__lag4": y4,
        "target__diff1": y1 - y2 if np.isfinite(y1) and np.isfinite(y2) else np.nan,
        "target__diff4": y1 - y5 if np.isfinite(y1) and np.isfinite(y5) else np.nan,
        "target__roll4_mean": float(np.nanmean(last4)) if np.isfinite(last4).sum() >= 2 else np.nan,
        "target__roll8_mean": float(np.nanmean(last8)) if np.isfinite(last8).sum() >= 4 else np.nan,
    }


def build_feature_row(period: pd.Period, sd: StreamData, y_hist: Dict[pd.Period, float], feature_names: Sequence[str], include_target_lags: bool) -> Dict[str, float]:
    row: Dict[str, float] = {}
    if period in sd.exog.index:
        row.update(sd.exog.loc[period].to_dict())
    if include_target_lags:
        row.update(target_lag_features(period, y_hist))
    return {f: row.get(f, np.nan) for f in feature_names}

# =============================================================================
# 4. Candidate definitions
# =============================================================================

@dataclass(frozen=True)
class CandidateConfig:
    stream: str
    name: str
    model_kind: str
    params_json: str
    window: Optional[int]
    feature_set: str
    include_target_lags: bool
    family_tag: str
    min_train_quarters: int = MIN_TRAIN_QUARTERS


LINEAR_KINDS = {"ols", "ridge", "bayesian_ridge", "elastic_net", "huber"}
RESIDUAL_KINDS = {"resid_ridge", "resid_bayesian_ridge", "resid_elastic_net", "resid_gbr", "resid_huber"}


def make_estimator(kind: str, params: Dict[str, Any]):
    if kind == "ols":
        return LinearRegression(**params)
    if kind == "ridge":
        return Ridge(**params)
    if kind == "bayesian_ridge":
        return BayesianRidge(**params)
    if kind == "elastic_net":
        return ElasticNet(**params)
    if kind == "huber":
        return HuberRegressor(**params)
    if kind == "gbr":
        return GradientBoostingRegressor(**params)
    if kind == "extra_trees":
        return ExtraTreesRegressor(**params)
    if kind == "random_forest":
        return RandomForestRegressor(**params)
    if kind == "xgboost":
        return XGBRegressor(**params)
    raise ValueError(kind)


def base_kind_for_resid(kind: str) -> str:
    return {
        "resid_ridge": "ridge",
        "resid_bayesian_ridge": "bayesian_ridge",
        "resid_elastic_net": "elastic_net",
        "resid_gbr": "gbr",
        "resid_huber": "huber",
    }[kind]


def make_pipeline(kind: str, params: Dict[str, Any]) -> Pipeline:
    est = make_estimator(kind, params)
    if kind in LINEAR_KINDS:
        return Pipeline([("imputer", SimpleImputer(strategy="median")), ("scaler", StandardScaler()), ("model", est)])
    return Pipeline([("imputer", SimpleImputer(strategy="median")), ("model", est)])


def clean_label_params(params: Dict[str, Any]) -> str:
    if not params:
        return ""
    parts = []
    for k in sorted(params):
        v = params[k]
        if k == "random_state" or k == "n_jobs" or k == "loss" or k == "max_iter" or k == "subsample":
            continue
        if isinstance(v, float):
            s = f"{v:g}".replace(".", "_")
        else:
            s = str(v).replace(".", "_")
        parts.append(f"{k}{s}")
    return "_" + "_".join(parts) if parts else ""


def add_candidate(configs: List[CandidateConfig], stream: str, kind: str, params: Dict[str, Any], window: Optional[int], feature_set: str, ylags: bool, tag: str) -> None:
    w = "exp" if window is None else str(window)
    params_label = clean_label_params(params)
    name = f"{stream}__{feature_set}__{tag}{params_label}__{'ylag' if ylags else 'noylag'}__w{w}"
    configs.append(CandidateConfig(
        stream=stream,
        name=name,
        model_kind=kind,
        params_json=json.dumps(params, sort_keys=True),
        window=window,
        feature_set=feature_set,
        include_target_lags=ylags,
        family_tag=tag,
    ))


def generate_targeted_candidates(stream: str) -> List[CandidateConfig]:
    """Generate a small Heavy RUC-only candidate set.

    This deliberately targets the family that produced the earlier 3.2766% / 2.4446%
    bespoke result, plus pure Schiff benchmarks and the main arbitration static-solver
    neighbour family. It should be fast relative to the broad all-stream solve.
    """
    if stream != "HEAVY_RUC":
        return []

    configs: List[CandidateConfig] = []

    # 0. Exact historical old-bespoke static-solver component family.
    # These are the components that were not fully represented in the later heavy reconciliation.
    # They allow us to rebuild the old static solver on the broad grid and test the rescue blend.
    add_candidate(configs, stream, "gbr", {
        "n_estimators": 400,
        "learning_rate": 0.08,
        "max_depth": 1,
        "subsample": 0.85,
        "loss": "squared_error",
        "random_state": RANDOM_SEED,
    }, 40, "struct", True, "GBR")
    add_candidate(configs, stream, "gbr", {
        "n_estimators": 250,
        "learning_rate": 0.08,
        "max_depth": 1,
        "subsample": 0.85,
        "loss": "squared_error",
        "random_state": RANDOM_SEED,
    }, 40, "struct", True, "GBR")
    add_candidate(configs, stream, "gbr", {
        "n_estimators": 150,
        "learning_rate": 0.08,
        "max_depth": 2,
        "subsample": 0.85,
        "loss": "squared_error",
        "random_state": RANDOM_SEED,
    }, 40, "struct", True, "GBR")

    for alpha, l1, window, ylags in [
        (0.05, 0.2, 52, False),
        (0.05, 0.2, None, False),
        (0.10, 0.2, 64, False),
        (0.10, 0.2, 52, False),
        (0.10, 0.2, 64, True),
        (0.10, 0.2, None, True),
    ]:
        add_candidate(configs, stream, "elastic_net", {
            "alpha": alpha,
            "l1_ratio": l1,
            "max_iter": 50000,
            "random_state": RANDOM_SEED,
        }, window, "rich", ylags, "Elastic")

    add_candidate(configs, stream, "elastic_net", {
        "alpha": 0.05,
        "l1_ratio": 0.2,
        "max_iter": 50000,
        "random_state": RANDOM_SEED,
    }, 40, "diff", False, "Elastic")

    # 1. Pure Schiff benchmarks.
    for window in [None, 40, 52, 64]:
        add_candidate(configs, stream, "ols", {}, window, "schiff", False, "SCHIFF_OLS")
        add_candidate(configs, stream, "ols", {}, window, "schiff_no_lead", False, "SCHIFF_OLS")

    # 2. Exact old bespoke prequential base family.
    # Old reported ensemble components:
    #   dynamic_no_leads GBM: n = 150/250/400, depth=1, lr=0.03/0.06/0.08, mostly no target lags, W40.
    #   dynamic_no_leads Ridge alpha=100 with several windows and target-lag options.
    #   schiff_no_lead GBM variants with target lags, W52/W64/exp.

    # Dynamic/no-leads shallow GBMs, old-style core.
    for n in [150, 250, 400]:
        for lr in [0.03, 0.06, 0.08]:
            params = {
                "n_estimators": n,
                "learning_rate": lr,
                "max_depth": 1,
                "subsample": 0.85,
                "loss": "squared_error",
                "random_state": RANDOM_SEED,
            }
            add_candidate(configs, stream, "gbr", params, 40, "dynamic_no_leads", False, "GBR")
            add_candidate(configs, stream, "gbr", params, 40, "dynamic_no_leads", True, "GBR")

    # A small adjacent dynamic/no-lead set to test robustness around the old winning family.
    for n in [400, 650]:
        for lr in [0.05, 0.06, 0.08]:
            for window in [40, 52, 64]:
                params = {
                    "n_estimators": n,
                    "learning_rate": lr,
                    "max_depth": 1,
                    "subsample": 0.85,
                    "loss": "squared_error",
                    "random_state": RANDOM_SEED,
                }
                add_candidate(configs, stream, "gbr", params, window, "dynamic_no_leads", False, "GBR")
                add_candidate(configs, stream, "gbr", params, window, "dynamic_no_leads", True, "GBR")

    # Ridge alpha=100 family used by the old prequential pool.
    for window in [40, 52, 64, None]:
        for ylags in [False, True]:
            add_candidate(configs, stream, "ridge", {"alpha": 100.0, "random_state": RANDOM_SEED}, window, "dynamic_no_leads", ylags, "Ridge")

    # Schiff-no-lead GBMs from old preq pool.
    for n in [250, 400]:
        for lr in [0.03, 0.06]:
            for window in [52, 64, None]:
                params = {
                    "n_estimators": n,
                    "learning_rate": lr,
                    "max_depth": 1,
                    "subsample": 0.85,
                    "loss": "squared_error",
                    "random_state": RANDOM_SEED,
                }
                add_candidate(configs, stream, "gbr", params, window, "schiff_no_lead", True, "GBR")

    # 3. Arbitration static-solver positive component family, included so the audit can compare
    # old bespoke-family candidates against the newer arbitration static-solver candidate.
    for fs, ylags, window, alpha, l1 in [
        ("dynamic_no_leads", True, 64, 0.005, 0.2),
        ("dynamic_no_leads", False, 64, 0.01, 0.2),
        ("dynamic_pruned", True, 64, 0.05, 0.2),
        ("dynamic_pruned", True, 64, 0.01, 0.8),
        ("rich", True, 64, 0.1, 0.2),
    ]:
        add_candidate(configs, stream, "elastic_net", {
            "alpha": alpha,
            "l1_ratio": l1,
            "max_iter": 50000,
            "random_state": RANDOM_SEED,
        }, window, fs, ylags, "Elastic")

    # Arbitration's strong Schiff GBM neighbours.
    for fs, ylags, window, n, lr in [
        ("schiff", False, 64, 650, 0.06),
        ("schiff", False, 64, 650, 0.05),
        ("schiff", False, 52, 650, 0.06),
        ("schiff", False, 52, 650, 0.05),
        ("schiff_no_lead", True, 64, 180, 0.045),
        ("schiff_no_lead", True, 40, 450, 0.055),
    ]:
        params = {
            "n_estimators": n,
            "learning_rate": lr,
            "max_depth": 1,
            "subsample": 0.85,
            "loss": "squared_error",
            "random_state": RANDOM_SEED,
        }
        add_candidate(configs, stream, "gbr", params, window, fs, ylags, "GBR")

    # 4. Residual-correction sanity check: small targeted GBM residual corrections only.
    for fs in ["dynamic_no_leads", "schiff_no_lead"]:
        for window in [40, 52, 64]:
            for n in [80, 150, 250]:
                params = {
                    "n_estimators": n,
                    "learning_rate": 0.05,
                    "max_depth": 1,
                    "subsample": 0.85,
                    "loss": "squared_error",
                    "random_state": RANDOM_SEED,
                }
                add_candidate(configs, stream, "resid_gbr", params, window, fs, False, "RESID_SchiffResidual_GBR")

    unique: Dict[str, CandidateConfig] = {}
    for c in configs:
        unique[c.name] = c
    return list(unique.values())

# =============================================================================
# 5. Rolling-origin evaluation
# =============================================================================

def valid_periods(sd: StreamData) -> List[pd.Period]:
    idx = set(sd.y_raw.dropna().index).intersection(set(sd.exog.index))
    return sorted(idx, key=period_sort_value)


def build_training_matrix(sd: StreamData, train_periods: Sequence[pd.Period], feature_names: Sequence[str], include_target_lags: bool) -> Tuple[pd.DataFrame, pd.Series]:
    y_hist = {p: float(sd.y_log.loc[p]) for p in sd.y_log.index if pd.notna(sd.y_log.loc[p])}
    rows, ys, used = [], [], []
    for p in train_periods:
        if p not in sd.y_log.index or pd.isna(sd.y_log.loc[p]):
            continue
        rows.append(build_feature_row(p, sd, y_hist, feature_names, include_target_lags))
        ys.append(float(sd.y_log.loc[p]))
        used.append(p)
    X = pd.DataFrame(rows, index=used).reindex(columns=feature_names)
    y = pd.Series(ys, index=used, name="y_log")
    return X, y


def fit_model(cfg: CandidateConfig, X: pd.DataFrame, y: pd.Series, X_base: Optional[pd.DataFrame] = None):
    params = json.loads(cfg.params_json)
    if cfg.model_kind in RESIDUAL_KINDS:
        if X_base is None:
            raise ValueError("Residual model requires X_base")
        base = make_pipeline("ols", {})
        base.fit(X_base, y)
        resid = y.values - base.predict(X_base)
        rk = base_kind_for_resid(cfg.model_kind)
        rm = make_pipeline(rk, params)
        rm.fit(X, resid)
        return {"kind": "residual", "base": base, "resid": rm}
    pipe = make_pipeline(cfg.model_kind, params)
    pipe.fit(X, y)
    return pipe


def predict_model(model: Any, X_pred: pd.DataFrame, X_pred_base: Optional[pd.DataFrame] = None) -> float:
    if isinstance(model, dict) and model.get("kind") == "residual":
        if X_pred_base is None:
            raise ValueError("Residual model requires X_pred_base")
        return float(model["base"].predict(X_pred_base)[0] + model["resid"].predict(X_pred)[0])
    return float(model.predict(X_pred)[0])


def evaluate_candidate(sd: StreamData, cfg: CandidateConfig) -> pd.DataFrame:
    feature_names = feature_names_for_set(sd, cfg.feature_set, cfg.include_target_lags)
    feature_names = [f for f in feature_names if f in sd.exog.columns or f.startswith("target__")]
    if not feature_names:
        return pd.DataFrame()

    base_features = feature_names_for_set(sd, "schiff_no_lead" if cfg.stream in {"LIGHT_RUC", "HEAVY_RUC"} else "schiff", False)
    periods = valid_periods(sd)
    period_set = set(periods)
    origins = []
    for origin in periods:
        train_periods = [p for p in periods if period_sort_value(p) <= period_sort_value(origin)]
        if len(train_periods) >= cfg.min_train_quarters and any((origin + h) in period_set for h in range(1, MAX_HORIZON + 1)):
            origins.append(origin)
    origins = origins[::ORIGIN_STRIDE]
    if MAX_ORIGINS_PER_STREAM is not None:
        origins = origins[-MAX_ORIGINS_PER_STREAM:]

    records: List[Dict[str, Any]] = []
    for origin in origins:
        train_periods = [p for p in periods if period_sort_value(p) <= period_sort_value(origin)]
        if cfg.window is not None:
            train_periods = train_periods[-int(cfg.window):]
        if len(train_periods) < cfg.min_train_quarters:
            continue

        X, y = build_training_matrix(sd, train_periods, feature_names, cfg.include_target_lags)
        mask = y.notna()
        X, y = X.loc[mask], y.loc[mask]
        if len(X) < max(20, int(cfg.min_train_quarters * 0.60)):
            continue
        all_na_cols = [c for c in X.columns if X[c].isna().all()]
        if all_na_cols:
            X = X.copy()
            X[all_na_cols] = 0.0

        X_base = None
        base_all_na: List[str] = []
        if cfg.model_kind in RESIDUAL_KINDS:
            X_base, _ = build_training_matrix(sd, train_periods, base_features, False)
            X_base = X_base.reindex(index=X.index)
            base_all_na = [c for c in X_base.columns if X_base[c].isna().all()]
            if base_all_na:
                X_base = X_base.copy()
                X_base[base_all_na] = 0.0

        try:
            model = fit_model(cfg, X, y, X_base)
        except Exception:
            continue

        y_hist = {p: float(sd.y_log.loc[p]) for p in sd.y_log.index if pd.notna(sd.y_log.loc[p]) and period_sort_value(p) <= period_sort_value(origin)}
        for h in range(1, MAX_HORIZON + 1):
            target_p = origin + h
            if target_p not in period_set or target_p not in sd.y_raw.index or pd.isna(sd.y_raw.loc[target_p]):
                continue
            row = build_feature_row(target_p, sd, y_hist, feature_names, cfg.include_target_lags)
            Xp = pd.DataFrame([row]).reindex(columns=feature_names)
            if all_na_cols:
                Xp[all_na_cols] = 0.0
            Xpb = None
            if cfg.model_kind in RESIDUAL_KINDS:
                brow = build_feature_row(target_p, sd, y_hist, base_features, False)
                Xpb = pd.DataFrame([brow]).reindex(columns=base_features)
                if base_all_na:
                    Xpb[base_all_na] = 0.0
            try:
                pred_log = predict_model(model, Xp, Xpb)
            except Exception:
                pred_log = np.nan
            pred = safe_exp(pred_log)
            actual = float(sd.y_raw.loc[target_p])
            if np.isfinite(pred_log):
                y_hist[target_p] = pred_log
            records.append({
                "stream": cfg.stream,
                "model": cfg.name,
                "source_family": "bespoke_residual_correction" if cfg.model_kind in RESIDUAL_KINDS else "bespoke_sklearn",
                "model_kind": base_kind_for_resid(cfg.model_kind) if cfg.model_kind in RESIDUAL_KINDS else cfg.model_kind,
                "feature_set": cfg.feature_set,
                "family_tag": cfg.family_tag,
                "include_target_lags": cfg.include_target_lags,
                "window": "expanding" if cfg.window is None else cfg.window,
                "params_json": cfg.params_json,
                "origin": str(origin),
                "target_period": str(target_p),
                "horizon": h,
                "actual": actual,
                "pred": pred,
                "actual_log": float(sd.y_log.loc[target_p]) if pd.notna(sd.y_log.loc[target_p]) else np.nan,
                "pred_log": pred_log,
            })
    return pd.DataFrame(records)

# =============================================================================
# 6. Summaries, annual aggregation and stress tests
# =============================================================================

def summarise_quarterly(preds: pd.DataFrame) -> pd.DataFrame:
    if preds.empty:
        return pd.DataFrame()
    rows = []
    group_cols = ["stream", "model"]
    for (stream, model), g in preds.groupby(group_cols, sort=False):
        row = {
            "stream": stream,
            "model": model,
            "n_quarterly_pairs": int(len(g)),
            "n_origins": int(g["origin"].nunique()),
            "quarterly_mape": mape(g["actual"], g["pred"]),
            "quarterly_bias_pct": bias_pct(g["actual"], g["pred"]),
            "quarterly_p90_ape": p90_ape(g["actual"], g["pred"]),
            "quarterly_rmse": rmse(g["actual"], g["pred"]),
        }
        for h in range(1, MAX_HORIZON + 1):
            gh = g[g["horizon"] == h]
            row[f"mape_h{h:02d}"] = mape(gh["actual"], gh["pred"]) if not gh.empty else np.nan
        for label, lo, hi in [("h01_04", 1, 4), ("h05_08", 5, 8), ("h09_12", 9, 12)]:
            gh = g[g["horizon"].between(lo, hi)]
            row[f"mape_{label}"] = mape(gh["actual"], gh["pred"]) if not gh.empty else np.nan
        for c in ["source_family", "model_kind", "feature_set", "family_tag", "include_target_lags", "window", "params_json"]:
            row[c] = g[c].iloc[0] if c in g.columns else ""
        rows.append(row)
    return pd.DataFrame(rows)


def annualise(preds: pd.DataFrame) -> pd.DataFrame:
    if preds.empty:
        return pd.DataFrame()
    dfp = preds.copy()
    dfp["target_period_obj"] = dfp["target_period"].map(lambda s: pd.Period(str(s), freq="Q-DEC"))
    dfp["june_year"] = dfp["target_period_obj"].map(june_year_from_period)
    rows = []
    group_cols = ["stream", "model", "origin", "june_year"]
    for (stream, model, origin, jy), g in dfp.groupby(group_cols):
        if g["target_period_obj"].nunique() < 4:
            continue
        if stream == "PED":
            actual, pred = float(g["actual"].mean()), float(g["pred"].mean())
        else:
            actual, pred = float(g["actual"].sum()), float(g["pred"].sum())
        rows.append({"stream": stream, "model": model, "origin": origin, "june_year": int(jy), "actual": actual, "pred": pred, "n_quarters": int(g["target_period_obj"].nunique())})
    return pd.DataFrame(rows)


def summarise_annual(annual_preds: pd.DataFrame) -> pd.DataFrame:
    if annual_preds.empty:
        return pd.DataFrame()
    rows = []
    for (stream, model), g in annual_preds.groupby(["stream", "model"], sort=False):
        rows.append({
            "stream": stream,
            "model": model,
            "n_annual_pairs": int(len(g)),
            "annual_mape": mape(g["actual"], g["pred"]),
            "annual_bias_pct": bias_pct(g["actual"], g["pred"]),
            "annual_p90_ape": p90_ape(g["actual"], g["pred"]),
        })
    return pd.DataFrame(rows)


def build_final_summary(preds: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    qsum = summarise_quarterly(preds)
    annual_preds = annualise(preds)
    asum = summarise_annual(annual_preds)
    final = qsum.merge(asum, on=["stream", "model"], how="left")
    # stream-specific score, lower is better.
    score_parts = []
    for stream, g in final.groupby("stream"):
        x = g.copy()
        if stream == "PED":
            score = 0.55 * x["quarterly_mape"].rank(pct=True) + 0.25 * x["annual_mape"].rank(pct=True) + 0.10 * x["quarterly_p90_ape"].rank(pct=True) + 0.10 * x["quarterly_bias_pct"].abs().rank(pct=True)
        elif stream == "LIGHT_RUC":
            stress_col = "mape_h09_12" if "mape_h09_12" in x else "quarterly_p90_ape"
            score = 0.35 * x["quarterly_mape"].rank(pct=True) + 0.35 * x["annual_mape"].rank(pct=True) + 0.15 * x[stress_col].rank(pct=True) + 0.15 * x["quarterly_bias_pct"].abs().rank(pct=True)
        else:
            score = 0.35 * x["quarterly_mape"].rank(pct=True) + 0.45 * x["annual_mape"].rank(pct=True) + 0.10 * x["quarterly_p90_ape"].rank(pct=True) + 0.10 * x["quarterly_bias_pct"].abs().rank(pct=True)
        x["governance_score"] = score
        score_parts.append(x)
    final = pd.concat(score_parts, ignore_index=True)
    final = final.sort_values(["stream", "governance_score", "quarterly_mape"], na_position="last").reset_index(drop=True)
    return final, annual_preds

# =============================================================================
# 7. Ensembles
# =============================================================================

def common_matrix(preds: pd.DataFrame, stream: str, model_names: Sequence[str]) -> Tuple[pd.DataFrame, pd.Series, pd.DataFrame]:
    key_cols = ["stream", "origin", "target_period", "horizon"]
    sub = preds[(preds["stream"] == stream) & (preds["model"].isin(model_names))].copy()
    if sub.empty:
        return pd.DataFrame(), pd.Series(dtype=float), pd.DataFrame()
    pivot = sub.pivot_table(index=key_cols, columns="model", values="pred", aggfunc="first")
    actual = sub.groupby(key_cols)["actual"].first()
    pivot = pivot.reindex(columns=list(model_names))
    common = pivot.dropna(axis=0, how="any")
    if common.empty:
        return pd.DataFrame(), pd.Series(dtype=float), pd.DataFrame()
    actual = actual.loc[common.index]
    keys = pd.DataFrame(index=common.index).reset_index()
    return common.reset_index(drop=True), actual.reset_index(drop=True), keys


def optimise_convex_mape(P: np.ndarray, y: np.ndarray) -> np.ndarray:
    P = np.asarray(P, dtype=float)
    y = np.asarray(y, dtype=float)
    mask = np.isfinite(y) & (np.abs(y) > 1e-12) & np.isfinite(P).all(axis=1)
    P, y = P[mask], y[mask]
    n, k = P.shape
    if n == 0 or k == 0:
        return np.ones(k) / max(k, 1)
    c = np.zeros(k + n)
    c[k:] = 1.0 / np.maximum(np.abs(y), 1e-12)
    A_ub = np.zeros((2 * n, k + n))
    b_ub = np.zeros(2 * n)
    A_ub[:n, :k] = P
    A_ub[:n, k:] = -np.eye(n)
    b_ub[:n] = y
    A_ub[n:, :k] = -P
    A_ub[n:, k:] = -np.eye(n)
    b_ub[n:] = -y
    A_eq = np.zeros((1, k + n))
    A_eq[0, :k] = 1.0
    b_eq = np.array([1.0])
    bounds = [(0.0, 1.0)] * k + [(0.0, None)] * n
    res = linprog(c=c, A_ub=A_ub, b_ub=b_ub, A_eq=A_eq, b_eq=b_eq, bounds=bounds, method="highs")
    if res.success:
        w = np.maximum(np.asarray(res.x[:k], dtype=float), 0)
        if w.sum() > 0:
            return w / w.sum()
    def obj(w):
        return np.mean(np.abs(P @ w - y) / np.maximum(np.abs(y), 1e-12))
    res2 = minimize(obj, x0=np.ones(k) / k, bounds=[(0.0, 1.0)] * k, constraints=[{"type": "eq", "fun": lambda w: np.sum(w) - 1.0}], method="SLSQP")
    if res2.success:
        w = np.maximum(np.asarray(res2.x, dtype=float), 0)
        if w.sum() > 0:
            return w / w.sum()
    return np.ones(k) / k


def make_ensemble_frame(keys: pd.DataFrame, y: pd.Series, pred: np.ndarray, stream: str, name: str, kind: str, params: Dict[str, Any]) -> pd.DataFrame:
    out = keys.copy()
    out["stream"] = stream
    out["model"] = name
    out["source_family"] = "posthoc_ensemble"
    out["model_kind"] = kind
    out["feature_set"] = "ensemble"
    out["family_tag"] = kind
    out["include_target_lags"] = "mixed"
    out["window"] = "mixed"
    out["params_json"] = json.dumps(params, sort_keys=True)
    out["actual"] = y.values
    out["pred"] = pred
    out["actual_log"] = np.log(np.maximum(out["actual"].astype(float), 1e-12))
    out["pred_log"] = np.log(np.maximum(out["pred"].astype(float), 1e-12))
    return out



def build_fixed_weight_named_ensemble(
    preds: pd.DataFrame,
    stream: str,
    component_weights: Dict[str, float],
    ensemble_name: str,
    method: str,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Build a fixed-weight named ensemble from already evaluated component predictions."""
    models = [m for m, w in component_weights.items() if abs(float(w)) > 1e-12]
    if len(models) == 0:
        return pd.DataFrame(), pd.DataFrame()
    P, y, keys = common_matrix(preds, stream, models)
    if P.empty:
        missing = [m for m in models if m not in set(preds[preds["stream"] == stream]["model"])]
        print(f"WARNING: could not build {ensemble_name}; missing components: {missing}")
        return pd.DataFrame(), pd.DataFrame()
    w = np.asarray([float(component_weights[m]) for m in models], dtype=float)
    if w.sum() <= 0:
        w = np.ones(len(models)) / len(models)
    else:
        w = w / w.sum()
    pred = P.values @ w
    frame = make_ensemble_frame(keys, y, pred, stream, ensemble_name, method, {"models": models, "weights": dict(zip(models, [float(x) for x in w])), "method": method})
    weights = pd.DataFrame([
        {"stream": stream, "ensemble": ensemble_name, "component_model": m, "weight": float(weight), "method": method, "origin": ""}
        for m, weight in zip(models, w)
        if abs(weight) > 1e-12
    ])
    return frame, weights


def build_rescue_old_recon_blends(preds: pd.DataFrame, stream: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Build old-bespoke static, recon static and old/recon blend candidates on common full grid."""
    frames: List[pd.DataFrame] = []
    weights: List[pd.DataFrame] = []

    old_name = f"{stream}__OLD_BESPOKE_STATIC_REBUILT"
    rec_name = f"{stream}__RECON_STATIC_REBUILT"

    old_frame, old_w = build_fixed_weight_named_ensemble(
        preds,
        stream,
        OLD_BESPOKE_STATIC_COMPONENT_WEIGHTS,
        old_name,
        "fixed_old_bespoke_static_rebuilt",
    )
    if not old_frame.empty:
        frames.append(old_frame)
        weights.append(old_w)

    rec_frame, rec_w = build_fixed_weight_named_ensemble(
        preds,
        stream,
        HEAVY_RECON_STATIC_COMPONENT_WEIGHTS,
        rec_name,
        "fixed_recon_static_rebuilt",
    )
    if not rec_frame.empty:
        frames.append(rec_frame)
        weights.append(rec_w)

    # Build fixed blends between the rebuilt old and rebuilt reconciliation static ensembles.
    # This is the true full-grid test of the rescue insight that combined these two families.
    if not old_frame.empty and not rec_frame.empty:
        base_aug = pd.concat([preds, old_frame, rec_frame], ignore_index=True)
        for w_old in RESCUE_OLD_RECON_BLEND_STEPS:
            P, y, keys = common_matrix(base_aug, stream, [old_name, rec_name])
            if P.empty:
                continue
            w = np.asarray([float(w_old), float(1.0 - w_old)], dtype=float)
            pred = P.values @ w
            nm = f"{stream}__FULLGRID_RESCUE_BLEND_old{w_old:.3f}_recon{1.0 - w_old:.3f}"
            params = {"models": [old_name, rec_name], "weights": [float(w[0]), float(w[1])], "method": "fixed_old_recon_fullgrid_blend"}
            frames.append(make_ensemble_frame(keys, y, pred, stream, nm, "fixed_old_recon_fullgrid_blend", params))
            weights.append(pd.DataFrame([
                {"stream": stream, "ensemble": nm, "component_model": old_name, "weight": float(w[0]), "method": "fixed_old_recon_fullgrid_blend", "origin": ""},
                {"stream": stream, "ensemble": nm, "component_model": rec_name, "weight": float(w[1]), "method": "fixed_old_recon_fullgrid_blend", "origin": ""},
            ]))

        # Also solve the optimal two-component static convex blend directly.
        P, y, keys = common_matrix(base_aug, stream, [old_name, rec_name])
        if not P.empty:
            opt_w = optimise_convex_mape(P.values, y.values)
            pred = P.values @ opt_w
            nm = f"{stream}__FULLGRID_RESCUE_OPTIMAL_old_recon_static"
            frames.append(make_ensemble_frame(keys, y, pred, stream, nm, "solver_static_old_recon_only", {"models": [old_name, rec_name], "weights": [float(x) for x in opt_w], "method": "static_convex_old_recon_only"}))
            weights.append(pd.DataFrame([
                {"stream": stream, "ensemble": nm, "component_model": old_name, "weight": float(opt_w[0]), "method": "static_convex_old_recon_only", "origin": ""},
                {"stream": stream, "ensemble": nm, "component_model": rec_name, "weight": float(opt_w[1]), "method": "static_convex_old_recon_only", "origin": ""},
            ]))

    return (pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(),
            pd.concat(weights, ignore_index=True) if weights else pd.DataFrame())

def build_ensembles(preds: pd.DataFrame, summary: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    ens_frames = []
    weight_rows = []

    # First build explicit old-bespoke and reconciliation fixed ensembles, plus their full-grid rescue blends.
    for stream in STREAM_ORDER:
        named_frames, named_weights = build_rescue_old_recon_blends(preds, stream)
        if not named_frames.empty:
            ens_frames.append(named_frames)
        if not named_weights.empty:
            weight_rows.extend(named_weights.to_dict("records"))

    for stream in STREAM_ORDER:
        ssum = summary[summary["stream"] == stream].copy()
        if ssum.empty:
            continue
        # Ensemble pool includes top by several criteria and pure Schiff models.
        candidates = []
        for sort_cols in [["quarterly_mape"], ["annual_mape"], ["quarterly_p90_ape"], ["governance_score"]]:
            candidates.extend(ssum.sort_values(sort_cols, na_position="last")["model"].head(TOP_MODELS_FOR_ENSEMBLES).tolist())
        schiff = ssum[ssum["family_tag"].astype(str).str.contains("SCHIFF_OLS", case=False, na=False)]
        candidates.extend(schiff.sort_values(["quarterly_mape"], na_position="last")["model"].head(4).tolist())
        models = dedupe(candidates)[:max(TOP_MODELS_FOR_ENSEMBLES, 8)]
        if len(models) < 2:
            continue

        # Static convex.
        P, y, keys = common_matrix(preds, stream, models)
        if not P.empty:
            w = optimise_convex_mape(P.values, y.values)
            pred = P.values @ w
            nm = f"{stream}__solver_static_convex_top{len(models)}"
            ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, "solver_static_convex", {"base_models": models, "weights": dict(zip(models, [float(x) for x in w])), "method": "static_convex_mape"}))
            for m, weight in zip(models, w):
                if abs(weight) > 1e-10:
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": m, "weight": float(weight), "method": "static_convex_mape", "origin": ""})

        # Prequential convex.
        P, y, keys = common_matrix(preds, stream, models)
        if not P.empty:
            dfm = keys.copy()
            for m in models:
                dfm[m] = P[m].values
            dfm["actual"] = y.values
            dfm["origin_period"] = dfm["origin"].map(lambda s: pd.Period(str(s), freq="Q-DEC"))
            dfm["origin_sort"] = dfm["origin_period"].map(period_sort_value)
            dfm = dfm.sort_values(["origin_sort", "horizon", "target_period"]).reset_index(drop=True)
            preds_out = []
            eq_w = np.ones(len(models)) / len(models)
            for _, row in dfm.iterrows():
                hist = dfm[dfm["origin_sort"] < row["origin_sort"]]
                if len(hist) >= 40:
                    w = optimise_convex_mape(hist[models].values, hist["actual"].values)
                else:
                    w = eq_w
                preds_out.append(float(row[models].values.astype(float) @ w))
                for m, weight in zip(models, w):
                    if abs(weight) > 1e-10:
                        weight_rows.append({"stream": stream, "ensemble": f"{stream}__solver_preq_convex_top{len(models)}", "component_model": m, "weight": float(weight), "method": "prequential_convex_mape", "origin": row["origin"]})
            nm = f"{stream}__solver_preq_convex_top{len(models)}"
            ens_frames.append(make_ensemble_frame(keys, y, np.asarray(preds_out), stream, nm, "solver_preq_convex", {"base_models": models, "method": "prequential_convex_mape"}))

        # Top-k mean/median/trimmed.
        for k in [3, 5, 8, 10]:
            submodels = models[:k]
            if len(submodels) < 2:
                continue
            P, y, keys = common_matrix(preds, stream, submodels)
            if P.empty:
                continue
            arr = P.values
            for method in ["mean", "median", "trimmed_mean"]:
                if method == "mean":
                    pred = arr.mean(axis=1)
                elif method == "median":
                    pred = np.median(arr, axis=1)
                else:
                    if arr.shape[1] >= 4:
                        pred = np.sort(arr, axis=1)[:, 1:-1].mean(axis=1)
                    else:
                        pred = arr.mean(axis=1)
                nm = f"{stream}__top{k}_{method}"
                ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, f"topk_{method}", {"models": submodels, "method": method}))

        # Fixed blends with best pure Schiff.
        if not schiff.empty:
            schiff_model = schiff.sort_values(["quarterly_mape"], na_position="last")["model"].iloc[0]
            challenger_models = [m for m in models[:8] if m != schiff_model]
            for ch in challenger_models:
                P, y, keys = common_matrix(preds, stream, [schiff_model, ch])
                if P.empty:
                    continue
                for w_s in FIXED_BLEND_STEPS:
                    w = np.array([w_s, 1 - w_s], dtype=float)
                    pred = P.values @ w
                    nm = f"{stream}__fixedblend_schiff{w_s:.2f}__{ch}"
                    ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, "fixed_schiff_blend", {"models": [schiff_model, ch], "weights": [float(w[0]), float(w[1])], "method": "fixed_schiff_challenger_blend"}))
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": schiff_model, "weight": float(w[0]), "method": "fixed_schiff_challenger_blend", "origin": ""})
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": ch, "weight": float(w[1]), "method": "fixed_schiff_challenger_blend", "origin": ""})

    ensemble_preds = pd.concat(ens_frames, ignore_index=True) if ens_frames else pd.DataFrame()
    weights = pd.DataFrame(weight_rows)
    return ensemble_preds, weights

# =============================================================================
# 8. Paired comparisons, stress tests, recommendations
# =============================================================================

def choose_schiff_benchmark(summary: pd.DataFrame, stream: str) -> Optional[str]:
    s = summary[(summary["stream"] == stream) & (summary["family_tag"].astype(str).str.contains("SCHIFF_OLS", case=False, na=False))].copy()
    if s.empty:
        return None
    # Prefer W40 if close; otherwise choose best pure Schiff by quarterly MAPE.
    w40 = s[s["window"].astype(str).eq("40")]
    if not w40.empty:
        return w40.sort_values(["quarterly_mape", "annual_mape"], na_position="last")["model"].iloc[0]
    return s.sort_values(["quarterly_mape", "annual_mape"], na_position="last")["model"].iloc[0]


def paired_compare(preds: pd.DataFrame, candidate: str, benchmark: str) -> Dict[str, Any]:
    key = ["stream", "origin", "target_period", "horizon"]
    c = preds[preds["model"] == candidate][key + ["actual", "pred"]].rename(columns={"pred": "candidate_pred"})
    b = preds[preds["model"] == benchmark][key + ["actual", "pred"]].rename(columns={"actual": "benchmark_actual", "pred": "benchmark_pred"})
    m = c.merge(b, on=key, how="inner")
    if m.empty:
        return {"candidate": candidate, "benchmark": benchmark, "n_common_pairs": 0}
    actual = m["actual"].astype(float)
    ce = np.abs(m["candidate_pred"] - actual) / np.abs(actual)
    be = np.abs(m["benchmark_pred"] - actual) / np.abs(actual)
    valid = np.isfinite(ce) & np.isfinite(be) & (np.abs(actual) > 1e-12)
    if valid.sum() == 0:
        return {"candidate": candidate, "benchmark": benchmark, "n_common_pairs": 0}
    stream = m["stream"].iloc[0]
    return {
        "stream": stream,
        "baseline": benchmark,
        "challenger": candidate,
        "n_common_pairs": int(valid.sum()),
        "baseline_mape": float(be[valid].mean() * 100),
        "challenger_mape": float(ce[valid].mean() * 100),
        "mape_improvement_pct_points": float((be[valid].mean() - ce[valid].mean()) * 100),
        "challenger_win_rate": float((ce[valid] < be[valid]).mean() * 100),
    }


def build_stress_tests(preds: pd.DataFrame, annual_preds: pd.DataFrame, models: Sequence[str]) -> pd.DataFrame:
    rows = []
    dfp = preds.copy()
    dfp["target_period_obj"] = dfp["target_period"].map(lambda s: pd.Period(str(s), freq="Q-DEC"))
    for model in models:
        g = dfp[dfp["model"] == model].copy()
        if g.empty:
            continue
        stream = g["stream"].iloc[0]
        row = {"stream": stream, "model": model}
        slices = {
            "h1_4": g["horizon"].between(1, 4),
            "h5_8": g["horizon"].between(5, 8),
            "h9_12": g["horizon"].between(9, 12),
            "recent_2024_plus": g["target_period_obj"] >= pd.Period("2024Q1", freq="Q-DEC"),
            "stress_2022_23": g["target_period_obj"].between(pd.Period("2022Q1", freq="Q-DEC"), pd.Period("2023Q4", freq="Q-DEC")),
        }
        for label, mask in slices.items():
            gg = g[mask]
            row[f"{label}_n"] = int(len(gg))
            row[f"{label}_mape"] = mape(gg["actual"], gg["pred"]) if not gg.empty else np.nan
            row[f"{label}_bias"] = bias_pct(gg["actual"], gg["pred"]) if not gg.empty else np.nan
        a = annual_preds[annual_preds["model"] == model]
        row["annual_n"] = int(len(a))
        row["annual_mape"] = mape(a["actual"], a["pred"]) if not a.empty else np.nan
        row["annual_bias"] = bias_pct(a["actual"], a["pred"]) if not a.empty else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


def recommend_models(summary: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    primary = []
    by_q = []
    for stream in STREAM_ORDER:
        s = summary[summary["stream"] == stream].copy()
        if s.empty:
            continue
        primary.append(s.sort_values(["governance_score", "quarterly_mape", "annual_mape"], na_position="last").head(1))
        by_q.append(s.sort_values(["quarterly_mape", "annual_mape", "quarterly_p90_ape"], na_position="last").head(1))
    return pd.concat(primary, ignore_index=True), pd.concat(by_q, ignore_index=True)

# =============================================================================
# 9. Outputs and figures
# =============================================================================

def make_figures(run_dir: Path, summary: pd.DataFrame, primary: pd.DataFrame) -> None:
    try:
        import matplotlib.pyplot as plt
        fig_dir = run_dir / "figures"
        fig_dir.mkdir(exist_ok=True)

        # Finalist vs expected PDF bar chart.
        rows = []
        for _, r in primary.iterrows():
            stream = r["stream"]
            rows.append({"stream": stream, "type": "Latest quarterly", "mape": r["quarterly_mape"]})
            rows.append({"stream": stream, "type": "Latest annual", "mape": r["annual_mape"]})
            if stream in PDF_EXPECTED:
                rows.append({"stream": stream, "type": "PDF quarterly", "mape": PDF_EXPECTED[stream]["quarterly_mape"]})
                rows.append({"stream": stream, "type": "PDF annual", "mape": PDF_EXPECTED[stream]["annual_mape"]})
        d = pd.DataFrame(rows)
        if not d.empty:
            order = STREAM_ORDER
            types = ["PDF quarterly", "Latest quarterly", "PDF annual", "Latest annual"]
            x = np.arange(len(order))
            width = 0.18
            plt.figure(figsize=(10, 5))
            for i, t in enumerate(types):
                vals = [d[(d["stream"] == s) & (d["type"] == t)]["mape"].mean() for s in order]
                plt.bar(x + (i - 1.5) * width, vals, width, label=t)
                for j, v in enumerate(vals):
                    if np.isfinite(v):
                        plt.text(x[j] + (i - 1.5) * width, v + 0.05, f"{v:.2f}", ha="center", va="bottom", fontsize=8)
            plt.xticks(x, [STREAM_LABELS[s] for s in order])
            plt.ylabel("MAPE (%)")
            plt.title("Recommended finalists vs earlier PDF finalist metrics")
            plt.legend()
            plt.tight_layout()
            plt.savefig(fig_dir / "finalists_vs_pdf.png", dpi=160)
            plt.close()

        # Candidate landscape.
        plt.figure(figsize=(10, 5.5))
        colors = {"PED": "tab:blue", "LIGHT_RUC": "tab:orange", "HEAVY_RUC": "tab:green"}
        for stream in STREAM_ORDER:
            s = summary[summary["stream"] == stream]
            plt.scatter(s["quarterly_mape"], s["annual_mape"], s=20, alpha=0.30, label=STREAM_LABELS[stream], color=colors[stream])
            p = primary[primary["stream"] == stream]
            if not p.empty:
                plt.scatter(p["quarterly_mape"], p["annual_mape"], s=120, marker="*", color=colors[stream], edgecolor="black")
        plt.xlabel("Quarterly MAPE (%)")
        plt.ylabel("Annual MAPE (%)")
        plt.title("Candidate arbitration landscape")
        plt.legend()
        plt.tight_layout()
        plt.savefig(fig_dir / "candidate_landscape.png", dpi=160)
        plt.close()
    except Exception as e:
        print(f"Figure generation skipped: {type(e).__name__}: {e}")


def write_excel(run_dir: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    path = run_dir / "stage1_finalist_arbitration_results.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None or df.empty:
                continue
            safe_name = name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        wb = load_workbook(path)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top")
            for i, col_cells in enumerate(ws.columns, 1):
                max_len = 10
                for cell in col_cells:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(val), 70))
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2
        wb.save(path)
    except Exception as e:
        print(f"Excel styling skipped: {type(e).__name__}: {e}")


def write_report(run_dir: Path, primary: pd.DataFrame, by_q: pd.DataFrame, paired: pd.DataFrame, pdf_comp: pd.DataFrame, stress: pd.DataFrame) -> None:
    lines = []
    lines.append("# Stage 1 Finalist Arbitration Report\n")
    lines.append(f"Run folder: `{run_dir}`\n")
    lines.append("\n## Purpose\n")
    lines.append("This run arbitrates between the earlier PDF-style finalist families, the latest bespoke Heavy RUC gains, and the pure Schiff structural benchmarks on a common full rolling-origin validation grid.\n")
    lines.append("\n## Recommended finalists by stream-specific governance score\n")
    lines.append(primary[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "annual_bias_pct", "governance_score"]].to_markdown(index=False))
    lines.append("\n\n## Recommended finalists by quarterly MAPE\n")
    lines.append(by_q[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "governance_score"]].to_markdown(index=False))
    lines.append("\n\n## Comparison with earlier PDF metrics\n")
    lines.append(pdf_comp.to_markdown(index=False))
    lines.append("\n\n## Paired comparison against Schiff benchmarks\n")
    lines.append(paired.head(80).to_markdown(index=False))
    lines.append("\n\n## Stress tests\n")
    lines.append(stress.to_markdown(index=False))
    lines.append("\n\n## Interpretation guidance\n")
    lines.append("- If a new candidate beats the PDF metric but is a static solver, inspect the prequential solver as a production-safer comparator.\n")
    lines.append("- If Light RUC has weaker quarterly MAPE but better annual MAPE, inspect the 2022--23 stress window before choosing.\n")
    lines.append("- If a model uses lead price features, treat it as actual-driver/model-discovery unless the relevant policy path is known at the forecast origin.\n")
    (run_dir / "stage1_finalist_arbitration_report.md").write_text("\n".join(lines), encoding="utf-8")

# =============================================================================
# 10. Main
# =============================================================================

def main() -> None:
    t0 = time.time()
    run_dir = ensure_run_dir(OUTPUT_ROOT)
    print("=" * 100)
    print("HEAVY RUC BESPOKE RECONCILIATION FULL-GRID WORKBENCH")
    print("=" * 100)
    print(f"Workbook: {INPUT_XLSX}")
    print(f"Run folder: {run_dir}")

    df = load_input_sheet(INPUT_XLSX)
    input_sheet = df.attrs.get("input_sheet", "")
    targets = {s: detect_target_col(df, s) for s in STREAM_ORDER}
    all_targets = [v[0] for v in targets.values()]

    stream_data: Dict[str, StreamData] = {}
    feature_inventory = []
    for stream in STREAM_ORDER:
        target_col, target_is_log = targets[stream]
        fcols = detect_feature_cols(df, stream, all_targets)
        y_raw, y_log = build_target_series(df, target_col, target_is_log)
        exog, groups, primary_log = build_exog(df, stream, fcols)
        sd = StreamData(stream, target_col, target_is_log, fcols, y_raw, y_log, exog, groups, primary_log)
        stream_data[stream] = sd
        for c in fcols:
            feature_inventory.append({"stream": stream, "feature_column": c, "group": assign_base_group(stream, c)})
        print(f"{stream}: target={target_col}; rows={y_raw.dropna().shape[0]}; period={min(y_raw.dropna().index)} to {max(y_raw.dropna().index)}; features={len(fcols)}")

    pd.DataFrame(feature_inventory).to_csv(run_dir / "feature_inventory.csv", index=False)

    # Generate and evaluate curated candidates.
    all_pred_frames = []
    config_rows = []
    for stream in STREAM_ORDER:
        configs = generate_targeted_candidates(stream)
        print(f"{stream}: candidate configs={len(configs):,}")
        config_rows.extend([asdict(c) for c in configs])
        sd = stream_data[stream]
        for i, cfg in enumerate(configs, 1):
            if i % 250 == 0:
                print(f"  {stream}: evaluated {i:,}/{len(configs):,}")
            pred = evaluate_candidate(sd, cfg)
            if not pred.empty:
                all_pred_frames.append(pred)
    pd.DataFrame(config_rows).to_csv(run_dir / "candidate_config_inventory.csv", index=False)

    base_preds = pd.concat(all_pred_frames, ignore_index=True) if all_pred_frames else pd.DataFrame()
    if base_preds.empty:
        raise RuntimeError("No base predictions were generated.")
    base_preds.to_csv(run_dir / "base_quarterly_predictions.csv", index=False)

    base_summary, base_annual = build_final_summary(base_preds)
    base_summary.to_csv(run_dir / "base_summary.csv", index=False)
    base_annual.to_csv(run_dir / "base_annual_predictions.csv", index=False)

    print("Building ensembles...")
    ens_preds, weights = build_ensembles(base_preds, base_summary)
    weights.to_csv(run_dir / "ensemble_weights.csv", index=False)
    ens_preds.to_csv(run_dir / "ensemble_quarterly_predictions.csv", index=False)

    all_preds = pd.concat([base_preds, ens_preds], ignore_index=True) if not ens_preds.empty else base_preds.copy()
    all_preds.to_csv(run_dir / "all_quarterly_predictions.csv", index=False)
    final_summary, all_annual = build_final_summary(all_preds)
    final_summary.to_csv(run_dir / "final_summary.csv", index=False)
    all_annual.to_csv(run_dir / "all_annual_predictions.csv", index=False)

    primary, by_q = recommend_models(final_summary)
    primary.to_csv(run_dir / "recommended_finalists_primary.csv", index=False)
    by_q.to_csv(run_dir / "recommended_finalists_by_quarterly.csv", index=False)
    # Backward-compatible name: primary recommendation.
    primary.to_csv(run_dir / "recommended_finalists.csv", index=False)

    # Pair against Schiff and top candidates.
    paired_rows = []
    stress_models = set(primary["model"].tolist() + by_q["model"].tolist())
    for stream in STREAM_ORDER:
        schiff = choose_schiff_benchmark(final_summary, stream)
        if schiff:
            stress_models.add(schiff)
            top_models = final_summary[final_summary["stream"] == stream].sort_values(["governance_score", "quarterly_mape"], na_position="last")["model"].head(30).tolist()
            for model in dedupe(primary[primary["stream"] == stream]["model"].tolist() + by_q[by_q["stream"] == stream]["model"].tolist() + top_models):
                if model != schiff:
                    paired_rows.append(paired_compare(all_preds, model, schiff))
    paired = pd.DataFrame(paired_rows)
    paired.to_csv(run_dir / "paired_vs_schiff.csv", index=False)

    stress = build_stress_tests(all_preds, all_annual, list(stress_models))
    stress.to_csv(run_dir / "stress_tests.csv", index=False)

    # PDF expected comparison.
    comp_rows = []
    for _, r in primary.iterrows():
        stream = r["stream"]
        expected = PDF_EXPECTED.get(stream, {})
        latest = LATEST_BESPOKE_EXPECTED.get(stream, {})
        comp_rows.append({
            "stream": stream,
            "selected_model": r["model"],
            "selected_quarterly_mape": r["quarterly_mape"],
            "pdf_quarterly_mape": expected.get("quarterly_mape", np.nan),
            "selected_minus_pdf_q_pp": r["quarterly_mape"] - expected.get("quarterly_mape", np.nan),
            "selected_annual_mape": r["annual_mape"],
            "pdf_annual_mape": expected.get("annual_mape", np.nan),
            "selected_minus_pdf_a_pp": r["annual_mape"] - expected.get("annual_mape", np.nan),
            "latest_bespoke_q_mape": latest.get("quarterly_mape", np.nan),
            "selected_minus_latest_q_pp": r["quarterly_mape"] - latest.get("quarterly_mape", np.nan),
            "latest_bespoke_a_mape": latest.get("annual_mape", np.nan),
            "selected_minus_latest_a_pp": r["annual_mape"] - latest.get("annual_mape", np.nan),
        })
    pdf_comp = pd.DataFrame(comp_rows)
    pdf_comp.to_csv(run_dir / "pdf_expected_comparison.csv", index=False)

    top50 = pd.concat([final_summary[final_summary["stream"] == s].sort_values(["governance_score", "quarterly_mape"], na_position="last").head(50) for s in STREAM_ORDER], ignore_index=True)
    top50.to_csv(run_dir / "top50_by_stream.csv", index=False)

    make_figures(run_dir, final_summary, primary)
    write_excel(run_dir, {
        "Recommended Primary": primary,
        "Recommended By Q MAPE": by_q,
        "PDF Comparison": pdf_comp,
        "Paired Vs Schiff": paired,
        "Stress Tests": stress,
        "Final Summary": final_summary,
        "Top 50 By Stream": top50,
        "Ensemble Weights": weights,
        "Feature Inventory": pd.DataFrame(feature_inventory),
    })
    write_report(run_dir, primary, by_q, paired, pdf_comp, stress)

    elapsed = (time.time() - t0) / 60
    print("\nRUN COMPLETE")
    print("=" * 100)
    print(f"Elapsed minutes: {elapsed:.1f}")
    print(f"Input sheet: {input_sheet}")
    print(f"Run folder: {run_dir}")
    print("Recommended finalists by primary governance score:")
    print(primary[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "governance_score"]].to_markdown(index=False))
    print("\nKey files:")
    for fn in ["recommended_finalists.csv", "recommended_finalists_by_quarterly.csv", "final_summary.csv", "paired_vs_schiff.csv", "stress_tests.csv", "ensemble_weights.csv", "pdf_expected_comparison.csv", "stage1_finalist_arbitration_results.xlsx", "stage1_finalist_arbitration_report.md"]:
        print(run_dir / fn)


if __name__ == "__main__":
    main()
