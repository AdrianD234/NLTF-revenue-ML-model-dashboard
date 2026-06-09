# -*- coding: utf-8 -*-
"""
Stage 1 Finalist Arbitration Workbench
======================================

Purpose
-------
Run a narrow, decision-grade Stage 1 actual-driver validation that arbitrates
between:

1. the earlier PDF-style finalist families that performed well for PED and Light RUC;
2. the latest bespoke-solver finalist families that improved Heavy RUC;
3. the pure Schiff structural benchmarks; and
4. targeted near-neighbour models around those known-good candidates.

This is deliberately NOT a broad AutoGluon search and NOT a full brute-force search.
It is a targeted full-grid validation over a curated candidate set, with the same
forecast origins / target periods / horizons used for fair model comparisons.

The models are fitted on log targets and scored after exponentiating forecasts
back to real-world units.

Expected use
------------
Run from PowerShell, for example:

    conda activate agts312
    cd "<repo-root>"
    python "stage1_finalist_arbitration.py"

Before running, confirm INPUT_XLSX points to the latest Master Copy workbook.

Outputs
-------
A timestamped run folder is created under OUTPUT_ROOT containing:

    recommended_finalists_primary.csv
    recommended_finalists_by_quarterly.csv
    final_summary.csv
    paired_vs_schiff.csv
    paired_vs_pdf_style.csv
    stress_tests.csv
    ensemble_weights.csv
    all_quarterly_predictions.csv
    all_annual_predictions.csv
    pdf_expected_comparison.csv
    top50_by_stream.csv
    stage1_finalist_arbitration_results.xlsx
    stage1_finalist_arbitration_report.md

Notes
-----
Stage 1 actual-driver testing uses realised future explanatory variables. It is
therefore a model-structure / volume-model test, not yet a full production test
with vintage macro or fuel-price forecast-error propagation.
"""

from __future__ import annotations

import json
import math
import re
import shutil
import time
import warnings
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from scipy.optimize import linprog, minimize
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, BayesianRidge, ElasticNet, HuberRegressor
from sklearn.ensemble import GradientBoostingRegressor, ExtraTreesRegressor, RandomForestRegressor

warnings.filterwarnings("ignore")

try:
    from xgboost import XGBRegressor
    HAS_XGBOOST = True
except Exception:
    HAS_XGBOOST = False

# =============================================================================
# 0. User configuration
# =============================================================================

# Sanitized vendored copy: override these repo-relative placeholders before rerun.
INPUT_XLSX = Path("data/source_workbooks/Master Copy revenue modelling workbook.xlsx")
OUTPUT_ROOT = Path("artifacts/ped_inner_hpo/finalist_arbitration_outputs")

RUN_PROFILE = "arbitration"  # keep this script narrow and full-grid by design
RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)

MAX_HORIZON = 12
MIN_TRAIN_QUARTERS = 40

# Full-grid validation for curated candidates. This is much smaller than the broad search,
# so origin_stride=1 is intentional.
ORIGIN_STRIDE = 1
MAX_ORIGINS_PER_STREAM: Optional[int] = None

# Feature/timing assumptions.
ALLOW_LEAD_PRICE_FEATURES_FOR_DISCOVERY = True
INCLUDE_XGBOOST_IF_AVAILABLE = False  # keep off by default; sklearn GBM was sufficient and faster/stabler
TREE_N_JOBS = 1

# Candidate generation controls.
ADD_RANDOM_LOCAL_NEIGHBOURS = True
N_RANDOM_LOCAL_NEIGHBOURS_PER_STREAM = 80

# Ensemble controls.
TOP_MODELS_FOR_ENSEMBLES = 18
FIXED_BLEND_STEPS = np.round(np.linspace(0.0, 1.0, 21), 2)

# Expected PDF-finalist metrics from the earlier scientific report. These are not used
# for model fitting; they are reference values in the output comparison table.
PDF_EXPECTED = {
    "PED": {"quarterly_mape": 2.48, "annual_mape": 2.42, "label": "PDF finalist"},
    "LIGHT_RUC": {"quarterly_mape": 9.16, "annual_mape": 6.25, "label": "PDF finalist"},
    "HEAVY_RUC": {"quarterly_mape": 3.80, "annual_mape": 3.07, "label": "PDF finalist"},
}

LATEST_BESPOKE_EXPECTED = {
    "PED": {"quarterly_mape": 2.65701, "annual_mape": 2.60088, "label": "Latest bespoke finalist"},
    "LIGHT_RUC": {"quarterly_mape": 10.3145, "annual_mape": 6.70964, "label": "Latest bespoke finalist"},
    "HEAVY_RUC": {"quarterly_mape": 3.27657, "annual_mape": 2.44456, "label": "Latest bespoke finalist"},
}

POLICY_WINDOWS = {
    "ruc_discount_active": ("2022Q2", "2023Q2"),
    "ruc_discount_reversal": ("2023Q3", "2023Q4"),
    "post_discount_normalisation": ("2024Q1", "2025Q4"),
}

STREAM_ORDER = ["PED", "LIGHT_RUC", "HEAVY_RUC"]
STREAM_LABELS = {
    "PED": "PED VKT per capita",
    "LIGHT_RUC": "Light RUC volume",
    "HEAVY_RUC": "Heavy RUC volume",
}

# =============================================================================
# 1. Utilities
# =============================================================================

def normalise_name(x: Any) -> str:
    s = str(x).strip().lower().replace("\n", " ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def snake_name(x: Any) -> str:
    s = normalise_name(x)
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


def safe_float_series(s: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")
    return pd.to_numeric(
        s.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip(),
        errors="coerce",
    )


def parse_quarter_value(x: Any) -> Optional[pd.Period]:
    if pd.isna(x):
        return None
    if isinstance(x, pd.Period):
        return pd.Period(str(x), freq="Q-DEC")
    if isinstance(x, pd.Timestamp):
        return x.to_period("Q-DEC")
    s = str(x).strip()
    if not s:
        return None
    m = re.search(r"(19\d{2}|20\d{2})\s*[- ]?\s*[Qq]\s*([1-4])", s)
    if m:
        return pd.Period(f"{int(m.group(1))}Q{int(m.group(2))}", freq="Q-DEC")
    m = re.search(r"[Qq]\s*([1-4])\s*[- ]?\s*(19\d{2}|20\d{2})", s)
    if m:
        return pd.Period(f"{int(m.group(2))}Q{int(m.group(1))}", freq="Q-DEC")
    try:
        if re.fullmatch(r"\d+(\.\d+)?", s):
            val = float(s)
            if 20000 < val < 60000:
                return (pd.Timestamp("1899-12-30") + pd.Timedelta(days=val)).to_period("Q-DEC")
    except Exception:
        pass
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.notna(ts):
            return ts.to_period("Q-DEC")
    except Exception:
        pass
    return None


def period_sort_value(p: pd.Period) -> int:
    return int(p.year * 4 + p.quarter)


def june_year_from_period(p: pd.Period) -> int:
    return int(p.year + 1) if p.quarter in (3, 4) else int(p.year)


def period_between(p: pd.Period, start: str, end: str) -> bool:
    return pd.Period(start, freq="Q-DEC") <= p <= pd.Period(end, freq="Q-DEC")


def mape(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    return float(np.mean(np.abs(p[mask] - a[mask]) / np.abs(a[mask])) * 100.0)


def bias_pct(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    return float(np.mean((p[mask] - a[mask]) / np.abs(a[mask])) * 100.0)


def p90_ape(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    ape = np.abs(p[mask] - a[mask]) / np.abs(a[mask]) * 100.0
    return float(np.nanpercentile(ape, 90))


def rmse(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p)
    if mask.sum() == 0:
        return np.nan
    return float(np.sqrt(np.mean((p[mask] - a[mask]) ** 2)))


def safe_exp(x: float) -> float:
    if not np.isfinite(x):
        return np.nan
    return float(np.exp(np.clip(x, -50, 50)))


def dedupe(seq: Iterable[str]) -> List[str]:
    out, seen = [], set()
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def boolish(x: Any) -> bool:
    if isinstance(x, bool):
        return x
    if pd.isna(x):
        return False
    return str(x).strip().lower() in {"true", "1", "yes", "y"}


def ensure_run_dir(root: Path) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    run_dir = root / f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    run_dir.mkdir(parents=True, exist_ok=False)
    (run_dir / "figures").mkdir(exist_ok=True)
    return run_dir

# =============================================================================
# 2. Workbook loading and schema detection
# =============================================================================

TARGET_PATTERNS = {
    "PED": [
        r"light.*petrol.*vkt.*cap",
        r"petrol.*vkt.*cap",
        r"vkt.*per.*cap",
        r"vkt.*pc",
    ],
    "LIGHT_RUC": [
        r"light.*ruc.*volume",
        r"light.*ruc.*net.*km",
        r"light.*ruc.*kilomet",
        r"light.*iuc.*volume",
        r"light.*iuc.*net.*km",
        r"light.*net.*km",
    ],
    "HEAVY_RUC": [
        r"heavy.*ruc.*volume",
        r"heavy.*ruc.*net.*km",
        r"heavy.*ruc.*kilomet",
        r"heavy.*iuc.*volume",
        r"heavy.*iuc.*net.*km",
        r"heavy.*net.*km",
    ],
}

FEATURE_BLACKLIST = [
    "household", "hes", "hils", "survey", "hies", "disposable", "rent", "mortgage",
    "housing cost", "income decile", "brent", "wti", "crude", "barrel", "importer",
    "margin", "spread", "gst", "excise", "levy", "carbon", "ets", "nzu", "emissions",
    "shipping", "refining", "refinery", "forecast", "projection", "error", "residual",
    "mape", "ape", "p50", "p80", "p90", "confidence", "upper", "lower", "oracle", "freeze",
]

FEATURE_ALLOW_TERMS = {
    "PED": ["petrol", "gdp", "gross domestic", "unemployment", "unemploy", "cpi", "population", "per capita", "gdppc", "gdp pc"],
    "LIGHT_RUC": ["diesel", "light ruc", "light iuc", "ruc price", "iuc price", "gdp", "gross domestic", "unemployment", "unemploy", "cpi"],
    "HEAVY_RUC": ["diesel", "heavy ruc", "heavy iuc", "ruc price", "iuc price", "gdp", "gross domestic", "unemployment", "unemploy", "cpi"],
}


def is_log_col(col: str) -> bool:
    n = normalise_name(col)
    return n.startswith("ln ") or " ln " in f" {n} " or "log" in n or "natural log" in n


def load_input_sheet(xlsx_path: Path) -> pd.DataFrame:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")
    xl = pd.ExcelFile(xlsx_path)
    preferred = ["Stage 1 Inputs", "Stage1 Inputs", "Stage 1", "Model Inputs", "Inputs", "Data"]
    norm_to_actual = {normalise_name(s): s for s in xl.sheet_names}
    sheet = None
    for p in preferred:
        if normalise_name(p) in norm_to_actual:
            sheet = norm_to_actual[normalise_name(p)]
            break
    if sheet is None:
        best_score, best_sheet = -1, None
        for s in xl.sheet_names:
            try:
                tmp = pd.read_excel(xlsx_path, sheet_name=s, nrows=8)
            except Exception:
                continue
            joined = " ".join(normalise_name(c) for c in tmp.columns)
            score = sum(kw in joined for kw in ["quarter", "vkt", "ruc", "iuc", "petrol", "diesel", "gdp", "unemployment", "cpi"])
            if score > best_score:
                best_score, best_sheet = score, s
        sheet = best_sheet
    if sheet is None:
        raise RuntimeError("Could not detect a model input sheet.")

    df = pd.read_excel(xlsx_path, sheet_name=sheet)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

    norm_cols = {c: normalise_name(c) for c in df.columns}
    quarter_col = None
    for c, n in norm_cols.items():
        if "quarter" in n and df[c].notna().sum() >= 10:
            quarter_col = c
            break
    if quarter_col is None:
        for c, n in norm_cols.items():
            if ("date" in n or "period" in n) and df[c].notna().sum() >= 10:
                quarter_col = c
                break
    if quarter_col is None:
        year_cols = [c for c, n in norm_cols.items() if n in {"year", "financial year", "fiscal year"}]
        q_cols = [c for c, n in norm_cols.items() if n in {"quarter", "qtr", "q"}]
        if not year_cols or not q_cols:
            raise RuntimeError("Could not detect quarter/date column.")
        periods = (df[year_cols[0]].astype(str) + "Q" + df[q_cols[0]].astype(str).str.extract(r"([1-4])", expand=False).fillna(""))
        periods = periods.map(parse_quarter_value)
    else:
        periods = df[quarter_col].map(parse_quarter_value)

    df = df.copy()
    df["__period__"] = periods
    df = df[df["__period__"].notna()].copy()
    df["__sort__"] = df["__period__"].map(period_sort_value)
    df = df.sort_values("__sort__").drop_duplicates("__period__", keep="last")
    df = df.drop(columns=["__sort__"]).reset_index(drop=True)
    df.attrs["input_sheet"] = sheet
    return df


def detect_target_col(df: pd.DataFrame, stream: str) -> Tuple[str, bool]:
    patterns = TARGET_PATTERNS[stream]
    candidates = []
    for c in df.columns:
        if c == "__period__":
            continue
        n = normalise_name(c)
        if "revenue" in n:
            continue
        if any(re.search(p, n) for p in patterns):
            s = safe_float_series(df[c])
            non_na = s.notna().sum()
            if non_na >= 20:
                score = 100 + min(non_na, 200) / 10 - (5 if is_log_col(c) else 0)
                candidates.append((score, c, is_log_col(c), non_na))
    if not candidates:
        raise RuntimeError(f"Could not detect target column for {stream}.")
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1], candidates[0][2]


def detect_feature_cols(df: pd.DataFrame, stream: str, all_targets: Sequence[str]) -> List[str]:
    allow = FEATURE_ALLOW_TERMS[stream]
    cols = []
    for c in df.columns:
        if c == "__period__" or c in all_targets:
            continue
        n = normalise_name(c)
        if any(b in n for b in FEATURE_BLACKLIST):
            continue
        if "revenue" in n:
            continue
        # Avoid pulling in other target-like series unless explicitly a price/rate.
        if any(t in n for t in ["volume", "vkt", "net km", "net kilomet", "kilometres", "kilometers", "kms"]):
            if "price" not in n and "rate" not in n:
                continue
        if not any(term in n for term in allow):
            continue
        s = safe_float_series(df[c])
        if s.notna().sum() >= 20:
            cols.append(c)
    if not cols:
        raise RuntimeError(f"No feature columns detected for {stream}.")
    return cols

# =============================================================================
# 3. Feature engineering
# =============================================================================

@dataclass
class StreamData:
    stream: str
    target_col: str
    target_is_log: bool
    feature_cols: List[str]
    y_raw: pd.Series
    y_log: pd.Series
    exog: pd.DataFrame
    groups: Dict[str, List[str]]
    primary_log: Dict[str, str]


def assign_base_group(stream: str, col: str) -> str:
    n = normalise_name(col)
    if "petrol" in n and ("price" in n or "cents" in n or "cost" in n):
        return "petrol_price"
    if "diesel" in n and ("price" in n or "cents" in n or "cost" in n):
        return "diesel_price"
    if "light" in n and ("ruc" in n or "iuc" in n) and ("price" in n or "rate" in n or "cost" in n):
        return "light_ruc_price"
    if "heavy" in n and ("ruc" in n or "iuc" in n) and ("price" in n or "rate" in n or "cost" in n):
        return "heavy_ruc_price"
    if ("ruc price" in n or "iuc price" in n) and stream == "LIGHT_RUC":
        return "light_ruc_price"
    if ("ruc price" in n or "iuc price" in n) and stream == "HEAVY_RUC":
        return "heavy_ruc_price"
    if "gdp" in n or "gross domestic" in n:
        if "per capita" in n or "pc" in n:
            return "gdp_pc"
        return "gdp"
    if "unemployment" in n or "unemploy" in n:
        return "unemployment"
    if "cpi" in n:
        return "cpi"
    if "population" in n:
        return "population"
    return "other_core"


def add_group(groups: Dict[str, List[str]], group: str, name: str) -> None:
    groups.setdefault(group, [])
    if name not in groups[group]:
        groups[group].append(name)


def build_target_series(df: pd.DataFrame, target_col: str, target_is_log: bool) -> Tuple[pd.Series, pd.Series]:
    s = safe_float_series(df[target_col])
    s.index = df["__period__"]
    if target_is_log:
        y_log = s.copy()
        y_raw = np.exp(y_log)
    else:
        y_raw = s.copy()
        y_log = np.log(y_raw.where(y_raw > 0))
    return y_raw, y_log


def build_exog(df: pd.DataFrame, stream: str, feature_cols: Sequence[str]) -> Tuple[pd.DataFrame, Dict[str, List[str]], Dict[str, str]]:
    idx = df["__period__"]
    exog = pd.DataFrame(index=idx)
    groups: Dict[str, List[str]] = {"time": [], "policy_window": [], "policy_price_signal": [], "policy_interaction": []}
    primary_log: Dict[str, str] = {}

    for col in feature_cols:
        s = safe_float_series(df[col]).copy()
        s.index = idx
        base = snake_name(col)
        group = assign_base_group(stream, col)
        positive_share = (s > 0).mean(skipna=True)
        col_is_log = is_log_col(col)

        level_name = f"{base}__level"
        exog[level_name] = s
        add_group(groups, "core_level", level_name)
        add_group(groups, f"{group}_level", level_name)

        log_s = None
        if col_is_log:
            log_s = s.copy()
            log_name = f"{base}__log"
            exog[log_name] = log_s
            add_group(groups, "core_level", log_name)
            add_group(groups, f"{group}_level", log_name)
            primary_log.setdefault(group, log_name)
        elif positive_share > 0.90:
            log_s = np.log(s.where(s > 0))
            log_name = f"{base}__log"
            exog[log_name] = log_s
            add_group(groups, "core_level", log_name)
            add_group(groups, f"{group}_level", log_name)
            primary_log.setdefault(group, log_name)

        # Lags/diffs for all core variables.
        for lag in [1, 2, 4, 8]:
            nm = f"{base}__lag{lag}"
            exog[nm] = s.shift(lag)
            add_group(groups, "core_lag", nm)
            add_group(groups, f"{group}_lag", nm)
            if log_s is not None:
                nml = f"{base}__log_lag{lag}"
                exog[nml] = log_s.shift(lag)
                add_group(groups, "core_lag", nml)
                add_group(groups, f"{group}_lag", nml)
        for d in [1, 4]:
            nm = f"{base}__diff{d}"
            exog[nm] = s.diff(d)
            add_group(groups, "core_diff", nm)
            add_group(groups, f"{group}_diff", nm)
            if log_s is not None:
                nml = f"{base}__log_diff{d}"
                exog[nml] = log_s.diff(d)
                add_group(groups, "core_diff", nml)
                add_group(groups, f"{group}_diff", nml)

        # Lead price features are kept only for feature sets that explicitly allow leads.
        if ALLOW_LEAD_PRICE_FEATURES_FOR_DISCOVERY and group in {"light_ruc_price", "heavy_ruc_price", "diesel_price", "petrol_price"}:
            for lead in [1, 2, 4]:
                nm = f"{base}__lead{lead}"
                exog[nm] = s.shift(-lead)
                add_group(groups, "core_lead", nm)
                add_group(groups, f"{group}_lead", nm)
                if log_s is not None:
                    nml = f"{base}__log_lead{lead}"
                    exog[nml] = log_s.shift(-lead)
                    add_group(groups, "core_lead", nml)
                    add_group(groups, f"{group}_lead", nml)
                    chg = f"{base}__log_lead{lead}_minus_current"
                    exog[chg] = log_s.shift(-lead) - log_s
                    add_group(groups, "core_lead", chg)
                    add_group(groups, "policy_price_signal", chg)
                    add_group(groups, f"{group}_lead", chg)

    periods = pd.Series(list(exog.index), index=exog.index)
    q = periods.map(lambda p: int(p.quarter))
    exog["time__trend"] = np.arange(len(exog), dtype=float) + 1.0
    exog["time__trend_sq"] = exog["time__trend"] ** 2
    exog["time__post2011"] = periods.map(lambda p: 1.0 if p >= pd.Period("2011Q1", freq="Q-DEC") else 0.0)
    exog["time__post2011_trend"] = exog["time__post2011"] * exog["time__trend"]
    exog["time__post2020"] = periods.map(lambda p: 1.0 if p >= pd.Period("2020Q1", freq="Q-DEC") else 0.0)
    exog["time__covid2020"] = periods.map(lambda p: 1.0 if pd.Period("2020Q1", freq="Q-DEC") <= p <= pd.Period("2020Q4", freq="Q-DEC") else 0.0)
    for qq in [1, 2, 3]:
        nm = f"time__q{qq}"
        exog[nm] = (q == qq).astype(float)
        add_group(groups, "time", nm)
    for nm in ["time__trend", "time__trend_sq", "time__post2011", "time__post2011_trend", "time__post2020", "time__covid2020"]:
        add_group(groups, "time", nm)

    # RUC-specific policy windows and price-shock terms.
    if stream in {"LIGHT_RUC", "HEAVY_RUC"}:
        for key, (start, end) in POLICY_WINDOWS.items():
            nm = f"policy__{key}"
            exog[nm] = periods.map(lambda p, s=start, e=end: 1.0 if period_between(p, s, e) else 0.0)
            add_group(groups, "policy_window", nm)

        ruc_key = "light_ruc_price" if stream == "LIGHT_RUC" else "heavy_ruc_price"
        ruc_log = primary_log.get(ruc_key)
        diesel_log = primary_log.get("diesel_price")
        for label, src in [("ruc", ruc_log), ("diesel", diesel_log)]:
            if src is not None and src in exog.columns:
                z = exog[src]
                exog[f"policy__{label}_log_change_1"] = z.diff(1)
                exog[f"policy__{label}_log_change_4"] = z.diff(4)
                exog[f"policy__{label}_price_jump_up_1"] = z.diff(1).clip(lower=0)
                exog[f"policy__{label}_price_cut_1"] = (-z.diff(1)).clip(lower=0)
                exog[f"policy__abs_{label}_log_change_1"] = z.diff(1).abs()
                for nm in [f"policy__{label}_log_change_1", f"policy__{label}_log_change_4", f"policy__{label}_price_jump_up_1", f"policy__{label}_price_cut_1", f"policy__abs_{label}_log_change_1"]:
                    add_group(groups, "policy_price_signal", nm)
                for lag in [1, 2, 4]:
                    for src2 in [f"policy__{label}_price_jump_up_1", f"policy__{label}_price_cut_1", f"policy__abs_{label}_log_change_1"]:
                        nm = f"{src2}_lag{lag}"
                        exog[nm] = exog[src2].shift(lag)
                        add_group(groups, "policy_price_signal", nm)

        for policy_col in groups.get("policy_window", []):
            for price_col in groups.get("policy_price_signal", [])[:10]:
                nm = f"interact__{policy_col.replace('policy__','')}__{price_col.replace('policy__','')}"
                exog[nm] = exog[policy_col] * exog[price_col]
                add_group(groups, "policy_interaction", nm)

    return exog.sort_index(), groups, primary_log


def pick_group(groups: Dict[str, List[str]], group: str, prefer_log: bool = True, max_n: int = 2) -> List[str]:
    candidates = groups.get(group, [])
    if not candidates:
        return []
    if prefer_log:
        log_cands = [c for c in candidates if "__log" in c and "__diff" not in c and "__lag" not in c and "__lead" not in c]
        if log_cands:
            return log_cands[:max_n]
    clean = [c for c in candidates if "__diff" not in c and "__lag" not in c and "__lead" not in c]
    return clean[:max_n]


def feature_names_for_set(sd: StreamData, feature_set: str, include_target_lags: bool) -> List[str]:
    g = sd.groups
    stream = sd.stream

    time_base = [c for c in g.get("time", []) if c in sd.exog.columns]
    quarter_time = [c for c in time_base if c.startswith("time__q")]

    schiff: List[str] = []
    if stream == "PED":
        schiff += pick_group(g, "petrol_price_level", max_n=1)
        schiff += pick_group(g, "gdp_pc_level", max_n=1) or pick_group(g, "gdp_level", max_n=1)
        schiff += pick_group(g, "unemployment_level", prefer_log=False, max_n=1)
        schiff += ["time__trend", "time__post2011_trend", "time__post2020", "time__covid2020"] + quarter_time
    elif stream == "LIGHT_RUC":
        schiff += pick_group(g, "diesel_price_level", max_n=1)
        schiff += pick_group(g, "light_ruc_price_level", max_n=1)
        # lagged light RUC price is part of the structural RUC intuition.
        schiff += [c for c in g.get("light_ruc_price_lag", []) if "lag1" in c and "__log" in c][:1]
        schiff += pick_group(g, "gdp_level", max_n=1)
        schiff += ["time__post2020"] + quarter_time
    elif stream == "HEAVY_RUC":
        schiff += pick_group(g, "gdp_level", max_n=1)
        schiff += pick_group(g, "heavy_ruc_price_level", max_n=1)
        if feature_set == "schiff" and ALLOW_LEAD_PRICE_FEATURES_FOR_DISCOVERY:
            schiff += [c for c in g.get("heavy_ruc_price_lead", []) if "lead1" in c and "__log" in c][:1]
        schiff += quarter_time

    schiff = [c for c in dedupe(schiff) if c in sd.exog.columns]

    if feature_set in {"schiff", "schiff_no_lead"}:
        cols = schiff
        if feature_set == "schiff_no_lead":
            cols = [c for c in cols if "lead" not in c]
    elif feature_set == "struct":
        cols = [c for c in g.get("core_level", []) if "__diff" not in c and "__lag" not in c and "__lead" not in c] + time_base
    elif feature_set == "diff":
        cols = g.get("core_level", []) + g.get("core_diff", []) + time_base
        cols = [c for c in cols if "__lead" not in c]
    elif feature_set == "rich":
        cols = g.get("core_level", []) + g.get("core_diff", []) + g.get("core_lag", []) + time_base
        cols = [c for c in cols if "__lead" not in c]
    elif feature_set == "dynamic_pruned":
        cols = schiff + g.get("core_diff", []) + g.get("core_lag", []) + g.get("policy_window", []) + g.get("policy_price_signal", [])
        cols = [c for c in cols if "lead" not in c]
    elif feature_set == "dynamic_no_leads":
        cols = g.get("core_level", []) + g.get("core_diff", []) + g.get("core_lag", []) + time_base + g.get("policy_window", []) + g.get("policy_price_signal", []) + g.get("policy_interaction", [])
        cols = [c for c in cols if "lead" not in c]
    elif feature_set == "dynamic_rich":
        cols = g.get("core_level", []) + g.get("core_diff", []) + g.get("core_lag", []) + g.get("core_lead", []) + time_base + g.get("policy_window", []) + g.get("policy_price_signal", []) + g.get("policy_interaction", [])
    elif feature_set == "price_distributed_lags":
        price_groups = []
        for key in ["petrol_price", "diesel_price", "light_ruc_price", "heavy_ruc_price"]:
            price_groups += g.get(f"{key}_level", []) + g.get(f"{key}_diff", []) + g.get(f"{key}_lag", []) + g.get(f"{key}_lead", [])
        cols = price_groups + g.get("gdp_level", []) + g.get("gdp_lag", []) + g.get("gdp_pc_level", []) + g.get("unemployment_level", []) + time_base + g.get("policy_window", []) + g.get("policy_price_signal", [])
        if "no_leads" in feature_set:
            cols = [c for c in cols if "lead" not in c]
    else:
        raise ValueError(f"Unknown feature_set: {feature_set}")

    cols = [c for c in dedupe(cols) if c in sd.exog.columns]
    if include_target_lags:
        cols += ["target__lag1", "target__lag2", "target__lag4", "target__diff1", "target__diff4", "target__roll4_mean", "target__roll8_mean"]
    return dedupe(cols)


def target_lag_features(period: pd.Period, y_hist: Dict[pd.Period, float]) -> Dict[str, float]:
    def get(p: pd.Period) -> float:
        v = y_hist.get(p, np.nan)
        return float(v) if np.isfinite(v) else np.nan
    y1, y2, y4, y5 = get(period - 1), get(period - 2), get(period - 4), get(period - 5)
    last4 = np.array([get(period - i) for i in range(1, 5)], dtype=float)
    last8 = np.array([get(period - i) for i in range(1, 9)], dtype=float)
    return {
        "target__lag1": y1,
        "target__lag2": y2,
        "target__lag4": y4,
        "target__diff1": y1 - y2 if np.isfinite(y1) and np.isfinite(y2) else np.nan,
        "target__diff4": y1 - y5 if np.isfinite(y1) and np.isfinite(y5) else np.nan,
        "target__roll4_mean": float(np.nanmean(last4)) if np.isfinite(last4).sum() >= 2 else np.nan,
        "target__roll8_mean": float(np.nanmean(last8)) if np.isfinite(last8).sum() >= 4 else np.nan,
    }


def build_feature_row(period: pd.Period, sd: StreamData, y_hist: Dict[pd.Period, float], feature_names: Sequence[str], include_target_lags: bool) -> Dict[str, float]:
    row: Dict[str, float] = {}
    if period in sd.exog.index:
        row.update(sd.exog.loc[period].to_dict())
    if include_target_lags:
        row.update(target_lag_features(period, y_hist))
    return {f: row.get(f, np.nan) for f in feature_names}

# =============================================================================
# 4. Candidate definitions
# =============================================================================

@dataclass(frozen=True)
class CandidateConfig:
    stream: str
    name: str
    model_kind: str
    params_json: str
    window: Optional[int]
    feature_set: str
    include_target_lags: bool
    family_tag: str
    min_train_quarters: int = MIN_TRAIN_QUARTERS


LINEAR_KINDS = {"ols", "ridge", "bayesian_ridge", "elastic_net", "huber"}
RESIDUAL_KINDS = {"resid_ridge", "resid_bayesian_ridge", "resid_elastic_net", "resid_gbr", "resid_huber"}


def make_estimator(kind: str, params: Dict[str, Any]):
    if kind == "ols":
        return LinearRegression(**params)
    if kind == "ridge":
        return Ridge(**params)
    if kind == "bayesian_ridge":
        return BayesianRidge(**params)
    if kind == "elastic_net":
        return ElasticNet(**params)
    if kind == "huber":
        return HuberRegressor(**params)
    if kind == "gbr":
        return GradientBoostingRegressor(**params)
    if kind == "extra_trees":
        return ExtraTreesRegressor(**params)
    if kind == "random_forest":
        return RandomForestRegressor(**params)
    if kind == "xgboost":
        return XGBRegressor(**params)
    raise ValueError(kind)


def base_kind_for_resid(kind: str) -> str:
    return {
        "resid_ridge": "ridge",
        "resid_bayesian_ridge": "bayesian_ridge",
        "resid_elastic_net": "elastic_net",
        "resid_gbr": "gbr",
        "resid_huber": "huber",
    }[kind]


def make_pipeline(kind: str, params: Dict[str, Any]) -> Pipeline:
    est = make_estimator(kind, params)
    if kind in LINEAR_KINDS:
        return Pipeline([("imputer", SimpleImputer(strategy="median")), ("scaler", StandardScaler()), ("model", est)])
    return Pipeline([("imputer", SimpleImputer(strategy="median")), ("model", est)])


def clean_label_params(params: Dict[str, Any]) -> str:
    if not params:
        return ""
    parts = []
    for k in sorted(params):
        v = params[k]
        if k == "random_state" or k == "n_jobs" or k == "loss" or k == "max_iter" or k == "subsample":
            continue
        if isinstance(v, float):
            s = f"{v:g}".replace(".", "_")
        else:
            s = str(v).replace(".", "_")
        parts.append(f"{k}{s}")
    return "_" + "_".join(parts) if parts else ""


def add_candidate(configs: List[CandidateConfig], stream: str, kind: str, params: Dict[str, Any], window: Optional[int], feature_set: str, ylags: bool, tag: str) -> None:
    w = "exp" if window is None else str(window)
    params_label = clean_label_params(params)
    name = f"{stream}__{feature_set}__{tag}{params_label}__{'ylag' if ylags else 'noylag'}__w{w}"
    configs.append(CandidateConfig(
        stream=stream,
        name=name,
        model_kind=kind,
        params_json=json.dumps(params, sort_keys=True),
        window=window,
        feature_set=feature_set,
        include_target_lags=ylags,
        family_tag=tag,
    ))


def generate_targeted_candidates(stream: str) -> List[CandidateConfig]:
    configs: List[CandidateConfig] = []

    # Pure Schiff benchmarks: always include expanding and sliding windows.
    for window in [None, 40, 52, 64]:
        add_candidate(configs, stream, "ols", {}, window, "schiff", False, "SCHIFF_OLS")
        add_candidate(configs, stream, "ols", {}, window, "schiff_no_lead", False, "SCHIFF_OLS")

    # Stream-specific candidate grid. This intentionally preserves the older PDF-style
    # families and the latest Heavy RUC dynamic-no-lead family.
    if stream == "PED":
        # Old PDF-style PED family: shallow GBM, diff features, target lags, W40/W64.
        for fs, ylags, windows in [
            ("diff", True, [40, 64]),
            ("rich", False, [40, 64]),
            ("struct", False, [None, 40, 64]),
            ("schiff", True, [40, 64]),
            ("schiff_no_lead", True, [40, 64]),
        ]:
            for window in windows:
                for n in [150, 250, 400, 650]:
                    for lr in [0.03, 0.05, 0.06, 0.08]:
                        for depth in [1, 2]:
                            params = {"n_estimators": n, "learning_rate": lr, "max_depth": depth, "subsample": 0.85, "loss": "squared_error", "random_state": RANDOM_SEED}
                            add_candidate(configs, stream, "gbr", params, window, fs, ylags, "GBR")
        # Regularised linear stabilisers.
        for fs in ["schiff", "diff", "rich"]:
            for window in [40, 52, 64, None]:
                for ylags in [False, True]:
                    for alpha in [0.001, 0.01, 0.1, 1.0, 10.0]:
                        add_candidate(configs, stream, "ridge", {"alpha": alpha, "random_state": RANDOM_SEED}, window, fs, ylags, "Ridge")
                    add_candidate(configs, stream, "bayesian_ridge", {}, window, fs, ylags, "BayesianRidge")
                    for alpha in [0.005, 0.01, 0.05]:
                        for l1 in [0.2, 0.5, 0.8]:
                            add_candidate(configs, stream, "elastic_net", {"alpha": alpha, "l1_ratio": l1, "max_iter": 50000, "random_state": RANDOM_SEED}, window, fs, ylags, "Elastic")

    elif stream == "LIGHT_RUC":
        # Light RUC: focus on 40Q-ish windows, elastic regularisation, residual correction and policy/price lags.
        for fs in ["schiff", "dynamic_pruned", "dynamic_no_leads", "price_distributed_lags", "diff", "rich"]:
            for window in [36, 40, 44, 52]:
                for ylags in [False, True]:
                    for alpha in [0.001, 0.005, 0.01, 0.05, 0.1, 0.2]:
                        for l1 in [0.15, 0.3, 0.5, 0.7, 0.85]:
                            add_candidate(configs, stream, "elastic_net", {"alpha": alpha, "l1_ratio": l1, "max_iter": 50000, "random_state": RANDOM_SEED}, window, fs, ylags, "Elastic")
                    for alpha in [0.01, 0.1, 1.0, 10.0, 100.0]:
                        add_candidate(configs, stream, "ridge", {"alpha": alpha, "random_state": RANDOM_SEED}, window, fs, ylags, "Ridge")
                    add_candidate(configs, stream, "bayesian_ridge", {}, window, fs, ylags, "BayesianRidge")
                # Residual correction around Schiff base.
                for ylags in [False, True]:
                    for n in [80, 150, 250, 400]:
                        for depth in [1, 2]:
                            for lr in [0.03, 0.05, 0.06, 0.08]:
                                params = {"n_estimators": n, "learning_rate": lr, "max_depth": depth, "subsample": 0.85, "loss": "squared_error", "random_state": RANDOM_SEED}
                                add_candidate(configs, stream, "resid_gbr", params, window, fs, ylags, "RESID_SchiffResidual_GBR")
                    for alpha in [0.01, 0.05, 0.1, 1.0, 10.0]:
                        add_candidate(configs, stream, "resid_ridge", {"alpha": alpha, "random_state": RANDOM_SEED}, window, fs, ylags, "RESID_SchiffResidual_Ridge")
                    for alpha in [0.005, 0.01, 0.05]:
                        for l1 in [0.3, 0.5, 0.7, 0.85]:
                            add_candidate(configs, stream, "resid_elastic_net", {"alpha": alpha, "l1_ratio": l1, "max_iter": 50000, "random_state": RANDOM_SEED}, window, fs, ylags, "RESID_SchiffResidual_Elastic")
                # Direct GBM challengers.
                for ylags in [False, True]:
                    for n in [80, 150, 250, 400]:
                        for depth in [1, 2, 3]:
                            for lr in [0.03, 0.05, 0.06, 0.08]:
                                params = {"n_estimators": n, "learning_rate": lr, "max_depth": depth, "subsample": 0.85, "loss": "squared_error", "random_state": RANDOM_SEED}
                                add_candidate(configs, stream, "gbr", params, window, fs, ylags, "GBR")

    elif stream == "HEAVY_RUC":
        # Heavy RUC: preserve latest winning dynamic-no-leads family and adjacent variants.
        for fs in ["dynamic_no_leads", "price_distributed_lags", "dynamic_pruned", "diff", "rich", "schiff_no_lead", "schiff"]:
            for window in [40, 52, 64]:
                for ylags in [False, True]:
                    for n in [150, 250, 400, 650]:
                        for depth in [1, 2]:
                            for lr in [0.03, 0.05, 0.06, 0.08]:
                                params = {"n_estimators": n, "learning_rate": lr, "max_depth": depth, "subsample": 0.85, "loss": "squared_error", "random_state": RANDOM_SEED}
                                add_candidate(configs, stream, "gbr", params, window, fs, ylags, "GBR")
                    for alpha in [0.001, 0.01, 0.05, 0.1, 1.0, 10.0]:
                        add_candidate(configs, stream, "ridge", {"alpha": alpha, "random_state": RANDOM_SEED}, window, fs, ylags, "Ridge")
                    for alpha in [0.005, 0.01, 0.05, 0.1]:
                        for l1 in [0.2, 0.5, 0.8]:
                            add_candidate(configs, stream, "elastic_net", {"alpha": alpha, "l1_ratio": l1, "max_iter": 50000, "random_state": RANDOM_SEED}, window, fs, ylags, "Elastic")
                    add_candidate(configs, stream, "bayesian_ridge", {}, window, fs, ylags, "BayesianRidge")
                # A few residual corrections too.
                for ylags in [False, True]:
                    for n in [80, 150, 250]:
                        for depth in [1, 2]:
                            params = {"n_estimators": n, "learning_rate": 0.05, "max_depth": depth, "subsample": 0.85, "loss": "squared_error", "random_state": RANDOM_SEED}
                            add_candidate(configs, stream, "resid_gbr", params, window, fs, ylags, "RESID_SchiffResidual_GBR")

    # Optional random local neighbours around the same known-good regions. This is not broad HPO;
    # it samples small perturbations around already plausible families.
    if ADD_RANDOM_LOCAL_NEIGHBOURS:
        rng = np.random.default_rng(RANDOM_SEED + hash(stream) % 10_000)
        fs_pool = {
            "PED": ["diff", "rich", "schiff", "schiff_no_lead"],
            "LIGHT_RUC": ["dynamic_pruned", "dynamic_no_leads", "price_distributed_lags", "schiff"],
            "HEAVY_RUC": ["dynamic_no_leads", "price_distributed_lags", "schiff_no_lead"],
        }[stream]
        window_pool = {
            "PED": [40, 52, 64, None],
            "LIGHT_RUC": [36, 40, 44, 52],
            "HEAVY_RUC": [40, 52, 64],
        }[stream]
        for i in range(N_RANDOM_LOCAL_NEIGHBOURS_PER_STREAM):
            fs = str(rng.choice(fs_pool))
            window = rng.choice(window_pool)
            if isinstance(window, np.integer):
                window = int(window)
            ylags = bool(rng.choice([False, True]))
            depth = int(rng.choice([1, 2, 3 if stream == "LIGHT_RUC" else 2]))
            n = int(rng.choice([120, 180, 250, 320, 450, 600]))
            lr = float(rng.choice([0.025, 0.035, 0.045, 0.055, 0.07, 0.09]))
            params = {"n_estimators": n, "learning_rate": lr, "max_depth": depth, "subsample": float(rng.choice([0.75, 0.85, 0.95, 1.0])), "loss": "squared_error", "random_state": RANDOM_SEED + i}
            add_candidate(configs, stream, "gbr", params, window, fs, ylags, "GBRLocal")

    # Deduplicate.
    unique: Dict[str, CandidateConfig] = {}
    for c in configs:
        unique[c.name] = c
    return list(unique.values())

# =============================================================================
# 5. Rolling-origin evaluation
# =============================================================================

def valid_periods(sd: StreamData) -> List[pd.Period]:
    idx = set(sd.y_raw.dropna().index).intersection(set(sd.exog.index))
    return sorted(idx, key=period_sort_value)


def build_training_matrix(sd: StreamData, train_periods: Sequence[pd.Period], feature_names: Sequence[str], include_target_lags: bool) -> Tuple[pd.DataFrame, pd.Series]:
    y_hist = {p: float(sd.y_log.loc[p]) for p in sd.y_log.index if pd.notna(sd.y_log.loc[p])}
    rows, ys, used = [], [], []
    for p in train_periods:
        if p not in sd.y_log.index or pd.isna(sd.y_log.loc[p]):
            continue
        rows.append(build_feature_row(p, sd, y_hist, feature_names, include_target_lags))
        ys.append(float(sd.y_log.loc[p]))
        used.append(p)
    X = pd.DataFrame(rows, index=used).reindex(columns=feature_names)
    y = pd.Series(ys, index=used, name="y_log")
    return X, y


def fit_model(cfg: CandidateConfig, X: pd.DataFrame, y: pd.Series, X_base: Optional[pd.DataFrame] = None):
    params = json.loads(cfg.params_json)
    if cfg.model_kind in RESIDUAL_KINDS:
        if X_base is None:
            raise ValueError("Residual model requires X_base")
        base = make_pipeline("ols", {})
        base.fit(X_base, y)
        resid = y.values - base.predict(X_base)
        rk = base_kind_for_resid(cfg.model_kind)
        rm = make_pipeline(rk, params)
        rm.fit(X, resid)
        return {"kind": "residual", "base": base, "resid": rm}
    pipe = make_pipeline(cfg.model_kind, params)
    pipe.fit(X, y)
    return pipe


def predict_model(model: Any, X_pred: pd.DataFrame, X_pred_base: Optional[pd.DataFrame] = None) -> float:
    if isinstance(model, dict) and model.get("kind") == "residual":
        if X_pred_base is None:
            raise ValueError("Residual model requires X_pred_base")
        return float(model["base"].predict(X_pred_base)[0] + model["resid"].predict(X_pred)[0])
    return float(model.predict(X_pred)[0])


def evaluate_candidate(sd: StreamData, cfg: CandidateConfig) -> pd.DataFrame:
    feature_names = feature_names_for_set(sd, cfg.feature_set, cfg.include_target_lags)
    feature_names = [f for f in feature_names if f in sd.exog.columns or f.startswith("target__")]
    if not feature_names:
        return pd.DataFrame()

    base_features = feature_names_for_set(sd, "schiff_no_lead" if cfg.stream in {"LIGHT_RUC", "HEAVY_RUC"} else "schiff", False)
    periods = valid_periods(sd)
    period_set = set(periods)
    origins = []
    for origin in periods:
        train_periods = [p for p in periods if period_sort_value(p) <= period_sort_value(origin)]
        if len(train_periods) >= cfg.min_train_quarters and any((origin + h) in period_set for h in range(1, MAX_HORIZON + 1)):
            origins.append(origin)
    origins = origins[::ORIGIN_STRIDE]
    if MAX_ORIGINS_PER_STREAM is not None:
        origins = origins[-MAX_ORIGINS_PER_STREAM:]

    records: List[Dict[str, Any]] = []
    for origin in origins:
        train_periods = [p for p in periods if period_sort_value(p) <= period_sort_value(origin)]
        if cfg.window is not None:
            train_periods = train_periods[-int(cfg.window):]
        if len(train_periods) < cfg.min_train_quarters:
            continue

        X, y = build_training_matrix(sd, train_periods, feature_names, cfg.include_target_lags)
        mask = y.notna()
        X, y = X.loc[mask], y.loc[mask]
        if len(X) < max(20, int(cfg.min_train_quarters * 0.60)):
            continue
        all_na_cols = [c for c in X.columns if X[c].isna().all()]
        if all_na_cols:
            X = X.copy()
            X[all_na_cols] = 0.0

        X_base = None
        base_all_na: List[str] = []
        if cfg.model_kind in RESIDUAL_KINDS:
            X_base, _ = build_training_matrix(sd, train_periods, base_features, False)
            X_base = X_base.reindex(index=X.index)
            base_all_na = [c for c in X_base.columns if X_base[c].isna().all()]
            if base_all_na:
                X_base = X_base.copy()
                X_base[base_all_na] = 0.0

        try:
            model = fit_model(cfg, X, y, X_base)
        except Exception:
            continue

        y_hist = {p: float(sd.y_log.loc[p]) for p in sd.y_log.index if pd.notna(sd.y_log.loc[p]) and period_sort_value(p) <= period_sort_value(origin)}
        for h in range(1, MAX_HORIZON + 1):
            target_p = origin + h
            if target_p not in period_set or target_p not in sd.y_raw.index or pd.isna(sd.y_raw.loc[target_p]):
                continue
            row = build_feature_row(target_p, sd, y_hist, feature_names, cfg.include_target_lags)
            Xp = pd.DataFrame([row]).reindex(columns=feature_names)
            if all_na_cols:
                Xp[all_na_cols] = 0.0
            Xpb = None
            if cfg.model_kind in RESIDUAL_KINDS:
                brow = build_feature_row(target_p, sd, y_hist, base_features, False)
                Xpb = pd.DataFrame([brow]).reindex(columns=base_features)
                if base_all_na:
                    Xpb[base_all_na] = 0.0
            try:
                pred_log = predict_model(model, Xp, Xpb)
            except Exception:
                pred_log = np.nan
            pred = safe_exp(pred_log)
            actual = float(sd.y_raw.loc[target_p])
            if np.isfinite(pred_log):
                y_hist[target_p] = pred_log
            records.append({
                "stream": cfg.stream,
                "model": cfg.name,
                "source_family": "bespoke_residual_correction" if cfg.model_kind in RESIDUAL_KINDS else "bespoke_sklearn",
                "model_kind": base_kind_for_resid(cfg.model_kind) if cfg.model_kind in RESIDUAL_KINDS else cfg.model_kind,
                "feature_set": cfg.feature_set,
                "family_tag": cfg.family_tag,
                "include_target_lags": cfg.include_target_lags,
                "window": "expanding" if cfg.window is None else cfg.window,
                "params_json": cfg.params_json,
                "origin": str(origin),
                "target_period": str(target_p),
                "horizon": h,
                "actual": actual,
                "pred": pred,
                "actual_log": float(sd.y_log.loc[target_p]) if pd.notna(sd.y_log.loc[target_p]) else np.nan,
                "pred_log": pred_log,
            })
    return pd.DataFrame(records)

# =============================================================================
# 6. Summaries, annual aggregation and stress tests
# =============================================================================

def summarise_quarterly(preds: pd.DataFrame) -> pd.DataFrame:
    if preds.empty:
        return pd.DataFrame()
    rows = []
    group_cols = ["stream", "model"]
    for (stream, model), g in preds.groupby(group_cols, sort=False):
        row = {
            "stream": stream,
            "model": model,
            "n_quarterly_pairs": int(len(g)),
            "n_origins": int(g["origin"].nunique()),
            "quarterly_mape": mape(g["actual"], g["pred"]),
            "quarterly_bias_pct": bias_pct(g["actual"], g["pred"]),
            "quarterly_p90_ape": p90_ape(g["actual"], g["pred"]),
            "quarterly_rmse": rmse(g["actual"], g["pred"]),
        }
        for h in range(1, MAX_HORIZON + 1):
            gh = g[g["horizon"] == h]
            row[f"mape_h{h:02d}"] = mape(gh["actual"], gh["pred"]) if not gh.empty else np.nan
        for label, lo, hi in [("h01_04", 1, 4), ("h05_08", 5, 8), ("h09_12", 9, 12)]:
            gh = g[g["horizon"].between(lo, hi)]
            row[f"mape_{label}"] = mape(gh["actual"], gh["pred"]) if not gh.empty else np.nan
        for c in ["source_family", "model_kind", "feature_set", "family_tag", "include_target_lags", "window", "params_json"]:
            row[c] = g[c].iloc[0] if c in g.columns else ""
        rows.append(row)
    return pd.DataFrame(rows)


def annualise(preds: pd.DataFrame) -> pd.DataFrame:
    if preds.empty:
        return pd.DataFrame()
    dfp = preds.copy()
    dfp["target_period_obj"] = dfp["target_period"].map(lambda s: pd.Period(str(s), freq="Q-DEC"))
    dfp["june_year"] = dfp["target_period_obj"].map(june_year_from_period)
    rows = []
    group_cols = ["stream", "model", "origin", "june_year"]
    for (stream, model, origin, jy), g in dfp.groupby(group_cols):
        if g["target_period_obj"].nunique() < 4:
            continue
        if stream == "PED":
            actual, pred = float(g["actual"].mean()), float(g["pred"].mean())
        else:
            actual, pred = float(g["actual"].sum()), float(g["pred"].sum())
        rows.append({"stream": stream, "model": model, "origin": origin, "june_year": int(jy), "actual": actual, "pred": pred, "n_quarters": int(g["target_period_obj"].nunique())})
    return pd.DataFrame(rows)


def summarise_annual(annual_preds: pd.DataFrame) -> pd.DataFrame:
    if annual_preds.empty:
        return pd.DataFrame()
    rows = []
    for (stream, model), g in annual_preds.groupby(["stream", "model"], sort=False):
        rows.append({
            "stream": stream,
            "model": model,
            "n_annual_pairs": int(len(g)),
            "annual_mape": mape(g["actual"], g["pred"]),
            "annual_bias_pct": bias_pct(g["actual"], g["pred"]),
            "annual_p90_ape": p90_ape(g["actual"], g["pred"]),
        })
    return pd.DataFrame(rows)


def build_final_summary(preds: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    qsum = summarise_quarterly(preds)
    annual_preds = annualise(preds)
    asum = summarise_annual(annual_preds)
    final = qsum.merge(asum, on=["stream", "model"], how="left")
    # stream-specific score, lower is better.
    score_parts = []
    for stream, g in final.groupby("stream"):
        x = g.copy()
        if stream == "PED":
            score = 0.55 * x["quarterly_mape"].rank(pct=True) + 0.25 * x["annual_mape"].rank(pct=True) + 0.10 * x["quarterly_p90_ape"].rank(pct=True) + 0.10 * x["quarterly_bias_pct"].abs().rank(pct=True)
        elif stream == "LIGHT_RUC":
            stress_col = "mape_h09_12" if "mape_h09_12" in x else "quarterly_p90_ape"
            score = 0.35 * x["quarterly_mape"].rank(pct=True) + 0.35 * x["annual_mape"].rank(pct=True) + 0.15 * x[stress_col].rank(pct=True) + 0.15 * x["quarterly_bias_pct"].abs().rank(pct=True)
        else:
            score = 0.35 * x["quarterly_mape"].rank(pct=True) + 0.45 * x["annual_mape"].rank(pct=True) + 0.10 * x["quarterly_p90_ape"].rank(pct=True) + 0.10 * x["quarterly_bias_pct"].abs().rank(pct=True)
        x["governance_score"] = score
        score_parts.append(x)
    final = pd.concat(score_parts, ignore_index=True)
    final = final.sort_values(["stream", "governance_score", "quarterly_mape"], na_position="last").reset_index(drop=True)
    return final, annual_preds

# =============================================================================
# 7. Ensembles
# =============================================================================

def common_matrix(preds: pd.DataFrame, stream: str, model_names: Sequence[str]) -> Tuple[pd.DataFrame, pd.Series, pd.DataFrame]:
    key_cols = ["stream", "origin", "target_period", "horizon"]
    sub = preds[(preds["stream"] == stream) & (preds["model"].isin(model_names))].copy()
    if sub.empty:
        return pd.DataFrame(), pd.Series(dtype=float), pd.DataFrame()
    pivot = sub.pivot_table(index=key_cols, columns="model", values="pred", aggfunc="first")
    actual = sub.groupby(key_cols)["actual"].first()
    pivot = pivot.reindex(columns=list(model_names))
    common = pivot.dropna(axis=0, how="any")
    if common.empty:
        return pd.DataFrame(), pd.Series(dtype=float), pd.DataFrame()
    actual = actual.loc[common.index]
    keys = pd.DataFrame(index=common.index).reset_index()
    return common.reset_index(drop=True), actual.reset_index(drop=True), keys


def optimise_convex_mape(P: np.ndarray, y: np.ndarray) -> np.ndarray:
    P = np.asarray(P, dtype=float)
    y = np.asarray(y, dtype=float)
    mask = np.isfinite(y) & (np.abs(y) > 1e-12) & np.isfinite(P).all(axis=1)
    P, y = P[mask], y[mask]
    n, k = P.shape
    if n == 0 or k == 0:
        return np.ones(k) / max(k, 1)
    c = np.zeros(k + n)
    c[k:] = 1.0 / np.maximum(np.abs(y), 1e-12)
    A_ub = np.zeros((2 * n, k + n))
    b_ub = np.zeros(2 * n)
    A_ub[:n, :k] = P
    A_ub[:n, k:] = -np.eye(n)
    b_ub[:n] = y
    A_ub[n:, :k] = -P
    A_ub[n:, k:] = -np.eye(n)
    b_ub[n:] = -y
    A_eq = np.zeros((1, k + n))
    A_eq[0, :k] = 1.0
    b_eq = np.array([1.0])
    bounds = [(0.0, 1.0)] * k + [(0.0, None)] * n
    res = linprog(c=c, A_ub=A_ub, b_ub=b_ub, A_eq=A_eq, b_eq=b_eq, bounds=bounds, method="highs")
    if res.success:
        w = np.maximum(np.asarray(res.x[:k], dtype=float), 0)
        if w.sum() > 0:
            return w / w.sum()
    def obj(w):
        return np.mean(np.abs(P @ w - y) / np.maximum(np.abs(y), 1e-12))
    res2 = minimize(obj, x0=np.ones(k) / k, bounds=[(0.0, 1.0)] * k, constraints=[{"type": "eq", "fun": lambda w: np.sum(w) - 1.0}], method="SLSQP")
    if res2.success:
        w = np.maximum(np.asarray(res2.x, dtype=float), 0)
        if w.sum() > 0:
            return w / w.sum()
    return np.ones(k) / k


def make_ensemble_frame(keys: pd.DataFrame, y: pd.Series, pred: np.ndarray, stream: str, name: str, kind: str, params: Dict[str, Any]) -> pd.DataFrame:
    out = keys.copy()
    out["stream"] = stream
    out["model"] = name
    out["source_family"] = "posthoc_ensemble"
    out["model_kind"] = kind
    out["feature_set"] = "ensemble"
    out["family_tag"] = kind
    out["include_target_lags"] = "mixed"
    out["window"] = "mixed"
    out["params_json"] = json.dumps(params, sort_keys=True)
    out["actual"] = y.values
    out["pred"] = pred
    out["actual_log"] = np.log(np.maximum(out["actual"].astype(float), 1e-12))
    out["pred_log"] = np.log(np.maximum(out["pred"].astype(float), 1e-12))
    return out


def build_ensembles(preds: pd.DataFrame, summary: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    ens_frames = []
    weight_rows = []
    for stream in STREAM_ORDER:
        ssum = summary[summary["stream"] == stream].copy()
        if ssum.empty:
            continue
        # Ensemble pool includes top by several criteria and pure Schiff models.
        candidates = []
        for sort_cols in [["quarterly_mape"], ["annual_mape"], ["quarterly_p90_ape"], ["governance_score"]]:
            candidates.extend(ssum.sort_values(sort_cols, na_position="last")["model"].head(TOP_MODELS_FOR_ENSEMBLES).tolist())
        schiff = ssum[ssum["family_tag"].astype(str).str.contains("SCHIFF_OLS", case=False, na=False)]
        candidates.extend(schiff.sort_values(["quarterly_mape"], na_position="last")["model"].head(4).tolist())
        models = dedupe(candidates)[:max(TOP_MODELS_FOR_ENSEMBLES, 8)]
        if len(models) < 2:
            continue

        # Static convex.
        P, y, keys = common_matrix(preds, stream, models)
        if not P.empty:
            w = optimise_convex_mape(P.values, y.values)
            pred = P.values @ w
            nm = f"{stream}__solver_static_convex_top{len(models)}"
            ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, "solver_static_convex", {"base_models": models, "weights": dict(zip(models, [float(x) for x in w])), "method": "static_convex_mape"}))
            for m, weight in zip(models, w):
                if abs(weight) > 1e-10:
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": m, "weight": float(weight), "method": "static_convex_mape", "origin": ""})

        # Prequential convex.
        P, y, keys = common_matrix(preds, stream, models)
        if not P.empty:
            dfm = keys.copy()
            for m in models:
                dfm[m] = P[m].values
            dfm["actual"] = y.values
            dfm["origin_period"] = dfm["origin"].map(lambda s: pd.Period(str(s), freq="Q-DEC"))
            dfm["origin_sort"] = dfm["origin_period"].map(period_sort_value)
            dfm = dfm.sort_values(["origin_sort", "horizon", "target_period"]).reset_index(drop=True)
            preds_out = []
            eq_w = np.ones(len(models)) / len(models)
            for _, row in dfm.iterrows():
                hist = dfm[dfm["origin_sort"] < row["origin_sort"]]
                if len(hist) >= 40:
                    w = optimise_convex_mape(hist[models].values, hist["actual"].values)
                else:
                    w = eq_w
                preds_out.append(float(row[models].values.astype(float) @ w))
                for m, weight in zip(models, w):
                    if abs(weight) > 1e-10:
                        weight_rows.append({"stream": stream, "ensemble": f"{stream}__solver_preq_convex_top{len(models)}", "component_model": m, "weight": float(weight), "method": "prequential_convex_mape", "origin": row["origin"]})
            nm = f"{stream}__solver_preq_convex_top{len(models)}"
            ens_frames.append(make_ensemble_frame(keys, y, np.asarray(preds_out), stream, nm, "solver_preq_convex", {"base_models": models, "method": "prequential_convex_mape"}))

        # Top-k mean/median/trimmed.
        for k in [3, 5, 8, 10]:
            submodels = models[:k]
            if len(submodels) < 2:
                continue
            P, y, keys = common_matrix(preds, stream, submodels)
            if P.empty:
                continue
            arr = P.values
            for method in ["mean", "median", "trimmed_mean"]:
                if method == "mean":
                    pred = arr.mean(axis=1)
                elif method == "median":
                    pred = np.median(arr, axis=1)
                else:
                    if arr.shape[1] >= 4:
                        pred = np.sort(arr, axis=1)[:, 1:-1].mean(axis=1)
                    else:
                        pred = arr.mean(axis=1)
                nm = f"{stream}__top{k}_{method}"
                ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, f"topk_{method}", {"models": submodels, "method": method}))

        # Fixed blends with best pure Schiff.
        if not schiff.empty:
            schiff_model = schiff.sort_values(["quarterly_mape"], na_position="last")["model"].iloc[0]
            challenger_models = [m for m in models[:8] if m != schiff_model]
            for ch in challenger_models:
                P, y, keys = common_matrix(preds, stream, [schiff_model, ch])
                if P.empty:
                    continue
                for w_s in FIXED_BLEND_STEPS:
                    w = np.array([w_s, 1 - w_s], dtype=float)
                    pred = P.values @ w
                    nm = f"{stream}__fixedblend_schiff{w_s:.2f}__{ch}"
                    ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, "fixed_schiff_blend", {"models": [schiff_model, ch], "weights": [float(w[0]), float(w[1])], "method": "fixed_schiff_challenger_blend"}))
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": schiff_model, "weight": float(w[0]), "method": "fixed_schiff_challenger_blend", "origin": ""})
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": ch, "weight": float(w[1]), "method": "fixed_schiff_challenger_blend", "origin": ""})

    ensemble_preds = pd.concat(ens_frames, ignore_index=True) if ens_frames else pd.DataFrame()
    weights = pd.DataFrame(weight_rows)
    return ensemble_preds, weights

# =============================================================================
# 8. Paired comparisons, stress tests, recommendations
# =============================================================================

def choose_schiff_benchmark(summary: pd.DataFrame, stream: str) -> Optional[str]:
    s = summary[(summary["stream"] == stream) & (summary["family_tag"].astype(str).str.contains("SCHIFF_OLS", case=False, na=False))].copy()
    if s.empty:
        return None
    # Prefer W40 if close; otherwise choose best pure Schiff by quarterly MAPE.
    w40 = s[s["window"].astype(str).eq("40")]
    if not w40.empty:
        return w40.sort_values(["quarterly_mape", "annual_mape"], na_position="last")["model"].iloc[0]
    return s.sort_values(["quarterly_mape", "annual_mape"], na_position="last")["model"].iloc[0]


def paired_compare(preds: pd.DataFrame, candidate: str, benchmark: str) -> Dict[str, Any]:
    key = ["stream", "origin", "target_period", "horizon"]
    c = preds[preds["model"] == candidate][key + ["actual", "pred"]].rename(columns={"pred": "candidate_pred"})
    b = preds[preds["model"] == benchmark][key + ["actual", "pred"]].rename(columns={"actual": "benchmark_actual", "pred": "benchmark_pred"})
    m = c.merge(b, on=key, how="inner")
    if m.empty:
        return {"candidate": candidate, "benchmark": benchmark, "n_common_pairs": 0}
    actual = m["actual"].astype(float)
    ce = np.abs(m["candidate_pred"] - actual) / np.abs(actual)
    be = np.abs(m["benchmark_pred"] - actual) / np.abs(actual)
    valid = np.isfinite(ce) & np.isfinite(be) & (np.abs(actual) > 1e-12)
    if valid.sum() == 0:
        return {"candidate": candidate, "benchmark": benchmark, "n_common_pairs": 0}
    stream = m["stream"].iloc[0]
    return {
        "stream": stream,
        "baseline": benchmark,
        "challenger": candidate,
        "n_common_pairs": int(valid.sum()),
        "baseline_mape": float(be[valid].mean() * 100),
        "challenger_mape": float(ce[valid].mean() * 100),
        "mape_improvement_pct_points": float((be[valid].mean() - ce[valid].mean()) * 100),
        "challenger_win_rate": float((ce[valid] < be[valid]).mean() * 100),
    }


def build_stress_tests(preds: pd.DataFrame, annual_preds: pd.DataFrame, models: Sequence[str]) -> pd.DataFrame:
    rows = []
    dfp = preds.copy()
    dfp["target_period_obj"] = dfp["target_period"].map(lambda s: pd.Period(str(s), freq="Q-DEC"))
    for model in models:
        g = dfp[dfp["model"] == model].copy()
        if g.empty:
            continue
        stream = g["stream"].iloc[0]
        row = {"stream": stream, "model": model}
        slices = {
            "h1_4": g["horizon"].between(1, 4),
            "h5_8": g["horizon"].between(5, 8),
            "h9_12": g["horizon"].between(9, 12),
            "recent_2024_plus": g["target_period_obj"] >= pd.Period("2024Q1", freq="Q-DEC"),
            "stress_2022_23": g["target_period_obj"].between(pd.Period("2022Q1", freq="Q-DEC"), pd.Period("2023Q4", freq="Q-DEC")),
        }
        for label, mask in slices.items():
            gg = g[mask]
            row[f"{label}_n"] = int(len(gg))
            row[f"{label}_mape"] = mape(gg["actual"], gg["pred"]) if not gg.empty else np.nan
            row[f"{label}_bias"] = bias_pct(gg["actual"], gg["pred"]) if not gg.empty else np.nan
        a = annual_preds[annual_preds["model"] == model]
        row["annual_n"] = int(len(a))
        row["annual_mape"] = mape(a["actual"], a["pred"]) if not a.empty else np.nan
        row["annual_bias"] = bias_pct(a["actual"], a["pred"]) if not a.empty else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


def recommend_models(summary: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    primary = []
    by_q = []
    for stream in STREAM_ORDER:
        s = summary[summary["stream"] == stream].copy()
        if s.empty:
            continue
        primary.append(s.sort_values(["governance_score", "quarterly_mape", "annual_mape"], na_position="last").head(1))
        by_q.append(s.sort_values(["quarterly_mape", "annual_mape", "quarterly_p90_ape"], na_position="last").head(1))
    return pd.concat(primary, ignore_index=True), pd.concat(by_q, ignore_index=True)

# =============================================================================
# 9. Outputs and figures
# =============================================================================

def make_figures(run_dir: Path, summary: pd.DataFrame, primary: pd.DataFrame) -> None:
    try:
        import matplotlib.pyplot as plt
        fig_dir = run_dir / "figures"
        fig_dir.mkdir(exist_ok=True)

        # Finalist vs expected PDF bar chart.
        rows = []
        for _, r in primary.iterrows():
            stream = r["stream"]
            rows.append({"stream": stream, "type": "Latest quarterly", "mape": r["quarterly_mape"]})
            rows.append({"stream": stream, "type": "Latest annual", "mape": r["annual_mape"]})
            if stream in PDF_EXPECTED:
                rows.append({"stream": stream, "type": "PDF quarterly", "mape": PDF_EXPECTED[stream]["quarterly_mape"]})
                rows.append({"stream": stream, "type": "PDF annual", "mape": PDF_EXPECTED[stream]["annual_mape"]})
        d = pd.DataFrame(rows)
        if not d.empty:
            order = STREAM_ORDER
            types = ["PDF quarterly", "Latest quarterly", "PDF annual", "Latest annual"]
            x = np.arange(len(order))
            width = 0.18
            plt.figure(figsize=(10, 5))
            for i, t in enumerate(types):
                vals = [d[(d["stream"] == s) & (d["type"] == t)]["mape"].mean() for s in order]
                plt.bar(x + (i - 1.5) * width, vals, width, label=t)
                for j, v in enumerate(vals):
                    if np.isfinite(v):
                        plt.text(x[j] + (i - 1.5) * width, v + 0.05, f"{v:.2f}", ha="center", va="bottom", fontsize=8)
            plt.xticks(x, [STREAM_LABELS[s] for s in order])
            plt.ylabel("MAPE (%)")
            plt.title("Recommended finalists vs earlier PDF finalist metrics")
            plt.legend()
            plt.tight_layout()
            plt.savefig(fig_dir / "finalists_vs_pdf.png", dpi=160)
            plt.close()

        # Candidate landscape.
        plt.figure(figsize=(10, 5.5))
        colors = {"PED": "tab:blue", "LIGHT_RUC": "tab:orange", "HEAVY_RUC": "tab:green"}
        for stream in STREAM_ORDER:
            s = summary[summary["stream"] == stream]
            plt.scatter(s["quarterly_mape"], s["annual_mape"], s=20, alpha=0.30, label=STREAM_LABELS[stream], color=colors[stream])
            p = primary[primary["stream"] == stream]
            if not p.empty:
                plt.scatter(p["quarterly_mape"], p["annual_mape"], s=120, marker="*", color=colors[stream], edgecolor="black")
        plt.xlabel("Quarterly MAPE (%)")
        plt.ylabel("Annual MAPE (%)")
        plt.title("Candidate arbitration landscape")
        plt.legend()
        plt.tight_layout()
        plt.savefig(fig_dir / "candidate_landscape.png", dpi=160)
        plt.close()
    except Exception as e:
        print(f"Figure generation skipped: {type(e).__name__}: {e}")


def write_excel(run_dir: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    path = run_dir / "stage1_finalist_arbitration_results.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None or df.empty:
                continue
            safe_name = name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        wb = load_workbook(path)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top")
            for i, col_cells in enumerate(ws.columns, 1):
                max_len = 10
                for cell in col_cells:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(val), 70))
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2
        wb.save(path)
    except Exception as e:
        print(f"Excel styling skipped: {type(e).__name__}: {e}")


def write_report(run_dir: Path, primary: pd.DataFrame, by_q: pd.DataFrame, paired: pd.DataFrame, pdf_comp: pd.DataFrame, stress: pd.DataFrame) -> None:
    lines = []
    lines.append("# Stage 1 Finalist Arbitration Report\n")
    lines.append(f"Run folder: `{run_dir}`\n")
    lines.append("\n## Purpose\n")
    lines.append("This run arbitrates between the earlier PDF-style finalist families, the latest bespoke Heavy RUC gains, and the pure Schiff structural benchmarks on a common full rolling-origin validation grid.\n")
    lines.append("\n## Recommended finalists by stream-specific governance score\n")
    lines.append(primary[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "annual_bias_pct", "governance_score"]].to_markdown(index=False))
    lines.append("\n\n## Recommended finalists by quarterly MAPE\n")
    lines.append(by_q[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "governance_score"]].to_markdown(index=False))
    lines.append("\n\n## Comparison with earlier PDF metrics\n")
    lines.append(pdf_comp.to_markdown(index=False))
    lines.append("\n\n## Paired comparison against Schiff benchmarks\n")
    lines.append(paired.head(80).to_markdown(index=False))
    lines.append("\n\n## Stress tests\n")
    lines.append(stress.to_markdown(index=False))
    lines.append("\n\n## Interpretation guidance\n")
    lines.append("- If a new candidate beats the PDF metric but is a static solver, inspect the prequential solver as a production-safer comparator.\n")
    lines.append("- If Light RUC has weaker quarterly MAPE but better annual MAPE, inspect the 2022--23 stress window before choosing.\n")
    lines.append("- If a model uses lead price features, treat it as actual-driver/model-discovery unless the relevant policy path is known at the forecast origin.\n")
    (run_dir / "stage1_finalist_arbitration_report.md").write_text("\n".join(lines), encoding="utf-8")


# =============================================================================
# 10. Scoped HyperOpt finalist refinement
# =============================================================================

# This section replaces the broad arbitration main loop with a scoped HPO audit.
# It deliberately reuses the latest arbitration run outputs rather than rerunning
# the solved broad candidate universe. HyperOpt is used only around known-good
# finalist families, and the best HPO proposals are then validated on the full
# rolling-origin grid.

try:
    from hyperopt import STATUS_OK, Trials, fmin, hp, space_eval, tpe
    HAS_HYPEROPT = True
except Exception as _hyperopt_exc:  # pragma: no cover - informative runtime guard
    HAS_HYPEROPT = False
    HYPEROPT_IMPORT_ERROR = _hyperopt_exc
else:
    HYPEROPT_IMPORT_ERROR = None

# -----------------------------------------------------------------------------
# User-facing HPO refinement configuration
# -----------------------------------------------------------------------------

LATEST_ARBITRATION_RUN_DIR = Path(
    "data/dashboard_evidence_pack_reproducibility/ped_inner_hpo/source_artifacts/finalist_arbitration_run_20260520_002339"
)

HPO_OUTPUT_ROOT = Path(
    "artifacts/ped_inner_hpo/hpo_finalist_refinement_outputs"
)

# These are deliberately much smaller than the full solve. They are targeted
# around the top-performing families already found in run_20260520_002339.
RUN_SCOPED_HPO = True
HPO_ALGO = "tpe"  # currently only TPE is used; leave as label for report clarity.
HPO_RANDOM_SEED = 20260520

HPO_TRIALS_BY_STREAM = {
    "PED": 35,
    "LIGHT_RUC": 55,
    "HEAVY_RUC": 55,
}

# HPO screening is deliberately cheaper than final validation. Final candidates
# are then evaluated with origin_stride=1 and all available origins.
HPO_SCREEN_ORIGIN_STRIDE = 3
HPO_SCREEN_MAX_ORIGINS_PER_STREAM: Optional[int] = 24
HPO_SCREEN_MAX_HORIZON = 12

HPO_TOP_TO_FULL_VALIDATE = {
    "PED": 8,
    "LIGHT_RUC": 12,
    "HEAVY_RUC": 12,
}

# Existing latest-run predictions are treated as anchor candidates. This avoids
# rerunning the solved candidate universe and keeps the HPO run focused.
ANCHOR_TOP_N_BY_METRIC = 15
ANCHOR_INCLUDE_ALL_LATEST_RECOMMENDED = True
ANCHOR_INCLUDE_TOP_BY_QUARTERLY = True
ANCHOR_INCLUDE_TOP_BY_ANNUAL = True
ANCHOR_INCLUDE_TOP_BY_GOVERNANCE = True
ANCHOR_INCLUDE_PURE_SCHIFF = True
ANCHOR_INCLUDE_PDF_STYLE_REFERENCES = True

# Final full-grid validation settings for HPO proposals.
FULL_EVAL_ORIGIN_STRIDE = 1
FULL_EVAL_MAX_ORIGINS_PER_STREAM: Optional[int] = None

# New refined ensembles use a unique prefix so they do not collide with the
# existing arbitration solver names loaded from the anchor run.
REFINED_ENSEMBLE_PREFIX = "HPOREFINE"
TOP_MODELS_FOR_REFINED_ENSEMBLES = 18

# Candidate rejection guardrails for the HPO screen. Keep these loose enough not
# to discard promising oddballs, but tight enough to prevent pathological trials.
MAX_ACCEPTABLE_SCREEN_QMAPE = {
    "PED": 8.0,
    "LIGHT_RUC": 20.0,
    "HEAVY_RUC": 12.0,
}

LATEST_EXPECTED_FINALISTS = {
    "PED": {"model": "PED__solver_static_convex_top18", "quarterly_mape": 2.47358, "annual_mape": 2.38709, "quarterly_bias_pct": 1.50491},
    "LIGHT_RUC": {"model": "LIGHT_RUC__solver_static_convex_top18", "quarterly_mape": 9.14755, "annual_mape": 5.99950, "quarterly_bias_pct": 0.738125},
    "HEAVY_RUC": {"model": "HEAVY_RUC__solver_static_convex_top18", "quarterly_mape": 3.56092, "annual_mape": 3.17141, "quarterly_bias_pct": 0.16585},
}

# -----------------------------------------------------------------------------
# Scoped HPO utilities
# -----------------------------------------------------------------------------

def evaluate_candidate_grid(sd: StreamData, cfg: CandidateConfig, origin_stride: int, max_origins: Optional[int], max_horizon: int = MAX_HORIZON) -> pd.DataFrame:
    """Evaluate a candidate while temporarily overriding grid globals.

    The base arbitration function uses global ORIGIN_STRIDE, MAX_ORIGINS_PER_STREAM
    and MAX_HORIZON. This wrapper lets HPO screening run on a cheaper grid while
    final validation uses the full grid.
    """
    global ORIGIN_STRIDE, MAX_ORIGINS_PER_STREAM, MAX_HORIZON
    old_stride = ORIGIN_STRIDE
    old_max_origins = MAX_ORIGINS_PER_STREAM
    old_horizon = MAX_HORIZON
    try:
        ORIGIN_STRIDE = int(origin_stride)
        MAX_ORIGINS_PER_STREAM = max_origins
        MAX_HORIZON = int(max_horizon)
        return evaluate_candidate(sd, cfg)
    finally:
        ORIGIN_STRIDE = old_stride
        MAX_ORIGINS_PER_STREAM = old_max_origins
        MAX_HORIZON = old_horizon


def safe_metric(value: Any, fallback: float = 1e6) -> float:
    try:
        v = float(value)
    except Exception:
        return fallback
    return v if np.isfinite(v) else fallback


def hpo_screen_loss(row: pd.Series, stream: str) -> float:
    """Stream-specific scalar loss used only for HPO screening.

    This is not the final production decision. Final candidates are later
    re-evaluated on the full grid and ranked by the normal final_summary logic.
    """
    q = safe_metric(row.get("quarterly_mape"))
    a = safe_metric(row.get("annual_mape"), fallback=q)
    p90 = safe_metric(row.get("quarterly_p90_ape"), fallback=q)
    bias = abs(safe_metric(row.get("quarterly_bias_pct"), fallback=0.0))
    h9_12 = safe_metric(row.get("mape_h09_12"), fallback=q)

    if stream == "PED":
        return 0.62 * q + 0.22 * a + 0.08 * p90 + 0.08 * bias
    if stream == "LIGHT_RUC":
        return 0.36 * q + 0.34 * a + 0.12 * h9_12 + 0.10 * p90 + 0.08 * bias
    # Heavy RUC: weight annual a bit more heavily because the latest bespoke
    # run did very well annually and we want to see if HPO can recover it.
    return 0.32 * q + 0.46 * a + 0.10 * p90 + 0.12 * bias


def stream_hpo_space(stream: str) -> Dict[str, Any]:
    """HyperOpt search spaces scoped around known-good families."""
    if stream == "PED":
        return {
            "model_kind": hp.choice("ped_model_kind", ["gbr", "elastic_net", "ridge", "bayesian_ridge"]),
            "feature_set": hp.choice("ped_feature_set", ["diff", "rich", "struct", "schiff", "schiff_no_lead"]),
            "window": hp.choice("ped_window", [40, 52, 64, None]),
            "include_target_lags": hp.choice("ped_ylags", [True, False]),
            "n_estimators": hp.quniform("ped_n_estimators", 120, 750, 10),
            "learning_rate": hp.loguniform("ped_learning_rate", math.log(0.018), math.log(0.12)),
            "max_depth": hp.choice("ped_max_depth", [1, 2]),
            "subsample": hp.choice("ped_subsample", [0.75, 0.85, 0.95, 1.0]),
            "alpha": hp.loguniform("ped_alpha", math.log(0.0005), math.log(50.0)),
            "l1_ratio": hp.uniform("ped_l1_ratio", 0.10, 0.90),
        }

    if stream == "LIGHT_RUC":
        return {
            "model_kind": hp.choice("light_model_kind", ["elastic_net", "ridge", "bayesian_ridge", "gbr", "resid_gbr", "resid_elastic_net", "resid_ridge", "huber"]),
            "feature_set": hp.choice("light_feature_set", ["rich", "diff", "schiff", "schiff_no_lead", "dynamic_pruned", "dynamic_no_leads", "price_distributed_lags"]),
            "window": hp.choice("light_window", [36, 40, 44, 52]),
            "include_target_lags": hp.choice("light_ylags", [True, False]),
            "n_estimators": hp.quniform("light_n_estimators", 80, 550, 10),
            "learning_rate": hp.loguniform("light_learning_rate", math.log(0.018), math.log(0.12)),
            "max_depth": hp.choice("light_max_depth", [1, 2, 3]),
            "subsample": hp.choice("light_subsample", [0.75, 0.85, 0.95, 1.0]),
            "alpha": hp.loguniform("light_alpha", math.log(0.0005), math.log(100.0)),
            "l1_ratio": hp.uniform("light_l1_ratio", 0.10, 0.92),
            "huber_epsilon": hp.uniform("light_huber_epsilon", 1.15, 2.20),
        }

    return {
        "model_kind": hp.choice("heavy_model_kind", ["gbr", "elastic_net", "ridge", "bayesian_ridge", "resid_gbr", "resid_elastic_net", "resid_ridge"]),
        "feature_set": hp.choice("heavy_feature_set", ["dynamic_no_leads", "price_distributed_lags", "dynamic_pruned", "schiff_no_lead", "schiff", "diff", "rich"]),
        "window": hp.choice("heavy_window", [40, 52, 64]),
        "include_target_lags": hp.choice("heavy_ylags", [True, False]),
        "n_estimators": hp.quniform("heavy_n_estimators", 100, 750, 10),
        "learning_rate": hp.loguniform("heavy_learning_rate", math.log(0.018), math.log(0.12)),
        "max_depth": hp.choice("heavy_max_depth", [1, 2]),
        "subsample": hp.choice("heavy_subsample", [0.75, 0.85, 0.95, 1.0]),
        "alpha": hp.loguniform("heavy_alpha", math.log(0.0005), math.log(100.0)),
        "l1_ratio": hp.uniform("heavy_l1_ratio", 0.10, 0.90),
    }


def candidate_from_hpo_sample(stream: str, sample: Dict[str, Any], trial_id: int) -> CandidateConfig:
    kind = str(sample["model_kind"])
    feature_set = str(sample["feature_set"])
    window = sample.get("window")
    if isinstance(window, np.integer):
        window = int(window)
    include_target_lags = bool(sample.get("include_target_lags"))

    params: Dict[str, Any]
    tag: str
    if kind in {"gbr", "resid_gbr"}:
        params = {
            "n_estimators": int(round(float(sample.get("n_estimators", 250)))),
            "learning_rate": float(sample.get("learning_rate", 0.05)),
            "max_depth": int(sample.get("max_depth", 1)),
            "subsample": float(sample.get("subsample", 0.85)),
            "loss": "squared_error",
            "random_state": HPO_RANDOM_SEED + trial_id,
        }
        tag = "HPO_RESID_GBR" if kind == "resid_gbr" else "HPO_GBR"
    elif kind in {"elastic_net", "resid_elastic_net"}:
        params = {
            "alpha": float(sample.get("alpha", 0.01)),
            "l1_ratio": float(sample.get("l1_ratio", 0.5)),
            "max_iter": 50000,
            "random_state": HPO_RANDOM_SEED + trial_id,
        }
        tag = "HPO_RESID_Elastic" if kind == "resid_elastic_net" else "HPO_Elastic"
    elif kind in {"ridge", "resid_ridge"}:
        params = {"alpha": float(sample.get("alpha", 1.0)), "random_state": HPO_RANDOM_SEED + trial_id}
        tag = "HPO_RESID_Ridge" if kind == "resid_ridge" else "HPO_Ridge"
    elif kind == "huber":
        params = {"epsilon": float(sample.get("huber_epsilon", 1.5)), "alpha": float(sample.get("alpha", 0.0001)), "max_iter": 5000}
        tag = "HPO_Huber"
    elif kind == "bayesian_ridge":
        params = {}
        tag = "HPO_BayesianRidge"
    else:
        raise ValueError(f"Unsupported HPO model kind: {kind}")

    w = "exp" if window is None else str(window)
    name = f"{stream}__HPO_trial{trial_id:04d}__{feature_set}__{tag}{clean_label_params(params)}__{'ylag' if include_target_lags else 'noylag'}__w{w}"
    return CandidateConfig(
        stream=stream,
        name=name,
        model_kind=kind,
        params_json=json.dumps(params, sort_keys=True),
        window=window,
        feature_set=feature_set,
        include_target_lags=include_target_lags,
        family_tag=tag,
    )


def run_hpo_for_stream(sd: StreamData, run_dir: Path, trials_to_run: int) -> Tuple[pd.DataFrame, List[CandidateConfig]]:
    """Run scoped HyperOpt for one stream and return trial records + top configs."""
    if not HAS_HYPEROPT:
        raise RuntimeError(f"HyperOpt is not importable: {HYPEROPT_IMPORT_ERROR}")

    stream = sd.stream
    space = stream_hpo_space(stream)
    trials = Trials()
    records: List[Dict[str, Any]] = []
    pred_cache: Dict[str, pd.DataFrame] = {}
    cfg_cache: Dict[str, CandidateConfig] = {}

    def objective(sample: Dict[str, Any]) -> Dict[str, Any]:
        trial_id = len(records) + 1
        cfg = candidate_from_hpo_sample(stream, sample, trial_id)
        if cfg.name in pred_cache:
            pred = pred_cache[cfg.name]
        else:
            pred = evaluate_candidate_grid(sd, cfg, HPO_SCREEN_ORIGIN_STRIDE, HPO_SCREEN_MAX_ORIGINS_PER_STREAM, HPO_SCREEN_MAX_HORIZON)
            pred_cache[cfg.name] = pred
            cfg_cache[cfg.name] = cfg

        if pred.empty:
            loss = 1e6
            row_metrics = {}
        else:
            fs, _ = build_final_summary(pred)
            if fs.empty:
                loss = 1e6
                row_metrics = {}
            else:
                row = fs.iloc[0]
                row_metrics = row.to_dict()
                if safe_metric(row.get("quarterly_mape")) > MAX_ACCEPTABLE_SCREEN_QMAPE[stream]:
                    # Soft penalty for obviously poor candidates.
                    loss = 1e5 + safe_metric(row.get("quarterly_mape"))
                else:
                    loss = hpo_screen_loss(row, stream)

        rec = {
            "stream": stream,
            "trial_id": trial_id,
            "candidate_name": cfg.name,
            "loss": float(loss),
            "model_kind": cfg.model_kind,
            "feature_set": cfg.feature_set,
            "include_target_lags": cfg.include_target_lags,
            "window": "expanding" if cfg.window is None else cfg.window,
            "params_json": cfg.params_json,
            "status": "ok" if loss < 1e6 else "failed_or_empty",
        }
        for k in ["quarterly_mape", "annual_mape", "quarterly_bias_pct", "annual_bias_pct", "quarterly_p90_ape", "governance_score", "mape_h01_04", "mape_h05_08", "mape_h09_12"]:
            rec[k] = row_metrics.get(k, np.nan)
        records.append(rec)

        # Persist progressively so an interrupted HPO run still has useful results.
        pd.DataFrame(records).to_csv(run_dir / f"hpo_trials_{stream}.csv", index=False)
        return {"loss": float(loss), "status": STATUS_OK, "candidate_name": cfg.name}

    print(f"{stream}: starting scoped HyperOpt trials={trials_to_run}, screen_origin_stride={HPO_SCREEN_ORIGIN_STRIDE}, screen_max_origins={HPO_SCREEN_MAX_ORIGINS_PER_STREAM}")
    rng = np.random.default_rng(HPO_RANDOM_SEED + {"PED": 101, "LIGHT_RUC": 202, "HEAVY_RUC": 303}.get(stream, 0))
    fmin(
        fn=objective,
        space=space,
        algo=tpe.suggest,
        max_evals=int(trials_to_run),
        trials=trials,
        rstate=rng,
        show_progressbar=True,
    )

    trial_df = pd.DataFrame(records)
    trial_df.to_csv(run_dir / f"hpo_trials_{stream}.csv", index=False)
    good = trial_df[trial_df["status"].eq("ok")].sort_values(["loss", "quarterly_mape", "annual_mape"], na_position="last")
    top_n = int(HPO_TOP_TO_FULL_VALIDATE[stream])
    top_names = good["candidate_name"].head(top_n).tolist()
    top_configs = [cfg_cache[n] for n in top_names if n in cfg_cache]
    pd.DataFrame([asdict(c) for c in top_configs]).to_csv(run_dir / f"hpo_top_configs_{stream}.csv", index=False)
    return trial_df, top_configs


def select_anchor_models(latest_summary: pd.DataFrame) -> Dict[str, List[str]]:
    """Select a small but rich anchor set from the solved arbitration run."""
    anchors: Dict[str, List[str]] = {}
    for stream in STREAM_ORDER:
        s = latest_summary[latest_summary["stream"].eq(stream)].copy()
        if s.empty:
            anchors[stream] = []
            continue
        models: List[str] = []
        expected = LATEST_EXPECTED_FINALISTS.get(stream, {})
        if expected.get("model") in set(s["model"]):
            models.append(expected["model"])

        if ANCHOR_INCLUDE_TOP_BY_GOVERNANCE and "governance_score" in s.columns:
            models.extend(s.sort_values(["governance_score", "quarterly_mape"], na_position="last")["model"].head(ANCHOR_TOP_N_BY_METRIC).tolist())
        if ANCHOR_INCLUDE_TOP_BY_QUARTERLY:
            models.extend(s.sort_values(["quarterly_mape", "annual_mape"], na_position="last")["model"].head(ANCHOR_TOP_N_BY_METRIC).tolist())
        if ANCHOR_INCLUDE_TOP_BY_ANNUAL:
            models.extend(s.sort_values(["annual_mape", "quarterly_mape"], na_position="last")["model"].head(ANCHOR_TOP_N_BY_METRIC).tolist())
        if ANCHOR_INCLUDE_PURE_SCHIFF:
            schiff = s[s["family_tag"].astype(str).str.contains("SCHIFF_OLS", case=False, na=False)].copy()
            # Pure Schiff only, not residuals/blends/ensembles.
            schiff = schiff[~schiff["model"].astype(str).str.contains("RESID|fixedblend|solver|convex|ensemble|top|median|mean", case=False, na=False)]
            models.extend(schiff.sort_values(["quarterly_mape", "annual_mape"], na_position="last")["model"].head(4).tolist())
        if ANCHOR_INCLUDE_PDF_STYLE_REFERENCES:
            # PDF-style finalists were usually solver/static/top-k around current candidate families.
            pdf_like = s[s["model"].astype(str).str.contains("solver_static_convex_top18|top10_median|top8_median|top10_mean", case=False, na=False)]
            models.extend(pdf_like.sort_values(["quarterly_mape", "annual_mape"], na_position="last")["model"].head(8).tolist())

        anchors[stream] = dedupe(models)
    return anchors


def read_filtered_predictions(pred_path: Path, anchor_models: Dict[str, List[str]], chunksize: int = 300_000) -> pd.DataFrame:
    """Read only selected anchor model predictions from a large predictions CSV."""
    if not pred_path.exists():
        print(f"Anchor prediction file not found: {pred_path}")
        return pd.DataFrame()
    wanted = set(m for models in anchor_models.values() for m in models)
    if not wanted:
        return pd.DataFrame()
    frames = []
    for chunk in pd.read_csv(pred_path, chunksize=chunksize):
        if "model" not in chunk.columns:
            raise ValueError(f"{pred_path} does not contain a model column")
        sub = chunk[chunk["model"].isin(wanted)].copy()
        if not sub.empty:
            frames.append(sub)
    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return out


def build_refined_ensembles(preds: pd.DataFrame, summary: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Build HPO-refined ensembles with unique names to avoid anchor collisions."""
    ens_frames = []
    weight_rows = []
    for stream in STREAM_ORDER:
        ssum = summary[summary["stream"] == stream].copy()
        if ssum.empty:
            continue
        candidates: List[str] = []
        for sort_cols in [["quarterly_mape"], ["annual_mape"], ["quarterly_p90_ape"], ["governance_score"]]:
            if all(c in ssum.columns for c in sort_cols):
                candidates.extend(ssum.sort_values(sort_cols, na_position="last")["model"].head(TOP_MODELS_FOR_REFINED_ENSEMBLES).tolist())
        schiff = ssum[ssum["family_tag"].astype(str).str.contains("SCHIFF_OLS", case=False, na=False)].copy() if "family_tag" in ssum.columns else pd.DataFrame()
        if not schiff.empty:
            schiff = schiff[~schiff["model"].astype(str).str.contains("RESID|fixedblend|solver|convex|ensemble|top|median|mean", case=False, na=False)]
            candidates.extend(schiff.sort_values(["quarterly_mape"], na_position="last")["model"].head(4).tolist())
        models = dedupe(candidates)[:max(TOP_MODELS_FOR_REFINED_ENSEMBLES, 8)]
        if len(models) < 2:
            continue

        P, y, keys = common_matrix(preds, stream, models)
        if not P.empty:
            w = optimise_convex_mape(P.values, y.values)
            pred = P.values @ w
            nm = f"{stream}__{REFINED_ENSEMBLE_PREFIX}_solver_static_convex_top{len(models)}"
            ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, f"{REFINED_ENSEMBLE_PREFIX.lower()}_solver_static_convex", {"base_models": models, "weights": dict(zip(models, [float(x) for x in w])), "method": "static_convex_mape_hpo_refined"}))
            for m, weight in zip(models, w):
                if abs(weight) > 1e-10:
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": m, "weight": float(weight), "method": "static_convex_mape_hpo_refined", "origin": ""})

        # Prequential.
        P, y, keys = common_matrix(preds, stream, models)
        if not P.empty:
            dfm = keys.copy()
            for m in models:
                dfm[m] = P[m].values
            dfm["actual"] = y.values
            dfm["origin_period"] = dfm["origin"].map(lambda s: pd.Period(str(s), freq="Q-DEC"))
            dfm["origin_sort"] = dfm["origin_period"].map(period_sort_value)
            dfm = dfm.sort_values(["origin_sort", "horizon", "target_period"]).reset_index(drop=True)
            preds_out = []
            eq_w = np.ones(len(models)) / len(models)
            nm = f"{stream}__{REFINED_ENSEMBLE_PREFIX}_solver_preq_convex_top{len(models)}"
            for _, row in dfm.iterrows():
                hist = dfm[dfm["origin_sort"] < row["origin_sort"]]
                if len(hist) >= 40:
                    w = optimise_convex_mape(hist[models].values, hist["actual"].values)
                else:
                    w = eq_w
                preds_out.append(float(row[models].values.astype(float) @ w))
                for m, weight in zip(models, w):
                    if abs(weight) > 1e-10:
                        weight_rows.append({"stream": stream, "ensemble": nm, "component_model": m, "weight": float(weight), "method": "prequential_convex_mape_hpo_refined", "origin": row["origin"]})
            ens_frames.append(make_ensemble_frame(keys, y, np.asarray(preds_out), stream, nm, f"{REFINED_ENSEMBLE_PREFIX.lower()}_solver_preq_convex", {"base_models": models, "method": "prequential_convex_mape_hpo_refined"}))

        # Top-k robust ensembles.
        for k in [3, 5, 8, 10]:
            submodels = models[:k]
            if len(submodels) < 2:
                continue
            P, y, keys = common_matrix(preds, stream, submodels)
            if P.empty:
                continue
            arr = P.values
            for method in ["mean", "median", "trimmed_mean"]:
                if method == "mean":
                    pred = arr.mean(axis=1)
                elif method == "median":
                    pred = np.median(arr, axis=1)
                else:
                    pred = np.sort(arr, axis=1)[:, 1:-1].mean(axis=1) if arr.shape[1] >= 4 else arr.mean(axis=1)
                nm = f"{stream}__{REFINED_ENSEMBLE_PREFIX}_top{k}_{method}"
                ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, f"{REFINED_ENSEMBLE_PREFIX.lower()}_topk_{method}", {"models": submodels, "method": method}))

        # Fixed Schiff/challenger blends using refined pool.
        if not schiff.empty:
            schiff_model = schiff.sort_values(["quarterly_mape"], na_position="last")["model"].iloc[0]
            challenger_models = [m for m in models[:8] if m != schiff_model]
            for ch in challenger_models:
                P, y, keys = common_matrix(preds, stream, [schiff_model, ch])
                if P.empty:
                    continue
                for w_s in FIXED_BLEND_STEPS:
                    w = np.array([w_s, 1 - w_s], dtype=float)
                    pred = P.values @ w
                    nm = f"{stream}__{REFINED_ENSEMBLE_PREFIX}_fixedblend_schiff{w_s:.2f}__{ch}"
                    ens_frames.append(make_ensemble_frame(keys, y, pred, stream, nm, f"{REFINED_ENSEMBLE_PREFIX.lower()}_fixed_schiff_blend", {"models": [schiff_model, ch], "weights": [float(w[0]), float(w[1])], "method": "fixed_schiff_challenger_blend_hpo_refined"}))
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": schiff_model, "weight": float(w[0]), "method": "fixed_schiff_challenger_blend_hpo_refined", "origin": ""})
                    weight_rows.append({"stream": stream, "ensemble": nm, "component_model": ch, "weight": float(w[1]), "method": "fixed_schiff_challenger_blend_hpo_refined", "origin": ""})

    ensemble_preds = pd.concat(ens_frames, ignore_index=True) if ens_frames else pd.DataFrame()
    weights = pd.DataFrame(weight_rows)
    return ensemble_preds, weights


def write_hpo_report(run_dir: Path, primary: pd.DataFrame, by_q: pd.DataFrame, paired: pd.DataFrame, pdf_comp: pd.DataFrame, stress: pd.DataFrame, hpo_trials: pd.DataFrame) -> None:
    lines = []
    lines.append("# Stage 1 Scoped HPO Finalist Refinement Report\n")
    lines.append(f"Run folder: `{run_dir}`\n")
    lines.append(f"Anchor run folder: `{LATEST_ARBITRATION_RUN_DIR}`\n")
    lines.append("\n## Purpose\n")
    lines.append("This run does not repeat the broad solve. It loads the latest arbitration run as anchors, then uses HyperOpt only around known-good finalist families. Top HPO proposals are full-grid validated and then combined with anchor candidates in refined solver ensembles.\n")
    lines.append("\n## HPO settings\n")
    lines.append(pd.DataFrame([{"stream": k, "trials": v, "top_full_validated": HPO_TOP_TO_FULL_VALIDATE[k]} for k, v in HPO_TRIALS_BY_STREAM.items()]).to_markdown(index=False))
    lines.append("\n\n## Recommended finalists by governance score\n")
    lines.append(primary[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "annual_bias_pct", "governance_score"]].to_markdown(index=False))
    lines.append("\n\n## Recommended finalists by quarterly MAPE\n")
    lines.append(by_q[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "governance_score"]].to_markdown(index=False))
    lines.append("\n\n## HPO trial summary\n")
    if not hpo_trials.empty:
        agg = hpo_trials.groupby("stream").agg(n_trials=("trial_id", "count"), best_loss=("loss", "min"), best_q_mape=("quarterly_mape", "min"), best_a_mape=("annual_mape", "min")).reset_index()
        lines.append(agg.to_markdown(index=False))
        lines.append("\n\n### Top HPO trial rows\n")
        lines.append(hpo_trials.sort_values(["stream", "loss"]).groupby("stream", group_keys=False).head(10)[["stream", "candidate_name", "loss", "model_kind", "feature_set", "include_target_lags", "window", "quarterly_mape", "annual_mape", "quarterly_bias_pct"]].to_markdown(index=False))
    else:
        lines.append("No HPO trials were run or all trials failed.\n")
    lines.append("\n\n## Comparison with earlier PDF/latest metrics\n")
    lines.append(pdf_comp.to_markdown(index=False))
    lines.append("\n\n## Paired comparison against Schiff benchmarks\n")
    lines.append(paired.head(80).to_markdown(index=False))
    lines.append("\n\n## Stress tests\n")
    lines.append(stress.to_markdown(index=False))
    lines.append("\n\n## Interpretation guidance\n")
    lines.append("- Treat any static convex solver gain as diagnostic until prequential/locked-weight alternatives are also acceptable.\n")
    lines.append("- This run is a targeted HPO refinement of known-good families, not a new broad AutoML/neural search.\n")
    lines.append("- If HPO does not materially beat the arbitration anchors, prefer the simpler arbitration finalist set.\n")
    (run_dir / "stage1_scoped_hpo_finalist_report.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    t0 = time.time()
    if RUN_SCOPED_HPO and not HAS_HYPEROPT:
        raise RuntimeError(f"HyperOpt requested but import failed: {HYPEROPT_IMPORT_ERROR}")
    if not LATEST_ARBITRATION_RUN_DIR.exists():
        raise FileNotFoundError(f"Latest arbitration run folder not found: {LATEST_ARBITRATION_RUN_DIR}")

    run_dir = ensure_run_dir(HPO_OUTPUT_ROOT)
    print("=" * 100)
    print("STAGE 1 SCOPED HPO FINALIST REFINEMENT")
    print("=" * 100)
    print(f"Workbook: {INPUT_XLSX}")
    print(f"Anchor arbitration run: {LATEST_ARBITRATION_RUN_DIR}")
    print(f"Run folder: {run_dir}")
    print("This run loads existing top arbitration predictions and only HPO-tunes known-good families.")

    latest_summary_path = LATEST_ARBITRATION_RUN_DIR / "final_summary.csv"
    latest_pred_path = LATEST_ARBITRATION_RUN_DIR / "all_quarterly_predictions.csv"
    if not latest_summary_path.exists():
        raise FileNotFoundError(f"Missing anchor final_summary.csv: {latest_summary_path}")
    latest_summary = pd.read_csv(latest_summary_path)
    latest_summary.to_csv(run_dir / "anchor_final_summary_copy.csv", index=False)

    # Select and load anchor predictions from the solved arbitration run.
    anchor_models = select_anchor_models(latest_summary)
    pd.DataFrame([{"stream": s, "model": m} for s, models in anchor_models.items() for m in models]).to_csv(run_dir / "anchor_model_inventory.csv", index=False)
    anchor_preds = read_filtered_predictions(latest_pred_path, anchor_models)
    if anchor_preds.empty:
        raise RuntimeError("No anchor predictions were loaded. Ensure all_quarterly_predictions.csv exists in the latest arbitration run folder.")
    anchor_preds.to_csv(run_dir / "anchor_quarterly_predictions.csv", index=False)
    anchor_summary, anchor_annual = build_final_summary(anchor_preds)
    anchor_summary.to_csv(run_dir / "anchor_summary.csv", index=False)
    anchor_annual.to_csv(run_dir / "anchor_annual_predictions.csv", index=False)
    print("Loaded anchor predictions:")
    print(anchor_summary.groupby("stream").size().reset_index(name="n_anchor_models").to_markdown(index=False))

    # Build workbook stream data once for HPO proposals.
    df = load_input_sheet(INPUT_XLSX)
    input_sheet = df.attrs.get("input_sheet", "")
    targets = {s: detect_target_col(df, s) for s in STREAM_ORDER}
    all_targets = [v[0] for v in targets.values()]
    stream_data: Dict[str, StreamData] = {}
    feature_inventory = []
    for stream in STREAM_ORDER:
        target_col, target_is_log = targets[stream]
        fcols = detect_feature_cols(df, stream, all_targets)
        y_raw, y_log = build_target_series(df, target_col, target_is_log)
        exog, groups, primary_log = build_exog(df, stream, fcols)
        sd = StreamData(stream, target_col, target_is_log, fcols, y_raw, y_log, exog, groups, primary_log)
        stream_data[stream] = sd
        for c in fcols:
            feature_inventory.append({"stream": stream, "feature_column": c, "group": assign_base_group(stream, c)})
        print(f"{stream}: target={target_col}; rows={y_raw.dropna().shape[0]}; period={min(y_raw.dropna().index)} to {max(y_raw.dropna().index)}; features={len(fcols)}")
    pd.DataFrame(feature_inventory).to_csv(run_dir / "feature_inventory.csv", index=False)

    # Run scoped HPO by stream.
    hpo_trial_frames = []
    hpo_top_configs: List[CandidateConfig] = []
    if RUN_SCOPED_HPO:
        for stream in STREAM_ORDER:
            trials_to_run = int(HPO_TRIALS_BY_STREAM[stream])
            trial_df, top_configs = run_hpo_for_stream(stream_data[stream], run_dir, trials_to_run)
            hpo_trial_frames.append(trial_df)
            hpo_top_configs.extend(top_configs)
    hpo_trials = pd.concat(hpo_trial_frames, ignore_index=True) if hpo_trial_frames else pd.DataFrame()
    hpo_trials.to_csv(run_dir / "hpo_trials_all_streams.csv", index=False)
    pd.DataFrame([asdict(c) for c in hpo_top_configs]).to_csv(run_dir / "hpo_top_configs_all_streams.csv", index=False)

    # Full-grid validation of top HPO proposals.
    hpo_full_frames = []
    for i, cfg in enumerate(hpo_top_configs, 1):
        print(f"Full validating HPO proposal {i:03d}/{len(hpo_top_configs):03d}: {cfg.name}")
        pred = evaluate_candidate_grid(stream_data[cfg.stream], cfg, FULL_EVAL_ORIGIN_STRIDE, FULL_EVAL_MAX_ORIGINS_PER_STREAM, MAX_HORIZON)
        if not pred.empty:
            hpo_full_frames.append(pred)
    hpo_full_preds = pd.concat(hpo_full_frames, ignore_index=True) if hpo_full_frames else pd.DataFrame()
    hpo_full_preds.to_csv(run_dir / "hpo_full_validation_quarterly_predictions.csv", index=False)
    hpo_full_summary, hpo_full_annual = build_final_summary(hpo_full_preds) if not hpo_full_preds.empty else (pd.DataFrame(), pd.DataFrame())
    hpo_full_summary.to_csv(run_dir / "hpo_full_validation_summary.csv", index=False)
    hpo_full_annual.to_csv(run_dir / "hpo_full_validation_annual_predictions.csv", index=False)

    # Combine anchor + HPO proposal predictions, then build refined ensembles.
    combined_base_preds = pd.concat([anchor_preds, hpo_full_preds], ignore_index=True) if not hpo_full_preds.empty else anchor_preds.copy()
    combined_base_preds.to_csv(run_dir / "combined_base_quarterly_predictions.csv", index=False)
    combined_base_summary, combined_base_annual = build_final_summary(combined_base_preds)
    combined_base_summary.to_csv(run_dir / "combined_base_summary.csv", index=False)
    combined_base_annual.to_csv(run_dir / "combined_base_annual_predictions.csv", index=False)

    print("Building HPO-refined ensembles...")
    refined_ens_preds, refined_weights = build_refined_ensembles(combined_base_preds, combined_base_summary)
    refined_weights.to_csv(run_dir / "hpo_refined_ensemble_weights.csv", index=False)
    refined_ens_preds.to_csv(run_dir / "hpo_refined_ensemble_quarterly_predictions.csv", index=False)

    all_preds = pd.concat([combined_base_preds, refined_ens_preds], ignore_index=True) if not refined_ens_preds.empty else combined_base_preds.copy()
    all_preds.to_csv(run_dir / "all_quarterly_predictions.csv", index=False)
    final_summary, all_annual = build_final_summary(all_preds)
    final_summary.to_csv(run_dir / "final_summary.csv", index=False)
    all_annual.to_csv(run_dir / "all_annual_predictions.csv", index=False)

    primary, by_q = recommend_models(final_summary)
    primary.to_csv(run_dir / "recommended_finalists.csv", index=False)
    by_q.to_csv(run_dir / "recommended_finalists_by_quarterly.csv", index=False)

    # Paired comparisons against pure Schiff.
    paired_rows = []
    stress_models = set(primary["model"].tolist() + by_q["model"].tolist())
    for stream in STREAM_ORDER:
        schiff = choose_schiff_benchmark(final_summary, stream)
        if schiff:
            stress_models.add(schiff)
            top_models = final_summary[final_summary["stream"] == stream].sort_values(["governance_score", "quarterly_mape"], na_position="last")["model"].head(40).tolist()
            for model in dedupe(primary[primary["stream"] == stream]["model"].tolist() + by_q[by_q["stream"] == stream]["model"].tolist() + top_models):
                if model != schiff:
                    paired_rows.append(paired_compare(all_preds, model, schiff))
    paired = pd.DataFrame(paired_rows)
    paired.to_csv(run_dir / "paired_vs_schiff.csv", index=False)

    stress = build_stress_tests(all_preds, all_annual, list(stress_models))
    stress.to_csv(run_dir / "stress_tests.csv", index=False)

    # Compare against PDF and latest arbitration expected values.
    comp_rows = []
    for _, r in primary.iterrows():
        stream = r["stream"]
        pdf = PDF_EXPECTED.get(stream, {})
        latest = LATEST_EXPECTED_FINALISTS.get(stream, {})
        latest_bespoke = LATEST_BESPOKE_EXPECTED.get(stream, {})
        comp_rows.append({
            "stream": stream,
            "selected_model": r["model"],
            "selected_quarterly_mape": r["quarterly_mape"],
            "latest_arbitration_q_mape": latest.get("quarterly_mape", np.nan),
            "selected_minus_latest_arbitration_q_pp": r["quarterly_mape"] - latest.get("quarterly_mape", np.nan),
            "pdf_quarterly_mape": pdf.get("quarterly_mape", np.nan),
            "selected_minus_pdf_q_pp": r["quarterly_mape"] - pdf.get("quarterly_mape", np.nan),
            "selected_annual_mape": r["annual_mape"],
            "latest_arbitration_a_mape": latest.get("annual_mape", np.nan),
            "selected_minus_latest_arbitration_a_pp": r["annual_mape"] - latest.get("annual_mape", np.nan),
            "pdf_annual_mape": pdf.get("annual_mape", np.nan),
            "selected_minus_pdf_a_pp": r["annual_mape"] - pdf.get("annual_mape", np.nan),
            "latest_bespoke_q_mape": latest_bespoke.get("quarterly_mape", np.nan),
            "selected_minus_latest_bespoke_q_pp": r["quarterly_mape"] - latest_bespoke.get("quarterly_mape", np.nan),
            "latest_bespoke_a_mape": latest_bespoke.get("annual_mape", np.nan),
            "selected_minus_latest_bespoke_a_pp": r["annual_mape"] - latest_bespoke.get("annual_mape", np.nan),
        })
    pdf_comp = pd.DataFrame(comp_rows)
    pdf_comp.to_csv(run_dir / "expected_metric_comparison.csv", index=False)
    pdf_comp.to_csv(run_dir / "pdf_expected_comparison.csv", index=False)  # backward-compatible name

    top50 = pd.concat([final_summary[final_summary["stream"] == s].sort_values(["governance_score", "quarterly_mape"], na_position="last").head(50) for s in STREAM_ORDER], ignore_index=True)
    top50.to_csv(run_dir / "top50_by_stream.csv", index=False)

    make_figures(run_dir, final_summary, primary)
    write_excel(run_dir, {
        "Recommended Finalists": primary,
        "Recommended By Q MAPE": by_q,
        "Expected Comparison": pdf_comp,
        "Paired Vs Schiff": paired,
        "Stress Tests": stress,
        "Final Summary": final_summary,
        "Top 50 By Stream": top50,
        "HPO Trials": hpo_trials,
        "HPO Full Summary": hpo_full_summary,
        "Anchor Summary": anchor_summary,
        "Refined Ensemble Weights": refined_weights,
        "Feature Inventory": pd.DataFrame(feature_inventory),
    })
    # Rename the default Excel file to reflect HPO refinement if it exists.
    default_xlsx = run_dir / "stage1_finalist_arbitration_results.xlsx"
    hpo_xlsx = run_dir / "stage1_scoped_hpo_finalist_results.xlsx"
    if default_xlsx.exists():
        default_xlsx.replace(hpo_xlsx)

    write_hpo_report(run_dir, primary, by_q, paired, pdf_comp, stress, hpo_trials)

    elapsed = (time.time() - t0) / 60
    print("\nRUN COMPLETE")
    print("=" * 100)
    print(f"Elapsed minutes: {elapsed:.1f}")
    print(f"Input sheet: {input_sheet}")
    print(f"Anchor run folder: {LATEST_ARBITRATION_RUN_DIR}")
    print(f"Run folder: {run_dir}")
    print("Recommended finalists by primary governance score:")
    print(primary[["stream", "model", "source_family", "model_kind", "feature_set", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "governance_score"]].to_markdown(index=False))
    print("\nKey files:")
    for fn in ["recommended_finalists.csv", "recommended_finalists_by_quarterly.csv", "final_summary.csv", "paired_vs_schiff.csv", "stress_tests.csv", "hpo_trials_all_streams.csv", "hpo_full_validation_summary.csv", "hpo_refined_ensemble_weights.csv", "expected_metric_comparison.csv", "stage1_scoped_hpo_finalist_results.xlsx", "stage1_scoped_hpo_finalist_report.md"]:
        print(run_dir / fn)


if __name__ == "__main__":
    main()
