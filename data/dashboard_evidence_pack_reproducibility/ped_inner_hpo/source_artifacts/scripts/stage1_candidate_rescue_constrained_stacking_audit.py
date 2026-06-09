#!/usr/bin/env python
# -*- coding: utf-8 -*-
r"""
Stage 1 candidate-rescue + constrained-stacking audit
=====================================================

Purpose
-------
This script is a narrow, production-oriented audit layer designed to sit AFTER the
expensive broad model searches have already been completed.

It does NOT rerun the broad candidate search. Instead, it:

1. Reads completed run folders, especially:
   - final arbitration run
   - scoped HPO run
   - Heavy RUC reconciliation run
   - optional earlier bespoke solver run

2. Builds a rescued candidate registry from the best models in those runs:
   - governance finalists
   - top quarterly-MAPE candidates
   - top annual-MAPE candidates
   - low-bias candidates
   - pure Schiff benchmarks
   - prequential/static solver candidates
   - diverse candidate families

3. Loads only the selected models' quarterly predictions from each run folder.

4. Recomputes metrics on a common basis where predictions are available.

5. Builds new constrained ensembles using exact linear-programming solvers:
   - quarterly-MAPE objective
   - quarterly + annual objective
   - bias-constrained objective
   - stress-aware objective
   - max-weight-capped objective
   - prequential / cross-fitted stacking

6. Produces stream-level recommendations and governance diagnostics.

Why this is useful
------------------
This script addresses the failure mode where a broad search may contain a good
candidate family but the final solver pool does not include the right diversified
set of models. It also avoids re-running thousands of expensive base-model fits.

Default paths in this vendored copy are repo-relative placeholders. You can
override them with CLI arguments.

Run
---
PowerShell example:

    conda activate agts312
    cd "<repo-root>"
    python "stage1_candidate_rescue_constrained_stacking_audit.py"

Outputs
-------
A timestamped output folder containing:

    rescued_candidate_registry.csv
    selected_quarterly_predictions.csv
    anchor_candidate_metrics.csv
    generated_ensemble_metrics.csv
    final_summary.csv
    final_recommendations.csv
    paired_vs_schiff.csv
    stress_tests.csv
    ensemble_weights.csv
    stage1_candidate_rescue_results.xlsx
    stage1_candidate_rescue_report.md
    figures/*.png

Dependencies
------------
Required:
    pandas, numpy, scipy, openpyxl, matplotlib

Optional:
    xlsxwriter (not required)

"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from scipy.optimize import linprog, minimize

# Matplotlib is only used to save simple audit charts.
import matplotlib.pyplot as plt


# =============================================================================
# 0. Defaults and configuration
# =============================================================================

DEFAULT_BASE = Path("data/dashboard_evidence_pack_reproducibility/ped_inner_hpo/source_artifacts")

DEFAULT_RUNS = {
    "arbitration": DEFAULT_BASE / "finalist_arbitration_run_20260520_002339",
    "hpo": DEFAULT_BASE / "hpo_refinement_core_outputs",
    "heavy_recon": DEFAULT_BASE / "heavy_ruc_reconciliation_outputs/run_20260521_151642",
    # Optional. If missing, it is skipped.
    "bespoke_solver": DEFAULT_BASE / "bespoke_solver_stage1_outputs/run_20260519_150434",
}

DEFAULT_OUTPUT_ROOT = Path("artifacts/ped_inner_hpo/candidate_rescue_constrained_stacking_outputs")

STREAMS = ["PED", "LIGHT_RUC", "HEAVY_RUC"]
STREAM_LABELS = {
    "PED": "PED VKT per capita",
    "LIGHT_RUC": "Light RUC volume",
    "HEAVY_RUC": "Heavy RUC volume",
}

# PED annual aggregation uses average VKT-per-capita. RUC annual aggregation uses
# summed volumes/net km.
ANNUAL_AGG = {
    "PED": "mean",
    "LIGHT_RUC": "sum",
    "HEAVY_RUC": "sum",
}

# Candidate rescue controls
TOP_N_BY_METRIC = 25
TOP_N_LOW_BIAS = 20
TOP_N_FOR_ENSEMBLE = 30
MAX_CANDIDATES_PER_STREAM_FOR_PRED_LOAD = 90

# Solver controls
ENSEMBLE_POOL_SIZES = [8, 12, 18, 25]
MAX_WEIGHT_CAPS = [None, 0.60, 0.50, 0.40]
MIN_PREQ_HISTORY_ROWS = 60

# Stress windows
STRESS_WINDOWS = {
    "recent_2024_plus": ("2024Q1", "2099Q4"),
    "policy_2022_23": ("2022Q1", "2023Q4"),
}

# Reference values from prior runs. These are used for context only; the script
# ranks based on actual loaded data.
REFERENCE_EXPECTED = {
    "pdf": {
        "PED": {"quarterly_mape": 2.480, "annual_mape": 2.420},
        "LIGHT_RUC": {"quarterly_mape": 9.160, "annual_mape": 6.250},
        "HEAVY_RUC": {"quarterly_mape": 3.800, "annual_mape": 3.070},
    },
    "arbitration": {
        "PED": {"quarterly_mape": 2.47358, "annual_mape": 2.38709},
        "LIGHT_RUC": {"quarterly_mape": 9.14755, "annual_mape": 5.99950},
        "HEAVY_RUC": {"quarterly_mape": 3.56092, "annual_mape": 3.17141},
    },
    "hpo": {
        "PED": {"quarterly_mape": 2.47324, "annual_mape": 2.38562},
        "LIGHT_RUC": {"quarterly_mape": 9.11760, "annual_mape": 5.99244},
        "HEAVY_RUC": {"quarterly_mape": 3.56092, "annual_mape": 3.17141},
    },
    "heavy_recon": {
        "HEAVY_RUC": {"quarterly_mape": 3.48437, "annual_mape": 3.01998},
    },
    "old_bespoke_heavy": {
        "HEAVY_RUC": {"quarterly_mape": 3.27657, "annual_mape": 2.445},
    },
}


# =============================================================================
# 1. Utility functions
# =============================================================================

def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def normalise_colname(col: Any) -> str:
    text = str(col).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def normalise_stream(x: Any) -> str:
    text = str(x).strip().upper().replace(" ", "_")
    if text in {"PED", "PED_VKT", "PED_VKT_PER_CAPITA", "PED_VKT_PC"}:
        return "PED"
    if "LIGHT" in text and ("RUC" in text or "IUC" in text):
        return "LIGHT_RUC"
    if "HEAVY" in text and ("RUC" in text or "IUC" in text):
        return "HEAVY_RUC"
    return text


def stream_label(stream: str) -> str:
    return STREAM_LABELS.get(normalise_stream(stream), str(stream))


def human_label(value: Any) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    text = str(value).replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def short_model_name(model: str, max_len: int = 70) -> str:
    text = str(model)
    replacements = {
        "HEAVY_RUC__": "Heavy ",
        "LIGHT_RUC__": "Light ",
        "PED__": "PED ",
        "solver_static_convex_top18": "Static solver top18",
        "solver_preq_convex_top18": "Prequential solver top18",
        "HPOREFINE_solver_static_convex_top18": "HPO static solver top18",
        "HPOREFINE_top3_mean": "HPO top3 mean",
        "dynamic_no_leads": "Dynamic no-leads",
        "price_distributed_lags": "Price lags",
        "SCHIFF_OLS": "Schiff OLS",
        "SCHIFF_RESID": "Schiff residual",
        "GBR": "GBM",
        "Elastic": "ElasticNet",
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = text.replace("__", " · ").replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > max_len:
        return text[: max_len - 1].rstrip() + "…"
    return text


def fmt_pct(x: Any, dp: int = 3) -> str:
    try:
        if pd.isna(x):
            return ""
        return f"{float(x):.{dp}f}%"
    except Exception:
        return str(x)


def safe_float(x: Any) -> float:
    try:
        if pd.isna(x):
            return np.nan
        return float(x)
    except Exception:
        return np.nan


def read_csv_optional(path: Path) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception as exc:
        print(f"WARNING: could not read {path}: {type(exc).__name__}: {exc}")
        return pd.DataFrame()


def find_first_existing(run_dir: Path, names: Sequence[str]) -> Optional[Path]:
    for name in names:
        p = run_dir / name
        if p.exists() and p.stat().st_size > 0:
            return p
    return None


def period_sort_value(period: Any) -> int:
    p = parse_period(period)
    if p is None:
        return -1
    return int(p.year * 4 + p.quarter)


def parse_period(x: Any) -> Optional[pd.Period]:
    if pd.isna(x):
        return None
    if isinstance(x, pd.Period):
        return pd.Period(str(x), freq="Q-DEC")
    text = str(x).strip()
    m = re.search(r"(19\d{2}|20\d{2})\s*[- ]?\s*[Qq]\s*([1-4])", text)
    if m:
        return pd.Period(f"{m.group(1)}Q{m.group(2)}", freq="Q-DEC")
    m = re.search(r"[Qq]\s*([1-4])\s*[- ]?\s*(19\d{2}|20\d{2})", text)
    if m:
        return pd.Period(f"{m.group(2)}Q{m.group(1)}", freq="Q-DEC")
    try:
        ts = pd.to_datetime(text, errors="coerce")
        if pd.notna(ts):
            return ts.to_period("Q-DEC")
    except Exception:
        pass
    return None


def june_year_from_period(period: Any) -> Optional[int]:
    p = parse_period(period)
    if p is None:
        return None
    return int(p.year + 1 if p.quarter in (3, 4) else p.year)


def in_period_window(period: Any, start: str, end: str) -> bool:
    p = parse_period(period)
    if p is None:
        return False
    return pd.Period(start, freq="Q-DEC") <= p <= pd.Period(end, freq="Q-DEC")


def mape(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    return float(np.mean(np.abs(p[mask] - a[mask]) / np.abs(a[mask])) * 100.0)


def bias_pct(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    return float(np.mean((p[mask] - a[mask]) / np.abs(a[mask])) * 100.0)


def p90_ape(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p) & (np.abs(a) > 1e-12)
    if mask.sum() == 0:
        return np.nan
    ape = np.abs(p[mask] - a[mask]) / np.abs(a[mask]) * 100.0
    return float(np.nanpercentile(ape, 90))


def rmse(actual: Sequence[float], pred: Sequence[float]) -> float:
    a = np.asarray(actual, dtype=float)
    p = np.asarray(pred, dtype=float)
    mask = np.isfinite(a) & np.isfinite(p)
    if mask.sum() == 0:
        return np.nan
    return float(np.sqrt(np.mean((p[mask] - a[mask]) ** 2)))


def make_output_dir(root: Path) -> Path:
    out = root / f"run_{now_stamp()}"
    out.mkdir(parents=True, exist_ok=True)
    (out / "figures").mkdir(parents=True, exist_ok=True)
    return out


# =============================================================================
# 2. Data model
# =============================================================================

@dataclass
class RunSpec:
    run_id: str
    path: Path


@dataclass
class CandidateRef:
    run_id: str
    stream: str
    model: str
    model_uid: str
    include_reason: str
    source_family: str = ""
    model_kind: str = ""
    feature_set: str = ""


# =============================================================================
# 3. Loading run outputs
# =============================================================================

def normalise_summary(df: pd.DataFrame, run_id: str) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()
    # Map common columns to canonical names.
    norm_map = {c: normalise_colname(c) for c in out.columns}
    out = out.rename(columns={c: norm_map[c] for c in out.columns})

    # Column aliasing.
    aliases = {
        "q_mape": "quarterly_mape",
        "quarterly_mape_pct": "quarterly_mape",
        "a_mape": "annual_mape",
        "annual_mape_pct": "annual_mape",
        "bias": "quarterly_bias_pct",
        "quarterly_bias": "quarterly_bias_pct",
        "annual_bias": "annual_bias_pct",
        "p90_ape": "quarterly_p90_ape",
        "n_pairs": "n_quarterly_pairs",
    }
    for old, new in aliases.items():
        if old in out.columns and new not in out.columns:
            out = out.rename(columns={old: new})

    required = ["stream", "model"]
    for col in required:
        if col not in out.columns:
            raise ValueError(f"Summary from {run_id} is missing required column: {col}")

    out["run_id"] = run_id
    out["stream"] = out["stream"].map(normalise_stream)
    out["model"] = out["model"].astype(str)
    out["model_uid"] = out["run_id"].astype(str) + "::" + out["model"].astype(str)

    for col in [
        "quarterly_mape",
        "annual_mape",
        "quarterly_bias_pct",
        "annual_bias_pct",
        "quarterly_p90_ape",
        "annual_p90_ape",
        "governance_score",
        "n_quarterly_pairs",
        "n_annual_pairs",
        "n_origins",
    ]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")

    for col in ["source_family", "model_kind", "feature_set", "variant", "stage"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].astype(str)

    out["abs_quarterly_bias"] = out.get("quarterly_bias_pct", np.nan).abs()
    out["is_pure_schiff"] = out["model"].map(is_pure_schiff_model)
    out["model_short"] = out["model"].map(short_model_name)
    return out


def load_run_summaries(run_specs: Sequence[RunSpec]) -> pd.DataFrame:
    frames = []
    summary_names = [
        "final_summary.csv",
        "all_model_summary.csv",
        "quarterly_summary.csv",
        "recommended_finalists.csv",
    ]
    for spec in run_specs:
        if not spec.path.exists():
            print(f"Skipping missing run folder: {spec.run_id}: {spec.path}")
            continue
        p = find_first_existing(spec.path, summary_names)
        if p is None:
            print(f"No summary file found in {spec.run_id}: {spec.path}")
            continue
        df = read_csv_optional(p)
        if df.empty:
            continue
        try:
            norm = normalise_summary(df, spec.run_id)
            norm["summary_file"] = str(p)
            frames.append(norm)
            print(f"Loaded summary {spec.run_id}: {len(norm):,} rows from {p.name}")
        except Exception as exc:
            print(f"WARNING: failed to normalise summary {p}: {type(exc).__name__}: {exc}")
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def is_pure_schiff_model(model: str) -> bool:
    text = str(model).upper()
    if "SCHIFF" not in text:
        return False
    if "OLS" not in text:
        return False
    bad = [
        "RESID",
        "RESIDUAL",
        "FIXEDBLEND",
        "SOLVER",
        "CONVEX",
        "ENSEMBLE",
        "TOP",
        "MEDIAN",
        "MEAN",
        "GBR",
        "ELASTIC",
        "RIDGE",
    ]
    return not any(b in text for b in bad)


# =============================================================================
# 4. Candidate rescue registry
# =============================================================================

def add_candidate(rows: List[CandidateRef], row: pd.Series, reason: str) -> None:
    try:
        rows.append(
            CandidateRef(
                run_id=str(row.get("run_id", "")),
                stream=normalise_stream(row.get("stream", "")),
                model=str(row.get("model", "")),
                model_uid=str(row.get("model_uid", f"{row.get('run_id','')}::{row.get('model','')}")),
                include_reason=reason,
                source_family=str(row.get("source_family", "")),
                model_kind=str(row.get("model_kind", "")),
                feature_set=str(row.get("feature_set", "")),
            )
        )
    except Exception:
        pass


def build_candidate_registry(summary: pd.DataFrame) -> pd.DataFrame:
    if summary.empty:
        return pd.DataFrame()

    rows: List[CandidateRef] = []

    for stream in STREAMS:
        s = summary[summary["stream"] == stream].copy()
        if s.empty:
            continue

        # Known named finalists and expected candidates.
        name_patterns = [
            "solver_static_convex_top18",
            "solver_preq_convex_top18",
            "HPOREFINE_solver_static_convex_top18",
            "HPOREFINE_top3_mean",
            "top10_median",
            "top8_median",
            "top5_mean",
            "top5_median",
            "fixedblend_schiff0.50",
        ]
        for pat in name_patterns:
            matched = s[s["model"].str.contains(pat, case=False, na=False)]
            for _, r in matched.iterrows():
                add_candidate(rows, r, f"named_pattern:{pat}")

        # Pure Schiff benchmarks.
        for _, r in s[s["is_pure_schiff"]].iterrows():
            add_candidate(rows, r, "pure_schiff_benchmark")

        # Top by quarterly, annual, governance, p90, bias.
        ranking_specs = [
            ("quarterly_mape", "top_quarterly"),
            ("annual_mape", "top_annual"),
            ("governance_score", "top_governance"),
            ("quarterly_p90_ape", "top_p90"),
            ("abs_quarterly_bias", "low_bias"),
        ]
        for col, reason in ranking_specs:
            if col in s.columns:
                sub = s[np.isfinite(pd.to_numeric(s[col], errors="coerce"))].sort_values(col).head(
                    TOP_N_LOW_BIAS if reason == "low_bias" else TOP_N_BY_METRIC
                )
                for _, r in sub.iterrows():
                    add_candidate(rows, r, reason)

        # Family diversity: top few from each feature/model kind/source.
        for group_cols, reason in [
            (["source_family"], "family_diversity_source"),
            (["model_kind"], "family_diversity_kind"),
            (["feature_set"], "family_diversity_feature"),
            (["source_family", "feature_set"], "family_diversity_source_feature"),
        ]:
            available = [c for c in group_cols if c in s.columns]
            if not available or "quarterly_mape" not in s.columns:
                continue
            for _, g in s.groupby(available, dropna=False):
                gg = g.sort_values(["quarterly_mape", "annual_mape"], na_position="last").head(5)
                for _, r in gg.iterrows():
                    add_candidate(rows, r, reason)

    reg = pd.DataFrame([asdict(r) for r in rows])
    if reg.empty:
        return reg

    # Merge duplicate include reasons.
    reg = (
        reg.groupby(["run_id", "stream", "model", "model_uid", "source_family", "model_kind", "feature_set"], as_index=False)
        .agg(include_reason=("include_reason", lambda x: ";".join(sorted(set(map(str, x))))))
    )

    # Cap per stream to avoid huge prediction loads. Keep known/pure and top-ranked first.
    priority_terms = [
        "named_pattern",
        "pure_schiff",
        "top_quarterly",
        "top_annual",
        "top_governance",
        "low_bias",
        "family_diversity",
    ]
    reg["priority"] = 0
    for i, term in enumerate(priority_terms[::-1], 1):
        reg.loc[reg["include_reason"].str.contains(term, case=False, na=False), "priority"] += i

    capped = []
    for stream, g in reg.sort_values(["priority"], ascending=False).groupby("stream"):
        capped.append(g.head(MAX_CANDIDATES_PER_STREAM_FOR_PRED_LOAD))
    reg = pd.concat(capped, ignore_index=True) if capped else reg

    reg["stream_label"] = reg["stream"].map(stream_label)
    reg["model_short"] = reg["model"].map(short_model_name)
    return reg.sort_values(["stream", "priority"], ascending=[True, False]).reset_index(drop=True)


# =============================================================================
# 5. Prediction loading
# =============================================================================

def prediction_file_for_run(run_dir: Path) -> Optional[Path]:
    return find_first_existing(
        run_dir,
        [
            "quarterly_predictions.csv",
            "all_quarterly_predictions.csv",
            "base_model_quarterly_predictions.csv",
        ],
    )


def normalise_prediction_chunk(chunk: pd.DataFrame, run_id: str) -> pd.DataFrame:
    norm_map = {c: normalise_colname(c) for c in chunk.columns}
    df = chunk.rename(columns={c: norm_map[c] for c in chunk.columns}).copy()

    # Aliases.
    aliases = {
        "target": "actual",
        "prediction": "pred",
        "forecast": "pred",
        "y": "actual",
        "yhat": "pred",
        "period": "target_period",
        "quarter": "target_period",
        "target_quarter": "target_period",
        "forecast_origin": "origin",
    }
    for old, new in aliases.items():
        if old in df.columns and new not in df.columns:
            df = df.rename(columns={old: new})

    required = ["stream", "model", "origin", "target_period", "horizon", "actual", "pred"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Prediction chunk missing required columns: {missing}")

    out = df[required].copy()
    out["run_id"] = run_id
    out["stream"] = out["stream"].map(normalise_stream)
    out["model"] = out["model"].astype(str)
    out["model_uid"] = out["run_id"].astype(str) + "::" + out["model"].astype(str)
    out["origin"] = out["origin"].astype(str)
    out["target_period"] = out["target_period"].astype(str)
    out["horizon"] = pd.to_numeric(out["horizon"], errors="coerce")
    out["actual"] = pd.to_numeric(out["actual"], errors="coerce")
    out["pred"] = pd.to_numeric(out["pred"], errors="coerce")
    return out


def load_selected_predictions(run_specs: Sequence[RunSpec], registry: pd.DataFrame) -> pd.DataFrame:
    if registry.empty:
        return pd.DataFrame()

    frames = []
    by_run = registry.groupby("run_id")
    for spec in run_specs:
        if spec.run_id not in by_run.groups:
            continue
        run_models = set(registry.loc[by_run.groups[spec.run_id], "model"].astype(str))
        if not run_models:
            continue
        p = prediction_file_for_run(spec.path)
        if p is None:
            print(f"No quarterly prediction file found for {spec.run_id}; skipping prediction load.")
            continue

        print(f"Loading selected predictions from {spec.run_id}: {p.name}; selected models={len(run_models):,}")
        chunks = []
        try:
            for chunk in pd.read_csv(p, chunksize=300_000):
                # Fast filter using the original model column if present.
                model_cols = [c for c in chunk.columns if normalise_colname(c) == "model"]
                if not model_cols:
                    continue
                model_col = model_cols[0]
                sub = chunk[chunk[model_col].astype(str).isin(run_models)]
                if sub.empty:
                    continue
                norm = normalise_prediction_chunk(sub, spec.run_id)
                chunks.append(norm)
        except Exception as exc:
            print(f"WARNING: failed to load predictions from {p}: {type(exc).__name__}: {exc}")
            continue

        if chunks:
            df = pd.concat(chunks, ignore_index=True)
            frames.append(df)
            print(f"  loaded {len(df):,} prediction rows from {spec.run_id}")
        else:
            print(f"  no selected prediction rows found in {spec.run_id}")

    if not frames:
        return pd.DataFrame()

    pred = pd.concat(frames, ignore_index=True)
    pred = pred.dropna(subset=["actual", "pred", "horizon"])
    pred = pred[np.isfinite(pred["actual"]) & np.isfinite(pred["pred"]) & (pred["actual"].abs() > 1e-12)]
    pred["target_sort"] = pred["target_period"].map(period_sort_value)
    pred["origin_sort"] = pred["origin"].map(period_sort_value)
    return pred.reset_index(drop=True)


# =============================================================================
# 6. Metric computation
# =============================================================================

def annualise_quarterly_predictions(pred: pd.DataFrame) -> pd.DataFrame:
    if pred.empty:
        return pd.DataFrame()

    df = pred.copy()
    df["june_year"] = df["target_period"].map(june_year_from_period)
    df = df[df["june_year"].notna()].copy()
    df["target_period_obj"] = df["target_period"].map(parse_period)

    rows = []
    group_cols = ["stream", "model_uid", "model", "run_id", "origin", "june_year"]
    for keys, g in df.groupby(group_cols, dropna=False):
        stream, model_uid, model, run_id, origin, june_year = keys
        if g["target_period"].nunique() < 4:
            continue
        agg = ANNUAL_AGG.get(stream, "sum")
        if agg == "mean":
            actual = float(g["actual"].mean())
            predv = float(g["pred"].mean())
        else:
            actual = float(g["actual"].sum())
            predv = float(g["pred"].sum())
        rows.append(
            {
                "stream": stream,
                "model_uid": model_uid,
                "model": model,
                "run_id": run_id,
                "origin": origin,
                "june_year": int(june_year),
                "actual": actual,
                "pred": predv,
                "n_quarters": int(g["target_period"].nunique()),
            }
        )
    return pd.DataFrame(rows)


def compute_metrics_for_predictions(pred: pd.DataFrame, annual_pred: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    if pred.empty:
        return pd.DataFrame()
    if annual_pred is None:
        annual_pred = annualise_quarterly_predictions(pred)

    rows = []
    for model_uid, g in pred.groupby("model_uid"):
        stream = g["stream"].iloc[0]
        model = g["model"].iloc[0]
        run_id = g["run_id"].iloc[0]
        annual_g = annual_pred[annual_pred["model_uid"] == model_uid] if not annual_pred.empty else pd.DataFrame()

        row = {
            "stream": stream,
            "stream_label": stream_label(stream),
            "model_uid": model_uid,
            "model": model,
            "model_short": short_model_name(model),
            "run_id": run_id,
            "n_quarterly_pairs": int(len(g)),
            "n_origins": int(g["origin"].nunique()),
            "quarterly_mape": mape(g["actual"], g["pred"]),
            "quarterly_bias_pct": bias_pct(g["actual"], g["pred"]),
            "quarterly_p90_ape": p90_ape(g["actual"], g["pred"]),
            "quarterly_rmse": rmse(g["actual"], g["pred"]),
            "n_annual_pairs": int(len(annual_g)),
            "annual_mape": mape(annual_g["actual"], annual_g["pred"]) if not annual_g.empty else np.nan,
            "annual_bias_pct": bias_pct(annual_g["actual"], annual_g["pred"]) if not annual_g.empty else np.nan,
            "annual_p90_ape": p90_ape(annual_g["actual"], annual_g["pred"]) if not annual_g.empty else np.nan,
        }

        # Horizon buckets.
        for lo, hi, name in [(1, 4, "h1_4"), (5, 8, "h5_8"), (9, 12, "h9_12")]:
            sub = g[g["horizon"].between(lo, hi)]
            row[f"{name}_mape"] = mape(sub["actual"], sub["pred"]) if not sub.empty else np.nan
            row[f"{name}_n"] = int(len(sub))

        # Stress windows.
        for wname, (start, end) in STRESS_WINDOWS.items():
            mask = g["target_period"].map(lambda p, s=start, e=end: in_period_window(p, s, e))
            sub = g[mask]
            row[f"{wname}_mape"] = mape(sub["actual"], sub["pred"]) if not sub.empty else np.nan
            row[f"{wname}_bias_pct"] = bias_pct(sub["actual"], sub["pred"]) if not sub.empty else np.nan
            row[f"{wname}_n"] = int(len(sub))

        rows.append(row)

    out = pd.DataFrame(rows)
    return add_governance_scores(out)


def add_governance_scores(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out["abs_quarterly_bias"] = out["quarterly_bias_pct"].abs()
    out["abs_annual_bias"] = out.get("annual_bias_pct", np.nan).abs()

    parts = []
    weights = []
    score_spec = [
        ("quarterly_mape", 0.35),
        ("annual_mape", 0.25),
        ("quarterly_p90_ape", 0.15),
        ("abs_quarterly_bias", 0.15),
        ("policy_2022_23_mape", 0.10),
    ]
    for col, w in score_spec:
        if col in out.columns and out[col].notna().sum() > 0:
            parts.append(out[col].rank(pct=True, na_option="bottom") * w)
            weights.append(w)
    if parts:
        out["governance_score"] = sum(parts) / sum(weights)
    else:
        out["governance_score"] = np.nan
    return out


# =============================================================================
# 7. Common prediction matrix and optimisation
# =============================================================================

def common_prediction_matrix(pred: pd.DataFrame, stream: str, model_uids: Sequence[str]) -> Tuple[pd.DataFrame, pd.Series, pd.DataFrame]:
    key_cols = ["stream", "origin", "target_period", "horizon"]
    sub = pred[(pred["stream"] == stream) & (pred["model_uid"].isin(model_uids))].copy()
    if sub.empty:
        return pd.DataFrame(), pd.Series(dtype=float), pd.DataFrame()

    pivot = sub.pivot_table(index=key_cols, columns="model_uid", values="pred", aggfunc="first")
    actual = sub.groupby(key_cols)["actual"].first()
    pivot = pivot.reindex(columns=list(model_uids))
    common = pivot.dropna(axis=0, how="any")
    if common.empty:
        return pd.DataFrame(), pd.Series(dtype=float), pd.DataFrame()
    actual = actual.loc[common.index]
    keys = pd.DataFrame(index=common.index).reset_index()
    return common.reset_index(drop=True), actual.reset_index(drop=True), keys


def annual_matrix_from_quarterly_matrix(
    P: pd.DataFrame,
    y: pd.Series,
    keys: pd.DataFrame,
    stream: str,
) -> Tuple[np.ndarray, np.ndarray, pd.DataFrame]:
    if P.empty or keys.empty:
        return np.empty((0, P.shape[1])), np.empty((0,)), pd.DataFrame()
    df = keys.copy()
    df["actual"] = y.values
    df["june_year"] = df["target_period"].map(june_year_from_period)
    df = df[df["june_year"].notna()].copy()
    if df.empty:
        return np.empty((0, P.shape[1])), np.empty((0,)), pd.DataFrame()

    P_arr = P.values
    p_df = pd.DataFrame(P_arr, columns=P.columns)
    p_df["origin"] = keys["origin"].values
    p_df["june_year"] = keys["target_period"].map(june_year_from_period).values
    p_df["target_period"] = keys["target_period"].values
    p_df["actual"] = y.values

    rows = []
    parr = []
    yarr = []
    for (origin, jy), g in p_df.groupby(["origin", "june_year"]):
        if g["target_period"].nunique() < 4:
            continue
        if ANNUAL_AGG.get(stream, "sum") == "mean":
            parr.append(g[P.columns].mean(axis=0).values)
            yarr.append(float(g["actual"].mean()))
        else:
            parr.append(g[P.columns].sum(axis=0).values)
            yarr.append(float(g["actual"].sum()))
        rows.append({"origin": origin, "june_year": int(jy), "n_quarters": int(g["target_period"].nunique())})
    if not parr:
        return np.empty((0, P.shape[1])), np.empty((0,)), pd.DataFrame()
    return np.vstack(parr), np.asarray(yarr, dtype=float), pd.DataFrame(rows)


def build_objective_matrix(
    P: pd.DataFrame,
    y: pd.Series,
    keys: pd.DataFrame,
    stream: str,
    profile: str,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Build row-stacked matrix for LP objective."""
    matrices = []
    targets = []
    weights = []

    # Quarterly rows.
    Pq = P.values.astype(float)
    yq = y.values.astype(float)
    matrices.append(Pq)
    targets.append(yq)
    q_weight = 1.0
    weights.append(np.ones(len(yq)) * q_weight)

    # Annual rows.
    Pa, ya, _ = annual_matrix_from_quarterly_matrix(P, y, keys, stream)
    if Pa.shape[0] > 0:
        if profile in {"quarterly_annual", "governance", "annual_weighted", "stress_aware", "bias_constrained"}:
            a_weight = 0.60 if profile != "annual_weighted" else 1.25
            matrices.append(Pa)
            targets.append(ya)
            weights.append(np.ones(len(ya)) * a_weight)

    # Stress rows.
    if profile in {"stress_aware", "governance"}:
        for wname, (start, end) in STRESS_WINDOWS.items():
            mask = keys["target_period"].map(lambda p, s=start, e=end: in_period_window(p, s, e)).values
            if mask.sum() > 0:
                stress_weight = 0.50 if wname == "policy_2022_23" else 0.30
                matrices.append(Pq[mask])
                targets.append(yq[mask])
                weights.append(np.ones(mask.sum()) * stress_weight)

    P_obj = np.vstack(matrices) if matrices else Pq
    y_obj = np.concatenate(targets) if targets else yq
    row_w = np.concatenate(weights) if weights else np.ones(len(yq))
    return P_obj, y_obj, row_w


def optimise_convex_l1_weights(
    P: np.ndarray,
    y: np.ndarray,
    row_weight: Optional[np.ndarray] = None,
    max_weight: Optional[float] = None,
    bias_penalty: float = 0.0,
    l2_penalty: float = 0.0,
) -> np.ndarray:
    """Exact LP for weighted relative L1 error, with optional absolute bias penalty.

    Objective is sum_i row_w_i * |P_i w - y_i| / |y_i| + bias_penalty * |mean rel bias|.
    L2 penalty is handled by SLSQP fallback because it is quadratic; set to 0 for exact LP.
    """
    P = np.asarray(P, dtype=float)
    y = np.asarray(y, dtype=float)
    n, k = P.shape
    if n == 0 or k == 0:
        return np.ones(k) / max(k, 1)

    mask = np.isfinite(y) & (np.abs(y) > 1e-12) & np.isfinite(P).all(axis=1)
    P = P[mask]
    y = y[mask]
    if row_weight is None:
        row_weight = np.ones(n)
    row_weight = np.asarray(row_weight, dtype=float)[mask]
    n, k = P.shape
    if n == 0:
        return np.ones(k) / k

    if max_weight is None:
        max_weight = 1.0
    if max_weight * k < 1.0:
        max_weight = 1.0

    # If a quadratic regulariser is requested, use SLSQP directly.
    if l2_penalty and l2_penalty > 0:
        return optimise_slsqp_weights(P, y, row_weight, max_weight=max_weight, bias_penalty=bias_penalty, l2_penalty=l2_penalty)

    include_bias = bias_penalty > 0
    num_bias_vars = 1 if include_bias else 0

    # Variables: w_k, e_n, [b]
    total_vars = k + n + num_bias_vars
    c = np.zeros(total_vars)
    c[k : k + n] = row_weight / np.maximum(np.abs(y), 1e-12)
    if include_bias:
        c[-1] = bias_penalty

    # Absolute error constraints.
    A_ub = []
    b_ub = []

    # P w - e <= y
    A = np.zeros((n, total_vars))
    A[:, :k] = P
    A[:, k : k + n] = -np.eye(n)
    A_ub.append(A)
    b_ub.append(y)

    # -P w - e <= -y
    A = np.zeros((n, total_vars))
    A[:, :k] = -P
    A[:, k : k + n] = -np.eye(n)
    A_ub.append(A)
    b_ub.append(-y)

    # Bias absolute value constraints: |mean((Pw-y)/|y|)| <= b
    if include_bias:
        rel_denom = np.maximum(np.abs(y), 1e-12)
        coef = np.mean(P / rel_denom[:, None], axis=0)
        const = np.mean(y / rel_denom)
        A = np.zeros((1, total_vars))
        A[0, :k] = coef
        A[0, -1] = -1.0
        A_ub.append(A)
        b_ub.append(np.array([const]))

        A = np.zeros((1, total_vars))
        A[0, :k] = -coef
        A[0, -1] = -1.0
        A_ub.append(A)
        b_ub.append(np.array([-const]))

    A_ub_arr = np.vstack(A_ub)
    b_ub_arr = np.concatenate(b_ub)

    A_eq = np.zeros((1, total_vars))
    A_eq[0, :k] = 1.0
    b_eq = np.array([1.0])

    bounds = [(0.0, max_weight)] * k + [(0.0, None)] * n
    if include_bias:
        bounds.append((0.0, None))

    res = linprog(
        c=c,
        A_ub=A_ub_arr,
        b_ub=b_ub_arr,
        A_eq=A_eq,
        b_eq=b_eq,
        bounds=bounds,
        method="highs",
    )
    if res.success:
        w = np.maximum(np.asarray(res.x[:k], dtype=float), 0)
        if w.sum() > 0:
            return w / w.sum()

    # Fallback.
    return optimise_slsqp_weights(P, y, row_weight, max_weight=max_weight, bias_penalty=bias_penalty, l2_penalty=l2_penalty)


def optimise_slsqp_weights(
    P: np.ndarray,
    y: np.ndarray,
    row_weight: np.ndarray,
    max_weight: float = 1.0,
    bias_penalty: float = 0.0,
    l2_penalty: float = 0.0,
) -> np.ndarray:
    P = np.asarray(P, dtype=float)
    y = np.asarray(y, dtype=float)
    n, k = P.shape
    if n == 0 or k == 0:
        return np.ones(k) / max(k, 1)
    row_weight = np.asarray(row_weight, dtype=float)

    def obj(w: np.ndarray) -> float:
        pred = P @ w
        rel_abs = np.abs(pred - y) / np.maximum(np.abs(y), 1e-12)
        value = float(np.mean(row_weight * rel_abs))
        if bias_penalty > 0:
            bias = float(np.mean((pred - y) / np.maximum(np.abs(y), 1e-12)))
            value += bias_penalty * abs(bias)
        if l2_penalty > 0:
            value += l2_penalty * float(np.sum(w * w))
        return value

    cons = [{"type": "eq", "fun": lambda w: float(np.sum(w) - 1.0)}]
    bounds = [(0.0, max_weight)] * k
    x0 = np.ones(k) / k
    res = minimize(obj, x0=x0, bounds=bounds, constraints=cons, method="SLSQP", options={"maxiter": 2000})
    if res.success:
        w = np.maximum(np.asarray(res.x, dtype=float), 0)
        if w.sum() > 0:
            return w / w.sum()
    return x0


def build_ensemble_predictions(
    stream: str,
    name: str,
    P: pd.DataFrame,
    y: pd.Series,
    keys: pd.DataFrame,
    weights: np.ndarray,
    method: str,
    components: Sequence[str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    pred_values = P.values @ weights
    out = keys.copy()
    out["stream"] = stream
    out["model"] = name
    out["model_uid"] = f"RESCUE::{name}"
    out["run_id"] = "RESCUE"
    out["actual"] = y.values
    out["pred"] = pred_values
    out["method"] = method

    wdf = pd.DataFrame(
        {
            "stream": stream,
            "ensemble": name,
            "ensemble_uid": f"RESCUE::{name}",
            "component_model_uid": list(components),
            "weight": weights,
            "method": method,
        }
    )
    return out, wdf


def build_static_ensembles_for_stream(
    pred: pd.DataFrame,
    metrics: pd.DataFrame,
    stream: str,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    s_metrics = metrics[metrics["stream"] == stream].copy()
    if s_metrics.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Multi-criteria pool.
    pools: List[str] = []
    for col in ["quarterly_mape", "annual_mape", "governance_score", "quarterly_p90_ape", "abs_quarterly_bias"]:
        if col in s_metrics.columns:
            pools.extend(s_metrics.sort_values(col, na_position="last")["model_uid"].head(TOP_N_FOR_ENSEMBLE).tolist())
    # Pure Schiff and known solver candidates for anchor diversity.
    pools.extend(s_metrics[s_metrics["model"].str.contains("SCHIFF_OLS", case=False, na=False)]["model_uid"].tolist())
    pools.extend(s_metrics[s_metrics["model"].str.contains("solver", case=False, na=False)]["model_uid"].tolist())
    pools = list(dict.fromkeys([p for p in pools if isinstance(p, str)]))

    ens_frames = []
    weight_frames = []

    for k in ENSEMBLE_POOL_SIZES:
        component_pool = pools[: min(k, len(pools))]
        if len(component_pool) < 2:
            continue
        P, y, keys = common_prediction_matrix(pred, stream, component_pool)
        if P.empty or P.shape[1] < 2:
            continue

        profiles = [
            ("quarterly", "static_q", 0.0, 0.0, None),
            ("quarterly_annual", "static_q_annual", 0.0, 0.0, None),
            ("governance", "static_governance", 0.50, 0.0, None),
            ("annual_weighted", "static_annual_weighted", 0.10, 0.0, None),
            ("stress_aware", "static_stress_aware", 0.25, 0.0, None),
            ("bias_constrained", "static_bias_penalty", 1.00, 0.0, None),
        ]
        for profile, method, bias_pen, l2_pen, cap_override in profiles:
            P_obj, y_obj, row_w = build_objective_matrix(P, y, keys, stream, profile)
            for cap in MAX_WEIGHT_CAPS:
                cap_to_use = cap_override if cap_override is not None else cap
                w = optimise_convex_l1_weights(P_obj, y_obj, row_w, max_weight=cap_to_use, bias_penalty=bias_pen, l2_penalty=l2_pen)
                cap_label = "none" if cap_to_use is None else str(cap_to_use).replace(".", "p")
                name = f"{stream}__RESCUE_{method}_top{k}_cap{cap_label}"
                ens, wdf = build_ensemble_predictions(stream, name, P, y, keys, w, method, component_pool)
                ens_frames.append(ens)
                weight_frames.append(wdf)

        # Simple robust ensembles.
        arr = P.values
        for method in ["top_mean", "top_median", "top_trimmed_mean"]:
            if method == "top_mean":
                pred_values = np.mean(arr, axis=1)
            elif method == "top_median":
                pred_values = np.median(arr, axis=1)
            else:
                if arr.shape[1] >= 4:
                    sorted_arr = np.sort(arr, axis=1)
                    pred_values = sorted_arr[:, 1:-1].mean(axis=1)
                else:
                    pred_values = arr.mean(axis=1)
            name = f"{stream}__RESCUE_{method}_top{k}"
            out = keys.copy()
            out["stream"] = stream
            out["model"] = name
            out["model_uid"] = f"RESCUE::{name}"
            out["run_id"] = "RESCUE"
            out["actual"] = y.values
            out["pred"] = pred_values
            out["method"] = method
            ens_frames.append(out)

    return (
        pd.concat(ens_frames, ignore_index=True) if ens_frames else pd.DataFrame(),
        pd.concat(weight_frames, ignore_index=True) if weight_frames else pd.DataFrame(),
    )


def build_prequential_ensembles_for_stream(
    pred: pd.DataFrame,
    metrics: pd.DataFrame,
    stream: str,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    s_metrics = metrics[metrics["stream"] == stream].copy()
    if s_metrics.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Choose a compact, diverse pool.
    pool = []
    for col in ["quarterly_mape", "annual_mape", "governance_score", "abs_quarterly_bias"]:
        if col in s_metrics.columns:
            pool.extend(s_metrics.sort_values(col, na_position="last")["model_uid"].head(10).tolist())
    pool.extend(s_metrics[s_metrics["model"].str.contains("solver|SCHIFF_OLS", case=False, na=False)]["model_uid"].tolist())
    pool = list(dict.fromkeys(pool))[:25]
    if len(pool) < 2:
        return pd.DataFrame(), pd.DataFrame()

    ens_frames = []
    w_records = []

    for k in [8, 12, 18]:
        components = pool[: min(k, len(pool))]
        P, y, keys = common_prediction_matrix(pred, stream, components)
        if P.empty:
            continue
        dfm = keys.copy()
        for m in components:
            dfm[m] = P[m].values
        dfm["actual"] = y.values
        dfm["origin_sort"] = dfm["origin"].map(period_sort_value)
        dfm = dfm.sort_values(["origin_sort", "horizon", "target_period"]).reset_index(drop=True)

        pred_values = []
        for _, row in dfm.iterrows():
            hist = dfm[dfm["origin_sort"] < row["origin_sort"]]
            if len(hist) >= MIN_PREQ_HISTORY_ROWS:
                P_hist = hist[components]
                y_hist = hist["actual"]
                P_obj, y_obj, row_w = build_objective_matrix(P_hist, y_hist, hist[["stream", "origin", "target_period", "horizon"]], stream, "governance")
                w = optimise_convex_l1_weights(P_obj, y_obj, row_w, max_weight=0.60, bias_penalty=0.50)
            else:
                w = np.ones(len(components)) / len(components)
            p_row = row[components].values.astype(float)
            pred_values.append(float(p_row @ w))
            for comp, weight in zip(components, w):
                if abs(weight) > 1e-9:
                    w_records.append(
                        {
                            "stream": stream,
                            "ensemble": f"{stream}__RESCUE_preq_governance_top{k}",
                            "ensemble_uid": f"RESCUE::{stream}__RESCUE_preq_governance_top{k}",
                            "component_model_uid": comp,
                            "weight": float(weight),
                            "method": "prequential_governance",
                            "origin": row["origin"],
                        }
                    )
        out = keys.copy()
        out["stream"] = stream
        out["model"] = f"{stream}__RESCUE_preq_governance_top{k}"
        out["model_uid"] = "RESCUE::" + out["model"]
        out["run_id"] = "RESCUE"
        out["actual"] = y.values
        out["pred"] = pred_values
        out["method"] = "prequential_governance"
        ens_frames.append(out)

    return (
        pd.concat(ens_frames, ignore_index=True) if ens_frames else pd.DataFrame(),
        pd.DataFrame(w_records),
    )


# =============================================================================
# 8. Comparisons and summaries
# =============================================================================

def paired_comparisons(all_pred: pd.DataFrame, final_metrics: pd.DataFrame) -> pd.DataFrame:
    rows = []
    key_cols = ["stream", "origin", "target_period", "horizon"]

    for stream in STREAMS:
        s_metrics = final_metrics[final_metrics["stream"] == stream].copy()
        if s_metrics.empty:
            continue
        # Find pure Schiff baseline and best models.
        schiff_candidates = s_metrics[s_metrics["model"].map(is_pure_schiff_model)]
        if schiff_candidates.empty:
            schiff_candidates = s_metrics[s_metrics["model"].str.contains("SCHIFF_OLS", case=False, na=False)]
        if schiff_candidates.empty:
            continue
        baseline_uid = schiff_candidates.sort_values("quarterly_mape", na_position="last").iloc[0]["model_uid"]

        challenger_uids = s_metrics.sort_values("governance_score", na_position="last")["model_uid"].head(25).tolist()
        if baseline_uid not in challenger_uids:
            challenger_uids.append(baseline_uid)

        base = all_pred[all_pred["model_uid"] == baseline_uid][key_cols + ["actual", "pred"]].rename(
            columns={"pred": "pred_baseline"}
        )
        for uid in challenger_uids:
            if uid == baseline_uid:
                continue
            ch = all_pred[all_pred["model_uid"] == uid][key_cols + ["actual", "pred"]].rename(columns={"pred": "pred_challenger"})
            m = base.merge(ch, on=key_cols + ["actual"], how="inner")
            if m.empty:
                continue
            base_ape = np.abs(m["pred_baseline"] - m["actual"]) / np.abs(m["actual"])
            ch_ape = np.abs(m["pred_challenger"] - m["actual"]) / np.abs(m["actual"])
            rows.append(
                {
                    "stream": stream,
                    "baseline_model_uid": baseline_uid,
                    "challenger_model_uid": uid,
                    "n_common_pairs": int(len(m)),
                    "baseline_mape": float(base_ape.mean() * 100.0),
                    "challenger_mape": float(ch_ape.mean() * 100.0),
                    "mape_improvement_pct_points": float((base_ape.mean() - ch_ape.mean()) * 100.0),
                    "challenger_win_rate": float((ch_ape < base_ape).mean() * 100.0),
                }
            )
    return pd.DataFrame(rows)


def stress_table(metrics: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if metrics.empty:
        return pd.DataFrame()
    for _, r in metrics.iterrows():
        for bucket, col in [
            ("1-4 qtrs", "h1_4_mape"),
            ("5-8 qtrs", "h5_8_mape"),
            ("9-12 qtrs", "h9_12_mape"),
            ("2024+", "recent_2024_plus_mape"),
            ("2022-23", "policy_2022_23_mape"),
            ("Annual", "annual_mape"),
        ]:
            rows.append(
                {
                    "stream": r.get("stream"),
                    "stream_label": r.get("stream_label"),
                    "model_uid": r.get("model_uid"),
                    "model": r.get("model"),
                    "model_short": r.get("model_short"),
                    "stress_bucket": bucket,
                    "mape": r.get(col, np.nan),
                }
            )
    return pd.DataFrame(rows)


def select_recommendations(metrics: pd.DataFrame) -> pd.DataFrame:
    if metrics.empty:
        return pd.DataFrame()
    rows = []
    for stream, g in metrics.groupby("stream"):
        gg = g.copy()
        # Prefer rescue/constrained ensembles if they are materially competitive.
        gg = gg.sort_values(["governance_score", "quarterly_mape", "annual_mape"], na_position="last")
        rows.append(gg.iloc[0])
    return pd.DataFrame(rows).reset_index(drop=True)


def compare_to_references(metrics: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if metrics.empty:
        return pd.DataFrame()
    rec = select_recommendations(metrics)
    for _, r in rec.iterrows():
        stream = r["stream"]
        for ref_name, ref in REFERENCE_EXPECTED.items():
            if stream not in ref:
                continue
            rows.append(
                {
                    "stream": stream,
                    "stream_label": stream_label(stream),
                    "recommended_model": r["model"],
                    "reference": ref_name,
                    "recommended_quarterly_mape": r["quarterly_mape"],
                    "reference_quarterly_mape": ref[stream].get("quarterly_mape", np.nan),
                    "quarterly_difference_pp": r["quarterly_mape"] - ref[stream].get("quarterly_mape", np.nan),
                    "recommended_annual_mape": r["annual_mape"],
                    "reference_annual_mape": ref[stream].get("annual_mape", np.nan),
                    "annual_difference_pp": r["annual_mape"] - ref[stream].get("annual_mape", np.nan),
                }
            )
    return pd.DataFrame(rows)


# =============================================================================
# 9. Charts
# =============================================================================

def save_recommendation_chart(rec: pd.DataFrame, out_dir: Path) -> None:
    if rec.empty:
        return
    fig, ax = plt.subplots(figsize=(9, 5))
    labels = [stream_label(s) for s in rec["stream"]]
    x = np.arange(len(labels))
    width = 0.36
    q = rec["quarterly_mape"].values
    a = rec["annual_mape"].values
    ax.bar(x - width / 2, q, width, label="Quarterly MAPE")
    ax.bar(x + width / 2, a, width, label="Annual MAPE")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=0)
    ax.set_ylabel("MAPE (%)")
    ax.set_title("Candidate-rescue recommended finalists")
    ax.legend()
    for i, val in enumerate(q):
        ax.text(i - width / 2, val + 0.05, f"{val:.2f}%", ha="center", va="bottom", fontsize=8)
    for i, val in enumerate(a):
        ax.text(i + width / 2, val + 0.05, f"{val:.2f}%", ha="center", va="bottom", fontsize=8)
    fig.tight_layout()
    fig.savefig(out_dir / "figures" / "recommended_finalists_mape.png", dpi=170)
    plt.close(fig)


def save_candidate_landscape(metrics: pd.DataFrame, rec: pd.DataFrame, out_dir: Path) -> None:
    if metrics.empty:
        return
    fig, ax = plt.subplots(figsize=(9, 6))
    colors = {"PED": "tab:blue", "LIGHT_RUC": "tab:orange", "HEAVY_RUC": "tab:green"}
    for stream, g in metrics.groupby("stream"):
        g = g.dropna(subset=["quarterly_mape", "annual_mape"])
        if g.empty:
            continue
        ax.scatter(g["quarterly_mape"], g["annual_mape"], s=26, alpha=0.35, label=stream_label(stream), color=colors.get(stream))
    if not rec.empty:
        for _, r in rec.iterrows():
            ax.scatter(r["quarterly_mape"], r["annual_mape"], marker="*", s=220, color=colors.get(r["stream"], "black"), edgecolor="black")
            ax.annotate(stream_label(r["stream"]).split()[0], (r["quarterly_mape"], r["annual_mape"]), xytext=(5, 5), textcoords="offset points", fontsize=8)
    schiff = metrics[metrics["model"].map(is_pure_schiff_model)]
    if not schiff.empty:
        ax.scatter(schiff["quarterly_mape"], schiff["annual_mape"], marker="^", s=90, facecolors="none", edgecolors="black", label="Pure Schiff")
    ax.set_xlabel("Quarterly MAPE (%)")
    ax.set_ylabel("Annual MAPE (%)")
    ax.set_title("Candidate rescue landscape")
    ax.grid(alpha=0.25)
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(out_dir / "figures" / "candidate_rescue_landscape.png", dpi=170)
    plt.close(fig)


def save_weight_chart(weights: pd.DataFrame, rec: pd.DataFrame, metrics: pd.DataFrame, out_dir: Path) -> None:
    if weights.empty or rec.empty:
        return
    rec_uids = set(rec["model_uid"].astype(str))
    w = weights[weights["ensemble_uid"].isin(rec_uids)].copy()
    if w.empty:
        # Try by ensemble name.
        rec_models = set(rec["model"].astype(str))
        w = weights[weights["ensemble"].isin(rec_models)].copy()
    if w.empty:
        return
    w = w[pd.to_numeric(w["weight"], errors="coerce").fillna(0) > 1e-8].copy()
    if w.empty:
        return
    lookup = metrics[["model_uid", "model_short"]].drop_duplicates().set_index("model_uid")["model_short"].to_dict()
    w["component_short"] = w["component_model_uid"].map(lookup).fillna(w["component_model_uid"].map(lambda x: short_model_name(str(x).split("::")[-1])))
    # Limit chart rows.
    w = w.sort_values(["stream", "ensemble", "weight"], ascending=[True, True, False]).groupby(["stream", "ensemble"], group_keys=False).head(8)
    streams = list(w["stream"].dropna().unique())
    fig, axes = plt.subplots(len(streams), 1, figsize=(10, 3.2 * max(1, len(streams))))
    if len(streams) == 1:
        axes = [axes]
    for ax, stream in zip(axes, streams):
        g = w[w["stream"] == stream].copy().sort_values("weight")
        labels = g["component_short"].tolist()
        ax.barh(labels, g["weight"] * 100)
        ax.set_title(stream_label(stream))
        ax.set_xlabel("Weight (%)")
        for i, val in enumerate(g["weight"] * 100):
            ax.text(val + 0.5, i, f"{val:.1f}%", va="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(out_dir / "figures" / "recommended_ensemble_weights.png", dpi=170)
    plt.close(fig)


# =============================================================================
# 10. Excel/report output
# =============================================================================

def write_excel(out_path: Path, tables: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in tables.items():
            if df is None or df.empty:
                continue
            sheet = re.sub(r"[^A-Za-z0-9 ]+", "", name)[:31] or "Sheet"
            df.to_excel(writer, sheet_name=sheet, index=False)
    # Light formatting without tables.
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(out_path)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top")
            for col_idx, cells in enumerate(ws.columns, start=1):
                max_len = 10
                for cell in cells:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(val), 70))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        wb.save(out_path)
    except Exception as exc:
        print(f"WARNING: Excel formatting skipped: {exc}")


def write_report(out_path: Path, rec: pd.DataFrame, refs: pd.DataFrame, metrics: pd.DataFrame, paired: pd.DataFrame) -> None:
    lines = []
    lines.append("# Stage 1 candidate-rescue + constrained-stacking audit\n")
    lines.append(f"Generated: {datetime.now().isoformat(timespec='seconds')}\n")

    lines.append("\n## Recommended finalists\n")
    if not rec.empty:
        cols = ["stream", "model", "run_id", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "quarterly_p90_ape", "governance_score"]
        lines.append(rec[[c for c in cols if c in rec.columns]].to_markdown(index=False))
    else:
        lines.append("No recommendations produced.")

    lines.append("\n\n## Comparison to references\n")
    if not refs.empty:
        lines.append(refs.to_markdown(index=False))
    else:
        lines.append("No reference comparison available.")

    lines.append("\n\n## Top candidates by stream\n")
    if not metrics.empty:
        for stream, g in metrics.sort_values("governance_score", na_position="last").groupby("stream"):
            lines.append(f"\n### {stream_label(stream)}\n")
            cols = ["model", "run_id", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "quarterly_p90_ape", "governance_score"]
            lines.append(g[[c for c in cols if c in g.columns]].head(15).to_markdown(index=False))
    else:
        lines.append("No candidate metrics available.")

    lines.append("\n\n## Paired-vs-Schiff check\n")
    if not paired.empty:
        lines.append(paired.head(60).to_markdown(index=False))
    else:
        lines.append("No paired comparison available.")

    lines.append("\n\n## Interpretation\n")
    lines.append(
        "This audit reuses existing predictions and optimises ensemble weights across rescued candidate pools. "
        "It is intended to detect whether strong candidates from prior runs were omitted from the final solver pool, "
        "and whether constrained stacking can improve the quarterly/annual/bias/stress trade-off without rerunning the broad search."
    )
    out_path.write_text("\n".join(lines), encoding="utf-8")


# =============================================================================
# 11. Main
# =============================================================================

def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Stage 1 candidate rescue + constrained stacking audit")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT), help="Output root folder")
    parser.add_argument("--run", action="append", default=[], help="Run spec as run_id=path. Can be repeated.")
    parser.add_argument("--top-n-by-metric", type=int, default=TOP_N_BY_METRIC)
    parser.add_argument("--max-candidates-per-stream", type=int, default=MAX_CANDIDATES_PER_STREAM_FOR_PRED_LOAD)
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> None:
    global TOP_N_BY_METRIC, MAX_CANDIDATES_PER_STREAM_FOR_PRED_LOAD
    args = parse_args(argv)
    TOP_N_BY_METRIC = args.top_n_by_metric
    MAX_CANDIDATES_PER_STREAM_FOR_PRED_LOAD = args.max_candidates_per_stream

    # Runs.
    run_specs: List[RunSpec] = []
    if args.run:
        for item in args.run:
            if "=" not in item:
                raise ValueError("--run must be run_id=path")
            run_id, path = item.split("=", 1)
            run_specs.append(RunSpec(run_id=run_id.strip(), path=Path(path.strip().strip('"'))))
    else:
        run_specs = [RunSpec(run_id=k, path=v) for k, v in DEFAULT_RUNS.items()]

    out_dir = make_output_dir(Path(args.output_root))
    print("=" * 100)
    print("STAGE 1 CANDIDATE RESCUE + CONSTRAINED STACKING AUDIT")
    print("=" * 100)
    print(f"Output folder: {out_dir}")
    print("Run folders:")
    for rs in run_specs:
        print(f"  {rs.run_id}: {rs.path} {'[OK]' if rs.path.exists() else '[missing]'}")

    t0 = time.time()

    # Load summaries and rescue candidate registry.
    summary = load_run_summaries(run_specs)
    if summary.empty:
        raise RuntimeError("No run summaries were loaded. Check run folder paths.")
    summary.to_csv(out_dir / "all_loaded_summaries.csv", index=False)

    registry = build_candidate_registry(summary)
    if registry.empty:
        raise RuntimeError("Candidate registry is empty.")
    registry.to_csv(out_dir / "rescued_candidate_registry.csv", index=False)
    print(f"Candidate registry: {len(registry):,} candidates")

    # Load selected predictions.
    pred = load_selected_predictions(run_specs, registry)
    if pred.empty:
        raise RuntimeError("No selected quarterly predictions were loaded. Need quarterly_predictions.csv/all_quarterly_predictions.csv in run folders.")
    pred.to_csv(out_dir / "selected_quarterly_predictions.csv", index=False)
    print(f"Selected quarterly predictions: {len(pred):,} rows, models={pred['model_uid'].nunique():,}")

    annual_pred = annualise_quarterly_predictions(pred)
    annual_pred.to_csv(out_dir / "selected_annual_predictions.csv", index=False)

    anchor_metrics = compute_metrics_for_predictions(pred, annual_pred)
    # Attach metadata where possible.
    meta_cols = ["model_uid", "source_family", "model_kind", "feature_set", "include_reason"]
    meta = registry[[c for c in meta_cols if c in registry.columns]].drop_duplicates("model_uid")
    if not meta.empty:
        anchor_metrics = anchor_metrics.merge(meta, on="model_uid", how="left")
    anchor_metrics = add_governance_scores(anchor_metrics)
    anchor_metrics.to_csv(out_dir / "anchor_candidate_metrics.csv", index=False)

    # Build constrained ensembles stream by stream.
    ens_frames = []
    weight_frames = []
    for stream in STREAMS:
        print(f"Building rescue ensembles for {stream}...")
        static_pred, static_w = build_static_ensembles_for_stream(pred, anchor_metrics, stream)
        preq_pred, preq_w = build_prequential_ensembles_for_stream(pred, anchor_metrics, stream)
        if not static_pred.empty:
            ens_frames.append(static_pred)
        if not preq_pred.empty:
            ens_frames.append(preq_pred)
        if not static_w.empty:
            weight_frames.append(static_w)
        if not preq_w.empty:
            weight_frames.append(preq_w)

    ensemble_pred = pd.concat(ens_frames, ignore_index=True) if ens_frames else pd.DataFrame()
    ensemble_weights = pd.concat(weight_frames, ignore_index=True) if weight_frames else pd.DataFrame()
    if not ensemble_pred.empty:
        ensemble_pred.to_csv(out_dir / "generated_ensemble_quarterly_predictions.csv", index=False)
    if not ensemble_weights.empty:
        ensemble_weights.to_csv(out_dir / "ensemble_weights.csv", index=False)

    # Combine anchors + generated ensembles.
    all_pred = pd.concat([pred, ensemble_pred], ignore_index=True) if not ensemble_pred.empty else pred.copy()
    all_pred.to_csv(out_dir / "all_quarterly_predictions_selected_and_rescue.csv", index=False)
    all_annual = annualise_quarterly_predictions(all_pred)
    all_annual.to_csv(out_dir / "all_annual_predictions_selected_and_rescue.csv", index=False)

    final_metrics = compute_metrics_for_predictions(all_pred, all_annual)
    # Add source metadata for generated ensembles.
    final_metrics["source_family"] = final_metrics["run_id"].map(lambda x: "rescued_ensemble" if x == "RESCUE" else "anchor_candidate")
    # Restore anchor source metadata.
    meta2 = registry[["model_uid", "source_family", "model_kind", "feature_set", "include_reason"]].drop_duplicates("model_uid")
    final_metrics = final_metrics.merge(meta2, on="model_uid", how="left", suffixes=("", "_anchor"))
    final_metrics["source_family"] = np.where(final_metrics["run_id"] == "RESCUE", "rescued_ensemble", final_metrics["source_family_anchor"].fillna(final_metrics["source_family"]))
    final_metrics["model_kind"] = np.where(final_metrics["run_id"] == "RESCUE", final_metrics["model"].map(lambda m: "rescued_ensemble"), final_metrics.get("model_kind", ""))
    final_metrics["feature_set"] = np.where(final_metrics["run_id"] == "RESCUE", "ensemble", final_metrics.get("feature_set", ""))
    final_metrics = add_governance_scores(final_metrics)
    final_metrics.to_csv(out_dir / "final_summary.csv", index=False)

    recommendations = select_recommendations(final_metrics)
    recommendations.to_csv(out_dir / "final_recommendations.csv", index=False)
    recommendations.sort_values(["stream", "quarterly_mape"]).to_csv(out_dir / "final_recommendations_by_quarterly.csv", index=False)

    paired = paired_comparisons(all_pred, final_metrics)
    paired.to_csv(out_dir / "paired_vs_schiff.csv", index=False)

    stress = stress_table(final_metrics)
    stress.to_csv(out_dir / "stress_tests.csv", index=False)

    refs = compare_to_references(final_metrics)
    refs.to_csv(out_dir / "reference_comparison.csv", index=False)

    # Charts.
    save_recommendation_chart(recommendations, out_dir)
    save_candidate_landscape(final_metrics, recommendations, out_dir)
    save_weight_chart(ensemble_weights, recommendations, final_metrics, out_dir)

    # Excel and report.
    tables = {
        "Final Recommendations": recommendations,
        "Final Summary": final_metrics,
        "Anchor Candidate Metrics": anchor_metrics,
        "Rescued Candidate Registry": registry,
        "Paired vs Schiff": paired,
        "Stress Tests": stress,
        "Ensemble Weights": ensemble_weights,
        "Reference Comparison": refs,
    }
    write_excel(out_dir / "stage1_candidate_rescue_results.xlsx", tables)
    write_report(out_dir / "stage1_candidate_rescue_report.md", recommendations, refs, final_metrics, paired)

    elapsed = (time.time() - t0) / 60.0
    print("\n" + "=" * 100)
    print("RUN COMPLETE")
    print("=" * 100)
    print(f"Elapsed minutes: {elapsed:.1f}")
    print(f"Run folder: {out_dir}")
    print("\nRecommended finalists:")
    show_cols = ["stream", "model", "run_id", "quarterly_mape", "annual_mape", "quarterly_bias_pct", "quarterly_p90_ape", "governance_score"]
    print(recommendations[[c for c in show_cols if c in recommendations.columns]].to_markdown(index=False))
    print("\nKey files:")
    for name in [
        "final_recommendations.csv",
        "final_summary.csv",
        "paired_vs_schiff.csv",
        "stress_tests.csv",
        "ensemble_weights.csv",
        "reference_comparison.csv",
        "stage1_candidate_rescue_results.xlsx",
        "stage1_candidate_rescue_report.md",
    ]:
        print(out_dir / name)


if __name__ == "__main__":
    main()
