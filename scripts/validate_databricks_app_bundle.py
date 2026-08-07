"""Validate the slim Databricks App runtime bundle, fail-closed.

Checks the bundle produced by ``scripts/build_databricks_app_bundle.py``
against both the platform constraints (10 MiB per-file limit, no ``.git``,
no tests/docs/deliverables) and the repository's governed runtime contracts:
manifest hash integrity, CSV->Parquet replacement equivalence, Revenue Outlook
pack loads for both engines, replay-cache / policy-runtime / quarterly-display
currency, a real AppTest render of Revenue Outlook from the bundle, the XLSX
forecast extract, and value parity of the governed chart rows and extract
workbook between the bundle and the full source checkout.

Usage:

    python scripts/validate_databricks_app_bundle.py \
        --bundle build/databricks_app/app \
        --source .

No model fitting and no pack rebuilding happens here: everything is loaded
from committed content, exactly as the deployed app would.

The runtime probes run against DISPOSABLE COPIES of the bundle and of the
source checkout, never against either original. Rendering the app writes into
``artifacts/`` (the r2-ladder chart sources are rewritten on every render), so
probing in place would mutate the bundle after its hashes were verified and
the publish workflow would ship content that no longer matches its manifest.
Structure and manifest hashes are therefore re-verified after the probes, and
the source checkout's ``git status`` must be unchanged.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path, PurePosixPath

REPO_ROOT = Path(__file__).resolve().parents[1]

MANIFEST_NAME = "databricks_app_bundle_manifest.json"
MAX_FILE_BYTES = 10 * 1024 * 1024
FORBIDDEN_ROOTS = (".git", "tests", "docs", "deliverables")

PARITY_SERIES = (
    "total_nltf_net_revenue",
    "ped_volume",
    "ped_vkt_per_capita",
    "light_ruc_net_km",
)
PARITY_YEARS = (2026, 2030, 2031, 2050)
PARITY_TRACE = "Current finalist Base case"

# Runs inside a subprocess rooted at either the bundle or the source checkout;
# prints one JSON document between the marker lines. Loads everything the
# deployed app loads, renders Revenue Outlook via AppTest, and builds the
# governed XLSX extract.
_PROBE = r"""
import json, sys
from pathlib import Path

root = Path.cwd()
report = {"root": str(root), "errors": []}

def fail(msg):
    report["errors"].append(str(msg))

try:
    from model_dashboard.revenue_outlook import (
        CURRENT_REVENUE_OUTLOOK_DIR,
        load_revenue_outlook_pack,
    )
    from model_dashboard.official_vintage import bridge_vintage_id_from_manifest
    from model_dashboard.revenue_outlook_replay_cache import replay_cache_status
    from model_dashboard.revenue_outlook_policy_runtime import (
        policy_runtime_status,
        upstream_manifests,
    )
    from model_dashboard.revenue_outlook_series_coverage import (
        load_quarterly_display_pack,
    )

    packs = {}
    for engine, pack_dir in (
        ("ensemble", root / "data" / "current_revenue_outlook"),
        ("ar1", root / "data" / "engine_ar1" / "current_revenue_outlook"),
    ):
        pack = load_revenue_outlook_pack(pack_dir, repo_root=root)
        if pack is None:
            fail(f"{engine}: Revenue Outlook pack did not load from {pack_dir}")
            continue
        packs[engine] = pack
        for required in ("revenue_chart_rows", "future_revenue_forecasts",
                         "revenue_stack_components", "scenario_feature_lineage"):
            frame = getattr(pack, required)
            if frame is None or frame.empty:
                fail(f"{engine}: runtime table {required} is empty")

        pack_manifest, replay_manifest, uncertainty_manifest = upstream_manifests(
            engine, root
        )
        status, detail = replay_cache_status(
            engine=engine,
            pack_manifest=pack_manifest,
            bridge_vintage_id=bridge_vintage_id_from_manifest(pack_manifest, root),
            repo_root=root,
        )
        report[f"replay_cache_status_{engine}"] = status
        if status != "ok":
            fail(f"{engine}: replay cache status {status}: {detail}")
        status, detail = policy_runtime_status(engine=engine, repo_root=root)
        report[f"policy_runtime_status_{engine}"] = status
        if status != "ok":
            fail(f"{engine}: policy runtime status {status}: {detail}")

    try:
        quarterly = load_quarterly_display_pack(root)
        report["quarterly_display_rows"] = int(len(quarterly.quarterly_rows))
    except Exception as error:
        fail(f"quarterly display pack: {error}")

    # Governed chart-row values used for bundle-vs-source parity.
    parity = {}
    ar1 = packs.get("ar1")
    if ar1 is not None:
        rows = ar1.revenue_chart_rows
        subset = rows[
            rows["series_id"].isin(%(series)s)
            & rows["june_year"].isin(%(years)s)
            & (rows["trace_name"] == %(trace)r)
        ]
        for record in subset.itertuples():
            key = f"{record.series_id}|FY{int(record.june_year)}"
            parity[key] = None if record.value != record.value else float(record.value)
    report["chart_row_values"] = dict(sorted(parity.items()))

    # AppTest: render the app, then the Revenue Outlook page.
    from streamlit.testing.v1 import AppTest

    at = AppTest.from_file(str(root / "app.py"), default_timeout=600)
    at.run()
    if at.exception:
        fail(f"AppTest initial render raised: {at.exception[0].value}")
    elif not at.radio:
        fail("AppTest rendered no navigation radio")
    else:
        options = list(at.radio[0].options)
        target = "Revenue Outlook" if "Revenue Outlook" in options else options[-1]
        at.radio[0].set_value(target).run()
        if at.exception:
            fail(f"AppTest Revenue Outlook render raised: {at.exception[0].value}")
        report["apptest_page"] = target

    # XLSX extract from the governed default selection.
    import app as app_module
    from model_dashboard.revenue_outlook import (
        PED_BRIDGE_DEFAULT_MODE,
        revenue_outlook_signature,
    )
    from openpyxl import load_workbook
    import io

    engine_dir = root / "data" / "engine_ar1" / "current_revenue_outlook"
    signature = revenue_outlook_signature(engine_dir, root)
    key = app_module.selected_sensitivity_key("Off", "Off", "Off")
    uptake_key = (app_module.DEFAULT_EV_UPTAKE_MODE, (), (), 0, 0)
    traces = ("Actual", "BEFU26 official", "Current finalist Base case")
    result = app_module.cached_revenue_outlook_extract_bytes(
        signature, key, PED_BRIDGE_DEFAULT_MODE, uptake_key, traces,
        packs["ar1"],
    )
    workbook = load_workbook(io.BytesIO(result.workbook_bytes))
    cells = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header = {}
        for column in range(2, worksheet.max_column + 1):
            value = worksheet.cell(row=1, column=column).value
            if isinstance(value, (int, float)):
                header[int(value)] = column
        for year in %(years)s:
            column = header.get(year)
            if column is None:
                continue
            value = worksheet.cell(row=65, column=column).value
            cells[f"{sheet_name}|row65|FY{year}"] = (
                None if value is None else float(value)
            )
    report["extract_row65_values"] = dict(sorted(cells.items()))
    report["extract_sheets"] = list(workbook.sheetnames)
except Exception as error:  # noqa: BLE001 - a probe must report, not crash
    import traceback

    fail("probe exception: " + "".join(traceback.format_exception(error)))

print("===PROBE-JSON-BEGIN===")
print(json.dumps(report, sort_keys=True))
print("===PROBE-JSON-END===")
""" % {
    "series": repr(tuple(PARITY_SERIES)),
    "years": repr(tuple(PARITY_YEARS)),
    "trace": PARITY_TRACE,
}


class BundleValidationError(RuntimeError):
    pass


def sha256_of(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def walk_files(root: Path) -> list[str]:
    return sorted(
        path.relative_to(root).as_posix()
        for path in root.rglob("*")
        if path.is_file()
    )


def check_structure(bundle: Path, manifest: dict, errors: list[str]) -> None:
    on_disk = walk_files(bundle)
    manifest_paths = {record["path"] for record in manifest["files"]}

    for relative in on_disk:
        head = PurePosixPath(relative).parts[0]
        if head in FORBIDDEN_ROOTS:
            errors.append(f"forbidden path in bundle: {relative}")
        size = (bundle / relative).stat().st_size
        if size > MAX_FILE_BYTES:
            errors.append(f"bundle file over 10 MiB: {relative} ({size} bytes)")
        if size > int(manifest.get("warning_file_bytes", MAX_FILE_BYTES)):
            errors.append(f"bundle file over the warning threshold: {relative}")

    extra = set(on_disk) - manifest_paths - {MANIFEST_NAME}
    missing = manifest_paths - set(on_disk)
    for relative in sorted(extra):
        errors.append(f"file on disk but not in the bundle manifest: {relative}")
    for relative in sorted(missing):
        errors.append(f"file in the bundle manifest but not on disk: {relative}")


def check_manifest_hashes(bundle: Path, manifest: dict, errors: list[str]) -> None:
    for record in manifest["files"]:
        path = bundle / record["path"]
        if not path.is_file():
            continue  # already reported by check_structure
        if path.stat().st_size != int(record["size_bytes"]):
            errors.append(f"size mismatch vs manifest: {record['path']}")
        elif sha256_of(path) != record["sha256"]:
            errors.append(f"sha256 mismatch vs manifest: {record['path']}")


def check_required(bundle: Path, policy: dict, errors: list[str]) -> None:
    for required in policy["required_root_files"]:
        if not (bundle / required).is_file():
            errors.append(f"required bundle file missing: {required}")


_ABSENT_SPELLINGS = {"", "nan", "none", "null", "<na>"}


def _absent_normalised(value) -> str:
    import pandas as pd

    if pd.isna(value):
        return ""
    text = str(value)
    return "" if text.strip().lower() in _ABSENT_SPELLINGS else text


def check_replacements(
    bundle: Path, source: Path, policy: dict, errors: list[str]
) -> None:
    import numpy as np
    import pandas as pd

    for entry in policy["parquet_replacements"]:
        parquet_rel = entry["parquet"]
        csv_rel = entry["csv"]
        parquet_path = bundle / parquet_rel
        csv_path = source / csv_rel
        if not parquet_path.is_file():
            errors.append(f"parquet replacement missing from bundle: {parquet_rel}")
            continue
        if (bundle / csv_rel).exists():
            errors.append(f"replaced CSV still present in bundle: {csv_rel}")
        if not csv_path.is_file():
            errors.append(f"source CSV for equivalence check missing: {csv_rel}")
            continue

        parquet_frame = pd.read_parquet(parquet_path)
        csv_frame = pd.read_csv(csv_path)
        if list(parquet_frame.columns) != list(csv_frame.columns):
            errors.append(f"column mismatch {csv_rel} vs {parquet_rel}")
            continue
        if len(parquet_frame) != len(csv_frame):
            errors.append(
                f"row count mismatch {csv_rel} ({len(csv_frame)}) vs "
                f"{parquet_rel} ({len(parquet_frame)})"
            )
            continue
        for column in parquet_frame.columns:
            left = parquet_frame[column]
            right = csv_frame[column]
            if pd.api.types.is_float_dtype(left) or pd.api.types.is_float_dtype(right):
                left_values = pd.to_numeric(left, errors="coerce").to_numpy(dtype=float)
                right_values = pd.to_numeric(right, errors="coerce").to_numpy(dtype=float)
                both_nan = np.isnan(left_values) & np.isnan(right_values)
                close = np.isclose(
                    left_values, right_values, rtol=1e-12, atol=1e-12, equal_nan=True
                )
                if not bool(np.all(close | both_nan)):
                    errors.append(
                        f"value divergence in float column {column!r}: "
                        f"{csv_rel} vs {parquet_rel}"
                    )
                    break
            else:
                # The pack writer stringifies missing values as "nan" in
                # Parquet object columns while the CSV round-trip yields real
                # NaN; both spellings mean "absent", so normalise them before
                # comparing rather than flagging the serialization artifact.
                left_text = left.map(_absent_normalised)
                right_text = right.map(_absent_normalised)
                if not left_text.eq(right_text).all():
                    mismatch = int((~left_text.eq(right_text)).sum())
                    errors.append(
                        f"value divergence in column {column!r} "
                        f"({mismatch} cells): {csv_rel} vs {parquet_rel}"
                    )
                    break


def check_compiles(bundle: Path, errors: list[str]) -> None:
    """In-memory syntax check; never writes __pycache__ into the bundle."""
    sources = [bundle / "app.py", bundle / "sitecustomize.py"]
    for package in ("model_dashboard", "pipeline"):
        sources.extend(sorted((bundle / package).rglob("*.py")))
    for source in sources:
        if not source.is_file():
            continue
        try:
            compile(source.read_text(encoding="utf-8"), str(source), "exec")
        except SyntaxError as error:
            errors.append(f"bundle source failed to compile: {source.name}: {error}")


def tracked_relatives(source: Path) -> list[str]:
    """Tracked files of the source checkout, as bundle-style relative paths."""
    result = subprocess.run(
        ["git", "-C", str(source), "ls-files", "-z"],
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        raise BundleValidationError(
            f"git ls-files failed in {source}; the source probe needs a git checkout"
        )
    return sorted(
        entry.decode("utf-8") for entry in result.stdout.split(b"\x00") if entry
    )


def copy_for_probe(root: Path, relatives: list[str], destination: Path) -> None:
    for relative in relatives:
        source_path = root / relative
        if not source_path.is_file():
            continue
        target = destination / relative
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, target)


def run_isolated_probe(root: Path, relatives: list[str], label: str) -> dict:
    """Probe a DISPOSABLE copy, never the publishable bundle or the checkout.

    Rendering the app writes into ``artifacts/`` (the r2-ladder chart sources
    are rewritten on every render). Probing the real bundle would therefore
    mutate content after its manifest hashes were verified, and the publish
    workflow would ship a bundle that no longer matches its own manifest.
    """
    with tempfile.TemporaryDirectory(prefix=f"nltf-{label}-probe-") as scratch:
        workspace = Path(scratch) / label
        workspace.mkdir(parents=True)
        copy_for_probe(root, relatives, workspace)
        return run_probe(workspace)


def git_status(source: Path) -> str | None:
    result = subprocess.run(
        ["git", "-C", str(source), "status", "--porcelain"],
        capture_output=True,
        text=True,
        check=False,
    )
    return result.stdout if result.returncode == 0 else None


def probe_env(root: Path) -> dict:
    """The exact environment the probe subprocess runs under.

    A function rather than inline construction so the shutdown-abort regression
    test can assert against the very environment the gate uses, not a copy that
    could drift.
    """
    env = os.environ.copy()
    env["PYTHONPATH"] = str(root)
    env["NLTF_DISABLE_RUNTIME_PYARROW24"] = "1"
    # Keep probe imports from writing __pycache__ into the bundle, which the
    # structural check would then (correctly) reject as unmanifested files.
    env["PYTHONDONTWRITEBYTECODE"] = "1"
    env.setdefault("STREAMLIT_GATHER_USAGE_STATS", "false")
    # The app spawns daemon cache-warmer threads per AppTest run
    # (app.py: _start_revenue_outlook_cache_warmer). That is a warm-start
    # optimisation for a live server; in a one-shot probe it only burns CPU
    # and - measured, not theorised - the warmers were still inside native
    # numpy/pyarrow code at interpreter exit in EVERY instrumented run,
    # crashing teardown in about half of them: "terminate called without an
    # active exception", then SIGABRT or SIGSEGV, and a publish gate failing
    # on a coin flip (hosted runs 31219276724 and 31224208227; diagnosis under
    # artifacts/ci_optimisation/probe_matrix/). tests/conftest.py disables the
    # same thread for the same reason. Probe values are unaffected: the warmer
    # precomputes caches for views the probe computes on demand anyway.
    #
    # A hard assignment, not setdefault: there is no configuration in which a
    # deterministic gate should host a background warmer.
    env["REVENUE_OUTLOOK_CACHE_WARMER"] = "0"
    return env


def run_probe(root: Path) -> dict:
    env = probe_env(root)
    result = subprocess.run(
        [sys.executable, "-c", _PROBE],
        cwd=root,
        env=env,
        text=True,
        capture_output=True,
        encoding="utf-8",
        errors="replace",
    )
    output = result.stdout or ""
    begin = output.find("===PROBE-JSON-BEGIN===")
    end = output.find("===PROBE-JSON-END===")
    if result.returncode != 0 or begin < 0 or end < 0:
        raise BundleValidationError(
            f"runtime probe failed in {root}\nstdout tail:\n{output[-2000:]}\n"
            f"stderr tail:\n{(result.stderr or '')[-2000:]}"
        )
    payload = output[begin + len("===PROBE-JSON-BEGIN==="):end].strip()
    return json.loads(payload)


def compare_parity(bundle_report: dict, source_report: dict, errors: list[str]) -> dict:
    summary = {}
    for field in ("chart_row_values", "extract_row65_values"):
        bundle_values = bundle_report.get(field, {})
        source_values = source_report.get(field, {})
        if not bundle_values:
            errors.append(f"bundle probe produced no {field}")
            continue
        if bundle_values != source_values:
            diff_keys = sorted(
                set(bundle_values) ^ set(source_values)
                | {
                    key
                    for key in set(bundle_values) & set(source_values)
                    if bundle_values[key] != source_values[key]
                }
            )
            errors.append(f"{field} differ between bundle and source: {diff_keys[:10]}")
        summary[field] = len(bundle_values)
    return summary


def validate(
    bundle: Path,
    source: Path,
    policy_path: Path,
    *,
    skip_runtime: bool = False,
) -> tuple[list[str], dict, dict]:
    """(errors, parity summary, bundle manifest).

    Ordering matters: the runtime probes run against disposable copies, and
    the bundle's structure and manifest hashes are re-verified AFTERWARDS, so
    a probe that mutated its workspace can never leave the publishable bundle
    disagreeing with the manifest the publish workflow ships.
    """
    errors: list[str] = []
    manifest = json.loads((bundle / MANIFEST_NAME).read_text(encoding="utf-8"))
    policy = json.loads(policy_path.read_text(encoding="utf-8"))

    check_structure(bundle, manifest, errors)
    check_manifest_hashes(bundle, manifest, errors)
    check_required(bundle, policy, errors)
    check_replacements(bundle, source, policy, errors)
    check_compiles(bundle, errors)

    parity_summary: dict = {}
    if skip_runtime or errors:
        return errors, parity_summary, manifest

    status_before = git_status(source)
    bundle_report = run_isolated_probe(bundle, walk_files(bundle), "bundle")
    source_report = run_isolated_probe(source, tracked_relatives(source), "source")
    for report, label in ((bundle_report, "bundle"), (source_report, "source")):
        for problem in report.get("errors", []):
            errors.append(f"{label} probe: {problem}")
    if not errors:
        parity_summary = compare_parity(bundle_report, source_report, errors)
        for engine in ("ensemble", "ar1"):
            for check in ("replay_cache_status", "policy_runtime_status"):
                status = bundle_report.get(f"{check}_{engine}")
                if status != "ok":
                    errors.append(f"bundle {check} for {engine}: {status!r}")

    # Post-probe: prove the publishable bundle is byte-identical to what was
    # hashed above, and that validation left the checkout alone.
    contamination: list[str] = []
    check_structure(bundle, manifest, contamination)
    check_manifest_hashes(bundle, manifest, contamination)
    errors.extend(
        f"bundle was modified during validation: {problem}" for problem in contamination
    )
    status_after = git_status(source)
    if status_before is not None and status_after != status_before:
        moved = sorted(
            set((status_after or "").splitlines()) ^ set(status_before.splitlines())
        )
        errors.append(
            "validation modified the source checkout; these entries changed: "
            + ", ".join(entry.strip() for entry in moved)
        )
    return errors, parity_summary, manifest


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--bundle", default="build/databricks_app/app")
    parser.add_argument("--source", default=".")
    parser.add_argument(
        "--policy",
        default="deployment/databricks_app_bundle_policy.json",
        help="Bundle policy JSON (relative to --source when not absolute)",
    )
    parser.add_argument(
        "--skip-runtime",
        action="store_true",
        help="Only structural checks (sizes, hashes, replacements, compile)",
    )
    args = parser.parse_args(argv)

    source = Path(args.source).resolve()
    bundle = Path(args.bundle)
    if not bundle.is_absolute():
        bundle = (source / bundle).resolve()
    policy_path = Path(args.policy)
    if not policy_path.is_absolute():
        policy_path = source / policy_path

    manifest_path = bundle / MANIFEST_NAME
    if not manifest_path.is_file():
        print(f"VALIDATION FAILED: bundle manifest missing at {manifest_path}", file=sys.stderr)
        return 1

    errors, parity_summary, manifest = validate(
        bundle, source, policy_path, skip_runtime=args.skip_runtime
    )

    if errors:
        print("BUNDLE VALIDATION FAILED", file=sys.stderr)
        for problem in errors:
            print(f"  - {problem}", file=sys.stderr)
        return 1

    print("Databricks App bundle validation passed")
    print(f"  bundle           : {bundle}")
    print(f"  files            : {manifest['file_count']} (+ manifest)")
    print(f"  total size       : {manifest['total_size_bytes'] / 1048576:.2f} MiB")
    print(f"  largest file     : {manifest['largest_file']['path']}")
    if parity_summary:
        print(f"  parity checks    : {parity_summary}")
    if args.skip_runtime:
        print("  runtime checks   : SKIPPED (--skip-runtime)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
