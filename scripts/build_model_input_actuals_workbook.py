"""Emit a fill-in model-input SOURCE TEMPLATE for the next quarterly actual.

SCOPE - this is a provenance/source-template utility, NOT an ingestion path.

    Authoritative quarterly refresh command:
        scripts/refresh_model_actuals.py

    That command is the only governed way to admit a new quarter into
    ``data/model_input_history``. It is header- and period-driven, detects
    periods newer than the accepted canonical history without code edits,
    regenerates every engineered feature centrally, and fails closed on
    schema/continuity/unit/identity violations. See
    ``artifacts/actuals_refresh_2026q1/`` for its evidence contract.

This script only goes the other way: it renders the CURRENT canonical history
into a spreadsheet whose trailing row is blank, so a data owner has somewhere
to enter the next quarter's raw source values. It writes nothing into
``data/`` and is not invoked by the runtime, the pack rebuild or CI. A
workbook it produces (or any workbook with the same three main sheet schemas)
becomes input to ``refresh_model_actuals.py``, never a substitute for it.

The workbook mirrors the vendored `Master Copy revenue modelling workbook.xlsx`
input sheets that `data/model_input_history/*.parquet` were extracted from:

- rows 2002Q1 through the last accepted actual carry official values verbatim;
- one trailing blank row is left on every raw-source input cell so a user can
  fill it, with every derived column written as a live Excel formula that
  recomputes from those cells.

Every derivation encoded here was verified numerically against the historical
actuals before being written as a formula. The formulas are a convenience for
the person filling the sheet; ingestion never trusts them - it re-derives
every engineered value in code and checks parity against the sheet.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

HISTORY_DIR = ROOT / "data" / "model_input_history"

LAST_ACTUAL_PERIOD = "2025Q4"
BLANK_PERIOD = "2026Q1"

SHEETS = {
    "PED": "PED Inputs",
    "LIGHT_RUC": "Light RUC Inputs",
    "HEAVY_RUC": "Heavy RUC Inputs",
}
PARQUETS = {
    "PED": "ped_inputs.parquet",
    "LIGHT_RUC": "light_ruc_inputs.parquet",
    "HEAVY_RUC": "heavy_ruc_inputs.parquet",
}

QUARTER_END_MONTH = {1: "Mar", 2: "Jun", 3: "Sep", 4: "Dec"}

# role, unit, description. Role drives fill colour, protection and whether the
# 2026Q1 cell is blank (input) or a formula (derived).
ID, INPUT, DERIVED, NOTE = "id", "input", "derived", "note"

SPEC: dict[str, dict[str, tuple[str, str, str]]] = {
    "PED": {
        "period": (ID, "", "Quarter label."),
        "period_label": (ID, "", "Quarter end month and year."),
        "year": (ID, "", "Calendar year."),
        "quarter": (ID, "", "Calendar quarter, 1-4."),
        "period_index": (ID, "", "Sequential quarter index, 2002Q1 = 1."),
        "trend_index": (ID, "", "Time trend used by the fitted specs. Equals period_index."),
        "target": (DERIVED, "km/person", "PED dependent variable: light petrol VKT per capita = light_petrol_vkt_total_km / population."),
        "log_target": (DERIVED, "log", "Natural log of target."),
        "target_lag_1": (DERIVED, "km/person", "Target one quarter earlier."),
        "target_lag_4": (DERIVED, "km/person", "Target four quarters earlier."),
        "log_trend": (DERIVED, "log", "Natural log of trend_index."),
        "nominal_gdp_sa_nzd": (INPUT, "NZD", "Nominal seasonally adjusted quarterly GDP."),
        "population": (INPUT, "people", "Estimated resident population."),
        "nominal_gdp_per_capita_nzd": (DERIVED, "NZD/person", "nominal_gdp_sa_nzd / population."),
        "gdp_deflator_sa": (INPUT, "index", "Seasonally adjusted GDP deflator index."),
        "gdp_rebasing_factor": (DERIVED, "ratio", "2025Q4 deflator / this quarter's deflator. Rebases GDP into 2025Q4 dollars."),
        "real_gdp_per_capita_nzd": (DERIVED, "NZD/person", "nominal_gdp_per_capita_nzd * gdp_rebasing_factor."),
        "log_real_gdp_per_capita": (DERIVED, "log", "Natural log of real_gdp_per_capita_nzd."),
        "petrol_price_nominal_cents_per_litre": (INPUT, "cents/litre", "Nominal retail petrol price (Stats NZ CPI detailed table)."),
        "cpi_rebasing_factor": (INPUT, "ratio", "CPI rebasing factor. Base quarter is 2026Q1, so 2026Q1 = 1."),
        "real_petrol_price_cents_per_litre": (DERIVED, "cents/litre", "petrol_price_nominal_cents_per_litre * cpi_rebasing_factor."),
        "log_real_petrol_price": (DERIVED, "log", "Natural log of real_petrol_price_cents_per_litre."),
        "light_petrol_vkt_total_km": (INPUT, "km", "Total light petrol vehicle kilometres travelled."),
        "log_total_vkt": (DERIVED, "log", "Natural log of light_petrol_vkt_total_km."),
        "unemployment_percent": (INPUT, "percent", "Unemployment rate in percentage points, for example 5.4."),
        "unemployment_rate": (DERIVED, "fraction", "unemployment_percent / 100. The fitted models use this fraction, not the percentage."),
        "log_unemployment_rate": (DERIVED, "log", "Natural log of unemployment_rate."),
        "post_2011_dummy": (DERIVED, "0/1", "1 from 2012Q1 onwards."),
        "dummy_2020": (DERIVED, "0/1", "1 for the four 2020 quarters."),
        "post_2020_dummy": (DERIVED, "0/1", "1 from 2021Q1 onwards."),
        "q2_dummy": (DERIVED, "0/1", "1 when quarter is 2."),
        "q3_dummy": (DERIVED, "0/1", "1 when quarter is 3."),
        "q4_dummy": (DERIVED, "0/1", "1 when quarter is 4."),
        "post_2011_x_log_trend": (DERIVED, "log", "post_2011_dummy * log_trend."),
        "data_status": (NOTE, "", "Provenance status for the row."),
        "notes": (NOTE, "", "Free-text provenance note."),
        "log_target_lag_1": (DERIVED, "log", "Natural log of target_lag_1."),
        "log_target_lag_4": (DERIVED, "log", "Natural log of target_lag_4."),
        "diff_log_target_lag_1_lag_4": (DERIVED, "log", "log_target_lag_1 minus log_target_lag_4."),
    },
    "LIGHT_RUC": {
        "period": (ID, "", "Quarter label."),
        "period_label": (ID, "", "Quarter end month and year."),
        "year": (ID, "", "Calendar year."),
        "quarter": (ID, "", "Calendar quarter, 1-4."),
        "period_index": (ID, "", "Sequential quarter index, 2002Q1 = 1."),
        "trend_index": (ID, "", "Time trend used by the fitted specs. Equals period_index."),
        "target": (INPUT, "km", "Light RUC dependent variable: Light RUC distance purchased. Zero before 2009Q3, where the series does not exist."),
        "log_target": (DERIVED, "log", "Natural log of target."),
        "target_lag_1": (DERIVED, "km", "Target one quarter earlier."),
        "target_lag_4": (DERIVED, "km", "Target four quarters earlier."),
        "log_trend": (DERIVED, "log", "Natural log of trend_index."),
        "nominal_gdp_sa_nzd": (INPUT, "NZD", "Nominal seasonally adjusted quarterly GDP."),
        "gdp_deflator_sa": (INPUT, "index", "Seasonally adjusted GDP deflator index."),
        "gdp_rebasing_factor": (DERIVED, "ratio", "2025Q4 deflator / this quarter's deflator."),
        "real_gdp_sa_nzd": (DERIVED, "NZD", "nominal_gdp_sa_nzd * gdp_rebasing_factor."),
        "log_real_gdp": (DERIVED, "log", "Natural log of real_gdp_sa_nzd."),
        "diesel_price_nominal_cents_per_litre": (INPUT, "cents/litre", "Nominal retail diesel price."),
        "cpi_rebasing_factor": (INPUT, "ratio", "CPI rebasing factor. Base quarter is 2026Q1, so 2026Q1 = 1."),
        "real_diesel_price_cents_per_litre": (DERIVED, "cents/litre", "diesel_price_nominal_cents_per_litre * cpi_rebasing_factor."),
        "log_real_diesel_price": (DERIVED, "log", "Natural log of real_diesel_price_cents_per_litre."),
        "light_ruc_revenue_nzd": (INPUT, "NZD", "Light RUC revenue. Zero before 2009Q3."),
        "light_ruc_price_nominal_nzd_per_1000km": (DERIVED, "NZD/1000km", "light_ruc_revenue_nzd / target * 1000. Effective average nominal rate."),
        "real_light_ruc_price_nzd_per_1000km": (DERIVED, "NZD/1000km", "light_ruc_price_nominal_nzd_per_1000km * cpi_rebasing_factor."),
        "log_real_light_ruc_price": (DERIVED, "log", "Natural log of real_light_ruc_price_nzd_per_1000km."),
        "lagged_real_light_ruc_price_nzd_per_1000km": (DERIVED, "NZD/1000km", "real_light_ruc_price_nzd_per_1000km one quarter earlier."),
        "log_lagged_real_light_ruc_price": (DERIVED, "log", "Natural log of the lagged real Light RUC price."),
        "q2_dummy": (DERIVED, "0/1", "1 when quarter is 2."),
        "q3_dummy": (DERIVED, "0/1", "1 when quarter is 3."),
        "q4_dummy": (DERIVED, "0/1", "1 when quarter is 4."),
        "post_2020_dummy": (DERIVED, "0/1", "1 from 2021Q1 onwards."),
        "data_status": (NOTE, "", "Provenance status for the row."),
        "notes": (NOTE, "", "Free-text provenance note."),
        "log_target_lag_1": (DERIVED, "log", "Natural log of target_lag_1."),
        "log_target_lag_4": (DERIVED, "log", "Natural log of target_lag_4."),
        "diff_log_target_lag_1_lag_4": (DERIVED, "log", "log_target_lag_1 minus log_target_lag_4."),
        "diesel_x_ruc_price": (DERIVED, "log", "log_real_diesel_price * log_real_light_ruc_price."),
        "gdp_x_post2020": (DERIVED, "log", "log_real_gdp * post_2020_dummy."),
        "ruc_x_post2020": (DERIVED, "log", "log_real_light_ruc_price * post_2020_dummy."),
        "diesel_x_post2020": (DERIVED, "log", "log_real_diesel_price * post_2020_dummy."),
        "time_trend": (DERIVED, "", "Equals trend_index."),
        "log_real_diesel_price_diff1": (DERIVED, "log", "Quarterly change in log_real_diesel_price."),
        "log_real_diesel_price_lag1": (DERIVED, "log", "log_real_diesel_price one quarter earlier."),
        "log_real_diesel_price_lag4": (DERIVED, "log", "log_real_diesel_price four quarters earlier."),
        "log_real_light_ruc_price_diff1": (DERIVED, "log", "Quarterly change in log_real_light_ruc_price."),
        "log_real_light_ruc_price_lag1": (DERIVED, "log", "log_real_light_ruc_price one quarter earlier."),
        "log_real_light_ruc_price_lag4": (DERIVED, "log", "log_real_light_ruc_price four quarters earlier."),
        "log_real_gdp_diff1": (DERIVED, "log", "Quarterly change in log_real_gdp."),
        "log_real_gdp_lag1": (DERIVED, "log", "log_real_gdp one quarter earlier."),
        "log_real_gdp_lag4": (DERIVED, "log", "log_real_gdp four quarters earlier."),
    },
    "HEAVY_RUC": {
        "period": (ID, "", "Quarter label."),
        "period_label": (ID, "", "Quarter end month and year."),
        "year": (ID, "", "Calendar year."),
        "quarter": (ID, "", "Calendar quarter, 1-4."),
        "period_index": (ID, "", "Sequential quarter index, 2002Q1 = 1."),
        "trend_index": (ID, "", "Time trend used by the fitted specs. Equals period_index."),
        "target": (INPUT, "km", "Heavy RUC dependent variable: Heavy RUC distance purchased. Zero before 2009Q3."),
        "log_target": (DERIVED, "log", "Natural log of target."),
        "target_lag_1": (DERIVED, "km", "Target one quarter earlier."),
        "target_lag_4": (DERIVED, "km", "Target four quarters earlier."),
        "log_trend": (DERIVED, "log", "Natural log of trend_index."),
        "nominal_gdp_sa_nzd": (INPUT, "NZD", "Nominal seasonally adjusted quarterly GDP."),
        "gdp_deflator_sa": (INPUT, "index", "Seasonally adjusted GDP deflator index."),
        "gdp_rebasing_factor": (DERIVED, "ratio", "2025Q4 deflator / this quarter's deflator."),
        "real_gdp_sa_nzd": (DERIVED, "NZD", "nominal_gdp_sa_nzd * gdp_rebasing_factor."),
        "log_real_gdp": (DERIVED, "log", "Natural log of real_gdp_sa_nzd."),
        "heavy_ruc_revenue_nzd": (INPUT, "NZD", "Heavy RUC revenue. Zero before 2009Q3."),
        "heavy_ruc_price_nominal_nzd_per_1000km": (DERIVED, "NZD/1000km", "heavy_ruc_revenue_nzd / target * 1000. Effective average nominal rate."),
        "real_heavy_ruc_price_nzd_per_1000km": (DERIVED, "NZD/1000km", "heavy_ruc_price_nominal_nzd_per_1000km * the Light RUC sheet cpi_rebasing_factor."),
        "log_real_heavy_ruc_price": (DERIVED, "log", "Natural log of real_heavy_ruc_price_nzd_per_1000km."),
        "lead_real_heavy_ruc_price_nzd_per_1000km": (DERIVED, "NZD/1000km", "real_heavy_ruc_price_nzd_per_1000km one quarter LATER. 2025Q4 stays blank until 2026Q1 is filled."),
        "log_lead_real_heavy_ruc_price": (DERIVED, "log", "Natural log of the lead real Heavy RUC price."),
        "q2_dummy": (DERIVED, "0/1", "1 when quarter is 2."),
        "q3_dummy": (DERIVED, "0/1", "1 when quarter is 3."),
        "q4_dummy": (DERIVED, "0/1", "1 when quarter is 4."),
        "data_status": (NOTE, "", "Provenance status for the row."),
        "notes": (NOTE, "", "Free-text provenance note."),
        "real_diesel_price_cents_per_litre": (DERIVED, "cents/litre", "Pulled from the Light RUC Inputs sheet, same row."),
        "log_real_diesel_price": (DERIVED, "log", "Natural log of real_diesel_price_cents_per_litre."),
        "real_light_ruc_price_nzd_per_1000km": (DERIVED, "NZD/1000km", "Pulled from the Light RUC Inputs sheet, same row."),
        "lagged_real_light_ruc_price_nzd_per_1000km": (DERIVED, "NZD/1000km", "real_light_ruc_price_nzd_per_1000km one quarter earlier."),
        "log_real_light_ruc_price": (DERIVED, "log", "Natural log of real_light_ruc_price_nzd_per_1000km."),
        "log_lagged_real_light_ruc_price": (DERIVED, "log", "Natural log of the lagged real Light RUC price."),
        "unemployment_rate": (DERIVED, "fraction", "Pulled from the PED Inputs sheet, same row."),
        "log_unemployment_rate": (DERIVED, "log", "Natural log of unemployment_rate."),
        "log_target_lag_1": (DERIVED, "log", "Natural log of target_lag_1."),
        "log_target_lag_4": (DERIVED, "log", "Natural log of target_lag_4."),
        "diff_log_target_lag_1_lag_4": (DERIVED, "log", "log_target_lag_1 minus log_target_lag_4."),
    },
}

PED_SHEET = f"'{SHEETS['PED']}'"
LIGHT_SHEET = f"'{SHEETS['LIGHT_RUC']}'"


def _ref(col_map: dict[str, int], name: str, row: int, *, abs_row: bool = False) -> str:
    return f"{get_column_letter(col_map[name])}{'$' if abs_row else ''}{row}"


def _log(source: str) -> str:
    return f'=IF(N({source})>0,LN({source}),"")'


def _mul(a: str, b: str) -> str:
    return f'=IF(OR({a}="",{b}=""),"",{a}*{b})'


def _blank_formulas(stream: str, col_map: dict[str, int], row: int, base_row: int) -> dict[str, str]:
    """Excel formulas for the trailing blank quarter. `base_row` holds 2025Q4."""

    def c(name: str, offset: int = 0) -> str:
        return _ref(col_map, name, row + offset)

    year, quarter = c("year"), c("quarter")
    deflator_base = _ref(col_map, "gdp_deflator_sa", base_row, abs_row=True)

    common = {
        "log_trend": _log(c("trend_index")),
        "gdp_rebasing_factor": f'=IF(N({c("gdp_deflator_sa")})>0,{deflator_base}/{c("gdp_deflator_sa")},"")',
        "log_target": _log(c("target")),
        "target_lag_1": f'=IF({c("target", -1)}="","",{c("target", -1)})',
        "target_lag_4": f'=IF({c("target", -4)}="","",{c("target", -4)})',
        "log_target_lag_1": _log(c("target_lag_1")),
        "log_target_lag_4": _log(c("target_lag_4")),
        "diff_log_target_lag_1_lag_4": (
            f'=IF(OR({c("log_target_lag_1")}="",{c("log_target_lag_4")}=""),"",'
            f'{c("log_target_lag_1")}-{c("log_target_lag_4")})'
        ),
        "q2_dummy": f"=IF({quarter}=2,1,0)",
        "q3_dummy": f"=IF({quarter}=3,1,0)",
        "q4_dummy": f"=IF({quarter}=4,1,0)",
        "post_2020_dummy": f"=IF({year}>=2021,1,0)",
    }

    if stream == "PED":
        common.update(
            {
                "nominal_gdp_per_capita_nzd": (
                    f'=IF(OR({c("nominal_gdp_sa_nzd")}="",N({c("population")})=0),"",'
                    f'{c("nominal_gdp_sa_nzd")}/{c("population")})'
                ),
                "real_gdp_per_capita_nzd": _mul(c("nominal_gdp_per_capita_nzd"), c("gdp_rebasing_factor")),
                "log_real_gdp_per_capita": _log(c("real_gdp_per_capita_nzd")),
                "real_petrol_price_cents_per_litre": _mul(
                    c("petrol_price_nominal_cents_per_litre"), c("cpi_rebasing_factor")
                ),
                "log_real_petrol_price": _log(c("real_petrol_price_cents_per_litre")),
                "target": (
                    f'=IF(OR({c("light_petrol_vkt_total_km")}="",N({c("population")})=0),"",'
                    f'{c("light_petrol_vkt_total_km")}/{c("population")})'
                ),
                "log_total_vkt": _log(c("light_petrol_vkt_total_km")),
                "unemployment_rate": f'=IF({c("unemployment_percent")}="","",{c("unemployment_percent")}/100)',
                "log_unemployment_rate": _log(c("unemployment_rate")),
                "post_2011_dummy": f"=IF({year}>=2012,1,0)",
                "dummy_2020": f"=IF({year}=2020,1,0)",
                "post_2011_x_log_trend": _mul(c("post_2011_dummy"), c("log_trend")),
            }
        )
        return common

    common.update(
        {
            "real_gdp_sa_nzd": _mul(c("nominal_gdp_sa_nzd"), c("gdp_rebasing_factor")),
            "log_real_gdp": _log(c("real_gdp_sa_nzd")),
        }
    )

    if stream == "LIGHT_RUC":
        common.update(
            {
                "real_diesel_price_cents_per_litre": _mul(
                    c("diesel_price_nominal_cents_per_litre"), c("cpi_rebasing_factor")
                ),
                "log_real_diesel_price": _log(c("real_diesel_price_cents_per_litre")),
                "light_ruc_price_nominal_nzd_per_1000km": (
                    f'=IF(OR({c("light_ruc_revenue_nzd")}="",N({c("target")})=0),"",'
                    f'{c("light_ruc_revenue_nzd")}/{c("target")}*1000)'
                ),
                "real_light_ruc_price_nzd_per_1000km": _mul(
                    c("light_ruc_price_nominal_nzd_per_1000km"), c("cpi_rebasing_factor")
                ),
                "log_real_light_ruc_price": _log(c("real_light_ruc_price_nzd_per_1000km")),
                "lagged_real_light_ruc_price_nzd_per_1000km": (
                    f'=IF({c("real_light_ruc_price_nzd_per_1000km", -1)}="","",'
                    f'{c("real_light_ruc_price_nzd_per_1000km", -1)})'
                ),
                "log_lagged_real_light_ruc_price": _log(c("lagged_real_light_ruc_price_nzd_per_1000km")),
                "diesel_x_ruc_price": _mul(c("log_real_diesel_price"), c("log_real_light_ruc_price")),
                "gdp_x_post2020": _mul(c("log_real_gdp"), c("post_2020_dummy")),
                "ruc_x_post2020": _mul(c("log_real_light_ruc_price"), c("post_2020_dummy")),
                "diesel_x_post2020": _mul(c("log_real_diesel_price"), c("post_2020_dummy")),
                "time_trend": f'={c("trend_index")}',
            }
        )
        for base in ("log_real_diesel_price", "log_real_light_ruc_price", "log_real_gdp"):
            common[f"{base}_lag1"] = f'=IF({c(base, -1)}="","",{c(base, -1)})'
            common[f"{base}_lag4"] = f'=IF({c(base, -4)}="","",{c(base, -4)})'
            common[f"{base}_diff1"] = (
                f'=IF(OR({c(base)}="",{c(base, -1)}=""),"",{c(base)}-{c(base, -1)})'
            )
        return common

    light_col = SPEC["LIGHT_RUC"]
    light_map = {name: idx for idx, name in enumerate(light_col, start=1)}
    ped_map = {name: idx for idx, name in enumerate(SPEC["PED"], start=1)}
    light_cpi = f'{LIGHT_SHEET}!{get_column_letter(light_map["cpi_rebasing_factor"])}{row}'
    light_diesel = f'{LIGHT_SHEET}!{get_column_letter(light_map["real_diesel_price_cents_per_litre"])}{row}'
    light_price = f'{LIGHT_SHEET}!{get_column_letter(light_map["real_light_ruc_price_nzd_per_1000km"])}{row}'
    ped_unemp = f'{PED_SHEET}!{get_column_letter(ped_map["unemployment_rate"])}{row}'

    common.update(
        {
            "heavy_ruc_price_nominal_nzd_per_1000km": (
                f'=IF(OR({c("heavy_ruc_revenue_nzd")}="",N({c("target")})=0),"",'
                f'{c("heavy_ruc_revenue_nzd")}/{c("target")}*1000)'
            ),
            "real_heavy_ruc_price_nzd_per_1000km": _mul(
                c("heavy_ruc_price_nominal_nzd_per_1000km"), light_cpi
            ),
            "log_real_heavy_ruc_price": _log(c("real_heavy_ruc_price_nzd_per_1000km")),
            "lead_real_heavy_ruc_price_nzd_per_1000km": '=""',
            "log_lead_real_heavy_ruc_price": _log(c("lead_real_heavy_ruc_price_nzd_per_1000km")),
            "real_diesel_price_cents_per_litre": f'=IF({light_diesel}="","",{light_diesel})',
            "log_real_diesel_price": _log(c("real_diesel_price_cents_per_litre")),
            "real_light_ruc_price_nzd_per_1000km": f'=IF({light_price}="","",{light_price})',
            "lagged_real_light_ruc_price_nzd_per_1000km": (
                f'=IF({c("real_light_ruc_price_nzd_per_1000km", -1)}="","",'
                f'{c("real_light_ruc_price_nzd_per_1000km", -1)})'
            ),
            "log_real_light_ruc_price": _log(c("real_light_ruc_price_nzd_per_1000km")),
            "log_lagged_real_light_ruc_price": _log(c("lagged_real_light_ruc_price_nzd_per_1000km")),
            "unemployment_rate": f'=IF({ped_unemp}="","",{ped_unemp})',
            "log_unemployment_rate": _log(c("unemployment_rate")),
        }
    )
    return common


HEADER_FILL = PatternFill("solid", fgColor="002B5C")
FILLS = {
    ID: PatternFill("solid", fgColor="E6EDF5"),
    INPUT: PatternFill("solid", fgColor="FFF2B2"),
    DERIVED: PatternFill("solid", fgColor="E7F5EC"),
    NOTE: PatternFill("solid", fgColor="F2F2F2"),
}
ROLE_LABEL = {
    ID: "IDENTIFIER (generated)",
    INPUT: "INPUT (you fill this for 2026Q1)",
    DERIVED: "DERIVED (formula recomputes)",
    NOTE: "NOTE (free text)",
}


def _number_format(name: str, unit: str) -> str:
    if unit in {"log", "ratio", "fraction"}:
        return "0.000000"
    if unit in {"0/1", ""}:
        return "General"
    if unit in {"NZD", "km"}:
        return "#,##0"
    if unit == "people":
        return "#,##0"
    return "#,##0.000000"


def _write_sheet(ws, stream: str, frame: pd.DataFrame) -> None:
    spec = SPEC[stream]
    col_map = {name: idx for idx, name in enumerate(spec, start=1)}

    for name, (role, unit, description) in spec.items():
        cell = ws.cell(row=1, column=col_map[name], value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        unit_text = f"Unit: {unit}\n" if unit else ""
        cell.comment = Comment(f"{ROLE_LABEL[role]}\n{unit_text}{description}", "NLTF model input sheet")
        ws.column_dimensions[get_column_letter(col_map[name])].width = max(14, min(30, len(name) + 2))

    for offset, (_, record) in enumerate(frame.iterrows()):
        row = offset + 2
        for name, (role, unit, _desc) in spec.items():
            cell = ws.cell(row=row, column=col_map[name])
            value = record.get(name)
            if pd.isna(value):
                value = None
            elif hasattr(value, "item"):
                value = value.item()
            cell.value = value
            cell.fill = FILLS[role]
            if role in {ID, INPUT, DERIVED} and isinstance(value, (int, float)):
                cell.number_format = _number_format(name, unit)

    base_row = len(frame) + 1
    blank_row = base_row + 1
    year, quarter = int(BLANK_PERIOD[:4]), int(BLANK_PERIOD[-1])
    identifiers = {
        "period": BLANK_PERIOD,
        "period_label": f"{QUARTER_END_MONTH[quarter]} {year}",
        "year": year,
        "quarter": quarter,
        "period_index": int(frame["period_index"].iloc[-1]) + 1,
        "trend_index": int(frame["trend_index"].iloc[-1]) + 1,
    }
    formulas = _blank_formulas(stream, col_map, blank_row, base_row)

    for name, (role, unit, _desc) in spec.items():
        cell = ws.cell(row=blank_row, column=col_map[name])
        if role == ID:
            cell.value = identifiers[name]
            cell.fill = FILLS[ID]
        elif role == DERIVED:
            cell.value = formulas.get(name, '=""')
            cell.fill = FILLS[DERIVED]
            cell.number_format = _number_format(name, unit)
        elif role == INPUT:
            if name == "cpi_rebasing_factor":
                # 2026Q1 is the CPI base quarter, so this factor is 1 by construction.
                cell.value = 1.0
            cell.fill = FILLS[INPUT]
            cell.font = Font(bold=True)
            cell.protection = Protection(locked=False)
            cell.number_format = _number_format(name, unit)
        else:
            cell.fill = FILLS[NOTE]
            cell.protection = Protection(locked=False)

    # 2025Q4's lead Heavy RUC price is only knowable once 2026Q1 is entered; the
    # vendored source carries a placeholder zero there.
    if stream == "HEAVY_RUC":
        real_price = _ref(col_map, "real_heavy_ruc_price_nzd_per_1000km", blank_row)
        lead_cell = ws.cell(row=base_row, column=col_map["lead_real_heavy_ruc_price_nzd_per_1000km"])
        lead_cell.value = f'=IF(N({real_price})>0,{real_price},"")'
        lead_cell.fill = FILLS[DERIVED]
        lead_cell.number_format = "#,##0.000000"
        lead_cell.comment = Comment(
            "Source workbook carried a placeholder 0 here. Replaced with a formula that "
            "fills once the 2026Q1 Heavy RUC revenue and volume are entered.",
            "NLTF model input sheet",
        )
        log_lead = ws.cell(row=base_row, column=col_map["log_lead_real_heavy_ruc_price"])
        log_lead.value = _log(_ref(col_map, "lead_real_heavy_ruc_price_nzd_per_1000km", base_row))
        log_lead.fill = FILLS[DERIVED]
        log_lead.number_format = "0.000000"

    ws.freeze_panes = "G2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(spec))}{blank_row}"


def build(output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for stream, sheet_name in SHEETS.items():
        frame = pd.read_parquet(HISTORY_DIR / PARQUETS[stream])
        actuals = frame[frame["period"] <= LAST_ACTUAL_PERIOD].reset_index(drop=True)
        missing = [name for name in SPEC[stream] if name not in actuals.columns]
        if missing:
            raise RuntimeError(f"{stream}: history is missing columns {missing}")
        extra = [name for name in actuals.columns if name not in SPEC[stream]]
        if extra:
            raise RuntimeError(f"{stream}: history has unmapped columns {extra}")
        _write_sheet(wb.create_sheet(sheet_name), stream, actuals)
    wb.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT / "deliverables" / "NLTF_model_input_sheet_actuals_to_2025Q4.xlsx",
    )
    args = parser.parse_args()
    print(build(args.output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
