"""Materialize the normalized NLTF revenue source pack from the distilled workbook.

The raw revenue workbook is lineage only and is never copied into the repo. This
script verifies its SHA256, then exports the compact normalized contract sheets
from the distilled workbook into data/revenue_model_source_pack/2026_05_19.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


EXPECTED_RAW_SHA256 = "00c6070694818d27d7c402749354d8175de999894846dce45a4abdd7f5eb3e6b"
SOURCE_PACK_VERSION = "2026_05_19"
DEFAULT_OUTPUT_DIR = Path("data") / "revenue_model_source_pack" / SOURCE_PACK_VERSION

SHEET_EXPORTS = {
    "Series Master": ("series_master.csv", "Series ID"),
    "Aggregation Rules": ("aggregation_rules.csv", "Rule ID"),
    "Current Selections": ("current_selections.csv", "Control ID"),
    "Model Coefficients": ("model_coefficients.csv", "Forecast cell"),
    "Annual Actuals": ("annual_actuals.csv", "Group"),
    "Annual Model Paths": ("annual_model_paths.csv", "Model basis"),
    "Release Registry": ("release_registry.csv", "Source sheet"),
    "Open Decisions": ("unresolved_decisions.csv", "Priority"),
    "Formula Errors": ("formula_errors.csv", "Sheet"),
}

SOURCE_PACK_CREATED_AT = "2026-05-19T00:00:00+00:00"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _slug(value: str) -> str:
    chars = []
    for char in value.strip().lower():
        if char.isalnum():
            chars.append(char)
        elif chars and chars[-1] != "_":
            chars.append("_")
    return "".join(chars).strip("_")


def _clean_header(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").lower()
    return text or "unnamed"


def _as_int(value: Any) -> int | None:
    try:
        if value is None:
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _june_year(period: Any) -> int | None:
    text = str(period or "").strip().upper()
    if not text:
        return None
    if text.startswith("FY"):
        return _as_int(text.replace("FY", ""))
    if "Q" not in text:
        return _as_int(text)
    try:
        year_text, quarter_text = text.split("Q", 1)
        year = int(year_text)
        quarter = int(quarter_text)
    except ValueError:
        return None
    return year if quarter <= 2 else year + 1


def _is_quarter(period: Any) -> bool:
    return bool(re.match(r"^\d{4}Q[1-4]$", str(period or "").strip().upper()))


def _infer_unit(series: Any) -> str:
    text = str(series or "").strip().lower()
    if not text:
        return ""
    if "rate" in text and ("nzd/l" in text or "petrol excise duty" in text):
        return "NZD/L"
    if "exgst" in text or "ex gst" in text or "actual incl crown" in text:
        return "$ nominal ex GST"
    if "revenue" in text or "refund" in text or "top-up" in text or "top up" in text or "forecast incl ev/phev" in text:
        return "$m nominal ex GST"
    if "ped volume" in text or "litre" in text or "liter" in text:
        return "million L"
    if "gross km" in text or "net km conventional" in text or "bev net km" in text or "phev net km" in text:
        return "km"
    if "vkt per capita" in text:
        return "km/person/FY"
    if "vkt" in text or "net km" in text:
        return "million km"
    if "gtk" in text:
        return "tonne-km"
    return ""


def _revenue_basis(series: Any, unit: Any = "") -> str:
    text = str(series or "").strip().lower()
    unit_text = str(unit or "").strip().lower()
    if "$" not in unit_text and "revenue" not in text and "refund" not in text and "top-up" not in text and "top up" not in text:
        return "activity"
    if "gross" in text:
        return "gross"
    if "refund" in text:
        return "deduction"
    if "admin" in text:
        return "admin"
    if "net" in text or "total ruc+ped" in text or "total nltf" in text or "forecast incl ev/phev" in text or "top-up" in text or "top up" in text:
        return "net"
    return "nominal_ex_gst"


def _release_family(release_round: Any) -> str:
    text = str(release_round or "").strip()
    match = re.match(r"^([A-Za-z ]+)", text)
    return match.group(1).strip() if match else ""


def _sheet_rows(workbook: Any, sheet_name: str) -> list[list[Any]]:
    sheet = workbook[sheet_name]
    return [[_cell_value(value) for value in row] for row in sheet.iter_rows(values_only=True)]


def _find_header_index(rows: list[list[Any]], first_header: str) -> int:
    for index, row in enumerate(rows):
        if row and str(row[0]).strip() == first_header:
            return index
    raise ValueError(f"Could not find header {first_header!r}")


def _trim(row: list[Any], width: int) -> list[Any]:
    row = row[:width]
    if len(row) < width:
        row = row + [""] * (width - len(row))
    return row


def export_table(workbook: Any, sheet_name: str, output_path: Path, first_header: str) -> int:
    rows = _sheet_rows(workbook, sheet_name)
    header_index = _find_header_index(rows, first_header)
    header = [str(value).strip() for value in rows[header_index]]
    while header and not header[-1]:
        header.pop()
    width = len(header)
    data_rows = []
    for row in rows[header_index + 1 :]:
        trimmed = _trim(row, width)
        if not any(str(value).strip() for value in trimmed):
            continue
        data_rows.append(trimmed)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(data_rows)
    return len(data_rows)


def _write_csv(output_path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> int:
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field, "")) for field in fieldnames})
    return len(rows)


def export_forecast_archive(workbook: Any, output_path: Path) -> int:
    sheet_name = "Forecast archive"
    sheet = workbook[sheet_name]
    header_row = 4
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [_clean_header(value) for value in header_values]
    rows: list[dict[str, Any]] = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(value not in (None, "") for value in values):
            continue
        record = {
            headers[index]: _cell_value(value)
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        record["source_workbook_sheet"] = sheet_name
        record["source_workbook_row"] = row_idx
        record["source_cell"] = f"J{row_idx}"
        record["FY"] = record.get("fiscal_year", "")
        record["release_round"] = record.get("release", "")
        record["value_status"] = "forecast_archive"
        record["revenue_basis"] = _revenue_basis(record.get("series", ""), record.get("unit", ""))
        rows.append(record)
    fieldnames = [
        "source_sheet",
        "release_round",
        "release_family",
        "release_year",
        "FY",
        "horizon",
        "series",
        "metric",
        "unit",
        "forecast",
        "raw_actual",
        "crown_top_up",
        "benchmark_actual",
        "error",
        "pct_error",
        "source_cell",
        "source_workbook_sheet",
        "source_workbook_row",
        "value_status",
        "revenue_basis",
    ]
    return _write_csv(output_path, fieldnames, rows)


def export_release_values(workbook: Any, output_path: Path) -> int:
    sheet_name = "MOT release series"
    sheet = workbook[sheet_name]
    header_row = 1
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [_clean_header(value) for value in header_values]
    rows: list[dict[str, Any]] = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        release_round = values[0] if len(values) > 0 else None
        series = values[1] if len(values) > 1 else None
        value = values[3] if len(values) > 3 else None
        if not str(release_round or "").strip() or not str(series or "").strip():
            continue
        record = {
            headers[index]: _cell_value(item)
            for index, item in enumerate(values)
            if index < len(headers) and headers[index]
        }
        unit = _infer_unit(series)
        rows.append(
            {
                "release_round": record.get("release_key", ""),
                "release_family": _release_family(record.get("release_key", "")),
                "release_year": record.get("release_year", ""),
                "FY": record.get("fiscal_year", ""),
                "horizon": record.get("horizon", ""),
                "series": record.get("series", ""),
                "value": value,
                "unit": unit,
                "revenue_basis": _revenue_basis(series, unit),
                "scenario": "MOT release round",
                "fed_path": "",
                "value_status": "forecast" if _as_int(record.get("horizon")) and _as_int(record.get("horizon")) > 0 else "benchmark_or_release_value",
                "source_sheet": record.get("source_sheet", ""),
                "source_label": record.get("source_label", ""),
                "source_unit": record.get("source_unit", ""),
                "derivation": record.get("derivation", ""),
                "source_workbook_sheet": sheet_name,
                "source_workbook_row": row_idx,
                "source_cell": f"D{row_idx}",
            }
        )
    fieldnames = [
        "release_round",
        "release_family",
        "release_year",
        "FY",
        "horizon",
        "series",
        "value",
        "unit",
        "revenue_basis",
        "scenario",
        "fed_path",
        "value_status",
        "source_sheet",
        "source_label",
        "source_unit",
        "derivation",
        "source_workbook_sheet",
        "source_workbook_row",
        "source_cell",
    ]
    return _write_csv(output_path, fieldnames, rows)


def export_quarterly_actuals(workbook: Any, output_path: Path) -> int:
    rows: list[dict[str, Any]] = []
    sheet_name = "Quarterly actuals"
    sheet = workbook[sheet_name]
    header_row = 4
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    quarters = {
        column: str(value).strip().upper()
        for column, value in enumerate(header_values[1:], start=2)
        if str(value or "").strip()
    }
    for row_idx, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        series = str((values[0] if len(values) > 0 else "") or "").strip()
        if not series:
            continue
        if series.lower() == "note":
            continue
        unit = _infer_unit(series)
        for column, quarter in quarters.items():
            value = values[column - 1] if len(values) >= column else None
            if value is None or str(value).strip() == "":
                continue
            rows.append(
                {
                    "series": series,
                    "quarter": quarter,
                    "FY": _june_year(quarter),
                    "value": value,
                    "unit": unit,
                    "revenue_basis": _revenue_basis(series, unit),
                    "value_status": "actual",
                    "source_sheet": sheet_name,
                    "source_workbook_sheet": sheet_name,
                    "source_workbook_row": row_idx,
                    "source_cell": f"{get_column_letter(column)}{row_idx}",
                }
            )

    top_up_sheet_name = "Crown top-up"
    top_up = workbook[top_up_sheet_name]
    top_up_header = 24
    top_up_header_values = next(top_up.iter_rows(min_row=top_up_header, max_row=top_up_header, values_only=True))
    top_up_columns = {
        column: str(value).strip()
        for column, value in enumerate(top_up_header_values[1:5], start=2)
        if str(value or "").strip()
    }
    top_up_series = {
        "FED / PED top-up": "FED / PED Crown top-up",
        "Light RUC top-up": "Light RUC Crown top-up",
        "Heavy RUC top-up": "Heavy RUC Crown top-up",
        "Total top-up": "Crown top-up",
    }
    for row_idx, values in enumerate(top_up.iter_rows(min_row=top_up_header + 1, values_only=True), start=top_up_header + 1):
        quarter = str((values[0] if len(values) > 0 else "") or "").strip().upper()
        if not quarter or "Q" not in quarter:
            continue
        for column, label in top_up_columns.items():
            value = values[column - 1] if len(values) >= column else None
            if value is None or str(value).strip() == "":
                continue
            series = top_up_series.get(label, label)
            rows.append(
                {
                    "series": series,
                    "quarter": quarter,
                    "FY": _june_year(quarter),
                    "value": value,
                    "unit": "$ nominal ex GST",
                    "revenue_basis": _revenue_basis(series, "$ nominal ex GST"),
                    "value_status": "policy_overlay_actual",
                    "source_sheet": top_up_sheet_name,
                    "source_workbook_sheet": top_up_sheet_name,
                    "source_workbook_row": row_idx,
                    "source_cell": f"{get_column_letter(column)}{row_idx}",
                }
            )
    fieldnames = [
        "series",
        "quarter",
        "FY",
        "value",
        "unit",
        "revenue_basis",
        "value_status",
        "source_sheet",
        "source_workbook_sheet",
        "source_workbook_row",
        "source_cell",
    ]
    return _write_csv(output_path, fieldnames, rows)


def export_fed_rate_paths(workbook: Any, output_path: Path) -> int:
    sheet_name = "FED rates"
    sheet = workbook[sheet_name]
    header_row = 4
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    quarters = {
        column: str(value).strip().upper()
        for column, value in enumerate(header_values[1:], start=2)
        if str(value or "").strip()
    }
    note_values = next(sheet.iter_rows(min_row=8, max_row=8, values_only=True))
    source_notes = {column: note_values[column - 1] if len(note_values) >= column else "" for column in quarters}
    rows: list[dict[str, Any]] = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=5, max_row=7, values_only=True), start=5):
        fed_path = str((values[0] if len(values) > 0 else "") or "").strip().replace(" NZD/L", "")
        if not fed_path:
            continue
        for column, quarter in quarters.items():
            value = values[column - 1] if len(values) >= column else None
            if value is None or str(value).strip() == "":
                continue
            rows.append(
                {
                    "fed_path": fed_path,
                    "quarter": quarter,
                    "FY": _june_year(quarter),
                    "rate_nzd_per_litre": value,
                    "unit": "NZD/L",
                    "value_status": "rate_path",
                    "source_note": source_notes.get(column) or "",
                    "source_sheet": sheet_name,
                    "source_workbook_sheet": sheet_name,
                    "source_workbook_row": row_idx,
                    "source_cell": f"{get_column_letter(column)}{row_idx}",
                }
            )
    fieldnames = [
        "fed_path",
        "quarter",
        "FY",
        "rate_nzd_per_litre",
        "unit",
        "value_status",
        "source_note",
        "source_sheet",
        "source_workbook_sheet",
        "source_workbook_row",
        "source_cell",
    ]
    return _write_csv(output_path, fieldnames, rows)


def _annual_sheet_values(workbook: Any, sheet_name: str, series_name: str, line_name: str) -> dict[int, dict[str, Any]]:
    sheet = workbook[sheet_name]
    header_row = 5
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    year_columns = {
        column: year
        for column, value in enumerate(header_values, start=1)
        if (year := _as_int(value)) is not None
    }
    values_by_fy: dict[int, dict[str, Any]] = {}
    for row_idx, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        series = str((values[1] if len(values) > 1 else "") or "").strip()
        line = str((values[2] if len(values) > 2 else "") or "").strip()
        if series != series_name or line != line_name:
            continue
        for column, fy in year_columns.items():
            value = values[column - 1] if len(values) >= column else None
            if value is None or str(value).strip() == "":
                continue
            values_by_fy[fy] = {
                "value": value,
                "source_sheet": sheet_name,
                "source_workbook_sheet": sheet_name,
                "source_workbook_row": row_idx,
                "source_cell": f"{get_column_letter(column)}{row_idx}",
            }
        break
    return values_by_fy


def _official_befu25_series(workbook: Any, series_name: str) -> dict[int, dict[str, Any]]:
    sheet_name = "Official BEFU25 annual"
    sheet = workbook[sheet_name]
    header_row = 5
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    status_values = next(sheet.iter_rows(min_row=6, max_row=6, values_only=True))
    year_columns = {
        column: year
        for column, value in enumerate(header_values, start=1)
        if (year := _as_int(value)) is not None
    }
    values_by_fy: dict[int, dict[str, Any]] = {}
    for row_idx, values in enumerate(sheet.iter_rows(min_row=10, values_only=True), start=10):
        series = str((values[1] if len(values) > 1 else "") or "").strip()
        if not series:
            continue
        if series != series_name:
            continue
        for column, fy in year_columns.items():
            value = values[column - 1] if len(values) >= column else None
            if value is None or str(value).strip() == "":
                continue
            status = status_values[column - 1] if len(status_values) >= column else ""
            values_by_fy[fy] = {
                "value": value,
                "status": status,
                "source_sheet": sheet_name,
                "source_workbook_sheet": sheet_name,
                "source_workbook_row": row_idx,
                "source_cell": f"{get_column_letter(column)}{row_idx}",
            }
        break
    return values_by_fy


def _forecast_input_population(workbook: Any) -> tuple[list[dict[str, Any]], dict[int, dict[str, Any]]]:
    sheet_name = "Forecast inputs"
    sheet = workbook[sheet_name]
    header_row = 11
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [_clean_header(value) for value in header_values]
    try:
        quarter_index = headers.index("quarter")
        status_index = headers.index("status")
        population_index = headers.index("population")
    except ValueError:
        return [], {}

    quarterly: list[dict[str, Any]] = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        quarter = str((values[quarter_index] if len(values) > quarter_index else "") or "").strip().upper()
        if not _is_quarter(quarter):
            continue
        population = values[population_index] if len(values) > population_index else None
        if population is None or str(population).strip() == "":
            continue
        quarterly.append(
            {
                "quarter": quarter,
                "FY": _june_year(quarter),
                "value": population,
                "status": values[status_index] if len(values) > status_index else "",
                "source_sheet": sheet_name,
                "source_workbook_sheet": sheet_name,
                "source_workbook_row": row_idx,
                "source_cell": f"{get_column_letter(population_index + 1)}{row_idx}",
            }
        )

    annual: dict[int, dict[str, Any]] = {}
    by_fy: dict[int, list[dict[str, Any]]] = {}
    for record in quarterly:
        fy = _as_int(record.get("FY"))
        if fy is None:
            continue
        by_fy.setdefault(fy, []).append(record)
    for fy, records in by_fy.items():
        ordered = sorted(records, key=lambda item: str(item["quarter"]))
        values = [float(record["value"]) for record in ordered]
        cells = [str(record["source_cell"]) for record in ordered]
        rows = [_as_int(record["source_workbook_row"]) for record in ordered]
        source_cell = cells[0] if len(cells) == 1 else f"{cells[0]}:{cells[-1]}"
        source_row = rows[0] if len(rows) == 1 else f"{rows[0]}:{rows[-1]}"
        annual[fy] = {
            "value": sum(values) / len(values),
            "quarter_count": len(values),
            "source_sheet": sheet_name,
            "source_workbook_sheet": sheet_name,
            "source_workbook_row": source_row,
            "source_cell": source_cell,
            "status": "forecast_input_full_june_year_average" if len(values) == 4 else f"forecast_input_partial_{len(values)}_quarter_average",
        }
    return quarterly, annual


def export_ped_bridge_inputs(workbook: Any, output_path: Path) -> int:
    quarterly_population, forecast_population = _forecast_input_population(workbook)
    official_vkt = _official_befu25_series(workbook, "Light petrol VKT")
    official_vkt_per_capita = _official_befu25_series(workbook, "Light petrol VKT per capita")
    ped_vkt_per_capita = _annual_sheet_values(workbook, "Model annual", "PED VKT per capita", "Model path")
    ped_volume = _annual_sheet_values(workbook, "Model annual", "PED volume", "Model path")

    rows: list[dict[str, Any]] = []
    for record in quarterly_population:
        rows.append(
            {
                "series": "Forecast input population count",
                "period": record["quarter"],
                "FY": record["FY"],
                "time_grain": "quarterly",
                "value": record["value"],
                "unit": "persons",
                "value_status": record["status"],
                "source_basis": "Forecast inputs population path",
                "source_sheet": record["source_sheet"],
                "source_workbook_sheet": record["source_workbook_sheet"],
                "source_workbook_row": record["source_workbook_row"],
                "source_cell": record["source_cell"],
                "derivation": "direct quarterly population path row",
                "quarter_count": 1,
            }
        )

    official_population: dict[int, dict[str, Any]] = {}
    for fy in sorted(set(official_vkt) & set(official_vkt_per_capita)):
        vkt = official_vkt[fy]
        vktpc = official_vkt_per_capita[fy]
        try:
            population = float(vkt["value"]) * 1_000_000.0 / float(vktpc["value"])
        except (TypeError, ValueError, ZeroDivisionError):
            continue
        official_population[fy] = {
            "value": population,
            "quarter_count": "",
            "source_sheet": "Official BEFU25 annual",
            "source_workbook_sheet": "Official BEFU25 annual",
            "source_workbook_row": f"{vkt['source_workbook_row']};{vktpc['source_workbook_row']}",
            "source_cell": f"{vkt['source_cell']};{vktpc['source_cell']}",
            "status": "official_befu25_implied_population",
            "derivation": "Light petrol VKT million km / Light petrol VKT per capita * 1,000,000",
        }

    common_fys = sorted(set(ped_vkt_per_capita) & set(ped_volume) & (set(official_population) | set(forecast_population)))
    for fy in common_fys:
        forecast = forecast_population.get(fy)
        official = official_population.get(fy)
        use_forecast = bool(forecast and _as_int(forecast.get("quarter_count")) == 4 and fy >= 2027)
        population = forecast if use_forecast else official or forecast
        if not population:
            continue
        try:
            population_value = float(population["value"])
            vktpc_value = float(ped_vkt_per_capita[fy]["value"])
            ped_volume_value = float(ped_volume[fy]["value"])
        except (TypeError, ValueError):
            continue
        total_vkt_million_km = population_value * vktpc_value / 1_000_000.0
        if total_vkt_million_km == 0:
            continue
        litres_per_100km = ped_volume_value / total_vkt_million_km * 100.0
        population_basis = (
            "Forecast inputs population path"
            if use_forecast
            else "Official BEFU25 implied annual population"
            if population is official
            else "Forecast inputs partial population path"
        )
        source_cells = {
            "population": population["source_cell"],
            "vktpc": ped_vkt_per_capita[fy]["source_cell"],
            "litres": ped_volume[fy]["source_cell"],
        }
        rows.extend(
            [
                {
                    "series": "Population count",
                    "period": f"FY{fy}",
                    "FY": fy,
                    "time_grain": "june_year",
                    "value": population_value,
                    "unit": "persons",
                    "value_status": population["status"],
                    "source_basis": population_basis,
                    "source_sheet": population["source_sheet"],
                    "source_workbook_sheet": population["source_workbook_sheet"],
                    "source_workbook_row": population["source_workbook_row"],
                    "source_cell": population["source_cell"],
                    "derivation": population.get("derivation", "annual average of quarterly population count"),
                    "quarter_count": population.get("quarter_count", ""),
                },
                {
                    "series": "PED total VKT",
                    "period": f"FY{fy}",
                    "FY": fy,
                    "time_grain": "june_year",
                    "value": total_vkt_million_km,
                    "unit": "million km",
                    "value_status": "derived_bridge",
                    "source_basis": population_basis,
                    "source_sheet": "Forecast inputs + Model annual" if use_forecast else "Official BEFU25 annual + Model annual",
                    "source_workbook_sheet": f"{population['source_workbook_sheet']};Model annual",
                    "source_workbook_row": f"{population['source_workbook_row']};{ped_vkt_per_capita[fy]['source_workbook_row']}",
                    "source_cell": f"{source_cells['population']};{source_cells['vktpc']}",
                    "derivation": "PED VKT per capita * population / 1,000,000",
                    "quarter_count": population.get("quarter_count", ""),
                },
                {
                    "series": "PED source-backed litres",
                    "period": f"FY{fy}",
                    "FY": fy,
                    "time_grain": "june_year",
                    "value": ped_volume_value,
                    "unit": "million L",
                    "value_status": "source_backed_litres",
                    "source_basis": "Model annual PED volume",
                    "source_sheet": ped_volume[fy]["source_sheet"],
                    "source_workbook_sheet": ped_volume[fy]["source_workbook_sheet"],
                    "source_workbook_row": ped_volume[fy]["source_workbook_row"],
                    "source_cell": ped_volume[fy]["source_cell"],
                    "derivation": "direct source-backed PED litres row",
                    "quarter_count": "",
                },
                {
                    "series": "PED litres per 100km",
                    "period": f"FY{fy}",
                    "FY": fy,
                    "time_grain": "june_year",
                    "value": litres_per_100km,
                    "unit": "L/100km",
                    "value_status": "derived_bridge",
                    "source_basis": "source-backed litres divided by derived PED total VKT",
                    "source_sheet": "Model annual + PED bridge",
                    "source_workbook_sheet": f"{ped_volume[fy]['source_workbook_sheet']};{population['source_workbook_sheet']};Model annual",
                    "source_workbook_row": f"{ped_volume[fy]['source_workbook_row']};{population['source_workbook_row']};{ped_vkt_per_capita[fy]['source_workbook_row']}",
                    "source_cell": f"{source_cells['litres']};{source_cells['population']};{source_cells['vktpc']}",
                    "derivation": "PED source-backed litres / PED total VKT * 100",
                    "quarter_count": population.get("quarter_count", ""),
                },
            ]
        )

    fieldnames = [
        "series",
        "period",
        "FY",
        "time_grain",
        "value",
        "unit",
        "value_status",
        "source_basis",
        "source_sheet",
        "source_workbook_sheet",
        "source_workbook_row",
        "source_cell",
        "derivation",
        "quarter_count",
    ]
    return _write_csv(output_path, fieldnames, rows)


def export_error_bands(workbook: Any, output_path: Path) -> int:
    sheet_name = "Error bands"
    sheet = workbook[sheet_name]
    header_row = 4
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [_clean_header(value) for value in header_values]
    rows: list[dict[str, Any]] = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        series = str((values[0] if len(values) > 0 else "") or "").strip()
        if not series:
            continue
        record = {
            headers[index]: _cell_value(value)
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        record["source_sheet"] = sheet_name
        record["source_workbook_sheet"] = sheet_name
        record["source_workbook_row"] = row_idx
        record["source_cell"] = f"A{row_idx}:L{row_idx}"
        rows.append(record)
    fieldnames = headers + ["source_sheet", "source_workbook_sheet", "source_workbook_row", "source_cell"]
    return _write_csv(output_path, fieldnames, rows)


def export_official_befu25_annual(workbook: Any, output_path: Path) -> int:
    sheet_name = "Official BEFU25 annual"
    sheet = workbook[sheet_name]
    year_row = 5
    status_row = 6
    first_value_column = 5
    year_values = next(sheet.iter_rows(min_row=year_row, max_row=year_row, values_only=True))
    status_values = next(sheet.iter_rows(min_row=status_row, max_row=status_row, values_only=True))
    rows: list[dict[str, Any]] = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=8, values_only=True), start=8):
        group = str((values[0] if len(values) > 0 else "") or "").strip()
        series = str((values[1] if len(values) > 1 else "") or "").strip()
        unit = str((values[2] if len(values) > 2 else "") or "").strip()
        if not group or not series:
            continue
        if group.lower() == "group" and series.lower() == "series":
            continue
        source_row_index = values[3] if len(values) > 3 else ""
        for column, value in enumerate(values[first_value_column - 1 :], start=first_value_column):
            fy = _as_int(year_values[column - 2] if len(year_values) >= column - 1 else None)
            if fy is None:
                continue
            if value is None or str(value).strip() == "":
                continue
            status = str((status_values[column - 2] if len(status_values) >= column - 1 else "") or "").strip()
            rows.append(
                {
                    "release_round": "BEFU25",
                    "release_family": "BEFU",
                    "release_year": 2025,
                    "FY": fy,
                    "horizon": fy - 2025,
                    "status": status,
                    "group": group,
                    "series": series,
                    "value": value,
                    "unit": unit,
                    "revenue_basis": _revenue_basis(series, unit),
                    "scenario": "Official BEFU25 annual",
                    "value_status": status.lower() if status else "official_befu25_annual",
                    "source_row_index": source_row_index,
                    "source_sheet": sheet_name,
                    "source_workbook_sheet": sheet_name,
                    "source_workbook_row": row_idx,
                    "source_year_cell": f"{get_column_letter(column - 1)}{year_row}",
                    "source_status_cell": f"{get_column_letter(column - 1)}{status_row}",
                    "source_cell": f"{get_column_letter(column)}{row_idx}",
                }
            )
    fieldnames = [
        "release_round",
        "release_family",
        "release_year",
        "FY",
        "horizon",
        "status",
        "group",
        "series",
        "value",
        "unit",
        "revenue_basis",
        "scenario",
        "value_status",
        "source_row_index",
        "source_sheet",
        "source_workbook_sheet",
        "source_workbook_row",
        "source_year_cell",
        "source_status_cell",
        "source_cell",
    ]
    return _write_csv(output_path, fieldnames, rows)


def export_markdown(workbook: Any, sheet_name: str, output_path: Path, title: str) -> None:
    rows = _sheet_rows(workbook, sheet_name)
    lines = [f"# {title}", ""]
    for row in rows:
        values = [str(value).strip() for value in row if str(value).strip()]
        if not values:
            lines.append("")
        elif len(values) == 1:
            lines.append(f"## {values[0]}")
        else:
            lines.append(f"- **{values[0]}:** {' | '.join(values[1:])}")
    while lines and lines[-1] == "":
        lines.pop()
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def dashboard_controls(workbook: Any) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows = _sheet_rows(workbook, "Dashboard Controls")
    header_index = _find_header_index(rows, "Control ID")
    header = [str(value).strip() for value in rows[header_index]]
    controls: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        record = dict(zip(header, _trim(row, len(header)), strict=False))
        if not str(record.get("Control ID", "")).strip():
            continue
        options_raw = record.get("Options", "")
        try:
            options = json.loads(options_raw) if isinstance(options_raw, str) and options_raw.strip().startswith("[") else []
        except json.JSONDecodeError:
            options = []
        controls.append(
            {
                "control_id": record.get("Control ID", ""),
                "legacy_cell": record.get("Legacy cell", ""),
                "source_list": record.get("Source list", ""),
                "option_count": record.get("Option count", ""),
                "options": options,
            }
        )

    selection_rows = _sheet_rows(workbook, "Current Selections")
    selection_header_index = _find_header_index(selection_rows, "Control ID")
    selection_header = [str(value).strip() for value in selection_rows[selection_header_index]]
    selections: dict[str, Any] = {}
    for row in selection_rows[selection_header_index + 1 :]:
        record = dict(zip(selection_header, _trim(row, len(selection_header)), strict=False))
        control_id = str(record.get("Control ID", "")).strip()
        if control_id:
            selections[control_id] = {
                "workbook_cell": record.get("Workbook cell", ""),
                "current_value": record.get("Current value", ""),
            }
    return controls, selections


def write_front_end_config(workbook: Any, output_path: Path) -> None:
    controls, current_selections = dashboard_controls(workbook)
    payload = {
        "source": "Dashboard Controls and Current Selections sheets",
        "controls": controls,
        "current_selections": current_selections,
    }
    output_path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8")


def materialize(raw_workbook: Path, distilled_workbook: Path, output_dir: Path) -> dict[str, Any]:
    raw_hash = sha256(raw_workbook)
    if raw_hash.lower() != EXPECTED_RAW_SHA256:
        raise ValueError(f"Raw workbook SHA256 mismatch: expected {EXPECTED_RAW_SHA256}, got {raw_hash}")
    distilled_hash = sha256(distilled_workbook)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(distilled_workbook, read_only=True, data_only=True)
    raw = load_workbook(raw_workbook, read_only=True, data_only=True)
    try:
        row_counts: dict[str, int] = {}
        for sheet_name, (filename, first_header) in SHEET_EXPORTS.items():
            row_counts[filename] = export_table(workbook, sheet_name, output_dir / filename, first_header)
        raw_exports = {
            "release_values.csv": export_release_values(raw, output_dir / "release_values.csv"),
            "forecast_archive.csv": export_forecast_archive(raw, output_dir / "forecast_archive.csv"),
            "quarterly_actuals.csv": export_quarterly_actuals(raw, output_dir / "quarterly_actuals.csv"),
            "fed_rate_paths.csv": export_fed_rate_paths(raw, output_dir / "fed_rate_paths.csv"),
            "ped_bridge_inputs.csv": export_ped_bridge_inputs(raw, output_dir / "ped_bridge_inputs.csv"),
            "mot_error_bands.csv": export_error_bands(raw, output_dir / "mot_error_bands.csv"),
            "official_befu25_annual.csv": export_official_befu25_annual(raw, output_dir / "official_befu25_annual.csv"),
        }
        row_counts.update(raw_exports)
        export_markdown(workbook, "Read Me", output_dir / "README.md", "NLTF Revenue Model Distilled Source Pack")
        export_markdown(workbook, "Overview", output_dir / "MODEL_WORKFLOW.md", "NLTF Revenue Model Workflow")
        write_front_end_config(workbook, output_dir / "front_end_config.json")
        manifest = {
            "schema_version": "nltf-revenue-source-pack-v1",
            "source_pack_version": SOURCE_PACK_VERSION,
            "created_at": SOURCE_PACK_CREATED_AT,
            "repo_relative_output_dir": output_dir.as_posix(),
            "raw_workbook": {
                "basename": raw_workbook.name,
                "sha256": raw_hash.lower(),
                "status": "verified_lineage_only_not_committed",
            },
            "distilled_workbook": {
                "basename": distilled_workbook.name,
                "sha256": distilled_hash.lower(),
                "status": "normalized_contract_source",
            },
            "source_policy": "runtime uses repo-local normalized files only; raw workbook is lineage-only and never loaded by Streamlit",
            "normalized_files": {
                filename: {
                    "sha256": sha256(output_dir / filename),
                    "row_count": row_counts[filename],
                    "source_sheet": sheet_name,
                }
                for sheet_name, (filename, _first_header) in SHEET_EXPORTS.items()
            },
            "config_files": {
                "README.md": {"sha256": sha256(output_dir / "README.md")},
                "MODEL_WORKFLOW.md": {"sha256": sha256(output_dir / "MODEL_WORKFLOW.md")},
                "front_end_config.json": {"sha256": sha256(output_dir / "front_end_config.json")},
            },
            "workbook_sheets": workbook.sheetnames,
            "raw_workbook_extracted_sheets": raw.sheetnames,
        }
        for filename, source_sheet in {
            "release_values.csv": "MOT release series",
            "forecast_archive.csv": "Forecast archive",
            "quarterly_actuals.csv": "Quarterly actuals + Crown top-up",
            "fed_rate_paths.csv": "FED rates",
            "ped_bridge_inputs.csv": "Forecast inputs + Official BEFU25 annual + Model annual",
            "mot_error_bands.csv": "Error bands",
            "official_befu25_annual.csv": "Official BEFU25 annual",
        }.items():
            manifest["normalized_files"][filename] = {
                "sha256": sha256(output_dir / filename),
                "row_count": row_counts[filename],
                "source_sheet": source_sheet,
            }
        manifest_path = output_dir / "manifest.json"
        manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8")
        return manifest
    finally:
        workbook.close()
        raw.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw-workbook", type=Path, required=True)
    parser.add_argument("--distilled-workbook", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    manifest = materialize(args.raw_workbook, args.distilled_workbook, args.output_dir)
    print(json.dumps({"output_dir": manifest["repo_relative_output_dir"], "source_pack_version": manifest["source_pack_version"]}, sort_keys=True))


if __name__ == "__main__":
    main()
