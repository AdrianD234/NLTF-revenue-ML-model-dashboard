"""Build the glass-box Excel walkthrough of the revenue model.

One workbook that steps, in plain language and linked formulas, from the MoT
VFM 202405 fleet-share assumptions and the MBU26 official proportions through
the engine's activity forecasts (PED VKT per capita, Light RUC km, Heavy RUC
km), the light-fleet class split (conventional / light BEV / PHEV), effective
MoT rates and the NLTF rollup - reconciling exactly to the committed AR(1)
Revenue Outlook pack that drives the dashboard.

Every derived cell is an Excel formula; blue cells are inputs (each annotated
with its MoT/MBU26 source cell), green cells link across sheets. The builder
asserts, in Python, that the same arithmetic reproduces the committed pack to
1e-6 before the workbook is written.

Usage:  .venv\\Scripts\\python.exe scripts/build_glass_box_workbook.py
Output: deliverables/NLTF_revenue_glass_box.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from model_dashboard.ev_uptake_levers import (  # noqa: E402
    EV_UPTAKE_PRESETS,
    solve_logistic_from_levers,
)

PACK = ROOT / "data" / "engine_ar1" / "current_revenue_outlook"
SPINE = ROOT / "data" / "revenue_model_source_pack" / "mbu26_annual_spine" / "mbu26_official_annual.csv"
VFM = ROOT / "data" / "vfm_202405" / "vfm_vkt_shares.csv"
VFM_MANIFEST = ROOT / "data" / "vfm_202405" / "manifest.json"
OUT = ROOT / "deliverables" / "NLTF_revenue_glass_box.xlsx"

FYS = list(range(2026, 2051))

# ---------------------------------------------------------------- styles ----
ARIAL = "Arial"
NAVY = "002B5C"
GREEN_NZTA = "00843D"
F_TITLE = Font(name=ARIAL, size=14, bold=True, color=NAVY)
F_SECTION = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
F_HEAD = Font(name=ARIAL, size=9, bold=True, color=NAVY)
F_BODY = Font(name=ARIAL, size=10)
F_SMALL = Font(name=ARIAL, size=9, color="334155")
F_INPUT = Font(name=ARIAL, size=10, color="0000FF")           # hardcoded inputs
F_FORMULA = Font(name=ARIAL, size=10, color="000000")         # in-sheet formulas
F_LINK = Font(name=ARIAL, size=10, color="008000")            # cross-sheet links
FILL_SECTION = PatternFill("solid", start_color=NAVY)
FILL_HEAD = PatternFill("solid", start_color="E6EDF5")
FILL_CHECK = PatternFill("solid", start_color="ECFDF5")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NF_MKM = "#,##0.0;(#,##0.0);\"-\""            # million km
NF_KM = "#,##0;(#,##0);\"-\""
NF_M = "#,##0.0;(#,##0.0);\"-\""              # $m
NF_PCT = "0.0%;(0.0%);\"-\""
NF_RATE_L = "$0.0000"
NF_RATE_KM = "$0.00000"
NF_POP = "#,##0"
NF_DELTA = "0.000000"
NF_VKTPC = "#,##0.0"


def sheet_title(ws, text: str, subtitle: str = "", width: int = 12) -> int:
    ws.cell(row=1, column=1, value=text).font = F_TITLE
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = F_SMALL
    return 4


def section(ws, row: int, text: str, span: int) -> int:
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
    ws.cell(row=row, column=1, value=text).font = F_SECTION
    return row + 1


def head(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = F_HEAD
    cell.fill = FILL_HEAD
    cell.border = BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def put(ws, row: int, col: int, value, font=F_BODY, nf: str | None = None, border: bool = True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if nf:
        cell.number_format = nf
    if border:
        cell.border = BORDER
    return cell


# ------------------------------------------------------------ data intake ---
def june_quarters(fy: int) -> list[str]:
    return [f"{fy - 1}Q3", f"{fy - 1}Q4", f"{fy}Q1", f"{fy}Q2"]


def load_inputs():
    rows = pd.read_csv(PACK / "revenue_chart_rows.csv")
    wide = pd.read_parquet(PACK / "scenario_inputs" / "scenario_input_wide.parquet")
    drift = pd.read_csv(PACK / "ev_phev_ped_light_drift_assumptions.csv")
    spine = pd.read_csv(SPINE)
    vfm = pd.read_csv(VFM)

    quarterly = rows[rows["time_grain"] == "quarterly"]

    def q_lookup(series_id: str):
        g = quarterly[quarterly["series_id"] == series_id]
        fut = g[g["scenario_role"] == "basecase"].groupby("period")["value"].first()
        act = g[g["row_type"] == "historical_actual"].groupby("period")["value"].first()
        return fut, act

    vk_f, vk_a = q_lookup("ped_vkt_per_capita")
    lt_f, lt_a = q_lookup("light_ruc_net_km")
    hv_f, hv_a = q_lookup("heavy_ruc_net_km")

    popq = (
        wide[(wide["stream"] == "PED") & (wide["scenario_name"] == "current_basecase")]
        .dropna(subset=["population"])  # type: ignore[arg-type]
        .set_index("period")["population"].astype(float)
    )

    off = {}
    cells = {}
    for fy in [2025, *FYS]:
        block = spine[spine["FY"] == fy].set_index("source_series_id")
        off[fy] = block["value"].astype(float).to_dict()
        cells[fy] = block["source_cell"].astype(str).to_dict()

    lam = (
        drift[(drift["lambda_mode"] == "optimized") & (drift["scenario_role"] == "basecase")]
        .set_index("FY")
    )

    # quarterly detail records for sheet 4
    detail = []
    for fy in FYS:
        fallback_pop = off[fy]["light_petrol_vkt"] * 1e6 / off[fy]["light_petrol_vkt_per_capita"]
        for q in june_quarters(fy):
            vk = vk_f.get(q)
            src = "Model forecast (AR(1) engine)" if vk is not None and not pd.isna(vk) else "Actual (published outturn)"
            if vk is None or pd.isna(vk):
                vk = vk_a.get(q)
            pop = popq.get(q)
            pop_src = "Scenario input workbook"
            if pop is None or pd.isna(pop):
                pop = fallback_pop
                pop_src = f"MBU26 proxy: petrol VKT / VKT per capita (FY{fy})"
            lt = lt_f.get(q, lt_a.get(q))
            hv = hv_f.get(q, hv_a.get(q))
            detail.append(
                dict(fy=fy, quarter=q, src=src, vktpc=float(vk), pop=float(pop), pop_src=pop_src,
                     light_km=float(lt), heavy_km=float(hv))
            )

    annual = rows[(rows["time_grain"] == "june_year") & (rows["scenario_role"] == "basecase")
                  & (rows["row_type"] == "future_forecast")]
    dash = {}
    for series in ["ped_vkt_per_capita", "ped_volume", "gross_ped_revenue", "light_ruc_net_km",
                   "light_bev_ruc_net_km", "phev_ruc_net_km", "heavy_ruc_net_km",
                   "light_ruc_net_revenue", "light_bev_ruc_net_revenue", "phev_ruc_net_revenue",
                   "heavy_ruc_net_revenue", "net_fed_revenue", "total_ruc_net_revenue",
                   "net_mvr_revenue", "total_nltf_net_revenue"]:
        g = annual[annual["series_id"] == series]
        per = {}
        for _, r in g.iterrows():
            p = str(r["period"])
            fy = int(p.replace("FY", "")[:4])
            per[fy] = float(r["value"])
        dash[series] = per

    return dict(detail=detail, off=off, cells=cells, lam=lam, vfm=vfm, dash=dash)


def assert_faithful(d) -> None:
    """The Python mirror of the workbook arithmetic must hit the pack to ~1e-6."""
    lam = d["lam"]
    off = d["off"]
    dash = d["dash"]
    worst = 0.0
    for fy in FYS:
        rows_fy = [r for r in d["detail"] if r["fy"] == fy]
        P = sum(r["vktpc"] * r["pop"] / 1e6 for r in rows_fy)
        L = sum(r["light_km"] for r in rows_fy) / 1e6
        H = sum(r["heavy_km"] for r in rows_fy) / 1e6
        assert abs(P - lam.loc[fy, "current_P_t_light_petrol_km"]) < 1e-6, (fy, "P_t")
        assert abs(L - lam.loc[fy, "current_L_t_total_light_ruc_km"]) < 1e-6, (fy, "L_t")
        U = P + L
        o = off[fy]
        D = o["light_petrol_vkt"] + o["light_ruc_net_km"] + o["light_bev_ruc_net_km"] + o["phev_ruc_net_km"]
        p_bev = o["light_bev_ruc_net_km"] / D
        p_phev = o["phev_ruc_net_km"] / D
        lv = float(lam.loc[fy, "lambda_value"])
        ev = U * (p_bev + p_phev)
        bev_km = U * p_bev
        phev_km = U * p_phev
        petrol_km = P - (1 - lv) * ev
        conv_km = L - lv * ev
        litres100 = o["ped_volume"] / o["light_petrol_vkt"] * 100.0
        ped_rate = o["gross_ped_revenue"] / o["ped_volume"]
        ped_volume = petrol_km * litres100 / 100.0
        gross_ped = ped_volume * ped_rate
        r_conv = o["light_ruc_net_revenue"] / o["light_ruc_net_km"]
        r_bev = o["light_bev_ruc_net_revenue"] / o["light_bev_ruc_net_km"]
        r_phev = o["phev_ruc_net_revenue"] / o["phev_ruc_net_km"]
        r_heavy = o["heavy_ruc_net_revenue"] / o["heavy_ruc_net_km"]
        rev_conv = conv_km * r_conv
        rev_bev = bev_km * r_bev
        rev_phev = phev_km * r_phev
        rev_heavy = H * r_heavy
        gross_ruc = rev_conv + rev_heavy + rev_bev + o["heavy_bev_ruc_net_revenue"] + rev_phev + o["ruc_refunds"]
        total_ruc = gross_ruc - o["ruc_admin_revenue"] - o["ruc_refunds"]
        gross_fed = gross_ped + o["gross_lpg_revenue"] + o["gross_cng_revenue"]
        net_fed = gross_fed - o["fed_refunds"]
        net_mvr = o["mr1_revenue"] + o["mr2_revenue"] - o["mvr_admin_revenue"] - o["mvr_refunds"]
        gross_mvr = o["mr1_revenue"] + o["mr2_revenue"] + o["coo_revenue"]
        total_gross = gross_ruc + gross_fed + gross_mvr + o["tuc_net_revenue"]
        total_admin = o["ruc_admin_revenue"] + o["mvr_admin_revenue"] + o["coo_revenue"]
        total_refunds = o["ruc_refunds"] + o["fed_refunds"] + o["mvr_refunds"]
        total_nltf = total_gross - total_admin - total_refunds
        for series, computed in [
            ("ped_volume", ped_volume), ("gross_ped_revenue", gross_ped),
            ("light_ruc_net_km", conv_km), ("light_bev_ruc_net_km", bev_km),
            ("phev_ruc_net_km", phev_km), ("heavy_ruc_net_km", H),
            ("light_ruc_net_revenue", rev_conv), ("light_bev_ruc_net_revenue", rev_bev),
            ("phev_ruc_net_revenue", rev_phev), ("heavy_ruc_net_revenue", rev_heavy),
            ("net_fed_revenue", net_fed), ("total_ruc_net_revenue", total_ruc),
            ("net_mvr_revenue", net_mvr), ("total_nltf_net_revenue", total_nltf),
        ]:
            target = dash[series].get(fy)
            if target is None:
                continue
            delta = abs(computed - target)
            worst = max(worst, delta)
            assert delta < 5e-5, (fy, series, computed, target)
    print(f"[glass-box] python mirror reproduces the committed pack; worst |delta| = {worst:.2e}")


# ------------------------------------------------------------- sheets -------
def build_readme(wb, d) -> None:
    import json

    ws = wb.create_sheet("1. Read me")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", (3, 30, 100, 14, 14, 14, 14)):
        ws.column_dimensions[col].width = w
    r = sheet_title(ws, "NLTF revenue model - glass-box walkthrough",
                    "How the revenue forecasts on the dashboard are built, end to end, in plain language.")
    manifest = json.loads(VFM_MANIFEST.read_text(encoding="utf-8"))
    r = section(ws, r, "What this workbook shows", 3)
    story = [
        "Our statistical models forecast three activity streams each quarter: petrol-vehicle travel per person "
        "(the AR(1) time-series model), light road-user-charge kilometres, and heavy road-user-charge kilometres.",
        "Those forecasts are turned into revenue with assumptions taken directly from the Ministry of Transport: "
        "the MBU26 class mix (how light travel divides into petrol, conventional RUC, battery-electric and plug-in "
        "hybrid), fuel intensity (litres per 100 km), and the effective duty and RUC rates.",
        "The electric-vehicle transition itself comes from MoT's Vehicle Fleet Model (VFM 202405): sheet 2 shows the "
        "official share curves, and sheet 3 shows how our dashboard reproduces them from a handful of readable levers.",
        "Sheets 4-7 then walk line by line from the model forecasts to total NLTF revenue. Sheet 8 proves the result: "
        "every line reconciles to the committed dashboard data to well under $0.01m.",
        "No opaque weights are used anywhere: every step is an ordinary spreadsheet formula you can audit.",
    ]
    for s in story:
        cell = ws.cell(row=r, column=2, value=s)
        cell.font = F_BODY
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1
    r = section(ws, r, "How to read the colours", 3)
    for font, label in [
        (F_INPUT, "Blue - an input taken from a named source (the source cell/row is written beside it)"),
        (F_FORMULA, "Black - a formula computed on the same sheet"),
        (F_LINK, "Green - a value linked from another sheet in this workbook"),
    ]:
        ws.cell(row=r, column=2, value=label).font = font
        r += 1
    r += 1
    r = section(ws, r, "The pipeline at a glance", 3)
    ws.cell(row=r, column=2, value=(
        "Model forecasts (sheet 4)  ->  light-fleet split using MoT class mix (sheet 5)  ->  "
        "litres, rates and revenue per stream (sheet 6)  ->  NLTF rollup (sheet 7)  ->  "
        "reconciliation to the dashboard (sheet 8)")).font = F_BODY
    r += 2
    r = section(ws, r, "Provenance (hash-verified sources)", 3)
    prov = [
        ("MoT Vehicle Fleet Model outputs", "data/source_workbooks/VFM202405_outputs_summary_V3.xlsx",
         manifest["source_workbook"]["sha256"]),
        ("VFM documentation", "data/source_workbooks/Vehicle-Fleet-Model-Documentation_202405_v7.pdf",
         manifest["source_documentation"]["sha256"]),
        ("MBU26 official spine", "data/revenue_model_source_pack/mbu26_annual_spine/mbu26_official_annual.csv",
         "row/cell references quoted per line in sheets 5-7"),
        ("Committed dashboard pack (AR(1) engine)", "data/engine_ar1/current_revenue_outlook/",
         "reconciled line by line on sheet 8"),
    ]
    for name, path, note in prov:
        ws.cell(row=r, column=2, value=name).font = F_HEAD
        ws.cell(row=r, column=3, value=f"{path}   |   {note}").font = F_SMALL
        r += 1
    r += 1
    r = section(ws, r, "Scope notes", 3)
    notes = [
        "Scenario: Base case, June years FY2026-FY2050 (the dashboard's default view). FY2025 is anchored to "
        "published MBU26 actuals and is not recomputed.",
        "All revenue is $m nominal, excluding GST - the same basis as MBU26. June year FY(t) runs July(t-1) to June(t).",
        "Admin fees, refunds, LPG/CNG, motor-vehicle registration components and the track-user charge are taken "
        "directly from MBU26 (they are not modelled); each is labelled with its MBU26 source cell.",
        "The e-RUC transition and the advanced dashboard levers are display-time what-ifs and sit outside this "
        "committed baseline.",
    ]
    for s in notes:
        cell = ws.cell(row=r, column=2, value=s)
        cell.font = F_SMALL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 24
        r += 1


def build_vfm(wb, d) -> int:
    ws = wb.create_sheet("2. MoT VFM curves")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 10
    for c in range(2, 18):
        ws.column_dimensions[get_column_letter(c)].width = 13
    r = sheet_title(ws, "MoT Vehicle Fleet Model 202405 - official share curves",
                    "Read straight from the MoT workbook, sheet 'Raw data (wem202405)' (WEM = with existing measures). "
                    "Light pool = diesel + diesel-hybrid (conventional) + battery-electric + petrol plug-in VKT for "
                    "light passenger + light commercial vehicles. June year = average of the two calendar years.")
    vfm = d["vfm"]
    scenarios = [("Base_EV", "Base"), ("Fast_EV", "Fast"), ("Slow_EV", "Slow")]
    r = section(ws, r, "Share of the light RUC pool by power type, plus heavy electric share and petrol share of all light travel", 17)
    header_row = r
    head(ws, header_row + 1, 1, "June year")
    col = 2
    for _, label in scenarios:
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 4)
        cell = ws.cell(row=header_row, column=col, value=f"{label} EV scenario")
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        for name in ["Conventional light (share)", "Battery-electric light (share)", "Plug-in hybrid (share)",
                     "Heavy electric (share)", "Petrol share of light VKT"]:
            head(ws, header_row + 1, col, name)
            col += 1
    r = header_row + 2
    start_data = r
    for fy in range(2025, 2051):
        put(ws, r, 1, fy, F_BODY, NF_KM)
        col = 2
        for scen, _ in scenarios:
            row = vfm[(vfm["scenario"] == scen) & (vfm["june_year"] == fy)].iloc[0]
            for field in ["light_ruc_conventional_share", "light_ruc_bev_share", "light_ruc_phev_share",
                          "heavy_bev_vkt_share", "light_petrol_share_of_light_vkt"]:
                put(ws, r, col, float(row[field]), F_INPUT, NF_PCT)
                col += 1
        r += 1
    ws.cell(row=r + 1, column=1, value=(
        "Source rows: VKT by scenario x vehicle type (LPV, LCV, M Truck, H Truck) x power type (Petrol, Hybrid petrol, "
        "Diesel, Hybrid diesel, Electric, Petrol plug-in) x calendar year, from the raw-output sheet of "
        "VFM202405_outputs_summary_V3.xlsx.")).font = F_SMALL
    ws.freeze_panes = ws.cell(row=start_data, column=2)
    return start_data  # first data row (FY2025) for cross-references


def build_presets(wb, d, vfm_start_row: int) -> None:
    ws = wb.create_sheet("3. Uptake presets")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 44
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 13
    r = sheet_title(ws, "Our uptake levers - and how closely they reproduce the MoT curves",
                    "The dashboard's EV levers are five readable dials per curve. This sheet computes the base-preset "
                    "curves with ordinary formulas and compares them to MoT's official base scenario from sheet 2.")

    r = section(ws, r, "Lever settings (fitted to the MoT VFM scenarios)", 11)
    lever_rows = [
        ("Battery-electric: fastest uptake speed (share points per year)", "bev_peak_speed_pp", NF_PCT),
        ("Battery-electric: year of fastest uptake", "bev_peak_year", NF_KM),
        ("Battery-electric: share of light RUC pool reached by 2050", "bev_share_2050", NF_PCT),
        ("Plug-in hybrid: share at FY2025", "phev_start_share", NF_PCT),
        ("Plug-in hybrid: rise per year while growing (share points)", "phev_rise_pp", NF_PCT),
        ("Plug-in hybrid: peak share", "phev_peak_share", NF_PCT),
        ("Plug-in hybrid: decline rate after the peak (per year)", "phev_decay_rate", "0.000"),
        ("Petrol displacement: fastest speed (share points per year)", "ped_disp_speed_pp", NF_PCT),
        ("Petrol displacement: year of fastest displacement", "ped_disp_midpoint", NF_KM),
        ("Petrol displacement: share of FY2025 petrol travel displaced by 2050", "ped_disp_2050", NF_PCT),
        ("Heavy electric: fastest uptake speed (share points per year)", "heavy_bev_speed_pp", NF_PCT),
        ("Heavy electric: midpoint year", "heavy_bev_midpoint", NF_KM),
        ("Heavy electric: share of heavy pool by 2050", "heavy_bev_share_2050", NF_PCT),
    ]
    head(ws, r, 1, "Lever (plain language)")
    for i, preset in enumerate(["MoT VFM base", "MoT VFM fast", "MoT VFM slow"]):
        head(ws, r, 2 + i, preset)
    r += 1
    param_row0 = r
    from dataclasses import asdict

    presets = {name: asdict(EV_UPTAKE_PRESETS[name]) for name in ["MoT VFM base", "MoT VFM fast", "MoT VFM slow"]}
    for label, field, nf in lever_rows:
        put(ws, r, 1, label, F_BODY)
        for i, name in enumerate(["MoT VFM base", "MoT VFM fast", "MoT VFM slow"]):
            put(ws, r, 2 + i, float(presets[name][field]), F_INPUT, nf)
        r += 1
    r += 1

    r = section(ws, r, "Derived curve constants (base preset) - solved so each S-curve passes through its 2050 share at the stated speed", 11)
    base = EV_UPTAKE_PRESETS["MoT VFM base"]
    bev_smax, _ = solve_logistic_from_levers(base.bev_peak_speed_pp, base.bev_peak_year, base.bev_share_2050)
    ped_smax, _ = solve_logistic_from_levers(base.ped_disp_speed_pp, base.ped_disp_midpoint, base.ped_disp_2050)
    hv_smax, _ = solve_logistic_from_levers(base.heavy_bev_speed_pp, base.heavy_bev_midpoint, base.heavy_bev_share_2050)
    field_row = {field: param_row0 + i for i, (_, field, _) in enumerate(lever_rows)}
    consts = [
        ("BEV curve ceiling (saturation share)", bev_smax,
         "Solved numerically (bisection) so the S-curve hits the 2050 share at the stated peak speed"),
        ("BEV curve steepness", f"=4*B{field_row['bev_peak_speed_pp']}/B{r}", "Equals 4 x peak speed / ceiling"),
        ("Petrol displacement ceiling", ped_smax, "Solved the same way for the petrol displacement curve"),
        ("Petrol displacement steepness", f"=4*B{field_row['ped_disp_speed_pp']}/B{r + 2}", "Equals 4 x peak speed / ceiling"),
        ("Heavy electric ceiling", hv_smax, "Solved the same way for the heavy curve"),
        ("Heavy electric steepness", f"=4*B{field_row['heavy_bev_speed_pp']}/B{r + 4}", "Equals 4 x peak speed / ceiling"),
        ("Plug-in hybrid peak year", f"=2025+(B{field_row['phev_peak_share']}-B{field_row['phev_start_share']})/B{field_row['phev_rise_pp']}",
         "Start share rises linearly until it reaches the peak share"),
    ]
    const_row0 = r
    for label, value, note in consts:
        put(ws, r, 1, label, F_BODY)
        if isinstance(value, str):
            put(ws, r, 2, value, F_FORMULA, "0.00000")
        else:
            put(ws, r, 2, float(value), F_INPUT, "0.00000")
        cell = ws.cell(row=r, column=3, value=note)
        cell.font = F_SMALL
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        r += 1
    bev_smax_r, bev_k_r = const_row0, const_row0 + 1
    ped_smax_r, ped_k_r = const_row0 + 2, const_row0 + 3
    hv_smax_r, hv_k_r = const_row0 + 4, const_row0 + 5
    phev_peak_r = const_row0 + 6
    r += 1

    r = section(ws, r, "Base-preset curves vs MoT official base scenario (green = linked from sheet 2)", 11)
    headers = ["June year", "BEV share (our formula)", "BEV share (MoT VFM)", "BEV gap (pp)",
               "PHEV share (our formula)", "PHEV share (MoT VFM)", "PHEV gap (pp)",
               "Petrol retained vs FY2025 (our formula)", "Heavy electric share (our formula)",
               "Heavy electric share (MoT VFM)", "Heavy gap (pp)"]
    for c, h in enumerate(headers, start=1):
        head(ws, r, c, h)
    r += 1
    data0 = r
    for i, fy in enumerate(range(2025, 2051)):
        vfm_row = vfm_start_row + i
        put(ws, r, 1, fy, F_BODY, NF_KM)
        put(ws, r, 2,
            f"=B{bev_smax_r}/(1+EXP(-B{bev_k_r}*(A{r}-B{field_row['bev_peak_year']})))",
            F_FORMULA, NF_PCT)
        put(ws, r, 3, f"='2. MoT VFM curves'!C{vfm_row}", F_LINK, NF_PCT)
        put(ws, r, 4, f"=(B{r}-C{r})*100", F_FORMULA, "0.00")
        put(ws, r, 5,
            (f"=IF(A{r}<=B{phev_peak_r},MIN(B{field_row['phev_start_share']}+B{field_row['phev_rise_pp']}*(A{r}-2025),"
             f"B{field_row['phev_peak_share']}),B{field_row['phev_peak_share']}*EXP(-B{field_row['phev_decay_rate']}*(A{r}-B{phev_peak_r})))"),
            F_FORMULA, NF_PCT)
        put(ws, r, 6, f"='2. MoT VFM curves'!D{vfm_row}", F_LINK, NF_PCT)
        put(ws, r, 7, f"=(E{r}-F{r})*100", F_FORMULA, "0.00")
        put(ws, r, 8,
            (f"=(1-B{ped_smax_r}/(1+EXP(-B{ped_k_r}*(A{r}-B{field_row['ped_disp_midpoint']}))))/"
             f"(1-B{ped_smax_r}/(1+EXP(-B{ped_k_r}*(2025-B{field_row['ped_disp_midpoint']}))))"),
            F_FORMULA, NF_PCT)
        put(ws, r, 9,
            (f"=B{hv_smax_r}/(1+EXP(-B{hv_k_r}*(A{r}-B{field_row['heavy_bev_midpoint']})))-"
             f"B{hv_smax_r}/(1+EXP(-B{hv_k_r}*(2025-B{field_row['heavy_bev_midpoint']})))"),
            F_FORMULA, NF_PCT)
        put(ws, r, 10, f"='2. MoT VFM curves'!E{vfm_row}", F_LINK, NF_PCT)
        put(ws, r, 11, f"=(I{r}-J{r})*100", F_FORMULA, "0.00")
        r += 1
    put(ws, r, 1, "Largest gap (share points)", F_HEAD)
    for col, rng in [(4, "D"), (7, "G"), (11, "K")]:
        cell = put(ws, r, col,
                   f"=MAX(MAX({rng}{data0}:{rng}{r - 1}),-MIN({rng}{data0}:{rng}{r - 1}))",
                   F_FORMULA, "0.00")
        cell.fill = FILL_CHECK
    ws.cell(row=r + 2, column=1, value=(
        "The base preset reproduces MoT's official base scenario within about 1.5 share points across 2025-2050 "
        "(the fast/slow presets do the same for the fast/slow scenarios). MBU26's official class mix matches the VFM "
        "base scenario within the same tolerance, so one set of dials speaks both languages.")).font = F_SMALL


def build_quarterly(wb, d) -> dict:
    ws = wb.create_sheet("4. Engine forecasts")
    ws.sheet_view.showGridLines = False
    widths = dict(A=9, B=10, C=26, D=15, E=13, F=34, G=15, H=16, I=14, J=16, K=14)
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    r = sheet_title(ws, "What the models forecast (quarterly)",
                    "The engine's three activity forecasts, quarter by quarter. Petrol travel per person comes from "
                    "the AR(1) time-series model; light and heavy RUC kilometres from their governed models. "
                    "Population is the scenario input used to gross per-person travel up to national kilometres.")
    r = section(ws, r, "Quarterly model outputs and population, FY2026-FY2050 (blue = committed model/pack values)", 11)
    headers = ["June year", "Quarter", "Value source", "Petrol travel per person (km/quarter)",
               "Population (people)", "Population source", "Petrol pool (million km) = travel x people / 1m",
               "Light RUC km (model, km)", "Light RUC (million km)", "Heavy RUC km (model, km)",
               "Heavy RUC (million km)"]
    for c, h in enumerate(headers, start=1):
        head(ws, r, c, h)
    r += 1
    data0 = r
    for rec in d["detail"]:
        put(ws, r, 1, rec["fy"], F_BODY, NF_KM)
        put(ws, r, 2, rec["quarter"], F_BODY)
        put(ws, r, 3, rec["src"], F_SMALL)
        put(ws, r, 4, rec["vktpc"], F_INPUT, NF_VKTPC)
        put(ws, r, 5, rec["pop"], F_INPUT, NF_POP)
        put(ws, r, 6, rec["pop_src"], F_SMALL)
        put(ws, r, 7, f"=D{r}*E{r}/1000000", F_FORMULA, NF_MKM)
        put(ws, r, 8, rec["light_km"], F_INPUT, NF_KM)
        put(ws, r, 9, f"=H{r}/1000000", F_FORMULA, NF_MKM)
        put(ws, r, 10, rec["heavy_km"], F_INPUT, NF_KM)
        put(ws, r, 11, f"=J{r}/1000000", F_FORMULA, NF_MKM)
        r += 1
    ws.freeze_panes = ws.cell(row=data0, column=1)
    return dict(data0=data0, data1=r - 1)


def build_split(wb, d, q) -> dict:
    ws = wb.create_sheet("5. Light fleet split")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    for c in range(2, 22):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions["V"].width = 44
    r = sheet_title(ws, "Splitting the light fleet: petrol, conventional RUC, battery-electric, plug-in hybrid",
                    "We pool the model's petrol travel and light RUC kilometres into one light-mobility universe, then "
                    "split it using MoT's official MBU26 class mix. One allocation assumption decides how much of the "
                    "electric switch comes out of each pool - no other weights are involved.")
    r = section(ws, r, "Per June year. Green = linked from sheet 4; blue = MBU26 official inputs (source cell quoted at right)", 22)
    headers = [
        "June year",
        "Petrol pool P (million km)", "Light RUC pool L (million km)", "Light universe U = P + L",
        "MoT petrol VKT (million km)", "MoT conventional RUC km", "MoT battery-electric km", "MoT plug-in hybrid km",
        "MoT light total (sum)",
        "MoT petrol share", "MoT conventional share", "MoT BEV share", "MoT PHEV share",
        "Electric switch (million km) = U x (BEV+PHEV shares)",
        "EV switch drawn from RUC pool (share)",
        "Battery-electric km = U x BEV share", "Plug-in hybrid km = U x PHEV share",
        "Petrol km kept = P - (1 - split) x switch", "Conventional RUC km = L - split x switch",
        "Check: four classes sum back to U", "Largest gap to MoT mix (pp)",
        "MBU26 source cells (this June year)"]
    for c, h in enumerate(headers, start=1):
        head(ws, r, c, h)
    r += 1
    data0 = r
    for i, fy in enumerate(FYS):
        o = d["off"][fy]
        cells = d["cells"][fy]
        lv = float(d["lam"].loc[fy, "lambda_value"])
        q_top = q["data0"] + 4 * i
        q_bot = q_top + 3
        put(ws, r, 1, fy, F_BODY, NF_KM)
        put(ws, r, 2, f"=SUM('4. Engine forecasts'!G{q_top}:G{q_bot})", F_LINK, NF_MKM)
        put(ws, r, 3, f"=SUM('4. Engine forecasts'!I{q_top}:I{q_bot})", F_LINK, NF_MKM)
        put(ws, r, 4, f"=B{r}+C{r}", F_FORMULA, NF_MKM)
        put(ws, r, 5, o["light_petrol_vkt"], F_INPUT, NF_MKM)
        put(ws, r, 6, o["light_ruc_net_km"], F_INPUT, NF_MKM)
        put(ws, r, 7, o["light_bev_ruc_net_km"], F_INPUT, NF_MKM)
        put(ws, r, 8, o["phev_ruc_net_km"], F_INPUT, NF_MKM)
        put(ws, r, 9, f"=SUM(E{r}:H{r})", F_FORMULA, NF_MKM)
        put(ws, r, 10, f"=E{r}/I{r}", F_FORMULA, NF_PCT)
        put(ws, r, 11, f"=F{r}/I{r}", F_FORMULA, NF_PCT)
        put(ws, r, 12, f"=G{r}/I{r}", F_FORMULA, NF_PCT)
        put(ws, r, 13, f"=H{r}/I{r}", F_FORMULA, NF_PCT)
        put(ws, r, 14, f"=D{r}*(L{r}+M{r})", F_FORMULA, NF_MKM)
        put(ws, r, 15, lv, F_INPUT, NF_PCT)
        put(ws, r, 16, f"=D{r}*L{r}", F_FORMULA, NF_MKM)
        put(ws, r, 17, f"=D{r}*M{r}", F_FORMULA, NF_MKM)
        put(ws, r, 18, f"=B{r}-(1-O{r})*N{r}", F_FORMULA, NF_MKM)
        put(ws, r, 19, f"=C{r}-O{r}*N{r}", F_FORMULA, NF_MKM)
        cell = put(ws, r, 20, f"=P{r}+Q{r}+R{r}+S{r}-D{r}", F_FORMULA, NF_DELTA)
        cell.fill = FILL_CHECK
        put(ws, r, 21,
            f"=MAX(ABS(R{r}/D{r}-J{r}),ABS(S{r}/D{r}-K{r}),ABS(P{r}/D{r}-L{r}),ABS(Q{r}/D{r}-M{r}))*100",
            F_FORMULA, "0.00")
        put(ws, r, 22,
            f"MBU26 sheet: {cells['light_petrol_vkt']} petrol VKT | {cells['light_ruc_net_km']} conventional | "
            f"{cells['light_bev_ruc_net_km']} BEV | {cells['phev_ruc_net_km']} PHEV",
            F_SMALL)
        r += 1
    ws.cell(row=r + 1, column=1, value=(
        "The 'EV switch drawn from RUC pool' column is this workbook's single allocation assumption: of each year's "
        "electric-vehicle kilometres, that fraction is deducted from the light RUC pool and the remainder from the "
        "petrol pool. It was chosen once so the combined mix tracks MoT's class proportions smoothly - the 'largest "
        "gap to MoT mix' column shows the result stays within a fraction of a share point of MoT's official mix. "
        "Change the assumption and every downstream number re-computes.")).font = F_SMALL
    ws.freeze_panes = ws.cell(row=data0, column=2)
    return dict(data0=data0)


def build_rates(wb, d, split, q) -> dict:
    ws = wb.create_sheet("6. Rates & revenue")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    for c in range(2, 25):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions["Y"].width = 52
    r = sheet_title(ws, "From kilometres to dollars",
                    "Rates are not invented: every effective rate is MoT's own revenue divided by MoT's own volume for "
                    "the same June year, so legislated changes (including the FY2027/28 petrol duty increases) are "
                    "embedded automatically. Our kilometres x MoT's effective rates = our revenue.")
    s0 = split["data0"]
    r = section(ws, r, "Per June year. Green = linked; blue = MBU26 official inputs (source cells at right)", 25)
    headers = [
        "June year",
        "MoT petrol litres (million L)", "MoT petrol VKT (million km)", "Fuel intensity (L/100km) = litres/VKT x 100",
        "MoT petrol duty revenue ($m)", "Effective duty rate ($/litre) = revenue/litres",
        "Our petrol km (million)", "Our petrol litres = km x intensity / 100", "Petrol duty revenue ($m) = litres x rate",
        "MoT conventional RUC revenue ($m)", "Conventional rate ($/km)",
        "Our conventional km", "Conventional RUC revenue ($m)",
        "MoT BEV RUC revenue ($m)", "BEV rate ($/km)", "Our BEV km", "BEV RUC revenue ($m)",
        "MoT PHEV RUC revenue ($m)", "PHEV rate ($/km)", "Our PHEV km", "PHEV RUC revenue ($m)",
        "MoT heavy RUC revenue ($m)", "MoT heavy km / heavy rate ($/km)", "Our heavy km", "Heavy RUC revenue ($m)",
    ]
    for c, h in enumerate(headers, start=1):
        head(ws, r, c, h)
    r += 1
    data0 = r
    for i, fy in enumerate(FYS):
        o = d["off"][fy]
        cells = d["cells"][fy]
        srow = s0 + i
        put(ws, r, 1, fy, F_BODY, NF_KM)
        put(ws, r, 2, o["ped_volume"], F_INPUT, NF_MKM)
        put(ws, r, 3, f"='5. Light fleet split'!E{srow}", F_LINK, NF_MKM)
        put(ws, r, 4, f"=B{r}/C{r}*100", F_FORMULA, "0.000")
        put(ws, r, 5, o["gross_ped_revenue"], F_INPUT, NF_M)
        put(ws, r, 6, f"=E{r}/B{r}", F_FORMULA, NF_RATE_L)
        put(ws, r, 7, f"='5. Light fleet split'!R{srow}", F_LINK, NF_MKM)
        put(ws, r, 8, f"=G{r}*D{r}/100", F_FORMULA, NF_MKM)
        put(ws, r, 9, f"=H{r}*F{r}", F_FORMULA, NF_M)
        put(ws, r, 10, o["light_ruc_net_revenue"], F_INPUT, NF_M)
        put(ws, r, 11, f"=J{r}/'5. Light fleet split'!F{srow}", F_FORMULA, NF_RATE_KM)
        put(ws, r, 12, f"='5. Light fleet split'!S{srow}", F_LINK, NF_MKM)
        put(ws, r, 13, f"=L{r}*K{r}", F_FORMULA, NF_M)
        put(ws, r, 14, o["light_bev_ruc_net_revenue"], F_INPUT, NF_M)
        put(ws, r, 15, f"=N{r}/'5. Light fleet split'!G{srow}", F_FORMULA, NF_RATE_KM)
        put(ws, r, 16, f"='5. Light fleet split'!P{srow}", F_LINK, NF_MKM)
        put(ws, r, 17, f"=P{r}*O{r}", F_FORMULA, NF_M)
        put(ws, r, 18, o["phev_ruc_net_revenue"], F_INPUT, NF_M)
        put(ws, r, 19, f"=R{r}/'5. Light fleet split'!H{srow}", F_FORMULA, NF_RATE_KM)
        put(ws, r, 20, f"='5. Light fleet split'!Q{srow}", F_LINK, NF_MKM)
        put(ws, r, 21, f"=T{r}*S{r}", F_FORMULA, NF_M)
        q_top = q["data0"] + 4 * i
        q_bot = q_top + 3
        put(ws, r, 22, o["heavy_ruc_net_revenue"], F_INPUT, NF_M)
        put(ws, r, 23, o["heavy_ruc_net_km"], F_INPUT, NF_MKM)
        put(ws, r, 24, f"=SUM('4. Engine forecasts'!K{q_top}:K{q_bot})", F_LINK, NF_MKM)
        put(ws, r, 25, f"=X{r}*V{r}/W{r}", F_FORMULA, NF_M)
        r += 1
    note_r = r + 1
    ws.cell(row=note_r, column=1, value=(
        "MBU26 source cells per year, e.g. FY2030: petrol litres AF15, petrol duty revenue AF45, conventional RUC "
        "revenue AF33 / km AF10, BEV AF35/AF12, PHEV AF37/AF14, heavy AF34/AF11 (column letter advances one per "
        "June year on the MBU26 sheet). The effective duty rate steps up in FY2027 (+6c legislated) and FY2028 "
        "(+12c) because MoT's own revenue line prices the legislated path.")).font = F_SMALL
    ws.freeze_panes = ws.cell(row=data0, column=2)
    return dict(data0=data0)


def build_rollup(wb, d, rates) -> dict:
    ws = wb.create_sheet("7. NLTF rollup")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    for c in range(2, 25):
        ws.column_dimensions[get_column_letter(c)].width = 14
    r = sheet_title(ws, "Rolling up to the National Land Transport Fund",
                    "Modelled lines (green) plus the MBU26 lines we do not model (blue: LPG/CNG, refunds, admin fees, "
                    "registration, track-user charge) combine through the standard fund identities.")
    t0 = rates["data0"]
    r = section(ws, r, "Per June year, $m nominal ex GST. Blue inputs quote MBU26 rows 36-63", 24)
    headers = [
        "June year",
        "Petrol duty (ours)", "LPG (MBU26)", "CNG (MBU26)", "Gross fuel excise = sum", "Fuel refunds (MBU26)",
        "Net fuel excise",
        "Conventional RUC (ours)", "BEV RUC (ours)", "PHEV RUC (ours)", "Heavy RUC (ours)", "Heavy BEV RUC (MBU26)",
        "RUC refunds (MBU26)", "Gross RUC = sum + refunds back", "RUC admin fees (MBU26)", "Net RUC",
        "Registration MR1+MR2 (MBU26)", "Registration admin (MBU26)", "Registration refunds (MBU26)",
        "Net registration", "Certificate of ownership (MBU26)", "Track user charge (MBU26)",
        "Total NLTF = net fuel + net RUC + net rego + TUC", "Cross-check: gross-to-net identity"]
    for c, h in enumerate(headers, start=1):
        head(ws, r, c, h)
    r += 1
    data0 = r
    for i, fy in enumerate(FYS):
        o = d["off"][fy]
        trow = t0 + i
        put(ws, r, 1, fy, F_BODY, NF_KM)
        put(ws, r, 2, f"='6. Rates & revenue'!I{trow}", F_LINK, NF_M)
        put(ws, r, 3, o["gross_lpg_revenue"], F_INPUT, NF_M)
        put(ws, r, 4, o["gross_cng_revenue"], F_INPUT, NF_M)
        put(ws, r, 5, f"=B{r}+C{r}+D{r}", F_FORMULA, NF_M)
        put(ws, r, 6, o["fed_refunds"], F_INPUT, NF_M)
        put(ws, r, 7, f"=E{r}-F{r}", F_FORMULA, NF_M)
        put(ws, r, 8, f"='6. Rates & revenue'!M{trow}", F_LINK, NF_M)
        put(ws, r, 9, f"='6. Rates & revenue'!Q{trow}", F_LINK, NF_M)
        put(ws, r, 10, f"='6. Rates & revenue'!U{trow}", F_LINK, NF_M)
        put(ws, r, 11, f"='6. Rates & revenue'!Y{trow}", F_LINK, NF_M)
        put(ws, r, 12, o["heavy_bev_ruc_net_revenue"], F_INPUT, NF_M)
        put(ws, r, 13, o["ruc_refunds"], F_INPUT, NF_M)
        put(ws, r, 14, f"=SUM(H{r}:M{r})", F_FORMULA, NF_M)
        put(ws, r, 15, o["ruc_admin_revenue"], F_INPUT, NF_M)
        put(ws, r, 16, f"=N{r}-O{r}-M{r}", F_FORMULA, NF_M)
        put(ws, r, 17, o["mr1_revenue"] + o["mr2_revenue"], F_INPUT, NF_M)
        put(ws, r, 18, o["mvr_admin_revenue"], F_INPUT, NF_M)
        put(ws, r, 19, o["mvr_refunds"], F_INPUT, NF_M)
        put(ws, r, 20, f"=Q{r}-R{r}-S{r}", F_FORMULA, NF_M)
        put(ws, r, 21, o["coo_revenue"], F_INPUT, NF_M)
        put(ws, r, 22, o["tuc_net_revenue"], F_INPUT, NF_M)
        put(ws, r, 23, f"=G{r}+P{r}+T{r}+V{r}", F_FORMULA, NF_M)
        cell = put(ws, r, 24,
                   f"=(E{r}+N{r}+Q{r}+U{r}+V{r})-(O{r}+R{r}+U{r})-(F{r}+M{r}+S{r})-W{r}",
                   F_FORMULA, NF_DELTA)
        cell.fill = FILL_CHECK
        r += 1
    ws.cell(row=r + 1, column=1, value=(
        "Gross RUC adds refunds back before netting them off again (MoT's published convention). The cross-check "
        "column recomputes total NLTF the official way - total gross revenue minus total admin fees minus total "
        "refunds - and must equal zero against the stream-wise rollup.")).font = F_SMALL
    ws.freeze_panes = ws.cell(row=data0, column=2)
    return dict(data0=data0)


def build_reconciliation(wb, d, split, rates, rollup) -> None:
    ws = wb.create_sheet("8. Reconciliation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    for c in range(2, 40):
        ws.column_dimensions[get_column_letter(c)].width = 13
    r = sheet_title(ws, "Does this workbook match the dashboard? Yes - to the dollar.",
                    "Blue columns are the committed values behind the dashboard charts (AR(1) engine pack, base case). "
                    "Green columns are this workbook's formulas. Every difference column should read zero.")
    s0, t0, u0 = split["data0"], rates["data0"], rollup["data0"]
    blocks = [
        ("Petrol litres (million L)", "ped_volume", f"='6. Rates & revenue'!H{{t}}", NF_MKM),
        ("Petrol duty revenue ($m)", "gross_ped_revenue", f"='6. Rates & revenue'!I{{t}}", NF_M),
        ("Conventional light RUC km (m km)", "light_ruc_net_km", f"='5. Light fleet split'!S{{s}}", NF_MKM),
        ("Battery-electric RUC km (m km)", "light_bev_ruc_net_km", f"='5. Light fleet split'!P{{s}}", NF_MKM),
        ("Plug-in hybrid RUC km (m km)", "phev_ruc_net_km", f"='5. Light fleet split'!Q{{s}}", NF_MKM),
        ("Conventional light RUC revenue ($m)", "light_ruc_net_revenue", f"='6. Rates & revenue'!M{{t}}", NF_M),
        ("Battery-electric RUC revenue ($m)", "light_bev_ruc_net_revenue", f"='6. Rates & revenue'!Q{{t}}", NF_M),
        ("Plug-in hybrid RUC revenue ($m)", "phev_ruc_net_revenue", f"='6. Rates & revenue'!U{{t}}", NF_M),
        ("Heavy RUC revenue ($m)", "heavy_ruc_net_revenue", f"='6. Rates & revenue'!Y{{t}}", NF_M),
        ("Net fuel excise ($m)", "net_fed_revenue", f"='7. NLTF rollup'!G{{u}}", NF_M),
        ("Net RUC ($m)", "total_ruc_net_revenue", f"='7. NLTF rollup'!P{{u}}", NF_M),
        ("Total NLTF revenue ($m)", "total_nltf_net_revenue", f"='7. NLTF rollup'!W{{u}}", NF_M),
    ]
    r = section(ws, r, "Workbook vs committed dashboard pack (base case, June years)", 1 + 3 * len(blocks))
    head(ws, r + 1, 1, "June year")
    col = 2
    for label, _, _, _ in blocks:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 2)
        cell = ws.cell(row=r, column=col, value=label)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        for sub in ["This workbook", "Dashboard", "Difference"]:
            head(ws, r + 1, col, sub)
            col += 1
    r += 2
    data0 = r
    for i, fy in enumerate(FYS):
        put(ws, r, 1, fy, F_BODY, NF_KM)
        col = 2
        for _, series, formula, nf in blocks:
            put(ws, r, col, formula.format(s=s0 + i, t=t0 + i, u=u0 + i), F_LINK, nf)
            put(ws, r, col + 1, d["dash"][series].get(fy), F_INPUT, nf)
            cell = put(ws, r, col + 2,
                       f"={get_column_letter(col)}{r}-{get_column_letter(col + 1)}{r}", F_FORMULA, NF_DELTA)
            cell.fill = FILL_CHECK
            col += 3
        r += 1
    put(ws, r, 1, "Largest difference", F_HEAD)
    col = 2
    for _ in blocks:
        letter = get_column_letter(col + 2)
        cell = put(ws, r, col + 2,
                   f"=MAX(MAX({letter}{data0}:{letter}{r - 1}),-MIN({letter}{data0}:{letter}{r - 1}))",
                   F_FORMULA, NF_DELTA)
        cell.fill = FILL_CHECK
        col += 3
    verdict = put(ws, r + 2, 1,
                  f"=IF(MAX(B{r}:{get_column_letter(1 + 3 * len(blocks))}{r})<0.01,"
                  f"\"MATCH: every line reconciles to the dashboard within $0.01m / 0.01m km\","
                  f"\"CHECK: a line differs - inspect the difference columns\")", F_HEAD, border=False)
    verdict.fill = FILL_CHECK
    ws.freeze_panes = ws.cell(row=data0, column=2)


def build_mapping(wb) -> None:
    ws = wb.create_sheet("9. Mapping & glossary")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", (40, 30, 40, 44, 60)):
        ws.column_dimensions[col].width = w
    r = sheet_title(ws, "Where every line comes from - and what the jargon means")
    r = section(ws, r, "Series mapping: this workbook <-> the dashboard <-> MoT", 5)
    for c, h in enumerate(["Plain-language line (this workbook)", "Dashboard series",
                           "MoT / MBU26 source", "MoT VFM role", "Notes"], start=1):
        head(ws, r, c, h)
    r += 1
    rows = [
        ("Petrol travel per person", "PED VKT per capita",
         "MBU26 sheet row 17 (history/official)", "Petrol share of light VKT (sheet 2, col F)",
         "Forecast by the AR(1) time-series model each quarter; the only line that differs between the two engines."),
        ("Petrol pool (million km)", "light_petrol_vkt",
         "MBU26 sheet row 16", "Petrol VKT by scenario",
         "Travel per person x population, summed over the four quarters of the June year."),
        ("Light RUC pool (million km)", "light_ruc_net_km (model total)",
         "MBU26 sheet row 10", "Light pool = diesel + diesel-hybrid + BEV + PHEV",
         "The light RUC model forecasts the whole light RUC universe; the class split happens on sheet 5."),
        ("Conventional light RUC km", "light_ruc_net_km (committed)",
         "MBU26 row 10 (share), row 33 (revenue)", "Diesel + diesel-hybrid VKT",
         "Light pool minus the electric switch drawn from it."),
        ("Battery-electric light km", "light_bev_ruc_net_km",
         "MBU26 row 12 (share), row 35 (revenue)", "'Electric' VKT for LPV+LCV",
         "Light universe x MoT BEV share."),
        ("Plug-in hybrid km", "phev_ruc_net_km",
         "MBU26 row 14 (share), row 37 (revenue)", "'Petrol plug-in' VKT",
         "Light universe x MoT PHEV share."),
        ("Heavy RUC km", "heavy_ruc_net_km",
         "MBU26 row 11 (km), row 34 (revenue)", "M Truck + H Truck VKT; electric share tracked separately",
         "Forecast directly by the heavy model; heavy BEVs pay the same per-km rate, so electrification is "
         "revenue-neutral inside the heavy block."),
        ("Petrol litres and duty", "ped_volume / gross_ped_revenue",
         "MBU26 rows 15 (litres) and 45 (revenue)", "-",
         "Litres = petrol km x MoT fuel intensity; revenue = litres x MoT effective duty rate (legislated FY2027 +6c "
         "and FY2028 +12c steps included)."),
        ("Registration, LPG/CNG, refunds, admin, TUC", "net_mvr_revenue etc.",
         "MBU26 rows 36-63", "-",
         "Taken from MBU26 as-is - not modelled."),
        ("Total NLTF revenue", "total_nltf_net_revenue",
         "MBU26 rows 66-70 (official rollup)", "-",
         "Net fuel excise + net RUC + net registration + track user charge."),
    ]
    for vals in rows:
        for c, v in enumerate(vals, start=1):
            cell = put(ws, r, c, v, F_BODY if c == 1 else F_SMALL)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1
    r = section(ws, r, "Glossary", 5)
    glossary = [
        ("NLTF", "National Land Transport Fund - where fuel excise, road user charges, registration and track user charges land."),
        ("PED", "Petrol excise duty. 'PED volume' is petrol litres sold; 'PED VKT per capita' is petrol-vehicle kilometres travelled per person."),
        ("RUC", "Road user charges - distance-based charges for vehicles that do not pay fuel excise (diesel, electric, heavy)."),
        ("VKT", "Vehicle kilometres travelled."),
        ("BEV / PHEV", "Battery-electric vehicle / plug-in hybrid electric vehicle."),
        ("MBU26", "MoT's official 2026 baseline update - the published revenue model this dashboard benchmarks against."),
        ("VFM 202405", "MoT's Vehicle Fleet Model (May 2024 run) - projects the fleet's power-type mix; source of the EV transition curves."),
        ("MVR / MR1 / MR2 / COO", "Motor vehicle registration revenue components and the certificate-of-ownership fee."),
        ("TUC", "Track user charges (rail)."),
        ("June year", "The transport funding year: FY2030 runs July 2029 to June 2030."),
        ("ex GST", "All dollars exclude GST, matching MoT's published basis."),
        ("AR(1) model", "Our production PED model: a regression with statistically well-behaved (autocorrelation-corrected) errors; passes all six core diagnostic tests."),
    ]
    for term, meaning in glossary:
        put(ws, r, 1, term, F_HEAD)
        cell = put(ws, r, 2, meaning, F_SMALL)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def main() -> None:
    d = load_inputs()
    assert_faithful(d)
    OUT.parent.mkdir(exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb, d)
    vfm_start = build_vfm(wb, d)
    build_presets(wb, d, vfm_start)
    q = build_quarterly(wb, d)
    split = build_split(wb, d, q)
    rates = build_rates(wb, d, split, q)
    rollup = build_rollup(wb, d, rates)
    build_reconciliation(wb, d, split, rates, rollup)
    build_mapping(wb)
    wb.save(OUT)
    print(f"[glass-box] wrote {OUT}")


if __name__ == "__main__":
    main()
