"""Inventory and hash the structural source workbooks.

Section 1 of the brief. The three workbooks are immutable source/audit inputs:
they are read here to reconstruct the legacy structural method and to derive
the committed canonical artefacts, and they are never read by runtime code.

Filenames are resolved by case-insensitive glob and MUST match uniquely, so a
renamed or duplicated drop fails closed rather than silently picking one.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

OUT = REPO_ROOT / "artifacts" / "anchored_structural_shape_transition"
REFERENCES = REPO_ROOT / "references"

# (logical role, case-insensitive glob). Each must match exactly one file.
WORKBOOK_PATTERNS: tuple[tuple[str, str], ...] = (
    ("mbu26_vs_vfm_structural", "*mbu26*vfm202405*"),
    ("vfm202405_outputs_summary", "vfm202405_outputs_summary*"),
    ("befu26_revenue_forecast", "befu26 revenue forecast*"),
)


class WorkbookInventoryError(RuntimeError):
    """A source workbook is missing, duplicated or unreadable."""


def sha256_of(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_workbook(role: str, pattern: str) -> Path:
    """Case-insensitive unique resolution; fail closed on 0 or >1 matches."""

    lowered = pattern.lower()
    matches = sorted(
        path
        for path in REFERENCES.glob("*.xlsx")
        if path.is_file() and Path(path.name.lower()).match(lowered)
    )
    if not matches:
        raise WorkbookInventoryError(
            f"{role}: no workbook in references/ matches {pattern!r}."
        )
    if len(matches) > 1:
        raise WorkbookInventoryError(
            f"{role}: {pattern!r} matches {len(matches)} workbooks "
            f"({[p.name for p in matches]}); the source must be unambiguous."
        )
    return matches[0]


def _used_range(sheet) -> dict[str, object]:
    return {
        "min_row": sheet.min_row,
        "max_row": sheet.max_row,
        "min_column": sheet.min_column,
        "max_column": sheet.max_column,
        "dimensions": sheet.dimensions,
        "cell_count": (sheet.max_row - sheet.min_row + 1)
        * (sheet.max_column - sheet.min_column + 1),
    }


def inspect_workbook(role: str, path: Path) -> tuple[dict[str, object], list[dict], list[dict]]:
    """Structure, formula inventory and cached-value availability."""

    stat = path.stat()
    formulas = load_workbook(path, data_only=False, read_only=False)
    cached = load_workbook(path, data_only=True, read_only=False)

    schema_rows: list[dict[str, object]] = []
    formula_rows: list[dict[str, object]] = []
    sheet_summaries: list[dict[str, object]] = []

    for name in formulas.sheetnames:
        f_sheet = formulas[name]
        c_sheet = cached[name]
        formula_count = 0
        cached_present = 0
        cached_missing = 0
        populated = 0
        for row in f_sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                populated += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    cached_value = c_sheet[cell.coordinate].value
                    if cached_value is None:
                        cached_missing += 1
                    else:
                        cached_present += 1
                    formula_rows.append(
                        {
                            "workbook_role": role,
                            "workbook": path.name,
                            "sheet": name,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "cached_value": cached_value,
                            "cached_value_available": cached_value is not None,
                        }
                    )
        used = _used_range(f_sheet)
        schema_rows.append(
            {
                "workbook_role": role,
                "workbook": path.name,
                "sheet": name,
                "sheet_state": f_sheet.sheet_state,
                "populated_cells": populated,
                "formula_count": formula_count,
                "cached_values_present": cached_present,
                "cached_values_missing": cached_missing,
                **used,
            }
        )
        sheet_summaries.append(
            {
                "sheet": name,
                "dimensions": used["dimensions"],
                "formula_count": formula_count,
                "cached_values_missing": cached_missing,
            }
        )

    formulas.close()
    cached.close()

    inventory = {
        "workbook_role": role,
        "path": path.relative_to(REPO_ROOT).as_posix(),
        "filename": path.name,
        "sha256": sha256_of(path),
        "size_bytes": stat.st_size,
        "modified_utc": dt.datetime.fromtimestamp(stat.st_mtime, dt.UTC).isoformat(),
        "sheet_names": [s["sheet"] for s in sheet_summaries],
        "sheet_count": len(sheet_summaries),
        "sheets": sheet_summaries,
        "total_formula_count": sum(s["formula_count"] for s in sheet_summaries),
        "total_cached_values_missing": sum(s["cached_values_missing"] for s in sheet_summaries),
        "role": "immutable source/audit input; never read by runtime code",
    }
    return inventory, schema_rows, formula_rows


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    inventories: list[dict[str, object]] = []
    schema_rows: list[dict] = []
    formula_rows: list[dict] = []

    for role, pattern in WORKBOOK_PATTERNS:
        path = resolve_workbook(role, pattern)
        inventory, schema, formula = inspect_workbook(role, path)
        inventories.append(inventory)
        schema_rows.extend(schema)
        formula_rows.extend(formula)
        print(
            f"{role}: {path.name} "
            f"({inventory['size_bytes']} bytes, {inventory['sheet_count']} sheets, "
            f"{inventory['total_formula_count']} formulas)"
        )

    pd.DataFrame(schema_rows).to_csv(OUT / "source_workbook_schema.csv", index=False)
    pd.DataFrame(formula_rows).to_csv(OUT / "source_formula_inventory.csv", index=False)
    (OUT / "source_workbook_inventory.json").write_text(
        json.dumps({"workbooks": inventories}, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    manifest = {
        "manifest_id": "anchored_structural_shape_transition_source_workbooks_v1",
        "description": (
            "Immutable source/audit workbooks for the anchored structural shape "
            "transition. Runtime code reads committed CSV/parquet/registry "
            "artefacts only; these workbooks are never loaded by Streamlit."
        ),
        "workbooks": [
            {
                "workbook_role": inv["workbook_role"],
                "filename": inv["filename"],
                "path": inv["path"],
                "sha256": inv["sha256"],
                "size_bytes": inv["size_bytes"],
                "sheet_names": inv["sheet_names"],
            }
            for inv in inventories
        ],
    }
    (OUT / "source_workbook_manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
