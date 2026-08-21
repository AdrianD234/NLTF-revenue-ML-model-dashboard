"""Official-vintage reconciliation and bridge-refresh impact builder.

Generic over the vintages registered in
``data/revenue_model_source_pack/official_vintage_registry.json``. For one
official (comparator) vintage and one bridge-assumption vintage it produces a
complete source-lineage, reconciliation, cross-vintage delta, bridge-impact,
policy-basis and selector-contract artifact set under
``artifacts/official_vintage_<vintage>/``.

Governance rules enforced here:

- Published source values are surfaced, never corrected. Known published
  residuals (the gross-RUC closure defect family) are classified as
  ``published_source_residual``; any OTHER residual_reported row is an
  unexpected source inconsistency and the script exits non-zero.
- Bridge-refresh impact is computed ONLY against the pinned pre-refresh
  baseline files (``pre_bridge_refresh_chart_rows_<engine>.csv``); their
  hashes are verified against the pinned baseline manifest first. Impact is
  never computed against a live regeneration.
- The financial decomposition is a FINANCIAL closure, not causal driver
  attribution: official GDP/unemployment/price/judgment inputs are not
  supplied and receive no fabricated dollar attribution.
- Artifacts contain repo-relative paths only.

Usage:
    python scripts/build_official_vintage_reconciliation.py \
        --official-vintage BEFU26 --bridge-vintage BEFU26 \
        [--output-dir artifacts/official_vintage_befu26]

Prints ``OFFICIAL_VINTAGE_RECONCILIATION_OK <n_artifacts>`` on success.
"""
from __future__ import annotations

import argparse
import json
import math
import os
import shutil
import sys
from pathlib import Path
from typing import Any

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
# The policy-basis stage replays the app's cached overlay helpers; the engine
# default is set before the (deferred) app import exactly as the corrected
# MBU26 reconciliation precedent does.
os.environ.setdefault("DASHBOARD_ENGINE_DEFAULT", "ar1")

from model_dashboard.official_vintage import (  # noqa: E402
    CANONICAL_SERIES_DEFINITIONS,
    PERCENTAGE_CHANGE_ANCHOR,
    SECTION_ANCHOR_LABELS,
    OfficialVintageError,
    default_bridge_vintage_id,
    default_comparator_vintage_id,
    extract_annual_percentage_changes,
    load_official_vintage,
    official_comparator_trace_name,
    official_vintage_entry,
    sha256,
)

TOL = 1e-6
FIXED_ROW_TOL = 1e-9
ENGINES: dict[str, Path] = {
    "ensemble": Path("data") / "current_revenue_outlook",
    "ar1": Path("data") / "engine_ar1" / "current_revenue_outlook",
}
CURRENT_SCENARIOS = ("current_basecase", "current_comparison_1")
CURRENT_BASE_TRACE = "Current finalist Base case"

# Known published-source residual rows per vintage (the gross-RUC closure
# defect family). These are published facts, surfaced and classified; any
# other residual_reported row is unexpected and fails the run.
KNOWN_PUBLISHED_SOURCE_RESIDUALS: dict[str, set[tuple[str, int]]] = {
    "MBU26": {("gross_ruc_revenue", fy) for fy in (2027, 2028, 2029, 2030)},
    "BEFU26": {("gross_ruc_revenue", fy) for fy in (2027, 2028, 2029, 2030)},
    # PREBU26's ST_FORECAST block runs FY2027-FY2031, and every year of it
    # carries the same published gross-RUC closure defect.
    "PREBU26": {("gross_ruc_revenue", fy) for fy in (2027, 2028, 2029, 2030, 2031)},
}

GAP_SERIES = (
    "total_nltf_net_revenue",
    "total_ruc_net_revenue",
    "net_fed_revenue",
    "net_mvr_revenue",
    "tuc_net_revenue",
    "gross_ped_revenue",
    "light_ruc_net_revenue",
    "heavy_ruc_net_revenue",
)
SUMMARY_SERIES_GROUPS = (
    ("total_nltf_net_revenue", "Totals"),
    ("total_ruc_net_revenue", "RUC"),
    ("net_fed_revenue", "FED"),
    ("gross_ped_revenue", "FED"),
    ("net_mvr_revenue", "MVR"),
    ("tuc_net_revenue", "TUC"),
)
HORIZON_BUCKETS = (
    ("FY2026-2030", 2026, 2030),
    ("FY2031-2050", 2031, 2050),
    ("FY2051-2055", 2051, 2055),
)
RUC_CLASS_LEAVES = (
    "light_ruc_net_revenue",
    "heavy_ruc_net_revenue",
    "light_bev_ruc_net_revenue",
    "heavy_bev_ruc_net_revenue",
    "phev_ruc_net_revenue",
)
FIXED_ROWS = (
    "heavy_bev_ruc_net_revenue",
    "gross_lpg_revenue",
    "gross_cng_revenue",
    "fed_refunds",
    "ruc_admin_revenue",
    "ruc_refunds",
    "net_mvr_revenue",
    "tuc_net_revenue",
)
POLICY_STREAMS = (
    ("total_nltf", "total_nltf_net_revenue"),
    ("net_fed", "net_fed_revenue"),
    ("gross_ped", "gross_ped_revenue"),
    ("total_ruc", "total_ruc_net_revenue"),
    ("light_ruc_conventional", "light_ruc_net_revenue"),
    ("light_bev", "light_bev_ruc_net_revenue"),
    ("phev", "phev_ruc_net_revenue"),
    ("heavy_ruc_conventional", "heavy_ruc_net_revenue"),
)
POLICY_FYS = (2026, 2027, 2028, 2029, 2030)
DECOMP_NOTE = (
    "financial_decomposition_not_causal_attribution: official "
    "GDP/unemployment/price/judgment inputs are not supplied by the published "
    "vintage; no causal driver dollars are fabricated. See "
    "driver_availability_matrix.csv."
)
COMPOSITION_NOTE = (
    "OPT-IN CANDIDATE ONLY: this audit surfaces the official vintage's "
    "embedded Light-RUC class mix against the exact-VFM shares. The Current "
    "model composition is UNCHANGED pending separate approval; nothing in "
    "this artifact alters any runtime pack."
)

FORBIDDEN_PATH_FRAGMENTS = ("c:\\users", "c:/users", "downloads")


class Checks:
    """Ordered check register feeding the validation report and exit code."""

    def __init__(self) -> None:
        self.rows: list[dict[str, Any]] = []
        self.failures: list[str] = []

    def record(self, name: str, passed: bool, detail: str = "") -> None:
        self.rows.append({"check": name, "status": "PASS" if passed else "FAIL", "detail": detail})
        if not passed:
            self.failures.append(f"{name}: {detail}")


def repo_rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def numeric(value: Any) -> float | None:
    coerced = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return float(coerced) if pd.notna(coerced) else None


def values_map(official_annual: pd.DataFrame) -> dict[tuple[str, int], float]:
    out: dict[tuple[str, int], float] = {}
    for row in official_annual.itertuples():
        value = numeric(row.value)
        if value is not None:
            out[(str(row.series_id), int(row.FY))] = value
    return out


def period_status_map(official_annual: pd.DataFrame) -> dict[tuple[str, int], str]:
    return {
        (str(row.series_id), int(row.FY)): str(row.period_status)
        for row in official_annual.itertuples()
    }


def write_csv(frame: pd.DataFrame, path: Path) -> None:
    frame.to_csv(path, index=False)


def write_json(payload: dict[str, Any], path: Path) -> None:
    path.write_text(json.dumps(payload, indent=2, allow_nan=False) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# Workbook artifacts (1-5, 7)
# ---------------------------------------------------------------------------


def open_workbook(entry: dict[str, Any]):
    import openpyxl

    workbook_path = ROOT / str(entry["source_workbook"])
    if not workbook_path.exists():
        raise OfficialVintageError(
            f"{entry['vintage_id']}: source workbook missing at {repo_rel(workbook_path)}"
        )
    observed = sha256(workbook_path)
    expected = str(entry["workbook_sha256"]).lower()
    if observed.lower() != expected:
        raise OfficialVintageError(
            f"{entry['vintage_id']}: workbook sha256 {observed} != registered {expected}"
        )
    values_wb = openpyxl.load_workbook(workbook_path, data_only=True)
    formulas_wb = openpyxl.load_workbook(workbook_path, data_only=False)
    return workbook_path, observed, values_wb, formulas_wb


def section_spans(sheet: Any, label_column: int) -> dict[str, tuple[int, int]]:
    anchors: list[tuple[int, str]] = []
    for row in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row, label_column).value or "").strip()
        if label in SECTION_ANCHOR_LABELS:
            anchors.append((row, label))
    spans: dict[str, tuple[int, int]] = {}
    for index, (row, label) in enumerate(anchors):
        end = anchors[index + 1][0] - 1 if index + 1 < len(anchors) else sheet.max_row
        spans[label] = (row + 1, end)
    return spans


def build_workbook_inventory(
    entry: dict[str, Any],
    pack,
    workbook_path: Path,
    workbook_hash: str,
    values_wb,
    formulas_wb,
    formula_cells: list[dict[str, Any]],
    out: Path,
) -> None:
    sheet = values_wb[str(entry["source_sheet"])]
    layout = entry.get("layout") or {}
    label_column = int(layout.get("label_column", 1))
    year_row = int(layout.get("year_header_row", 1))
    status_row = int(layout.get("period_status_row", 2))
    numeric_cells = 0
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                numeric_cells += 1
    years: list[int] = []
    blocks: dict[str, list[int]] = {}
    for column in range(1, sheet.max_column + 1):
        year = numeric(sheet.cell(year_row, column).value)
        if year is None:
            continue
        fy = int(year)
        years.append(fy)
        status = str(sheet.cell(status_row, column).value or "").strip()
        blocks.setdefault(status, []).append(fy)
    spans = section_spans(sheet, label_column)
    payload = {
        "vintage_id": str(entry["vintage_id"]),
        "release_round": str(entry["release_round"]),
        "workbook": {
            "path": repo_rel(workbook_path),
            "basename": workbook_path.name,
            "sha256": workbook_hash,
            "size_bytes": workbook_path.stat().st_size,
        },
        "sheet_names": list(values_wb.sheetnames),
        "source_sheet": str(entry["source_sheet"]),
        "used_range": sheet.calculate_dimension(),
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "year_span": {
            "min_fy": min(years),
            "max_fy": max(years),
            "n_year_columns": len(years),
        },
        "period_block_spans": {
            status: {"start_fy": min(fys), "end_fy": max(fys), "n_years": len(fys)}
            for status, fys in sorted(blocks.items())
        },
        "section_row_spans": {
            label: {"start_row": span[0], "end_row": span[1]} for label, span in spans.items()
        },
        "numeric_cell_count": numeric_cells,
        "formula_count": len(formula_cells),
        "formula_policy": "static published values; no Excel formulas stored in the grid",
        "pack": {
            "pack_dir": repo_rel(pack.pack_dir),
            "schema_version": str(pack.manifest.get("schema_version")),
            "extracted_at": str(pack.manifest.get("extracted_at")),
            "extracted_by": str(pack.manifest.get("extracted_by")),
            "sentinels_validated": pack.manifest.get("sentinels_validated"),
        },
    }
    write_json(payload, out / "workbook_inventory.json")


def build_workbook_schema(
    entry: dict[str, Any], pack, values_wb, out: Path
) -> None:
    sheet = values_wb[str(entry["source_sheet"])]
    layout = entry.get("layout") or {}
    label_column = int(layout.get("label_column", 1))
    spans = section_spans(sheet, label_column)
    row_to_series = {
        int(row.source_row): str(row.series_id)
        for row in pack.annual_spine[["source_row", "series_id"]].drop_duplicates().itertuples()
    }
    first_anchor_row = min(start - 1 for start, _end in spans.values())
    rows: list[dict[str, Any]] = []
    for row_index in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row_index, label_column).value or "").strip()
        if not label:
            continue
        if label in spans:
            rows.append(
                {
                    "row_index": row_index,
                    "label": label,
                    "section_anchor": label,
                    "series_id": "",
                    "block_type": "anchor",
                }
            )
            continue
        if row_index < first_anchor_row:
            rows.append(
                {
                    "row_index": row_index,
                    "label": label,
                    "section_anchor": "",
                    "series_id": "",
                    "block_type": "header",
                }
            )
            continue
        containing = ""
        for anchor, (start, end) in spans.items():
            if start <= row_index <= end:
                containing = anchor
                break
        if containing == PERCENTAGE_CHANGE_ANCHOR:
            block_type, series_id = "pct_change", ""
        else:
            block_type = "level"
            series_id = row_to_series.get(row_index, "")
        rows.append(
            {
                "row_index": row_index,
                "label": label,
                "section_anchor": containing,
                "series_id": series_id,
                "block_type": block_type,
            }
        )
    write_csv(pd.DataFrame(rows), out / "workbook_schema.csv")


def build_source_lineage(entry: dict[str, Any], pack, workbook_hash: str, out: Path) -> None:
    definitions = {str(item["series_id"]): item for item in CANONICAL_SERIES_DEFINITIONS}
    rows: list[dict[str, Any]] = []
    spine = pack.annual_spine
    for series_id, group in spine.groupby("series_id", sort=False):
        if str(series_id) not in definitions:
            continue
        ordered = group.sort_values("FY")
        rows.append(
            {
                "series_id": str(series_id),
                "source_label": str(ordered["source_label"].iloc[0]),
                "worksheet_row": int(ordered["source_row"].iloc[0]),
                "cell_range": f"{ordered['source_cell'].iloc[0]}:{ordered['source_cell'].iloc[-1]}",
                "sheet": str(entry["source_sheet"]),
                "workbook_basename": Path(str(entry["source_workbook"])).name,
                "workbook_sha256": workbook_hash,
            }
        )
    frame = pd.DataFrame(rows).sort_values("worksheet_row").reset_index(drop=True)
    write_csv(frame, out / "source_lineage.csv")


def build_source_workbook_manifest(
    entry: dict[str, Any], workbook_path: Path, workbook_hash: str, out: Path
) -> None:
    snapshot_fields = (
        "vintage_id",
        "display_name",
        "release_round",
        "release_date",
        "schema_version",
        "actual_start_fy",
        "actual_end_fy",
        "short_forecast_start_fy",
        "short_forecast_end_fy",
        "long_forecast_start_fy",
        "long_forecast_end_fy",
        "source_horizon_fy",
        "source_pack_path",
        "pack_format",
        "available",
        "is_latest",
        "is_default_comparator",
        "is_default_bridge_vintage",
        "status",
    )
    payload = {
        "workbook_path": repo_rel(workbook_path),
        "workbook_basename": workbook_path.name,
        "size_bytes": workbook_path.stat().st_size,
        "sha256": workbook_hash,
        "sheet": str(entry["source_sheet"]),
        "registry_entry_snapshot": {key: entry.get(key) for key in snapshot_fields},
    }
    write_json(payload, out / "source_workbook_manifest.json")


def scan_formula_cells(entry: dict[str, Any], formulas_wb) -> list[dict[str, Any]]:
    sheet = formulas_wb[str(entry["source_sheet"])]
    found: list[dict[str, Any]] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                found.append(
                    {
                        "sheet": str(entry["source_sheet"]),
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "note": "excel_formula_found",
                    }
                )
    return found


def build_formula_inventory(
    entry: dict[str, Any], formula_cells: list[dict[str, Any]], out: Path
) -> None:
    rows = list(formula_cells)
    rows.append(
        {
            "sheet": str(entry["source_sheet"]),
            "cell": "",
            "formula": "",
            "note": (
                f"policy: {entry['release_round']} publishes static values; no Excel "
                "formulas are stored in the grid. Aggregate identities are asserted "
                "from the governed formula registry and residuals reported without "
                "force-balancing."
            ),
        }
    )
    write_csv(
        pd.DataFrame(rows, columns=["sheet", "cell", "formula", "note"]),
        out / "formula_inventory.csv",
    )


def build_official_annual_changes(
    entry: dict[str, Any], pack, values_wb, out: Path, checks: Checks
) -> None:
    sheet = values_wb[str(entry["source_sheet"])]
    published = extract_annual_percentage_changes(sheet, entry)
    if published.empty:
        checks.record("annual_percentage_change_block", False, "no percentage-change rows extracted")
        return
    levels = values_map(pack.official_annual)
    fys = sorted({int(fy) for fy in published["FY"]})

    def recomputed_fraction(series_id: str, fy: int) -> float | None:
        current = levels.get((series_id, fy))
        prior = levels.get((series_id, fy - 1))
        if current is None or prior is None or prior == 0.0:
            return None
        return current / prior - 1.0

    # Detect whether the workbook stores fractions or percents by magnitude.
    ratios: list[float] = []
    for row in published.itertuples():
        pub = numeric(row.pct_change)
        frac = recomputed_fraction(str(row.series_id), int(row.FY))
        if pub is not None and frac is not None and abs(frac) > 1e-9:
            ratios.append(abs(pub) / abs(frac))
    if not ratios:
        detected_unit, scale = "undetermined", 1.0
    else:
        median_ratio = sorted(ratios)[len(ratios) // 2]
        if 0.2 < median_ratio < 5.0:
            detected_unit, scale = "fraction", 1.0
        elif 20.0 < median_ratio < 500.0:
            detected_unit, scale = "percent", 100.0
        else:
            detected_unit, scale = f"undetermined_median_ratio_{median_ratio:.3g}", 1.0
    rows: list[dict[str, Any]] = []
    for row in published.itertuples():
        series_id = str(row.series_id)
        fy = int(row.FY)
        pub = numeric(row.pct_change)
        frac = recomputed_fraction(series_id, fy)
        recomputed = frac * scale if frac is not None else None
        if pub is None or recomputed is None:
            residual = None
            relative = None
            status = "not_computable"
        else:
            residual = pub - recomputed
            relative = abs(residual) / max(abs(recomputed), 1e-12)
            status = "reconciled" if (relative <= TOL or abs(residual) <= 1e-9) else "residual_reported"
        rows.append(
            {
                "series_id": series_id,
                "source_label": str(row.source_label),
                "source_row": int(row.source_row),
                "source_cell": str(row.source_cell),
                "FY": fy,
                "period_status": str(row.period_status),
                "published_pct_change": pub,
                "recomputed_pct_change": recomputed,
                "detected_workbook_unit": detected_unit,
                "residual": residual,
                "relative_residual": relative,
                "status": status,
            }
        )
    frame = pd.DataFrame(rows)
    write_csv(frame, out / "official_annual_changes.csv")
    n_reported = int(frame["status"].eq("residual_reported").sum())
    checks.record(
        "annual_percentage_change_block",
        True,
        f"unit={detected_unit}; {len(frame)} rows over FY{min(fys)}-FY{max(fys)}; "
        f"{n_reported} published rows differ from level-implied change beyond 1e-6 "
        "relative (published rounding; reported, not corrected)",
    )


# ---------------------------------------------------------------------------
# Reconciliation classification (8)
# ---------------------------------------------------------------------------


def classify_reconciliation(
    vintage_id: str, frame: pd.DataFrame, checks: Checks, artifact_name: str
) -> pd.DataFrame:
    known = KNOWN_PUBLISHED_SOURCE_RESIDUALS.get(vintage_id, set())
    classified = frame.copy()
    labels: list[str] = []
    unexpected: list[str] = []
    for row in classified.itertuples():
        status = str(row.status)
        key = (str(row.output_series_id), int(row.FY))
        if status == "reconciled":
            labels.append("reconciled")
        elif status == "residual_reported" and key in known:
            labels.append("published_source_residual")
        else:
            labels.append("unexpected_source_inconsistency")
            unexpected.append(f"{key[0]} FY{key[1]} status={status} residual={row.residual}")
    classified["classification"] = labels
    checks.record(
        f"{artifact_name}_no_unexpected_residuals",
        not unexpected,
        "; ".join(unexpected) if unexpected else f"{len(classified)} rows classified",
    )
    return classified


# ---------------------------------------------------------------------------
# Effective rates (9) and class shares (10)
# ---------------------------------------------------------------------------


def safe_ratio(numerator: float | None, denominator: float | None, scale: float = 1.0) -> float | None:
    if numerator is None or denominator is None or denominator == 0.0:
        return None
    value = numerator / denominator * scale
    return value if math.isfinite(value) else None


def build_effective_rate_audit(
    official_id: str,
    compare_id: str,
    official_values: dict[tuple[str, int], float],
    compare_values: dict[tuple[str, int], float],
    fys: list[int],
    out: Path,
    checks: Checks,
) -> None:
    metrics = (
        ("ped_effective_rate", "gross_ped_revenue", "ped_volume", 1.0, "$/L"),
        ("petrol_fleet_intensity", "ped_volume", "light_petrol_vkt", 100.0, "L/100km"),
        ("light_ruc_effective_rate", "light_ruc_net_revenue", "light_ruc_net_km", 1.0, "$/km"),
        ("heavy_ruc_effective_rate", "heavy_ruc_net_revenue", "heavy_ruc_net_km", 1.0, "$/km"),
        ("light_bev_ruc_effective_rate", "light_bev_ruc_net_revenue", "light_bev_ruc_net_km", 1.0, "$/km"),
        ("heavy_bev_ruc_effective_rate", "heavy_bev_ruc_net_revenue", "heavy_bev_ruc_net_km", 1.0, "$/km"),
        ("phev_ruc_effective_rate", "phev_ruc_net_revenue", "phev_ruc_net_km", 1.0, "$/km"),
    )
    rows: list[dict[str, Any]] = []
    non_finite = 0
    for fy in fys:
        for metric, num_series, den_series, scale, unit in metrics:
            official = safe_ratio(
                official_values.get((num_series, fy)), official_values.get((den_series, fy)), scale
            )
            compare = safe_ratio(
                compare_values.get((num_series, fy)), compare_values.get((den_series, fy)), scale
            )
            delta = official - compare if official is not None and compare is not None else None
            for value in (official, compare, delta):
                if value is not None and not math.isfinite(value):
                    non_finite += 1
            rows.append(
                {
                    "FY": fy,
                    "metric": metric,
                    "unit": unit,
                    "official_vintage": official_id,
                    "comparison_vintage": compare_id,
                    f"{official_id.lower()}_value": official,
                    f"{compare_id.lower()}_value": compare,
                    f"delta_{official_id.lower()}_minus_{compare_id.lower()}": delta,
                }
            )
    write_csv(pd.DataFrame(rows), out / "effective_rate_audit.csv")
    checks.record(
        "effective_rate_audit_finite",
        non_finite == 0,
        f"{non_finite} non-finite rate values" if non_finite else "all rates finite or NA (div-by-zero guarded)",
    )


def build_class_share_audit(
    official_id: str,
    official_values: dict[tuple[str, int], float],
    fys: list[int],
    out: Path,
    checks: Checks,
) -> None:
    vfm_path = ROOT / "data" / "vfm_202405" / "vfm_vkt_shares.csv"
    vfm = pd.read_csv(vfm_path)
    vfm_shares: dict[tuple[str, int], dict[str, float]] = {}
    for row in vfm.itertuples():
        vfm_shares[(str(row.scenario), int(row.june_year))] = {
            "conventional": float(row.light_ruc_conventional_share),
            "bev": float(row.light_ruc_bev_share),
            "phev": float(row.light_ruc_phev_share),
        }
    rows: list[dict[str, Any]] = []
    bad_sums = 0
    any_candidate = False
    for fy in fys:
        conventional = official_values.get(("light_ruc_net_km", fy))
        bev = official_values.get(("light_bev_ruc_net_km", fy))
        phev = official_values.get(("phev_ruc_net_km", fy))
        pool = None
        shares = {"conventional": None, "bev": None, "phev": None}
        if conventional is not None and bev is not None and phev is not None:
            pool = conventional + bev + phev
            if pool > 0:
                shares = {
                    "conventional": conventional / pool,
                    "bev": bev / pool,
                    "phev": phev / pool,
                }
                if abs(sum(shares.values()) - 1.0) > 1e-9:
                    bad_sums += 1
        record: dict[str, Any] = {
            "FY": fy,
            "official_vintage": official_id,
            "light_pool_million_km": pool,
            f"{official_id.lower()}_conventional_share": shares["conventional"],
            f"{official_id.lower()}_bev_share": shares["bev"],
            f"{official_id.lower()}_phev_share": shares["phev"],
        }
        for scenario, prefix in (("Base_EV", "vfm_base"), ("Fast_EV", "vfm_fast"), ("Slow_EV", "vfm_slow")):
            vfm_row = vfm_shares.get((scenario, fy))
            for cls in ("conventional", "bev", "phev"):
                record[f"{prefix}_{cls}_share"] = vfm_row[cls] if vfm_row else None
        base = vfm_shares.get(("Base_EV", fy))
        candidate = False
        for cls in ("conventional", "bev", "phev"):
            delta = (
                shares[cls] - base[cls]
                if base is not None and shares[cls] is not None
                else None
            )
            record[f"delta_vs_vfm_base_{cls}"] = delta
            if delta is not None and abs(delta) > 0.01:
                candidate = True
        record["composition_refresh_candidate"] = candidate if base is not None and pool else False
        record["notes"] = COMPOSITION_NOTE
        any_candidate = any_candidate or bool(record["composition_refresh_candidate"])
        rows.append(record)
    write_csv(pd.DataFrame(rows), out / "class_share_audit.csv")
    checks.record(
        "class_share_sums_to_one",
        bad_sums == 0,
        f"{bad_sums} FYs violate share sum" if bad_sums else "shares sum to 1 within 1e-9 where all classes present",
    )
    checks.record(
        "composition_refresh_candidate_flagged",
        True,
        ("candidate deltas > 0.01 vs VFM Base exist (opt-in only; Current composition unchanged)"
         if any_candidate
         else "no share delta vs VFM Base exceeds 0.01; composition refresh not indicated"),
    )


# ---------------------------------------------------------------------------
# Cross-vintage deltas (11, 12)
# ---------------------------------------------------------------------------


def build_vintage_deltas(
    official_id: str,
    compare_id: str,
    official_annual: pd.DataFrame,
    compare_annual: pd.DataFrame,
    out: Path,
    checks: Checks,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    off_values = values_map(official_annual)
    cmp_values = values_map(compare_annual)
    off_status = period_status_map(official_annual)
    cmp_status = period_status_map(compare_annual)
    off_keys = set(off_status)
    cmp_keys = set(cmp_status)
    checks.record(
        "vintage_series_fy_key_parity",
        off_keys == cmp_keys,
        f"official-only={len(off_keys - cmp_keys)}, compare-only={len(cmp_keys - off_keys)}"
        if off_keys != cmp_keys
        else f"{len(off_keys)} shared series/FY keys",
    )
    meta = {
        str(row.series_id): (str(row.section), str(row.unit))
        for row in official_annual[["series_id", "section", "unit"]].drop_duplicates().itertuples()
    }
    rows: list[dict[str, Any]] = []
    revisions: list[dict[str, Any]] = []
    for series_id, fy in sorted(off_keys & cmp_keys):
        official = off_values.get((series_id, fy))
        compare = cmp_values.get((series_id, fy))
        delta = official - compare if official is not None and compare is not None else None
        pct = (
            100.0 * delta / abs(compare)
            if delta is not None and compare not in (None, 0.0)
            else None
        )
        status_off = off_status.get((series_id, fy), "")
        status_cmp = cmp_status.get((series_id, fy), "")
        note = ""
        if (
            delta is not None
            and abs(delta) > 1e-9
            and status_off == "ACTUAL"
            and status_cmp == "ACTUAL"
        ):
            note = "published_actual_revision"
            revisions.append({"series_id": series_id, "FY": fy, "delta": delta})
        section, unit = meta.get(series_id, ("", ""))
        rows.append(
            {
                "series_id": series_id,
                "section": section,
                "unit": unit,
                "FY": fy,
                "official_vintage": official_id,
                "comparison_vintage": compare_id,
                f"{official_id.lower()}_value": official,
                f"{compare_id.lower()}_value": compare,
                f"{official_id.lower()}_period_status": status_off,
                f"{compare_id.lower()}_period_status": status_cmp,
                "delta": delta,
                "pct_delta": pct,
                "note": note,
            }
        )
    frame = pd.DataFrame(rows)
    write_csv(frame, out / f"{official_id.lower()}_vs_{compare_id.lower()}_by_series_fy.csv")
    checks.record(
        "published_actual_revisions_surfaced",
        True,
        "; ".join(f"{r['series_id']} FY{r['FY']} {r['delta']:+.6f}" for r in revisions)
        or "no ACTUAL-period revisions between the vintages",
    )

    summary_rows: list[dict[str, Any]] = []
    for series_id, group in SUMMARY_SERIES_GROUPS:
        for bucket, start, end in HORIZON_BUCKETS:
            deltas = [
                off_values[(series_id, fy)] - cmp_values[(series_id, fy)]
                for fy in range(start, end + 1)
                if (series_id, fy) in off_values and (series_id, fy) in cmp_values
            ]
            summary_rows.append(
                {
                    "series_id": series_id,
                    "section_group": group,
                    "horizon_bucket": bucket,
                    "fy_start": start,
                    "fy_end": end,
                    "n_years": len(deltas),
                    "sum_delta": sum(deltas) if deltas else None,
                    "mean_delta": sum(deltas) / len(deltas) if deltas else None,
                    "official_vintage": official_id,
                    "comparison_vintage": compare_id,
                }
            )
    write_csv(
        pd.DataFrame(summary_rows),
        out / f"{official_id.lower()}_vs_{compare_id.lower()}_summary.csv",
    )
    return frame, revisions


# ---------------------------------------------------------------------------
# Runtime pack readers
# ---------------------------------------------------------------------------


def load_chart_rows(pack_dir: Path) -> pd.DataFrame:
    return pd.read_csv(ROOT / pack_dir / "revenue_chart_rows.csv", low_memory=False)


def load_line_reconciliation(pack_dir: Path) -> pd.DataFrame:
    return pd.read_csv(ROOT / pack_dir / "revenue_line_reconciliation.csv", low_memory=False)


def load_runtime_manifest(pack_dir: Path) -> dict[str, Any]:
    return json.loads((ROOT / pack_dir / "manifest.json").read_text(encoding="utf-8"))


def current_june_rows(chart_rows: pd.DataFrame) -> pd.DataFrame:
    return chart_rows[
        chart_rows["scenario_name"].astype(str).isin(CURRENT_SCENARIOS)
        & chart_rows["time_grain"].astype(str).eq("june_year")
    ].copy()


def line_values(line: pd.DataFrame, source_path: str) -> dict[tuple[str, int], float]:
    subset = line[line["source_path"].astype(str).eq(source_path)]
    out: dict[tuple[str, int], float] = {}
    for row in subset.itertuples():
        value = numeric(row.value)
        if value is not None:
            out[(str(row.series_id), int(row.FY))] = value
    return out


# ---------------------------------------------------------------------------
# Bridge impact vs pinned baseline (13)
# ---------------------------------------------------------------------------


def build_bridge_impact(
    bridge_id: str, out: Path, checks: Checks
) -> tuple[pd.DataFrame, dict[str, dict[str, int]]]:
    manifest_path = out / "pre_bridge_refresh_baseline_manifest.json"
    if not manifest_path.exists():
        checks.record("pinned_baseline_manifest_present", False, f"missing {repo_rel(manifest_path)}")
        return pd.DataFrame(), {}
    baseline_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    checks.record(
        "pinned_baseline_manifest_present",
        True,
        f"baseline bridge vintage {baseline_manifest.get('bridge_vintage')} at "
        f"{baseline_manifest.get('base_sha')}",
    )
    impact_rows: list[pd.DataFrame] = []
    coverage: dict[str, dict[str, int]] = {}
    for engine, pack_dir in ENGINES.items():
        baseline_path = out / f"pre_bridge_refresh_chart_rows_{engine}.csv"
        expected_hash = str(baseline_manifest.get(f"{engine}_sha256", "")).lower()
        if not baseline_path.exists():
            checks.record(f"pinned_baseline_{engine}", False, f"missing {repo_rel(baseline_path)}")
            continue
        observed_hash = sha256(baseline_path)
        checks.record(
            f"pinned_baseline_{engine}",
            observed_hash.lower() == expected_hash,
            "sha256 matches pinned manifest (impact computed against the pinned file, "
            "never a live regeneration)"
            if observed_hash.lower() == expected_hash
            else f"sha256 {observed_hash} != pinned {expected_hash}",
        )
        if observed_hash.lower() != expected_hash:
            continue
        baseline_all = pd.read_csv(baseline_path, low_memory=False)
        baseline_official = int(
            baseline_all["scenario_role"].astype(str).eq("official_comparator").sum()
        )
        baseline = current_june_rows(baseline_all)
        refreshed = current_june_rows(load_chart_rows(pack_dir))
        keys = ["scenario_name", "series_id", "june_year"]
        for name, frame in (("baseline", baseline), ("refreshed", refreshed)):
            dupes = int(frame.duplicated(subset=keys).sum())
            if dupes:
                checks.record(
                    f"impact_join_uniqueness_{engine}_{name}", False, f"{dupes} duplicate keys"
                )
        merged = baseline.merge(
            refreshed,
            on=keys,
            how="outer",
            suffixes=("_baseline", "_refreshed"),
            indicator=True,
        )
        both = merged[merged["_merge"].eq("both")].copy()
        only_baseline = int(merged["_merge"].eq("left_only").sum())
        only_refreshed = int(merged["_merge"].eq("right_only").sum())
        coverage[engine] = {
            "matched_rows": int(len(both)),
            "rows_only_in_pinned_baseline": only_baseline,
            "rows_only_in_refreshed_pack": only_refreshed,
            "baseline_official_comparator_rows_excluded": baseline_official,
        }
        baseline_values = pd.to_numeric(both["value_baseline"], errors="coerce")
        refreshed_values = pd.to_numeric(both["value_refreshed"], errors="coerce")
        delta = refreshed_values - baseline_values
        pct = pd.Series(
            [
                100.0 * d / abs(b) if pd.notna(d) and pd.notna(b) and b != 0.0 else None
                for d, b in zip(delta, baseline_values)
            ],
            index=both.index,
        )
        impact_rows.append(
            pd.DataFrame(
                {
                    "engine": engine,
                    "bridge_vintage": bridge_id,
                    "scenario_name": both["scenario_name"],
                    "trace_name": both["trace_name_refreshed"],
                    "series_id": both["series_id"],
                    "june_year": both["june_year"].astype(int),
                    "baseline_value": baseline_values,
                    "refreshed_value": refreshed_values,
                    "delta": delta,
                    "pct_delta": pct,
                }
            )
        )
    impact = (
        pd.concat(impact_rows, ignore_index=True)
        if impact_rows
        else pd.DataFrame(
            columns=[
                "engine",
                "bridge_vintage",
                "scenario_name",
                "trace_name",
                "series_id",
                "june_year",
                "baseline_value",
                "refreshed_value",
                "delta",
                "pct_delta",
            ]
        )
    )
    write_csv(impact, out / "current_bridge_vintage_impact.csv")
    checks.record(
        "bridge_impact_non_empty_both_engines",
        set(impact["engine"].unique()) == set(ENGINES) and not impact.empty,
        f"{len(impact)} matched impact rows across {sorted(impact['engine'].unique())}",
    )
    return impact, coverage


# ---------------------------------------------------------------------------
# Current-vs-official gap and decomposition (14, 15)
# ---------------------------------------------------------------------------


def build_current_vs_official(
    official_id: str,
    bridge_id: str,
    official_values: dict[tuple[str, int], float],
    line_by_engine: dict[str, dict[tuple[str, int], float]],
    out: Path,
    checks: Checks,
) -> None:
    gap_fys = list(range(2026, 2051))
    gap_rows: list[dict[str, Any]] = []
    missing: list[str] = []
    for engine, current in line_by_engine.items():
        for series_id in GAP_SERIES:
            for fy in gap_fys:
                cur = current.get((series_id, fy))
                off = official_values.get((series_id, fy))
                if cur is None or off is None:
                    missing.append(f"{engine} {series_id} FY{fy}")
                    continue
                gap = cur - off
                gap_rows.append(
                    {
                        "engine": engine,
                        "series_id": series_id,
                        "FY": fy,
                        "current_value": cur,
                        "official_value": off,
                        "gap_current_minus_official": gap,
                        "pct_gap": 100.0 * gap / abs(off) if off != 0.0 else None,
                        "official_vintage": official_id,
                        "bridge_vintage": bridge_id,
                    }
                )
    write_csv(
        pd.DataFrame(gap_rows),
        out / f"current_vs_{official_id.lower()}_gap_by_stream_fy.csv",
    )
    checks.record(
        "current_vs_official_gap_coverage",
        not missing,
        f"missing series/FY: {missing[:5]}" if missing else
        f"{len(gap_rows)} rows: {len(GAP_SERIES)} streams x FY2026-2050 x {len(line_by_engine)} engines",
    )

    fixed_gaps_zero = bridge_id == official_id
    decomposition_rows: list[dict[str, Any]] = []
    worst = {"nltf": 0.0, "ruc": 0.0, "fed": 0.0, "fixed": 0.0, "current_closure": 0.0}
    for engine, current in line_by_engine.items():
        for fy in gap_fys:

            def gap(series_id: str) -> float:
                cur = current.get((series_id, fy))
                off = official_values.get((series_id, fy))
                if cur is None or off is None:
                    raise OfficialVintageError(f"{engine}: missing {series_id} FY{fy} for decomposition")
                return cur - off

            def closure(values: dict[tuple[str, int], float]) -> float:
                return values[("total_ruc_net_revenue", fy)] - (
                    sum(values[(series_id, fy)] for series_id in RUC_CLASS_LEAVES)
                    - values[("ruc_admin_revenue", fy)]
                )

            current_closure = closure(current)
            official_closure = closure({k: v for k, v in official_values.items()})
            nltf_gap = gap("total_nltf_net_revenue")
            fed_residual = gap("net_fed_revenue") - (
                gap("gross_ped_revenue") + gap("gross_lpg_revenue") + gap("gross_cng_revenue") - gap("fed_refunds")
            )
            class_gaps = {series_id: gap(series_id) for series_id in RUC_CLASS_LEAVES}
            ruc_residual = gap("total_ruc_net_revenue") - (
                sum(class_gaps.values()) - gap("ruc_admin_revenue") + current_closure - official_closure
            )
            nltf_residual = nltf_gap - (
                gap("net_fed_revenue") + gap("total_ruc_net_revenue") + gap("net_mvr_revenue") + gap("tuc_net_revenue")
            )
            fixed_abs = max(abs(gap(series_id)) for series_id in FIXED_ROWS)
            worst["nltf"] = max(worst["nltf"], abs(nltf_residual))
            worst["ruc"] = max(worst["ruc"], abs(ruc_residual))
            worst["fed"] = max(worst["fed"], abs(fed_residual))
            worst["fixed"] = max(worst["fixed"], fixed_abs)
            worst["current_closure"] = max(worst["current_closure"], abs(current_closure))
            decomposition_rows.append(
                {
                    "engine": engine,
                    "FY": fy,
                    "total_nltf_gap": nltf_gap,
                    "activity_ped_gap": gap("gross_ped_revenue"),
                    "activity_light_ruc_conventional_gap": class_gaps["light_ruc_net_revenue"],
                    "activity_light_bev_gap": class_gaps["light_bev_ruc_net_revenue"],
                    "activity_phev_gap": class_gaps["phev_ruc_net_revenue"],
                    "activity_heavy_ruc_gap": class_gaps["heavy_ruc_net_revenue"],
                    "fixed_heavy_bev_gap": class_gaps["heavy_bev_ruc_net_revenue"],
                    "fixed_lpg_gap": gap("gross_lpg_revenue"),
                    "fixed_cng_gap": gap("gross_cng_revenue"),
                    "fixed_fed_refunds_gap": gap("fed_refunds"),
                    "fixed_ruc_admin_gap": gap("ruc_admin_revenue"),
                    "fixed_ruc_refunds_gap": gap("ruc_refunds"),
                    "mvr_gap": gap("net_mvr_revenue"),
                    "tuc_gap": gap("tuc_net_revenue"),
                    "official_published_source_residual": official_closure,
                    "current_ruc_closure_residual": current_closure,
                    "net_fed_identity_residual": fed_residual,
                    "ruc_identity_residual": ruc_residual,
                    "residual_closure_term": nltf_residual,
                    "fixed_rows_zero_by_construction": fixed_gaps_zero and fixed_abs <= FIXED_ROW_TOL,
                    "notes": DECOMP_NOTE,
                }
            )
    write_csv(
        pd.DataFrame(decomposition_rows),
        out / f"current_vs_{official_id.lower()}_financial_decomposition.csv",
    )
    checks.record(
        "decomposition_residual_closure",
        worst["nltf"] <= TOL and worst["ruc"] <= TOL and worst["fed"] <= TOL,
        f"max |residual|: NLTF {worst['nltf']:.2e}, RUC {worst['ruc']:.2e}, FED {worst['fed']:.2e} (tolerance 1e-6)",
    )
    checks.record(
        "decomposition_current_side_closes",
        worst["current_closure"] <= TOL,
        f"max |current RUC closure| {worst['current_closure']:.2e}",
    )
    if fixed_gaps_zero:
        checks.record(
            "decomposition_fixed_rows_zero",
            worst["fixed"] <= FIXED_ROW_TOL,
            f"max |fixed/carried row gap| {worst['fixed']:.2e} "
            f"(zero by construction: bridge vintage == official vintage == {official_id})",
        )
    else:
        checks.record(
            "decomposition_fixed_rows_reported",
            True,
            f"bridge vintage {bridge_id} != official vintage {official_id}; fixed-row gaps "
            f"reported (max {worst['fixed']:.6f}), not asserted zero",
        )


# ---------------------------------------------------------------------------
# Policy-basis comparison (16)
# ---------------------------------------------------------------------------


def build_policy_basis(
    official_id: str,
    official_values: dict[tuple[str, int], float],
    chart_by_engine: dict[str, pd.DataFrame],
    out: Path,
    checks: Checks,
) -> None:
    import inspect

    import app  # deferred: heavy streamlit import
    from model_dashboard.revenue_outlook import (
        PED_BRIDGE_DEFAULT_MODE,
        load_revenue_outlook_pack,
    )

    default_policy = str(
        inspect.signature(app._session_fed_policy_state).parameters["default"].default
    )
    bases_coincide = default_policy == app.FED_POLICY_PUBLISHED
    checks.record(
        "policy_basis_app_default_state",
        True,
        f"app default UI Current policy state is '{default_policy}'"
        + ("; policy_normalised and actual_default_ui coincide" if bases_coincide else ""),
    )

    def overlay_values(pack, engine: str, policy: str) -> dict[tuple[str, int], float]:
        sensitivity_key = app.selected_sensitivity_key("Off", "Off", "Off", freight_rail_shift="Off")
        signature = ((f"official_vintage_reconciliation_{engine}", 0, 0),)
        key = (app.DEFAULT_EV_UPTAKE_MODE, (), (), policy, app.FED_POLICY_PUBLISHED, False, False)
        rows, *_ = app.cached_scenario_overlay_rows(
            signature, sensitivity_key, PED_BRIDGE_DEFAULT_MODE, key, pack
        )
        annual = rows[
            rows["time_grain"].astype(str).eq("june_year")
            & rows["scenario_role"].astype(str).eq("basecase")
        ]
        values: dict[tuple[str, int], float] = {}
        for row in annual.itertuples():
            fy = numeric(row.june_year)
            value = numeric(row.value)
            if fy is not None and value is not None:
                values[(str(row.series_id), int(fy))] = value
        return values

    rows: list[dict[str, Any]] = []
    for engine, pack_dir in ENGINES.items():
        committed = current_june_rows(chart_by_engine[engine])
        committed = committed[committed["scenario_name"].astype(str).eq("current_basecase")]
        committed_values: dict[tuple[str, int], float] = {}
        for row in committed.itertuples():
            value = numeric(row.value)
            fy = numeric(row.june_year)
            if value is not None and fy is not None:
                committed_values[(str(row.series_id), int(fy))] = value
        pack = load_revenue_outlook_pack(ROOT / pack_dir, repo_root=ROOT)
        overlay_published = overlay_values(pack, engine, app.FED_POLICY_PUBLISHED)
        overlay_default = (
            overlay_published if bases_coincide else overlay_values(pack, engine, default_policy)
        )
        for fy in POLICY_FYS:
            for stream, series_id in POLICY_STREAMS:
                official = official_values.get((series_id, fy))
                committed_value = committed_values.get((series_id, fy))
                default_value = overlay_default.get((series_id, fy))
                published_overlay_value = overlay_published.get((series_id, fy))
                rows.append(
                    {
                        "engine": engine,
                        "comparison_basis": "policy_normalised",
                        "fy": fy,
                        "stream": stream,
                        "series_id": series_id,
                        "current_value": committed_value,
                        "official_value": official,
                        "gap_current_minus_official": (
                            committed_value - official
                            if committed_value is not None and official is not None
                            else None
                        ),
                        "current_basis": "committed_pack_rows_published_policy",
                        "official_basis": f"{official_id}_published",
                        "app_default_policy_state": default_policy,
                        "bases_coincide": bases_coincide,
                        "note": (
                            "The committed pack rows ARE the published-policy basis. "
                            "Official side is always the published vintage; "
                            "delayed-vs-delayed is never the standard comparison."
                        ),
                    }
                )
                rows.append(
                    {
                        "engine": engine,
                        "comparison_basis": "actual_default_ui",
                        "fy": fy,
                        "stream": stream,
                        "series_id": series_id,
                        "current_value": default_value,
                        "official_value": official,
                        "gap_current_minus_official": (
                            default_value - official
                            if default_value is not None and official is not None
                            else None
                        ),
                        "current_basis": (
                            f"app_cached_overlay_default_policy_{default_policy}"
                        ),
                        "official_basis": f"{official_id}_published",
                        "app_default_policy_state": default_policy,
                        "bases_coincide": bases_coincide,
                        "current_same_chain_published_value": published_overlay_value,
                        "policy_effect": (
                            default_value - published_overlay_value
                            if default_value is not None and published_overlay_value is not None
                            else None
                        ),
                        "display_chain_effect_vs_committed": (
                            published_overlay_value - committed_value
                            if published_overlay_value is not None and committed_value is not None
                            else None
                        ),
                        "note": (
                            "Derived via the app's cached overlay helpers (same route as "
                            "scripts/build_corrected_mbu26_reconciliation.py actual_default_ui). "
                            "Official side stays published; delayed-vs-delayed is never "
                            "the standard comparison."
                            + (
                                " App default policy is 'published', so this coincides with "
                                "policy_normalised on the overlay chain."
                                if bases_coincide
                                else ""
                            )
                        ),
                    }
                )
    frame = pd.DataFrame(rows)
    write_csv(frame, out / "policy_basis_comparison.csv")
    checks.record(
        "policy_basis_both_bases_present",
        {"policy_normalised", "actual_default_ui"} <= set(frame["comparison_basis"].unique()),
        f"bases: {sorted(frame['comparison_basis'].unique())}",
    )


# ---------------------------------------------------------------------------
# Driver availability (17), horizons (18), selector audit (19)
# ---------------------------------------------------------------------------


def build_driver_availability(official_id: str, out: Path) -> None:
    unavailable = "unavailable_source_not_supplied"
    supplied = "supplied_published_output"
    rows = [
        ("gdp_path", unavailable, "not published at input grain by the official vintage", "unavailable"),
        ("population", unavailable, "only output-implied (VKT / VKT-per-capita); not independently published", "unavailable"),
        ("unemployment", unavailable, "not published at input grain by the official vintage", "unavailable"),
        ("consumer_producer_prices", unavailable, "not published at input grain by the official vintage", "unavailable"),
        ("fuel_prices", unavailable, "not published at input grain by the official vintage", "unavailable"),
        ("judgment_adjustments", unavailable, "analyst overlays are not published", "unavailable"),
        ("activity_levels", supplied, "annual activity rows (km, litres, GTK) are published", "available_observable"),
        ("effective_rates", supplied, "derivable as revenue/activity from published rows", "available_derived"),
        ("class_mix", supplied, "embedded Light-RUC class mix observable from published km rows", "available_derived"),
        ("fixed_rows", supplied, "admin/refunds/MVR/TUC/LPG/CNG rows are published", "available_observable"),
    ]
    frame = pd.DataFrame(
        [
            {
                "driver": driver,
                "official_vintage": official_id,
                "supplied_by_official_vintage": status == supplied,
                "availability": status,
                "causal_attribution": attribution,
                "note": note,
            }
            for driver, status, note, attribution in rows
        ]
    )
    write_csv(frame, out / "driver_availability_matrix.csv")


def build_source_horizon_audit(
    entries: dict[str, dict[str, Any]],
    manifests: dict[str, dict[str, Any]],
    chart_by_engine: dict[str, pd.DataFrame],
    out: Path,
) -> None:
    rows: list[dict[str, Any]] = []
    for vintage_id, entry in entries.items():
        rows.append(
            {
                "entity": vintage_id,
                "kind": "official_vintage",
                "source_horizon_fy": int(entry["source_horizon_fy"]),
                "actual_block": f"FY{entry['actual_start_fy']}-FY{entry['actual_end_fy']}",
                "short_forecast_block": f"FY{entry['short_forecast_start_fy']}-FY{entry['short_forecast_end_fy']}",
                "long_forecast_block": f"FY{entry['long_forecast_start_fy']}-FY{entry['long_forecast_end_fy']}",
                "runtime_cutoff_fy": None,
                "post_model_extension_end_fy": None,
                "official_comparator_cutoff_by_vintage": "",
                "source": repo_rel(ROOT / str(entry["source_pack_path"])),
            }
        )
    for engine, manifest in manifests.items():
        period_rule = manifest.get("period_rule") or {}
        current = current_june_rows(chart_by_engine[engine])
        rows.append(
            {
                "entity": f"current_{engine}",
                "kind": "current_runtime_pack",
                "source_horizon_fy": None,
                "actual_block": "",
                "short_forecast_block": "",
                "long_forecast_block": "",
                "runtime_cutoff_fy": period_rule.get("runtime_cutoff_fy"),
                "post_model_extension_end_fy": int(pd.to_numeric(current["june_year"]).max()),
                "official_comparator_cutoff_by_vintage": json.dumps(
                    period_rule.get("official_comparator_cutoff_by_vintage") or {}
                ),
                "source": repo_rel(ROOT / ENGINES[engine] / "manifest.json"),
            }
        )
    write_csv(pd.DataFrame(rows), out / "source_horizon_audit.csv")


def build_front_end_selector_audit(
    manifests: dict[str, dict[str, Any]], out: Path, checks: Checks, official_id: str
) -> None:
    rows: list[dict[str, Any]] = []
    for engine, manifest in manifests.items():
        block = manifest.get("official_vintages") or {}
        available = block.get("available") or {}
        for vintage_id, info in available.items():
            rows.append(
                {
                    "engine": engine,
                    "vintage_id": vintage_id,
                    "display_name": info.get("display_name"),
                    "release_round": info.get("release_round"),
                    "status": info.get("status"),
                    "trace_name": info.get("trace_name"),
                    "scenario_name": info.get("scenario_name"),
                    "source_horizon_fy": info.get("source_horizon_fy"),
                    "workbook_sha256": info.get("workbook_sha256"),
                    "source_pack_path": info.get("source_pack_path"),
                    "pack_manifest_sha256": info.get("manifest_sha256"),
                    "official_comparator_vintage_id": block.get("official_comparator_vintage_id"),
                    "bridge_assumption_vintage_id": block.get("bridge_assumption_vintage_id"),
                    "default_official_comparator_trace": block.get("default_official_comparator_trace"),
                    "registry": block.get("registry"),
                    "allowed_traces": "; ".join(manifest.get("allowed_traces") or []),
                }
            )
    frame = pd.DataFrame(rows)
    write_csv(frame, out / "front_end_selector_audit.csv")
    ok = all(
        official_id in set(frame.loc[frame["engine"].eq(engine), "vintage_id"])
        for engine in manifests
    )
    checks.record(
        "front_end_selector_official_vintage_available",
        ok,
        f"{official_id} present in official_vintages.available for engines {sorted(manifests)}",
    )


# ---------------------------------------------------------------------------
# Validation report (20)
# ---------------------------------------------------------------------------


def build_validation_report(
    official_id: str,
    bridge_id: str,
    compare_id: str,
    checks: Checks,
    classified: pd.DataFrame,
    revisions: list[dict[str, Any]],
    impact: pd.DataFrame,
    coverage: dict[str, dict[str, int]],
    class_share_path: Path,
    out: Path,
) -> None:
    lines = [
        f"# Official vintage reconciliation: {official_id} (bridge: {bridge_id})",
        "",
        f"Comparator vintage for cross-vintage deltas: {compare_id}.",
        "All artifact paths are repo-relative; published source values are surfaced,",
        "never corrected.",
        "",
        "## Checks",
        "",
        "| Check | Status | Detail |",
        "|---|---|---|",
    ]
    for row in checks.rows:
        detail = str(row["detail"]).replace("|", "\\|")
        lines.append(f"| {row['check']} | {row['status']} | {detail} |")

    lines += ["", "## Published-source residuals (named, quantified, retained)", ""]
    published = classified[classified["classification"].eq("published_source_residual")]
    if published.empty:
        lines.append("None found in the official vintage's formula reconciliation.")
    else:
        lines.append(
            f"{official_id} `gross_ruc_revenue` fails to close to its class leaves plus "
            "refunds (same defect family as the MBU26 FY2027 +0.627 residual). The "
            "published values are retained; nothing is force-balanced:"
        )
        lines.append("")
        for row in published.sort_values("FY").itertuples():
            lines.append(
                f"- FY{int(row.FY)}: residual {float(row.residual):+.6f} $m "
                f"({row.output_series_id}, status {row.status}, classified published_source_residual)"
            )

    lines += ["", "## Published ACTUAL revisions vs the prior vintage", ""]
    if revisions:
        for revision in revisions:
            lines.append(
                f"- {revision['series_id']} FY{revision['FY']}: {official_id} - {compare_id} = "
                f"{revision['delta']:+.6f} (note `published_actual_revision`)"
            )
    else:
        lines.append("No ACTUAL-period value revisions between the two vintages.")

    lines += ["", "## Bridge-refresh impact headline (Current finalist Base case, Total NLTF)", ""]
    if impact.empty:
        lines.append("Impact frame empty; see failed checks above.")
    else:
        base = impact[
            impact["scenario_name"].astype(str).eq("current_basecase")
            & impact["series_id"].astype(str).eq("total_nltf_net_revenue")
        ]
        for engine in sorted(ENGINES):
            subset = base[base["engine"].eq(engine)]
            near = subset[(subset["june_year"] >= 2026) & (subset["june_year"] <= 2030)]["delta"].sum()
            far = subset[(subset["june_year"] >= 2031) & (subset["june_year"] <= 2050)]["delta"].sum()
            lines.append(
                f"- {engine}: FY2026-2030 sum {near:+.3f} $m; FY2031-2050 sum {far:+.3f} $m "
                f"(refreshed {bridge_id} bridge minus pinned {compare_id}-bridge baseline)"
            )
    lines += ["", "### Impact coverage", ""]
    if coverage:
        for engine, stats in sorted(coverage.items()):
            lines.append(
                f"- {engine}: {stats['matched_rows']} matched rows; "
                f"{stats['rows_only_in_pinned_baseline']} only in pinned baseline; "
                f"{stats['rows_only_in_refreshed_pack']} only in refreshed pack; "
                f"{stats['baseline_official_comparator_rows_excluded']} baseline official-comparator "
                "rows excluded (comparator addition is not bridge impact)"
            )
    else:
        lines.append("No coverage stats (baseline verification failed).")

    lines += ["", "## Composition-refresh candidate verdict", ""]
    shares = pd.read_csv(class_share_path)
    candidates = shares[shares["composition_refresh_candidate"].astype(bool)]
    delta_columns = [c for c in shares.columns if c.startswith("delta_vs_vfm_base_")]
    max_abs_delta = (
        pd.concat([pd.to_numeric(shares[c], errors="coerce").abs() for c in delta_columns], axis=1)
        .max()
        .max()
        if delta_columns
        else float("nan")
    )
    if candidates.empty:
        lines.append(
            f"NOT MATERIAL vs VFM Base: no {official_id} embedded Light-RUC class share "
            f"deviates from the exact-VFM Base share by more than 0.01 "
            f"(max |delta| {max_abs_delta:.6f})."
        )
    else:
        fys = sorted(int(fy) for fy in candidates["FY"])
        lines.append(
            f"MATERIAL vs VFM Base in {len(fys)} FYs ({fys[0]}-{fys[-1]}): max |share delta| "
            f"{max_abs_delta:.6f} exceeds 0.01. OPT-IN CANDIDATE ONLY: the Current model "
            "composition is unchanged pending separate approval."
        )

    lines += [
        "",
        "## Scope notes",
        "",
        "- The financial decomposition is a financial closure, not causal driver",
        "  attribution: official GDP/unemployment/price/judgment inputs are not",
        "  supplied (driver_availability_matrix.csv) and receive no fabricated dollars.",
        "- The policy-basis comparison keeps `policy_normalised` and `actual_default_ui`",
        "  separate; the official side is always the published vintage and",
        "  delayed-vs-delayed is never constructed as the standard comparison.",
        "",
        f"Overall: {'PASS' if not checks.failures else 'FAIL'} "
        f"({sum(1 for r in checks.rows if r['status'] == 'PASS')}/{len(checks.rows)} checks passed).",
        "",
    ]
    (out / "validation_report.md").write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def scan_artifacts_for_absolute_paths(out: Path, artifact_names: list[str], checks: Checks) -> None:
    offenders: list[str] = []
    for name in artifact_names:
        text = (out / name).read_text(encoding="utf-8", errors="replace").lower()
        for fragment in FORBIDDEN_PATH_FRAGMENTS:
            if fragment in text:
                offenders.append(f"{name}: contains {fragment!r}")
    checks.record(
        "artifacts_repo_relative_paths_only",
        not offenders,
        "; ".join(offenders) if offenders else "no absolute user paths in any artifact",
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--official-vintage", default=None, help="registered vintage id (default: registry default comparator)")
    parser.add_argument("--bridge-vintage", default=None, help="registered vintage id (default: registry default bridge vintage)")
    parser.add_argument("--compare-vintage", default="MBU26", help="prior vintage for cross-vintage deltas")
    parser.add_argument("--output-dir", default=None, help="artifact directory (default: artifacts/official_vintage_<vintage>)")
    args = parser.parse_args(argv)

    checks = Checks()
    try:
        official_id = str(args.official_vintage or default_comparator_vintage_id(ROOT))
        bridge_id = str(args.bridge_vintage or default_bridge_vintage_id(ROOT))
        compare_id = str(args.compare_vintage)
        out = ROOT / (args.output_dir or f"artifacts/official_vintage_{official_id.lower()}")
        out.mkdir(parents=True, exist_ok=True)

        official_entry = official_vintage_entry(official_id, ROOT)
        compare_entry = official_vintage_entry(compare_id, ROOT)
        official_pack = load_official_vintage(official_id, repo_root=ROOT)
        compare_pack = load_official_vintage(compare_id, repo_root=ROOT)
        if official_pack is None or compare_pack is None:
            raise OfficialVintageError("official/compare vintage pack not materialized")
        for name, frame in (
            ("official_annual", official_pack.official_annual),
            ("formula_audit", official_pack.formula_audit),
            ("row_reconciliation", official_pack.row_reconciliation),
            ("annual_spine", official_pack.annual_spine),
            ("compare_official_annual", compare_pack.official_annual),
        ):
            if frame is None or frame.empty:
                raise OfficialVintageError(f"empty pack frame: {name}")
        checks.record(
            "pack_frames_loaded",
            True,
            f"{official_id} and {compare_id} packs loaded and hash-validated by the governed loader",
        )

        # Runtime pack manifests must be on the requested bridge vintage.
        manifests = {engine: load_runtime_manifest(pack_dir) for engine, pack_dir in ENGINES.items()}
        for engine, manifest in manifests.items():
            block = manifest.get("official_vintages") or {}
            observed_bridge = str(block.get("bridge_assumption_vintage_id"))
            checks.record(
                f"runtime_pack_bridge_vintage_{engine}",
                observed_bridge == bridge_id,
                f"manifest bridge_assumption_vintage_id={observed_bridge}, requested={bridge_id}",
            )

        # Workbook artifacts (1-5, 7).
        workbook_path, workbook_hash, values_wb, formulas_wb = open_workbook(official_entry)
        try:
            formula_cells = scan_formula_cells(official_entry, formulas_wb)
            checks.record(
                "workbook_formula_count",
                True,
                f"{len(formula_cells)} Excel formulas found (static published values expected)",
            )
            build_workbook_inventory(
                official_entry, official_pack, workbook_path, workbook_hash, values_wb, formulas_wb, formula_cells, out
            )
            build_workbook_schema(official_entry, official_pack, values_wb, out)
            build_source_lineage(official_entry, official_pack, workbook_hash, out)
            build_source_workbook_manifest(official_entry, workbook_path, workbook_hash, out)
            build_formula_inventory(official_entry, formula_cells, out)
            build_official_annual_changes(official_entry, official_pack, values_wb, out, checks)
        finally:
            values_wb.close()
            formulas_wb.close()

        # 6: verbatim copy of the pack's official_annual.
        stems = official_entry.get("file_stems") or {}
        official_annual_name = f"{stems.get('official_annual', 'official_annual')}.csv"
        shutil.copyfile(official_pack.pack_dir / official_annual_name, out / "official_annual.csv")
        checks.record(
            "official_annual_verbatim_copy",
            sha256(out / "official_annual.csv")
            == sha256(official_pack.pack_dir / official_annual_name),
            f"byte-identical to {repo_rel(official_pack.pack_dir / official_annual_name)}",
        )

        # 8: classified reconciliation frames.
        row_classified = classify_reconciliation(
            official_id, official_pack.row_reconciliation, checks, "row_reconciliation"
        )
        write_csv(row_classified, out / "row_reconciliation.csv")
        formula_classified = classify_reconciliation(
            official_id, official_pack.formula_audit, checks, "formula_reconciliation"
        )
        write_csv(formula_classified, out / "formula_reconciliation.csv")

        official_values = values_map(official_pack.official_annual)
        compare_values = values_map(compare_pack.official_annual)
        fys = sorted({fy for _series, fy in official_values})

        # 9, 10.
        build_effective_rate_audit(
            official_id, compare_id, official_values, compare_values, fys, out, checks
        )
        build_class_share_audit(official_id, official_values, fys, out, checks)

        # 11, 12.
        _deltas, revisions = build_vintage_deltas(
            official_id,
            compare_id,
            official_pack.official_annual,
            compare_pack.official_annual,
            out,
            checks,
        )

        # 13.
        impact, coverage = build_bridge_impact(bridge_id, out, checks)

        # 14, 15.
        chart_by_engine = {engine: load_chart_rows(pack_dir) for engine, pack_dir in ENGINES.items()}
        line_by_engine = {
            engine: line_values(load_line_reconciliation(pack_dir), CURRENT_BASE_TRACE)
            for engine, pack_dir in ENGINES.items()
        }
        for engine, values in line_by_engine.items():
            if not values:
                raise OfficialVintageError(
                    f"{engine}: no '{CURRENT_BASE_TRACE}' rows in revenue_line_reconciliation"
                )
        build_current_vs_official(
            official_id, bridge_id, official_values, line_by_engine, out, checks
        )

        # 16.
        build_policy_basis(official_id, official_values, chart_by_engine, out, checks)

        # 17, 18, 19.
        build_driver_availability(official_id, out)
        build_source_horizon_audit(
            {compare_id: compare_entry, official_id: official_entry},
            manifests,
            chart_by_engine,
            out,
        )
        build_front_end_selector_audit(manifests, out, checks, official_id)

        artifact_names = [
            "workbook_inventory.json",
            "workbook_schema.csv",
            "source_lineage.csv",
            "source_workbook_manifest.json",
            "formula_inventory.csv",
            "official_annual.csv",
            "official_annual_changes.csv",
            "row_reconciliation.csv",
            "formula_reconciliation.csv",
            "effective_rate_audit.csv",
            "class_share_audit.csv",
            f"{official_id.lower()}_vs_{compare_id.lower()}_by_series_fy.csv",
            f"{official_id.lower()}_vs_{compare_id.lower()}_summary.csv",
            "current_bridge_vintage_impact.csv",
            f"current_vs_{official_id.lower()}_gap_by_stream_fy.csv",
            f"current_vs_{official_id.lower()}_financial_decomposition.csv",
            "policy_basis_comparison.csv",
            "driver_availability_matrix.csv",
            "source_horizon_audit.csv",
            "front_end_selector_audit.csv",
        ]
        scan_artifacts_for_absolute_paths(out, artifact_names, checks)

        # 20.
        build_validation_report(
            official_id,
            bridge_id,
            compare_id,
            checks,
            formula_classified,
            revisions,
            impact,
            coverage,
            out / "class_share_audit.csv",
            out,
        )
        artifact_names.append("validation_report.md")
    except OfficialVintageError as error:
        print(f"FAIL: {error}")
        return 1

    if checks.failures:
        print("FAIL")
        for failure in checks.failures:
            print(f"  - {failure}")
        return 1
    print(f"OFFICIAL_VINTAGE_RECONCILIATION_OK {len(artifact_names)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
