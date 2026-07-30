"""Governed quarterly actuals refresh for the canonical model-input history.

Ingests a model-input actuals workbook (schema-equal to
``data/model_input_history/*.parquet``: sheets "PED Inputs", "Light RUC
Inputs", "Heavy RUC Inputs"), appends the governed new-quarter rows to the
canonical history parquets through one central derivation path, and emits the
full audit-artifact set under ``--output-dir``.

Contract (see AGENTS/Agent Build Spec sheet of the source workbook):

- Excel is an immutable source snapshot for ingestion only; runtime code keeps
  reading the committed parquet artifacts.
- Only authoritative RAW input fields are extracted per stream; every log,
  lag, difference, dummy, trend, interaction, real-price conversion and
  per-capita field is regenerated centrally and checked against the workbook
  row for parity (``feature_parity.csv``).
- Accepted history rows (data_status "Historical official/source data") are
  never modified: the refresh fails closed on any conflicting value.
- The PED target is governed by ``--ped-mode``:
    exclude                  do not admit the provisional bridge as history;
                             the PED target keeps the 0.0 placeholder.
    provisional_replay_only  same parquet outcome as ``exclude`` (the target
                             column NEVER carries a provisional value), but
                             the bridge is written to a governed sidecar
                             (``ped_provisional_bridge.json``) that replay
                             code may use as a recursive-history seed.
    accepted                 reserved for an explicit governance decision;
                             requires ``--governance-approval``.
- Nothing period-specific is hard-coded: new rows are identified as workbook
  periods newer than the last accepted canonical row (or placeholder rows
  eligible for upgrade), so a later workbook with the same three main sheet
  schemas refreshes the next quarter without code edits.

Usage:
    python scripts/refresh_model_actuals.py \
      --workbook "references/NLTF_model_input_sheet_actuals_to_2026Q1_complete1.xlsx" \
      --expected-period 2026Q1 \
      --ped-mode provisional_replay_only \
      --output-dir artifacts/actuals_refresh_2026q1
"""
from __future__ import annotations

import argparse
from datetime import datetime, timezone
import hashlib
import json
import math
import platform
import subprocess
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

HISTORY_DIR_DEFAULT = ROOT / "data" / "model_input_history"

STREAMS = ["ped", "light_ruc", "heavy_ruc"]
STREAM_SHEETS = {
    "ped": "PED Inputs",
    "light_ruc": "Light RUC Inputs",
    "heavy_ruc": "Heavy RUC Inputs",
}
HISTORY_FILES = {
    "ped": "ped_inputs.parquet",
    "light_ruc": "light_ruc_inputs.parquet",
    "heavy_ruc": "heavy_ruc_inputs.parquet",
}
PED_MODES = ("exclude", "provisional_replay_only", "accepted")

# Authoritative raw/source fields per stream. Everything else on the sheet is
# derived centrally by ``derive_stream_row`` and used only for parity checks.
RAW_INPUT_FIELDS = {
    "ped": [
        "nominal_gdp_sa_nzd",
        "population",
        "gdp_deflator_sa",
        "petrol_price_nominal_cents_per_litre",
        "cpi_rebasing_factor",
        "light_petrol_vkt_total_km",
        "unemployment_percent",
    ],
    "light_ruc": [
        "target",
        "nominal_gdp_sa_nzd",
        "gdp_deflator_sa",
        "diesel_price_nominal_cents_per_litre",
        "cpi_rebasing_factor",
        "light_ruc_revenue_nzd",
    ],
    "heavy_ruc": [
        "target",
        "nominal_gdp_sa_nzd",
        "gdp_deflator_sa",
        "heavy_ruc_revenue_nzd",
    ],
}
# Optional vintage-flagged raw fields (present only when retrospectively
# observed). Stored with explicit vintage status in the lead audit.
OPTIONAL_RAW_FIELDS = {
    "heavy_ruc": ["lead_real_heavy_ruc_price_nzd_per_1000km"],
}
NOTE_FIELDS = ["data_status", "notes"]

RAW_FIELD_UNITS = {
    "target": {"ped": "km/person", "light_ruc": "km", "heavy_ruc": "km"},
    "nominal_gdp_sa_nzd": "NZD",
    "population": "persons",
    "gdp_deflator_sa": "index",
    "petrol_price_nominal_cents_per_litre": "cents/litre",
    "diesel_price_nominal_cents_per_litre": "cents/litre",
    "cpi_rebasing_factor": "factor",
    "light_petrol_vkt_total_km": "km",
    "unemployment_percent": "percent",
    "light_ruc_revenue_nzd": "NZD",
    "heavy_ruc_revenue_nzd": "NZD",
    "lead_real_heavy_ruc_price_nzd_per_1000km": "NZD/1,000 km (real)",
}

ACCEPTED_STATUS_PREFIX = "historical official"
PLACEHOLDER_STATUS_PREFIX = "partial official"
PROVISIONAL_STATUS_PREFIX = "provisional"

PED_PROVISIONAL_STATUS = "Provisional quarterly PED bridge; macro inputs are official Q1 actuals"
PED_PROVISIONAL_METHOD = "mbu26_residual_core_ped_share"

# Tolerances.
HISTORY_CROSS_CHECK_RTOL = 1e-12   # committed history vs workbook history region
FEATURE_PARITY_RTOL = 1e-9         # centrally derived vs workbook derived cells
IDENTITY_RTOL = 1e-9               # revenue / km x 1,000 style identities
NOOP_RTOL = 1e-12                  # idempotency comparison
UNIT_SANITY_MAX_RATIO = 10.0       # new raw value vs last accepted non-zero value

QUARTER_END_MONTH = {1: "Mar", 2: "Jun", 3: "Sep", 4: "Dec"}


class RefreshError(RuntimeError):
    """Fail-closed ingestion error."""


# ---------------------------------------------------------------------------
# Small utilities
# ---------------------------------------------------------------------------

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_period(period: str) -> tuple[int, int]:
    text = str(period).strip().upper()
    year_text, quarter_text = text.split("Q", 1)
    year, quarter = int(year_text), int(quarter_text)
    if quarter not in {1, 2, 3, 4}:
        raise RefreshError(f"Invalid quarter number in period {period!r}")
    return year, quarter


def quarter_sort_key(period: str) -> int:
    year, quarter = parse_period(period)
    return year * 4 + quarter


def next_period(period: str) -> str:
    key = quarter_sort_key(period)
    return f"{key // 4}Q{(key % 4) + 1}"


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def safe_log(value: float) -> float:
    return math.log(value) if value is not None and np.isfinite(value) and value > 0 else np.nan


def is_num(value: Any) -> bool:
    try:
        return np.isfinite(float(value))
    except (TypeError, ValueError):
        return False


def rel_diff(a: float, b: float) -> float:
    if not is_num(a) or not is_num(b):
        if (not is_num(a)) and (not is_num(b)):
            return 0.0
        return float("inf")
    a, b = float(a), float(b)
    denom = max(abs(a), abs(b), 1e-12)
    return abs(a - b) / denom


# ---------------------------------------------------------------------------
# Workbook loading and inventory
# ---------------------------------------------------------------------------

def load_workbook_bundle(workbook_path: Path) -> dict[str, Any]:
    """Load values + formulas + structural inventory from the workbook."""
    import openpyxl

    wb_values = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)

    sheets: list[dict[str, Any]] = []
    formulas: list[dict[str, Any]] = []
    for name in wb_formulas.sheetnames:
        ws = wb_formulas[name]
        ws_vals = wb_values[name]
        headers = [c.value for c in ws[1]]
        n_formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    n_formulas += 1
                    cached = ws_vals.cell(cell.row, cell.column).value
                    formulas.append(
                        {
                            "sheet": name,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "cached_value": cached if cached is not None else "",
                        }
                    )
        sheets.append(
            {
                "sheet": name,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "headers": [str(h) if h is not None else "" for h in headers],
                "formula_cells": n_formulas,
            }
        )
    defined_names = sorted(wb_formulas.defined_names.keys()) if hasattr(wb_formulas.defined_names, "keys") else []

    frames: dict[str, pd.DataFrame] = {}
    for stream, sheet in STREAM_SHEETS.items():
        if sheet not in wb_values.sheetnames:
            raise RefreshError(f"Required sheet {sheet!r} is missing from the workbook.")
        ws = wb_values[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else "" for h in rows[0]]
        frame = pd.DataFrame(rows[1:], columns=headers)
        frame = frame.dropna(how="all")
        frames[stream] = frame

    return {
        "sheets": sheets,
        "formulas": formulas,
        "defined_names": defined_names,
        "frames": frames,
        "sheetnames": list(wb_values.sheetnames),
    }


def auxiliary_sheets(sheetnames: list[str]) -> dict[str, str | None]:
    """Discover governance sheets by suffix pattern, without period hardcodes."""
    def find(suffix: str) -> str | None:
        for name in sheetnames:
            if name not in STREAM_SHEETS.values() and name.lower().endswith(suffix):
                return name
        return None

    return {
        "filled_rows": find("filled rows"),
        "source_map": find("source map"),
        "reconciliation": find("reconciliation"),
        "build_spec": next((n for n in sheetnames if n.lower() == "agent build spec"), None),
    }


# ---------------------------------------------------------------------------
# Structural validation
# ---------------------------------------------------------------------------

def validate_structure(frames: dict[str, pd.DataFrame], history: dict[str, pd.DataFrame]) -> None:
    for stream in STREAMS:
        frame = frames[stream]
        hist = history[stream]
        expected = [c for c in hist.columns]
        # PED sheet may omit the note columns; everything else must match the
        # canonical schema exactly (order included: header-driven, not cells).
        observed = [str(c) for c in frame.columns]
        expected_required = [c for c in expected if c not in NOTE_FIELDS or c in observed]
        if [c for c in observed if c not in expected]:
            unknown = [c for c in observed if c not in expected]
            raise RefreshError(f"{STREAM_SHEETS[stream]}: unexpected columns {unknown}.")
        missing = [c for c in expected_required if c not in observed]
        if missing:
            raise RefreshError(f"{STREAM_SHEETS[stream]}: missing required columns {missing}.")

        periods = frame["period"].astype(str).str.strip().str.upper().tolist()
        if len(set(periods)) != len(periods):
            dupes = sorted({p for p in periods if periods.count(p) > 1})
            raise RefreshError(f"{STREAM_SHEETS[stream]}: duplicated periods {dupes}.")
        keys = [quarter_sort_key(p) for p in periods]
        if keys != sorted(keys):
            raise RefreshError(f"{STREAM_SHEETS[stream]}: periods are not in ascending order.")
        gaps = [periods[i] for i in range(1, len(keys)) if keys[i] != keys[i - 1] + 1]
        if gaps:
            raise RefreshError(f"{STREAM_SHEETS[stream]}: quarter sequence is discontinuous at {gaps}.")
        if periods[0] != str(hist["period"].iloc[0]):
            raise RefreshError(
                f"{STREAM_SHEETS[stream]}: first period {periods[0]} does not match canonical history "
                f"start {hist['period'].iloc[0]}."
            )


def cross_check_history(
    frames: dict[str, pd.DataFrame],
    history: dict[str, pd.DataFrame],
    lead_audit_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """The workbook's history region must reproduce committed history.

    Any relative difference above ``HISTORY_CROSS_CHECK_RTOL`` fails closed,
    except retrospective completion of a previously-missing (zero placeholder)
    Heavy lead value, which is recorded in the lead vintage audit and NOT
    applied to canonical history.
    """
    report: list[dict[str, Any]] = []
    for stream in STREAMS:
        frame = frames[stream].copy()
        frame["period"] = frame["period"].astype(str).str.strip().str.upper()
        hist = history[stream]
        hist_periods = set(hist["period"].astype(str))
        accepted = accepted_period_mask(hist)
        shared = frame[frame["period"].isin(hist_periods)]
        hist_idx = hist.set_index("period")
        for _, row in shared.iterrows():
            period = row["period"]
            if not bool(accepted.get(period, False)):
                continue  # placeholder rows are replaced, not cross-checked
            hrow = hist_idx.loc[period]
            for column in hist.columns:
                if column in ("period",) or column not in row.index:
                    continue
                if column in NOTE_FIELDS:
                    continue
                wv, hv = row[column], hrow[column]
                if not is_num(wv) and not is_num(hv):
                    continue
                lead_completion = (
                    stream == "heavy_ruc"
                    and column in ("lead_real_heavy_ruc_price_nzd_per_1000km", "log_lead_real_heavy_ruc_price")
                    and is_num(hv)
                    and float(hv) == 0.0
                    and is_num(wv)
                    and float(wv) > 0.0
                )
                if lead_completion:
                    lead_audit_rows.append(
                        {
                            "period": period,
                            "column": column,
                            "canonical_value": float(hv),
                            "workbook_value": float(wv),
                            "vintage_status": "retrospective_history",
                            "applied_to_canonical_history": False,
                            "note": (
                                "Previously-missing lead placeholder completed retrospectively in the "
                                "workbook; canonical accepted history is kept byte-identical, and the "
                                "retrospective value is recorded here as refit/evaluation evidence."
                            ),
                        }
                    )
                    continue
                drift = rel_diff(wv, hv)
                if drift > HISTORY_CROSS_CHECK_RTOL:
                    raise RefreshError(
                        f"{STREAM_SHEETS[stream]}: accepted history value changed at {period}/{column}: "
                        f"canonical={hv!r} workbook={wv!r} (rel diff {drift:.3e}). "
                        "Accepted history must never be modified by an actuals refresh."
                    )
                report.append(
                    {
                        "stream": stream.upper(),
                        "period": period,
                        "column": column,
                        "max_rel_diff": drift,
                    }
                )
    # Compact: keep only per-stream/column maxima.
    if report:
        df = pd.DataFrame(report)
        df = df.groupby(["stream", "column"], as_index=False)["max_rel_diff"].max()
        return df.to_dict(orient="records")
    return []


def accepted_period_mask(hist: pd.DataFrame) -> dict[str, bool]:
    status = hist.get("data_status")
    if status is None:
        return {str(p): True for p in hist["period"]}
    out = {}
    for period, text in zip(hist["period"].astype(str), status.fillna("").astype(str)):
        out[period] = text.strip().lower().startswith(ACCEPTED_STATUS_PREFIX)
    return out


def last_accepted_period(hist: pd.DataFrame) -> str:
    mask = accepted_period_mask(hist)
    accepted = [p for p, ok in mask.items() if ok]
    if not accepted:
        raise RefreshError("Canonical history has no accepted rows.")
    return max(accepted, key=quarter_sort_key)


# ---------------------------------------------------------------------------
# Central derivation of engineered rows (the one canonical code path)
# ---------------------------------------------------------------------------

def gdp_rebase_base_deflator(hist: pd.DataFrame) -> float:
    """The GDP rebasing base is the period whose factor is exactly 1."""
    factors = pd.to_numeric(hist["gdp_rebasing_factor"], errors="coerce")
    deflators = pd.to_numeric(hist["gdp_deflator_sa"], errors="coerce")
    candidates = [
        (str(p), float(d))
        for p, f, d in zip(hist["period"], factors, deflators)
        if is_num(f) and abs(float(f) - 1.0) < 1e-9 and is_num(d) and float(d) > 0
    ]
    if not candidates:
        raise RefreshError("Cannot locate the GDP rebasing base period (no gdp_rebasing_factor == 1 row).")
    # Latest such period is the governed base.
    period, deflator = max(candidates, key=lambda item: quarter_sort_key(item[0]))
    return deflator


def prior_row(hist: pd.DataFrame, new_rows: list[dict[str, Any]], period: str, lag: int) -> dict[str, Any] | None:
    """Row ``lag`` quarters before ``period`` from history plus pending new rows."""
    target_key = quarter_sort_key(period) - lag
    for row in new_rows:
        if quarter_sort_key(row["period"]) == target_key:
            return row
    matches = hist[hist["period"].map(quarter_sort_key) == target_key]
    if matches.empty:
        return None
    return matches.iloc[0].to_dict()


def derive_stream_row(
    stream: str,
    period: str,
    raw: dict[str, float],
    notes: dict[str, str],
    hist: pd.DataFrame,
    pending: dict[str, list[dict[str, Any]]],
    base_deflator: float,
    cross: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Compute every canonical column for one appended row from raw inputs.

    ``cross`` carries the already-derived same-period rows of other streams
    (heavy joins light prices / ped unemployment by canonical period).
    """
    year, quarter = parse_period(period)
    hist_last = hist.iloc[-1]
    period_index = int(hist_last["period_index"]) + (quarter_sort_key(period) - quarter_sort_key(str(hist_last["period"])))
    row: dict[str, Any] = {
        "period": period,
        "period_label": f"{QUARTER_END_MONTH[quarter]} {year}",
        "year": year,
        "quarter": quarter,
        "period_index": period_index,
        "trend_index": period_index,
        "log_trend": safe_log(period_index),
        "data_status": notes.get("data_status", ""),
        "notes": notes.get("notes", ""),
    }

    deflator = raw["gdp_deflator_sa"]
    if not is_num(deflator) or deflator <= 0:
        raise RefreshError(f"{stream}: gdp_deflator_sa at {period} is not a positive number.")
    row["gdp_deflator_sa"] = deflator
    row["nominal_gdp_sa_nzd"] = raw["nominal_gdp_sa_nzd"]
    row["gdp_rebasing_factor"] = base_deflator / deflator
    if "cpi_rebasing_factor" in raw:
        row["cpi_rebasing_factor"] = raw["cpi_rebasing_factor"]

    lag1 = prior_row(hist, pending.get(stream, []), period, 1)
    lag4 = prior_row(hist, pending.get(stream, []), period, 4)
    if lag1 is None or lag4 is None:
        raise RefreshError(f"{stream}: history is missing the lag-1 or lag-4 row for {period}.")

    def lag_target(lagrow: dict[str, Any]) -> float:
        return float(pd.to_numeric(pd.Series([lagrow.get("target")]), errors="coerce").iloc[0] or 0.0)

    row["target_lag_1"] = lag_target(lag1)
    row["target_lag_4"] = lag_target(lag4)
    row["log_target_lag_1"] = safe_log(row["target_lag_1"])
    row["log_target_lag_4"] = safe_log(row["target_lag_4"])
    if is_num(row["log_target_lag_1"]) and is_num(row["log_target_lag_4"]):
        row["diff_log_target_lag_1_lag_4"] = row["log_target_lag_1"] - row["log_target_lag_4"]
    else:
        row["diff_log_target_lag_1_lag_4"] = np.nan

    row["q2_dummy"] = 1 if quarter == 2 else 0
    row["q3_dummy"] = 1 if quarter == 3 else 0
    row["q4_dummy"] = 1 if quarter == 4 else 0
    row["post_2020_dummy"] = 1 if year >= 2021 else 0

    if stream == "ped":
        population = raw["population"]
        if not is_num(population) or population <= 0:
            raise RefreshError(f"ped: population at {period} is not a positive number.")
        row["population"] = population
        row["nominal_gdp_per_capita_nzd"] = raw["nominal_gdp_sa_nzd"] / population
        row["real_gdp_per_capita_nzd"] = row["nominal_gdp_per_capita_nzd"] * row["gdp_rebasing_factor"]
        row["log_real_gdp_per_capita"] = safe_log(row["real_gdp_per_capita_nzd"])
        row["petrol_price_nominal_cents_per_litre"] = raw["petrol_price_nominal_cents_per_litre"]
        row["real_petrol_price_cents_per_litre"] = (
            raw["petrol_price_nominal_cents_per_litre"] * raw["cpi_rebasing_factor"]
        )
        row["log_real_petrol_price"] = safe_log(row["real_petrol_price_cents_per_litre"])
        vkt = raw["light_petrol_vkt_total_km"]
        row["light_petrol_vkt_total_km"] = vkt
        row["log_total_vkt"] = safe_log(vkt) if vkt and vkt > 0 else 0.0
        row["target"] = (vkt / population) if vkt and vkt > 0 else 0.0
        row["log_target"] = safe_log(row["target"])
        row["unemployment_percent"] = raw["unemployment_percent"]
        row["unemployment_rate"] = raw["unemployment_percent"] / 100.0
        row["log_unemployment_rate"] = safe_log(row["unemployment_rate"])
        row["post_2011_dummy"] = 1 if year >= 2012 else 0
        row["dummy_2020"] = 1 if year == 2020 else 0
        row["post_2011_x_log_trend"] = row["post_2011_dummy"] * row["log_trend"]
        return row

    # RUC streams share real GDP.
    row["real_gdp_sa_nzd"] = raw["nominal_gdp_sa_nzd"] * row["gdp_rebasing_factor"]
    row["log_real_gdp"] = safe_log(row["real_gdp_sa_nzd"])

    if stream == "light_ruc":
        target = raw["target"]
        revenue = raw["light_ruc_revenue_nzd"]
        row["target"] = target
        row["log_target"] = safe_log(target)
        row["light_ruc_revenue_nzd"] = revenue
        row["diesel_price_nominal_cents_per_litre"] = raw["diesel_price_nominal_cents_per_litre"]
        row["real_diesel_price_cents_per_litre"] = (
            raw["diesel_price_nominal_cents_per_litre"] * raw["cpi_rebasing_factor"]
        )
        row["log_real_diesel_price"] = safe_log(row["real_diesel_price_cents_per_litre"])
        rate = (revenue / target * 1000.0) if target and target > 0 else 0.0
        row["light_ruc_price_nominal_nzd_per_1000km"] = rate
        row["real_light_ruc_price_nzd_per_1000km"] = rate * raw["cpi_rebasing_factor"]
        row["log_real_light_ruc_price"] = safe_log(row["real_light_ruc_price_nzd_per_1000km"])
        lag_rate = lag1.get("real_light_ruc_price_nzd_per_1000km", 0.0)
        row["lagged_real_light_ruc_price_nzd_per_1000km"] = float(lag_rate or 0.0)
        row["log_lagged_real_light_ruc_price"] = safe_log(row["lagged_real_light_ruc_price_nzd_per_1000km"])
        row["diesel_x_ruc_price"] = _nan_mul(row["log_real_diesel_price"], row["log_real_light_ruc_price"])
        row["gdp_x_post2020"] = _nan_mul(row["log_real_gdp"], row["post_2020_dummy"])
        row["ruc_x_post2020"] = _nan_mul(row["log_real_light_ruc_price"], row["post_2020_dummy"])
        row["diesel_x_post2020"] = _nan_mul(row["log_real_diesel_price"], row["post_2020_dummy"])
        row["time_trend"] = period_index
        for base in ("log_real_diesel_price", "log_real_light_ruc_price", "log_real_gdp"):
            lag1v = float(pd.to_numeric(pd.Series([lag1.get(base)]), errors="coerce").iloc[0] or np.nan)
            lag4v = float(pd.to_numeric(pd.Series([lag4.get(base)]), errors="coerce").iloc[0] or np.nan)
            row[f"{base}_lag1"] = lag1v
            row[f"{base}_lag4"] = lag4v
            row[f"{base}_diff1"] = row[base] - lag1v if is_num(row[base]) and is_num(lag1v) else np.nan
        return row

    # heavy_ruc
    target = raw["target"]
    revenue = raw["heavy_ruc_revenue_nzd"]
    row["target"] = target
    row["log_target"] = safe_log(target)
    row["heavy_ruc_revenue_nzd"] = revenue
    light_row = cross.get("light_ruc")
    ped_row = cross.get("ped")
    if light_row is None or ped_row is None:
        raise RefreshError("heavy_ruc row derivation requires same-period light_ruc and ped rows.")
    cpi_factor = float(light_row["cpi_rebasing_factor"])
    rate = (revenue / target * 1000.0) if target and target > 0 else 0.0
    row["heavy_ruc_price_nominal_nzd_per_1000km"] = rate
    row["real_heavy_ruc_price_nzd_per_1000km"] = rate * cpi_factor
    row["log_real_heavy_ruc_price"] = safe_log(row["real_heavy_ruc_price_nzd_per_1000km"])
    lead = raw.get("lead_real_heavy_ruc_price_nzd_per_1000km")
    row["lead_real_heavy_ruc_price_nzd_per_1000km"] = float(lead) if is_num(lead) and float(lead) > 0 else 0.0
    row["log_lead_real_heavy_ruc_price"] = (
        safe_log(row["lead_real_heavy_ruc_price_nzd_per_1000km"])
        if row["lead_real_heavy_ruc_price_nzd_per_1000km"] > 0
        else 0.0
    )
    row["real_diesel_price_cents_per_litre"] = float(light_row["real_diesel_price_cents_per_litre"])
    row["log_real_diesel_price"] = float(light_row["log_real_diesel_price"])
    row["real_light_ruc_price_nzd_per_1000km"] = float(light_row["real_light_ruc_price_nzd_per_1000km"])
    row["log_real_light_ruc_price"] = float(light_row["log_real_light_ruc_price"])
    row["lagged_real_light_ruc_price_nzd_per_1000km"] = float(light_row["lagged_real_light_ruc_price_nzd_per_1000km"])
    row["log_lagged_real_light_ruc_price"] = float(light_row["log_lagged_real_light_ruc_price"])
    row["unemployment_rate"] = float(ped_row["unemployment_rate"])
    row["log_unemployment_rate"] = float(ped_row["log_unemployment_rate"])
    return row


def _nan_mul(a: float, b: float) -> float:
    if is_num(a) and is_num(b):
        return float(a) * float(b)
    return np.nan


# ---------------------------------------------------------------------------
# Extraction, classification, validation of new rows
# ---------------------------------------------------------------------------

def extract_raw(stream: str, wb_row: pd.Series) -> dict[str, float]:
    raw: dict[str, float] = {}
    for field in RAW_INPUT_FIELDS[stream]:
        value = wb_row.get(field)
        if not is_num(value):
            raise RefreshError(
                f"{STREAM_SHEETS[stream]}: required raw field {field!r} at {wb_row.get('period')} is "
                f"non-numeric or non-finite ({value!r})."
            )
        raw[field] = float(value)
    for field in OPTIONAL_RAW_FIELDS.get(stream, []):
        value = wb_row.get(field)
        if is_num(value):
            raw[field] = float(value)
    return raw


def unit_sanity_check(stream: str, raw: dict[str, float], hist: pd.DataFrame) -> None:
    """New raw values must stay within an order of magnitude of the last
    accepted non-zero observation for the same field (unit-invariance guard)."""
    accepted = accepted_period_mask(hist)
    accepted_hist = hist[hist["period"].astype(str).map(lambda p: accepted.get(p, False))]
    for field, value in raw.items():
        if field not in accepted_hist.columns:
            continue
        series = pd.to_numeric(accepted_hist[field], errors="coerce")
        series = series[series > 0]
        if series.empty:
            continue
        last = float(series.iloc[-1])
        if value <= 0:
            if field in ("target", "light_petrol_vkt_total_km", "light_ruc_revenue_nzd", "heavy_ruc_revenue_nzd"):
                raise RefreshError(
                    f"{STREAM_SHEETS[stream]}: raw field {field!r} must be positive, got {value!r}."
                )
            continue
        ratio = max(value / last, last / value)
        if ratio > UNIT_SANITY_MAX_RATIO:
            unit = RAW_FIELD_UNITS.get(field)
            if isinstance(unit, dict):
                unit = unit.get(stream, "")
            raise RefreshError(
                f"{STREAM_SHEETS[stream]}: raw field {field!r} at value {value!r} is more than "
                f"{UNIT_SANITY_MAX_RATIO:.0f}x away from the last accepted observation {last!r} "
                f"(declared unit: {unit}). Unit-invalid or wrong-scale input; failing closed."
            )


def validate_identities(stream: str, row: dict[str, Any], wb_row: pd.Series) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []

    def check(name: str, lhs: float, rhs: float, tol: float = IDENTITY_RTOL) -> None:
        drift = rel_diff(lhs, rhs)
        ok = drift <= tol
        checks.append(
            {
                "stream": stream.upper(),
                "period": row["period"],
                "identity": name,
                "lhs": lhs,
                "rhs": rhs,
                "rel_diff": drift,
                "tolerance": tol,
                "status": "pass" if ok else "fail",
            }
        )
        if not ok:
            raise RefreshError(f"{stream}: identity {name} failed at {row['period']}: {lhs!r} vs {rhs!r}.")

    if stream == "light_ruc" and row["target"] > 0:
        check(
            "light_rate = revenue / km * 1000",
            row["light_ruc_price_nominal_nzd_per_1000km"],
            row["light_ruc_revenue_nzd"] / row["target"] * 1000.0,
        )
    if stream == "heavy_ruc" and row["target"] > 0:
        check(
            "heavy_rate = revenue / km * 1000",
            row["heavy_ruc_price_nominal_nzd_per_1000km"],
            row["heavy_ruc_revenue_nzd"] / row["target"] * 1000.0,
        )
    if stream in ("light_ruc", "heavy_ruc"):
        check(
            "real GDP = nominal GDP * rebasing factor",
            row["real_gdp_sa_nzd"],
            row["nominal_gdp_sa_nzd"] * row["gdp_rebasing_factor"],
        )
    if stream == "ped":
        if row["light_petrol_vkt_total_km"] > 0:
            check(
                "PED target = VKT / population",
                row["target"],
                row["light_petrol_vkt_total_km"] / row["population"],
            )
        check(
            "real GDP pc = nominal GDP / population * rebasing factor",
            row["real_gdp_per_capita_nzd"],
            row["nominal_gdp_sa_nzd"] / row["population"] * row["gdp_rebasing_factor"],
        )
    return checks


def feature_parity_rows(
    stream: str,
    derived: dict[str, Any],
    wb_row: pd.Series,
    ped_mode_suppressed: set[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for column, value in derived.items():
        if column in ("period", "period_label") or column in NOTE_FIELDS:
            continue
        wb_value = wb_row.get(column)
        if column in ped_mode_suppressed:
            rows.append(
                {
                    "stream": stream.upper(),
                    "period": derived["period"],
                    "column": column,
                    "derived_value": value,
                    "workbook_value": wb_value if wb_value is not None else "",
                    "rel_diff": "",
                    "status": "suppressed_by_ped_mode",
                }
            )
            continue
        drift = rel_diff(value, wb_value)
        status = "pass" if drift <= FEATURE_PARITY_RTOL else "fail"
        rows.append(
            {
                "stream": stream.upper(),
                "period": derived["period"],
                "column": column,
                "derived_value": value,
                "workbook_value": wb_value if wb_value is not None else "",
                "rel_diff": drift if np.isfinite(drift) else "",
                "status": status,
            }
        )
        if status == "fail":
            raise RefreshError(
                f"{stream}: feature parity failed at {derived['period']}/{column}: "
                f"derived={value!r} workbook={wb_value!r} (rel diff {drift:.3e})."
            )
    return rows


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------

def replace_or_append(hist: pd.DataFrame, row: dict[str, Any]) -> tuple[pd.DataFrame, str]:
    period = row["period"]
    columns = list(hist.columns)
    full_row = {c: row.get(c, np.nan) for c in columns}
    replaced = period in set(hist["period"].astype(str))
    kept = hist[hist["period"].astype(str) != period]
    out = pd.concat([kept, pd.DataFrame([full_row], columns=columns)], ignore_index=True)
    out = out.sort_values("period", key=lambda s: s.map(quarter_sort_key), kind="stable").reset_index(drop=True)
    return out, ("replaced_placeholder" if replaced else "appended")


def rows_equivalent(existing: pd.Series, candidate: dict[str, Any], columns: list[str]) -> bool:
    for column in columns:
        ev, cv = existing.get(column), candidate.get(column)
        if column in NOTE_FIELDS or column in ("period", "period_label"):
            if str(ev if ev is not None else "") != str(cv if cv is not None else ""):
                return False
            continue
        if not is_num(ev) and not is_num(cv):
            continue
        if rel_diff(ev, cv) > NOOP_RTOL:
            return False
    return True


# ---------------------------------------------------------------------------
# Main refresh
# ---------------------------------------------------------------------------

def run_refresh(args: argparse.Namespace) -> dict[str, Any]:
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise RefreshError(f"Workbook not found: {workbook_path}")
    history_dir = Path(args.history_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook_sha = sha256_file(workbook_path)
    if args.expected_sha256 and workbook_sha.lower() != args.expected_sha256.lower():
        raise RefreshError(
            f"Workbook SHA-256 mismatch: expected {args.expected_sha256}, got {workbook_sha}."
        )

    if args.ped_mode == "accepted" and not args.governance_approval:
        raise RefreshError(
            "--ped-mode accepted is reserved for an explicit governance decision; "
            "pass --governance-approval '<owner>: <decision reference>' to activate it."
        )

    history: dict[str, pd.DataFrame] = {
        stream: pd.read_parquet(history_dir / HISTORY_FILES[stream]) for stream in STREAMS
    }
    history_before_sha = {stream: sha256_file(history_dir / HISTORY_FILES[stream]) for stream in STREAMS}

    bundle = load_workbook_bundle(workbook_path)
    frames = bundle["frames"]
    aux = auxiliary_sheets(bundle["sheetnames"])

    validate_structure(frames, history)
    lead_audit_rows: list[dict[str, Any]] = []
    history_cross_check = cross_check_history(frames, history, lead_audit_rows)

    # Identify new/updatable periods per stream: newer than the last accepted
    # row (this includes existing non-accepted placeholder/provisional rows).
    # Streams may legitimately disagree - e.g. an accepted Light/Heavy quarter
    # coexists with a still-provisional PED quarter - so the apply loop works
    # over the union and serves cross-stream joins from canonical history when
    # a stream's same-period row is already accepted.
    new_periods: dict[str, list[str]] = {}
    for stream in STREAMS:
        hist = history[stream]
        last_ok = last_accepted_period(hist)
        frame = frames[stream]
        periods = frame["period"].astype(str).str.strip().str.upper()
        fresh = [p for p in periods if quarter_sort_key(p) > quarter_sort_key(last_ok)]
        # Continuity of the append itself.
        cursor = last_ok
        for p in fresh:
            if p != next_period(cursor):
                raise RefreshError(
                    f"{STREAM_SHEETS[stream]}: new periods are not continuous after {cursor}: found {p}."
                )
            cursor = p
        new_periods[stream] = fresh
    periods_to_apply = sorted(
        {p for fresh in new_periods.values() for p in fresh}, key=quarter_sort_key
    )
    if not periods_to_apply:
        raise RefreshError("Workbook contains no rows newer than the accepted canonical history.")
    if args.expected_period and args.expected_period.upper() not in periods_to_apply:
        raise RefreshError(
            f"Expected period {args.expected_period} is not among the new workbook periods "
            f"{periods_to_apply}."
        )

    # Derive rows centrally, per period, ped -> light -> heavy (cross joins).
    pending: dict[str, list[dict[str, Any]]] = {stream: [] for stream in STREAMS}
    parity_rows: list[dict[str, Any]] = []
    identity_rows: list[dict[str, Any]] = []
    raw_records: list[dict[str, Any]] = []
    lineage_rows: list[dict[str, Any]] = []
    observations: list[dict[str, Any]] = []
    ped_sidecar_entries: list[dict[str, Any]] = []

    base_deflators = {stream: gdp_rebase_base_deflator(history[stream]) for stream in STREAMS}

    for period in periods_to_apply:
        cross: dict[str, dict[str, Any]] = {}
        for stream in ("ped", "light_ruc", "heavy_ruc"):
            if period not in new_periods[stream]:
                # This stream's row is already accepted canonical history; use
                # it as the cross-join source and leave it untouched.
                hist = history[stream]
                existing = hist[hist["period"].astype(str) == period]
                if not existing.empty:
                    cross[stream] = existing.iloc[0].to_dict()
                continue
            frame = frames[stream]
            wb_row = frame[frame["period"].astype(str).str.strip().str.upper() == period].iloc[0]
            raw = extract_raw(stream, wb_row)
            unit_sanity_check(stream, raw, history[stream])

            notes = {
                "data_status": str(wb_row.get("data_status") or "").strip(),
                "notes": str(wb_row.get("notes") or "").strip(),
            }
            ped_suppressed: set[str] = set()
            classification: dict[str, Any]
            if stream == "ped":
                bridge_target = raw["light_petrol_vkt_total_km"] / raw["population"]
                if args.ped_mode == "accepted":
                    classification = {
                        "observation_status": "accepted_exact_actual",
                        "eligible_for_replay": True,
                        "eligible_for_refit": True,
                        "display_as_observed_actual": True,
                    }
                    if not notes["data_status"]:
                        notes["data_status"] = "Historical official/source data"
                else:
                    classification = {
                        "observation_status": "provisional_annual_bridge",
                        "target_method": PED_PROVISIONAL_METHOD,
                        "eligible_for_replay": "conditional",
                        "eligible_for_refit": False,
                        "display_as_observed_actual": False,
                    }
                    notes["data_status"] = PED_PROVISIONAL_STATUS
                    if not notes["notes"]:
                        notes["notes"] = (
                            "PED target is provisionally bridged from the MBU26 annual spine and Core "
                            "Gross PED revenue split; macro inputs are official quarterly actuals. The "
                            "provisional target is stored only in the governed sidecar, never in the "
                            "fitting target column."
                        )
                    ped_suppressed = {"target", "log_target", "light_petrol_vkt_total_km", "log_total_vkt"}
                    ped_sidecar_entries.append(
                        {
                            "period": period,
                            "provisional_target_km_per_person": bridge_target,
                            "provisional_light_petrol_vkt_total_km": raw["light_petrol_vkt_total_km"],
                            "population": raw["population"],
                            "target_method": PED_PROVISIONAL_METHOD,
                            "observation_status": "provisional_annual_bridge",
                            "eligible_for_replay": "conditional",
                            "eligible_for_refit": False,
                            "display_as_observed_actual": False,
                            "ped_mode": args.ped_mode,
                        }
                    )
            else:
                status_text = notes["data_status"].lower()
                if not status_text.startswith(ACCEPTED_STATUS_PREFIX):
                    raise RefreshError(
                        f"{STREAM_SHEETS[stream]}: new row {period} carries data_status "
                        f"{notes['data_status']!r}; an exact actual append requires an accepted status."
                    )
                classification = {
                    "observation_status": "accepted_exact_actual",
                    "eligible_for_replay": True,
                    "eligible_for_refit": True,
                    "display_as_observed_actual": True,
                }

            derived = derive_stream_row(
                stream, period, raw, notes, history[stream], pending, base_deflators[stream], cross
            )
            # Keep only canonical-schema columns (helper fields like the shared
            # post-2020 dummy are not part of every stream's schema).
            derived = {k: v for k, v in derived.items() if k in set(history[stream].columns)}
            if stream == "ped" and args.ped_mode != "accepted":
                # The provisional bridge never enters the fitting columns.
                derived["target"] = 0.0
                derived["log_target"] = np.nan
                derived["light_petrol_vkt_total_km"] = 0.0
                derived["log_total_vkt"] = 0.0

            identity_rows.extend(validate_identities(stream, derived, wb_row))
            parity_rows.extend(feature_parity_rows(stream, derived, wb_row, ped_suppressed))

            if stream == "heavy_ruc":
                lead_value = derived["lead_real_heavy_ruc_price_nzd_per_1000km"]
                lead_audit_rows.append(
                    {
                        "period": period,
                        "column": "lead_real_heavy_ruc_price_nzd_per_1000km",
                        "canonical_value": lead_value,
                        "workbook_value": raw.get("lead_real_heavy_ruc_price_nzd_per_1000km", 0.0),
                        "vintage_status": "retrospective_history" if lead_value > 0 else "real_time_unknown",
                        "applied_to_canonical_history": True,
                        "note": (
                            "Lead price at the appended quarter uses the subsequently observed "
                            "next-quarter effective rate (retrospective vintage); it was NOT known at "
                            "the forecast origin. Production vNext Heavy uses no lead features, so the "
                            "production replay is identical under retrospective and real-time vintages."
                            if lead_value > 0
                            else "No observed next-quarter rate was supplied; placeholder retained."
                        ),
                    }
                )

            for field, value in raw.items():
                unit = RAW_FIELD_UNITS.get(field, "")
                if isinstance(unit, dict):
                    unit = unit.get(stream, "")
                raw_records.append(
                    {
                        "stream": stream.upper(),
                        "period": period,
                        "field": field,
                        "value": value,
                        "unit": unit,
                        "source_sheet": STREAM_SHEETS[stream],
                        "workbook_sha256": workbook_sha,
                    }
                )
            observations.append(
                {
                    "stream": stream.upper(),
                    "period": period,
                    **classification,
                    "data_status": notes["data_status"],
                }
            )
            cross[stream] = derived
            pending[stream].append(derived)

    # Idempotency / conflict detection against the existing rows.
    applied_state: dict[str, str] = {}
    changed = False
    for stream in STREAMS:
        hist = history[stream]
        accepted = accepted_period_mask(hist)
        columns = list(hist.columns)
        for derived in pending[stream]:
            period = derived["period"]
            existing = hist[hist["period"].astype(str) == period]
            if existing.empty:
                applied_state[f"{stream}:{period}"] = "append"
                changed = True
                continue
            row = existing.iloc[0]
            if rows_equivalent(row, derived, columns):
                applied_state[f"{stream}:{period}"] = "no_op_identical"
                continue
            if accepted.get(period, False):
                raise RefreshError(
                    f"{stream}: refusing to modify accepted history row {period}; "
                    "conflicting values in the workbook (see history_before_after.csv)."
                )
            status = str(row.get("data_status") or "").strip().lower()
            upgradable = status.startswith(PLACEHOLDER_STATUS_PREFIX) or (
                stream == "ped"
                and status.startswith(PROVISIONAL_STATUS_PREFIX)
                and args.ped_mode == "accepted"
            )
            if stream == "ped" and status.startswith(PROVISIONAL_STATUS_PREFIX) and args.ped_mode != "accepted":
                raise RefreshError(
                    f"ped: existing provisional row {period} differs from the workbook values; "
                    "replacing a provisional PED row requires --ped-mode accepted with governance "
                    "approval (see history_before_after.csv for the diff)."
                )
            if not upgradable:
                raise RefreshError(
                    f"{stream}: existing row {period} (status {row.get('data_status')!r}) differs from "
                    "the workbook and is not an upgradable placeholder; failing closed."
                )
            applied_state[f"{stream}:{period}"] = "replace_placeholder"
            changed = True

    # Build before/after evidence and apply.
    before_after_rows: list[dict[str, Any]] = []
    updated_history: dict[str, pd.DataFrame] = {}
    for stream in STREAMS:
        hist = history[stream]
        out = hist
        for derived in pending[stream]:
            period = derived["period"]
            key = f"{stream}:{period}"
            existing = hist[hist["period"].astype(str) == period]
            old_row = existing.iloc[0] if not existing.empty else None
            for column in hist.columns:
                old_value = old_row.get(column) if old_row is not None else ""
                new_value = derived.get(column, "")
                before_after_rows.append(
                    {
                        "stream": stream.upper(),
                        "period": period,
                        "column": column,
                        "before": old_value if old_value is not None else "",
                        "after": new_value if new_value is not None else "",
                        "action": applied_state.get(key, ""),
                    }
                )
            if applied_state.get(key) != "no_op_identical":
                out, _how = replace_or_append(out, derived)
        # Preserve dtypes of integer-like identity columns.
        for column in ("year", "quarter", "period_index", "trend_index"):
            if column in out.columns and pd.api.types.is_integer_dtype(hist[column].dtype):
                out[column] = pd.to_numeric(out[column]).astype(hist[column].dtype)
        updated_history[stream] = out

    no_op = not changed
    apply_requested = not args.check_only

    history_after_sha: dict[str, str] = dict(history_before_sha)
    if apply_requested and changed:
        for stream in STREAMS:
            path = history_dir / HISTORY_FILES[stream]
            updated_history[stream].to_parquet(path, index=False)
            history_after_sha[stream] = sha256_file(path)
        write_history_manifest(history_dir, workbook_path, workbook_sha, periods_to_apply, args.ped_mode)

    # PED sidecar (written under both exclude and provisional_replay_only so
    # the governed bridge value and its status are always inspectable).
    sidecar_path = history_dir / "ped_provisional_bridge.json"
    if apply_requested and ped_sidecar_entries:
        sidecar = {
            "created_at": utc_now(),
            "source_workbook": workbook_path.name,
            "source_workbook_sha256": workbook_sha,
            "ped_mode": args.ped_mode,
            "entries": ped_sidecar_entries,
            "governance": {
                "never_fit_on_provisional": True,
                "display_as_observed_actual": False,
                "replay_seed_allowed_modes": ["provisional_replay_only"],
            },
        }
        sidecar_path.write_text(json.dumps(sidecar, indent=2) + "\n", encoding="utf-8")

    result = {
        "workbook": str(workbook_path),
        "workbook_sha256": workbook_sha,
        "workbook_size_bytes": workbook_path.stat().st_size,
        "expected_period": args.expected_period,
        "ped_mode": args.ped_mode,
        "periods_applied": periods_to_apply,
        "applied_state": applied_state,
        "no_op": no_op,
        "check_only": bool(args.check_only),
        "history_before_sha256": history_before_sha,
        "history_after_sha256": history_after_sha,
        "auxiliary_sheets": aux,
    }

    write_artifacts(
        output_dir=output_dir,
        args=args,
        bundle=bundle,
        aux=aux,
        workbook_path=workbook_path,
        workbook_sha=workbook_sha,
        history_cross_check=history_cross_check,
        parity_rows=parity_rows,
        identity_rows=identity_rows,
        raw_records=raw_records,
        before_after_rows=before_after_rows,
        lead_audit_rows=lead_audit_rows,
        observations=observations,
        updated_history=updated_history,
        pending=pending,
        result=result,
    )
    return result


def write_history_manifest(
    history_dir: Path,
    workbook_path: Path,
    workbook_sha: str,
    periods: list[str],
    ped_mode: str,
) -> None:
    manifest_path = history_dir / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8")) if manifest_path.exists() else {}
    rows = []
    for stream in STREAMS:
        path = history_dir / HISTORY_FILES[stream]
        frame = pd.read_parquet(path)
        rows.append(
            {
                "stream": stream.upper(),
                "source_sheet": STREAM_SHEETS[stream],
                "repo_relative_path": path.relative_to(ROOT).as_posix(),
                "rows": int(len(frame)),
                "columns": list(frame.columns),
                "first_period": str(frame["period"].iloc[0]),
                "last_period": str(frame["period"].iloc[-1]),
                "positive_target_rows": int(pd.to_numeric(frame["target"], errors="coerce").gt(0).sum()),
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
            }
        )
    manifest["artifacts"] = rows
    refresh_history = manifest.get("refresh_history", [])
    refresh_history.append(
        {
            "refreshed_at": utc_now(),
            "workbook": workbook_path.name,
            "workbook_sha256": workbook_sha,
            "periods": periods,
            "ped_mode": ped_mode,
            "tool": "scripts/refresh_model_actuals.py",
        }
    )
    manifest["refresh_history"] = refresh_history
    manifest["notes"] = (
        str(manifest.get("notes", ""))
        + f" Refreshed through {periods[-1]} from {workbook_path.name} (sha256 {workbook_sha[:12]}...)."
    ).strip()
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def write_artifacts(**kw: Any) -> None:
    output_dir: Path = kw["output_dir"]
    args = kw["args"]
    bundle = kw["bundle"]
    aux = kw["aux"]
    workbook_path: Path = kw["workbook_path"]
    workbook_sha: str = kw["workbook_sha"]
    result = kw["result"]

    inventory = {
        "created_at": utc_now(),
        "workbook": workbook_path.name,
        "workbook_sha256": workbook_sha,
        "workbook_size_bytes": workbook_path.stat().st_size,
        "sheet_count": len(bundle["sheets"]),
        "sheets": bundle["sheets"],
        "defined_names": bundle["defined_names"],
        "auxiliary_sheets": aux,
        "formula_cell_count": len(bundle["formulas"]),
    }
    (output_dir / "workbook_inventory.json").write_text(json.dumps(inventory, indent=2), encoding="utf-8")

    schema_rows = []
    for sheet in bundle["sheets"]:
        for index, header in enumerate(sheet["headers"], start=1):
            schema_rows.append(
                {
                    "sheet": sheet["sheet"],
                    "column_index": index,
                    "header": header,
                    "max_row": sheet["max_row"],
                }
            )
    pd.DataFrame(schema_rows).to_csv(output_dir / "workbook_schema.csv", index=False)
    pd.DataFrame(bundle["formulas"]).to_csv(output_dir / "workbook_formula_inventory.csv", index=False)

    source_manifest = {
        "created_at": utc_now(),
        "source_workbook": workbook_path.name,
        "source_workbook_repo_path": _repo_relative(workbook_path),
        "sha256": workbook_sha,
        "size_bytes": workbook_path.stat().st_size,
        "immutable_source_snapshot": True,
        "runtime_reads_excel": False,
        "expected_period": args.expected_period,
        "ped_mode": args.ped_mode,
        "history_before_sha256": result["history_before_sha256"],
        "history_after_sha256": result["history_after_sha256"],
    }
    (output_dir / "source_workbook_manifest.json").write_text(
        json.dumps(source_manifest, indent=2), encoding="utf-8"
    )

    pd.DataFrame(kw["history_cross_check"]).to_csv(output_dir / "history_cross_check.csv", index=False)
    pd.DataFrame(kw["parity_rows"]).to_csv(output_dir / "feature_parity.csv", index=False)
    pd.DataFrame(kw["identity_rows"]).to_csv(output_dir / "identity_checks.csv", index=False)
    pd.DataFrame(kw["raw_records"]).to_csv(output_dir / "canonical_raw_actuals.csv", index=False)
    pd.DataFrame(kw["raw_records"]).to_csv(output_dir / "extracted_q1_raw_rows.csv", index=False)
    pd.DataFrame(kw["before_after_rows"]).to_csv(output_dir / "history_before_after.csv", index=False)
    pd.DataFrame(kw["lead_audit_rows"]).to_csv(output_dir / "heavy_lead_vintage_audit.csv", index=False)
    pd.DataFrame(kw["observations"]).to_csv(output_dir / "observation_classification.csv", index=False)

    engineered_frames = []
    for stream in STREAMS:
        frame = kw["updated_history"][stream].copy()
        frame.insert(0, "stream", stream.upper())
        engineered_frames.append(frame)
    pd.concat(engineered_frames, ignore_index=True, sort=False).to_csv(
        output_dir / "canonical_engineered_history.csv", index=False
    )

    lineage_rows: list[dict[str, Any]] = []
    if aux.get("source_map"):
        try:
            raw_map = pd.read_excel(kw["workbook_path"], sheet_name=aux["source_map"], header=2)
            raw_map.columns = [str(c).strip() for c in raw_map.columns]
            raw_map = raw_map.dropna(how="all")
            raw_map.to_csv(output_dir / "source_lineage.csv", index=False)
        except Exception as exc:  # pragma: no cover - defensive
            pd.DataFrame([{"error": str(exc)}]).to_csv(output_dir / "source_lineage.csv", index=False)
    else:
        pd.DataFrame(lineage_rows).to_csv(output_dir / "source_lineage.csv", index=False)

    if aux.get("reconciliation"):
        try:
            recon = pd.read_excel(kw["workbook_path"], sheet_name=aux["reconciliation"], header=2)
            recon.columns = [str(c).strip() for c in recon.columns]
            recon = recon.dropna(how="all")
            recon.to_csv(output_dir / "q1_reconciliation.csv", index=False)
        except Exception as exc:  # pragma: no cover - defensive
            pd.DataFrame([{"error": str(exc)}]).to_csv(output_dir / "q1_reconciliation.csv", index=False)

    run_manifest = {
        "created_at": utc_now(),
        "tool": "scripts/refresh_model_actuals.py",
        "arguments": {
            "workbook": str(args.workbook),
            "expected_period": args.expected_period,
            "ped_mode": args.ped_mode,
            "output_dir": str(args.output_dir),
            "check_only": bool(args.check_only),
        },
        "result": result,
        "environment": {
            "python": sys.version,
            "platform": platform.platform(),
            "pandas": pd.__version__,
            "numpy": np.__version__,
        },
        "git": _git_state(),
        "authoritative_production_engines": {
            "PED": "AR(1) GLSAR (pipeline/ar1_engine.py, ped_ar1 promoted state)",
            "LIGHT_RUC": "dynamic_RESID_GBR_n150_d1_lr0.05_w36 (light_ruc_vnext promoted state)",
            "HEAVY_RUC": "HEAVY_RUC__VNEXT_SOLVED_CONVEX_TOP4 (heavy_ruc_vnext promoted state)",
        },
    }
    (output_dir / "run_manifest.json").write_text(json.dumps(run_manifest, indent=2), encoding="utf-8")


def _repo_relative(path: Path) -> str | None:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except ValueError:
        return None


def _git_state() -> dict[str, str]:
    out = {}
    for label, cmd in {
        "head_sha": ["git", "rev-parse", "HEAD"],
        "branch": ["git", "rev-parse", "--abbrev-ref", "HEAD"],
    }.items():
        try:
            out[label] = subprocess.run(
                cmd, cwd=ROOT, capture_output=True, text=True, timeout=30, check=True
            ).stdout.strip()
        except Exception:
            out[label] = ""
    return out


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Governed quarterly actuals refresh for the model-input history.")
    parser.add_argument("--workbook", required=True, help="Source actuals workbook (immutable snapshot).")
    parser.add_argument("--expected-period", default=None, help="Optional guard: the period the refresh must add.")
    parser.add_argument("--ped-mode", choices=PED_MODES, default="exclude")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--history-dir", default=str(HISTORY_DIR_DEFAULT))
    parser.add_argument("--expected-sha256", default=None)
    parser.add_argument("--check-only", action="store_true", help="Validate and emit artifacts without writing history.")
    parser.add_argument(
        "--governance-approval",
        default=None,
        help="Required to activate --ped-mode accepted (owner + decision reference).",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        result = run_refresh(args)
    except RefreshError as exc:
        print(f"ACTUALS_REFRESH_FAILED: {exc}", file=sys.stderr)
        return 1
    status = "NO_OP" if result["no_op"] else ("CHECK_ONLY" if result["check_only"] else "APPLIED")
    print(f"ACTUALS_REFRESH_{status} periods={result['periods_applied']} ped_mode={result['ped_mode']}")
    for key, state in sorted(result["applied_state"].items()):
        print(f"  {key}: {state}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
