"""Quarterly glass-box XLSX for one Current-model Single Scenario path.

A formula-driven quarterly workbook (calendar quarters 2000Q3-2050Q2, matching
the annual extract's FY2001-FY2050 horizon) whose top block preserves the
annual extract's rows 1-65 labels and order, and whose detail sections below
row 75 rebuild every governed line from named model parameters, committed
scenario inputs and exact imported nonlinear component predictions:

- PED is formula-reproduced from the AR(1) engine's fitted coefficients,
  including the geometric AR(1) error recursion seeded from the committed
  last residual.
- Light RUC combines an Excel-reproduced OLS base with the exact imported
  residual-GBR log component from the production scorer.
- Heavy RUC uses the exact imported component predictions and the governed
  ensemble weights (level-space blend, per the production forward scorer).
- The scenario/policy overlay chain, the FY2031+ PREBU growth handover, the
  VFM fleet allocation and every revenue identity follow the production
  formula registry; each identity is asserted in Python against the governed
  displayed values before the workbook is written.

The workbook is an exact fixed-scenario replay of the committed artifacts.
It never reads source Excel workbooks and never refits a model.
"""

from __future__ import annotations

import io
import json
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Mapping, Sequence

import numpy as np
import pandas as pd

from model_dashboard.engine import ENGINE_AR1
from model_dashboard.revenue_outlook_excel_extract import (
    ROW_SERIES,
    TEMPLATE_SHEET_NAME,
)

__all__ = [
    "GLASSBOX_BUTTON_LABEL",
    "GLASSBOX_FIRST_QUARTER",
    "GLASSBOX_LAST_QUARTER",
    "GLASSBOX_SUPPORTED_TRACE",
    "GlassboxResult",
    "RevenueOutlookGlassboxError",
    "build_quarterly_glassbox_workbook",
    "glassbox_quarter_grid",
    "glassbox_supported_selection",
    "glassbox_supported_trace",
]


class RevenueOutlookGlassboxError(RuntimeError):
    """The glass-box workbook cannot be built for this selection."""


GLASSBOX_BUTTON_LABEL = "Download quarterly glass-box XLSX"

#: The quarter grid mirrors the annual extract horizon exactly: FY2001-FY2050
#: is 2000Q3 through 2050Q2 in calendar quarters (June fiscal years).
GLASSBOX_FIRST_QUARTER = "2000Q3"
GLASSBOX_LAST_QUARTER = "2050Q2"

#: v1 scope: the central Current-model path only. Conflict reforecasts, the
#: High-population comparison, official vintages and the PREBU deferral
#: workbook are declared out of scope rather than approximated.
#: The dashboard's central view: the "Current conditions baseline" trace
#: (the Low conflict-severity path re-labelled as current conditions). Its
#: committed replay scenarios are the middle_east_low family.
GLASSBOX_SUPPORTED_TRACE = "Current conditions baseline"
GLASSBOX_LAST_ACTUAL_QUARTER = "2025Q2"
_SCENARIO_PREFIX = "middle_east_low"
GLASSBOX_SUPPORTED_ENGINE = ENGINE_AR1
_BASE_SCENARIO = "current_basecase"

#: Where the committed per-scenario replay artifacts live, per engine.
_REPLAY_CACHE_REL = "data/revenue_outlook_replay_cache"
_PED_AR1_STATE_REL = (
    "data/dashboard_evidence_pack_reproducibility/ped_ar1/ar1_fitted_state.json"
)
_HEAVY_MANIFEST_REL = (
    "data/dashboard_evidence_pack_reproducibility/heavy_ruc_vnext/fitted_model_manifest.json"
)
_MODEL_HISTORY_REL = "data/model_input_history"

#: The dashboard's chart-level series (governed quarterly display exists).
CHART_QUARTERLY_SERIES: tuple[str, ...] = (
    "ped_vkt_per_capita",
    "light_petrol_vkt",
    "ped_volume",
    "light_ruc_net_km",
    "heavy_ruc_net_km",
    "light_bev_ruc_net_km",
    "phev_ruc_net_km",
    "gross_ped_revenue",
    "gross_fed_revenue",
    "net_fed_revenue",
    "light_ruc_net_revenue",
    "heavy_ruc_net_revenue",
    "light_bev_ruc_net_revenue",
    "phev_ruc_net_revenue",
    "total_ruc_net_revenue",
    "net_mvr_revenue",
    "total_fed_ruc_net_revenue",
    "total_nltf_net_revenue",
)

#: Leaves with no governed quarterly display contract. Their quarters are a
#: builder-side neutral-flat Denton allocation of the governed annual value -
#: the same construction the display contract itself declares for MVR and the
#: Total NLTF line - and every such row is labelled GOVERNED CARRY.
CARRIED_ALLOCATED_SERIES: tuple[str, ...] = (
    "gross_lpg_revenue",
    "gross_cng_revenue",
    "fed_refunds",
    "heavy_bev_ruc_net_km",
    "heavy_bev_ruc_net_revenue",
    "ruc_refunds",
    "gross_ruc_revenue",
    "ruc_admin_revenue",
    "ruc_revenue_net_admin",
    "mr1_revenue",
    "mr2_revenue",
    "coo_revenue",
    "gross_mvr_revenue",
    "mvr_admin_revenue",
    "mvr_revenue_net_admin_coo",
    "mvr_refunds",
    "tuc_gtk",
    "tuc_net_revenue",
    "total_gross_revenue",
    "total_admin_fees",
    "total_revenue_net_admin",
    "total_refunds",
)

#: PED AR(1) coefficient defined-name map, in fitted-state feature order.
PED_COEFFICIENT_NAMES: tuple[tuple[str, str], ...] = (
    ("petrol__log", "PED_BPet"),
    ("gdp_pc__log", "PED_BGDP"),
    ("unemp__level", "PED_BUnemp"),
    ("time__trend", "PED_BTrend"),
    ("time__post2011_trend", "PED_BPost11"),
    ("time__post2020", "PED_BPost20"),
    ("time__covid2020", "PED_BCovid"),
    ("time__q2", "PED_BQ2"),
    ("time__q3", "PED_BQ3"),
    ("time__q4", "PED_BQ4"),
)

#: Light RUC OLS coefficient defined-name map, in governed feature order.
LIGHT_COEFFICIENT_NAMES: tuple[tuple[str, str], ...] = (
    ("log_real_diesel_price", "LR_BDiesel"),
    ("log_real_light_ruc_price", "LR_BRuc"),
    ("log_lagged_real_light_ruc_price", "LR_BRucLag"),
    ("log_real_gdp", "LR_BGDP"),
    ("post_2020_dummy", "LR_BPost20"),
    ("q2_dummy", "LR_BQ2"),
    ("q3_dummy", "LR_BQ3"),
    ("q4_dummy", "LR_BQ4"),
)

_PARITY_TOLERANCE_REL = 1e-9
_ANNUAL_CLOSE_REL = 5e-9


# ---------------------------------------------------------------------------
# Quarter arithmetic
# ---------------------------------------------------------------------------


def _quarter_tuple(quarter: str) -> tuple[int, int]:
    text = str(quarter).strip()
    return int(text[:4]), int(text[-1])


def _quarter_str(year: int, q: int) -> str:
    return f"{year}Q{q}"


def _quarter_add(quarter: str, n: int) -> str:
    year, q = _quarter_tuple(quarter)
    index = year * 4 + (q - 1) + n
    return _quarter_str(index // 4, index % 4 + 1)


def _quarter_index(quarter: str) -> int:
    year, q = _quarter_tuple(quarter)
    return year * 4 + (q - 1)


def fiscal_year_of_quarter(quarter: str) -> int:
    """June fiscal year of a calendar quarter (FYn = (n-1)Q3 .. nQ2)."""
    year, q = _quarter_tuple(quarter)
    return year + 1 if q >= 3 else year


def quarters_of_fiscal_year(fy: int) -> tuple[str, str, str, str]:
    return (f"{fy - 1}Q3", f"{fy - 1}Q4", f"{fy}Q1", f"{fy}Q2")


def glassbox_quarter_grid() -> tuple[str, ...]:
    """2000Q3 .. 2050Q2, one column per calendar quarter (200 quarters)."""
    quarters: list[str] = []
    quarter = GLASSBOX_FIRST_QUARTER
    while _quarter_index(quarter) <= _quarter_index(GLASSBOX_LAST_QUARTER):
        quarters.append(quarter)
        quarter = _quarter_add(quarter, 1)
    return tuple(quarters)


def glassbox_supported_trace(trace_name: str, engine: str) -> str:
    """Empty string when supported; otherwise the reader-facing refusal."""
    trace = str(trace_name or "").strip()
    if trace != GLASSBOX_SUPPORTED_TRACE:
        return (
            "Current-model scenarios only: the quarterly glass-box workbook "
            f"replays the Current model's coefficient chain, so '{trace or '(none)'}' "
            "is not supported. Select the Current conditions / Central baseline path."
        )
    if str(engine) != GLASSBOX_SUPPORTED_ENGINE:
        return (
            "The quarterly glass-box workbook reproduces the production AR(1) "
            "PED engine; switch the engine to AR(1) to export it."
        )
    return ""


def glassbox_supported_selection(
    *,
    trace_name: str,
    engine: str,
    uptake_basis: str,
    custom_ev_levers: Sequence[Any],
    eruc_levers: Sequence[Any],
    ped_retention_sensitivity: bool,
    fed_ruc_transition: str,
    sensitivities_off: bool,
) -> str:
    """Reader-facing refusal for the whole page selection, or empty string.

    v1 replays the Central baseline under any governed 12c rate state. Every
    other value-changing lever would put the displayed values outside the
    identities this workbook proves, so the export refuses rather than
    approximating (the parity layer would fail closed anyway).
    """
    refusal = glassbox_supported_trace(trace_name, engine)
    if refusal:
        return refusal
    from model_dashboard.ev_uptake_levers import DEFAULT_EV_UPTAKE_MODE

    if str(uptake_basis) != str(DEFAULT_EV_UPTAKE_MODE) or tuple(custom_ev_levers):
        return (
            "v1 replays the MoT VFM base uptake basis only; set EV/PHEV uptake "
            "back to the default to export the glass-box workbook."
        )
    if tuple(eruc_levers):
        return "v1 does not replay the e-RUC transition lever; switch it Off to export."
    if bool(ped_retention_sensitivity):
        return "v1 does not replay the PED retention sensitivity; switch it Off to export."
    if str(fed_ruc_transition) not in ("", "off"):
        return (
            "v1 does not replay the fleetwide FED->RUC transition (rows 66-69); "
            "switch it Off to export."
        )
    if not sensitivities_off:
        return (
            "v1 replays the central path with sensitivities Off (fleet efficiency, "
            "PT mode shift, freight shift, demand elasticity)."
        )
    return ""


# ---------------------------------------------------------------------------
# Result container
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class GlassboxResult:
    """Built workbook plus the Python-side parity evidence."""

    workbook_bytes: bytes
    sheet_names: tuple[str, ...]
    trace_name: str
    scenario_name: str
    parity: Mapping[str, float]
    warnings: tuple[str, ...] = ()


# ---------------------------------------------------------------------------
# Committed-artifact access
# ---------------------------------------------------------------------------


def _read_parquet(path: Path, label: str) -> pd.DataFrame:
    if not path.exists():
        raise RevenueOutlookGlassboxError(f"{label} is missing at {path}.")
    return pd.read_parquet(path)


def _numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").astype(float)


def _replay_frame(repo_root: Path, engine: str, name: str) -> pd.DataFrame:
    return _read_parquet(
        Path(repo_root) / _REPLAY_CACHE_REL / engine / "frames" / f"{name}.parquet",
        f"replay cache frame {name} ({engine})",
    )


def _policy_scenario_name(policy_state: str) -> str:
    """The committed replay scenario for the central view under one 12c state.

    The Current conditions baseline is the middle_east_low family; its policy
    variants join with a double underscore (middle_east_low__mcert).
    """
    state = str(policy_state).strip()
    if state in ("", "published"):
        return _SCENARIO_PREFIX
    if state == "off":
        return f"{_SCENARIO_PREFIX}__12c_no_uplift"
    if state.startswith("delayed_"):
        months = state.removeprefix("delayed_")
        return f"{_SCENARIO_PREFIX}__12c_delay_{months}"
    return f"{_SCENARIO_PREFIX}__{state}"


def _scenario_stream_frame(
    frame: pd.DataFrame, scenario: str, stream: str, label: str
) -> pd.DataFrame:
    scoped = frame[
        frame["scenario_name"].astype(str).eq(scenario)
        & frame["stream"].astype(str).eq(stream)
    ].copy()
    if scoped.empty:
        raise RevenueOutlookGlassboxError(
            f"{label} carries no rows for scenario {scenario!r} stream {stream}."
        )
    period_column = "target_period" if "target_period" in scoped.columns else "period"
    scoped["__q"] = scoped[period_column].astype(str)
    return scoped.sort_values("__q").set_index("__q")


# ---------------------------------------------------------------------------
# Neutral-flat Denton allocation for annual-only carried lines
# ---------------------------------------------------------------------------


def _neutral_flat_quarters(annual_by_fy: Mapping[int, float]) -> dict[str, float]:
    """Minimum quarter-to-quarter movement subject to exact annual closure.

    The same construction the quarterly display contract declares for lines
    with no governed quarterly activity indicator (MVR, the Total NLTF line):
    a Denton proportional-first-difference split of each June year against a
    flat seasonal basis, chained across years so year boundaries stay smooth.
    """
    from model_dashboard import revenue_outlook_series_coverage as coverage

    years = sorted(int(fy) for fy, value in annual_by_fy.items() if pd.notna(value))
    if not years:
        return {}
    out: dict[str, float] = {}
    run: list[int] = []
    for fy in years + [None]:  # type: ignore[list-item]
        if run and (fy is None or fy != run[-1] + 1):
            annual = np.asarray([float(annual_by_fy[y]) for y in run], dtype=float)
            indicator = np.ones(4 * len(run), dtype=float)
            split = coverage._denton_quarterly_split(annual, indicator, average=False)
            run_quarters = [q for y in run for q in quarters_of_fiscal_year(y)]
            out.update({q: float(v) for q, v in zip(run_quarters, split)})
            run = []
        if fy is not None:
            run.append(fy)
    return out


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------


@dataclass
class _PedModel:
    beta0: float
    coefficients: dict[str, float]  # feature -> beta, fitted-state order
    beta_ylag: float
    rho: float
    last_resid: float
    latest_actual: str
    log_target_latest: float
    state_sha256: str
    train_window: tuple[str, str, int]
    features: dict[str, dict[str, float]]  # quarter -> feature -> value
    raw_prediction: dict[str, float]  # quarter -> committed reference level


@dataclass
class _LightModel:
    intercept: float
    coefficients: dict[str, float]
    state_sha256: str
    train_window: tuple[str, str, int]
    residual_model_id: str
    features: dict[str, dict[str, float]]
    base_log: dict[str, float]  # committed OLS base log prediction
    residual_log: dict[str, float]  # committed exact GBR residual log
    raw_prediction: dict[str, float]


@dataclass
class _HeavyComponent:
    label: str
    model_id: str
    kind: str
    weight: float
    state_sha256: str
    train_window: tuple[str, str, int]
    log_value: dict[str, float]
    level: dict[str, float]


@dataclass
class _HeavyModel:
    finalist_id: str
    components: list[_HeavyComponent]
    raw_prediction: dict[str, float]


@dataclass
class _PolicyOverlay:
    state_id: str
    state_label: str
    pair_id: str
    elasticity: dict[str, float]  # stream -> governed Med elasticity
    price_ratio: dict[str, dict[str, float]]  # stream -> quarter -> ratio
    gdp_factor: dict[str, dict[str, float]]  # stream -> quarter -> GDP factor
    reference_price: dict[str, dict[str, float]]
    variant_price: dict[str, dict[str, float]]
    calibrated: dict[str, dict[str, float]]  # stream -> quarter -> displayed forecast
    ped_nominal: dict[str, dict[str, float]]  # column -> quarter -> value
    ruc_detail: dict[str, dict[str, dict[str, float]]]  # stream -> column -> quarter
    diesel_litres_per_100km: dict[str, float]  # stream -> constant
    annual_pair_factor: dict[str, dict[int, float]]  # series -> fy -> factor
    annual_rate_factor: dict[int, float]  # fy -> governed scalar rate ratio
    rate_priced_long_run_series: frozenset[str]
    last_decision_grade_fy: int


@dataclass
class _PostModel:
    anchor_fy: int
    completion_fy: int
    fys: list[int]
    weight: dict[int, float]
    current_index: dict[str, dict[int, float]]
    structural_index: dict[str, dict[int, float]]
    hybrid_index: dict[str, dict[int, float]]
    anchors: dict[str, float]
    scenario_population: dict[int, float]
    vfm_shares: dict[str, dict[int, float]]
    vfm_pool_million_km: dict[int, float]
    raw_vktpc_fy: dict[int, float]
    raw_petrol_fy: dict[int, float]
    raw_heavy_fy: dict[int, float]
    shape_vintage_id: str
    schedule_id: str


@dataclass
class _GlassboxData:
    trace_name: str
    scenario_name: str
    engine: str
    quarters: tuple[str, ...]
    fys: tuple[int, ...]
    row_labels: dict[int, str]  # template rows 1..65
    annual_values: dict[str, dict[int, float]]  # series -> fy -> displayed annual
    quarterly_values: dict[str, dict[str, float]]  # series -> quarter -> displayed
    quarterly_source: dict[str, str]  # series -> "governed" | "carried_allocated"
    native_quarters: dict[str, set[str]]
    actual_quarters: dict[str, dict[str, float]]  # series -> quarter -> actual value
    actual_kind: dict[str, dict[str, str]]  # series -> quarter -> native|derived
    actual_annual: dict[str, dict[int, float]]  # Actual-trace annual values
    status: dict[str, str]  # quarter -> ACTUAL/MODEL FORECAST/POST-MODEL
    drivers: dict[str, dict[str, float]]  # driver row key -> quarter -> value
    ped: _PedModel
    light: _LightModel
    heavy: _HeavyModel
    policy: _PolicyOverlay
    post_model: _PostModel
    annual_bridge: dict[str, dict[int, float]]  # series -> fy -> bridge value
    policy_rate_factor_by_fy: dict[int, float]
    effective_rates: dict[str, dict[int, float]]  # rate id -> fy -> $/unit
    intensity_l_per_100km: dict[int, float]
    official_spine: dict[str, dict[int, float]]  # bridge-vintage series -> fy
    bridge_vintage_id: str
    pack_central_total_ruc: dict[int, float]
    macro_terminal_factor_total_ruc: float
    provenance: dict[str, str]
    parity: dict[str, float] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


def _record_parity(data: _GlassboxData, check: str, delta: float, limit: float) -> None:
    data.parity[check] = float(delta)
    if not math.isfinite(delta) or delta > limit:
        raise RevenueOutlookGlassboxError(
            f"glass-box parity check failed: {check} delta {delta:.3e} exceeds {limit:.0e}."
        )


def _template_row_labels(template_path: Path) -> dict[int, str]:
    from openpyxl import load_workbook

    if not template_path.exists():
        raise RevenueOutlookGlassboxError(
            f"extract template is missing at {template_path}."
        )
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
        raise RevenueOutlookGlassboxError(
            f"extract template has no {TEMPLATE_SHEET_NAME!r} sheet."
        )
    sheet = workbook[TEMPLATE_SHEET_NAME]
    labels = {
        row: ("" if sheet.cell(row=row, column=1).value is None else str(sheet.cell(row=row, column=1).value))
        for row in range(1, 66)
    }
    workbook.close()
    if labels[65] != "Total net revenues (m $)":
        raise RevenueOutlookGlassboxError(
            "extract template row 65 is not 'Total net revenues (m $)'; refusing to build."
        )
    return labels


def _annual_value_lookup(
    value_frames: Sequence[pd.DataFrame], trace_name: str
) -> dict[str, dict[int, float]]:
    """series -> FY -> value with first-frame-wins precedence (extract rule).

    History years fall back to the Base-case source path exactly as the
    annual extract does: FY2001-FY2024 history is scenario-invariant and the
    committed pack stack carries it under that path only.
    """
    from model_dashboard.revenue_outlook_excel_extract import (
        HISTORY_FALLBACK_SOURCE_PATH,
    )

    out: dict[str, dict[int, float]] = {}
    for paths in ((trace_name,), (HISTORY_FALLBACK_SOURCE_PATH,)):
        history_only = paths[0] != trace_name
        for frame in value_frames:
            if frame is None or frame.empty:
                continue
            required = {"source_path", "series_id", "FY", "value"}
            if not required.issubset(frame.columns):
                continue
            scoped = frame[frame["source_path"].astype(str).isin(paths)]
            for row in scoped.itertuples():
                fy = pd.to_numeric(pd.Series([row.FY]), errors="coerce").iloc[0]
                value = pd.to_numeric(pd.Series([row.value]), errors="coerce").iloc[0]
                if pd.isna(fy) or pd.isna(value):
                    continue
                if history_only and int(fy) >= 2026:
                    continue
                out.setdefault(str(row.series_id), {}).setdefault(int(fy), float(value))
    return out


def _collect_quarterly_display(
    data: _GlassboxData, quarterly_rows: pd.DataFrame
) -> None:
    """Governed quarterly display values per chart series, from the view gate."""
    if quarterly_rows is None or quarterly_rows.empty:
        raise RevenueOutlookGlassboxError(
            "no governed quarterly display rows were supplied for the selected path."
        )
    scoped = quarterly_rows[
        quarterly_rows["trace_name"].astype(str).eq(data.trace_name)
        & quarterly_rows["time_grain"].astype(str).eq("quarterly")
    ]
    values = _numeric(scoped["value"])
    for (series_id, period), value in zip(
        zip(scoped["series_id"].astype(str), scoped["period"].astype(str)), values
    ):
        if pd.isna(value):
            continue
        data.quarterly_values.setdefault(series_id, {})[period] = float(value)
    for series_id in CHART_QUARTERLY_SERIES:
        if series_id not in data.quarterly_values:
            raise RevenueOutlookGlassboxError(
                f"governed quarterly display is missing series {series_id} for "
                f"{data.trace_name}."
            )
        data.quarterly_source[series_id] = "governed"


def _collect_carried_allocations(data: _GlassboxData) -> None:
    for series_id in CARRIED_ALLOCATED_SERIES:
        annual = data.annual_values.get(series_id, {})
        forecast_years = {
            fy: value for fy, value in annual.items() if fy >= 2026 and pd.notna(value)
        }
        if not forecast_years:
            data.warnings.append(
                f"{series_id}: no governed annual values to allocate; row stays blank."
            )
            continue
        data.quarterly_values[series_id] = _neutral_flat_quarters(forecast_years)
        data.quarterly_source[series_id] = "carried_allocated"


def _collect_actual_history(
    data: _GlassboxData, chart_rows: pd.DataFrame, repo_root: Path
) -> None:
    """Actual-period quarters for EVERY top-block row (owner request).

    Three layers, each labelled so the workbook can colour them apart:
    1. natively published actual quarters (the three modelled activity series);
    2. the governed quarterly-display presentation of annual actuals (the
       display pack's Actual trace, derived and labelled as such);
    3. a neutral-flat allocation of the remaining annual-only actual lines
       (LPG, refunds, admin, MVR, TUC, aggregates), the same construction the
       display contract itself uses for lines without quarterly evidence.
    """
    annual_actual = chart_rows[
        chart_rows["trace_name"].astype(str).eq("Actual")
        & chart_rows["time_grain"].astype(str).eq("june_year")
    ]
    annual_fy = pd.to_numeric(annual_actual["june_year"], errors="coerce")
    for (series_id, fy), value in zip(
        zip(annual_actual["series_id"].astype(str), annual_fy),
        _numeric(annual_actual["value"]),
    ):
        if pd.isna(fy) or pd.isna(value):
            continue
        data.actual_annual.setdefault(series_id, {})[int(fy)] = float(value)

    actual = chart_rows[
        chart_rows["trace_name"].astype(str).eq("Actual")
        & chart_rows["time_grain"].astype(str).eq("quarterly")
    ]
    values = _numeric(actual["value"])
    for (series_id, period), value in zip(
        zip(actual["series_id"].astype(str), actual["period"].astype(str)), values
    ):
        if pd.isna(value):
            continue
        data.actual_quarters.setdefault(series_id, {})[period] = float(value)
        data.actual_kind.setdefault(series_id, {})[period] = "native"

    pack_path = (
        Path(repo_root) / "data" / "revenue_outlook_quarterly_display" / "quarterly_rows.parquet"
    )
    if pack_path.exists():
        pack_rows = pd.read_parquet(pack_path)
        pack_actual = pack_rows[pack_rows["trace_name"].astype(str).eq("Actual")]
        pack_values = _numeric(pack_actual["value"])
        for (series_id, period), value in zip(
            zip(pack_actual["series_id"].astype(str), pack_actual["period"].astype(str)),
            pack_values,
        ):
            if pd.isna(value):
                continue
            series_quarters = data.actual_quarters.setdefault(series_id, {})
            if period in series_quarters:
                continue  # never overwrite a published native quarter
            series_quarters[period] = float(value)
            data.actual_kind.setdefault(series_id, {})[period] = "derived"

    # Complete-year gaps before the first published/derived quarter are
    # filled with the neutral-flat presentation of the annual actuals, so
    # every top-block row carries its full actual history (owner request).
    # Years already covered by a published or pack-derived quarter are never
    # touched, and a partially covered June year is left as published.
    last_actual_fy = fiscal_year_of_quarter(GLASSBOX_LAST_ACTUAL_QUARTER)
    for series_id in tuple(CHART_QUARTERLY_SERIES) + tuple(CARRIED_ALLOCATED_SERIES):
        existing = data.actual_quarters.get(series_id, {})
        covered_fys = {fiscal_year_of_quarter(q) for q in existing}
        annual_source = data.actual_annual.get(
            series_id, data.annual_values.get(series_id, {})
        )
        history = {
            fy: value
            for fy, value in annual_source.items()
            if fy <= last_actual_fy and fy not in covered_fys and pd.notna(value)
        }
        if not history:
            continue
        allocated = _neutral_flat_quarters(history)
        if not allocated:
            continue
        series_quarters = data.actual_quarters.setdefault(series_id, {})
        series_kind = data.actual_kind.setdefault(series_id, {})
        for quarter, value in allocated.items():
            if quarter in series_quarters:
                continue
            series_quarters[quarter] = value
            series_kind[quarter] = "derived"


def _collect_ped(data: _GlassboxData, repo_root: Path, replay_inputs: pd.DataFrame,
                 future: pd.DataFrame) -> None:
    state_path = Path(repo_root) / _PED_AR1_STATE_REL
    if not state_path.exists():
        raise RevenueOutlookGlassboxError(f"PED AR(1) fitted state missing at {state_path}.")
    state = json.loads(state_path.read_text(encoding="utf-8"))
    beta = [float(value) for value in state["beta"]]
    features = [str(name) for name in state["features"]]
    expected = [name for name, _ in PED_COEFFICIENT_NAMES]
    if features != expected:
        raise RevenueOutlookGlassboxError(
            "PED AR(1) fitted-state features do not match the glass-box coefficient "
            f"map: {features} vs {expected}."
        )
    manifest = json.loads(
        (state_path.parent / "fitted_model_manifest.json").read_text(encoding="utf-8")
    )
    ar1_meta = (manifest.get("production_states") or {}).get("AR1") or {}

    history = _read_parquet(
        Path(repo_root) / _MODEL_HISTORY_REL / "ped_inputs.parquet",
        "PED model input history",
    ).set_index("period")
    latest_actual = str(state["latest_actual"])
    log_target_latest = float(history.loc[latest_actual, "log_target"])

    inputs = _scenario_stream_frame(replay_inputs, _BASE_SCENARIO, "PED", "replay inputs")
    committed = _scenario_stream_frame(future, _BASE_SCENARIO, "PED", "future forecasts")
    raw = _numeric(committed["prediction"]).to_dict()

    feature_rows: dict[str, dict[str, float]] = {}
    for quarter, row in inputs.iterrows():
        year, q = _quarter_tuple(quarter)
        trend = float((year - 2000) * 4 + q)
        feature_rows[str(quarter)] = {
            "petrol__log": math.log(float(row["real_petrol_price_cents_per_litre"])),
            "gdp_pc__log": math.log(float(row["real_gdp_per_capita_nzd"])),
            "unemp__level": float(row["unemployment_rate"]),
            "time__trend": trend,
            "time__post2011_trend": trend if year >= 2011 else 0.0,
            "time__post2020": 1.0 if year >= 2020 else 0.0,
            "time__covid2020": 1.0 if year == 2020 else 0.0,
            "time__q2": 1.0 if q == 2 else 0.0,
            "time__q3": 1.0 if q == 3 else 0.0,
            "time__q4": 1.0 if q == 4 else 0.0,
        }

    data.ped = _PedModel(
        beta0=beta[0],
        coefficients={name: beta[1 + i] for i, name in enumerate(features)},
        beta_ylag=beta[-1],
        rho=float(state["rho"][0]),
        last_resid=float(state["last_resid"]),
        latest_actual=latest_actual,
        log_target_latest=log_target_latest,
        state_sha256=str(ar1_meta.get("sha256", "")),
        train_window=(
            str(state.get("train_window_start", "")),
            str(state.get("train_window_end", "")),
            int(state.get("train_rows", 0)),
        ),
        features=feature_rows,
        raw_prediction={str(k): float(v) for k, v in raw.items() if pd.notna(v)},
    )

    # Replay the workbook arithmetic and prove it against the committed pack.
    err = data.ped.last_resid
    y_prev = log_target_latest
    max_rel = 0.0
    for quarter in sorted(feature_rows, key=_quarter_index):
        err = data.ped.rho * err
        log_linear = data.ped.beta0 + sum(
            data.ped.coefficients[name] * feature_rows[quarter][name]
            for name in data.ped.coefficients
        )
        log_pred = log_linear + data.ped.beta_ylag * y_prev + err
        level = math.exp(log_pred)
        committed_level = data.ped.raw_prediction.get(quarter)
        if committed_level is not None:
            max_rel = max(max_rel, abs(level - committed_level) / abs(committed_level))
        y_prev = log_pred
    _record_parity(data, "ped_ar1_recursion_vs_committed", max_rel, _PARITY_TOLERANCE_REL)


def _collect_light(data: _GlassboxData, repo_root: Path, replay_inputs: pd.DataFrame,
                   future: pd.DataFrame) -> None:
    import pipeline.vnext_forward as vnext_forward

    vnext_forward._register_legacy_sklearn_loss_module_alias()
    from model_dashboard.forecast_runner import (
        LIGHT_RUC_BASE_FEATURES,
        load_light_ruc_promoted_state,
    )

    state = load_light_ruc_promoted_state(repo_root)
    ols_beta = np.asarray(state.ols_beta, dtype=float)
    expected = [name for name, _ in LIGHT_COEFFICIENT_NAMES]
    if list(LIGHT_RUC_BASE_FEATURES) != expected:
        raise RevenueOutlookGlassboxError(
            "Light RUC governed base features do not match the glass-box "
            f"coefficient map: {list(LIGHT_RUC_BASE_FEATURES)} vs {expected}."
        )

    inputs = _scenario_stream_frame(replay_inputs, _BASE_SCENARIO, "LIGHT_RUC", "replay inputs")
    committed = _scenario_stream_frame(future, _BASE_SCENARIO, "LIGHT_RUC", "future forecasts")

    feature_rows: dict[str, dict[str, float]] = {}
    for quarter, row in inputs.iterrows():
        year, q = _quarter_tuple(quarter)
        feature_rows[str(quarter)] = {
            "log_real_diesel_price": math.log(float(row["real_diesel_price_cents_per_litre"])),
            "log_real_light_ruc_price": math.log(float(row["real_light_ruc_price_nzd_per_1000km"])),
            "log_lagged_real_light_ruc_price": math.log(
                float(row["lagged_real_light_ruc_price_nzd_per_1000km"])
            ),
            "log_real_gdp": math.log(float(row["real_gdp_sa_nzd"])),
            "post_2020_dummy": 1.0 if year >= 2021 else 0.0,
            "q2_dummy": 1.0 if q == 2 else 0.0,
            "q3_dummy": 1.0 if q == 3 else 0.0,
            "q4_dummy": 1.0 if q == 4 else 0.0,
        }

    base_level = _numeric(committed["base_forecast"])
    residual_log = _numeric(committed["residual_log_correction"])
    prediction = _numeric(committed["prediction"])
    base_log = {
        str(q): math.log(float(v)) for q, v in base_level.items() if pd.notna(v) and v > 0
    }
    data.light = _LightModel(
        intercept=float(ols_beta[0]),
        coefficients={
            name: float(ols_beta[1 + i]) for i, name in enumerate(LIGHT_RUC_BASE_FEATURES)
        },
        state_sha256=str(state.sha256),
        train_window=(state.train_window_start, state.train_window_end, state.train_rows),
        residual_model_id="dynamic_RESID_GBR_n150_d1_lr0.05_w36",
        features=feature_rows,
        base_log=base_log,
        residual_log={str(q): float(v) for q, v in residual_log.items() if pd.notna(v)},
        raw_prediction={str(q): float(v) for q, v in prediction.items() if pd.notna(v)},
    )

    max_rel = 0.0
    for quarter, committed_base_log in base_log.items():
        computed = data.light.intercept + sum(
            data.light.coefficients[name] * feature_rows[quarter][name]
            for name in data.light.coefficients
        )
        max_rel = max(max_rel, abs(computed - committed_base_log))
        combined = math.exp(computed + data.light.residual_log[quarter])
        expected_level = data.light.raw_prediction[quarter]
        max_rel = max(max_rel, abs(combined - expected_level) / abs(expected_level))
    _record_parity(data, "light_ols_plus_gbr_vs_committed", max_rel, _PARITY_TOLERANCE_REL)


def _collect_heavy(data: _GlassboxData, repo_root: Path, components: pd.DataFrame,
                   future: pd.DataFrame) -> None:
    manifest = json.loads(
        (Path(repo_root) / _HEAVY_MANIFEST_REL).read_text(encoding="utf-8")
    )
    finalist = str(manifest.get("finalist_model", ""))
    production_states = manifest.get("production_states") or {}

    scoped = components[
        components["scenario_name"].astype(str).eq(_BASE_SCENARIO)
        & components["stream"].astype(str).eq("HEAVY_RUC")
    ].copy()
    if scoped.empty:
        raise RevenueOutlookGlassboxError(
            "committed Heavy RUC component forecasts are missing for the central path."
        )
    scoped["component_forecast"] = _numeric(scoped["component_forecast"])
    scoped["component_log_value"] = _numeric(scoped["component_log_value"])
    scoped["component_weight"] = _numeric(scoped["component_weight"])
    scoped["period_q"] = scoped["target_period"].astype(str)

    committed = _scenario_stream_frame(future, _BASE_SCENARIO, "HEAVY_RUC", "future forecasts")
    prediction = {
        str(q): float(v)
        for q, v in _numeric(committed["prediction"]).items()
        if pd.notna(v)
    }

    members: list[_HeavyComponent] = []
    for label in sorted(scoped["component_label"].astype(str).unique()):
        if label == "FINAL":
            continue
        member_rows = scoped[scoped["component_label"].astype(str).eq(label)]
        weight = float(member_rows["component_weight"].iloc[0])
        model_id = str(member_rows["component_model"].iloc[0])
        meta = production_states.get(label) or {}
        kind = "ridge" if "ridge" in model_id else "gbr"
        members.append(
            _HeavyComponent(
                label=label,
                model_id=model_id,
                kind=kind,
                weight=weight,
                state_sha256=str(meta.get("sha256", "")),
                train_window=(
                    str(meta.get("train_window_start", "")),
                    str(meta.get("train_window_end", "")),
                    int(meta.get("train_rows", 0)),
                ),
                log_value={
                    str(row.period_q): float(row.component_log_value)
                    for row in member_rows.itertuples()
                    if pd.notna(row.component_log_value)
                },
                level={
                    str(row.period_q): float(row.component_forecast)
                    for row in member_rows.itertuples()
                    if pd.notna(row.component_forecast)
                },
            )
        )
    if len(members) < 2:
        raise RevenueOutlookGlassboxError(
            f"Heavy RUC finalist {finalist!r} exposes fewer than two committed components."
        )
    weight_sum = sum(member.weight for member in members)
    if abs(weight_sum - 1.0) > 1e-9:
        raise RevenueOutlookGlassboxError(
            f"Heavy RUC ensemble weights sum to {weight_sum!r}, not 1."
        )
    data.heavy = _HeavyModel(
        finalist_id=finalist, components=members, raw_prediction=prediction
    )

    max_rel = 0.0
    for quarter, final_level in prediction.items():
        blend = 0.0
        for member in members:
            log_value = member.log_value.get(quarter)
            if log_value is None:
                blend = math.nan
                break
            blend += member.weight * math.exp(log_value)
        if not math.isfinite(blend):
            continue
        max_rel = max(max_rel, abs(blend - final_level) / abs(final_level))
    _record_parity(data, "heavy_weighted_blend_vs_committed", max_rel, 1e-9)


def _collect_policy(
    data: _GlassboxData,
    repo_root: Path,
    policy_state: str,
    replay_inputs: pd.DataFrame,
    future: pd.DataFrame,
    pair_factors: pd.DataFrame,
    pack_chart_rows: pd.DataFrame,
) -> None:
    from model_dashboard.fed_policy_states import policy_spec as fed_policy_spec
    from model_dashboard.fuel_price_scenario import (
        _RUC_DIESEL_LITRES_PER_100KM,
        _governed_policy_demand_elasticities,
    )
    from model_dashboard.rate_paths import (
        _POLICY_PAIR_BY_STATE,
        _RATE_PRICED_LONG_RUN_SERIES,
        LAST_DECISION_GRADE_ANNUAL_FY,
        fed_policy_annual_factors,
    )

    spec = fed_policy_spec(policy_state)
    scenario = data.scenario_name
    elasticity_frame = _governed_policy_demand_elasticities(Path(repo_root))
    elasticities = {
        str(row.stream): float(row.value) for row in elasticity_frame.itertuples()
    }

    # The Current conditions baseline's committed calibration is ONE combined
    # layer: displayed = base raw prediction x (combined conflict+policy price
    # ratio) ^ elasticity x conflict GDP factor, all carried per quarter in
    # the demand_* audit columns. Proven exact for every stream and quarter.
    price_ratio: dict[str, dict[str, float]] = {}
    gdp_factor: dict[str, dict[str, float]] = {}
    reference_price: dict[str, dict[str, float]] = {}
    variant_price: dict[str, dict[str, float]] = {}
    calibrated: dict[str, dict[str, float]] = {}
    ruc_detail: dict[str, dict[str, dict[str, float]]] = {}
    max_rel = 0.0
    for stream in ("PED", "LIGHT_RUC", "HEAVY_RUC"):
        committed = _scenario_stream_frame(future, scenario, stream, "future forecasts")
        ratio = _numeric(committed["demand_price_ratio"])
        reference = _numeric(committed["demand_reference_forecast"])
        gdp = _numeric(committed["demand_gdp_model_factor"])
        displayed = _numeric(committed["forecast"])
        epsilon = elasticities[stream]
        price_ratio[stream] = {str(q): float(v) for q, v in ratio.items() if pd.notna(v)}
        gdp_factor[stream] = {str(q): float(v) for q, v in gdp.items() if pd.notna(v)}
        reference_price[stream] = {
            str(q): float(v)
            for q, v in _numeric(committed["demand_reference_price"]).items()
            if pd.notna(v)
        }
        variant_price[stream] = {
            str(q): float(v)
            for q, v in _numeric(committed["demand_variant_price"]).items()
            if pd.notna(v)
        }
        calibrated[stream] = {
            str(q): float(v) for q, v in displayed.items() if pd.notna(v)
        }
        for quarter in price_ratio[stream]:
            ref_value = reference.get(quarter)
            shown = calibrated[stream].get(quarter)
            gdp_value = gdp_factor[stream].get(quarter)
            if ref_value is None or shown is None or gdp_value is None or pd.isna(ref_value):
                continue
            rebuilt = (
                float(ref_value)
                * price_ratio[stream][quarter] ** epsilon
                * gdp_value
            )
            max_rel = max(max_rel, abs(rebuilt - shown) / abs(shown))
        if stream in ("LIGHT_RUC", "HEAVY_RUC"):
            detail: dict[str, dict[str, float]] = {}
            for column, key in (
                ("demand_reference_fuel_cost_nzd_per_1000km", "reference_fuel_cost"),
                ("demand_variant_fuel_cost_nzd_per_1000km", "variant_fuel_cost"),
                ("demand_reference_ruc_price_nzd_per_1000km", "reference_ruc_price"),
                ("demand_variant_ruc_price_nzd_per_1000km", "variant_ruc_price"),
            ):
                if column in committed.columns:
                    detail[key] = {
                        str(q): float(v)
                        for q, v in _numeric(committed[column]).items()
                        if pd.notna(v)
                    }
            ruc_detail[stream] = detail
    _record_parity(
        data, "policy_calibration_identity", max_rel, _PARITY_TOLERANCE_REL
    )

    ped_nominal: dict[str, dict[str, float]] = {}

    pair_id = str(_POLICY_PAIR_BY_STATE.get(str(spec.calculation_state_id), ""))
    annual_pair_factor: dict[str, dict[int, float]] = {}
    if pair_id and pair_factors is not None and not pair_factors.empty:
        annual_rows = pair_factors[
            pair_factors["pair_id"].astype(str).eq(pair_id)
            & pair_factors["time_grain"].astype(str).eq("june_year")
        ]
        factor = _numeric(annual_rows["factor"])
        fy = pd.to_numeric(annual_rows["june_year"], errors="coerce")
        for series_id, year, value in zip(
            annual_rows["series_id"].astype(str), fy, factor
        ):
            if pd.isna(year) or pd.isna(value):
                continue
            annual_pair_factor.setdefault(series_id, {})[int(year)] = float(value)

    annual_rate_factor = {
        int(fy): float(value)
        for fy, value in fed_policy_annual_factors(
            Path(repo_root), pack_chart_rows, str(spec.calculation_state_id)
        ).items()
    }

    data.policy = _PolicyOverlay(
        state_id=str(spec.state_id),
        state_label=str(spec.label),
        pair_id=pair_id,
        elasticity=elasticities,
        price_ratio=price_ratio,
        gdp_factor=gdp_factor,
        reference_price=reference_price,
        variant_price=variant_price,
        calibrated=calibrated,
        ped_nominal=ped_nominal,
        ruc_detail=ruc_detail,
        diesel_litres_per_100km={
            str(k): float(v) for k, v in _RUC_DIESEL_LITRES_PER_100KM.items()
        },
        annual_pair_factor=annual_pair_factor,
        annual_rate_factor=annual_rate_factor,
        rate_priced_long_run_series=frozenset(_RATE_PRICED_LONG_RUN_SERIES),
        last_decision_grade_fy=int(LAST_DECISION_GRADE_ANNUAL_FY),
    )

    # NATIVE governed quarters must equal the calibrated model path exactly.
    # Derived display quarters (the pre-model FY2026 completion and the
    # FY2031+ presentation quarters) reconcile to annual anchors instead and
    # are deliberately outside this identity.
    max_rel = 0.0
    checked = 0
    for series_id, stream, scale in (
        ("ped_vkt_per_capita", "PED", 1.0),
        ("light_ruc_net_km", "LIGHT_RUC", 1e6),
        ("heavy_ruc_net_km", "HEAVY_RUC", 1e6),
    ):
        values = data.quarterly_values.get(series_id, {})
        for quarter in data.native_quarters.get(series_id, set()):
            value = values.get(quarter)
            model_value = calibrated[stream].get(quarter)
            if value is None or model_value is None:
                continue
            expected = model_value / scale
            if abs(value) > 0:
                max_rel = max(max_rel, abs(expected - value) / abs(value))
                checked += 1
    if checked:
        _record_parity(
            data, "native_quarters_vs_calibrated_model", max_rel, 1e-9
        )


def _collect_annual_bridge(
    data: _GlassboxData, repo_root: Path, annual_bridge: pd.DataFrame
) -> None:
    # The replay's variant scenarios are bookkept under the comparison slot;
    # the scenario_name is the identity that matters here, and each scenario
    # carries exactly one source_path in the committed bridge.
    scoped = annual_bridge[
        annual_bridge["scenario_name"].astype(str).eq(data.scenario_name)
    ].copy()
    if scoped.empty:
        raise RevenueOutlookGlassboxError(
            f"annual bridge has no rows for {data.scenario_name!r}."
        )
    source_paths = sorted(scoped["source_path"].astype(str).unique())
    if len(source_paths) != 1:
        raise RevenueOutlookGlassboxError(
            f"annual bridge is ambiguous for {data.scenario_name!r}: {source_paths}."
        )
    scoped["value"] = _numeric(scoped["value"])
    scoped["FY"] = pd.to_numeric(scoped["FY"], errors="coerce")
    bridge: dict[str, dict[int, float]] = {}
    for row in scoped.itertuples():
        if pd.isna(row.FY) or pd.isna(row.value):
            continue
        bridge.setdefault(str(row.series_id), {})[int(row.FY)] = float(row.value)
    data.annual_bridge = bridge

    rate_factor = _numeric(scoped["policy_rate_factor"])
    factors: dict[int, float] = {}
    for fy, value in zip(scoped["FY"], rate_factor):
        if pd.isna(fy) or pd.isna(value):
            continue
        factors.setdefault(int(fy), float(value))
    data.policy_rate_factor_by_fy = factors

    data.effective_rates = {}


#: (activity series, revenue series) per RUC class, used for the rate chain.
_RUC_CLASS_PAIRS: tuple[tuple[str, str], ...] = (
    ("light_ruc_net_km", "light_ruc_net_revenue"),
    ("light_bev_ruc_net_km", "light_bev_ruc_net_revenue"),
    ("phev_ruc_net_km", "phev_ruc_net_revenue"),
    ("heavy_ruc_net_km", "heavy_ruc_net_revenue"),
    ("heavy_bev_ruc_net_km", "heavy_bev_ruc_net_revenue"),
)

#: Lines the production pipeline carries from the bridge vintage unchanged.
_SPINE_CARRIED_SERIES: tuple[str, ...] = (
    "gross_lpg_revenue",
    "gross_cng_revenue",
    "fed_refunds",
    "ruc_refunds",
    "ruc_admin_revenue",
    "mr1_revenue",
    "mr2_revenue",
    "coo_revenue",
    "mvr_admin_revenue",
    "mvr_refunds",
    "tuc_net_revenue",
    "tuc_gtk",
    "heavy_bev_ruc_net_km",
)


def _collect_official_spine(data: _GlassboxData, repo_root: Path) -> None:
    """Bridge-vintage official spine + the full-horizon rate-factor map.

    The verified production identity for every revenue leaf (FY2026-FY2050):

        displayed revenue = displayed activity
                            x (official spine revenue / official spine activity)
                            x selected-policy rate factor

    with the FY2031+ hidden Heavy-BEV line solved residually from the
    displayed Gross RUC closure (detail-frame behaviour).
    """
    from model_dashboard.official_vintage import (
        default_bridge_vintage_id,
        load_official_vintage,
    )

    vintage_id = default_bridge_vintage_id(repo_root)
    vintage = load_official_vintage(vintage_id, repo_root)
    if vintage is None:
        raise RevenueOutlookGlassboxError(
            f"bridge vintage {vintage_id!r} is not materialized."
        )
    pivot = vintage.annual_spine.pivot_table(
        index="FY", columns="series_id", values="value", aggfunc="first"
    )
    spine: dict[str, dict[int, float]] = {}
    for series in pivot.columns:
        values = _numeric(pivot[series]).dropna()
        spine[str(series)] = {int(fy): float(v) for fy, v in values.items()}
    data.official_spine = spine
    data.bridge_vintage_id = str(vintage_id)

    # One continuous rate-factor map FY2026-FY2050 from the governed scalar
    # schedule, proven equal to the replay bridge's own factor on the
    # decision-grade window.
    full = {
        fy: float(data.policy.annual_rate_factor.get(fy, 1.0))
        for fy in range(2026, 2051)
    }
    bridge_factors = dict(data.policy_rate_factor_by_fy)
    max_delta = 0.0
    for fy, factor in bridge_factors.items():
        if 2026 <= fy <= 2050:
            max_delta = max(max_delta, abs(full[fy] - factor))
    _record_parity(data, "rate_factor_scalar_vs_bridge", max_delta, 1e-12)
    data.policy_rate_factor_by_fy = full

    data.intensity_l_per_100km = {
        fy: 100.0 * spine["ped_volume"][fy] / spine["light_petrol_vkt"][fy]
        for fy in range(2026, 2051)
        if spine.get("ped_volume", {}).get(fy) and spine.get("light_petrol_vkt", {}).get(fy)
    }

    annual = data.annual_values
    max_rel = 0.0
    worst = ""

    def _norm(delta: float, base: float, where: str) -> None:
        nonlocal max_rel, worst
        rel = abs(delta) / max(abs(base), 1.0)
        if rel > max_rel:
            max_rel = rel
            worst = where

    for fy in range(2026, 2051):
        factor = full[fy]
        volume = annual.get("ped_volume", {}).get(fy)
        gross_ped = annual.get("gross_ped_revenue", {}).get(fy)
        if volume is not None and gross_ped is not None:
            rate = spine["gross_ped_revenue"][fy] / spine["ped_volume"][fy]
            _norm(volume * rate * factor - gross_ped, gross_ped, f"gross_ped FY{fy}")
        vkt = annual.get("light_petrol_vkt", {}).get(fy)
        if volume is not None and vkt is not None:
            _norm(
                vkt * data.intensity_l_per_100km[fy] / 100.0 - volume,
                volume,
                f"ped_volume FY{fy}",
            )
        for km_series, revenue_series in _RUC_CLASS_PAIRS:
            if revenue_series == "heavy_bev_ruc_net_revenue" and fy > data.policy.last_decision_grade_fy:
                continue  # hidden line: residual solve, checked below
            km = annual.get(km_series, {}).get(fy)
            revenue = annual.get(revenue_series, {}).get(fy)
            spine_km = spine.get(km_series, {}).get(fy)
            spine_rev = spine.get(revenue_series, {}).get(fy)
            if None in (km, revenue, spine_km, spine_rev) or abs(spine_km) < 1e-9:
                continue
            _norm(km * (spine_rev / spine_km) * factor - revenue, revenue,
                  f"{revenue_series} FY{fy}")
        for series in _SPINE_CARRIED_SERIES:
            shown = annual.get(series, {}).get(fy)
            carried = spine.get(series, {}).get(fy)
            if shown is None or carried is None:
                continue
            _norm(carried - shown, shown, f"carried {series} FY{fy}")
    _record_parity(data, "official_rate_and_carried_identities", max_rel, 1e-9)
    if worst:
        data.parity["official_rate_identity_worst"] = 0.0

    # FY2031+ RUC aggregates are built TOP-DOWN in production: the chart
    # carries Total RUC = pack central x Treasury macro terminal-carry factor
    # x governed rate ratio, and the detail layer solves the hidden Heavy-BEV
    # leaf residually so the registry identities close. Verify all three legs.
    max_rel = 0.0
    for fy in range(data.policy.last_decision_grade_fy + 1, 2051):
        factor = full[fy]
        pack_total = data.pack_central_total_ruc.get(fy)
        total_shown = annual.get("total_ruc_net_revenue", {}).get(fy)
        gross_shown = annual.get("gross_ruc_revenue", {}).get(fy)
        refunds = spine.get("ruc_refunds", {}).get(fy)
        admin = spine.get("ruc_admin_revenue", {}).get(fy)
        if None in (pack_total, total_shown, gross_shown, refunds, admin):
            continue
        rebuilt_total = pack_total * data.macro_terminal_factor_total_ruc * factor
        max_rel = max(
            max_rel, abs(rebuilt_total - total_shown) / max(abs(total_shown), 1.0)
        )
        rebuilt_gross = total_shown + admin + refunds
        max_rel = max(
            max_rel, abs(rebuilt_gross - gross_shown) / max(abs(gross_shown), 1.0)
        )
        hidden_shown = annual.get("heavy_bev_ruc_net_revenue", {}).get(fy)
        visible = sum(
            annual.get(revenue_series, {}).get(fy, math.nan)
            for _, revenue_series in _RUC_CLASS_PAIRS
            if revenue_series != "heavy_bev_ruc_net_revenue"
        )
        if hidden_shown is not None and math.isfinite(visible):
            residual = gross_shown - visible - refunds
            max_rel = max(
                max_rel, abs(residual - hidden_shown) / max(abs(hidden_shown), 1.0)
            )
    _record_parity(data, "post_model_total_ruc_and_hidden_bev", max_rel, 1e-9)


def _collect_post_model(data: _GlassboxData, repo_root: Path, pack_dir: Path) -> None:
    from model_dashboard.long_run_shape_transition import (
        growth_handover_index,
        structural_growth_indices,
        transition_weight,
    )
    from model_dashboard.official_vintage import (
        default_long_run_shape_vintage_id,
        load_official_vintage,
    )
    from model_dashboard.post_model_extrapolation import (
        ANCHOR_FY,
        FIRST_EXTRAPOLATION_FY,
        LAST_EXTRAPOLATION_FY,
    )

    raw_audit_path = Path(pack_dir) / "raw_quarterly_forecast_audit.csv"
    if not raw_audit_path.exists():
        raise RevenueOutlookGlassboxError(
            f"raw quarterly forecast audit missing at {raw_audit_path}."
        )
    audit = pd.read_csv(raw_audit_path)
    audit = audit[audit["scenario_name"].astype(str).eq(_BASE_SCENARIO)].copy()
    audit["value"] = _numeric(audit["value"])
    audit["fy"] = audit["period"].astype(str).map(fiscal_year_of_quarter)

    scenario_inputs = _read_parquet(
        Path(repo_root) / "data" / "current_revenue_outlook" / "scenario_inputs" / "scenario_input_wide.parquet",
        "scenario input wide",
    )
    population = scenario_inputs[
        scenario_inputs["scenario_name"].astype(str).eq(_BASE_SCENARIO)
        & scenario_inputs["stream"].astype(str).eq("PED")
    ][["canonical_period", "population"]].copy()
    population["population"] = _numeric(population["population"])
    pop_by_quarter = {
        str(row.canonical_period): float(row.population)
        for row in population.itertuples()
        if pd.notna(row.population)
    }

    ped = audit[audit["series_id"].astype(str).eq("ped_vkt_per_capita")]
    heavy = audit[audit["series_id"].astype(str).eq("heavy_ruc_net_km")]
    raw_vktpc_fy: dict[int, float] = {}
    raw_petrol_fy: dict[int, float] = {}
    for fy, group in ped.groupby("fy"):
        raw_vktpc_fy[int(fy)] = float(group["value"].sum())
        raw_petrol_fy[int(fy)] = float(
            sum(
                float(row.value) * pop_by_quarter.get(str(row.period), math.nan)
                for row in group.itertuples()
            )
        )
    raw_heavy_fy = {int(fy): float(g["value"].sum()) for fy, g in heavy.groupby("fy")}

    vfm = pd.read_csv(Path(repo_root) / "data" / "vfm_202405" / "vfm_vkt_shares.csv")
    vfm = vfm[vfm["scenario"].astype(str).eq("Base_EV")].set_index("june_year")
    vfm_pool = (
        _numeric(vfm["light_ruc_conventional_million_km"])
        + _numeric(vfm["light_ruc_bev_million_km"])
        + _numeric(vfm["light_ruc_phev_million_km"])
    )

    fys = list(range(ANCHOR_FY, LAST_EXTRAPOLATION_FY + 1))
    current_index = {
        "light_petrol_vkt": {fy: raw_petrol_fy[fy] / raw_petrol_fy[ANCHOR_FY] for fy in fys},
        "light_ruc_pool": {fy: float(vfm_pool.loc[fy] / vfm_pool.loc[ANCHOR_FY]) for fy in fys},
        "heavy_ruc_net_km": {fy: raw_heavy_fy[fy] / raw_heavy_fy[ANCHOR_FY] for fy in fys},
    }

    shape_vintage_id = default_long_run_shape_vintage_id(repo_root)
    vintage = load_official_vintage(shape_vintage_id, repo_root)
    if vintage is None:
        raise RevenueOutlookGlassboxError(
            f"long-run shape vintage {shape_vintage_id!r} is not materialized."
        )
    structural = structural_growth_indices(
        vintage.annual_spine,
        vintage_id=shape_vintage_id,
        first_fy=ANCHOR_FY,
        last_fy=LAST_EXTRAPOLATION_FY,
    ).set_index("fy")

    from model_dashboard.long_run_shape_transition import (
        PRODUCTION_LONG_RUN_TRANSITION_SCHEDULE_ID,
        resolve_schedule,
    )

    schedule = resolve_schedule(PRODUCTION_LONG_RUN_TRANSITION_SCHEDULE_ID)
    weight = {
        int(fy): float(w)
        for fy, w in zip(
            fys,
            transition_weight(
                np.asarray(fys, dtype=float),
                anchor_fy=schedule.anchor_fy,
                completion_fy=schedule.completion_fy,
            ),
        )
    }

    structural_index = {
        stream: {fy: float(structural.loc[fy, f"s_{stream}"]) for fy in fys}
        for stream in current_index
    }
    hybrid_index = {
        stream: {
            fy: float(value)
            for fy, value in zip(
                fys,
                growth_handover_index(
                    np.asarray([current_index[stream][fy] for fy in fys]),
                    np.asarray([structural_index[stream][fy] for fy in fys]),
                    np.asarray([weight[fy] for fy in fys]),
                    context=f"glassbox {stream}",
                ),
            )
        }
        for stream in current_index
    }

    shares = {
        "conventional": {},
        "bev": {},
        "phev": {},
    }
    for fy in fys:
        raw_shares = (
            float(vfm.loc[fy, "light_ruc_conventional_share"]),
            float(vfm.loc[fy, "light_ruc_bev_share"]),
            float(vfm.loc[fy, "light_ruc_phev_share"]),
        )
        total = sum(raw_shares)
        shares["conventional"][fy] = raw_shares[0] / total
        shares["bev"][fy] = raw_shares[1] / total
        shares["phev"][fy] = raw_shares[2] / total

    # The post-model tail anchors on the CENTRAL FY2030 level. Because the
    # displayed first post-model value IS anchor x hybrid at that year (the
    # tail carries no scenario or rate-path activity response), the anchor is
    # recovered exactly as displayed / hybrid at the first extrapolation year
    # - committed values only, no per-scenario factor bookkeeping.
    annual = data.annual_values
    first_post_fy = int(FIRST_EXTRAPOLATION_FY)

    def _anchor_from_first_post(series_ids: tuple[str, ...], stream: str) -> float:
        shown = sum(
            annual.get(series_id, {}).get(first_post_fy, math.nan)
            for series_id in series_ids
        )
        return shown / hybrid_index[stream][first_post_fy]

    anchors = {
        "light_petrol_vkt": _anchor_from_first_post(("light_petrol_vkt",), "light_petrol_vkt"),
        "heavy_ruc_net_km": _anchor_from_first_post(("heavy_ruc_net_km",), "heavy_ruc_net_km"),
        "light_ruc_pool": _anchor_from_first_post(
            ("light_ruc_net_km", "light_bev_ruc_net_km", "phev_ruc_net_km"),
            "light_ruc_pool",
        ),
    }
    scenario_population = {
        fy: raw_petrol_fy[fy] / raw_vktpc_fy[fy] for fy in fys
    }

    data.post_model = _PostModel(
        anchor_fy=int(ANCHOR_FY),
        completion_fy=int(schedule.completion_fy),
        fys=[fy for fy in fys if fy >= FIRST_EXTRAPOLATION_FY],
        weight=weight,
        current_index=current_index,
        structural_index=structural_index,
        hybrid_index=hybrid_index,
        anchors=anchors,
        scenario_population=scenario_population,
        vfm_shares=shares,
        vfm_pool_million_km={fy: float(vfm_pool.loc[fy]) for fy in fys},
        raw_vktpc_fy={fy: raw_vktpc_fy[fy] for fy in fys},
        raw_petrol_fy={fy: raw_petrol_fy[fy] for fy in fys},
        raw_heavy_fy={fy: raw_heavy_fy[fy] for fy in fys},
        shape_vintage_id=str(shape_vintage_id),
        schedule_id=str(schedule.schedule_id),
    )

    # Prove the handover reproduces the displayed post-model activity path:
    # displayed activity carries no policy demand response beyond FY2030 by
    # governed contract, so anchor x hybrid must equal the displayed annuals.
    max_rel = 0.0
    for fy in data.post_model.fys:
        petrol = anchors["light_petrol_vkt"] * hybrid_index["light_petrol_vkt"][fy]
        shown = annual.get("light_petrol_vkt", {}).get(fy)
        if shown:
            max_rel = max(max_rel, abs(petrol - shown) / abs(shown))
        heavy_level = anchors["heavy_ruc_net_km"] * hybrid_index["heavy_ruc_net_km"][fy]
        shown_heavy = annual.get("heavy_ruc_net_km", {}).get(fy)
        if shown_heavy:
            max_rel = max(max_rel, abs(heavy_level - shown_heavy) / abs(shown_heavy))
        pool = anchors["light_ruc_pool"] * hybrid_index["light_ruc_pool"][fy]
        for cls, series_id in (
            ("conventional", "light_ruc_net_km"),
            ("bev", "light_bev_ruc_net_km"),
            ("phev", "phev_ruc_net_km"),
        ):
            level = pool * shares[cls][fy]
            shown_cls = annual.get(series_id, {}).get(fy)
            if shown_cls:
                max_rel = max(max_rel, abs(level - shown_cls) / abs(shown_cls))
    _record_parity(data, "post_model_handover_vs_displayed", max_rel, 1e-6)


def _collect_drivers(data: _GlassboxData, replay_inputs: pd.DataFrame) -> None:
    """Quarterly Inputs sheet rows: exact selected-scenario production inputs."""
    driver_columns = {
        "population": ("PED", "population"),
        "real_gdp_sa_nzd": ("LIGHT_RUC", "real_gdp_sa_nzd"),
        "real_gdp_per_capita_nzd": ("PED", "real_gdp_per_capita_nzd"),
        "unemployment_rate": ("PED", "unemployment_rate"),
        "real_petrol_price_cents_per_litre": ("PED", "real_petrol_price_cents_per_litre"),
        "real_diesel_price_cents_per_litre": ("LIGHT_RUC", "real_diesel_price_cents_per_litre"),
        "real_light_ruc_price_nzd_per_1000km": ("LIGHT_RUC", "real_light_ruc_price_nzd_per_1000km"),
        "lagged_real_light_ruc_price_nzd_per_1000km": (
            "LIGHT_RUC",
            "lagged_real_light_ruc_price_nzd_per_1000km",
        ),
        "real_heavy_ruc_price_nzd_per_1000km": ("HEAVY_RUC", "real_heavy_ruc_price_nzd_per_1000km"),
        "lead_real_heavy_ruc_price_nzd_per_1000km": (
            "HEAVY_RUC",
            "lead_real_heavy_ruc_price_nzd_per_1000km",
        ),
    }
    for key, (stream, column) in driver_columns.items():
        for scenario, suffix in ((_BASE_SCENARIO, ""), (data.scenario_name, "__selected")):
            if scenario == _BASE_SCENARIO and suffix == "__selected":
                continue
            frame = _scenario_stream_frame(
                replay_inputs, scenario, stream, "replay inputs"
            )
            if column not in frame.columns:
                continue
            label = key if not suffix else f"{key}{suffix}"
            if scenario == data.scenario_name and data.scenario_name == _BASE_SCENARIO:
                continue
            data.drivers[label] = {
                str(q): float(v)
                for q, v in _numeric(frame[column]).items()
                if pd.notna(v)
            }
        if key not in data.drivers:
            frame = _scenario_stream_frame(replay_inputs, _BASE_SCENARIO, stream, "replay inputs")
            data.drivers[key] = {
                str(q): float(v)
                for q, v in _numeric(frame[column]).items()
                if pd.notna(v)
            }


def _collect_status(data: _GlassboxData) -> None:
    from model_dashboard.post_model_extrapolation import FIRST_EXTRAPOLATION_FY

    latest_actual_index = _quarter_index(data.ped.latest_actual)
    for quarter in data.quarters:
        fy = fiscal_year_of_quarter(quarter)
        if _quarter_index(quarter) <= latest_actual_index:
            data.status[quarter] = "ACTUAL"
        elif fy < FIRST_EXTRAPOLATION_FY:
            data.status[quarter] = "MODEL FORECAST"
        else:
            data.status[quarter] = "POST-MODEL"


def _assert_annual_closure(data: _GlassboxData) -> None:
    """Formula-registry identities on the displayed annual values."""
    from model_dashboard.mbu26_source_spine import FORMULA_DEFINITIONS

    max_rel = 0.0
    for definition in FORMULA_DEFINITIONS:
        output = str(definition["output_series_id"])
        terms = definition["terms"]
        shown = data.annual_values.get(output, {})
        for fy, value in shown.items():
            if fy < 2026:
                continue
            total = 0.0
            complete = True
            for series_id, sign in terms:
                term = data.annual_values.get(str(series_id), {}).get(fy)
                if term is None:
                    complete = False
                    break
                total += float(sign) * term
            if not complete or abs(value) < 1e-9:
                continue
            max_rel = max(max_rel, abs(total - value) / max(abs(value), 1.0))
    _record_parity(data, "annual_formula_registry_closure", max_rel, 1e-6)


def _assert_four_quarter_sums(data: _GlassboxData) -> None:
    """Every governed/allocated flow row must close to its annual value."""
    max_rel = 0.0
    worst = ""
    # Display years and actual-history years are checked against their own
    # source only: a June year straddling the actual/forecast seam mixes two
    # constructions and is deliberately outside this closure.
    for source_name, store in (
        ("display", data.quarterly_values),
        ("actual", data.actual_quarters),
    ):
        for series_id, quarters in store.items():
            if series_id == "ped_vkt_per_capita" and source_name == "actual":
                # Native actual VKT-per-capita quarters are published
                # observations; the annual line is the published annual
                # series, not their arithmetic sum. No closure is claimed.
                continue
            if source_name == "actual":
                # Actual-source quarters close to the ACTUAL trace's own
                # annuals; the Current trace's bridge anchor can legitimately
                # differ (e.g. the FY2025 BEV class reclassification).
                annual = data.actual_annual.get(
                    series_id, data.annual_values.get(series_id, {})
                )
            else:
                annual = data.annual_values.get(series_id, {})
            for fy in range(2002, 2051):
                expected = annual.get(fy)
                if expected is None or abs(expected) < 1e-9:
                    continue
                values = [quarters.get(q) for q in quarters_of_fiscal_year(fy)]
                if any(v is None for v in values):
                    continue
                total = sum(float(v) for v in values)
                rel = abs(total - expected) / max(abs(expected), 1.0)
                if rel > max_rel:
                    max_rel = rel
                    worst = f"{series_id} FY{fy} ({source_name})"
    data.parity["four_quarter_sum_worst_case"] = max_rel
    if max_rel > 1e-6:
        raise RevenueOutlookGlassboxError(
            f"four-quarter sums do not close to the annual display ({worst}: "
            f"{max_rel:.3e})."
        )


def collect_glassbox_data(
    *,
    trace_name: str,
    engine: str,
    current_fed_policy_state: str,
    quarterly_rows: pd.DataFrame,
    chart_rows: pd.DataFrame,
    pack_chart_rows: pd.DataFrame,
    annual_value_frames: Sequence[pd.DataFrame],
    template_path: Path,
    repo_root: Path,
    pack_dir: Path,
) -> _GlassboxData:
    """Assemble and Python-verify everything the workbook writer needs."""
    refusal = glassbox_supported_trace(trace_name, engine)
    if refusal:
        raise RevenueOutlookGlassboxError(refusal)

    repo_root = Path(repo_root)
    scenario = _policy_scenario_name(current_fed_policy_state)
    data = _GlassboxData(
        trace_name=str(trace_name),
        scenario_name=scenario,
        engine=str(engine),
        quarters=glassbox_quarter_grid(),
        fys=tuple(range(2001, 2051)),
        row_labels=_template_row_labels(Path(template_path)),
        annual_values={},
        quarterly_values={},
        quarterly_source={},
        native_quarters={},
        actual_quarters={},
        actual_kind={},
        actual_annual={},
        status={},
        drivers={},
        ped=None,  # type: ignore[arg-type]
        light=None,  # type: ignore[arg-type]
        heavy=None,  # type: ignore[arg-type]
        policy=None,  # type: ignore[arg-type]
        post_model=None,  # type: ignore[arg-type]
        annual_bridge={},
        policy_rate_factor_by_fy={},
        effective_rates={},
        intensity_l_per_100km={},
        official_spine={},
        bridge_vintage_id="",
        pack_central_total_ruc={},
        macro_terminal_factor_total_ruc=1.0,
        provenance={},
    )

    data.annual_values = _annual_value_lookup(annual_value_frames, data.trace_name)
    if not data.annual_values:
        raise RevenueOutlookGlassboxError(
            f"no annual values resolved for {data.trace_name!r}."
        )

    replay_inputs = _replay_frame(repo_root, engine, "fuel.replay_inputs")
    future = _replay_frame(repo_root, engine, "fuel.replay.future_forecasts")
    components = _replay_frame(repo_root, engine, "fuel.replay.component_forecasts")
    annual_bridge = _replay_frame(repo_root, engine, "fuel.annual_bridge")
    pair_factors = _replay_frame(repo_root, engine, "fuel.policy_pair_factors")

    if scenario != _BASE_SCENARIO:
        known = set(replay_inputs["scenario_name"].astype(str).unique())
        if scenario not in known:
            raise RevenueOutlookGlassboxError(
                f"the committed replay cache has no scenario {scenario!r} for the "
                f"selected 12c state {current_fed_policy_state!r}."
            )

    native = chart_rows[
        chart_rows["trace_name"].astype(str).eq(data.trace_name)
        & chart_rows["time_grain"].astype(str).eq("quarterly")
    ]
    for series_id, period in zip(
        native["series_id"].astype(str), native["period"].astype(str)
    ):
        data.native_quarters.setdefault(series_id, set()).add(period)

    _collect_quarterly_display(data, quarterly_rows)
    _collect_carried_allocations(data)
    _collect_actual_history(data, chart_rows, repo_root)
    _collect_ped(data, repo_root, replay_inputs, future)
    _collect_light(data, repo_root, replay_inputs, future)
    _collect_heavy(data, repo_root, components, future)
    _collect_policy(
        data, repo_root, current_fed_policy_state, replay_inputs, future,
        pair_factors, pack_chart_rows,
    )
    _collect_annual_bridge(data, repo_root, annual_bridge)

    # Pack central Total RUC + the Treasury macro terminal-carry factor: the
    # committed primitives of the FY2031+ top-down RUC aggregate construction.
    # The FY2031+ top-down RUC construction is CENTRAL regardless of the
    # displayed trace (the conflict effect ends at FY2030 and the tail rides
    # the central path exactly), so its primitives live on the base trace.
    pack_total = pack_chart_rows[
        pack_chart_rows["trace_name"].astype(str).eq("Current finalist Base case")
        & pack_chart_rows["time_grain"].astype(str).eq("june_year")
        & pack_chart_rows["series_id"].astype(str).eq("total_ruc_net_revenue")
    ]
    for fy, value in zip(
        pd.to_numeric(pack_total["june_year"], errors="coerce"),
        _numeric(pack_total["value"]),
    ):
        if pd.notna(fy) and pd.notna(value):
            data.pack_central_total_ruc[int(fy)] = float(value)
    macro_factors = _replay_frame(
        repo_root, engine, "macro.baseline_macro_annual_factors"
    )
    terminal = macro_factors[
        macro_factors["series_id"].astype(str).eq("total_ruc_net_revenue")
        & macro_factors["trace_name"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(macro_factors["june_year"], errors="coerce").eq(2030)
    ]
    if terminal.empty:
        raise RevenueOutlookGlassboxError(
            "Treasury macro annual factor for total_ruc_net_revenue at FY2030 is missing."
        )
    data.macro_terminal_factor_total_ruc = float(_numeric(terminal["factor"]).iloc[0])

    _collect_official_spine(data, repo_root)
    _collect_post_model(data, repo_root, Path(pack_dir))
    _collect_drivers(data, replay_inputs)
    _collect_status(data)
    _assert_annual_closure(data)
    _assert_four_quarter_sums(data)

    # Presentation scaling AFTER every identity is verified: TUC GTK is ~8e9
    # tonne-km a year, unreadable raw; the workbook shows it in millions.
    for store in (data.annual_values, data.official_spine):
        if "tuc_gtk" in store:
            store["tuc_gtk"] = {k: v / 1e6 for k, v in store["tuc_gtk"].items()}
    for store in (data.quarterly_values, data.actual_quarters):
        if "tuc_gtk" in store:
            store["tuc_gtk"] = {k: v / 1e6 for k, v in store["tuc_gtk"].items()}

    manifest_path = Path(pack_dir) / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8")) if manifest_path.exists() else {}
    period_rule = manifest.get("period_rule") or {}
    data.provenance = {
        "engine": data.engine,
        "scenario_name": data.scenario_name,
        "policy_state": str(current_fed_policy_state),
        "model_training_cutoff": str(period_rule.get("model_training_cutoff", "")),
        "pack_dir": str(pack_dir),
        "ped_state_sha256": data.ped.state_sha256,
        "light_state_sha256": data.light.state_sha256,
        "heavy_components": "; ".join(
            f"{member.label}={member.model_id} (w={member.weight:.6f}, sha={member.state_sha256[:12]})"
            for member in data.heavy.components
        ),
        "long_run_schedule": data.post_model.schedule_id,
        "long_run_shape_vintage": data.post_model.shape_vintage_id,
    }
    return data


# ---------------------------------------------------------------------------
# Public entry point (writer implemented in the same module, below)
# ---------------------------------------------------------------------------


def build_quarterly_glassbox_workbook(
    *,
    trace_name: str,
    engine: str,
    current_fed_policy_state: str,
    quarterly_rows: pd.DataFrame,
    chart_rows: pd.DataFrame,
    pack_chart_rows: pd.DataFrame,
    annual_value_frames: Sequence[pd.DataFrame],
    template_path: Path,
    repo_root: Path,
    pack_dir: Path,
    scenario_note: str = "",
    audit_sheets: bool = False,
) -> GlassboxResult:
    """Build the quarterly glass-box workbook for one Current-model path."""
    data = collect_glassbox_data(
        trace_name=trace_name,
        engine=engine,
        current_fed_policy_state=current_fed_policy_state,
        quarterly_rows=quarterly_rows,
        chart_rows=chart_rows,
        pack_chart_rows=pack_chart_rows,
        annual_value_frames=annual_value_frames,
        template_path=template_path,
        repo_root=repo_root,
        pack_dir=pack_dir,
    )
    from model_dashboard._quarterly_glassbox_writer import write_glassbox_workbook

    workbook_bytes, sheet_names = write_glassbox_workbook(
        data, scenario_note=scenario_note, audit_sheets=audit_sheets
    )
    return GlassboxResult(
        workbook_bytes=workbook_bytes,
        sheet_names=tuple(sheet_names),
        trace_name=data.trace_name,
        scenario_name=data.scenario_name,
        parity=dict(data.parity),
        warnings=tuple(data.warnings),
    )
