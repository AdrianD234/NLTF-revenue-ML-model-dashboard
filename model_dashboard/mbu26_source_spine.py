"""MBU26 annual source spine and Revenue Outlook bridge helpers.

The workbook is an offline source artifact only. Streamlit runtime code reads
the repo-local extracts produced under ``data/revenue_model_source_pack``.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import re
import sys
from typing import Any

_RUNTIME_PYARROW24 = Path(__file__).resolve().parents[1] / ".runtime_pyarrow24"
if _RUNTIME_PYARROW24.exists() and str(_RUNTIME_PYARROW24) not in sys.path:
    sys.path.insert(0, str(_RUNTIME_PYARROW24))

import pandas as pd

from .revenue_source_pack import (
    CURRENT_FINALIST_COMPOSITE_MODEL_ID,
    CURRENT_FINALIST_MODEL_IDS,
    REVENUE_FIRST_FORECAST_QUARTER,
    REVENUE_LAST_COMPLETE_ACTUAL_FY,
    REVENUE_MODEL_TRAINING_CUTOFF,
)


MBU26_SOURCE_PACK_DIR = Path("data") / "revenue_model_source_pack" / "mbu26_annual_spine"
MBU26_SCHEMA_VERSION = "nltf-revenue-mbu26-annual-spine-v1"
MBU26_SHEET_NAME = "MBU26"
MBU26_RELEASE_ROUND = "MBU26"
REVENUE_PARTIAL_ACTUAL_FY = 2026
CURRENT_MODEL_EXTENSION_START_FY = 2051
CURRENT_MODEL_EXTENSION_END_FY = 2055
CURRENT_MODEL_EXTENSION_BASE_START_FY = 2046
CURRENT_MODEL_EXTENSION_BASE_END_FY = 2050
CURRENT_LIGHT_TOTAL_SERIES_ID = "current_light_ruc_total_modelled_km"


@dataclass(frozen=True)
class MBU26AnnualSpinePack:
    pack_dir: Path
    manifest: dict[str, Any]
    annual_spine: pd.DataFrame
    official_annual: pd.DataFrame
    formula_audit: pd.DataFrame
    row_reconciliation: pd.DataFrame
    series_alias_audit: pd.DataFrame
    series_trace_contract: pd.DataFrame
    trace_source_contract: pd.DataFrame
    path_trace_status: pd.DataFrame


ROW_DEFINITIONS: tuple[dict[str, Any], ...] = (
    {"row": 10, "series_id": "light_ruc_net_km", "display_name": "Light RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 11, "series_id": "heavy_ruc_net_km", "display_name": "Heavy RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 12, "series_id": "light_bev_ruc_net_km", "display_name": "Light BEV RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 13, "series_id": "heavy_bev_ruc_net_km", "display_name": "Heavy BEV RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 14, "series_id": "phev_ruc_net_km", "display_name": "PHEV RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 15, "series_id": "ped_volume", "display_name": "PED volume", "section": "Key volumes", "unit": "million litres", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 16, "series_id": "light_petrol_vkt", "display_name": "Light petrol VKT", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 17, "series_id": "ped_vkt_per_capita", "source_series_id": "light_petrol_vkt_per_capita", "display_name": "PED VKT per capita", "source_display_name": "Light petrol VKT per capita", "section": "Key volumes", "unit": "km/person", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 18, "series_id": "tuc_gtk", "display_name": "TUC GTK", "section": "Key volumes", "unit": "tonne-km", "metric_type": "activity", "row_role": "bridge_input"},
    {"row": 33, "series_id": "light_ruc_net_revenue", "display_name": "Light RUC revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 34, "series_id": "heavy_ruc_net_revenue", "display_name": "Heavy RUC revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 35, "series_id": "light_bev_ruc_net_revenue", "display_name": "Light BEV RUC net revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 36, "series_id": "heavy_bev_ruc_net_revenue", "display_name": "Heavy BEV RUC net revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 37, "series_id": "phev_ruc_net_revenue", "display_name": "PHEV RUC net revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 38, "series_id": "ruc_refunds", "display_name": "RUC refunds", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"row": 39, "series_id": "gross_ruc_revenue", "display_name": "Gross RUC revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 40, "series_id": "ruc_admin_revenue", "display_name": "RUC admin revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"row": 41, "series_id": "ruc_revenue_net_admin", "display_name": "RUC revenues net of admin fees", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 42, "series_id": "total_ruc_net_revenue", "display_name": "Total RUC all classes", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 45, "series_id": "gross_ped_revenue", "display_name": "PED revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 46, "series_id": "gross_lpg_revenue", "display_name": "Gross LPG revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 47, "series_id": "gross_cng_revenue", "display_name": "Gross CNG revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 48, "series_id": "gross_fed_revenue", "display_name": "Gross FED revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 49, "series_id": "fed_refunds", "display_name": "FED refunds", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"row": 50, "series_id": "net_fed_revenue", "display_name": "Net FED revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 53, "series_id": "mr1_revenue", "display_name": "MR1 revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 54, "series_id": "mr2_revenue", "display_name": "MR2 revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 55, "series_id": "coo_revenue", "display_name": "MR13/COO revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"row": 56, "series_id": "gross_mvr_revenue", "display_name": "Gross MVR revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 57, "series_id": "mvr_admin_revenue", "display_name": "MVR admin revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"row": 58, "series_id": "mvr_revenue_net_admin_coo", "display_name": "MVR revenues net of admin fees and COO", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 59, "series_id": "mvr_refunds", "display_name": "MVR refunds", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"row": 60, "series_id": "net_mvr_revenue", "display_name": "Net MVR revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 63, "series_id": "tuc_net_revenue", "display_name": "TUC net revenue", "section": "TUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"row": 66, "series_id": "total_gross_revenue", "display_name": "Total gross revenues", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 67, "series_id": "total_admin_fees", "display_name": "Total admin fees", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 68, "series_id": "total_revenue_net_admin", "display_name": "Total revenues net of admin fees", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 69, "series_id": "total_refunds", "display_name": "Total refunds", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"row": 70, "series_id": "total_nltf_net_revenue", "display_name": "Total NLTF revenue", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
)

REVENUE_LINE_TABLE_SERIES_IDS: tuple[str, ...] = tuple(str(row["series_id"]) for row in ROW_DEFINITIONS) + (
    CURRENT_LIGHT_TOTAL_SERIES_ID,
)
REVENUE_LINE_SOURCE_PATHS: tuple[str, ...] = (
    "MBU26 official",
    "Current finalist Base case",
    "Current finalist High population/comparison",
)

SERIES_ALIAS_AUDIT_ROWS: tuple[dict[str, str], ...] = (
    {
        "source_label": "Light petrol VKT per capita (km)",
        "source_series_id": "light_petrol_vkt_per_capita",
        "runtime_series_id": "ped_vkt_per_capita",
        "dashboard_label": "PED VKT per capita",
        "unit": "km/person",
        "alias_reason": "MBU26 row 17 uses the worksheet label Light petrol VKT per capita; the dashboard and current finalist model use PED VKT per capita for the same annual activity concept.",
        "status": "canonical_mapping",
    },
    {
        "source_label": "PED VKT per capita",
        "source_series_id": "ped_vkt_per_capita",
        "runtime_series_id": "ped_vkt_per_capita",
        "dashboard_label": "PED VKT per capita",
        "unit": "km/person",
        "alias_reason": "Current finalist source rows already use the dashboard canonical series.",
        "status": "already_canonical",
    },
    {
        "source_label": "Gross PED revenue",
        "source_series_id": "gross_ped_revenue",
        "runtime_series_id": "gross_ped_revenue",
        "dashboard_label": "PED revenue",
        "unit": "$m nominal ex GST",
        "alias_reason": "Gross PED revenue and PED revenue are display aliases for the same MBU26 gross PED row.",
        "status": "display_alias",
    },
    {
        "source_label": "RUC revenues net of admin fees & refunds",
        "source_series_id": "total_ruc_net_revenue",
        "runtime_series_id": "total_ruc_net_revenue",
        "dashboard_label": "Total RUC all classes",
        "unit": "$m nominal ex GST",
        "alias_reason": "The dashboard selector keeps the MBU26 Total RUC all classes label; source-pack wording describes the same net-admin/refunds formula.",
        "status": "canonical_mapping",
    },
    {
        "source_label": "MR13 revenue",
        "source_series_id": "coo_revenue",
        "runtime_series_id": "coo_revenue",
        "dashboard_label": "MR13",
        "unit": "$m nominal ex GST",
        "alias_reason": "MR13 and COO revenue are label variants for the same MBU26 row 55 deduction component.",
        "status": "display_alias",
    },
    {
        "source_label": "MR13/COO revenue",
        "source_series_id": "coo_revenue",
        "runtime_series_id": "coo_revenue",
        "dashboard_label": "MR13",
        "unit": "$m nominal ex GST",
        "alias_reason": "The original MBU26/source label is retained in provenance while the compact dashboard line label is MR13.",
        "status": "display_alias",
    },
    {
        "source_label": "MVR revenues net of admin fees & COO",
        "source_series_id": "mvr_revenue_net_admin_coo",
        "runtime_series_id": "mvr_revenue_net_admin_coo",
        "dashboard_label": "MVR net admin & COO",
        "unit": "$m nominal ex GST",
        "alias_reason": "Long-form MBU26/source label and compact dashboard label describe the same MVR net-admin/COO component.",
        "status": "display_alias",
    },
    {
        "source_label": "MVR revenues net of admin fees, refunds & COO",
        "source_series_id": "net_mvr_revenue",
        "runtime_series_id": "net_mvr_revenue",
        "dashboard_label": "MVR net admin/refunds/COO",
        "unit": "$m nominal ex GST",
        "alias_reason": "This label variant includes refunds and maps to the final Net MVR revenue canonical row.",
        "status": "canonical_mapping",
    },
    {
        "source_label": "Total net revenues",
        "source_series_id": "total_nltf_net_revenue",
        "runtime_series_id": "total_nltf_net_revenue",
        "dashboard_label": "Total NLTF revenue",
        "unit": "$m nominal ex GST",
        "alias_reason": "Total net revenues is the source-pack wording for the final Total NLTF revenue line.",
        "status": "canonical_mapping",
    },
)


FORMULA_DEFINITIONS: tuple[dict[str, Any], ...] = (
    {"output_series_id": "gross_ruc_revenue", "expression": "light_ruc_net_revenue + heavy_ruc_net_revenue + light_bev_ruc_net_revenue + heavy_bev_ruc_net_revenue + phev_ruc_net_revenue + ruc_refunds", "terms": (("light_ruc_net_revenue", 1), ("heavy_ruc_net_revenue", 1), ("light_bev_ruc_net_revenue", 1), ("heavy_bev_ruc_net_revenue", 1), ("phev_ruc_net_revenue", 1), ("ruc_refunds", 1))},
    {"output_series_id": "ruc_revenue_net_admin", "expression": "gross_ruc_revenue - ruc_admin_revenue", "terms": (("gross_ruc_revenue", 1), ("ruc_admin_revenue", -1))},
    {"output_series_id": "total_ruc_net_revenue", "expression": "ruc_revenue_net_admin - ruc_refunds", "terms": (("ruc_revenue_net_admin", 1), ("ruc_refunds", -1))},
    {"output_series_id": "gross_fed_revenue", "expression": "gross_ped_revenue + gross_lpg_revenue + gross_cng_revenue", "terms": (("gross_ped_revenue", 1), ("gross_lpg_revenue", 1), ("gross_cng_revenue", 1))},
    {"output_series_id": "net_fed_revenue", "expression": "gross_fed_revenue - fed_refunds", "terms": (("gross_fed_revenue", 1), ("fed_refunds", -1))},
    {"output_series_id": "gross_mvr_revenue", "expression": "mr1_revenue + mr2_revenue + coo_revenue", "terms": (("mr1_revenue", 1), ("mr2_revenue", 1), ("coo_revenue", 1))},
    {"output_series_id": "mvr_revenue_net_admin_coo", "expression": "mr1_revenue + mr2_revenue - mvr_admin_revenue", "terms": (("mr1_revenue", 1), ("mr2_revenue", 1), ("mvr_admin_revenue", -1))},
    {"output_series_id": "net_mvr_revenue", "expression": "mvr_revenue_net_admin_coo - mvr_refunds", "terms": (("mvr_revenue_net_admin_coo", 1), ("mvr_refunds", -1))},
    {"output_series_id": "total_gross_revenue", "expression": "gross_ruc_revenue + gross_fed_revenue + gross_mvr_revenue + tuc_net_revenue", "terms": (("gross_ruc_revenue", 1), ("gross_fed_revenue", 1), ("gross_mvr_revenue", 1), ("tuc_net_revenue", 1))},
    {"output_series_id": "total_admin_fees", "expression": "ruc_admin_revenue + mvr_admin_revenue + coo_revenue", "terms": (("ruc_admin_revenue", 1), ("mvr_admin_revenue", 1), ("coo_revenue", 1))},
    {"output_series_id": "total_revenue_net_admin", "expression": "total_gross_revenue - total_admin_fees", "terms": (("total_gross_revenue", 1), ("total_admin_fees", -1))},
    {"output_series_id": "total_refunds", "expression": "ruc_refunds + fed_refunds + mvr_refunds", "terms": (("ruc_refunds", 1), ("fed_refunds", 1), ("mvr_refunds", 1))},
    {"output_series_id": "total_nltf_net_revenue", "expression": "total_revenue_net_admin - total_refunds", "terms": (("total_revenue_net_admin", 1), ("total_refunds", -1))},
    {"output_series_id": "total_fed_ruc_net_revenue", "output_label": "Total RUC+PED revenue", "expression": "net_fed_revenue + total_ruc_net_revenue", "terms": (("net_fed_revenue", 1), ("total_ruc_net_revenue", 1)), "source_kind": "derived_dashboard_subtotal"},
)


DISPLAY_SERIES_METADATA: tuple[dict[str, str], ...] = (
    {"canonical_id": "ped_vkt_per_capita", "display_name": "PED VKT per capita", "metric_type": "activity", "unit": "km/person"},
    {"canonical_id": "ped_volume", "display_name": "PED volume", "metric_type": "activity", "unit": "million litres"},
    {"canonical_id": "light_ruc_net_km", "display_name": "Light RUC net km", "metric_type": "activity", "unit": "million km"},
    {"canonical_id": "heavy_ruc_net_km", "display_name": "Heavy RUC net km", "metric_type": "activity", "unit": "million km"},
    {"canonical_id": "gross_ped_revenue", "display_name": "PED revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "light_ruc_net_revenue", "display_name": "Light RUC revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "heavy_ruc_net_revenue", "display_name": "Heavy RUC revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "gross_fed_revenue", "display_name": "Gross FED revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "net_fed_revenue", "display_name": "Net FED revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "total_ruc_net_revenue", "display_name": "Total RUC all classes", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "total_fed_ruc_net_revenue", "display_name": "Total RUC+PED revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "total_nltf_net_revenue", "display_name": "Total NLTF revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "net_mvr_revenue", "display_name": "Net MVR revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
    {"canonical_id": "tuc_net_revenue", "display_name": "TUC net revenue", "metric_type": "revenue", "unit": "$m nominal ex GST"},
)


def repo_root_from_here() -> Path:
    return Path(__file__).resolve().parents[1]


def load_mbu26_annual_spine(
    pack_dir: Path | str | None = None,
    repo_root: Path | str | None = None,
) -> MBU26AnnualSpinePack | None:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    base = Path(pack_dir) if pack_dir is not None else root / MBU26_SOURCE_PACK_DIR
    manifest_path = base / "manifest.json"
    if not manifest_path.exists():
        return None
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    _validate_manifest_hashes(base, manifest)
    return MBU26AnnualSpinePack(
        pack_dir=base,
        manifest=manifest,
        annual_spine=pd.read_csv(base / "mbu26_annual_spine.csv"),
        official_annual=pd.read_csv(base / "mbu26_official_annual.csv"),
        formula_audit=pd.read_csv(base / "mbu26_formula_audit.csv"),
        row_reconciliation=pd.read_csv(base / "row_reconciliation.csv"),
        series_alias_audit=pd.read_csv(base / "series_alias_audit.csv"),
        series_trace_contract=pd.read_csv(base / "series_trace_contract.csv"),
        trace_source_contract=pd.read_csv(base / "trace_source_contract.csv"),
        path_trace_status=pd.read_csv(base / "path_trace_status.csv"),
    )


def materialize_mbu26_annual_spine(
    workbook_path: Path | str,
    *,
    output_dir: Path | str | None = None,
    repo_root: Path | str | None = None,
    extracted_by: str = "codex_mbu26_rebuild",
) -> dict[str, Any]:
    import openpyxl

    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    output = Path(output_dir) if output_dir is not None else root / MBU26_SOURCE_PACK_DIR
    output.mkdir(parents=True, exist_ok=True)
    existing_manifest = _read_existing_manifest(output)
    workbook = Path(workbook_path)
    workbook_hash = sha256(workbook)
    values_wb = openpyxl.load_workbook(workbook, read_only=False, data_only=True)
    formulas_wb = openpyxl.load_workbook(workbook, read_only=False, data_only=False)
    try:
        values_sheet = values_wb[MBU26_SHEET_NAME]
        formulas_sheet = formulas_wb[MBU26_SHEET_NAME]
        annual_spine = _extract_annual_spine(values_sheet, formulas_sheet, workbook.name, workbook_hash)
    finally:
        values_wb.close()
        formulas_wb.close()

    formula_audit = _formula_audit_frame(annual_spine)
    official_annual = _official_annual_frame(annual_spine, formula_audit)
    alias_audit = revenue_series_alias_audit_frame(annual_spine)
    series_contract = _series_trace_contract_frame()
    trace_contract = _trace_source_contract_frame(workbook.name, workbook_hash)
    path_status = _path_trace_status_frame()

    frames = {
        "mbu26_annual_spine": annual_spine,
        "mbu26_formula_audit": formula_audit,
        "mbu26_official_annual": official_annual,
        "row_reconciliation": formula_audit.copy(),
        "series_alias_audit": alias_audit,
        "series_trace_contract": series_contract,
        "trace_source_contract": trace_contract,
        "path_trace_status": path_status,
    }
    file_hashes: dict[str, dict[str, Any]] = {}
    for stem, frame in frames.items():
        output_frame = _prepare_frame_for_output(frame)
        output_frame.to_csv(output / f"{stem}.csv", index=False)
        output_frame.to_parquet(output / f"{stem}.parquet", index=False)
        for suffix in ("csv", "parquet"):
            path = output / f"{stem}.{suffix}"
            file_hashes[path.name] = {"sha256": sha256(path), "size_bytes": path.stat().st_size}

    existing_workbook = existing_manifest.get("workbook") if isinstance(existing_manifest, dict) else {}
    extracted_at = (
        existing_manifest.get("extracted_at")
        if isinstance(existing_workbook, dict)
        and existing_workbook.get("sha256") == workbook_hash
        and existing_manifest.get("schema_version") == MBU26_SCHEMA_VERSION
        else datetime.now(timezone.utc).isoformat()
    )
    manifest = {
        "schema_version": MBU26_SCHEMA_VERSION,
        "source_release": MBU26_RELEASE_ROUND,
        "repo_relative_output_dir": _repo_relative(root, output),
        "source_policy": "MBU26 worksheet only; workbook is offline lineage and is never loaded at Streamlit runtime.",
        "workbook": {
            "basename": workbook.name,
            "sha256": workbook_hash,
            "size_bytes": workbook.stat().st_size,
            "sheet": MBU26_SHEET_NAME,
        },
        "extracted_at": extracted_at,
        "extracted_by": extracted_by,
        "row_count": {
            name: int(len(frame))
            for name, frame in frames.items()
        },
        "normalized_files": file_hashes,
        "formula_policy": (
            "MBU26 annual value cells are stored without Excel formulas in the grid; aggregate formula contracts "
            "are asserted from MBU26 row identities and residuals are reported without force-balancing."
        ),
    }
    (output / "manifest.json").write_text(json.dumps(manifest, indent=2, allow_nan=False) + "\n", encoding="utf-8")
    (output / "manifest.md").write_text(_manifest_markdown(manifest), encoding="utf-8")
    return manifest


def _prepare_frame_for_output(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.copy()
    for column in output.columns:
        if output[column].dtype == object:
            output[column] = output[column].where(output[column].notna(), "").astype(str)
    return output


def _read_existing_manifest(output: Path) -> dict[str, Any]:
    manifest_path = output / "manifest.json"
    if not manifest_path.exists():
        return {}
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def revenue_series_alias_audit_frame(annual_spine: pd.DataFrame | None = None) -> pd.DataFrame:
    columns = [
        "source_label",
        "source_series_id",
        "runtime_series_id",
        "dashboard_label",
        "unit",
        "source_row",
        "source_cell",
        "alias_reason",
        "status",
    ]
    row_by_runtime = {str(row["series_id"]): row for row in ROW_DEFINITIONS}
    row_by_source = {str(row.get("source_series_id") or row["series_id"]): row for row in ROW_DEFINITIONS}
    source_cells: dict[tuple[str, str], str] = {}
    source_rows: dict[tuple[str, str], str] = {}
    source_labels: dict[tuple[str, str], str] = {}
    if isinstance(annual_spine, pd.DataFrame) and not annual_spine.empty:
        spine = annual_spine.copy()
        for key, group in spine.groupby(
            [
                spine.get("source_series_id", spine.get("series_id", pd.Series("", index=spine.index))).astype(str),
                spine.get("series_id", pd.Series("", index=spine.index)).astype(str),
            ],
            dropna=False,
        ):
            source_series_id, runtime_series_id = str(key[0]), str(key[1])
            source_cells[(source_series_id, runtime_series_id)] = _source_cell_range(group)
            source_rows[(source_series_id, runtime_series_id)] = _first_text(group, "source_row")
            source_labels[(source_series_id, runtime_series_id)] = _first_text(group, "source_label")
    rows: list[dict[str, Any]] = []
    for item in SERIES_ALIAS_AUDIT_ROWS:
        source_series_id = item["source_series_id"]
        runtime_series_id = item["runtime_series_id"]
        key = (source_series_id, runtime_series_id)
        definition = row_by_source.get(source_series_id) or row_by_runtime.get(runtime_series_id, {})
        source_label = source_labels.get(key) if source_series_id != runtime_series_id else ""
        rows.append(
            {
                "source_label": source_label or item["source_label"],
                "source_series_id": source_series_id,
                "runtime_series_id": runtime_series_id,
                "dashboard_label": item["dashboard_label"],
                "unit": item["unit"],
                "source_row": source_rows.get(key) or str(definition.get("row") or ""),
                "source_cell": source_cells.get(key) or _definition_source_cell_hint(definition),
                "alias_reason": item["alias_reason"],
                "status": item["status"],
            }
        )
    return pd.DataFrame(rows, columns=columns).sort_values(["runtime_series_id", "source_label"], kind="stable").reset_index(drop=True)


def current_forecast_annual_from_mbu26(
    *,
    current_outlook_chart_rows: pd.DataFrame,
    mbu26_official_annual: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "FY",
        "period",
        "source_path",
        "fed_path",
        "scenario_name",
        "scenario_role",
        "series_id",
        "display_name",
        "section",
        "value",
        "unit",
        "row_role",
        "source_basis",
        "source_file",
        "source_cell",
        "source_status",
        "model_id",
        "formula",
        "replacement_only",
        "official_value",
        "residual_vs_official",
        "availability_status",
        "line",
        "model_basis",
        "forecast_path",
        "path_status",
        "value_status",
        "quarters_present",
        "actual_quarters",
        "forecast_quarters",
        "nowcast_flag",
    ]
    activity = _current_activity_annual_values(current_outlook_chart_rows)
    if activity.empty:
        return pd.DataFrame(columns=columns)
    quarter_lookup = _current_activity_quarter_lookup(current_outlook_chart_rows)
    official = _official_lookup(mbu26_official_annual)
    official_records = _official_record_lookup(mbu26_official_annual)
    rows: list[dict[str, Any]] = []
    keys = sorted({(str(row.scenario_name), int(row.FY)) for row in activity.itertuples()})
    for scenario_name, fy in keys:
        activity_values = {
            str(row.series_id): row
            for row in activity[activity["scenario_name"].eq(scenario_name) & activity["FY"].eq(fy)].itertuples()
        }
        required_activity = {"ped_vkt_per_capita", "light_ruc_net_km", "heavy_ruc_net_km"}
        if not required_activity.issubset(activity_values):
            continue
        off = official.get(fy, {})
        required_official = {
            "light_petrol_vkt",
            "ped_vkt_per_capita",
            "ped_volume",
            "gross_ped_revenue",
            "light_ruc_net_revenue",
            "light_ruc_net_km",
            "heavy_ruc_net_revenue",
            "heavy_ruc_net_km",
            "light_bev_ruc_net_km",
            "heavy_bev_ruc_net_km",
            "phev_ruc_net_km",
            "light_bev_ruc_net_revenue",
            "heavy_bev_ruc_net_revenue",
            "phev_ruc_net_revenue",
            "ruc_refunds",
            "gross_ruc_revenue",
            "ruc_admin_revenue",
            "ruc_revenue_net_admin",
            "gross_lpg_revenue",
            "gross_cng_revenue",
            "fed_refunds",
            "total_ruc_net_revenue",
            "mr1_revenue",
            "mr2_revenue",
            "coo_revenue",
            "gross_mvr_revenue",
            "mvr_admin_revenue",
            "mvr_revenue_net_admin_coo",
            "mvr_refunds",
            "net_mvr_revenue",
            "tuc_net_revenue",
            "tuc_gtk",
            "gross_fed_revenue",
            "net_fed_revenue",
            "total_gross_revenue",
            "total_admin_fees",
            "total_revenue_net_admin",
            "total_refunds",
            "total_nltf_net_revenue",
        }
        if not required_official.issubset(off):
            continue
        official_row = official_records.get(fy, {})
        ped_activity = activity_values["ped_vkt_per_capita"]
        light_activity = activity_values["light_ruc_net_km"]
        heavy_activity = activity_values["heavy_ruc_net_km"]
        ped_quarters = quarter_lookup.get((scenario_name, fy, "ped_vkt_per_capita"), {})
        light_quarters = quarter_lookup.get((scenario_name, fy, "light_ruc_net_km"), {})
        heavy_quarters = quarter_lookup.get((scenario_name, fy, "heavy_ruc_net_km"), {})
        if _is_extrapolated_activity_row(ped_activity):
            ped_quarters = _extrapolated_activity_quarter_payload(ped_activity, "ped_vkt_per_capita")
        if _is_extrapolated_activity_row(light_activity):
            light_quarters = _extrapolated_activity_quarter_payload(light_activity, "light_ruc_net_km")
        if _is_extrapolated_activity_row(heavy_activity):
            heavy_quarters = _extrapolated_activity_quarter_payload(heavy_activity, "heavy_ruc_net_km")
        if not ped_quarters or not light_quarters or not heavy_quarters:
            continue
        scenario_role = str(getattr(ped_activity, "scenario_role", "") or "")
        source_path = _current_source_path_label(scenario_name, scenario_role)
        value_status = (
            "Actual anchor"
            if fy == REVENUE_LAST_COMPLETE_ACTUAL_FY
            else "extrapolated_model_extension"
            if _is_extrapolated_activity_row(ped_activity)
            else "Current-finalist FY nowcast (2 actual + 2 forecast)"
            if bool(getattr(ped_activity, "nowcast_flag", False))
            else "current_finalist_forecast"
        )
        actual_quarters = str(getattr(ped_activity, "actual_quarters", "") or "")
        forecast_quarters = str(getattr(ped_activity, "forecast_quarters", "") or "")
        quarters_present = str(getattr(ped_activity, "quarters_present", "") or "")
        ped_source_cell = str(ped_quarters.get("source_cells") or f"current_revenue_outlook:{scenario_name}:FY{fy}:PED")
        light_source_cell = str(light_quarters.get("source_cells") or f"current_revenue_outlook:{scenario_name}:FY{fy}:LIGHT_RUC")
        heavy_source_cell = str(heavy_quarters.get("source_cells") or f"current_revenue_outlook:{scenario_name}:FY{fy}:HEAVY_RUC")
        current_source_cell = f"current_revenue_outlook:{scenario_name}:FY{fy}"
        ped_values = [float(value) for value in ped_quarters.get("values", [])]
        light_values = [float(value) for value in light_quarters.get("values", [])]
        heavy_values = [float(value) for value in heavy_quarters.get("values", [])]
        ped_vkt_per_capita = sum(ped_values)
        light_km_million = sum(light_values)
        heavy_km_million = sum(heavy_values)
        if abs(light_km_million) > 10_000_000:
            light_km_million /= 1_000_000.0
        if abs(heavy_km_million) > 10_000_000:
            heavy_km_million /= 1_000_000.0
        split = _light_ev_phev_split_from_official(off)
        if split["availability_status"] != "available":
            continue
        population_count = float(off["light_petrol_vkt"]) * 1_000_000.0 / float(off["ped_vkt_per_capita"])
        ped_litres_per_100km = float(off["ped_volume"]) / float(off["light_petrol_vkt"]) * 100.0
        ped_rate = float(off["gross_ped_revenue"]) / float(off["ped_volume"])
        ped_total_vkt = sum(value * population_count / 1_000_000.0 for value in ped_values)
        ped_volume = sum(value * population_count / 1_000_000.0 * ped_litres_per_100km / 100.0 for value in ped_values)
        ped_revenue = sum(value * population_count / 1_000_000.0 * ped_litres_per_100km / 100.0 * ped_rate for value in ped_values)
        light_total_modelled_km = light_km_million
        conventional_light_km = light_total_modelled_km * float(split["conventional_share"])
        current_light_bev_km = light_total_modelled_km * float(split["light_bev_share"])
        current_phev_km = light_total_modelled_km * float(split["phev_share"])
        allocation_residual = light_total_modelled_km - conventional_light_km - current_light_bev_km - current_phev_km
        if abs(allocation_residual) > 1e-6:
            conventional_light_km += allocation_residual
        light_rate = float(split["conventional_rate"])
        light_bev_rate = float(split["light_bev_rate"])
        phev_rate = float(split["phev_rate"])
        heavy_rate = float(off["heavy_ruc_net_revenue"]) / float(off["heavy_ruc_net_km"])
        light_revenue = conventional_light_km * light_rate
        light_bev_revenue = current_light_bev_km * light_bev_rate
        phev_revenue = current_phev_km * phev_rate
        heavy_revenue = heavy_km_million * heavy_rate
        gross_ruc = (
            light_revenue
            + heavy_revenue
            + light_bev_revenue
            + float(off["heavy_bev_ruc_net_revenue"])
            + phev_revenue
            + float(off["ruc_refunds"])
        )
        ruc_net_admin = gross_ruc - float(off["ruc_admin_revenue"])
        total_ruc = ruc_net_admin - float(off["ruc_refunds"])
        gross_fed = ped_revenue + float(off["gross_lpg_revenue"]) + float(off["gross_cng_revenue"])
        net_fed = gross_fed - float(off["fed_refunds"])
        gross_mvr = float(off["mr1_revenue"]) + float(off["mr2_revenue"]) + float(off["coo_revenue"])
        mvr_net_admin_coo = float(off["mr1_revenue"]) + float(off["mr2_revenue"]) - float(off["mvr_admin_revenue"])
        net_mvr = mvr_net_admin_coo - float(off["mvr_refunds"])
        total_gross = gross_ruc + gross_fed + gross_mvr + float(off["tuc_net_revenue"])
        total_admin = float(off["ruc_admin_revenue"]) + float(off["mvr_admin_revenue"]) + float(off["coo_revenue"])
        total_net_admin = total_gross - total_admin
        total_refunds = float(off["ruc_refunds"]) + float(off["fed_refunds"]) + float(off["mvr_refunds"])
        total_fed_ruc = net_fed + total_ruc
        total_nltf = total_net_admin - total_refunds
        common = {
            "FY": fy,
            "period": f"FY{fy}",
            "source_path": source_path,
            "fed_path": "Current planned path",
            "scenario_name": scenario_name,
            "scenario_role": scenario_role,
            "quarters_present": quarters_present,
            "actual_quarters": actual_quarters,
            "forecast_quarters": forecast_quarters,
            "nowcast_flag": bool(getattr(ped_activity, "nowcast_flag", False)),
            "value_status": value_status,
        }
        fixed = lambda series_id, **overrides: _current_fixed_row(
            common=common,
            official_row=official_row,
            series_id=series_id,
            **overrides,
        )
        split_cells = "; ".join(
            value
            for value in [
                _source_cell_for_series(official_row, "light_ruc_net_km"),
                _source_cell_for_series(official_row, "light_bev_ruc_net_km"),
                _source_cell_for_series(official_row, "phev_ruc_net_km"),
            ]
            if value
        )
        light_split_source_cell = f"{light_source_cell}; MBU26 split cells {split_cells}".strip("; ")
        light_split_basis = "current_finalist_light_total + MBU26 EV/PHEV split share"
        light_split_revenue_basis = "current_finalist_light_total + MBU26 EV/PHEV split share and effective rate"
        rows.extend(
            [
                _current_annual_row(**common, series_id="ped_vkt_per_capita", display_name="PED VKT per capita", section="Key volumes", value=ped_vkt_per_capita, unit="km/person", row_role="bridge_input", source_basis="current_finalist_model", source_file="forecast_scenario_comparison.parquet", source_cell=ped_source_cell, formula="sum quarterly current finalist PED VKT/capita", official_value=off["ped_vkt_per_capita"]),
                _current_annual_row(**common, series_id="ped_volume", display_name="PED volume", section="Key volumes", value=ped_volume, unit="million litres", row_role="bridge_input", source_basis="current finalist PED VKT/capita + MBU26 intensity", source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=ped_source_cell, formula="sum(quarterly PED VKT/capita * MBU26 population / 1,000,000 * MBU26 litres intensity / 100)", official_value=off["ped_volume"]),
                _current_annual_row(**common, series_id="light_petrol_vkt", display_name="Light petrol VKT", section="Key volumes", value=ped_total_vkt, unit="million km", row_role="bridge_input", source_basis="current_finalist_model + MBU26 population", source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=ped_source_cell, formula="sum(quarterly PED VKT/capita * MBU26 population / 1,000,000)", official_value=off["light_petrol_vkt"]),
                _current_annual_row(**common, series_id=CURRENT_LIGHT_TOTAL_SERIES_ID, display_name="Current finalist Light RUC total modelled km", section="Key volumes", value=light_total_modelled_km, unit="million km", row_role="audit_only", source_basis="current_finalist_model_total_light_ruc_universe", source_file="forecast_scenario_comparison.parquet", source_cell=light_source_cell, formula="sum quarterly current finalist Light RUC total net km before EV/PHEV allocation", official_value=pd.NA, model_id=CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]),
                _current_annual_row(**common, series_id="light_ruc_net_km", display_name="Light RUC net km", section="Key volumes", value=conventional_light_km, unit="million km", row_role="bridge_input", source_basis=light_split_basis, source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=light_split_source_cell, formula="current Light RUC total modelled km * MBU26 conventional Light share", official_value=off["light_ruc_net_km"], model_id=CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]),
                _current_annual_row(**common, series_id="heavy_ruc_net_km", display_name="Heavy RUC net km", section="Key volumes", value=heavy_km_million, unit="million km", row_role="bridge_input", source_basis="current_finalist_model", source_file="forecast_scenario_comparison.parquet", source_cell=heavy_source_cell, formula="sum quarterly current finalist Heavy RUC net km", official_value=off["heavy_ruc_net_km"]),
                _current_annual_row(**common, series_id="light_bev_ruc_net_km", display_name="Light BEV RUC net km", section="Key volumes", value=current_light_bev_km, unit="million km", row_role="bridge_input", source_basis=light_split_basis, source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=light_split_source_cell, formula="current Light RUC total modelled km * MBU26 Light BEV share", official_value=off["light_bev_ruc_net_km"], model_id=CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]),
                fixed("heavy_bev_ruc_net_km"),
                _current_annual_row(**common, series_id="phev_ruc_net_km", display_name="PHEV RUC net km", section="Key volumes", value=current_phev_km, unit="million km", row_role="bridge_input", source_basis=light_split_basis, source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=light_split_source_cell, formula="current Light RUC total modelled km * MBU26 PHEV share", official_value=off["phev_ruc_net_km"], model_id=CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]),
                fixed("tuc_gtk"),
                _current_annual_row(**common, series_id="gross_ped_revenue", display_name="PED revenue", section="FED", value=ped_revenue, unit="$m nominal ex GST", row_role="replacement_line", source_basis="current_finalist_model + MBU26 bridge", source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=ped_source_cell, formula="sum(quarterly PED VKT/capita * MBU26 population / 1,000,000 * MBU26 litres intensity / 100 * MBU26 PED rate)", replacement_only=True, official_value=off["gross_ped_revenue"]),
                _current_annual_row(**common, series_id="light_ruc_net_revenue", display_name="Light RUC revenue", section="RUC", value=light_revenue, unit="$m nominal ex GST", row_role="replacement_line", source_basis=light_split_revenue_basis, source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=light_split_source_cell, formula="current conventional Light RUC km * MBU26 conventional Light effective rate", replacement_only=True, official_value=off["light_ruc_net_revenue"], model_id=CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]),
                _current_annual_row(**common, series_id="heavy_ruc_net_revenue", display_name="Heavy RUC revenue", section="RUC", value=heavy_revenue, unit="$m nominal ex GST", row_role="replacement_line", source_basis="current_finalist_model + MBU26 effective rate", source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=heavy_source_cell, formula="sum quarterly current Heavy RUC net km * MBU26 effective rate", replacement_only=True, official_value=off["heavy_ruc_net_revenue"]),
                _current_annual_row(**common, series_id="light_bev_ruc_net_revenue", display_name="Light BEV RUC net revenue", section="RUC", value=light_bev_revenue, unit="$m nominal ex GST", row_role="replacement_line", source_basis=light_split_revenue_basis, source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=light_split_source_cell, formula="current Light BEV RUC km * MBU26 Light BEV effective rate", replacement_only=True, official_value=off["light_bev_ruc_net_revenue"], model_id=CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]),
                fixed("heavy_bev_ruc_net_revenue"),
                _current_annual_row(**common, series_id="phev_ruc_net_revenue", display_name="PHEV RUC net revenue", section="RUC", value=phev_revenue, unit="$m nominal ex GST", row_role="replacement_line", source_basis=light_split_revenue_basis, source_file="forecast_scenario_comparison.parquet; mbu26_official_annual.csv", source_cell=light_split_source_cell, formula="current PHEV RUC km * MBU26 PHEV effective rate", replacement_only=True, official_value=off["phev_ruc_net_revenue"], model_id=CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]),
                fixed("ruc_refunds"),
                _current_annual_row(**common, series_id="gross_ruc_revenue", display_name="Gross RUC revenue", section="RUC", value=gross_ruc, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="allocated light_ruc_net_revenue + heavy_ruc_net_revenue + allocated light_bev_ruc_net_revenue + MBU26 heavy_bev_ruc_net_revenue + allocated phev_ruc_net_revenue + MBU26 ruc_refunds", official_value=off["gross_ruc_revenue"]),
                fixed("ruc_admin_revenue"),
                _current_annual_row(**common, series_id="ruc_revenue_net_admin", display_name="RUC revenues net of admin fees", section="RUC", value=ruc_net_admin, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="gross_ruc_revenue - MBU26 ruc_admin_revenue", official_value=off["ruc_revenue_net_admin"]),
                _current_annual_row(**common, series_id="total_ruc_net_revenue", display_name="Total RUC all classes", section="RUC", value=total_ruc, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="ruc_revenue_net_admin - MBU26 ruc_refunds", official_value=off["total_ruc_net_revenue"]),
                fixed("gross_lpg_revenue"),
                fixed("gross_cng_revenue"),
                _current_annual_row(**common, series_id="gross_fed_revenue", display_name="Gross FED revenue", section="FED", value=gross_fed, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="gross_ped_revenue + MBU26 gross_lpg_revenue + MBU26 gross_cng_revenue", official_value=off["gross_fed_revenue"]),
                fixed("fed_refunds"),
                _current_annual_row(**common, series_id="net_fed_revenue", display_name="Net FED revenue", section="FED", value=net_fed, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="gross_fed_revenue - MBU26 fed_refunds", official_value=off["net_fed_revenue"]),
                fixed("mr1_revenue"),
                fixed("mr2_revenue"),
                fixed("coo_revenue", display_name="MR13"),
                _current_annual_row(**common, series_id="gross_mvr_revenue", display_name="Gross MVR revenue", section="MVR", value=gross_mvr, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="MBU26 official formula", source_file="mbu26_official_annual.csv", source_cell=_source_cell_for_series(official_row, "gross_mvr_revenue"), formula="mr1_revenue + mr2_revenue + coo_revenue", official_value=off["gross_mvr_revenue"]),
                fixed("mvr_admin_revenue"),
                _current_annual_row(**common, series_id="mvr_revenue_net_admin_coo", display_name="MVR revenues net of admin fees and COO", section="MVR", value=mvr_net_admin_coo, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="MBU26 official formula", source_file="mbu26_official_annual.csv", source_cell=_source_cell_for_series(official_row, "mvr_revenue_net_admin_coo"), formula="mr1_revenue + mr2_revenue - mvr_admin_revenue", official_value=off["mvr_revenue_net_admin_coo"]),
                fixed("mvr_refunds"),
                _current_annual_row(**common, series_id="net_mvr_revenue", display_name="Net MVR revenue", section="MVR", value=net_mvr, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="MBU26 official formula", source_file="mbu26_official_annual.csv", source_cell=_source_cell_for_series(official_row, "net_mvr_revenue"), formula="mvr_revenue_net_admin_coo - mvr_refunds", official_value=off["net_mvr_revenue"]),
                fixed("tuc_net_revenue"),
                _current_annual_row(**common, series_id="total_gross_revenue", display_name="Total gross revenues", section="Totals", value=total_gross, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="gross_ruc_revenue + gross_fed_revenue + gross_mvr_revenue + tuc_net_revenue", official_value=off["total_gross_revenue"]),
                _current_annual_row(**common, series_id="total_admin_fees", display_name="Total admin fees", section="Totals", value=total_admin, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="ruc_admin_revenue + mvr_admin_revenue + coo_revenue", official_value=off["total_admin_fees"]),
                _current_annual_row(**common, series_id="total_revenue_net_admin", display_name="Total revenues net of admin fees", section="Totals", value=total_net_admin, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="total_gross_revenue - total_admin_fees", official_value=off["total_revenue_net_admin"]),
                _current_annual_row(**common, series_id="total_refunds", display_name="Total refunds", section="Totals", value=total_refunds, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="ruc_refunds + fed_refunds + mvr_refunds", official_value=off["total_refunds"]),
                _current_annual_row(**common, series_id="total_nltf_net_revenue", display_name="Total NLTF revenue", section="Totals", value=total_nltf, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="total_revenue_net_admin - total_refunds", official_value=off["total_nltf_net_revenue"]),
                _current_annual_row(**common, series_id="total_fed_ruc_net_revenue", display_name="Total RUC+PED revenue", section="Derived totals", value=total_fed_ruc, unit="$m nominal ex GST", row_role="calculated_rollup", source_basis="current_hybrid_formula", source_file="current_hybrid_formula", source_cell=current_source_cell, formula="net_fed_revenue + total_ruc_net_revenue", official_value=off.get("total_fed_ruc_net_revenue")),
            ]
        )
    if not rows:
        return pd.DataFrame(columns=columns)
    out = pd.DataFrame(rows)
    anchor_lookup = _source_actual_anchor_lookup(mbu26_official_annual)
    anchor_mask = (
        pd.to_numeric(out.get("FY"), errors="coerce").eq(REVENUE_LAST_COMPLETE_ACTUAL_FY)
        & out.get("series_id", pd.Series("", index=out.index)).astype(str).isin(anchor_lookup)
    )
    for idx, row in out.loc[anchor_mask].iterrows():
        actual = anchor_lookup.get(str(row.get("series_id", "")), {})
        actual_value = pd.to_numeric(pd.Series([actual.get("value")]), errors="coerce").iloc[0]
        if pd.isna(actual_value):
            continue
        out.at[idx, "value"] = actual_value
        out.at[idx, "unit"] = actual.get("unit") or out.at[idx, "unit"]
        out.at[idx, "source_basis"] = "MBU26 actual anchor"
        out.at[idx, "source_file"] = actual.get("source_file", "mbu26_official_annual.csv")
        out.at[idx, "source_cell"] = actual.get("source_cell", "")
        out.at[idx, "source_status"] = actual.get("period_status", "ACTUAL")
        out.at[idx, "formula"] = actual.get("formula", "MBU26 complete annual actual anchor")
        out.at[idx, "value_status"] = "Actual anchor"
        out.at[idx, "quarters_present"] = "; ".join(_expected_june_year_quarters(REVENUE_LAST_COMPLETE_ACTUAL_FY))
        out.at[idx, "actual_quarters"] = "; ".join(_expected_june_year_quarters(REVENUE_LAST_COMPLETE_ACTUAL_FY))
        out.at[idx, "forecast_quarters"] = ""
        out.at[idx, "nowcast_flag"] = False
        out.at[idx, "official_value"] = actual_value
        out.at[idx, "residual_vs_official"] = 0.0
    for column in columns:
        if column not in out.columns:
            out[column] = pd.NA if column in {"official_value", "residual_vs_official"} else ""
    out["_line_order"] = out["series_id"].map(_line_order)
    out = out.sort_values(["FY", "scenario_name", "fed_path", "_line_order", "series_id"], kind="stable")
    return out[columns].reset_index(drop=True)


def _extract_annual_spine(values_sheet: Any, formulas_sheet: Any, workbook_name: str, workbook_hash: str) -> pd.DataFrame:
    year_columns = []
    for column in range(1, values_sheet.max_column + 1):
        year = _as_int(values_sheet.cell(6, column).value)
        if year is None:
            continue
        year_columns.append(
            {
                "column": column,
                "fy": year,
                "period_status": str(values_sheet.cell(7, column).value or "").strip(),
                "source_year_cell": values_sheet.cell(6, column).coordinate,
                "source_status_cell": values_sheet.cell(7, column).coordinate,
            }
        )
    formula_by_output = {item["output_series_id"]: item["expression"] for item in FORMULA_DEFINITIONS}
    rows: list[dict[str, Any]] = []
    for definition in ROW_DEFINITIONS:
        source_row = int(definition["row"])
        runtime_series_id = str(definition["series_id"])
        source_series_id = str(definition.get("source_series_id") or runtime_series_id)
        dashboard_label = str(definition["display_name"])
        source_display_name = str(definition.get("source_display_name") or dashboard_label)
        workbook_label = str(values_sheet.cell(source_row, 2).value or "").strip()
        for year_info in year_columns:
            column = int(year_info["column"])
            value_cell = values_sheet.cell(source_row, column)
            formula_cell = formulas_sheet.cell(source_row, column)
            value = pd.to_numeric(pd.Series([value_cell.value]), errors="coerce").iloc[0]
            source_formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else ""
            rows.append(
                {
                    "source_release": MBU26_RELEASE_ROUND,
                    "workbook_basename": workbook_name,
                    "workbook_sha256": workbook_hash,
                    "sheet": MBU26_SHEET_NAME,
                    "source_row": source_row,
                    "source_cell": value_cell.coordinate,
                    "source_year_cell": year_info["source_year_cell"],
                    "source_status_cell": year_info["source_status_cell"],
                    "source_formula": source_formula,
                    "asserted_formula": formula_by_output.get(runtime_series_id, ""),
                    "FY": int(year_info["fy"]),
                    "period": f"FY{int(year_info['fy'])}",
                    "period_status": year_info["period_status"],
                    "source_label": workbook_label,
                    "source_series_id": source_series_id,
                    "runtime_series_id": runtime_series_id,
                    "dashboard_label": dashboard_label,
                    "source_display_name": source_display_name,
                    "label": dashboard_label,
                    "series_id": runtime_series_id,
                    "display_name": dashboard_label,
                    "section": definition["section"],
                    "unit": definition["unit"],
                    "metric_type": definition["metric_type"],
                    "row_role": definition["row_role"],
                    "value": value if pd.notna(value) else pd.NA,
                    "value_status": "actual" if str(year_info["period_status"]).upper() == "ACTUAL" else "official_forecast",
                    "source_kind": "mbu26_source_row",
                }
            )
    return pd.DataFrame(rows).sort_values(["source_row", "FY"], kind="stable").reset_index(drop=True)


def _formula_audit_frame(spine: pd.DataFrame) -> pd.DataFrame:
    values = {
        (str(row.series_id), int(row.FY)): pd.to_numeric(row.value, errors="coerce")
        for row in spine.itertuples()
        if pd.notna(row.value)
    }
    labels = {str(row["series_id"]): str(row["display_name"]) for row in ROW_DEFINITIONS}
    source_rows = {str(row["series_id"]): str(row["row"]) for row in ROW_DEFINITIONS}
    source_cells = {
        (str(row.series_id), int(row.FY)): str(row.source_cell)
        for row in spine.itertuples()
    }
    statuses = {
        int(row.FY): str(row.period_status)
        for row in spine[["FY", "period_status"]].drop_duplicates().itertuples(index=False)
    }
    fys = sorted({int(value) for value in spine["FY"].dropna().unique()})
    rows: list[dict[str, Any]] = []
    for formula in FORMULA_DEFINITIONS:
        output = str(formula["output_series_id"])
        terms = tuple(formula["terms"])
        for fy in fys:
            missing = [series_id for series_id, _sign in terms if (series_id, fy) not in values]
            calculated = pd.NA if missing else sum(float(values[(series_id, fy)]) * sign for series_id, sign in terms)
            observed = values.get((output, fy), calculated if str(formula.get("source_kind", "")) == "derived_dashboard_subtotal" else pd.NA)
            residual = pd.NA
            if pd.notna(calculated) and pd.notna(observed):
                residual = float(observed) - float(calculated)
            status = "missing_inputs" if missing else "reconciled" if pd.notna(residual) and abs(float(residual)) <= 0.05 else "residual_reported"
            rows.append(
                {
                    "source_release": MBU26_RELEASE_ROUND,
                    "FY": fy,
                    "period": f"FY{fy}",
                    "period_status": statuses.get(fy, ""),
                    "output_series_id": output,
                    "output_label": str(formula.get("output_label") or labels.get(output, output)),
                    "row_role": "aggregate",
                    "formula": str(formula["expression"]),
                    "source_rows": "; ".join(source_rows.get(series_id, "") for series_id, _sign in terms),
                    "source_cells": "; ".join(source_cells.get((series_id, fy), "") for series_id, _sign in terms),
                    "observed_value": observed,
                    "calculated_value": calculated,
                    "residual": residual,
                    "residual_abs": abs(float(residual)) if pd.notna(residual) else pd.NA,
                    "status": status,
                    "missing_inputs": "; ".join(missing),
                    "source_kind": str(formula.get("source_kind") or "mbu26_source_formula_assertion"),
                }
            )
    return pd.DataFrame(rows).sort_values(["FY", "output_series_id"], kind="stable").reset_index(drop=True)


def _official_annual_frame(spine: pd.DataFrame, formula_audit: pd.DataFrame) -> pd.DataFrame:
    base = spine.copy()
    base["source_file"] = "mbu26_annual_spine.csv"
    base["formula"] = base["asserted_formula"].where(base["asserted_formula"].astype(str).ne(""), base["source_formula"])
    subtotal = formula_audit[formula_audit["output_series_id"].eq("total_fed_ruc_net_revenue")].copy()
    if not subtotal.empty:
        subtotal_rows = pd.DataFrame(
            {
                "source_release": MBU26_RELEASE_ROUND,
                "workbook_basename": base["workbook_basename"].iloc[0] if not base.empty else "",
                "workbook_sha256": base["workbook_sha256"].iloc[0] if not base.empty else "",
                "sheet": MBU26_SHEET_NAME,
                "source_row": "",
                "source_cell": subtotal["source_cells"],
                "source_year_cell": "",
                "source_status_cell": "",
                "source_formula": "",
                "asserted_formula": subtotal["formula"],
                "FY": subtotal["FY"],
                "period": subtotal["period"],
                "period_status": subtotal["period_status"],
                "source_label": "Derived dashboard subtotal from MBU26 Net FED and Total RUC",
                "label": "Total RUC+PED revenue",
                "series_id": "total_fed_ruc_net_revenue",
                "display_name": "Total RUC+PED revenue",
                "section": "Derived totals",
                "unit": "$m nominal ex GST",
                "metric_type": "revenue",
                "row_role": "aggregate",
                "value": subtotal["calculated_value"],
                "value_status": subtotal["period_status"].map(lambda value: "actual" if str(value).upper() == "ACTUAL" else "official_forecast"),
                "source_kind": "mbu26_formula_derived_dashboard_subtotal",
                "source_file": "row_reconciliation.csv",
                "formula": subtotal["formula"],
            }
        )
        base = pd.concat([base, subtotal_rows], ignore_index=True, sort=False)
    return base.sort_values(["series_id", "FY"], kind="stable").reset_index(drop=True)


def _current_activity_annual_values(chart_rows: pd.DataFrame) -> pd.DataFrame:
    columns = ["FY", "scenario_name", "scenario_role", "series_id", "value", "unit", "model_id", "quarters_present", "actual_quarters", "forecast_quarters", "nowcast_flag", "value_status"]
    if chart_rows is None or chart_rows.empty:
        return pd.DataFrame(columns=columns)
    data = chart_rows.copy()
    data = data[
        data.get("time_grain", pd.Series(dtype=str)).astype(str).eq("quarterly")
        & data.get("metric_type", pd.Series(dtype=str)).astype(str).eq("activity")
        & data.get("period", pd.Series(dtype=str)).astype(str).str.match(r"^\d{4}Q[1-4]$", na=False)
    ].copy()
    data["value_numeric"] = pd.to_numeric(data.get("value"), errors="coerce")
    data = data[data["value_numeric"].notna()].copy()
    if data.empty:
        return pd.DataFrame(columns=columns)
    stream_to_series = {"PED": "ped_vkt_per_capita", "LIGHT_RUC": "light_ruc_net_km", "HEAVY_RUC": "heavy_ruc_net_km"}
    data["series_id"] = data.get("stream", pd.Series("", index=data.index)).astype(str).map(stream_to_series)
    data = data[data["series_id"].notna()].copy()
    future = data[data["row_type"].astype(str).eq("future_forecast")].copy()
    historical = data[data["row_type"].astype(str).eq("historical_actual")].copy()
    if future.empty:
        return pd.DataFrame(columns=columns)
    hist_lookup = {(str(row.series_id), str(row.period)): row for row in historical.itertuples()}
    rows: list[dict[str, Any]] = []
    for (scenario_name, series_id), group in future.groupby(["scenario_name", "series_id"], dropna=False):
        scenario_role = _first_text(group, "scenario_role")
        future_lookup = {str(row.period): row for row in group.itertuples()}
        model_id = CURRENT_FINALIST_MODEL_IDS.get(_current_stream_for_series(str(series_id)), "")
        fys = sorted({_june_year_from_quarter(str(period)) for period in list(future_lookup) if _june_year_from_quarter(str(period)) is not None})
        for fy in fys:
            expected = _expected_june_year_quarters(int(fy))
            values: list[float] = []
            actual_quarters: list[str] = []
            forecast_quarters: list[str] = []
            for quarter in expected:
                row = future_lookup.get(quarter)
                if row is not None:
                    values.append(float(row.value_numeric))
                    forecast_quarters.append(quarter)
                    continue
                hist_row = hist_lookup.get((str(series_id), quarter))
                if hist_row is not None:
                    values.append(float(hist_row.value_numeric))
                    actual_quarters.append(quarter)
            if len(values) != 4:
                continue
            if str(series_id) == "ped_vkt_per_capita":
                annual_value = sum(values)
                unit = "km/person"
            else:
                annual_value = sum(values)
                unit = "million km"
                if abs(annual_value) > 10_000_000:
                    annual_value /= 1_000_000.0
            rows.append(
                {
                    "FY": int(fy),
                    "scenario_name": str(scenario_name),
                    "scenario_role": scenario_role,
                    "series_id": str(series_id),
                    "value": annual_value,
                    "unit": unit,
                    "model_id": model_id,
                    "quarters_present": "; ".join(expected),
                    "actual_quarters": "; ".join(actual_quarters),
                    "forecast_quarters": "; ".join(forecast_quarters),
                    "nowcast_flag": bool(actual_quarters and forecast_quarters),
                    "value_status": "current_finalist_forecast",
                }
            )
        anchor_expected = _expected_june_year_quarters(REVENUE_LAST_COMPLETE_ACTUAL_FY)
        anchor_values = []
        for quarter in anchor_expected:
            hist_row = hist_lookup.get((str(series_id), quarter))
            if hist_row is not None:
                anchor_values.append(float(hist_row.value_numeric))
        if len(anchor_values) == 4:
            anchor_value = sum(anchor_values) / 4.0 if str(series_id) == "ped_vkt_per_capita" else sum(anchor_values)
            if str(series_id) != "ped_vkt_per_capita" and abs(anchor_value) > 10_000_000:
                anchor_value /= 1_000_000.0
            rows.append(
                {
                    "FY": REVENUE_LAST_COMPLETE_ACTUAL_FY,
                    "scenario_name": str(scenario_name),
                    "scenario_role": scenario_role,
                    "series_id": str(series_id),
                    "value": anchor_value,
                    "unit": "km/person" if str(series_id) == "ped_vkt_per_capita" else "million km",
                    "model_id": model_id,
                    "quarters_present": "; ".join(anchor_expected),
                    "actual_quarters": "; ".join(anchor_expected),
                    "forecast_quarters": "",
                    "nowcast_flag": False,
                    "value_status": "Actual anchor",
                }
            )
    if not rows:
        return pd.DataFrame(columns=columns)
    out = pd.DataFrame(rows, columns=columns)
    out = _append_extrapolated_activity_extensions(out)
    return out.sort_values(["FY", "scenario_name", "series_id"], kind="stable").reset_index(drop=True)


def _append_extrapolated_activity_extensions(activity: pd.DataFrame) -> pd.DataFrame:
    if activity is None or activity.empty:
        return activity
    rows: list[dict[str, Any]] = []
    data = activity.copy()
    data["FY_numeric"] = pd.to_numeric(data.get("FY"), errors="coerce").astype("Int64")
    data["value_numeric"] = pd.to_numeric(data.get("value"), errors="coerce")
    for (scenario_name, series_id), group in data.groupby(["scenario_name", "series_id"], dropna=False):
        base = group[
            group["FY_numeric"].between(CURRENT_MODEL_EXTENSION_BASE_START_FY, CURRENT_MODEL_EXTENSION_BASE_END_FY, inclusive="both")
            & group["value_numeric"].notna()
        ].copy()
        if len(base.drop_duplicates("FY_numeric")) < (CURRENT_MODEL_EXTENSION_BASE_END_FY - CURRENT_MODEL_EXTENSION_BASE_START_FY + 1):
            continue
        by_year = base.sort_values("FY_numeric", kind="stable").drop_duplicates("FY_numeric", keep="last").set_index("FY_numeric")
        start_value = float(by_year.loc[CURRENT_MODEL_EXTENSION_BASE_START_FY, "value_numeric"])
        end_value = float(by_year.loc[CURRENT_MODEL_EXTENSION_BASE_END_FY, "value_numeric"])
        annual_slope = (end_value - start_value) / float(CURRENT_MODEL_EXTENSION_BASE_END_FY - CURRENT_MODEL_EXTENSION_BASE_START_FY)
        template = by_year.loc[CURRENT_MODEL_EXTENSION_BASE_END_FY].to_dict()
        for fy in range(CURRENT_MODEL_EXTENSION_START_FY, CURRENT_MODEL_EXTENSION_END_FY + 1):
            if group["FY_numeric"].eq(fy).any():
                continue
            value = end_value + annual_slope * float(fy - CURRENT_MODEL_EXTENSION_BASE_END_FY)
            rows.append(
                {
                    "FY": fy,
                    "scenario_name": str(scenario_name),
                    "scenario_role": str(template.get("scenario_role") or ""),
                    "series_id": str(series_id),
                    "value": value,
                    "unit": str(template.get("unit") or ("km/person" if str(series_id) == "ped_vkt_per_capita" else "million km")),
                    "model_id": str(template.get("model_id") or CURRENT_FINALIST_MODEL_IDS.get(_current_stream_for_series(str(series_id)), "")),
                    "quarters_present": f"FY{fy} extrapolated_model_extension",
                    "actual_quarters": "",
                    "forecast_quarters": f"FY{fy} extrapolated from FY{CURRENT_MODEL_EXTENSION_BASE_START_FY}-FY{CURRENT_MODEL_EXTENSION_BASE_END_FY} annual current-finalist gradient",
                    "nowcast_flag": False,
                    "value_status": "extrapolated_model_extension",
                }
            )
    if not rows:
        return activity
    return pd.concat([activity, pd.DataFrame(rows)], ignore_index=True, sort=False)


def _is_extrapolated_activity_row(row: Any) -> bool:
    return str(getattr(row, "value_status", "") or "").strip() == "extrapolated_model_extension"


def _extrapolated_activity_quarter_payload(row: Any, series_id: str) -> dict[str, Any]:
    annual_value = pd.to_numeric(pd.Series([getattr(row, "value", pd.NA)]), errors="coerce").iloc[0]
    if pd.isna(annual_value):
        return {}
    fy = int(getattr(row, "FY"))
    scenario_name = str(getattr(row, "scenario_name", "") or "")
    per_quarter = float(annual_value) / 4.0
    return {
        "values": [per_quarter, per_quarter, per_quarter, per_quarter],
        "quarters_present": f"FY{fy} extrapolated_model_extension",
        "actual_quarters": "",
        "forecast_quarters": f"FY{fy} extrapolated from FY{CURRENT_MODEL_EXTENSION_BASE_START_FY}-FY{CURRENT_MODEL_EXTENSION_BASE_END_FY} annual current-finalist gradient",
        "source_cells": f"extrapolated_model_extension:{scenario_name}:FY{fy}:{series_id}",
    }


def _light_ev_phev_split_from_official(off: dict[str, Any]) -> dict[str, Any]:
    required = [
        "light_ruc_net_km",
        "light_bev_ruc_net_km",
        "phev_ruc_net_km",
        "light_ruc_net_revenue",
        "light_bev_ruc_net_revenue",
        "phev_ruc_net_revenue",
    ]
    values = {key: pd.to_numeric(pd.Series([off.get(key)]), errors="coerce").iloc[0] for key in required}
    if any(pd.isna(value) for value in values.values()):
        return {"availability_status": "missing_required_mbu26_split_or_rate_input"}
    conventional = float(values["light_ruc_net_km"])
    light_bev = float(values["light_bev_ruc_net_km"])
    phev = float(values["phev_ruc_net_km"])
    denominator = conventional + light_bev + phev
    if denominator <= 0:
        return {"availability_status": "invalid_nonpositive_light_universe_denominator"}
    conventional_rate = _rate_or_zero(values["light_ruc_net_revenue"], conventional)
    light_bev_rate = _rate_or_zero(values["light_bev_ruc_net_revenue"], light_bev)
    phev_rate = _rate_or_zero(values["phev_ruc_net_revenue"], phev)
    if conventional > 0 and conventional_rate is None:
        return {"availability_status": "missing_conventional_light_rate"}
    if light_bev > 0 and light_bev_rate is None:
        return {"availability_status": "missing_light_bev_rate"}
    if phev > 0 and phev_rate is None:
        return {"availability_status": "missing_phev_rate"}
    conventional_share = conventional / denominator
    light_bev_share = light_bev / denominator
    phev_share = phev / denominator
    share_sum = conventional_share + light_bev_share + phev_share
    if abs(share_sum - 1.0) > 1e-9:
        return {"availability_status": "share_sum_failed"}
    return {
        "availability_status": "available",
        "conventional_share": conventional_share,
        "light_bev_share": light_bev_share,
        "phev_share": phev_share,
        "share_sum": share_sum,
        "total_light_universe_km": denominator,
        "conventional_rate": float(conventional_rate or 0.0),
        "light_bev_rate": float(light_bev_rate or 0.0),
        "phev_rate": float(phev_rate or 0.0),
    }


def _rate_or_zero(revenue: Any, km: float) -> float | None:
    if abs(float(km)) < 1e-12:
        return 0.0
    numeric = pd.to_numeric(pd.Series([revenue]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return None
    return float(numeric) / float(km)


def _current_activity_quarter_lookup(chart_rows: pd.DataFrame) -> dict[tuple[str, int, str], dict[str, Any]]:
    if chart_rows is None or chart_rows.empty:
        return {}
    data = chart_rows.copy()
    data = data[
        data.get("time_grain", pd.Series(dtype=str)).astype(str).eq("quarterly")
        & data.get("metric_type", pd.Series(dtype=str)).astype(str).eq("activity")
        & data.get("period", pd.Series(dtype=str)).astype(str).str.match(r"^\d{4}Q[1-4]$", na=False)
    ].copy()
    data["value_numeric"] = pd.to_numeric(data.get("value"), errors="coerce")
    data = data[data["value_numeric"].notna()].copy()
    if data.empty:
        return {}
    stream_to_series = {"PED": "ped_vkt_per_capita", "LIGHT_RUC": "light_ruc_net_km", "HEAVY_RUC": "heavy_ruc_net_km"}
    data["series_id"] = data.get("stream", pd.Series("", index=data.index)).astype(str).map(stream_to_series)
    data = data[data["series_id"].notna()].copy()
    future = data[data["row_type"].astype(str).eq("future_forecast")].copy()
    historical = data[data["row_type"].astype(str).eq("historical_actual")].copy()
    hist_lookup = {(str(row.series_id), str(row.period)): row for row in historical.itertuples()}
    output: dict[tuple[str, int, str], dict[str, Any]] = {}
    for (scenario_name, series_id), group in future.groupby(["scenario_name", "series_id"], dropna=False):
        future_lookup = {str(row.period): row for row in group.itertuples()}
        fys = sorted({_june_year_from_quarter(str(period)) for period in list(future_lookup) if _june_year_from_quarter(str(period)) is not None})
        fys.append(REVENUE_LAST_COMPLETE_ACTUAL_FY)
        for fy in sorted(set(int(value) for value in fys if value is not None)):
            expected = _expected_june_year_quarters(fy)
            values: list[float] = []
            actual_quarters: list[str] = []
            forecast_quarters: list[str] = []
            source_cells: list[str] = []
            for quarter in expected:
                row = future_lookup.get(quarter)
                if row is not None:
                    values.append(float(row.value_numeric))
                    forecast_quarters.append(quarter)
                    source_cells.append(str(getattr(row, "source_cell", "") or f"data/current_revenue_outlook/revenue_chart_rows.csv:{scenario_name}:{quarter}"))
                    continue
                hist_row = hist_lookup.get((str(series_id), quarter))
                if hist_row is not None:
                    values.append(float(hist_row.value_numeric))
                    actual_quarters.append(quarter)
                    source_cells.append(str(getattr(hist_row, "source_cell", "") or f"data/current_revenue_outlook/revenue_chart_rows.csv:historical_actual:{quarter}"))
            if len(values) != 4:
                continue
            output[(str(scenario_name), fy, str(series_id))] = {
                "values": values,
                "quarters_present": "; ".join(expected),
                "actual_quarters": "; ".join(actual_quarters),
                "forecast_quarters": "; ".join(forecast_quarters),
                "source_cells": "; ".join(source_cells),
            }
    return output


def _official_lookup(official: pd.DataFrame) -> dict[int, dict[str, Any]]:
    if official is None or official.empty:
        return {}
    data = official.copy()
    data["FY_int"] = pd.to_numeric(data.get("FY"), errors="coerce").astype("Int64")
    data["value_numeric"] = pd.to_numeric(data.get("value"), errors="coerce")
    data = data[data["FY_int"].notna() & data["value_numeric"].notna()].copy()
    output: dict[int, dict[str, Any]] = {}
    for fy, group in data.groupby("FY_int", dropna=True):
        output[int(fy)] = group.groupby("series_id")["value_numeric"].first().to_dict()
    return output


def _official_record_lookup(official: pd.DataFrame) -> dict[int, dict[str, dict[str, Any]]]:
    if official is None or official.empty:
        return {}
    data = official.copy()
    data["FY_int"] = pd.to_numeric(data.get("FY"), errors="coerce").astype("Int64")
    data["value_numeric"] = pd.to_numeric(data.get("value"), errors="coerce")
    data = data[data["FY_int"].notna()].copy()
    output: dict[int, dict[str, dict[str, Any]]] = {}
    for fy, group in data.groupby("FY_int", dropna=True):
        fy_records: dict[str, dict[str, Any]] = {}
        for row in group.to_dict("records"):
            series_id = str(row.get("series_id") or "")
            if series_id and series_id not in fy_records:
                fy_records[series_id] = row
        output[int(fy)] = fy_records
    return output


def _source_actual_anchor_lookup(official: pd.DataFrame) -> dict[str, dict[str, Any]]:
    rows = official[
        pd.to_numeric(official.get("FY"), errors="coerce").eq(REVENUE_LAST_COMPLETE_ACTUAL_FY)
        & official.get("period_status", pd.Series("", index=official.index)).astype(str).eq("ACTUAL")
    ].copy()
    rows["value_numeric"] = pd.to_numeric(rows.get("value"), errors="coerce")
    rows = rows[rows["value_numeric"].notna()].copy()
    output: dict[str, dict[str, Any]] = {}
    for series_id, group in rows.groupby(rows.get("series_id", pd.Series("", index=rows.index)).astype(str), dropna=False):
        row = group.iloc[0]
        output[str(series_id)] = {
            "value": row.get("value_numeric"),
            "unit": row.get("unit"),
            "source_file": row.get("source_file", "mbu26_official_annual.csv"),
            "source_cell": row.get("source_cell", ""),
            "period_status": row.get("period_status", ""),
            "formula": row.get("formula", ""),
        }
    return output


def _current_annual_row(
    *,
    FY: int,
    period: str,
    source_path: str,
    fed_path: str,
    scenario_name: str,
    scenario_role: str,
    series_id: str,
    display_name: str,
    section: str,
    value: Any,
    unit: str,
    row_role: str,
    source_basis: str,
    source_file: str,
    source_cell: str,
    value_status: str,
    quarters_present: str,
    actual_quarters: str,
    forecast_quarters: str,
    nowcast_flag: bool,
    formula: str = "",
    replacement_only: bool = False,
    official_value: Any = pd.NA,
    availability_status: str = "available",
    model_id: str = "",
) -> dict[str, Any]:
    numeric_value = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    numeric_official = pd.to_numeric(pd.Series([official_value]), errors="coerce").iloc[0]
    residual = float(numeric_value) - float(numeric_official) if pd.notna(numeric_value) and pd.notna(numeric_official) else pd.NA
    return {
        "FY": int(FY),
        "period": period,
        "source_path": source_path,
        "fed_path": fed_path,
        "scenario_name": scenario_name,
        "scenario_role": scenario_role,
        "series_id": series_id,
        "display_name": display_name,
        "section": section,
        "value": numeric_value,
        "unit": unit,
        "row_role": row_role,
        "source_basis": source_basis,
        "source_file": source_file,
        "source_cell": source_cell,
        "source_status": "source_backed",
        "model_id": model_id or _current_model_id_for_series(series_id),
        "formula": formula,
        "replacement_only": bool(replacement_only),
        "official_value": numeric_official if pd.notna(numeric_official) else pd.NA,
        "residual_vs_official": residual,
        "availability_status": availability_status,
        "line": "Model path",
        "model_basis": "current_finalist_model",
        "forecast_path": "current_finalist_model",
        "path_status": "current_model_forecast",
        "value_status": value_status,
        "quarters_present": quarters_present,
        "actual_quarters": actual_quarters,
        "forecast_quarters": forecast_quarters,
        "nowcast_flag": bool(nowcast_flag),
    }


def _current_fixed_row(
    *,
    common: dict[str, Any],
    official_row: dict[str, dict[str, Any]],
    series_id: str,
    display_name: str | None = None,
    row_role: str | None = None,
) -> dict[str, Any]:
    record = official_row.get(series_id, {})
    return _current_annual_row(
        **common,
        series_id=series_id,
        display_name=str(display_name or record.get("display_name") or record.get("label") or series_id),
        section=str(record.get("section") or ""),
        value=record.get("value_numeric", record.get("value")),
        unit=str(record.get("unit") or ""),
        row_role=str(row_role or record.get("row_role") or "fixed_mbu26_component"),
        source_basis="MBU26 official fixed component",
        source_file=str(record.get("source_file") or "mbu26_official_annual.csv"),
        source_cell=str(record.get("source_cell") or ""),
        formula=str(record.get("formula") or ""),
        official_value=record.get("value_numeric", record.get("value")),
    )


def _source_cell_for_series(official_row: dict[str, dict[str, Any]], series_id: str) -> str:
    return str(official_row.get(series_id, {}).get("source_cell") or "")


def _line_order(series_id: Any) -> int:
    text = str(series_id or "")
    try:
        return REVENUE_LINE_TABLE_SERIES_IDS.index(text)
    except ValueError:
        if text == "total_fed_ruc_net_revenue":
            return len(REVENUE_LINE_TABLE_SERIES_IDS)
        return len(REVENUE_LINE_TABLE_SERIES_IDS) + 1


def _current_source_path_label(scenario_name: str, scenario_role: str = "") -> str:
    if str(scenario_name) == "current_basecase" or str(scenario_role).lower() == "basecase":
        return "Current finalist Base case"
    return "Current finalist High population/comparison"


def revenue_line_reconciliation_frame(
    *,
    mbu26_official_annual: pd.DataFrame,
    current_forecast_annual: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "source_path",
        "FY",
        "period",
        "section",
        "line_label",
        "series_id",
        "value",
        "unit",
        "row_role",
        "source_file",
        "source_cell",
        "formula",
        "source_status",
        "source_basis",
        "model_id",
        "quarter_composition",
        "actual_quarters",
        "forecast_quarters",
        "replacement_flag",
        "value_status",
        "residual_vs_official",
        "scenario_name",
        "scenario_role",
        "fed_path",
        "availability_status",
    ]
    rows: list[dict[str, Any]] = []
    if isinstance(mbu26_official_annual, pd.DataFrame) and not mbu26_official_annual.empty:
        official = mbu26_official_annual[
            mbu26_official_annual.get("series_id", pd.Series(dtype=str)).astype(str).isin(REVENUE_LINE_TABLE_SERIES_IDS)
        ].copy()
        official["FY_numeric"] = pd.to_numeric(official.get("FY"), errors="coerce").astype("Int64")
        official["value_numeric"] = pd.to_numeric(official.get("value"), errors="coerce")
        official = official[official["FY_numeric"].notna() & official["value_numeric"].notna()].copy()
        for row in official.to_dict("records"):
            fy = int(row.get("FY_numeric"))
            quarters = "; ".join(_expected_june_year_quarters(fy))
            rows.append(
                {
                    "source_path": "MBU26 official",
                    "FY": fy,
                    "period": f"FY{fy}",
                    "section": row.get("section", ""),
                    "line_label": _line_display_label(row.get("series_id"), row.get("display_name") or row.get("label")),
                    "series_id": row.get("series_id", ""),
                    "value": row.get("value_numeric"),
                    "unit": row.get("unit", ""),
                    "row_role": row.get("row_role", ""),
                    "source_file": row.get("source_file", "mbu26_official_annual.csv"),
                    "source_cell": row.get("source_cell", ""),
                    "formula": row.get("formula", ""),
                    "source_status": row.get("period_status", ""),
                    "source_basis": "MBU26 official source row",
                    "model_id": "",
                    "quarter_composition": quarters,
                    "actual_quarters": quarters if str(row.get("period_status", "")).upper() == "ACTUAL" else "",
                    "forecast_quarters": "" if str(row.get("period_status", "")).upper() == "ACTUAL" else quarters,
                    "replacement_flag": False,
                    "value_status": row.get("value_status", ""),
                    "residual_vs_official": 0.0,
                    "scenario_name": "mbu26_official",
                    "scenario_role": "official_comparator",
                    "fed_path": MBU26_RELEASE_ROUND,
                    "availability_status": "available",
                }
            )
    if isinstance(current_forecast_annual, pd.DataFrame) and not current_forecast_annual.empty:
        current = current_forecast_annual[
            current_forecast_annual.get("series_id", pd.Series(dtype=str)).astype(str).isin(REVENUE_LINE_TABLE_SERIES_IDS)
        ].copy()
        current["FY_numeric"] = pd.to_numeric(current.get("FY"), errors="coerce").astype("Int64")
        current["value_numeric"] = pd.to_numeric(current.get("value"), errors="coerce")
        current = current[current["FY_numeric"].notna() & current["value_numeric"].notna()].copy()
        for row in current.to_dict("records"):
            fy = int(row.get("FY_numeric"))
            rows.append(
                {
                    "source_path": row.get("source_path") or _current_source_path_label(str(row.get("scenario_name", "")), str(row.get("scenario_role", ""))),
                    "FY": fy,
                    "period": f"FY{fy}",
                    "section": row.get("section", ""),
                    "line_label": _line_display_label(row.get("series_id"), row.get("display_name")),
                    "series_id": row.get("series_id", ""),
                    "value": row.get("value_numeric"),
                    "unit": row.get("unit", ""),
                    "row_role": row.get("row_role", ""),
                    "source_file": row.get("source_file", ""),
                    "source_cell": row.get("source_cell", ""),
                    "formula": row.get("formula", ""),
                    "source_status": row.get("source_status", ""),
                    "source_basis": row.get("source_basis", ""),
                    "model_id": row.get("model_id", ""),
                    "quarter_composition": row.get("quarters_present", ""),
                    "actual_quarters": row.get("actual_quarters", ""),
                    "forecast_quarters": row.get("forecast_quarters", ""),
                    "replacement_flag": bool(row.get("replacement_only", False)),
                    "value_status": row.get("value_status", ""),
                    "residual_vs_official": row.get("residual_vs_official", pd.NA),
                    "scenario_name": row.get("scenario_name", ""),
                    "scenario_role": row.get("scenario_role", ""),
                    "fed_path": row.get("fed_path", ""),
                    "availability_status": row.get("availability_status", "available"),
                }
            )
    if not rows:
        return pd.DataFrame(columns=columns)
    out = pd.DataFrame(rows)
    out["_source_order"] = out["source_path"].map(lambda value: REVENUE_LINE_SOURCE_PATHS.index(value) if value in REVENUE_LINE_SOURCE_PATHS else len(REVENUE_LINE_SOURCE_PATHS))
    out["_line_order"] = out["series_id"].map(_line_order)
    return out[columns + ["_source_order", "_line_order"]].sort_values(["_source_order", "FY", "_line_order", "scenario_name"], kind="stable").drop(columns=["_source_order", "_line_order"]).reset_index(drop=True)


def revenue_formula_residual_frame(line_reconciliation: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "source_path",
        "FY",
        "period",
        "scenario_name",
        "output_series_id",
        "output_label",
        "formula",
        "observed_value",
        "calculated_value",
        "residual",
        "residual_abs",
        "status",
        "missing_inputs",
    ]
    if line_reconciliation is None or line_reconciliation.empty:
        return pd.DataFrame(columns=columns)
    data = line_reconciliation.copy()
    data["FY_numeric"] = pd.to_numeric(data.get("FY"), errors="coerce").astype("Int64")
    data["value_numeric"] = pd.to_numeric(data.get("value"), errors="coerce")
    data = data[data["FY_numeric"].notna()].copy()
    rows: list[dict[str, Any]] = []
    for (source_path, scenario_name, fy), group in data.groupby(["source_path", "scenario_name", "FY_numeric"], dropna=False):
        values = {
            str(row.series_id): float(row.value_numeric)
            for row in group.itertuples()
            if pd.notna(row.value_numeric)
        }
        labels = {
            str(row["series_id"]): str(row["line_label"])
            for row in group[["series_id", "line_label"]].drop_duplicates().to_dict("records")
        }
        for formula in FORMULA_DEFINITIONS:
            output = str(formula["output_series_id"])
            if output == "total_fed_ruc_net_revenue":
                continue
            terms = tuple(formula["terms"])
            missing = [series_id for series_id, _sign in terms if series_id not in values]
            observed = values.get(output, pd.NA)
            calculated = pd.NA if missing else sum(values[series_id] * sign for series_id, sign in terms)
            residual = pd.NA
            if pd.notna(observed) and pd.notna(calculated):
                residual = float(observed) - float(calculated)
            status = "missing_inputs" if missing else "reconciled" if pd.notna(residual) and abs(float(residual)) <= 0.05 else "residual_reported"
            rows.append(
                {
                    "source_path": source_path,
                    "FY": int(fy),
                    "period": f"FY{int(fy)}",
                    "scenario_name": scenario_name,
                    "output_series_id": output,
                    "output_label": labels.get(output, str(formula.get("output_label") or output)),
                    "formula": str(formula["expression"]),
                    "observed_value": observed,
                    "calculated_value": calculated,
                    "residual": residual,
                    "residual_abs": abs(float(residual)) if pd.notna(residual) else pd.NA,
                    "status": status,
                    "missing_inputs": "; ".join(missing),
                }
            )
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns).sort_values(["source_path", "FY", "output_series_id", "scenario_name"], kind="stable").reset_index(drop=True)


def _line_display_label(series_id: Any, fallback: Any = "") -> str:
    text = str(series_id or "")
    requested_labels = {
        CURRENT_LIGHT_TOTAL_SERIES_ID: "Current finalist Light RUC total modelled km",
        "light_ruc_net_revenue": "Light RUC net revenue",
        "heavy_ruc_net_revenue": "Heavy RUC net revenue",
        "gross_ruc_revenue": "Gross RUC",
        "ruc_admin_revenue": "RUC admin",
        "ruc_revenue_net_admin": "RUC net admin",
        "total_ruc_net_revenue": "RUC net admin/refunds",
        "gross_ped_revenue": "Gross PED",
        "gross_lpg_revenue": "LPG",
        "gross_cng_revenue": "CNG",
        "gross_fed_revenue": "Gross FED",
        "net_fed_revenue": "Net FED",
        "mr1_revenue": "MR1",
        "mr2_revenue": "MR2",
        "coo_revenue": "MR13",
        "gross_mvr_revenue": "Gross MVR",
        "mvr_admin_revenue": "MVR admin",
        "mvr_revenue_net_admin_coo": "MVR net admin & COO",
        "net_mvr_revenue": "MVR net admin/refunds/COO",
    }
    if text in requested_labels:
        return requested_labels[text]
    for row in ROW_DEFINITIONS:
        if str(row["series_id"]) == text:
            return str(row["display_name"])
    return str(fallback or text)


def _series_trace_contract_frame() -> pd.DataFrame:
    rows = []
    for item in DISPLAY_SERIES_METADATA:
        metric = item["metric_type"]
        controls = "series; time_grain; horizon" if metric == "activity" else "series; time_grain; horizon; selected_fy"
        rows.append(
            {
                "series_option": item["display_name"],
                "canonical_id": item["canonical_id"],
                "display_name": item["display_name"],
                "metric_type": metric,
                "unit": item["unit"],
                "valid_bases": "not_applicable" if metric == "activity" else "Nominal ex GST",
                "valid_controls": controls,
                "actual_source": "data/revenue_model_source_pack/mbu26_annual_spine/mbu26_annual_spine.csv",
                "primary_forecast_source": "MBU26 official annual rows plus current-finalist model forecast replacements where applicable",
                "excluded_lineage_source": "Legacy Excel forecast paths are excluded from runtime traces",
                "bridge": _series_bridge_text(item["canonical_id"]),
                "last_complete_actual_fy": REVENUE_LAST_COMPLETE_ACTUAL_FY,
                "first_forecast_fy": f"FY{REVENUE_PARTIAL_ACTUAL_FY}",
                "first_forecast_quarter": REVENUE_FIRST_FORECAST_QUARTER,
                "model_training_cutoff": REVENUE_MODEL_TRAINING_CUTOFF,
                "availability_status": "mbu26_current_runtime_available",
                "interpretation": "Activity controls hide revenue-only selectors." if metric == "activity" else "Revenue trace is either MBU26 official or current-finalist replacement-only hybrid.",
            }
        )
    return pd.DataFrame(rows).sort_values("series_option", kind="stable").reset_index(drop=True)


def _trace_source_contract_frame(workbook_name: str, workbook_hash: str) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "trace_name": "Actual",
                "trace_type": "Actual",
                "trace_role": "source_actual",
                "source_file": "mbu26_annual_spine.csv",
                "model_id": "",
                "cutoff": f"last complete FY{REVENUE_LAST_COMPLETE_ACTUAL_FY}",
                "scenario": "actual",
                "period_status": "ACTUAL",
                "anchor_forecast_flag": "actual",
                "runtime_forecast_source": False,
                "displayed": True,
                "workbook_basename": workbook_name,
                "workbook_sha256": workbook_hash,
                "notes": "Connected grey actual line ends at FY2025.",
            },
            {
                "trace_name": "MBU26 official",
                "trace_type": "MBU26 official",
                "trace_role": "official_external_comparator",
                "source_file": "mbu26_official_annual.csv",
                "model_id": "",
                "cutoff": "MBU26 ST_FORECAST/LT_FORECAST rows",
                "scenario": "MBU26",
                "period_status": "ST_FORECAST; LT_FORECAST",
                "anchor_forecast_flag": "external_comparator",
                "runtime_forecast_source": False,
                "displayed": True,
                "workbook_basename": workbook_name,
                "workbook_sha256": workbook_hash,
                "notes": "Official MOT release comparator from MBU26 worksheet only.",
            },
            {
                "trace_name": "Current finalist Base case",
                "trace_type": "current finalist base",
                "trace_role": "in_house_current_finalist",
                "source_file": "forecast_scenario_comparison.parquet; mbu26_official_annual.csv",
                "model_id": CURRENT_FINALIST_COMPOSITE_MODEL_ID,
                "cutoff": REVENUE_MODEL_TRAINING_CUTOFF,
                "scenario": "current_basecase",
                "period_status": "FY2025 actual anchor; FY2026 nowcast; FY2027+ forecast",
                "anchor_forecast_flag": "actual_anchor_then_forecast",
                "runtime_forecast_source": True,
                "displayed": True,
                "workbook_basename": workbook_name,
                "workbook_sha256": workbook_hash,
                "notes": "Finalist model output replaces only PED, Light RUC and Heavy RUC revenue.",
            },
            {
                "trace_name": "Current finalist High population/comparison",
                "trace_type": "current finalist comparison",
                "trace_role": "in_house_current_finalist",
                "source_file": "forecast_scenario_comparison.parquet; mbu26_official_annual.csv",
                "model_id": CURRENT_FINALIST_COMPOSITE_MODEL_ID,
                "cutoff": REVENUE_MODEL_TRAINING_CUTOFF,
                "scenario": "current_comparison_1",
                "period_status": "FY2025 actual anchor; FY2026 nowcast; FY2027+ forecast",
                "anchor_forecast_flag": "actual_anchor_then_forecast",
                "runtime_forecast_source": True,
                "displayed": True,
                "workbook_basename": workbook_name,
                "workbook_sha256": workbook_hash,
                "notes": "Finalist model output replaces only PED, Light RUC and Heavy RUC revenue.",
            },
        ]
    )


def _path_trace_status_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"trace_id": "actual", "trace_label": "Actual", "availability_status": "available", "plotted": True, "data_scope": "MBU26 ACTUAL rows through FY2025", "blocking_gap_id": "", "current_selection": "MBU26", "user_visible_message": "Actuals are read from the MBU26 annual source spine."},
            {"trace_id": "mbu26_official", "trace_label": "MBU26 official", "availability_status": "available", "plotted": True, "data_scope": "MBU26 ST_FORECAST/LT_FORECAST rows", "blocking_gap_id": "", "current_selection": "MBU26", "user_visible_message": "Official comparator is the MBU26 worksheet only."},
            {"trace_id": "current_finalist_base", "trace_label": "Current finalist Base case", "availability_status": "available", "plotted": True, "data_scope": "Current finalist model output annualized to June years", "blocking_gap_id": "", "current_selection": "current_basecase", "user_visible_message": "Current finalist base replaces only PED, Light RUC and Heavy RUC revenue."},
            {"trace_id": "current_finalist_comparison", "trace_label": "Current finalist High population/comparison", "availability_status": "available", "plotted": True, "data_scope": "Current finalist comparison output annualized to June years", "blocking_gap_id": "", "current_selection": "current_comparison_1", "user_visible_message": "Current finalist comparison replaces only PED, Light RUC and Heavy RUC revenue."},
        ]
    )


def _series_bridge_text(series_id: str) -> str:
    if series_id == "gross_ped_revenue":
        return "Current PED VKT/capita forecast through MBU26 population, litres intensity and gross PED rate."
    if series_id == "light_ruc_net_revenue":
        return "Current Light RUC net-km forecast multiplied by MBU26 effective Light RUC rate."
    if series_id == "heavy_ruc_net_revenue":
        return "Current Heavy RUC net-km forecast multiplied by MBU26 effective Heavy RUC rate."
    if series_id == "total_fed_ruc_net_revenue":
        return "Net FED plus Total RUC from MBU26/current-finalist hybrid rows."
    if series_id == "total_nltf_net_revenue":
        return "Net FED plus Total RUC plus MBU26 Net MVR and TUC."
    return "MBU26 source row or direct current finalist activity forecast."


def _current_model_id_for_series(series_id: str) -> str:
    text = str(series_id)
    if text in {"ped_vkt_per_capita", "gross_ped_revenue", "gross_fed_revenue", "net_fed_revenue", "ped_volume"}:
        return CURRENT_FINALIST_MODEL_IDS["PED"]
    if text in {
        CURRENT_LIGHT_TOTAL_SERIES_ID,
        "light_ruc_net_km",
        "light_ruc_net_revenue",
        "light_bev_ruc_net_km",
        "light_bev_ruc_net_revenue",
        "phev_ruc_net_km",
        "phev_ruc_net_revenue",
    }:
        return CURRENT_FINALIST_MODEL_IDS["LIGHT_RUC"]
    if text in {"heavy_ruc_net_km", "heavy_ruc_net_revenue"}:
        return CURRENT_FINALIST_MODEL_IDS["HEAVY_RUC"]
    if text in {"total_ruc_net_revenue", "total_fed_ruc_net_revenue", "total_nltf_net_revenue"}:
        return CURRENT_FINALIST_COMPOSITE_MODEL_ID
    return ""


def _current_stream_for_series(series_id: str) -> str:
    if str(series_id).startswith("ped_"):
        return "PED"
    if str(series_id).startswith("light_"):
        return "LIGHT_RUC"
    if str(series_id).startswith("heavy_"):
        return "HEAVY_RUC"
    return ""


def _june_year_from_quarter(period: str) -> int | None:
    year, quarter = _parse_quarter(period)
    if year is None or quarter is None:
        return None
    return year if quarter in {1, 2} else year + 1


def _expected_june_year_quarters(fy: int) -> list[str]:
    return [f"{fy - 1}Q3", f"{fy - 1}Q4", f"{fy}Q1", f"{fy}Q2"]


def _parse_quarter(period: str) -> tuple[int | None, int | None]:
    match = re.match(r"^(\d{4})Q([1-4])$", str(period or "").upper().strip())
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _first_text(frame: pd.DataFrame, column: str) -> str:
    if not isinstance(frame, pd.DataFrame) or frame.empty or column not in frame.columns:
        return ""
    values = [str(value).strip() for value in frame[column].dropna() if str(value).strip()]
    return values[0] if values else ""


def _source_cell_range(frame: pd.DataFrame) -> str:
    if not isinstance(frame, pd.DataFrame) or frame.empty or "source_cell" not in frame.columns:
        return ""
    cells = [str(value).strip() for value in frame["source_cell"].dropna() if str(value).strip()]
    if not cells:
        return ""
    return cells[0] if cells[0] == cells[-1] else f"{cells[0]}:{cells[-1]}"


def _definition_source_cell_hint(definition: dict[str, Any]) -> str:
    row = definition.get("row") if isinstance(definition, dict) else ""
    if row in (None, ""):
        return ""
    return f"MBU26 row {row}"


def _as_int(value: Any) -> int | None:
    try:
        if pd.isna(value):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _validate_manifest_hashes(base: Path, manifest: dict[str, Any]) -> None:
    errors = []
    for filename, metadata in sorted((manifest.get("normalized_files") or {}).items()):
        path = base / filename
        expected = str(metadata.get("sha256", "")).strip() if isinstance(metadata, dict) else ""
        if not path.exists():
            errors.append(f"{filename} missing")
        elif expected and sha256(path) != expected:
            errors.append(f"{filename} hash mismatch")
    if errors:
        raise ValueError("MBU26 annual spine failed hash validation: " + "; ".join(errors))


def _repo_relative(repo_root: Path, path: Path) -> str:
    try:
        return path.resolve().relative_to(repo_root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def _manifest_markdown(manifest: dict[str, Any]) -> str:
    workbook = manifest.get("workbook") or {}
    rows = [
        "# MBU26 Annual Source Spine",
        "",
        f"- Schema: `{manifest.get('schema_version')}`",
        f"- Source release: `{manifest.get('source_release')}`",
        f"- Output: `{manifest.get('repo_relative_output_dir')}`",
        f"- Workbook: `{workbook.get('basename')}`",
        f"- Workbook SHA256: `{workbook.get('sha256')}`",
        f"- Worksheet: `{workbook.get('sheet')}`",
        "",
        "The workbook is offline lineage only. Streamlit reads the repo-local CSV/Parquet extracts.",
        "",
        "## Formula Policy",
        "",
        str(manifest.get("formula_policy") or ""),
    ]
    return "\n".join(rows) + "\n"
