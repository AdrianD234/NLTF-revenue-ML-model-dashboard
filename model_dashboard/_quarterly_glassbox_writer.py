"""Workbook writer for the quarterly glass-box XLSX.

Layout contract (see revenue_outlook_quarterly_glassbox for the data layer):

- Every value sheet shares one quarter-column grid: column B = 2000Q3 through
  column GS = 2050Q2 (200 columns, FY2001-FY2050 in June years).
- Scenario sheet preserves the annual extract's row labels and order MINUS the
  annual-percentage-change block (owner request 2026-09): template rows 1-14
  stay put, template rows 26-65 shift up 11 rows, so the Total net revenues
  line lands on row 54. Rows 56-59 stay reserved for the FED->RUC transition
  lines. Detail sections start at row 64 and never move.
- Row 1 = June FY; row 2 = calendar quarter with the status painted as the
  cell fill (grey = actual, orange = model forecast, peach = post-model),
  matching the owner's reference workbook.
- Annual constructs (the FY2031+ handover, VFM shares, per-FY rates and
  factors) live on FY-grid blocks (column B = FY2001 .. column AY = FY2050)
  and are read from quarter columns with XLOOKUP on the column's FY header
  (written as _xlfn.XLOOKUP, which is how a non-Excel writer must spell it).
- Styling follows the owner's reference extract: Calibri 11, #,##0 value
  formats (nothing beyond two decimals anywhere), grey section banners, bold
  subtotal rows. Blue = committed input, black = in-sheet formula, green =
  cross-sheet link. No Excel Table objects, no INDIRECT/OFFSET.
- Zero-valued indicator cells are omitted entirely rather than written as 0,
  and dummy/trend rows are painted calendar constants, not string-parsing
  formulas - the owner wants small, simple formulas.
"""

from __future__ import annotations

import io
import math
from typing import Iterable, Mapping, Sequence

from model_dashboard.revenue_outlook_quarterly_glassbox import (
    CARRIED_ALLOCATED_SERIES,
    CHART_QUARTERLY_SERIES,
    LIGHT_COEFFICIENT_NAMES,
    PED_COEFFICIENT_NAMES,
    RevenueOutlookGlassboxError,
    fiscal_year_of_quarter,
    quarters_of_fiscal_year,
)

SHEET_README = "README"
SHEET_SCENARIO = "Scenario"
SHEET_INPUTS = "Quarterly Inputs"
SHEET_PARAMS = "Model Parameters"
SHEET_CHECKS = "Checks"

_FIRST_DATA_COLUMN = 2
_ANNUAL_FIRST_FY = 2001
_ANNUAL_LAST_FY = 2050

#: Scenario top block: workbook row -> annual-template row. The template's
#: annual-percentage-change block (rows 15-25) is omitted by owner request;
#: everything below shifts up 11 rows.
TOP_ROW_TEMPLATE_MAP: dict[int, int] = {
    **{row: row for row in range(1, 15)},
    **{row - 11: row for row in range(26, 66)},
}
TOP_TOTAL_ROW = 54
#: Reserved for the FED->RUC transition lines when that overlay is active.
TRANSITION_RESERVED_ROWS = (56, 57, 58, 59)
_DETAIL_START_ROW = 64

#: Template rows whose values are bold in the reference extract (subtotals).
_TEMPLATE_BOLD_ROWS = frozenset({34, 36, 37, 43, 45, 51, 53, 55, 58, 65})

_CALIBRI = "Calibri"
_BLUE_INPUT = "0000CC"
_GREEN_LINK = "006633"
_GREY_NOTE = "595959"

_FILL_SECTION = "D0D0D0"
_FILL_SUBSECTION = "EFEFEF"
_FILL_ACTUAL = "CDD4DD"
_FILL_MODEL = "F0A072"
_FILL_POST = "FBDECD"

_NF_VALUE = "#,##0"
_NF_VALUE2 = "#,##0.00"
_NF_KM = "#,##0"
_NF_RATE = "0.00"
_NF_LOG = "0.00"
_NF_POP = "#,##0"
_NF_DELTA = "0.00"
_NF_FACTOR = "0.00"
_NF_INT = "0"

_STATUS_FILLS = {
    "ACTUAL": _FILL_ACTUAL,
    "MODEL FORECAST": _FILL_MODEL,
    "POST-MODEL": _FILL_POST,
}


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


class _Styles:
    def __init__(self) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill

        self.title = Font(name=_CALIBRI, size=14, bold=True)
        self.section = Font(name=_CALIBRI, size=11, bold=True)
        self.subhead = Font(name=_CALIBRI, size=11, bold=True)
        self.label = Font(name=_CALIBRI, size=11)
        self.label_bold = Font(name=_CALIBRI, size=11, bold=True)
        self.note = Font(name=_CALIBRI, size=9, color=_GREY_NOTE)
        self.header = Font(name=_CALIBRI, size=11, bold=True)
        self.input = Font(name=_CALIBRI, size=11, color=_BLUE_INPUT)
        self.input_bold = Font(name=_CALIBRI, size=11, bold=True, color=_BLUE_INPUT)
        self.formula = Font(name=_CALIBRI, size=11)
        self.formula_bold = Font(name=_CALIBRI, size=11, bold=True)
        self.link = Font(name=_CALIBRI, size=11, color=_GREEN_LINK)
        self.link_bold = Font(name=_CALIBRI, size=11, bold=True, color=_GREEN_LINK)
        self.section_fill = PatternFill("solid", start_color=_FILL_SECTION)
        self.subsection_fill = PatternFill("solid", start_color=_FILL_SUBSECTION)
        self.check_fill = PatternFill("solid", start_color="ECFDF5")
        self.status_fills = {
            status: PatternFill("solid", start_color=color)
            for status, color in _STATUS_FILLS.items()
        }
        self.actual_fill = PatternFill("solid", start_color=_FILL_ACTUAL)
        self.derived_actual_fill = PatternFill("solid", start_color="E9EDF2")
        self.wrap = Alignment(wrap_text=True, vertical="top")
        self.center = Alignment(horizontal="center")


class _SheetGrid:
    """Quarter-grid bookkeeping for one worksheet."""

    def __init__(self, worksheet, quarters: Sequence[str], styles: _Styles) -> None:
        self.ws = worksheet
        self.quarters = list(quarters)
        self.styles = styles
        self.col_of = {
            quarter: _FIRST_DATA_COLUMN + index
            for index, quarter in enumerate(self.quarters)
        }
        self.last_column = _FIRST_DATA_COLUMN + len(self.quarters) - 1
        self.rows: dict[str, int] = {}

    def letter(self, quarter: str) -> str:
        return _column_letter(self.col_of[quarter])

    def register(self, key: str, row: int) -> int:
        if key in self.rows:
            raise RevenueOutlookGlassboxError(f"duplicate glass-box row key {key!r}.")
        self.rows[key] = row
        return row

    def row(self, key: str) -> int:
        if key not in self.rows:
            raise RevenueOutlookGlassboxError(f"unknown glass-box row key {key!r}.")
        return self.rows[key]


def _fy_column(fy: int) -> int:
    return _FIRST_DATA_COLUMN + (int(fy) - _ANNUAL_FIRST_FY)


_ANNUAL_LAST_COLUMN = _fy_column(_ANNUAL_LAST_FY)


# ---------------------------------------------------------------------------
# Reader-facing names: no snake_case identifiers anywhere on a worksheet.
# ---------------------------------------------------------------------------

_EXTRA_DISPLAY_NAMES = {
    "ped_vkt_per_capita": "Light petrol VKT per capita (km)",
    "light_petrol_vkt": "Light petrol VKT (m km)",
    "ped_volume": "PED volume (m L)",
    "total_fed_ruc_net_revenue": "Total FED + RUC net revenue (m $)",
    "current_light_ruc_conventional_modelled_km": "Light RUC net km (m km)",
    "tuc_gtk": "TUC GTK (m Tonne-km)",
}


def _template_display_names(row_labels: Mapping[int, str]) -> dict[str, str]:
    from model_dashboard.revenue_outlook_excel_extract import ROW_SERIES

    names = dict(_EXTRA_DISPLAY_NAMES)
    for template_row, series_id in ROW_SERIES.items():
        label = str(row_labels.get(template_row, "")).strip()
        if label:
            names.setdefault(series_id, label)
    return names


def display_name(data, series_id: str) -> str:
    names = getattr(data, "_display_names", None)
    if names is None:
        names = _template_display_names(data.row_labels)
        data._display_names = names
    return names.get(series_id, series_id.replace("_", " "))


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _set_label(grid: _SheetGrid, row: int, text: str, *, font=None, indent: int = 0,
               fill=None) -> None:
    cell = grid.ws.cell(row=row, column=1)
    cell.value = ("    " * indent) + text if indent else text
    cell.font = font or grid.styles.label
    if fill is not None:
        cell.fill = fill


def _section_header(grid: _SheetGrid, row: int, text: str, *, fill=None) -> None:
    ws = grid.ws
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = grid.styles.section
    banner = fill or grid.styles.section_fill
    for column in range(1, grid.last_column + 1):
        ws.cell(row=row, column=column).fill = banner


def _paint_quarters(
    grid: _SheetGrid,
    row: int,
    values: Mapping[str, float],
    *,
    font,
    number_format: str,
    quarters: Iterable[str] | None = None,
    skip_zero: bool = False,
) -> None:
    ws = grid.ws
    for quarter in quarters if quarters is not None else grid.quarters:
        value = values.get(quarter)
        if value is None or (isinstance(value, float) and not math.isfinite(value)):
            continue
        if skip_zero and abs(float(value)) < 1e-12:
            continue
        cell = ws.cell(row=row, column=grid.col_of[quarter])
        cell.value = float(value)
        cell.font = font
        cell.number_format = number_format


def _formula_quarters(
    grid: _SheetGrid,
    row: int,
    template: str,
    *,
    font,
    number_format: str,
    quarters: Iterable[str] | None = None,
) -> None:
    """One formula per quarter column; {c}/{cprev} resolve to column letters."""
    ws = grid.ws
    for quarter in quarters if quarters is not None else grid.quarters:
        column = grid.col_of[quarter]
        letter = _column_letter(column)
        prev_letter = _column_letter(column - 1) if column > _FIRST_DATA_COLUMN else ""
        if "{cprev}" in template and not prev_letter:
            continue
        formula = template.format(c=letter, cprev=prev_letter)
        cell = ws.cell(row=row, column=column)
        cell.value = formula
        cell.font = font
        cell.number_format = number_format


def _paint_fy_row(
    grid: _SheetGrid,
    row: int,
    values: Mapping[int, float],
    *,
    font,
    number_format: str,
    skip_zero: bool = False,
) -> None:
    ws = grid.ws
    for fy, value in values.items():
        if fy < _ANNUAL_FIRST_FY or fy > _ANNUAL_LAST_FY:
            continue
        if value is None or (isinstance(value, float) and not math.isfinite(value)):
            continue
        if skip_zero and abs(float(value)) < 1e-12:
            continue
        cell = ws.cell(row=row, column=_fy_column(fy))
        cell.value = float(value)
        cell.font = font
        cell.number_format = number_format


def _fy_header_row(grid: _SheetGrid, row: int, label: str) -> None:
    ws = grid.ws
    _set_label(grid, row, label, font=grid.styles.header)
    for fy in range(_ANNUAL_FIRST_FY, _ANNUAL_LAST_FY + 1):
        cell = ws.cell(row=row, column=_fy_column(fy))
        cell.value = int(fy)
        cell.font = grid.styles.header
        cell.number_format = "0"


def _fy_range(row: int, sheet: str | None = None) -> str:
    prefix = f"'{sheet}'!" if sheet else ""
    return (
        f"{prefix}${_column_letter(_FIRST_DATA_COLUMN)}${row}:"
        f"${_column_letter(_ANNUAL_LAST_COLUMN)}${row}"
    )


def _xlookup(lookup: str, header_range: str, value_range: str) -> str:
    """XLOOKUP as a non-Excel writer must store it (_xlfn prefix; Excel shows
    a plain XLOOKUP). Lookups here always hit, so no if_not_found argument."""
    return f"_xlfn.XLOOKUP({lookup},{header_range},{value_range})"


# ---------------------------------------------------------------------------
# Shared sheet chrome
# ---------------------------------------------------------------------------


def _quarter_header_rows(grid: _SheetGrid, *, status: Mapping[str, str]) -> None:
    ws = grid.ws
    _set_label(grid, 1, "YE June", font=grid.styles.header)
    _set_label(grid, 2, "Period", font=grid.styles.header)
    for quarter in grid.quarters:
        column = grid.col_of[quarter]
        fy_cell = ws.cell(row=1, column=column)
        fy_cell.value = fiscal_year_of_quarter(quarter)
        fy_cell.number_format = "0"
        fy_cell.font = grid.styles.header
        fy_cell.alignment = grid.styles.center
        q_cell = ws.cell(row=2, column=column)
        q_cell.value = quarter
        q_cell.font = grid.styles.header
        q_cell.alignment = grid.styles.center
        fill = grid.styles.status_fills.get(status.get(quarter, ""))
        if fill is not None:
            q_cell.fill = fill
    ws.column_dimensions["A"].width = 44
    for column in range(_FIRST_DATA_COLUMN, grid.last_column + 1):
        ws.column_dimensions[_column_letter(column)].width = 11.5
    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Quarterly Inputs sheet (drivers + policy pricing + components + shares)
# ---------------------------------------------------------------------------


def _write_inputs_sheet(grid: _SheetGrid, data, *, audit_sheets: bool) -> None:
    styles = grid.styles
    _quarter_header_rows(grid, status=data.status)
    row = 4
    _section_header(grid, row, "Quarterly inputs - committed production scenario inputs (blue = committed value)")
    row += 1
    _set_label(
        grid, row,
        "Reference scenario: central base case replay; selected rate path: "
        f"{data.policy.state_label}.",
        font=styles.note,
    )
    row += 2

    driver_meta = (
        ("population", "Population (persons)", _NF_POP),
        ("real_gdp_sa_nzd", "Real GDP, s.a. ($)", _NF_VALUE),
        ("real_gdp_per_capita_nzd", "Real GDP per capita ($)", _NF_VALUE),
        ("unemployment_rate", "Unemployment rate (fraction)", "0.000"),
        ("real_petrol_price_cents_per_litre", "Real petrol price (c/L)", _NF_VALUE2),
        ("real_diesel_price_cents_per_litre", "Real diesel price (c/L)", _NF_VALUE2),
        ("real_light_ruc_price_nzd_per_1000km", "Real Light RUC model price ($/1,000 km)", _NF_VALUE2),
        (
            "lagged_real_light_ruc_price_nzd_per_1000km",
            "Lagged real Light RUC price ($/1,000 km)",
            _NF_VALUE2,
        ),
        ("real_heavy_ruc_price_nzd_per_1000km", "Real Heavy RUC model price ($/1,000 km)", _NF_VALUE2),
        (
            "lead_real_heavy_ruc_price_nzd_per_1000km",
            "Led real Heavy RUC price ($/1,000 km)",
            _NF_VALUE2,
        ),
    )
    for key, label, number_format in driver_meta:
        values = data.drivers.get(key)
        if not values:
            continue
        grid.register(f"in.{key}", row)
        _set_label(grid, row, label)
        _paint_quarters(grid, row, values, font=styles.input, number_format=number_format)
        row += 1
        selected = data.drivers.get(f"{key}__selected")
        if selected:
            grid.register(f"in.{key}__selected", row)
            _set_label(grid, row, f"{label} - selected rate path", indent=1)
            _paint_quarters(
                grid, row, selected, font=styles.input, number_format=number_format
            )
            row += 1
    row += 1

    _section_header(grid, row, "Selected rate-path pricing - committed pump and rate paths")
    row += 1
    for key, label in (
        ("policy_free_source_nominal_petrol_cpl", "Policy-free nominal petrol price (c/L)"),
        ("policy_published_fed_wedge_nominal_cpl", "Published FED wedge (c/L)"),
        ("policy_source_nominal_petrol_cpl", "Reference nominal pump price (c/L)"),
        ("policy_target_nominal_petrol_cpl", "Selected-path nominal pump price (c/L)"),
    ):
        values = data.policy.ped_nominal.get(key)
        if not values:
            continue
        grid.register(f"in.{key}", row)
        _set_label(grid, row, label)
        _paint_quarters(grid, row, values, font=styles.input, number_format=_NF_VALUE2)
        row += 1
    for stream, stream_label in (("LIGHT_RUC", "Light RUC"), ("HEAVY_RUC", "Heavy RUC")):
        detail = data.policy.ruc_detail.get(stream, {})
        for key, label in (
            ("reference_fuel_cost", f"{stream_label}: reference diesel cost ($/1,000 km)"),
            ("variant_fuel_cost", f"{stream_label}: selected diesel cost ($/1,000 km)"),
            ("reference_ruc_price", f"{stream_label}: reference RUC rate ($/1,000 km)"),
            ("variant_ruc_price", f"{stream_label}: selected RUC rate ($/1,000 km)"),
        ):
            values = detail.get(key)
            if not values:
                continue
            grid.register(f"in.{stream}.{key}", row)
            _set_label(grid, row, label)
            _paint_quarters(grid, row, values, font=styles.input, number_format=_NF_VALUE2)
            row += 1
    row += 1

    _section_header(grid, row, "Model components - exact imported predictions from the fitted models")
    row += 1
    _set_label(
        grid, row,
        "Tree-based components have no regression coefficients; these are their "
        "exact committed per-quarter outputs from the production scorer. The "
        "Scenario sheet links to these rows.",
        font=styles.note,
    )
    row += 1
    grid.register("cmp.light_residual_log", row)
    _set_label(grid, row, "Light RUC residual-GBR log component (exact import)")
    _paint_quarters(grid, row, data.light.residual_log, font=styles.input, number_format=_NF_LOG)
    row += 1
    for member in data.heavy.components:
        grid.register(f"cmp.heavy_{member.label}_log", row)
        kind = "Ridge" if member.kind == "ridge" else "GBR"
        _set_label(grid, row, f"Heavy RUC {member.label} log prediction - {kind} (exact import)")
        _paint_quarters(grid, row, member.log_value, font=styles.input, number_format=_NF_LOG)
        row += 1
    row += 1

    _section_header(grid, row, "Governed quarterly allocation shares - display quarter / June-year total")
    row += 1
    _set_label(
        grid, row,
        "Share of each June year's governed annual value allocated to the quarter "
        "by the quarterly display contract. Four-quarter shares sum to 1.",
        font=styles.note,
    )
    row += 1
    annual = data.annual_values
    for series_id in tuple(CHART_QUARTERLY_SERIES) + tuple(CARRIED_ALLOCATED_SERIES):
        quarters = data.quarterly_values.get(series_id)
        if quarters is None:
            continue
        series_annual = annual.get(series_id, {})
        shares: dict[str, float] = {}
        for quarter, value in quarters.items():
            fy = fiscal_year_of_quarter(quarter)
            total = series_annual.get(fy)
            if total is None or abs(total) < 1e-12:
                continue
            shares[quarter] = value / total
        grid.register(f"share.{series_id}", row)
        _set_label(grid, row, f"Quarterly share: {display_name(data, series_id)}", indent=1)
        _paint_quarters(grid, row, shares, font=styles.input, number_format=_NF_FACTOR)
        row += 1
    row += 1

    _section_header(grid, row, "Actual-period history (published actuals and derived quarterly presentation)")
    row += 1
    _set_label(
        grid, row,
        "Natively published actual quarters where they exist (the three modelled "
        "activity series); otherwise the governed quarterly-display presentation "
        "of the annual actuals. The Scenario sheet colours the two apart.",
        font=styles.note,
    )
    row += 1
    for series_id in sorted(data.actual_quarters):
        values = data.actual_quarters.get(series_id, {})
        if not values:
            continue
        grid.register(f"hist.{series_id}", row)
        _set_label(grid, row, f"Actual periods: {display_name(data, series_id)}")
        _paint_quarters(grid, row, values, font=styles.input, number_format=_NF_VALUE)
        row += 1

    if audit_sheets:
        row += 1
        _section_header(grid, row, "Verification values (audit build only)")
        row += 1
        for key, label, values, number_format in (
            ("chk.ped_committed", "PED committed raw prediction (km per capita)",
             data.ped.raw_prediction, _NF_VALUE2),
            ("chk.light_committed", "Light RUC committed raw prediction (km)",
             data.light.raw_prediction, _NF_VALUE),
            ("chk.heavy_committed", "Heavy RUC committed FINAL prediction (km)",
             data.heavy.raw_prediction, _NF_VALUE),
            ("chk.calibrated_PED", "PED calibrated displayed (km per capita)",
             data.policy.calibrated.get("PED", {}), _NF_VALUE2),
            ("chk.calibrated_LIGHT_RUC", "Light RUC calibrated displayed (km)",
             data.policy.calibrated.get("LIGHT_RUC", {}), _NF_VALUE),
            ("chk.calibrated_HEAVY_RUC", "Heavy RUC calibrated displayed (km)",
             data.policy.calibrated.get("HEAVY_RUC", {}), _NF_VALUE),
        ):
            grid.register(key, row)
            _set_label(grid, row, label)
            _paint_quarters(grid, row, values, font=styles.input, number_format=number_format)
            row += 1


# ---------------------------------------------------------------------------
# Model Parameters sheet
# ---------------------------------------------------------------------------


def _write_params_sheet(grid: _SheetGrid, data, defined_names: dict[str, str]) -> None:
    ws = grid.ws
    styles = grid.styles
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 52
    ws.freeze_panes = "A2"

    def _name_cell(row: int, name: str, value: float, *, number_format=_NF_LOG) -> None:
        cell = ws.cell(row=row, column=3)
        cell.value = float(value)
        cell.font = styles.input
        cell.number_format = number_format
        defined_names[name] = f"'{SHEET_PARAMS}'!$C${row}"
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=2).font = styles.label_bold

    def _meta(row: int, label: str, name: str, value: float, units: str, component: str,
              source: str, number_format=_NF_LOG) -> int:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = styles.label
        _name_cell(row, name, value, number_format=number_format)
        for column, text in ((4, units), (5, component), (6, source)):
            ws.cell(row=row, column=column).value = text
            ws.cell(row=row, column=column).font = styles.note
        return row + 1

    row = 1
    ws.cell(row=row, column=1).value = "Model parameters - governed fitted values and defined names"
    ws.cell(row=row, column=1).font = styles.title
    row += 2
    for column, header in enumerate(
        ("Parameter", "Defined name", "Value", "Units", "Model / component", "Source artifact"),
        start=1,
    ):
        cell = ws.cell(row=row, column=column)
        cell.value = header
        cell.font = styles.subhead
        cell.fill = styles.subsection_fill
    row += 1

    ped_src = "PED AR(1) fitted state (evidence pack)"
    _section_header(grid, row, "PED - AR(1) log-linear model")
    row += 1
    row = _meta(row, "β0 - intercept", "PED_B0", data.ped.beta0, "log points", "PED AR(1)", ped_src)
    ped_labels = {
        "PED_BPet": "β petrol - log real petrol price",
        "PED_BGDP": "β GDP - log real GDP per capita",
        "PED_BUnemp": "β unemp - unemployment rate",
        "PED_BTrend": "β trend - linear trend",
        "PED_BPost11": "β post-2011 - post-2011 trend",
        "PED_BPost20": "β post-2020 - level shift",
        "PED_BCovid": "β COVID - 2020 dummy",
        "PED_BQ2": "β Q2 - seasonal dummy",
        "PED_BQ3": "β Q3 - seasonal dummy",
        "PED_BQ4": "β Q4 - seasonal dummy",
    }
    for feature, name in PED_COEFFICIENT_NAMES:
        row = _meta(row, ped_labels[name], name, data.ped.coefficients[feature],
                    "per unit", "PED AR(1)", ped_src)
    row = _meta(row, "β lag - lagged log VKT per capita", "PED_BLagY",
                data.ped.beta_ylag, "per log point", "PED AR(1)", ped_src)
    row = _meta(row, "ρ - AR(1) error autocorrelation", "PED_Rho", data.ped.rho,
                "per quarter", "PED AR(1)", ped_src)
    row = _meta(row, f"Seed residual ({data.ped.latest_actual})", "PED_LastErr",
                data.ped.last_resid, "log points", "PED AR(1)", ped_src)
    row = _meta(row, f"Seed log VKT per capita ({data.ped.latest_actual})",
                "PED_SeedLogVKT", data.ped.log_target_latest, "log points", "PED AR(1)",
                "committed model input history")
    ws.cell(row=row, column=1).value = (
        f"Training window {data.ped.train_window[0]}-{data.ped.train_window[1]} "
        f"({data.ped.train_window[2]} rows); latest actual {data.ped.latest_actual}."
    )
    ws.cell(row=row, column=1).font = styles.note
    row += 2

    light_src = "Light RUC promoted fitted state (evidence pack)"
    _section_header(grid, row, "Light RUC - OLS base (coefficients) + residual GBR (exact import)")
    row += 1
    row = _meta(row, "β0 - OLS intercept", "LR_B0", data.light.intercept, "log points",
                "Light RUC OLS base", light_src)
    light_labels = {
        "LR_BDiesel": "β diesel - log real diesel price",
        "LR_BRuc": "β RUC - log real Light RUC price",
        "LR_BRucLag": "β RUC lag - log lagged Light RUC price",
        "LR_BGDP": "β GDP - log real GDP",
        "LR_BPost20": "β post-2020 - dummy (from CY2021)",
        "LR_BQ2": "β Q2 - seasonal dummy",
        "LR_BQ3": "β Q3 - seasonal dummy",
        "LR_BQ4": "β Q4 - seasonal dummy",
    }
    for feature, name in LIGHT_COEFFICIENT_NAMES:
        row = _meta(row, light_labels[name], name, data.light.coefficients[feature],
                    "per unit", "Light RUC OLS base", light_src)
    ws.cell(row=row, column=1).value = (
        "Residual model: gradient-boosted trees (150 trees, depth 1, learning rate "
        "0.05). No regression coefficients exist for the tree component; its exact "
        "per-quarter log output is imported on the Quarterly Inputs sheet. "
        f"Training window {data.light.train_window[0]}-{data.light.train_window[1]}."
    )
    ws.cell(row=row, column=1).font = styles.note
    row += 2

    _section_header(grid, row, "Heavy RUC - governed ensemble weights (level-space blend)")
    row += 1
    kind_names = {"ridge": "Ridge regression (linear)", "gbr": "gradient-boosted trees"}
    for index, member in enumerate(data.heavy.components, start=1):
        row = _meta(
            row,
            f"w{index} - {member.label} weight ({kind_names.get(member.kind, member.kind)}, "
            f"trained {member.train_window[0]}-{member.train_window[1]})",
            f"HR_W{index}",
            member.weight,
            "share",
            f"Heavy RUC component {member.label}",
            "Heavy RUC fitted-model manifest (evidence pack)",
            number_format=_NF_FACTOR,
        )
    ws.cell(row=row, column=1).value = (
        "Final Heavy RUC = w1 x EXP(M1 log) + w2 x EXP(M2 log) + w3 x EXP(M3 log). "
        "M1 is the linear (Ridge) component; the two tree components are exact "
        "imports. Weights are governed and sum to 1."
    )
    ws.cell(row=row, column=1).font = styles.note
    ws.cell(row=row, column=1).alignment = styles.wrap
    row += 2

    _section_header(grid, row, "Other governed parameters")
    row += 1
    elasticity_src = "governed sensitivity seed inputs (demand elasticity, Med)"
    row = _meta(row, "ε PED - demand elasticity (pump-price basis)", "PED_Elas",
                data.policy.elasticity["PED"], "elasticity", "policy overlay", elasticity_src,
                number_format="0.000")
    row = _meta(row, "ε Light RUC - demand elasticity (running-cost basis)", "LR_Elas",
                data.policy.elasticity["LIGHT_RUC"], "elasticity", "policy overlay", elasticity_src,
                number_format="0.000")
    row = _meta(row, "ε Heavy RUC - demand elasticity (running-cost basis)", "HR_Elas",
                data.policy.elasticity["HEAVY_RUC"], "elasticity", "policy overlay", elasticity_src,
                number_format="0.000")
    row = _meta(row, "Light RUC diesel intensity", "LR_L100", data.policy.diesel_litres_per_100km["LIGHT_RUC"],
                "L/100km", "policy overlay", "governed calibration constant", number_format="0.000")
    row = _meta(row, "Heavy RUC diesel intensity", "HR_L100", data.policy.diesel_litres_per_100km["HEAVY_RUC"],
                "L/100km", "policy overlay", "governed calibration constant", number_format="0.000")
    ws.cell(row=row, column=1).value = (
        "Treasury macro terminal-carry factor: Total RUC (FY2030 factor carried forward)"
    )
    ws.cell(row=row, column=1).font = styles.label
    _name_cell(row, "RUC_MacroFac", data.macro_terminal_factor_total_ruc,
               number_format="0.0000")
    row += 1
    for label, value in (
        ("Selected FED/RUC rate path", data.policy.state_label),
        ("Long-run handover schedule", "growth handover to FY2035"),
        ("Long-run shape vintage", data.post_model.shape_vintage_id),
        ("Bridge vintage", data.bridge_vintage_id),
        ("Model training cutoff", data.provenance.get("model_training_cutoff", "")),
        ("Engine", "AR(1) production engine"),
    ):
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = styles.label
        ws.cell(row=row, column=3).value = str(value)
        ws.cell(row=row, column=3).font = styles.input
        row += 1
    row += 1

    _section_header(grid, row, "Annual parameter tables (June-year grid; XLOOKUP targets)")
    row += 2
    grid.register("prm.fy_header", row)
    _fy_header_row(grid, row, "June fiscal year")
    row += 1

    annual_tables = [
        ("prm.policy_rate_factor",
         f"Rate factor: selected / planned ({data.policy.state_label})",
         data.policy_rate_factor_by_fy, "0.0000", False),
        ("prm.pack_total_ruc",
         "Pack central Total RUC net revenue (m $, pre-macro pre-policy)",
         data.pack_central_total_ruc, _NF_VALUE, False),
    ]
    spine_series = (
        "ped_volume", "light_petrol_vkt", "gross_ped_revenue",
        "light_ruc_net_km", "light_ruc_net_revenue",
        "light_bev_ruc_net_km", "light_bev_ruc_net_revenue",
        "phev_ruc_net_km", "phev_ruc_net_revenue",
        "heavy_ruc_net_km", "heavy_ruc_net_revenue",
        "heavy_bev_ruc_net_km", "heavy_bev_ruc_net_revenue",
    )
    for series_id in spine_series:
        annual_tables.append(
            (f"prm.spine.{series_id}",
             f"{data.bridge_vintage_id} official: {display_name(data, series_id)}",
             {fy: v for fy, v in data.official_spine.get(series_id, {}).items() if 2026 <= fy <= 2050},
             _NF_VALUE, True)
        )
    for series_id in (
        "ped_vkt_per_capita", "light_petrol_vkt", "light_ruc_net_km",
        "light_bev_ruc_net_km", "phev_ruc_net_km", "heavy_ruc_net_km",
        "gross_ped_revenue", "total_ruc_net_revenue",
    ):
        annual_tables.append(
            (
                f"prm.pair_factor.{series_id}",
                f"Rate-path activity factor FY2026-FY2030: {display_name(data, series_id)}",
                data.policy.annual_pair_factor.get(series_id, {}),
                "0.0000", False,
            )
        )
    for key, label, values, number_format, skip_zero in annual_tables:
        grid.register(key, row)
        _set_label(grid, row, label)
        _paint_fy_row(grid, row, values, font=styles.input, number_format=number_format,
                      skip_zero=skip_zero)
        row += 1

    def _fy_ratio_row(key: str, label: str, numerator_key: str, denominator_key: str,
                      *, scale: str = "", number_format=_NF_RATE) -> None:
        nonlocal row
        grid.register(key, row)
        _set_label(grid, row, label)
        numerator = grid.row(numerator_key)
        denominator = grid.row(denominator_key)
        numerator_series = numerator_key.removeprefix("prm.spine.")
        denominator_series = denominator_key.removeprefix("prm.spine.")
        for fy in range(2026, _ANNUAL_LAST_FY + 1):
            if fy not in data.official_spine.get(numerator_series, {}):
                continue
            if abs(data.official_spine.get(denominator_series, {}).get(fy, 0.0)) < 1e-12:
                continue
            column = _fy_column(fy)
            letter = _column_letter(column)
            cell = ws.cell(row=row, column=column)
            cell.value = f"={scale}{letter}{numerator}/{letter}{denominator}"
            cell.font = styles.formula
            cell.number_format = number_format
        row += 1

    _fy_ratio_row("prm.rate.ped", "Official PED rate ($/L) = gross PED / volume",
                  "prm.spine.gross_ped_revenue", "prm.spine.ped_volume")
    _fy_ratio_row("prm.intensity", "Petrol intensity (L/100km) = 100 x volume / VKT",
                  "prm.spine.ped_volume", "prm.spine.light_petrol_vkt", scale="100*")
    for rate_key, revenue_series, km_series, label in (
        ("prm.rate.light", "light_ruc_net_revenue", "light_ruc_net_km",
         "Official conventional Light RUC rate ($/1,000 km)"),
        ("prm.rate.light_bev", "light_bev_ruc_net_revenue", "light_bev_ruc_net_km",
         "Official Light BEV RUC rate ($/1,000 km)"),
        ("prm.rate.phev", "phev_ruc_net_revenue", "phev_ruc_net_km",
         "Official PHEV RUC rate ($/1,000 km)"),
        ("prm.rate.heavy", "heavy_ruc_net_revenue", "heavy_ruc_net_km",
         "Official Heavy RUC rate ($/1,000 km)"),
        ("prm.rate.heavy_bev", "heavy_bev_ruc_net_revenue", "heavy_bev_ruc_net_km",
         "Official Heavy BEV RUC rate ($/1,000 km)"),
    ):
        _fy_ratio_row(rate_key, f"{label} = 1000 x revenue / km",
                      f"prm.spine.{revenue_series}", f"prm.spine.{km_series}",
                      scale="1000*", number_format=_NF_VALUE2)
    row += 1

    _section_header(grid, row, "Governed annual values (displayed dashboard annuals)")
    row += 1
    for series_id in sorted(set(list(CHART_QUARTERLY_SERIES) + list(CARRIED_ALLOCATED_SERIES))):
        values = data.annual_values.get(series_id, {})
        if not values:
            continue
        grid.register(f"prm.annual.{series_id}", row)
        _set_label(grid, row, f"Annual: {display_name(data, series_id)}")
        _paint_fy_row(grid, row, values, font=styles.input, number_format=_NF_VALUE)
        row += 1
    for column in range(_FIRST_DATA_COLUMN, _ANNUAL_LAST_COLUMN + 1):
        ws.column_dimensions[_column_letter(column)].width = 11.5


# ---------------------------------------------------------------------------
# Scenario sheet: header block + model sections A-D
# ---------------------------------------------------------------------------


def _write_scenario_sheet(grid: _SheetGrid, data) -> None:
    ws = grid.ws
    styles = grid.styles
    _quarter_header_rows(grid, status=data.status)

    section_labels = {
        "Key volumes: Level", "Revenues: Level ($m ex GST)",
    }
    subsection_labels = {
        "Road User Charges", "Fuel Excise Duties", "Motor Vehicle Register",
        "Track User Charges", "TOTALS",
    }
    for workbook_row, template_row in TOP_ROW_TEMPLATE_MAP.items():
        if workbook_row <= 2:
            continue
        label = data.row_labels.get(template_row, "")
        if not label:
            continue
        cell = ws.cell(row=workbook_row, column=1)
        cell.value = label
        if label in section_labels:
            cell.font = styles.label_bold
            for column in range(1, grid.last_column + 1):
                ws.cell(row=workbook_row, column=column).fill = styles.section_fill
        elif label in subsection_labels:
            cell.font = styles.label_bold
            for column in range(1, grid.last_column + 1):
                ws.cell(row=workbook_row, column=column).fill = styles.subsection_fill
        elif template_row in _TEMPLATE_BOLD_ROWS:
            cell.font = styles.label_bold
        else:
            cell.font = styles.label

    # TUC GTK is displayed in millions of tonne-km (scaled at the data layer).
    ws.cell(row=13, column=1).value = "TUC GTK (m Tonne-km)"

    note_row = TRANSITION_RESERVED_ROWS[-1] + 2
    _set_label(
        grid, note_row,
        f"Quarterly glass-box replay - {data.trace_name} / {data.policy.state_label}. "
        "Every value above is a formula linked to the detail sections below or to "
        "another worksheet. Period colours: grey = actual, orange = model forecast, "
        "peach = post-model.",
        font=styles.note,
    )
    _set_label(
        grid, note_row + 1,
        f"Rows {TRANSITION_RESERVED_ROWS[0]}-{TRANSITION_RESERVED_ROWS[-1]} are "
        "reserved for the FED->RUC transition lines (transition Off in this scenario).",
        font=styles.note,
    )

    row = _DETAIL_START_ROW

    # ---------------------------------------------------------- A. PED model
    _section_header(grid, row, "A. PED / light petrol VKT per capita - AR(1) coefficient chain (raw reference model)")
    row += 1
    _set_label(
        grid, row,
        "Central base-case inputs; each contribution row is coefficient x "
        "transformed input. The selected rate path's demand response is applied "
        "in section D. Calendar indicators are painted constants.",
        font=styles.note,
    )
    row += 1

    ped_quarters = [q for q in grid.quarters if q in data.ped.features]
    seed_quarter = data.ped.latest_actual
    inputs_sheet = SHEET_INPUTS
    input_row_petrol = data._inputs_grid.row("in.real_petrol_price_cents_per_litre")
    input_row_gdppc = data._inputs_grid.row("in.real_gdp_per_capita_nzd")
    input_row_unemp = data._inputs_grid.row("in.unemployment_rate")

    _set_label(grid, row, "Transformed inputs", font=styles.subhead)
    row += 1
    link_transforms = (
        ("ped.x.petrol__log", "ln(real petrol price)",
         f"=LN('{inputs_sheet}'!{{c}}{input_row_petrol})", _NF_LOG),
        ("ped.x.gdp_pc__log", "ln(real GDP per capita)",
         f"=LN('{inputs_sheet}'!{{c}}{input_row_gdppc})", _NF_LOG),
        ("ped.x.unemp__level", "Unemployment rate (fraction)",
         f"='{inputs_sheet}'!{{c}}{input_row_unemp}", "0.000"),
    )
    for key, label, template, number_format in link_transforms:
        grid.register(key, row)
        _set_label(grid, row, label, indent=1)
        _formula_quarters(grid, row, template, font=styles.link, number_format=number_format,
                          quarters=ped_quarters)
        row += 1
    for key, label, number_format in (
        ("ped.x.time__trend", "Trend index t = (year - 2000) x 4 + quarter", _NF_INT),
        ("ped.x.time__post2011_trend", "Post-2011 trend (t if year >= 2011)", _NF_INT),
        ("ped.x.time__post2020", "Post-2020 dummy", _NF_INT),
        ("ped.x.time__covid2020", "COVID-2020 dummy", _NF_INT),
        ("ped.x.time__q2", "Q2 dummy", _NF_INT),
        ("ped.x.time__q3", "Q3 dummy", _NF_INT),
        ("ped.x.time__q4", "Q4 dummy", _NF_INT),
    ):
        feature = key.removeprefix("ped.x.")
        grid.register(key, row)
        _set_label(grid, row, label, indent=1)
        _paint_quarters(
            grid, row,
            {q: data.ped.features[q][feature] for q in ped_quarters},
            font=styles.input, number_format=number_format, skip_zero=True,
        )
        row += 1

    _set_label(grid, row, "Coefficient contributions (coefficient x input)", font=styles.subhead)
    row += 1
    grid.register("ped.c.intercept", row)
    _set_label(grid, row, "β0 (intercept)", indent=1)
    _formula_quarters(grid, row, "=PED_B0", font=styles.formula, number_format=_NF_LOG,
                      quarters=ped_quarters)
    row += 1
    ped_symbols = {
        "PED_BPet": "β petrol", "PED_BGDP": "β GDP", "PED_BUnemp": "β unemp",
        "PED_BTrend": "β trend", "PED_BPost11": "β post-2011",
        "PED_BPost20": "β post-2020", "PED_BCovid": "β COVID",
        "PED_BQ2": "β Q2", "PED_BQ3": "β Q3", "PED_BQ4": "β Q4",
    }
    for feature, name in PED_COEFFICIENT_NAMES:
        source_row = grid.row(f"ped.x.{feature}")
        grid.register(f"ped.c.{feature}", row)
        _set_label(grid, row, f"{ped_symbols[name]} x input", indent=1)
        nonzero = [q for q in ped_quarters if abs(data.ped.features[q][feature]) > 1e-12]
        _formula_quarters(grid, row, f"={name}*{{c}}{source_row}",
                          font=styles.formula, number_format=_NF_LOG, quarters=nonzero)
        row += 1
    lag_row = row
    grid.register("ped.c.lag", row)
    _set_label(grid, row, "β lag x lagged log VKT per capita", indent=1)
    row += 1

    _set_label(grid, row, "AR(1) recursion and raw forecast", font=styles.subhead)
    row += 1
    grid.register("ped.log_linear", row)
    _set_label(grid, row, "Log-linear sum of contributions (incl. lag term)", indent=1)
    contribution_first = grid.row("ped.c.intercept")
    _formula_quarters(
        grid, row,
        f"=SUM({{c}}{contribution_first}:{{c}}{lag_row})",
        font=styles.formula, number_format=_NF_LOG, quarters=ped_quarters,
    )
    row += 1
    grid.register("ped.ar_error", row)
    _set_label(grid, row, "AR(1) error: ρ x previous error (seed = committed residual)", indent=1)
    seed_cell = ws.cell(row=row, column=grid.col_of[seed_quarter])
    seed_cell.value = "=PED_LastErr"
    seed_cell.font = styles.formula
    seed_cell.number_format = _NF_LOG
    _formula_quarters(grid, row, "=PED_Rho*{cprev}" + str(row),
                      font=styles.formula, number_format=_NF_LOG, quarters=ped_quarters)
    row += 1
    grid.register("ped.log_forecast", row)
    _set_label(grid, row, "Log forecast = log-linear + AR(1) error", indent=1)
    log_linear_row = grid.row("ped.log_linear")
    ar_row = grid.row("ped.ar_error")
    _formula_quarters(grid, row, f"={{c}}{log_linear_row}+{{c}}{ar_row}",
                      font=styles.formula, number_format=_NF_LOG, quarters=ped_quarters)
    seed_cell = ws.cell(row=row, column=grid.col_of[seed_quarter])
    seed_cell.value = "=PED_SeedLogVKT"
    seed_cell.font = styles.formula
    seed_cell.number_format = _NF_LOG
    log_forecast_row = row
    row += 1
    _formula_quarters(grid, lag_row, f"=PED_BLagY*{{cprev}}{log_forecast_row}",
                      font=styles.formula, number_format=_NF_LOG, quarters=ped_quarters)
    grid.register("ped.raw_level", row)
    _set_label(grid, row, "Raw PED VKT per capita (km) = EXP(log forecast)",
               font=styles.label_bold, indent=1)
    _formula_quarters(grid, row, f"=EXP({{c}}{log_forecast_row})",
                      font=styles.formula_bold, number_format=_NF_KM, quarters=ped_quarters)
    row += 2

    # ------------------------------------------------------ B. Light RUC
    _section_header(grid, row, "B. Light RUC model - OLS base (coefficients) + exact residual-GBR import")
    row += 1
    light_quarters = [q for q in grid.quarters if q in data.light.raw_prediction]
    _set_label(grid, row, "Transformed inputs", font=styles.subhead)
    row += 1
    light_links = {
        "log_real_diesel_price": ("ln(real diesel price)", "in.real_diesel_price_cents_per_litre"),
        "log_real_light_ruc_price": ("ln(real Light RUC price)", "in.real_light_ruc_price_nzd_per_1000km"),
        "log_lagged_real_light_ruc_price": ("ln(lagged Light RUC price)", "in.lagged_real_light_ruc_price_nzd_per_1000km"),
        "log_real_gdp": ("ln(real GDP)", "in.real_gdp_sa_nzd"),
    }
    for feature, (label, input_key) in light_links.items():
        source_row = data._inputs_grid.row(input_key)
        grid.register(f"light.x.{feature}", row)
        _set_label(grid, row, label, indent=1)
        _formula_quarters(grid, row, f"=LN('{inputs_sheet}'!{{c}}{source_row})",
                          font=styles.link, number_format=_NF_LOG, quarters=light_quarters)
        row += 1
    for feature, label in (
        ("post_2020_dummy", "Post-2020 dummy (from CY2021)"),
        ("q2_dummy", "Q2 dummy"),
        ("q3_dummy", "Q3 dummy"),
        ("q4_dummy", "Q4 dummy"),
    ):
        grid.register(f"light.x.{feature}", row)
        _set_label(grid, row, label, indent=1)
        _paint_quarters(
            grid, row,
            {q: data.light.features[q][feature] for q in light_quarters},
            font=styles.input, number_format=_NF_INT, skip_zero=True,
        )
        row += 1
    _set_label(grid, row, "OLS coefficient contributions", font=styles.subhead)
    row += 1
    grid.register("light.c.intercept", row)
    _set_label(grid, row, "β0 (OLS intercept)", indent=1)
    _formula_quarters(grid, row, "=LR_B0", font=styles.formula, number_format=_NF_LOG,
                      quarters=light_quarters)
    row += 1
    light_symbols = {
        "LR_BDiesel": "β diesel", "LR_BRuc": "β RUC", "LR_BRucLag": "β RUC lag",
        "LR_BGDP": "β GDP", "LR_BPost20": "β post-2020",
        "LR_BQ2": "β Q2", "LR_BQ3": "β Q3", "LR_BQ4": "β Q4",
    }
    for feature, name in LIGHT_COEFFICIENT_NAMES:
        source_row = grid.row(f"light.x.{feature}")
        grid.register(f"light.c.{feature}", row)
        _set_label(grid, row, f"{light_symbols[name]} x input", indent=1)
        nonzero = [q for q in light_quarters if abs(data.light.features[q][feature]) > 1e-12]
        _formula_quarters(grid, row, f"={name}*{{c}}{source_row}",
                          font=styles.formula, number_format=_NF_LOG, quarters=nonzero)
        row += 1
    grid.register("light.base_log", row)
    _set_label(grid, row, "OLS base log prediction = sum of contributions", indent=1)
    first = grid.row("light.c.intercept")
    _formula_quarters(grid, row, f"=SUM({{c}}{first}:{{c}}{row - 1})",
                      font=styles.formula, number_format=_NF_LOG, quarters=light_quarters)
    row += 1
    grid.register("light.residual_log", row)
    _set_label(grid, row, "Residual-GBR log component (exact import)", indent=1)
    residual_row = data._inputs_grid.row("cmp.light_residual_log")
    _formula_quarters(grid, row, f"='{inputs_sheet}'!{{c}}{residual_row}",
                      font=styles.link, number_format=_NF_LOG, quarters=light_quarters)
    row += 1
    grid.register("light.total_log", row)
    _set_label(grid, row, "Total log prediction = OLS base + residual GBR", indent=1)
    base_row = grid.row("light.base_log")
    resid_row = grid.row("light.residual_log")
    _formula_quarters(grid, row, f"={{c}}{base_row}+{{c}}{resid_row}",
                      font=styles.formula, number_format=_NF_LOG, quarters=light_quarters)
    row += 1
    grid.register("light.raw_mkm", row)
    _set_label(grid, row, "Raw Light RUC forecast (m km) = EXP(total log) / 1e6",
               font=styles.label_bold, indent=1)
    total_row = grid.row("light.total_log")
    _formula_quarters(grid, row, f"=EXP({{c}}{total_row})/1000000",
                      font=styles.formula_bold, number_format=_NF_VALUE, quarters=light_quarters)
    row += 2

    # ------------------------------------------------------ C. Heavy RUC
    _section_header(grid, row, "C. Heavy RUC model - governed level-space weighted ensemble")
    row += 1
    heavy_quarters = [q for q in grid.quarters if q in data.heavy.raw_prediction]
    for index, member in enumerate(data.heavy.components, start=1):
        grid.register(f"heavy.m{index}_level", row)
        kind = "Ridge" if member.kind == "ridge" else "GBR"
        _set_label(grid, row, f"{member.label} ({kind}) level (km) = EXP(imported log)", indent=1)
        log_row = data._inputs_grid.row(f"cmp.heavy_{member.label}_log")
        _formula_quarters(grid, row, f"=EXP('{inputs_sheet}'!{{c}}{log_row})",
                          font=styles.link, number_format=_NF_VALUE, quarters=heavy_quarters)
        row += 1
    grid.register("heavy.raw_mkm", row)
    weight_terms = "+".join(
        f"HR_W{index}*{{c}}{grid.row(f'heavy.m{index}_level')}"
        for index in range(1, len(data.heavy.components) + 1)
    )
    _set_label(grid, row, "Raw Heavy RUC forecast (m km) = (w1 x M1 + w2 x M2 + w3 x M3) / 1e6",
               font=styles.label_bold, indent=1)
    _formula_quarters(grid, row, f"=({weight_terms})/1000000",
                      font=styles.formula_bold, number_format=_NF_VALUE, quarters=heavy_quarters)
    row += 2

    # --------------------------------------- D. policy / scenario overlay
    _section_header(grid, row, "D. Scenario and rate-path overlays - selected path demand response")
    row += 1
    _set_label(
        grid, row,
        f"Selected path: {data.policy.state_label}. Calibrated activity = raw "
        "reference forecast x (price ratio) ^ elasticity x GDP factor. Central "
        "baseline: GDP factor = 1. Beyond FY2030 the displayed activity carries "
        "no rate-path demand response; revenue carries the governed rate ratio.",
        font=styles.note,
    )
    row += 1

    _set_label(grid, row, "PED - total pump-price basis", font=styles.subhead)
    row += 1
    ped_free_row = data._inputs_grid.row("in.policy_free_source_nominal_petrol_cpl")
    ped_wedge_row = data._inputs_grid.row("in.policy_published_fed_wedge_nominal_cpl")
    ped_target_row = data._inputs_grid.row("in.policy_target_nominal_petrol_cpl")
    grid.register("pol.ped.reference_pump", row)
    _set_label(grid, row, "Reference nominal pump price = policy-free + published wedge", indent=1)
    _formula_quarters(
        grid, row,
        f"='{inputs_sheet}'!{{c}}{ped_free_row}+'{inputs_sheet}'!{{c}}{ped_wedge_row}",
        font=styles.link, number_format=_NF_VALUE2, quarters=ped_quarters,
    )
    row += 1
    grid.register("pol.ped.target_pump", row)
    _set_label(grid, row, "Selected-path nominal pump price (committed)", indent=1)
    _formula_quarters(grid, row, f"='{inputs_sheet}'!{{c}}{ped_target_row}",
                      font=styles.link, number_format=_NF_VALUE2, quarters=ped_quarters)
    row += 1
    grid.register("pol.ped.ratio", row)
    _set_label(grid, row, "Pump-price ratio = selected / reference", indent=1)
    ref_row = grid.row("pol.ped.reference_pump")
    tgt_row = grid.row("pol.ped.target_pump")
    _formula_quarters(grid, row, f"={{c}}{tgt_row}/{{c}}{ref_row}",
                      font=styles.formula, number_format="0.0000", quarters=ped_quarters)
    row += 1
    grid.register("pol.ped.factor", row)
    _set_label(grid, row, "Price-response factor = ratio ^ ε PED", indent=1)
    ratio_row = grid.row("pol.ped.ratio")
    _formula_quarters(grid, row, f"={{c}}{ratio_row}^PED_Elas",
                      font=styles.formula, number_format="0.0000", quarters=ped_quarters)
    row += 1
    grid.register("pol.ped.calibrated", row)
    _set_label(grid, row, "Calibrated PED VKT per capita (km)", font=styles.label_bold, indent=1)
    raw_row = grid.row("ped.raw_level")
    factor_row = grid.row("pol.ped.factor")
    _formula_quarters(grid, row, f"={{c}}{raw_row}*{{c}}{factor_row}",
                      font=styles.formula_bold, number_format=_NF_KM, quarters=ped_quarters)
    row += 2

    for stream, raw_key, elasticity_name, label in (
        ("LIGHT_RUC", "light.raw_mkm", "LR_Elas", "Light RUC"),
        ("HEAVY_RUC", "heavy.raw_mkm", "HR_Elas", "Heavy RUC"),
    ):
        _set_label(grid, row, f"{label} - generalized running-cost basis", font=styles.subhead)
        row += 1
        stream_quarters = light_quarters if stream == "LIGHT_RUC" else heavy_quarters
        ref_fuel = data._inputs_grid.row(f"in.{stream}.reference_fuel_cost")
        var_fuel = data._inputs_grid.row(f"in.{stream}.variant_fuel_cost")
        ref_ruc = data._inputs_grid.row(f"in.{stream}.reference_ruc_price")
        var_ruc = data._inputs_grid.row(f"in.{stream}.variant_ruc_price")
        grid.register(f"pol.{stream}.reference_cost", row)
        _set_label(grid, row, "Reference cost ($/1,000 km) = diesel cost + RUC rate", indent=1)
        _formula_quarters(
            grid, row,
            f"='{inputs_sheet}'!{{c}}{ref_fuel}+'{inputs_sheet}'!{{c}}{ref_ruc}",
            font=styles.link, number_format=_NF_VALUE2, quarters=stream_quarters,
        )
        row += 1
        grid.register(f"pol.{stream}.variant_cost", row)
        _set_label(grid, row, "Selected cost ($/1,000 km) = diesel cost + selected RUC rate", indent=1)
        _formula_quarters(
            grid, row,
            f"='{inputs_sheet}'!{{c}}{var_fuel}+'{inputs_sheet}'!{{c}}{var_ruc}",
            font=styles.link, number_format=_NF_VALUE2, quarters=stream_quarters,
        )
        row += 1
        grid.register(f"pol.{stream}.ratio", row)
        _set_label(grid, row, "Running-cost ratio = selected / reference", indent=1)
        rc = grid.row(f"pol.{stream}.reference_cost")
        vc = grid.row(f"pol.{stream}.variant_cost")
        _formula_quarters(grid, row, f"={{c}}{vc}/{{c}}{rc}",
                          font=styles.formula, number_format="0.0000", quarters=stream_quarters)
        row += 1
        grid.register(f"pol.{stream}.factor", row)
        symbol = "ε Light RUC" if stream == "LIGHT_RUC" else "ε Heavy RUC"
        _set_label(grid, row, f"Price-response factor = ratio ^ {symbol}", indent=1)
        rr = grid.row(f"pol.{stream}.ratio")
        _formula_quarters(grid, row, f"={{c}}{rr}^{elasticity_name}",
                          font=styles.formula, number_format="0.0000", quarters=stream_quarters)
        row += 1
        grid.register(f"pol.{stream}.calibrated_mkm", row)
        _set_label(grid, row, f"Calibrated {label} activity (m km)", font=styles.label_bold, indent=1)
        raw_row_s = grid.row(raw_key)
        fr = grid.row(f"pol.{stream}.factor")
        _formula_quarters(grid, row, f"={{c}}{raw_row_s}*{{c}}{fr}",
                          font=styles.formula_bold, number_format=_NF_VALUE, quarters=stream_quarters)
        row += 2

    data._scenario_row_cursor = row


def write_glassbox_workbook(
    data, *, scenario_note: str = "", audit_sheets: bool = False
) -> tuple[bytes, list[str]]:
    """Assemble the workbook; returns (bytes, sheet names).

    ``audit_sheets=True`` adds the Checks sheet plus the committed verification
    rows it compares against - the acceptance harness builds with it, the
    delivered download does not (owner request: checks run at build time and
    stay out of the shipped file).
    """
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    styles = _Styles()
    workbook = Workbook()
    workbook.remove(workbook.active)
    data._audit_sheets = audit_sheets

    quarters = list(data.quarters)
    defined_names: dict[str, str] = {}

    readme_ws = workbook.create_sheet(SHEET_README)
    scenario_ws = workbook.create_sheet(SHEET_SCENARIO)
    inputs_ws = workbook.create_sheet(SHEET_INPUTS)
    params_ws = workbook.create_sheet(SHEET_PARAMS)

    inputs_grid = _SheetGrid(inputs_ws, quarters, styles)
    params_grid = _SheetGrid(params_ws, quarters, styles)
    scenario_grid = _SheetGrid(scenario_ws, quarters, styles)

    data._inputs_grid = inputs_grid
    data._params_grid = params_grid

    _write_inputs_sheet(inputs_grid, data, audit_sheets=audit_sheets)
    _write_params_sheet(params_grid, data, defined_names)
    _write_scenario_sheet(scenario_grid, data)

    from model_dashboard._quarterly_glassbox_writer_sections import (
        write_checks_sheet,
        write_post_model_and_revenue_sections,
        write_readme_sheet,
        write_top_block_links,
    )

    write_post_model_and_revenue_sections(scenario_grid, data)
    write_top_block_links(scenario_grid, data)
    if audit_sheets:
        checks_ws = workbook.create_sheet(SHEET_CHECKS)
        checks_grid = _SheetGrid(checks_ws, quarters, styles)
        write_checks_sheet(checks_grid, scenario_grid, data)
    write_readme_sheet(readme_ws, styles, data, scenario_note)

    for name, target in defined_names.items():
        workbook.defined_names.add(DefinedName(name, attr_text=target))

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.active = workbook.sheetnames.index(SHEET_SCENARIO)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), list(workbook.sheetnames)
