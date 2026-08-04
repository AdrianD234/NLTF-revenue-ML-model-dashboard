"""Template-matched XLSX extract of the Revenue Outlook Single scenario view.

Reproduces the governed BEFU26 forecast-extract layout (``references/BEFU26
revenue forecast.xlsx``, sheet ``Baseline``, rows 1-65) one worksheet per
selected scenario path, populated from the CANONICAL aligned detail frames -
the same computation identity as the Single scenario chart, so every page
setting (official vintage, Fleet efficiency including Custom, PT mode shift,
EV/PHEV uptake basis, 12c policy state, bridge mode, conflict path) is in the
exported numbers.

Layout contract:

* rows 1-65 exactly, row 65 = Total net revenues (m $);
* years in columns B:AY = FY2001-FY2050 (the template's FY2051-FY2055
  columns AZ:BD are removed - the one intentional difference);
* row 2 period labels: ACTUAL through FY2025, ST_FORECAST FY2026-FY2030,
  LT_FORECAST FY2031-FY2050 (carried by the template);
* all values are hardcoded finals (the reference workbook itself carries no
  formulas), so no formula-error token can occur;
* styling, number formats, column widths and freeze panes come from cloning
  the template sheet, never from restyling.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

TEMPLATE_RELATIVE_PATH = Path("references") / "BEFU26 revenue forecast.xlsx"
TEMPLATE_SHEET_NAME = "Baseline"

FIRST_EXTRACT_FY = 2001
LAST_EXTRACT_FY = 2050
#: Column B carries FY2001; FY2050 lands in column AY (index 51).
FIRST_DATA_COLUMN = 2
LAST_DATA_COLUMN = FIRST_DATA_COLUMN + (LAST_EXTRACT_FY - FIRST_EXTRACT_FY)
#: The template runs to FY2055 in column BD (index 56).
TEMPLATE_LAST_COLUMN = 56
EXTRACT_LAST_ROW = 65
TOTAL_ROW_LABEL = "Total net revenues (m $)"

#: Key volumes: Level (rows 5-13).
LEVEL_ROW_SERIES: dict[int, str] = {
    5: "light_ruc_net_km",
    6: "heavy_ruc_net_km",
    7: "light_bev_ruc_net_km",
    8: "heavy_bev_ruc_net_km",
    9: "phev_ruc_net_km",
    10: "ped_volume",
    11: "light_petrol_vkt",
    12: "ped_vkt_per_capita",
    13: "tuc_gtk",
}
#: Annual percentage changes (rows 16-24) derive from the level rows above.
PCT_ROW_BY_LEVEL_ROW: dict[int, int] = {row: row + 11 for row in LEVEL_ROW_SERIES}

#: Revenues and totals. ``coo_revenue`` is the repository name for MR13, and
#: the rollup names follow the line-reconciliation contract proven against
#: the reference workbook cell-for-cell (FY2026 column AA).
REVENUE_ROW_SERIES: dict[int, str] = {
    28: "light_ruc_net_revenue",
    29: "heavy_ruc_net_revenue",
    30: "light_bev_ruc_net_revenue",
    31: "heavy_bev_ruc_net_revenue",
    32: "phev_ruc_net_revenue",
    33: "ruc_refunds",
    34: "gross_ruc_revenue",
    35: "ruc_admin_revenue",
    36: "ruc_revenue_net_admin",
    37: "total_ruc_net_revenue",
    40: "gross_ped_revenue",
    41: "gross_lpg_revenue",
    42: "gross_cng_revenue",
    43: "gross_fed_revenue",
    44: "fed_refunds",
    45: "net_fed_revenue",
    48: "mr1_revenue",
    49: "mr2_revenue",
    50: "coo_revenue",
    51: "gross_mvr_revenue",
    52: "mvr_admin_revenue",
    53: "mvr_revenue_net_admin_coo",
    54: "mvr_refunds",
    55: "net_mvr_revenue",
    58: "tuc_net_revenue",
    61: "total_gross_revenue",
    62: "total_admin_fees",
    63: "total_revenue_net_admin",
    64: "total_refunds",
    65: "total_nltf_net_revenue",
}
ROW_SERIES: dict[int, str] = {**LEVEL_ROW_SERIES, **REVENUE_ROW_SERIES}

#: Readable worksheet names for the governed traces; anything else is
#: sanitised from the trace name itself.
EXTRACT_SHEET_NAMES: dict[str, str] = {
    "Current finalist Base case": "Current Base",
    "Current finalist High population/comparison": "High Population",
    "Middle East conflict: Low": "Conflict Low",
    "Middle East conflict: Medium": "Conflict Medium",
    "Middle East conflict: High": "Conflict High",
    "BEFU26 official": "BEFU26 Official",
    "MBU26 official": "MBU26 Official",
}
#: Traces that are columns of every sheet, not scenario paths of their own.
NON_SCENARIO_TRACES = ("Actual",)

#: Actual-year history is scenario-invariant; paths without their own
#: pre-forecast rows (the conflict reforecasts) read it from Base.
HISTORY_FALLBACK_SOURCE_PATH = "Current finalist Base case"

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


class RevenueOutlookExtractError(ValueError):
    """The extract cannot be built without violating the layout contract."""


@dataclass
class ExtractResult:
    workbook_bytes: bytes
    sheet_names: list[str]
    exported_traces: list[str]
    skipped_traces: list[str]
    blank_rows: dict[str, list[str]] = field(default_factory=dict)


def extract_sheet_name(trace_name: str, taken: set[str]) -> str:
    """A valid, unique, readable worksheet name (max 31 chars)."""
    base = EXTRACT_SHEET_NAMES.get(str(trace_name))
    if base is None:
        base = _INVALID_SHEET_CHARS.sub(" ", str(trace_name)).strip() or "Scenario"
    base = base[:31]
    name = base
    suffix = 2
    while name in taken:
        stem = base[: 31 - len(f" {suffix}")]
        name = f"{stem} {suffix}"
        suffix += 1
    taken.add(name)
    return name


def _value_lookup(frame: pd.DataFrame) -> dict[tuple[str, str, int], float]:
    """(source_path, series_id, FY) -> value for one detail frame."""
    if frame is None or not isinstance(frame, pd.DataFrame) or frame.empty:
        return {}
    required = {"source_path", "series_id", "FY", "value"}
    if not required.issubset(frame.columns):
        return {}
    out: dict[tuple[str, str, int], float] = {}
    fy = pd.to_numeric(frame["FY"], errors="coerce")
    values = pd.to_numeric(frame["value"], errors="coerce")
    paths = frame["source_path"].astype(str)
    series = frame["series_id"].astype(str)
    for path, sid, year, value in zip(paths, series, fy, values):
        if pd.isna(year) or pd.isna(value):
            continue
        key = (path, sid, int(year))
        if key not in out:
            out[key] = float(value)
    return out


def _resolve_source_path(trace_name: str, available: set[str]) -> str | None:
    """Map a chart trace name onto a detail-frame source path."""
    text = str(trace_name)
    if text in available:
        return text
    # Official traces sometimes carry a parenthetical presentation suffix
    # (e.g. "MBU26 official (prior vintage)") that the detail frames do not.
    stripped = re.sub(r"\s*\(.*\)$", "", text).strip()
    if stripped in available:
        return stripped
    return None


def build_revenue_outlook_extract(
    *,
    selected_traces: tuple[str, ...] | list[str],
    line_reconciliation: pd.DataFrame,
    stack_components: pd.DataFrame,
    pack_stack_components: pd.DataFrame,
    template_path: Path,
) -> ExtractResult:
    """One template-cloned worksheet per selected scenario path.

    Values coalesce per (path, series, FY) from, in order: the ALIGNED line
    reconciliation (the final overlaid computation, FY2025+), the aligned
    stack components, the committed pack stack (which carries the
    FY2001-FY2024 history), then the Base-case history for reforecast paths
    that have no pre-forecast rows of their own. A cell with no governed
    value in any source is left blank - never zero-filled - and the whole-row
    blanks are reported per sheet.
    """
    from openpyxl import load_workbook

    template_path = Path(template_path)
    if not template_path.is_file():
        raise RevenueOutlookExtractError(f"extract template missing: {template_path}")

    lookups = [
        _value_lookup(line_reconciliation),
        _value_lookup(stack_components),
        _value_lookup(pack_stack_components),
    ]
    available_paths: set[str] = set()
    for lookup in lookups:
        available_paths.update(path for path, _, _ in lookup)

    ordered_traces = [
        str(trace)
        for trace in selected_traces
        if str(trace) not in NON_SCENARIO_TRACES
    ]
    exportable: list[tuple[str, str]] = []
    skipped: list[str] = []
    for trace in ordered_traces:
        source_path = _resolve_source_path(trace, available_paths)
        if source_path is None:
            skipped.append(trace)
        else:
            exportable.append((trace, source_path))
    if not exportable:
        raise RevenueOutlookExtractError(
            "No selected trace resolves to a governed scenario path; nothing to export."
        )

    def cell_value(source_path: str, series_id: str, year: int) -> float | None:
        for lookup in lookups:
            value = lookup.get((source_path, series_id, year))
            if value is not None:
                return value
        if year < 2026 and source_path != HISTORY_FALLBACK_SOURCE_PATH:
            for lookup in lookups:
                value = lookup.get((HISTORY_FALLBACK_SOURCE_PATH, series_id, year))
                if value is not None:
                    return value
        return None

    workbook = load_workbook(template_path)
    if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
        raise RevenueOutlookExtractError(
            f"template sheet {TEMPLATE_SHEET_NAME!r} missing from {template_path}"
        )
    template_ws = workbook[TEMPLATE_SHEET_NAME]
    if str(template_ws.cell(row=EXTRACT_LAST_ROW, column=1).value) != TOTAL_ROW_LABEL:
        raise RevenueOutlookExtractError(
            f"template row {EXTRACT_LAST_ROW} is not {TOTAL_ROW_LABEL!r}"
        )
    template_freeze = template_ws.freeze_panes

    taken: set[str] = set()
    sheet_names: list[str] = []
    blank_rows: dict[str, list[str]] = {}
    for trace, source_path in exportable:
        worksheet = workbook.copy_worksheet(template_ws)
        worksheet.title = extract_sheet_name(trace, taken)
        if template_freeze:
            worksheet.freeze_panes = template_freeze
        sheet_names.append(worksheet.title)

        # Trim the template's FY2051-FY2055 tail so FY2050 (AY) is final.
        if TEMPLATE_LAST_COLUMN > LAST_DATA_COLUMN:
            worksheet.delete_cols(
                LAST_DATA_COLUMN + 1, TEMPLATE_LAST_COLUMN - LAST_DATA_COLUMN
            )

        levels: dict[int, dict[int, float | None]] = {}
        missing: list[str] = []
        for row, series_id in ROW_SERIES.items():
            wrote_any = False
            row_values: dict[int, float | None] = {}
            for year in range(FIRST_EXTRACT_FY, LAST_EXTRACT_FY + 1):
                column = FIRST_DATA_COLUMN + (year - FIRST_EXTRACT_FY)
                value = cell_value(source_path, series_id, year)
                worksheet.cell(row=row, column=column).value = value
                row_values[year] = value
                wrote_any = wrote_any or value is not None
            if row in LEVEL_ROW_SERIES:
                levels[row] = row_values
            if not wrote_any:
                label = str(worksheet.cell(row=row, column=1).value)
                missing.append(label)
                worksheet.cell(row=row, column=1).comment = _blank_row_comment()

        for level_row, pct_row in PCT_ROW_BY_LEVEL_ROW.items():
            row_values = levels.get(level_row, {})
            for year in range(FIRST_EXTRACT_FY, LAST_EXTRACT_FY + 1):
                column = FIRST_DATA_COLUMN + (year - FIRST_EXTRACT_FY)
                current = row_values.get(year)
                prior = row_values.get(year - 1)
                change: float | None = None
                if current is not None and prior is not None and prior != 0.0:
                    change = current / prior - 1.0
                worksheet.cell(row=pct_row, column=column).value = change

        if missing:
            blank_rows[worksheet.title] = missing

    workbook.remove(template_ws)
    workbook.properties.description = (
        "Revenue Outlook forecast extract; layout mirrors the governed BEFU26 "
        "extract with the horizon ending FY2050. "
        + (
            "Rows left blank (no governed value for the path): "
            + "; ".join(
                f"{sheet}: {', '.join(labels)}" for sheet, labels in blank_rows.items()
            )
            if blank_rows
            else "All template rows populated."
        )
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return ExtractResult(
        workbook_bytes=buffer.getvalue(),
        sheet_names=sheet_names,
        exported_traces=[trace for trace, _ in exportable],
        skipped_traces=skipped,
        blank_rows=blank_rows,
    )


def _blank_row_comment():
    from openpyxl.comments import Comment

    return Comment(
        "No governed value is published for this line item on this scenario "
        "path; the cell range is intentionally blank rather than zero-filled.",
        "Revenue Outlook extract",
    )
