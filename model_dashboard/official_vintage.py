"""Generic governed official-forecast-vintage framework.

An "official vintage" is a published MoT/Treasury revenue forecast round
(MBU26, BEFU26, a future PREFU/BEFU/MBU release). Each vintage is registered in
``data/revenue_model_source_pack/official_vintage_registry.json`` and
materialized into a repo-local CSV/Parquet pack; the workbook itself is offline
lineage and is never loaded at Streamlit runtime.

Two distinct vintage roles are governed here and must not be conflated:

- ``official comparator vintage``: the published forecast displayed as the
  external comparator on charts and tables.
- ``bridge assumption vintage``: the source of annual effective rates, fuel
  intensity, administration, refunds, MVR, TUC and fixed-line assumptions used
  to turn Current-model activity into revenue.

Both default to the registry flags ``is_default_comparator`` and
``is_default_bridge_vintage`` and are selectable independently.

The formula authority remains ``FORMULA_DEFINITIONS`` (the governed registry of
row identities); this module asserts those identities per vintage and reports
residuals without force-balancing, exactly as the MBU26 spine does. Published
source inconsistencies are named and retained, never absorbed.

New runtime code must not depend on the MBU26-specific pack-dir/sheet/release
constants; the legacy MBU26 pack is reached through its registry entry instead.
"""

from __future__ import annotations

import hashlib
import json
import math
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd

# The governed formula registry: canonical row identities shared by every
# official vintage. Declared once next to the MBU26 spine (the documented
# authority) and asserted here for all vintages.
from .mbu26_source_spine import FORMULA_DEFINITIONS

OFFICIAL_VINTAGE_REGISTRY_PATH = (
    Path("data") / "revenue_model_source_pack" / "official_vintage_registry.json"
)
OFFICIAL_VINTAGE_REGISTRY_SCHEMA_VERSION = "nltf-official-vintage-registry-v1"
OFFICIAL_VINTAGE_PACK_SCHEMA_VERSION = "nltf-revenue-official-vintage-v1"

# Every pack, legacy or generic, exposes exactly these eight logical files.
PACK_FILE_KEYS: tuple[str, ...] = (
    "annual_spine",
    "official_annual",
    "formula_audit",
    "row_reconciliation",
    "series_alias_audit",
    "series_trace_contract",
    "trace_source_contract",
    "path_trace_status",
)

GENERIC_FILE_STEMS: dict[str, str] = {key: key for key in PACK_FILE_KEYS}

REQUIRED_REGISTRY_FIELDS: tuple[str, ...] = (
    "vintage_id",
    "display_name",
    "release_round",
    "release_date",
    "source_workbook",
    "workbook_sha256",
    "source_sheet",
    "schema_version",
    "actual_start_fy",
    "actual_end_fy",
    "short_forecast_start_fy",
    "short_forecast_end_fy",
    "long_forecast_start_fy",
    "long_forecast_end_fy",
    "source_horizon_fy",
    "source_pack_path",
    "is_latest",
    "is_default_comparator",
    "is_default_bridge_vintage",
    "status",
)

KNOWN_PERIOD_STATUSES: tuple[str, ...] = ("ACTUAL", "ST_FORECAST", "LT_FORECAST")

# Worksheet section anchors. Rows are resolved by exact label match inside the
# anchored section span, never by absolute row index, so a shifted or renamed
# layout fails closed instead of silently mapping the wrong row.
SECTION_ANCHOR_LABELS: tuple[str, ...] = (
    "Key volumes: Level",
    "Key volumes: Annual percentage changes",
    "Revenues: Level ($m ex GST)",
    "Road User Charges",
    "Fuel Excise Duties",
    "Motor Vehicle Register",
    "Track User Charges",
    "TOTALS",
)

SECTION_TO_ANCHOR: dict[str, str] = {
    "Key volumes": "Key volumes: Level",
    "RUC": "Road User Charges",
    "FED": "Fuel Excise Duties",
    "MVR": "Motor Vehicle Register",
    "TUC": "Track User Charges",
    "Totals": "TOTALS",
}

PERCENTAGE_CHANGE_ANCHOR = "Key volumes: Annual percentage changes"

# Canonical series definitions, label-keyed. ``source_label`` is the exact
# worksheet label (units included); the same wording is used by MBU26 and
# BEFU26 and is the required contract for future vintages with this schema.
# Order mirrors the MBU26 ROW_DEFINITIONS; a static test binds the two tables.
CANONICAL_SERIES_DEFINITIONS: tuple[dict[str, Any], ...] = (
    {"source_label": "Light RUC net km (m km)", "series_id": "light_ruc_net_km", "display_name": "Light RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "Heavy RUC net km (m km)", "series_id": "heavy_ruc_net_km", "display_name": "Heavy RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "Light BEV RUC net km (m km)", "series_id": "light_bev_ruc_net_km", "display_name": "Light BEV RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "Heavy BEV RUC net km (m km)", "series_id": "heavy_bev_ruc_net_km", "display_name": "Heavy BEV RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "PHEV RUC net km (m km)", "series_id": "phev_ruc_net_km", "display_name": "PHEV RUC net km", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "PED volume (m L)", "series_id": "ped_volume", "display_name": "PED volume", "section": "Key volumes", "unit": "million litres", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "Light petrol VKT (m km)", "series_id": "light_petrol_vkt", "display_name": "Light petrol VKT", "section": "Key volumes", "unit": "million km", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "Light petrol VKT per capita (km)", "series_id": "ped_vkt_per_capita", "source_series_id": "light_petrol_vkt_per_capita", "display_name": "PED VKT per capita", "source_display_name": "Light petrol VKT per capita", "section": "Key volumes", "unit": "km/person", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "TUC GTK (Tonne-km)", "series_id": "tuc_gtk", "display_name": "TUC GTK", "section": "Key volumes", "unit": "tonne-km", "metric_type": "activity", "row_role": "bridge_input"},
    {"source_label": "Light RUC net revenue (m $)", "series_id": "light_ruc_net_revenue", "display_name": "Light RUC revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "Heavy RUC net revenue (m $)", "series_id": "heavy_ruc_net_revenue", "display_name": "Heavy RUC revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "Light BEV RUC net revenue (m $)", "series_id": "light_bev_ruc_net_revenue", "display_name": "Light BEV RUC net revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "Heavy BEV RUC net revenue (m $)", "series_id": "heavy_bev_ruc_net_revenue", "display_name": "Heavy BEV RUC net revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "PHEV RUC net revenue (m $)", "series_id": "phev_ruc_net_revenue", "display_name": "PHEV RUC net revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "RUC refunds (m $)", "series_id": "ruc_refunds", "display_name": "RUC refunds", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"source_label": "Gross RUC revenue (m $)", "series_id": "gross_ruc_revenue", "display_name": "Gross RUC revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "RUC admin revenue (m $)", "series_id": "ruc_admin_revenue", "display_name": "RUC admin revenue", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"source_label": "RUC revenues net of admin fees (m $)", "series_id": "ruc_revenue_net_admin", "display_name": "RUC revenues net of admin fees", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "RUC revenues net of admin fees & refunds (m $)", "series_id": "total_ruc_net_revenue", "display_name": "Total RUC all classes", "section": "RUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "Gross PED revenue (m $)", "series_id": "gross_ped_revenue", "display_name": "PED revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "Gross LPG revenue (m $)", "series_id": "gross_lpg_revenue", "display_name": "Gross LPG revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "Gross CNG revenue (m $)", "series_id": "gross_cng_revenue", "display_name": "Gross CNG revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "Gross FED revenue (m $)", "series_id": "gross_fed_revenue", "display_name": "Gross FED revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "FED refunds (m $)", "series_id": "fed_refunds", "display_name": "FED refunds", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"source_label": "Net FED revenue (m $)", "series_id": "net_fed_revenue", "display_name": "Net FED revenue", "section": "FED", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "MR1 revenue (m $)", "series_id": "mr1_revenue", "display_name": "MR1 revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "MR2 revenue (m $)", "series_id": "mr2_revenue", "display_name": "MR2 revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "MR13 revenue (m $)", "series_id": "coo_revenue", "display_name": "MR13/COO revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"source_label": "Gross MVR revenue (m $)", "series_id": "gross_mvr_revenue", "display_name": "Gross MVR revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "MVR admin revenue (m $)", "series_id": "mvr_admin_revenue", "display_name": "MVR admin revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"source_label": "MVR revenues net of admin fees & COO (m $)", "series_id": "mvr_revenue_net_admin_coo", "display_name": "MVR revenues net of admin fees and COO", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "MVR refunds (m $)", "series_id": "mvr_refunds", "display_name": "MVR refunds", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "deduction"},
    {"source_label": "MVR revenues net of admin fees, refunds & COO (m $)", "series_id": "net_mvr_revenue", "display_name": "Net MVR revenue", "section": "MVR", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "TUC net revenue (m $)", "series_id": "tuc_net_revenue", "display_name": "TUC net revenue", "section": "TUC", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "leaf"},
    {"source_label": "Total gross revenues (m $)", "series_id": "total_gross_revenue", "display_name": "Total gross revenues", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "Total admin fees (m $)", "series_id": "total_admin_fees", "display_name": "Total admin fees", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "Total revenues net of admin fees (m $)", "series_id": "total_revenue_net_admin", "display_name": "Total revenues net of admin fees", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "Total refunds (m $)", "series_id": "total_refunds", "display_name": "Total refunds", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
    {"source_label": "Total net revenues (m $)", "series_id": "total_nltf_net_revenue", "display_name": "Total NLTF revenue", "section": "Totals", "unit": "$m nominal ex GST", "metric_type": "revenue", "row_role": "aggregate"},
)

# Alias-audit templates; ``{release}`` renders per vintage.
SERIES_ALIAS_TEMPLATES: tuple[dict[str, str], ...] = (
    {"source_label": "Light petrol VKT per capita (km)", "source_series_id": "light_petrol_vkt_per_capita", "runtime_series_id": "ped_vkt_per_capita", "dashboard_label": "PED VKT per capita", "unit": "km/person", "alias_reason": "The {release} worksheet label is Light petrol VKT per capita; the dashboard and current finalist model use PED VKT per capita for the same annual activity concept.", "status": "canonical_mapping"},
    {"source_label": "PED VKT per capita", "source_series_id": "ped_vkt_per_capita", "runtime_series_id": "ped_vkt_per_capita", "dashboard_label": "PED VKT per capita", "unit": "km/person", "alias_reason": "Current finalist source rows already use the dashboard canonical series.", "status": "already_canonical"},
    {"source_label": "Gross PED revenue", "source_series_id": "gross_ped_revenue", "runtime_series_id": "gross_ped_revenue", "dashboard_label": "PED revenue", "unit": "$m nominal ex GST", "alias_reason": "Gross PED revenue and PED revenue are display aliases for the same {release} gross PED row.", "status": "display_alias"},
    {"source_label": "RUC revenues net of admin fees & refunds", "source_series_id": "total_ruc_net_revenue", "runtime_series_id": "total_ruc_net_revenue", "dashboard_label": "Total RUC all classes", "unit": "$m nominal ex GST", "alias_reason": "The dashboard selector keeps the Total RUC all classes label; the {release} source wording describes the same net-admin/refunds formula.", "status": "canonical_mapping"},
    {"source_label": "MR13 revenue", "source_series_id": "coo_revenue", "runtime_series_id": "coo_revenue", "dashboard_label": "MR13", "unit": "$m nominal ex GST", "alias_reason": "MR13 and COO revenue are label variants for the same {release} deduction component.", "status": "display_alias"},
    {"source_label": "MR13/COO revenue", "source_series_id": "coo_revenue", "runtime_series_id": "coo_revenue", "dashboard_label": "MR13", "unit": "$m nominal ex GST", "alias_reason": "The original {release} source label is retained in provenance while the compact dashboard line label is MR13.", "status": "display_alias"},
    {"source_label": "MVR revenues net of admin fees & COO", "source_series_id": "mvr_revenue_net_admin_coo", "runtime_series_id": "mvr_revenue_net_admin_coo", "dashboard_label": "MVR net admin & COO", "unit": "$m nominal ex GST", "alias_reason": "Long-form {release} source label and compact dashboard label describe the same MVR net-admin/COO component.", "status": "display_alias"},
    {"source_label": "MVR revenues net of admin fees, refunds & COO", "source_series_id": "net_mvr_revenue", "runtime_series_id": "net_mvr_revenue", "dashboard_label": "MVR net admin/refunds/COO", "unit": "$m nominal ex GST", "alias_reason": "This label variant includes refunds and maps to the final Net MVR revenue canonical row.", "status": "canonical_mapping"},
    {"source_label": "Total net revenues", "source_series_id": "total_nltf_net_revenue", "runtime_series_id": "total_nltf_net_revenue", "dashboard_label": "Total NLTF revenue", "unit": "$m nominal ex GST", "alias_reason": "Total net revenues is the {release} source wording for the final Total NLTF revenue line.", "status": "canonical_mapping"},
)


class OfficialVintageError(ValueError):
    """Raised when the registry, workbook or pack fails a governed check."""


def official_comparator_trace_name(release_round: str) -> str:
    """Chart trace name for a published vintage, e.g. ``BEFU26 official``."""
    return f"{release_round} official"


def official_comparator_scenario_name(vintage_id: str) -> str:
    """Chart scenario name for a published vintage, e.g. ``befu26_official``."""
    return f"{str(vintage_id).lower()}_official"


def relabel_official_release_provenance(
    frame: pd.DataFrame,
    entry: dict[str, Any],
) -> pd.DataFrame:
    """Re-stamp MBU26-templated provenance text for another official vintage.

    The revenue bridge, extrapolation and reconciliation builders template
    every provenance string, source filename, trace/scenario identifier and
    formula note on the MBU26 release token. When a different registered
    vintage supplies the same data, this pass rewrites those tokens in one
    governed place so provenance can never claim MBU26 for values that came
    from another vintage. Values (numeric columns) are never touched.
    """
    release = str(entry["release_round"])
    vid = str(entry["vintage_id"])
    if vid == "MBU26" and release == "MBU26":
        return frame
    if frame is None or not isinstance(frame, pd.DataFrame) or frame.empty:
        return frame
    stems = _entry_file_stems(entry)
    replacements = (
        ("mbu26_official_annual.csv", f"{stems['official_annual']}.csv"),
        ("mbu26_annual_spine.csv", f"{stems['annual_spine']}.csv"),
        ("mbu26_formula_audit.csv", f"{stems['formula_audit']}.csv"),
        ("mbu26_", f"{vid.lower()}_"),
        ("MBU26", release),
    )

    def _relabel(value: Any) -> Any:
        if not isinstance(value, str):
            return value
        for old, new in replacements:
            if old in value:
                value = value.replace(old, new)
        return value

    out = frame.copy()
    for column in out.columns:
        dtype = out[column].dtype
        if dtype == object or pd.api.types.is_string_dtype(dtype):
            mapped = out[column].map(_relabel, na_action="ignore")
            try:
                out[column] = mapped.astype(dtype)
            except (TypeError, ValueError):
                out[column] = mapped
    return out


@dataclass(frozen=True)
class OfficialVintagePack:
    vintage_id: str
    display_name: str
    registry_entry: dict[str, Any]
    pack_dir: Path
    manifest: dict[str, Any]
    annual_spine: pd.DataFrame
    official_annual: pd.DataFrame
    formula_audit: pd.DataFrame
    row_reconciliation: pd.DataFrame
    series_alias_audit: pd.DataFrame
    series_trace_contract: pd.DataFrame
    trace_source_contract: pd.DataFrame
    path_trace_status: pd.DataFrame

    @property
    def source_horizon_fy(self) -> int:
        return int(self.registry_entry["source_horizon_fy"])

    @property
    def release_round(self) -> str:
        return str(self.registry_entry["release_round"])


def repo_root_from_here() -> Path:
    return Path(__file__).resolve().parents[1]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------


def registry_path(repo_root: Path | str | None = None) -> Path:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    return root / OFFICIAL_VINTAGE_REGISTRY_PATH


def load_official_vintage_registry(
    repo_root: Path | str | None = None,
    *,
    path: Path | str | None = None,
) -> dict[str, Any]:
    target = Path(path) if path is not None else registry_path(repo_root)
    if not target.exists():
        raise OfficialVintageError(f"official vintage registry missing: {target}")
    registry = json.loads(target.read_text(encoding="utf-8"))
    issues = validate_official_vintage_registry(registry)
    if issues:
        raise OfficialVintageError(
            "official vintage registry failed validation: " + "; ".join(issues)
        )
    return registry


def validate_official_vintage_registry(registry: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    if str(registry.get("schema_version")) != OFFICIAL_VINTAGE_REGISTRY_SCHEMA_VERSION:
        issues.append(
            f"schema_version must be {OFFICIAL_VINTAGE_REGISTRY_SCHEMA_VERSION}"
        )
    vintages = registry.get("vintages")
    if not isinstance(vintages, list) or not vintages:
        return issues + ["vintages must be a non-empty list"]
    seen: set[str] = set()
    for entry in vintages:
        if not isinstance(entry, dict):
            issues.append("vintage entries must be objects")
            continue
        vid = str(entry.get("vintage_id") or "")
        missing = [field for field in REQUIRED_REGISTRY_FIELDS if field not in entry]
        if missing:
            issues.append(f"{vid or '<unnamed>'}: missing fields {', '.join(missing)}")
        if vid in seen:
            issues.append(f"duplicate vintage_id {vid}")
        seen.add(vid)
        for flag in ("is_latest", "is_default_comparator", "is_default_bridge_vintage"):
            if flag in entry and not isinstance(entry[flag], bool):
                issues.append(f"{vid}: {flag} must be boolean")
    for flag in ("is_latest", "is_default_comparator", "is_default_bridge_vintage"):
        count = sum(1 for entry in vintages if isinstance(entry, dict) and entry.get(flag) is True)
        if count != 1:
            issues.append(f"exactly one vintage must set {flag}=true (found {count})")
    return issues


def official_vintage_entry(
    vintage_id: str,
    repo_root: Path | str | None = None,
    *,
    registry: dict[str, Any] | None = None,
) -> dict[str, Any]:
    data = registry if registry is not None else load_official_vintage_registry(repo_root)
    for entry in data.get("vintages", []):
        if str(entry.get("vintage_id")) == str(vintage_id):
            return entry
    known = ", ".join(str(item.get("vintage_id")) for item in data.get("vintages", []))
    raise OfficialVintageError(
        f"unknown official vintage {vintage_id!r}; registered vintages: {known}"
    )


def official_vintage_ids(repo_root: Path | str | None = None) -> tuple[str, ...]:
    registry = load_official_vintage_registry(repo_root)
    return tuple(str(entry["vintage_id"]) for entry in registry["vintages"])


def _single_flagged(registry: dict[str, Any], flag: str) -> str:
    for entry in registry.get("vintages", []):
        if entry.get(flag) is True:
            return str(entry["vintage_id"])
    raise OfficialVintageError(f"registry has no vintage with {flag}=true")


def latest_official_vintage_id(repo_root: Path | str | None = None) -> str:
    return _single_flagged(load_official_vintage_registry(repo_root), "is_latest")


def default_comparator_vintage_id(repo_root: Path | str | None = None) -> str:
    return _single_flagged(
        load_official_vintage_registry(repo_root), "is_default_comparator"
    )


def default_bridge_vintage_id(repo_root: Path | str | None = None) -> str:
    return _single_flagged(
        load_official_vintage_registry(repo_root), "is_default_bridge_vintage"
    )


def bridge_vintage_id_from_manifest(
    manifest: dict[str, Any] | None,
    repo_root: Path | str | None = None,
) -> str:
    """The bridge-assumption vintage an ALREADY-BUILT pack was built with.

    This is the authoritative value for every runtime calculation that turns
    Current activity into revenue. The live registry default selects a bridge
    only when a NEW pack is being constructed; it must never override the
    bridge recorded in a pack that already exists, or a pack built on one
    vintage would be re-bridged at runtime on another - the vintage-mixing
    defect this framework exists to prevent.

    Packs predating the ``official_vintages`` manifest block fall back to the
    registry default so an older pack still loads.
    """
    block = manifest.get("official_vintages") if isinstance(manifest, dict) else None
    vid = str((block or {}).get("bridge_assumption_vintage_id") or "").strip()
    if vid:
        return vid
    return default_bridge_vintage_id(repo_root)


def comparator_vintage_id_from_manifest(
    manifest: dict[str, Any] | None,
    repo_root: Path | str | None = None,
) -> str:
    """The default official comparator vintage recorded in a built pack."""
    block = manifest.get("official_vintages") if isinstance(manifest, dict) else None
    vid = str((block or {}).get("official_comparator_vintage_id") or "").strip()
    if vid:
        return vid
    return default_comparator_vintage_id(repo_root)


def official_vintage_spine_frame(
    vintage_id: str,
    repo_root: Path | str | None = None,
) -> pd.DataFrame:
    """FY x series_id numeric pivot of one vintage's official annual rows."""
    pack = load_official_vintage(vintage_id, repo_root=repo_root)
    if pack is None:
        raise OfficialVintageError(
            f"official vintage {vintage_id} is registered but not materialized"
        )
    return pack.official_annual.pivot_table(
        index="FY", columns="series_id", values="value", aggfunc="first"
    ).apply(pd.to_numeric, errors="coerce")


def official_vintage_pack_files(
    vintage_id: str,
    repo_root: Path | str | None = None,
) -> tuple[Path, ...]:
    """Pack CSV paths for a vintage, for cache-signature stat'ing.

    Returned rather than hard-coded at call sites so a cache key follows
    whichever vintage is actually in use.
    """
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    entry = official_vintage_entry(vintage_id, root)
    stems = _entry_file_stems(entry)
    base = root / str(entry["source_pack_path"])
    return tuple(base / f"{stems[key]}.csv" for key in ("annual_spine", "official_annual"))


def official_vintage_choices(
    repo_root: Path | str | None = None,
) -> tuple[tuple[str, str], ...]:
    """(vintage_id, display_name) pairs, default comparator first."""
    registry = load_official_vintage_registry(repo_root)
    default = _single_flagged(registry, "is_default_comparator")
    entries = list(registry["vintages"])
    entries.sort(key=lambda entry: (str(entry["vintage_id"]) != default,))
    return tuple(
        (str(entry["vintage_id"]), str(entry["display_name"]))
        for entry in entries
        if entry.get("available", True)
    )


def resolve_official_vintage_selection(
    selection: str | None,
    repo_root: Path | str | None = None,
) -> str:
    registry = load_official_vintage_registry(repo_root)
    if selection is None or not str(selection).strip():
        return _single_flagged(registry, "is_default_comparator")
    vid = str(selection).strip()
    official_vintage_entry(vid, registry=registry)
    return vid


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------


def _entry_file_stems(entry: dict[str, Any]) -> dict[str, str]:
    stems = entry.get("file_stems")
    if isinstance(stems, dict) and stems:
        resolved = {key: str(stems[key]) for key in PACK_FILE_KEYS if key in stems}
        missing = [key for key in PACK_FILE_KEYS if key not in resolved]
        if missing:
            raise OfficialVintageError(
                f"{entry.get('vintage_id')}: file_stems missing {', '.join(missing)}"
            )
        return resolved
    return dict(GENERIC_FILE_STEMS)


def _validate_pack_hashes(base: Path, manifest: dict[str, Any], vintage_id: str) -> None:
    errors = []
    for filename, metadata in sorted((manifest.get("normalized_files") or {}).items()):
        path = base / filename
        expected = str(metadata.get("sha256", "")).strip() if isinstance(metadata, dict) else ""
        if not path.exists():
            errors.append(f"{filename} missing")
        elif expected and sha256(path) != expected:
            errors.append(f"{filename} hash mismatch")
    if errors:
        raise OfficialVintageError(
            f"official vintage {vintage_id} failed pack hash validation: " + "; ".join(errors)
        )


def load_official_vintage(
    vintage_id: str | None = None,
    repo_root: Path | str | None = None,
) -> OfficialVintagePack | None:
    """Load a registered official vintage pack; ``None`` if not materialized.

    ``vintage_id=None`` resolves to the default comparator vintage.
    """
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    registry = load_official_vintage_registry(root)
    vid = (
        str(vintage_id)
        if vintage_id is not None
        else _single_flagged(registry, "is_default_comparator")
    )
    entry = official_vintage_entry(vid, registry=registry)
    pack_dir = root / Path(str(entry["source_pack_path"]))
    manifest_path = pack_dir / "manifest.json"
    if not manifest_path.exists():
        return None
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    _validate_pack_hashes(pack_dir, manifest, vid)
    manifest_workbook = manifest.get("workbook") or {}
    expected_sha = str(entry.get("workbook_sha256") or "").strip().lower()
    if expected_sha and str(manifest_workbook.get("sha256", "")).lower() != expected_sha:
        raise OfficialVintageError(
            f"official vintage {vid}: pack manifest workbook sha256 does not match the registry"
        )
    stems = _entry_file_stems(entry)
    frames = {
        key: pd.read_csv(pack_dir / f"{stems[key]}.csv") for key in PACK_FILE_KEYS
    }
    return OfficialVintagePack(
        vintage_id=vid,
        display_name=str(entry["display_name"]),
        registry_entry=entry,
        pack_dir=pack_dir,
        manifest=manifest,
        annual_spine=frames["annual_spine"],
        official_annual=frames["official_annual"],
        formula_audit=frames["formula_audit"],
        row_reconciliation=frames["row_reconciliation"],
        series_alias_audit=frames["series_alias_audit"],
        series_trace_contract=frames["series_trace_contract"],
        trace_source_contract=frames["trace_source_contract"],
        path_trace_status=frames["path_trace_status"],
    )


# ---------------------------------------------------------------------------
# Workbook extraction (label-driven, fail-closed)
# ---------------------------------------------------------------------------


def _as_int(value: Any) -> int | None:
    try:
        if pd.isna(value):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _entry_layout(entry: dict[str, Any]) -> dict[str, int]:
    layout = entry.get("layout") or {}
    return {
        "year_header_row": int(layout.get("year_header_row", 1)),
        "period_status_row": int(layout.get("period_status_row", 2)),
        "label_column": int(layout.get("label_column", 1)),
    }


def _locate_year_columns(sheet: Any, entry: dict[str, Any]) -> list[dict[str, Any]]:
    layout = _entry_layout(entry)
    year_row = layout["year_header_row"]
    status_row = layout["period_status_row"]
    vid = str(entry["vintage_id"])
    columns: list[dict[str, Any]] = []
    for column in range(1, sheet.max_column + 1):
        year = _as_int(sheet.cell(year_row, column).value)
        if year is None:
            continue
        status = str(sheet.cell(status_row, column).value or "").strip()
        columns.append(
            {
                "column": column,
                "fy": year,
                "period_status": status,
                "source_year_cell": sheet.cell(year_row, column).coordinate,
                "source_status_cell": sheet.cell(status_row, column).coordinate,
            }
        )
    if not columns:
        raise OfficialVintageError(f"{vid}: no fiscal-year columns found in header row {year_row}")
    years = [item["fy"] for item in columns]
    duplicates = sorted({year for year in years if years.count(year) > 1})
    if duplicates:
        raise OfficialVintageError(f"{vid}: duplicate year columns {duplicates}")
    if years != list(range(min(years), min(years) + len(years))):
        raise OfficialVintageError(f"{vid}: year columns are not continuous: {years}")
    unknown = sorted(
        {item["period_status"] for item in columns} - set(KNOWN_PERIOD_STATUSES)
    )
    if unknown:
        raise OfficialVintageError(f"{vid}: unknown Period values {unknown}")
    _validate_period_blocks(columns, entry)
    return columns


def _validate_period_blocks(columns: list[dict[str, Any]], entry: dict[str, Any]) -> None:
    vid = str(entry["vintage_id"])
    blocks = {
        "ACTUAL": ("actual_start_fy", "actual_end_fy"),
        "ST_FORECAST": ("short_forecast_start_fy", "short_forecast_end_fy"),
        "LT_FORECAST": ("long_forecast_start_fy", "long_forecast_end_fy"),
    }
    keys = [key for pair in blocks.values() for key in pair] + ["source_horizon_fy"]
    if any(key not in entry for key in keys):
        # New-entry inference mode: blocks are derived from the workbook after
        # this scan, so there is nothing registered to validate against yet.
        return
    for status, (start_key, end_key) in blocks.items():
        expected = set(range(int(entry[start_key]), int(entry[end_key]) + 1))
        observed = {item["fy"] for item in columns if item["period_status"] == status}
        if observed != expected:
            raise OfficialVintageError(
                f"{vid}: {status} years {sorted(observed)} do not match the registry "
                f"{entry[start_key]}-{entry[end_key]}"
            )
    horizon = max(item["fy"] for item in columns)
    if horizon != int(entry["source_horizon_fy"]):
        raise OfficialVintageError(
            f"{vid}: workbook horizon FY{horizon} does not match registry "
            f"source_horizon_fy FY{entry['source_horizon_fy']}"
        )


def _section_spans(sheet: Any, label_column: int, vid: str) -> dict[str, tuple[int, int]]:
    anchor_rows: list[tuple[int, str]] = []
    labels_seen: dict[str, int] = {}
    for row in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row, label_column).value or "").strip()
        if label in SECTION_ANCHOR_LABELS:
            if label in labels_seen:
                raise OfficialVintageError(
                    f"{vid}: duplicate section anchor {label!r} at rows "
                    f"{labels_seen[label]} and {row}"
                )
            labels_seen[label] = row
            anchor_rows.append((row, label))
    missing = [label for label in SECTION_ANCHOR_LABELS if label not in labels_seen]
    if missing:
        raise OfficialVintageError(f"{vid}: missing section anchors {missing}")
    spans: dict[str, tuple[int, int]] = {}
    for index, (row, label) in enumerate(anchor_rows):
        end = anchor_rows[index + 1][0] - 1 if index + 1 < len(anchor_rows) else sheet.max_row
        spans[label] = (row + 1, end)
    return spans


def _resolve_series_rows(sheet: Any, entry: dict[str, Any]) -> dict[str, int]:
    """Exact-label row resolution within anchored sections; fails closed."""
    layout = _entry_layout(entry)
    label_column = layout["label_column"]
    vid = str(entry["vintage_id"])
    spans = _section_spans(sheet, label_column, vid)
    resolved: dict[str, int] = {}
    problems: list[str] = []
    for definition in CANONICAL_SERIES_DEFINITIONS:
        series_id = str(definition["series_id"])
        expected_label = str(definition["source_label"])
        anchor = SECTION_TO_ANCHOR[str(definition["section"])]
        start, end = spans[anchor]
        matches = [
            row
            for row in range(start, end + 1)
            if str(sheet.cell(row, label_column).value or "").strip() == expected_label
        ]
        if not matches:
            problems.append(
                f"{series_id}: label {expected_label!r} not found in section {anchor!r}"
            )
        elif len(matches) > 1:
            problems.append(
                f"{series_id}: label {expected_label!r} duplicated at rows {matches}"
            )
        else:
            resolved[series_id] = matches[0]
    if problems:
        raise OfficialVintageError(
            f"{vid}: source schema mismatch (missing/renamed/duplicated rows): "
            + "; ".join(problems)
        )
    rows = list(resolved.values())
    duplicate_rows = sorted({row for row in rows if rows.count(row) > 1})
    if duplicate_rows:
        raise OfficialVintageError(
            f"{vid}: multiple series resolved to the same worksheet rows {duplicate_rows}"
        )
    return resolved


def _allowed_missing(entry: dict[str, Any]) -> set[tuple[str, int]]:
    allowed: set[tuple[str, int]] = set()
    for item in entry.get("allowed_missing", []) or []:
        for fy in item.get("fys", []):
            allowed.add((str(item["series_id"]), int(fy)))
    return allowed


def _extract_annual_spine(
    values_sheet: Any,
    formulas_sheet: Any,
    entry: dict[str, Any],
    workbook_name: str,
    workbook_hash: str,
) -> pd.DataFrame:
    vid = str(entry["vintage_id"])
    release = str(entry["release_round"])
    sheet_name = str(entry["source_sheet"])
    layout = _entry_layout(entry)
    year_columns = _locate_year_columns(values_sheet, entry)
    series_rows = _resolve_series_rows(values_sheet, entry)
    allowed_missing = _allowed_missing(entry)
    formula_by_output = {item["output_series_id"]: item["expression"] for item in FORMULA_DEFINITIONS}
    problems: list[str] = []
    rows: list[dict[str, Any]] = []
    for definition in CANONICAL_SERIES_DEFINITIONS:
        runtime_series_id = str(definition["series_id"])
        source_row = series_rows[runtime_series_id]
        source_series_id = str(definition.get("source_series_id") or runtime_series_id)
        dashboard_label = str(definition["display_name"])
        source_display_name = str(definition.get("source_display_name") or dashboard_label)
        workbook_label = str(values_sheet.cell(source_row, layout["label_column"]).value or "").strip()
        for year_info in year_columns:
            column = int(year_info["column"])
            fy = int(year_info["fy"])
            value_cell = values_sheet.cell(source_row, column)
            formula_cell = formulas_sheet.cell(source_row, column)
            source_formula = (
                formula_cell.value
                if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
                else ""
            )
            raw = value_cell.value
            value = pd.to_numeric(pd.Series([raw]), errors="coerce").iloc[0]
            if source_formula and raw is None:
                problems.append(
                    f"{runtime_series_id} FY{fy}: formula {source_formula!r} has no cached value"
                )
            if pd.isna(value):
                if raw is not None:
                    problems.append(
                        f"{runtime_series_id} FY{fy}: non-numeric value {raw!r} at "
                        f"{value_cell.coordinate}"
                    )
                elif (runtime_series_id, fy) not in allowed_missing:
                    problems.append(
                        f"{runtime_series_id} FY{fy}: missing required value at "
                        f"{value_cell.coordinate}"
                    )
            elif not math.isfinite(float(value)):
                problems.append(
                    f"{runtime_series_id} FY{fy}: non-finite value at {value_cell.coordinate}"
                )
            rows.append(
                {
                    "source_release": release,
                    "workbook_basename": workbook_name,
                    "workbook_sha256": workbook_hash,
                    "sheet": sheet_name,
                    "source_row": source_row,
                    "source_cell": value_cell.coordinate,
                    "source_year_cell": year_info["source_year_cell"],
                    "source_status_cell": year_info["source_status_cell"],
                    "source_formula": source_formula,
                    "asserted_formula": formula_by_output.get(runtime_series_id, ""),
                    "FY": fy,
                    "period": f"FY{fy}",
                    "period_status": year_info["period_status"],
                    "source_label": workbook_label,
                    "source_series_id": source_series_id,
                    "runtime_series_id": runtime_series_id,
                    "dashboard_label": dashboard_label,
                    "source_display_name": source_display_name,
                    "label": dashboard_label,
                    "series_id": runtime_series_id,
                    "display_name": dashboard_label,
                    "section": definition["section"],
                    "unit": definition["unit"],
                    "metric_type": definition["metric_type"],
                    "row_role": definition["row_role"],
                    "value": value if pd.notna(value) else pd.NA,
                    "value_status": (
                        "actual"
                        if str(year_info["period_status"]).upper() == "ACTUAL"
                        else "official_forecast"
                    ),
                    "source_kind": "official_source_row",
                }
            )
    if problems:
        raise OfficialVintageError(
            f"{vid}: workbook value validation failed: " + "; ".join(problems[:25])
            + (f"; ... {len(problems) - 25} more" if len(problems) > 25 else "")
        )
    frame = pd.DataFrame(rows).sort_values(["source_row", "FY"], kind="stable").reset_index(drop=True)
    duplicated = frame.duplicated(subset=["series_id", "FY"])
    if duplicated.any():
        raise OfficialVintageError(f"{vid}: duplicate canonical series/year keys extracted")
    return frame


def extract_annual_percentage_changes(
    values_sheet: Any,
    entry: dict[str, Any],
) -> pd.DataFrame:
    """The 'Annual percentage changes' block, for level-consistency audits.

    Not part of the runtime pack; consumed by the reconciliation builder to
    validate published percentage-change rows against the level rows.
    """
    layout = _entry_layout(entry)
    label_column = layout["label_column"]
    vid = str(entry["vintage_id"])
    spans = _section_spans(values_sheet, label_column, vid)
    start, end = spans[PERCENTAGE_CHANGE_ANCHOR]
    year_columns = _locate_year_columns(values_sheet, entry)
    label_to_series = {
        str(item["source_label"]): str(item["series_id"])
        for item in CANONICAL_SERIES_DEFINITIONS
        if item["section"] == "Key volumes"
    }
    rows: list[dict[str, Any]] = []
    for row in range(start, end + 1):
        label = str(values_sheet.cell(row, label_column).value or "").strip()
        if label not in label_to_series:
            continue
        for year_info in year_columns:
            cell = values_sheet.cell(row, int(year_info["column"]))
            value = pd.to_numeric(pd.Series([cell.value]), errors="coerce").iloc[0]
            rows.append(
                {
                    "source_release": str(entry["release_round"]),
                    "series_id": label_to_series[label],
                    "source_label": label,
                    "source_row": row,
                    "source_cell": cell.coordinate,
                    "FY": int(year_info["fy"]),
                    "period_status": year_info["period_status"],
                    "pct_change": value if pd.notna(value) else pd.NA,
                }
            )
    return pd.DataFrame(rows).sort_values(["source_row", "FY"], kind="stable").reset_index(drop=True)


# ---------------------------------------------------------------------------
# Derived frames (formula audit, official annual, contracts)
# ---------------------------------------------------------------------------


def _formula_audit_frame(spine: pd.DataFrame, entry: dict[str, Any]) -> pd.DataFrame:
    release = str(entry["release_round"])
    values = {
        (str(row.series_id), int(row.FY)): pd.to_numeric(row.value, errors="coerce")
        for row in spine.itertuples()
        if pd.notna(row.value)
    }
    labels = {str(item["series_id"]): str(item["display_name"]) for item in CANONICAL_SERIES_DEFINITIONS}
    source_rows = {
        str(row.series_id): str(row.source_row)
        for row in spine[["series_id", "source_row"]].drop_duplicates().itertuples(index=False)
    }
    source_cells = {
        (str(row.series_id), int(row.FY)): str(row.source_cell) for row in spine.itertuples()
    }
    statuses = {
        int(row.FY): str(row.period_status)
        for row in spine[["FY", "period_status"]].drop_duplicates().itertuples(index=False)
    }
    fys = sorted({int(value) for value in spine["FY"].dropna().unique()})
    rows: list[dict[str, Any]] = []
    for formula in FORMULA_DEFINITIONS:
        output = str(formula["output_series_id"])
        terms = tuple(formula["terms"])
        for fy in fys:
            missing = [series_id for series_id, _sign in terms if (series_id, fy) not in values]
            calculated = (
                pd.NA
                if missing
                else sum(float(values[(series_id, fy)]) * sign for series_id, sign in terms)
            )
            observed = values.get(
                (output, fy),
                calculated
                if str(formula.get("source_kind", "")) == "derived_dashboard_subtotal"
                else pd.NA,
            )
            residual = pd.NA
            if pd.notna(calculated) and pd.notna(observed):
                residual = float(observed) - float(calculated)
            status = (
                "missing_inputs"
                if missing
                else "reconciled"
                if pd.notna(residual) and abs(float(residual)) <= 0.05
                else "residual_reported"
            )
            rows.append(
                {
                    "source_release": release,
                    "FY": fy,
                    "period": f"FY{fy}",
                    "period_status": statuses.get(fy, ""),
                    "output_series_id": output,
                    "output_label": str(formula.get("output_label") or labels.get(output, output)),
                    "row_role": "aggregate",
                    "formula": str(formula["expression"]),
                    "source_rows": "; ".join(source_rows.get(series_id, "") for series_id, _sign in terms),
                    "source_cells": "; ".join(source_cells.get((series_id, fy), "") for series_id, _sign in terms),
                    "observed_value": observed,
                    "calculated_value": calculated,
                    "residual": residual,
                    "residual_abs": abs(float(residual)) if pd.notna(residual) else pd.NA,
                    "status": status,
                    "missing_inputs": "; ".join(missing),
                    "source_kind": str(formula.get("source_kind") or "official_source_formula_assertion"),
                }
            )
    frame = pd.DataFrame(rows).sort_values(["FY", "output_series_id"], kind="stable").reset_index(drop=True)
    if frame["status"].eq("missing_inputs").any():
        broken = sorted(frame.loc[frame["status"].eq("missing_inputs"), "output_series_id"].unique())
        raise OfficialVintageError(
            f"{entry['vintage_id']}: formula audit has missing inputs for {broken}"
        )
    return frame


def _official_annual_frame(
    spine: pd.DataFrame, formula_audit: pd.DataFrame, entry: dict[str, Any], stems: dict[str, str]
) -> pd.DataFrame:
    release = str(entry["release_round"])
    base = spine.copy()
    base["source_file"] = f"{stems['annual_spine']}.csv"
    base["formula"] = base["asserted_formula"].where(
        base["asserted_formula"].astype(str).ne(""), base["source_formula"]
    )
    subtotal = formula_audit[formula_audit["output_series_id"].eq("total_fed_ruc_net_revenue")].copy()
    if not subtotal.empty:
        subtotal_rows = pd.DataFrame(
            {
                "source_release": release,
                "workbook_basename": base["workbook_basename"].iloc[0] if not base.empty else "",
                "workbook_sha256": base["workbook_sha256"].iloc[0] if not base.empty else "",
                "sheet": str(entry["source_sheet"]),
                "source_row": "",
                "source_cell": subtotal["source_cells"],
                "source_year_cell": "",
                "source_status_cell": "",
                "source_formula": "",
                "asserted_formula": subtotal["formula"],
                "FY": subtotal["FY"],
                "period": subtotal["period"],
                "period_status": subtotal["period_status"],
                "source_label": f"Derived dashboard subtotal from {release} Net FED and Total RUC",
                "label": "Total RUC+PED revenue",
                "series_id": "total_fed_ruc_net_revenue",
                "display_name": "Total RUC+PED revenue",
                "section": "Derived totals",
                "unit": "$m nominal ex GST",
                "metric_type": "revenue",
                "row_role": "aggregate",
                "value": subtotal["calculated_value"],
                "value_status": subtotal["period_status"].map(
                    lambda value: "actual" if str(value).upper() == "ACTUAL" else "official_forecast"
                ),
                "source_kind": "official_formula_derived_dashboard_subtotal",
                "source_file": f"{stems['row_reconciliation']}.csv",
                "formula": subtotal["formula"],
            }
        )
        base = pd.concat([base, subtotal_rows], ignore_index=True, sort=False)
    return base.sort_values(["series_id", "FY"], kind="stable").reset_index(drop=True)


def _series_alias_audit_frame(spine: pd.DataFrame, entry: dict[str, Any]) -> pd.DataFrame:
    release = str(entry["release_round"])
    columns = [
        "source_label",
        "source_series_id",
        "runtime_series_id",
        "dashboard_label",
        "unit",
        "source_row",
        "source_cell",
        "alias_reason",
        "status",
    ]
    source_cells: dict[tuple[str, str], str] = {}
    source_rows: dict[tuple[str, str], str] = {}
    source_labels: dict[tuple[str, str], str] = {}
    for key, group in spine.groupby(
        [spine["source_series_id"].astype(str), spine["series_id"].astype(str)], dropna=False
    ):
        source_series_id, runtime_series_id = str(key[0]), str(key[1])
        cells = [str(value).strip() for value in group["source_cell"].dropna() if str(value).strip()]
        cell_range = "" if not cells else cells[0] if cells[0] == cells[-1] else f"{cells[0]}:{cells[-1]}"
        source_cells[(source_series_id, runtime_series_id)] = cell_range
        rows_text = [str(value).strip() for value in group["source_row"].dropna() if str(value).strip()]
        source_rows[(source_series_id, runtime_series_id)] = rows_text[0] if rows_text else ""
        labels_text = [str(value).strip() for value in group["source_label"].dropna() if str(value).strip()]
        source_labels[(source_series_id, runtime_series_id)] = labels_text[0] if labels_text else ""
    rows: list[dict[str, Any]] = []
    for item in SERIES_ALIAS_TEMPLATES:
        source_series_id = item["source_series_id"]
        runtime_series_id = item["runtime_series_id"]
        key = (source_series_id, runtime_series_id)
        source_label = source_labels.get(key) if source_series_id != runtime_series_id else ""
        rows.append(
            {
                "source_label": source_label or item["source_label"],
                "source_series_id": source_series_id,
                "runtime_series_id": runtime_series_id,
                "dashboard_label": item["dashboard_label"],
                "unit": item["unit"],
                "source_row": source_rows.get(key, ""),
                "source_cell": source_cells.get(key, ""),
                "alias_reason": item["alias_reason"].format(release=release),
                "status": item["status"],
            }
        )
    return (
        pd.DataFrame(rows, columns=columns)
        .sort_values(["runtime_series_id", "source_label"], kind="stable")
        .reset_index(drop=True)
    )


def _series_bridge_text(series_id: str, release: str) -> str:
    if series_id == "gross_ped_revenue":
        return (
            f"Current PED VKT/capita forecast through {release} population, litres intensity "
            "and gross PED rate."
        )
    if series_id == "light_ruc_net_revenue":
        return f"Current Light RUC net-km forecast multiplied by {release} effective Light RUC rate."
    if series_id == "heavy_ruc_net_revenue":
        return f"Current Heavy RUC net-km forecast multiplied by {release} effective Heavy RUC rate."
    if series_id == "total_fed_ruc_net_revenue":
        return f"Net FED plus Total RUC from {release}/current-finalist hybrid rows."
    if series_id == "total_nltf_net_revenue":
        return f"Net FED plus Total RUC plus {release} Net MVR and TUC."
    return f"{release} source row or direct current finalist activity forecast."


def _series_trace_contract_frame(entry: dict[str, Any], stems: dict[str, str]) -> pd.DataFrame:
    from .mbu26_source_spine import DISPLAY_SERIES_METADATA, REVENUE_PARTIAL_ACTUAL_FY
    from .revenue_source_pack import (
        REVENUE_FIRST_FORECAST_QUARTER,
        REVENUE_LAST_COMPLETE_ACTUAL_FY,
        REVENUE_MODEL_TRAINING_CUTOFF,
    )

    release = str(entry["release_round"])
    vid = str(entry["vintage_id"])
    pack_path = str(entry["source_pack_path"]).replace("\\", "/")
    rows = []
    for item in DISPLAY_SERIES_METADATA:
        metric = item["metric_type"]
        controls = (
            "series; time_grain; horizon"
            if metric == "activity"
            else "series; time_grain; horizon; selected_fy"
        )
        rows.append(
            {
                "series_option": item["display_name"],
                "canonical_id": item["canonical_id"],
                "display_name": item["display_name"],
                "metric_type": metric,
                "unit": item["unit"],
                "valid_bases": "not_applicable" if metric == "activity" else "Nominal ex GST",
                "valid_controls": controls,
                "actual_source": f"{pack_path}/{stems['annual_spine']}.csv",
                "primary_forecast_source": (
                    f"{release} official annual rows plus current-finalist model forecast "
                    "replacements where applicable"
                ),
                "excluded_lineage_source": "Legacy Excel forecast paths are excluded from runtime traces",
                "bridge": _series_bridge_text(item["canonical_id"], release),
                "last_complete_actual_fy": REVENUE_LAST_COMPLETE_ACTUAL_FY,
                "first_forecast_fy": f"FY{REVENUE_PARTIAL_ACTUAL_FY}",
                "first_forecast_quarter": REVENUE_FIRST_FORECAST_QUARTER,
                "model_training_cutoff": REVENUE_MODEL_TRAINING_CUTOFF,
                "availability_status": f"{vid.lower()}_current_runtime_available",
                "interpretation": (
                    "Activity controls hide revenue-only selectors."
                    if metric == "activity"
                    else f"Revenue trace is either {release} official or current-finalist "
                    "replacement-only hybrid."
                ),
            }
        )
    return pd.DataFrame(rows).sort_values("series_option", kind="stable").reset_index(drop=True)


def _trace_source_contract_frame(
    entry: dict[str, Any], stems: dict[str, str], workbook_name: str, workbook_hash: str
) -> pd.DataFrame:
    from .governance_constants import current_composite_model_id
    from .revenue_source_pack import (
        REVENUE_LAST_COMPLETE_ACTUAL_FY,
        REVENUE_MODEL_TRAINING_CUTOFF,
    )

    release = str(entry["release_round"])
    composite = current_composite_model_id()
    shared = {
        "workbook_basename": workbook_name,
        "workbook_sha256": workbook_hash,
    }
    return pd.DataFrame(
        [
            {
                "trace_name": "Actual",
                "trace_type": "Actual",
                "trace_role": "source_actual",
                "source_file": f"{stems['annual_spine']}.csv",
                "model_id": "",
                "cutoff": f"last complete FY{REVENUE_LAST_COMPLETE_ACTUAL_FY}",
                "scenario": "actual",
                "period_status": "ACTUAL",
                "anchor_forecast_flag": "actual",
                "runtime_forecast_source": False,
                "displayed": True,
                **shared,
                "notes": f"Connected grey actual line ends at FY{REVENUE_LAST_COMPLETE_ACTUAL_FY}.",
            },
            {
                "trace_name": f"{release} official",
                "trace_type": f"{release} official",
                "trace_role": "official_external_comparator",
                "source_file": f"{stems['official_annual']}.csv",
                "model_id": "",
                "cutoff": f"{release} ST_FORECAST/LT_FORECAST rows",
                "scenario": release,
                "period_status": "ST_FORECAST; LT_FORECAST",
                "anchor_forecast_flag": "external_comparator",
                "runtime_forecast_source": False,
                "displayed": True,
                **shared,
                "notes": f"Official MOT release comparator from the {release} worksheet only.",
            },
            {
                "trace_name": "Current finalist Base case",
                "trace_type": "current finalist base",
                "trace_role": "in_house_current_finalist",
                "source_file": f"forecast_scenario_comparison.parquet; {stems['official_annual']}.csv",
                "model_id": composite,
                "cutoff": REVENUE_MODEL_TRAINING_CUTOFF,
                "scenario": "current_basecase",
                "period_status": "FY2025 actual anchor; FY2026 nowcast; FY2027+ forecast",
                "anchor_forecast_flag": "actual_anchor_then_forecast",
                "runtime_forecast_source": True,
                "displayed": True,
                **shared,
                "notes": "Finalist model output replaces only PED, Light RUC and Heavy RUC revenue.",
            },
            {
                "trace_name": "Current finalist High population/comparison",
                "trace_type": "current finalist comparison",
                "trace_role": "in_house_current_finalist",
                "source_file": f"forecast_scenario_comparison.parquet; {stems['official_annual']}.csv",
                "model_id": composite,
                "cutoff": REVENUE_MODEL_TRAINING_CUTOFF,
                "scenario": "current_comparison_1",
                "period_status": "FY2025 actual anchor; FY2026 nowcast; FY2027+ forecast",
                "anchor_forecast_flag": "actual_anchor_then_forecast",
                "runtime_forecast_source": True,
                "displayed": True,
                **shared,
                "notes": "Finalist model output replaces only PED, Light RUC and Heavy RUC revenue.",
            },
        ]
    )


def _path_trace_status_frame(entry: dict[str, Any]) -> pd.DataFrame:
    release = str(entry["release_round"])
    vid = str(entry["vintage_id"])
    return pd.DataFrame(
        [
            {
                "trace_id": "actual",
                "trace_label": "Actual",
                "availability_status": "available",
                "plotted": True,
                "data_scope": f"{release} ACTUAL rows through FY{int(entry['actual_end_fy'])}",
                "blocking_gap_id": "",
                "current_selection": release,
                "user_visible_message": f"Actuals are read from the {release} annual source spine.",
            },
            {
                "trace_id": f"{vid.lower()}_official",
                "trace_label": f"{release} official",
                "availability_status": "available",
                "plotted": True,
                "data_scope": f"{release} ST_FORECAST/LT_FORECAST rows",
                "blocking_gap_id": "",
                "current_selection": release,
                "user_visible_message": f"Official comparator is the {release} worksheet only.",
            },
            {
                "trace_id": "current_finalist_base",
                "trace_label": "Current finalist Base case",
                "availability_status": "available",
                "plotted": True,
                "data_scope": "Current finalist model output annualized to June years",
                "blocking_gap_id": "",
                "current_selection": "current_basecase",
                "user_visible_message": "Current finalist base replaces only PED, Light RUC and Heavy RUC revenue.",
            },
            {
                "trace_id": "current_finalist_comparison",
                "trace_label": "Current finalist High population/comparison",
                "availability_status": "available",
                "plotted": True,
                "data_scope": "Current finalist comparison output annualized to June years",
                "blocking_gap_id": "",
                "current_selection": "current_comparison_1",
                "user_visible_message": "Current finalist comparison replaces only PED, Light RUC and Heavy RUC revenue.",
            },
        ]
    )


# ---------------------------------------------------------------------------
# Sentinels
# ---------------------------------------------------------------------------


def _validate_sentinels(spine: pd.DataFrame, entry: dict[str, Any]) -> int:
    sentinels = entry.get("sentinels") or []
    values = {
        (str(row.series_id), int(row.FY)): row.value
        for row in spine.itertuples()
        if pd.notna(row.value)
    }
    problems: list[str] = []
    for sentinel in sentinels:
        key = (str(sentinel["series_id"]), int(sentinel["fy"]))
        expected = float(sentinel["value"])
        observed = values.get(key)
        if observed is None:
            problems.append(f"{key[0]} FY{key[1]}: sentinel value missing from workbook")
        elif not math.isclose(float(observed), expected, rel_tol=1e-12, abs_tol=1e-9):
            problems.append(
                f"{key[0]} FY{key[1]}: observed {float(observed)!r} != pinned {expected!r}"
            )
    if problems:
        raise OfficialVintageError(
            f"{entry['vintage_id']}: sentinel validation failed (possible shifted row "
            "mapping or wrong workbook): " + "; ".join(problems)
        )
    return len(sentinels)


# ---------------------------------------------------------------------------
# Materializer
# ---------------------------------------------------------------------------


def _prepare_frame_for_output(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.copy()
    for column in output.columns:
        if output[column].dtype == object:
            output[column] = output[column].where(output[column].notna(), "").astype(str)
    return output


def _repo_relative(repo_root: Path, path: Path) -> str:
    try:
        return path.resolve().relative_to(repo_root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def _manifest_markdown(manifest: dict[str, Any]) -> str:
    workbook = manifest.get("workbook") or {}
    rows = [
        f"# {manifest.get('source_release')} Official Vintage Pack",
        "",
        f"- Schema: `{manifest.get('schema_version')}`",
        f"- Vintage: `{manifest.get('vintage_id')}`",
        f"- Source release: `{manifest.get('source_release')}`",
        f"- Output: `{manifest.get('repo_relative_output_dir')}`",
        f"- Workbook: `{workbook.get('basename')}`",
        f"- Workbook SHA256: `{workbook.get('sha256')}`",
        f"- Worksheet: `{workbook.get('sheet')}`",
        "",
        "The workbook is offline lineage only. Streamlit reads the repo-local CSV/Parquet extracts.",
        "",
        "## Formula Policy",
        "",
        str(manifest.get("formula_policy") or ""),
    ]
    return "\n".join(rows) + "\n"


def _write_registry(registry: dict[str, Any], path: Path) -> None:
    issues = validate_official_vintage_registry(registry)
    if issues:
        raise OfficialVintageError(
            "refusing to write invalid registry: " + "; ".join(issues)
        )
    path.write_text(json.dumps(registry, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def _infer_period_blocks(columns: list[dict[str, Any]]) -> dict[str, int]:
    blocks: dict[str, int] = {}
    for status, (start_key, end_key) in {
        "ACTUAL": ("actual_start_fy", "actual_end_fy"),
        "ST_FORECAST": ("short_forecast_start_fy", "short_forecast_end_fy"),
        "LT_FORECAST": ("long_forecast_start_fy", "long_forecast_end_fy"),
    }.items():
        years = sorted(item["fy"] for item in columns if item["period_status"] == status)
        if not years:
            raise OfficialVintageError(f"workbook has no {status} year columns")
        if years != list(range(years[0], years[-1] + 1)):
            raise OfficialVintageError(f"{status} year block is discontinuous: {years}")
        blocks[start_key] = years[0]
        blocks[end_key] = years[-1]
    return blocks


def materialize_official_vintage(
    vintage_id: str,
    *,
    workbook_path: Path | str | None = None,
    repo_root: Path | str | None = None,
    output_dir: Path | str | None = None,
    extracted_by: str = "official_vintage_materializer",
    display_name: str | None = None,
    sheet: str | None = None,
    release_round: str | None = None,
    source_pack_path: str | None = None,
    set_latest: bool = False,
    set_default_comparator: bool = False,
    set_default_bridge_vintage: bool = False,
) -> dict[str, Any]:
    """Validate, extract and materialize one official vintage pack.

    Idempotent for an identical workbook; fails closed on a conflicting
    re-materialization (same vintage, different workbook hash), on any schema
    drift, and on sentinel mismatches. Registry defaults flags are only moved
    when explicitly requested via ``set_*``.
    """
    import openpyxl

    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    reg_path = registry_path(root)
    registry = load_official_vintage_registry(root)
    vid = str(vintage_id)
    existing_ids = {str(item["vintage_id"]) for item in registry["vintages"]}

    workbook = (
        Path(workbook_path)
        if workbook_path is not None
        else (
            root / str(official_vintage_entry(vid, registry=registry)["source_workbook"])
            if vid in existing_ids
            else None
        )
    )
    if workbook is None:
        raise OfficialVintageError(f"{vid}: --workbook is required for a new vintage")
    if not workbook.exists():
        raise OfficialVintageError(f"{vid}: workbook not found: {workbook}")
    workbook_hash = sha256(workbook)

    if vid in existing_ids:
        entry = official_vintage_entry(vid, registry=registry)
        if str(entry.get("pack_format", "official_vintage_v1")) != "official_vintage_v1":
            raise OfficialVintageError(
                f"{vid}: pack_format {entry.get('pack_format')!r} is a legacy compatibility "
                "entry; it is not materialized through the generic path"
            )
        expected_sha = str(entry.get("workbook_sha256") or "").strip().lower()
        if expected_sha and workbook_hash.lower() != expected_sha:
            raise OfficialVintageError(
                f"{vid}: workbook sha256 {workbook_hash} does not match the registered "
                f"hash {expected_sha}; refusing to silently ingest a different workbook"
            )
        if sheet is not None and str(sheet) != str(entry["source_sheet"]):
            raise OfficialVintageError(
                f"{vid}: sheet {sheet!r} conflicts with registered sheet "
                f"{entry['source_sheet']!r}"
            )
    else:
        if sheet is None:
            raise OfficialVintageError(f"{vid}: --sheet is required for a new vintage")
        entry = {
            "vintage_id": vid,
            "display_name": str(display_name or f"{vid} official"),
            "release_round": str(release_round or vid),
            "release_date": None,
            "source_workbook": _repo_relative(root, workbook),
            "workbook_sha256": workbook_hash,
            "source_sheet": str(sheet),
            "schema_version": OFFICIAL_VINTAGE_PACK_SCHEMA_VERSION,
            "source_pack_path": str(
                source_pack_path
                or (Path("data") / "revenue_model_source_pack" / "official_vintages" / vid.lower()).as_posix()
            ),
            "pack_format": "official_vintage_v1",
            "file_stems": dict(GENERIC_FILE_STEMS),
            "layout": {"year_header_row": 1, "period_status_row": 2, "label_column": 1},
            "available": True,
            "is_latest": False,
            "is_default_comparator": False,
            "is_default_bridge_vintage": False,
            "status": "registered_official_vintage",
        }

    stems = _entry_file_stems(entry)
    output = (
        Path(output_dir)
        if output_dir is not None
        else root / Path(str(entry["source_pack_path"]))
    )

    # Conflicting re-materialization: an existing pack built from a different
    # workbook must never be silently replaced.
    existing_manifest_path = output / "manifest.json"
    existing_manifest: dict[str, Any] = {}
    if existing_manifest_path.exists():
        try:
            existing_manifest = json.loads(existing_manifest_path.read_text(encoding="utf-8"))
        except Exception:
            existing_manifest = {}
        existing_sha = str((existing_manifest.get("workbook") or {}).get("sha256", "")).lower()
        if existing_sha and existing_sha != workbook_hash.lower():
            raise OfficialVintageError(
                f"{vid}: existing pack at {output} was materialized from workbook sha256 "
                f"{existing_sha}; conflicting re-materialization refused. Remove the pack "
                "directory deliberately if the vintage source is being replaced."
            )

    values_wb = openpyxl.load_workbook(workbook, read_only=False, data_only=True)
    formulas_wb = openpyxl.load_workbook(workbook, read_only=False, data_only=False)
    try:
        sheet_name = str(entry["source_sheet"])
        if sheet_name not in values_wb.sheetnames:
            raise OfficialVintageError(
                f"{vid}: sheet {sheet_name!r} not found in workbook (sheets: "
                f"{values_wb.sheetnames})"
            )
        values_sheet = values_wb[sheet_name]
        formulas_sheet = formulas_wb[sheet_name]
        if vid not in existing_ids:
            entry.update(_infer_period_blocks(_locate_year_columns(values_sheet, entry)))
            entry["source_horizon_fy"] = max(
                item["fy"] for item in _locate_year_columns(values_sheet, entry)
            )
        annual_spine = _extract_annual_spine(
            values_sheet, formulas_sheet, entry, workbook.name, workbook_hash
        )
    finally:
        values_wb.close()
        formulas_wb.close()

    sentinel_count = _validate_sentinels(annual_spine, entry)
    formula_audit = _formula_audit_frame(annual_spine, entry)
    official_annual = _official_annual_frame(annual_spine, formula_audit, entry, stems)
    alias_audit = _series_alias_audit_frame(annual_spine, entry)
    series_contract = _series_trace_contract_frame(entry, stems)
    trace_contract = _trace_source_contract_frame(entry, stems, workbook.name, workbook_hash)
    path_status = _path_trace_status_frame(entry)

    frames = {
        stems["annual_spine"]: annual_spine,
        stems["formula_audit"]: formula_audit,
        stems["official_annual"]: official_annual,
        stems["row_reconciliation"]: formula_audit.copy(),
        stems["series_alias_audit"]: alias_audit,
        stems["series_trace_contract"]: series_contract,
        stems["trace_source_contract"]: trace_contract,
        stems["path_trace_status"]: path_status,
    }
    output.mkdir(parents=True, exist_ok=True)
    file_hashes: dict[str, dict[str, Any]] = {}
    for stem, frame in frames.items():
        output_frame = _prepare_frame_for_output(frame)
        output_frame.to_csv(output / f"{stem}.csv", index=False)
        output_frame.to_parquet(output / f"{stem}.parquet", index=False)
        for suffix in ("csv", "parquet"):
            path = output / f"{stem}.{suffix}"
            file_hashes[path.name] = {"sha256": sha256(path), "size_bytes": path.stat().st_size}

    existing_workbook = existing_manifest.get("workbook") if isinstance(existing_manifest, dict) else {}
    extracted_at = (
        existing_manifest.get("extracted_at")
        if isinstance(existing_workbook, dict)
        and existing_workbook.get("sha256") == workbook_hash
        and existing_manifest.get("schema_version") == str(entry["schema_version"])
        else datetime.now(timezone.utc).isoformat()
    )
    manifest = {
        "schema_version": str(entry["schema_version"]),
        "vintage_id": vid,
        "source_release": str(entry["release_round"]),
        "repo_relative_output_dir": _repo_relative(root, output),
        "registry": _repo_relative(root, reg_path),
        "source_policy": (
            f"{entry['source_sheet']} worksheet only; workbook is offline lineage and is "
            "never loaded at Streamlit runtime."
        ),
        "workbook": {
            "basename": workbook.name,
            "sha256": workbook_hash,
            "size_bytes": workbook.stat().st_size,
            "sheet": str(entry["source_sheet"]),
        },
        "extracted_at": extracted_at,
        "extracted_by": extracted_by,
        "sentinels_validated": sentinel_count,
        "row_count": {name: int(len(frame)) for name, frame in frames.items()},
        "normalized_files": file_hashes,
        "formula_policy": (
            f"{entry['release_round']} annual value cells are stored without Excel formulas "
            "in the grid; aggregate formula contracts are asserted from the governed formula "
            "registry and residuals are reported without force-balancing."
        ),
    }
    (output / "manifest.json").write_text(
        json.dumps(manifest, indent=2, allow_nan=False) + "\n", encoding="utf-8"
    )
    (output / "manifest.md").write_text(_manifest_markdown(manifest), encoding="utf-8")

    # Registry write-back: add the new entry and/or move default flags.
    changed = vid not in existing_ids
    if vid not in existing_ids:
        registry["vintages"].append(entry)
    for flag, requested in (
        ("is_latest", set_latest),
        ("is_default_comparator", set_default_comparator),
        ("is_default_bridge_vintage", set_default_bridge_vintage),
    ):
        if not requested:
            continue
        for item in registry["vintages"]:
            new_value = str(item["vintage_id"]) == vid
            if bool(item.get(flag)) != new_value:
                item[flag] = new_value
                changed = True
    if changed:
        _write_registry(registry, reg_path)
    return manifest
