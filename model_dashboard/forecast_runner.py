from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import hashlib
import json
from pathlib import Path
import re
from typing import Any, BinaryIO
import zipfile

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from .forward_scorer_governance import (
    ForwardScorerAudit,
    GOVERNED_GAP,
    NUMERIC_FORECAST_AVAILABLE,
    artifact_hashes,
    existing_basis,
    missing_paths,
)
from .governance_constants import current_finalist
from .heavy_ruc_forward import evaluate_heavy_ruc_forward_scorer
from .ped_forward import evaluate_ped_forward_scorer


DEFAULT_FORECAST_HORIZON_QUARTERS = 20
MIN_FORECAST_HORIZON_QUARTERS = 1
FORECAST_HORIZON_QUARTERS = DEFAULT_FORECAST_HORIZON_QUARTERS
TEMPLATE_FILENAME_PREFIX = "NLTF_forecast_input_template"
TEMPLATE_FILENAME = f"{TEMPLATE_FILENAME_PREFIX}_{DEFAULT_FORECAST_HORIZON_QUARTERS}q.xlsx"
FORECAST_RUNNER_VERSION = "forecast-runner-v5-forward-scorer-governance"
FORECAST_BUILDER_TITLE = "Forecast Builder"
FORECAST_BUILDER_NOTE = (
    "This workflow creates forward forecasts or governed missing-capability gaps from a user-supplied "
    "variable-horizon assumption workbook. It writes separate forecast-run artifacts and does not alter governance "
    "evidence, KPIs, MAPE/R2, chart sources, finalists, scenarios, stress tests or diagnostics."
)
BACKTEST_SUPPORTED_MAX_HORIZON = 12
HORIZON_SUPPORT_NOTE = (
    "H1-H12 are the validated backtest-supported horizon; H13-H100 are long-range extrapolation, "
    "not validated 2050 accuracy."
)
HIGH_POPULATION_SMOKE_FIXTURE_NOTE = (
    "The high_population workbook is a technical smoke-test fixture: every user input is 2% above base, "
    "including unemployment, prices and starting target lags. It is not a decision-grade population-only scenario."
)
SCENARIO_ROLE_BASECASE = "basecase"
SCENARIO_ROLE_COMPARISON = "comparison"
SCENARIO_ROLE_OPTIONS = (SCENARIO_ROLE_BASECASE, SCENARIO_ROLE_COMPARISON)
SMOKE_FIXTURE_EXPECTED_DELTA = 0.02
SMOKE_FIXTURE_DELTA_TOLERANCE = 1e-9

STREAM_ORDER = ["PED", "LIGHT_RUC", "HEAVY_RUC"]
STREAM_LABELS = {
    "PED": "PED VKT per capita",
    "LIGHT_RUC": "Light RUC volume",
    "HEAVY_RUC": "Heavy RUC volume",
}
SHEET_BY_STREAM = {
    "PED": "PED Inputs",
    "LIGHT_RUC": "Light RUC Inputs",
    "HEAVY_RUC": "Heavy RUC Inputs",
}

def governed_finalist(stream: str) -> str:
    return current_finalist(str(stream))


class _GovernedFinalistMap(dict[str, str]):
    def __getitem__(self, stream: str) -> str:
        return governed_finalist(stream)

    def get(self, stream: str, default: Any = None) -> Any:
        try:
            return governed_finalist(stream)
        except Exception:
            return default


FINALIST_BY_STREAM = _GovernedFinalistMap()
MISSING_CAPABILITY_BY_STREAM = {
    "PED": "ped_inner_hpo_static_solver_forward_scorer_missing",
    "LIGHT_RUC": "light_ruc_fixed_finalist_forward_scorer_missing",
    "HEAVY_RUC": "heavy_ruc_component_forward_scorers_missing",
}
MISSING_CAPABILITY_NOTES = {
    "PED": (
        "PED remains a governed gap for forward scoring: repo-vendored HPO/static-solver artifacts prove stored "
        "component replay and training-fit provenance, but do not provide a verified executable inner HPO/static-solver "
        "forward scorer for new assumption rows. Missing forward state: nested top-member refit registry, feature matrix "
        "builder parity tests, and inner weighted replay parity for new rows."
    ),
    "LIGHT_RUC": (
        "Light RUC requires repo-local model_input_history and the fixed dynamic_RESID_GBR_n150_d1_lr0.05_w36 recipe."
    ),
    "HEAVY_RUC": (
        "Heavy RUC remains a governed gap for forward scoring: stored historical weighted replay and training-fit R2 "
        "are available, but new-row Heavy forecasts require exact C3/C4 parent-state parity. Current status: governed "
        "gap because the triage pack did not contain original C3/C4 parent estimators or exact parent feature matrices."
    ),
}
LIGHT_RUC_FORWARD_SCORER_VERSION = "light-ruc-forward-scorer-v1"
CAPABILITY_METADATA_COLUMNS = [
    "capability_status",
    "scorer_version",
    "source_artifact_hashes",
    "parity_status",
    "max_parity_delta",
    "stored_replay_max_delta",
    "required_components",
    "missing_artifacts",
    "failing_component",
]
MODEL_INPUT_HISTORY_DIR = Path("data") / "model_input_history"
MODEL_INPUT_HISTORY_FILES = {
    "PED": "ped_inputs.parquet",
    "LIGHT_RUC": "light_ruc_inputs.parquet",
    "HEAVY_RUC": "heavy_ruc_inputs.parquet",
}
LIGHT_RUC_WINDOW = 36
LIGHT_RUC_BASE_FEATURES = [
    "log_real_diesel_price",
    "log_real_light_ruc_price",
    "log_lagged_real_light_ruc_price",
    "log_real_gdp",
    "post_2020_dummy",
    "q2_dummy",
    "q3_dummy",
    "q4_dummy",
]
LIGHT_RUC_RESIDUAL_FEATURES = [
    *LIGHT_RUC_BASE_FEATURES,
    "diesel_x_ruc_price",
    "gdp_x_post2020",
    "ruc_x_post2020",
    "diesel_x_post2020",
    "time_trend",
    "log_trend",
    "log_real_diesel_price_diff1",
    "log_real_diesel_price_lag1",
    "log_real_diesel_price_lag4",
    "log_real_light_ruc_price_diff1",
    "log_real_light_ruc_price_lag1",
    "log_real_light_ruc_price_lag4",
    "log_real_gdp_diff1",
    "log_real_gdp_lag1",
    "log_real_gdp_lag4",
]


@dataclass(frozen=True)
class TemplateColumn:
    name: str
    role: str
    required: bool
    description: str
    formula_key: str | None = None


COMMON_SYSTEM_COLUMNS = [
    TemplateColumn("period", "system", True, "Forecast quarter label generated by the template, for example 2026Q1."),
    TemplateColumn("year", "system", True, "Calendar year for the forecast quarter."),
    TemplateColumn("quarter", "system", True, "Calendar quarter number, 1 through 4."),
    TemplateColumn("horizon", "system", True, "Forecast horizon from the last known actual quarter."),
]

COMMON_FORMULA_COLUMNS = [
    TemplateColumn("q2_dummy", "formula", True, "1 when quarter is Q2, otherwise 0.", "q2_dummy"),
    TemplateColumn("q3_dummy", "formula", True, "1 when quarter is Q3, otherwise 0.", "q3_dummy"),
    TemplateColumn("q4_dummy", "formula", True, "1 when quarter is Q4, otherwise 0.", "q4_dummy"),
    TemplateColumn("post_2020_dummy", "formula", True, "1 for quarters after 2020, otherwise 0.", "post_2020_dummy"),
    TemplateColumn("trend_index", "formula", True, "Sequential time index used by static specifications.", "trend_index"),
]

PED_COLUMNS = [
    *COMMON_SYSTEM_COLUMNS,
    TemplateColumn("real_gdp_per_capita_nzd", "user", True, "User-entered real GDP per capita assumption."),
    TemplateColumn("population", "user", True, "User-entered population assumption."),
    TemplateColumn("unemployment_rate", "user", True, "User-entered unemployment rate assumption in percentage points."),
    TemplateColumn("real_petrol_price_cents_per_litre", "user", True, "User-entered real petrol price assumption."),
    TemplateColumn("target_lag_1", "user", True, "Most recent PED VKT per capita value available before this forecast quarter."),
    TemplateColumn("target_lag_4", "user", True, "PED VKT per capita value four quarters before this forecast quarter."),
    *COMMON_FORMULA_COLUMNS,
    TemplateColumn("log_real_gdp_per_capita", "formula", True, "Natural log of real GDP per capita.", "log_real_gdp_per_capita_nzd"),
    TemplateColumn("log_population", "formula", True, "Natural log of population.", "log_population"),
    TemplateColumn("log_unemployment_rate", "formula", True, "Natural log of unemployment rate.", "log_unemployment_rate"),
    TemplateColumn("log_real_petrol_price", "formula", True, "Natural log of real petrol price.", "log_real_petrol_price_cents_per_litre"),
    TemplateColumn("log_target_lag_1", "formula", True, "Natural log of target lag 1.", "log_target_lag_1"),
    TemplateColumn("log_target_lag_4", "formula", True, "Natural log of target lag 4.", "log_target_lag_4"),
    TemplateColumn("diff_log_target_lag_1_lag_4", "formula", True, "Log target lag 1 minus log target lag 4.", "diff_log_target"),
    TemplateColumn("gdp_petrol_interaction", "formula", True, "Log real GDP per capita multiplied by log real petrol price.", "gdp_petrol_interaction"),
    TemplateColumn("post_2011_x_log_trend", "formula", True, "Post-2011 dummy multiplied by log trend index.", "post_2011_x_log_trend"),
]

LIGHT_RUC_COLUMNS = [
    *COMMON_SYSTEM_COLUMNS,
    TemplateColumn("real_gdp_sa_nzd", "user", True, "User-entered real GDP assumption."),
    TemplateColumn("real_diesel_price_cents_per_litre", "user", True, "User-entered real diesel price assumption."),
    TemplateColumn("real_light_ruc_price_nzd_per_1000km", "user", True, "User-entered real Light RUC price assumption."),
    TemplateColumn("lagged_real_light_ruc_price_nzd_per_1000km", "user", True, "Real Light RUC price lag available before this forecast quarter."),
    TemplateColumn("target_lag_1", "user", True, "Most recent Light RUC volume available before this forecast quarter."),
    TemplateColumn("target_lag_4", "user", True, "Light RUC volume four quarters before this forecast quarter."),
    *COMMON_FORMULA_COLUMNS,
    TemplateColumn("log_real_gdp", "formula", True, "Natural log of real GDP.", "log_real_gdp_sa_nzd"),
    TemplateColumn("log_real_diesel_price", "formula", True, "Natural log of real diesel price.", "log_real_diesel_price_cents_per_litre"),
    TemplateColumn("log_real_light_ruc_price", "formula", True, "Natural log of real Light RUC price.", "log_real_light_ruc_price_nzd_per_1000km"),
    TemplateColumn("log_lagged_real_light_ruc_price", "formula", True, "Natural log of lagged real Light RUC price.", "log_lagged_real_light_ruc_price_nzd_per_1000km"),
    TemplateColumn("log_target_lag_1", "formula", True, "Natural log of target lag 1.", "log_target_lag_1"),
    TemplateColumn("log_target_lag_4", "formula", True, "Natural log of target lag 4.", "log_target_lag_4"),
    TemplateColumn("diff_log_real_gdp", "formula", True, "Quarterly difference in log real GDP.", "diff_log_real_gdp"),
    TemplateColumn("gdp_light_ruc_price_interaction", "formula", True, "Log real GDP multiplied by log real Light RUC price.", "gdp_light_ruc_price_interaction"),
]

HEAVY_RUC_COLUMNS = [
    *COMMON_SYSTEM_COLUMNS,
    TemplateColumn("real_gdp_sa_nzd", "user", True, "User-entered real GDP assumption."),
    TemplateColumn("real_diesel_price_cents_per_litre", "user", True, "User-entered real diesel price assumption."),
    TemplateColumn("unemployment_rate", "user", True, "User-entered unemployment rate assumption in percentage points."),
    TemplateColumn("real_light_ruc_price_nzd_per_1000km", "user", True, "User-entered real Light RUC price assumption used by Heavy components."),
    TemplateColumn("lagged_real_light_ruc_price_nzd_per_1000km", "user", True, "Real Light RUC price lag used by Heavy components."),
    TemplateColumn("real_heavy_ruc_price_nzd_per_1000km", "user", True, "User-entered real Heavy RUC price assumption."),
    TemplateColumn("lead_real_heavy_ruc_price_nzd_per_1000km", "user", True, "Expected next-quarter real Heavy RUC price assumption."),
    TemplateColumn("target_lag_1", "user", True, "Most recent Heavy RUC volume available before this forecast quarter."),
    TemplateColumn("target_lag_4", "user", True, "Heavy RUC volume four quarters before this forecast quarter."),
    *COMMON_FORMULA_COLUMNS,
    TemplateColumn("log_real_gdp", "formula", True, "Natural log of real GDP.", "log_real_gdp_sa_nzd"),
    TemplateColumn("log_real_diesel_price", "formula", True, "Natural log of real diesel price.", "log_real_diesel_price_cents_per_litre"),
    TemplateColumn("log_unemployment_rate", "formula", True, "Natural log of unemployment rate.", "log_unemployment_rate"),
    TemplateColumn("log_real_light_ruc_price", "formula", True, "Natural log of real Light RUC price.", "log_real_light_ruc_price_nzd_per_1000km"),
    TemplateColumn("log_lagged_real_light_ruc_price", "formula", True, "Natural log of lagged real Light RUC price.", "log_lagged_real_light_ruc_price_nzd_per_1000km"),
    TemplateColumn("log_real_heavy_ruc_price", "formula", True, "Natural log of real Heavy RUC price.", "log_real_heavy_ruc_price_nzd_per_1000km"),
    TemplateColumn("log_lead_real_heavy_ruc_price", "formula", True, "Natural log of lead real Heavy RUC price.", "log_lead_real_heavy_ruc_price_nzd_per_1000km"),
    TemplateColumn("log_target_lag_1", "formula", True, "Natural log of target lag 1.", "log_target_lag_1"),
    TemplateColumn("log_target_lag_4", "formula", True, "Natural log of target lag 4.", "log_target_lag_4"),
    TemplateColumn("diff_log_real_gdp", "formula", True, "Quarterly difference in log real GDP.", "diff_log_real_gdp"),
    TemplateColumn("gdp_heavy_ruc_price_interaction", "formula", True, "Log real GDP multiplied by log real Heavy RUC price.", "gdp_heavy_ruc_price_interaction"),
]

STREAM_COLUMNS = {
    "PED": PED_COLUMNS,
    "LIGHT_RUC": LIGHT_RUC_COLUMNS,
    "HEAVY_RUC": HEAVY_RUC_COLUMNS,
}


@dataclass
class ForecastValidationResult:
    valid: bool
    errors: list[str]
    warnings: list[str]
    assumptions: pd.DataFrame
    latest_actual_period: str
    forecast_periods: list[str]

    @property
    def forecast_horizon_quarters(self) -> int:
        return len(self.forecast_periods)

    @property
    def forecast_start_period(self) -> str | None:
        return self.forecast_periods[0] if self.forecast_periods else None

    @property
    def forecast_end_period(self) -> str | None:
        return self.forecast_periods[-1] if self.forecast_periods else None

    def report_frame(self) -> pd.DataFrame:
        rows = [
            {"severity": "error", "message": message}
            for message in self.errors
        ] + [
            {"severity": "warning", "message": message}
            for message in self.warnings
        ]
        if not rows:
            rows = [{"severity": "ok", "message": "Workbook inputs passed structural and required-value validation."}]
        return pd.DataFrame(rows)


@dataclass
class ForecastRunResult:
    output_dir: Path
    manifest: dict[str, Any]
    validation: ForecastValidationResult
    future_forecasts: pd.DataFrame
    component_forecasts: pd.DataFrame
    capability_report: pd.DataFrame
    forecast_chart_rows: pd.DataFrame
    assumptions: pd.DataFrame
    report_markdown: str

    @property
    def has_numeric_forecasts(self) -> bool:
        if self.future_forecasts.empty or "forecast" not in self.future_forecasts.columns:
            return False
        return pd.to_numeric(self.future_forecasts["forecast"], errors="coerce").notna().any()


@dataclass
class ForecastScenarioComparisonResult:
    output_dir: Path
    manifest: dict[str, Any]
    scenario_results: list[ForecastRunResult]
    future_forecasts: pd.DataFrame
    capability_report: pd.DataFrame
    forecast_chart_rows: pd.DataFrame
    scenario_input_delta_audit: pd.DataFrame


def repo_root_from_here() -> Path:
    return Path(__file__).resolve().parents[1]


def latest_known_actual_period(repo_root: Path | str | None = None) -> str:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    candidates = [
        root / "data" / "dashboard_evidence_pack" / "data" / "scorecard_predictions.parquet",
        root / "data" / "dashboard_evidence_pack" / "data" / "component_predictions.parquet",
    ]
    periods: list[str] = []
    for path in candidates:
        if not path.exists():
            continue
        try:
            frame = pd.read_parquet(path, columns=["target_period"])
        except Exception:
            frame = pd.read_parquet(path)
        if "target_period" in frame.columns:
            periods.extend(frame["target_period"].dropna().astype(str).tolist())
    if not periods:
        return "2025Q4"
    return max(periods, key=quarter_sort_key)


def quarter_sort_key(period: str) -> int:
    year, quarter = parse_period(period)
    return year * 4 + quarter


def parse_period(period: str) -> tuple[int, int]:
    text = str(period).strip().upper()
    if "Q" not in text:
        raise ValueError(f"Invalid quarter period: {period!r}")
    year_text, quarter_text = text.split("Q", 1)
    year = int(year_text)
    quarter = int(quarter_text)
    if quarter not in {1, 2, 3, 4}:
        raise ValueError(f"Invalid quarter number in period: {period!r}")
    return year, quarter


def add_quarters(period: str, quarters: int) -> str:
    year, quarter = parse_period(period)
    index = year * 4 + (quarter - 1) + quarters
    return f"{index // 4}Q{(index % 4) + 1}"


def future_quarters_after(period: str, horizon: int = FORECAST_HORIZON_QUARTERS) -> list[str]:
    if horizon < MIN_FORECAST_HORIZON_QUARTERS:
        raise ValueError(f"Forecast horizon must be at least {MIN_FORECAST_HORIZON_QUARTERS} quarter.")
    return [add_quarters(period, offset) for offset in range(1, horizon + 1)]


def horizon_support_status(horizon: Any) -> str:
    try:
        value = int(float(horizon))
    except Exception:
        return "unknown_horizon_scope"
    if 1 <= value <= BACKTEST_SUPPORTED_MAX_HORIZON:
        return "backtest_supported_h1_12"
    return "long_range_extrapolation_h13_plus"


def horizon_support_label(horizon: Any) -> str:
    status = horizon_support_status(horizon)
    if status == "backtest_supported_h1_12":
        return "H1-H12 backtest-supported horizon"
    if status == "long_range_extrapolation_h13_plus":
        return "H13-H100 long-range extrapolation"
    return "Unknown horizon scope"


def horizon_support_note(horizon: Any) -> str:
    return HORIZON_SUPPORT_NOTE


def _add_horizon_support_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame is None:
        return pd.DataFrame()
    if frame.empty or "horizon" not in frame.columns:
        return frame
    out = frame.copy()
    out["horizon_support_status"] = out["horizon"].map(horizon_support_status)
    out["horizon_support_label"] = out["horizon"].map(horizon_support_label)
    out["horizon_support_note"] = out["horizon"].map(horizon_support_note)
    out["backtest_supported_max_horizon"] = BACKTEST_SUPPORTED_MAX_HORIZON
    return out


def forecast_periods_after(
    period: str,
    *,
    quarters: int | None = None,
    end_period: str | None = None,
) -> list[str]:
    if quarters is not None and end_period is not None:
        raise ValueError("Use either quarters or end_period, not both.")
    if end_period is not None:
        end = str(end_period).strip().upper()
        parse_period(end)
        count = quarter_sort_key(end) - quarter_sort_key(period)
        if count < MIN_FORECAST_HORIZON_QUARTERS:
            raise ValueError(f"End period {end} must be after latest actual quarter {period}.")
        return future_quarters_after(period, count)
    horizon = DEFAULT_FORECAST_HORIZON_QUARTERS if quarters is None else quarters
    return future_quarters_after(period, horizon)


def forecast_template_filename(*, quarters: int | None = None, end_period: str | None = None) -> str:
    if quarters is not None and end_period is not None:
        raise ValueError("Use either quarters or end_period, not both.")
    if end_period is not None:
        return f"{TEMPLATE_FILENAME_PREFIX}_to_{str(end_period).strip().upper()}.xlsx"
    horizon = DEFAULT_FORECAST_HORIZON_QUARTERS if quarters is None else quarters
    if horizon < MIN_FORECAST_HORIZON_QUARTERS:
        raise ValueError(f"Forecast horizon must be at least {MIN_FORECAST_HORIZON_QUARTERS} quarter.")
    return f"{TEMPLATE_FILENAME_PREFIX}_{horizon}q.xlsx"


def scenario_name_from_filename(filename: str | Path | None) -> str:
    name = Path(str(filename or "")).stem
    text = name.strip()
    prefixes = [
        TEMPLATE_FILENAME_PREFIX + "_",
        "NLTF_forecast_input_template_",
        "NLTF_forecast_input_",
        "forecast_input_template_",
        "forecast_input_",
        "completed_forecast_input_",
        "forecast_builder_",
    ]
    for prefix in prefixes:
        if text.lower().startswith(prefix.lower()):
            text = text[len(prefix) :]
            break
    text = re.sub(r"^completed[_ -]+", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[_ -]+completed$", "", text, flags=re.IGNORECASE)
    return sanitize_scenario_name(text or "scenario")


def sanitize_scenario_name(value: str | None) -> str:
    text = str(value or "scenario").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "scenario"


def normalize_scenario_role(value: str | None) -> str | None:
    if value is None or not str(value).strip():
        return None
    text = sanitize_scenario_name(value)
    if text in {"base", "basecase", "base_case", "baseline", "reference"}:
        return SCENARIO_ROLE_BASECASE
    if text in {"comparison", "compare", "alternative", "alternate", "high_population", "upside", "downside"}:
        return SCENARIO_ROLE_COMPARISON
    return None


def infer_scenario_role(*values: str | Path | None) -> str | None:
    text = " ".join(str(value or "") for value in values)
    normalized = sanitize_scenario_name(text)
    tokens = {token for token in normalized.split("_") if token}
    has_base = bool(
        {"basecase", "baseline", "reference"}.intersection(tokens)
        or re.search(r"(?:^|_)base(?:_|$)", normalized)
        or "base_case" in normalized
    )
    has_comparison = bool(
        "high_population" in normalized
        or {"comparison", "compare", "alternative", "alternate", "upside", "downside"}.intersection(tokens)
    )
    if has_base == has_comparison:
        return None
    return SCENARIO_ROLE_BASECASE if has_base else SCENARIO_ROLE_COMPARISON


def resolve_scenario_role(
    *,
    scenario_role: str | None = None,
    scenario_name: str | None = None,
    workbook_filename: str | Path | None = None,
) -> tuple[str | None, str]:
    explicit = normalize_scenario_role(scenario_role)
    if explicit is not None:
        return explicit, "explicit"
    inferred = infer_scenario_role(scenario_name, workbook_filename)
    if inferred is not None:
        return inferred, "inferred_from_name"
    return None, "ambiguous"


def build_forecast_input_template(
    output_path: Path | str,
    repo_root: Path | str | None = None,
    *,
    quarters: int | None = None,
    end_period: str | None = None,
) -> Path:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    latest = latest_known_actual_period(root)
    periods = forecast_periods_after(latest, quarters=quarters, end_period=end_period)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    _write_readme_sheet(readme, latest, periods)
    for stream in STREAM_ORDER:
        ws = wb.create_sheet(SHEET_BY_STREAM[stream])
        _write_input_sheet(ws, stream, periods)
    wb.save(output)
    return output


def build_forecast_input_template_bytes(repo_root: Path | str | None = None) -> bytes:
    return build_forecast_input_template_bytes_for_horizon(repo_root=repo_root)


def build_forecast_input_template_bytes_for_horizon(
    repo_root: Path | str | None = None,
    *,
    quarters: int | None = None,
    end_period: str | None = None,
) -> bytes:
    stream = BytesIO()
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    latest = latest_known_actual_period(root)
    periods = forecast_periods_after(latest, quarters=quarters, end_period=end_period)
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    _write_readme_sheet(readme, latest, periods)
    for stream_code in STREAM_ORDER:
        ws = wb.create_sheet(SHEET_BY_STREAM[stream_code])
        _write_input_sheet(ws, stream_code, periods)
    wb.save(stream)
    return stream.getvalue()


def _write_readme_sheet(ws: Any, latest_period: str, periods: list[str]) -> None:
    ws["A1"] = "NLTF forecast input template"
    ws["A1"].font = Font(bold=True, size=14, color="002B5C")
    rows = [
        ("Latest known actual quarter", latest_period),
        ("Forecast horizon", f"{len(periods)} quarters: {periods[0]} to {periods[-1]}"),
        ("What to fill", "Only fill the user-entry columns in PED Inputs, Light RUC Inputs and Heavy RUC Inputs."),
        ("Variable horizon", "You may fill 1 quarter, the default 20 quarters, or any continuous horizon to a chosen end period."),
        ("Runner rule", "The runner scores only the continuous valid rows present across all three stream sheets."),
        ("What not to edit", "Do not edit period, year, quarter, horizon or protected formula columns."),
        ("Data rule", "Use real price/GDP assumptions in the units named in each header."),
        ("Governance rule", "The runner does not run a broad candidate search and never overwrites dashboard evidence packs."),
        (
            "Missing model-state rule",
            "If repo-local fitted finalist artifacts are unavailable, the runner writes governed gaps instead of fake forecasts.",
        ),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_index, column=2, value=value)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 118
    ws.freeze_panes = "A3"


def _write_input_sheet(ws: Any, stream: str, periods: list[str]) -> None:
    columns = STREAM_COLUMNS[stream]
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    user_fill = PatternFill("solid", fgColor="FFF7CC")
    formula_fill = PatternFill("solid", fgColor="E7F5EC")
    system_fill = PatternFill("solid", fgColor="E6EDF5")
    for col_index, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=column.name)
        cell.font = Font(bold=True, color="002B5C")
        cell.fill = header_fill
        cell.comment = Comment(f"{column.role.upper()}: {column.description}", "Codex")
        ws.column_dimensions[get_column_letter(col_index)].width = max(15, min(34, len(column.name) + 3))

    col_map = {column.name: idx for idx, column in enumerate(columns, start=1)}
    for row_offset, period in enumerate(periods, start=2):
        year, quarter = parse_period(period)
        system_values = {
            "period": period,
            "year": year,
            "quarter": quarter,
            "horizon": row_offset - 1,
        }
        for column in columns:
            cell = ws.cell(row=row_offset, column=col_map[column.name])
            if column.role == "system":
                cell.value = system_values[column.name]
                cell.fill = system_fill
            elif column.role == "formula":
                cell.value = _excel_formula(column.formula_key or column.name, row_offset, col_map)
                cell.fill = formula_fill
            else:
                cell.fill = user_fill
                cell.protection = Protection(locked=False)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.protection.sheet = True
    ws.protection.enable()


def _cell(row: int, col_map: dict[str, int], column: str) -> str:
    return f"{get_column_letter(col_map[column])}{row}"


def _excel_formula(formula_key: str, row: int, col_map: dict[str, int]) -> str:
    q = _cell(row, col_map, "quarter")
    year = _cell(row, col_map, "year")
    horizon = _cell(row, col_map, "horizon")
    if formula_key == "q2_dummy":
        return f'=IF({q}=2,1,0)'
    if formula_key == "q3_dummy":
        return f'=IF({q}=3,1,0)'
    if formula_key == "q4_dummy":
        return f'=IF({q}=4,1,0)'
    if formula_key == "post_2020_dummy":
        return f'=IF({year}>=2021,1,0)'
    if formula_key == "trend_index":
        return f"={horizon}+104"
    if formula_key == "diff_log_real_gdp":
        current = _cell(row, col_map, "log_real_gdp")
        if row == 2:
            return '=""'
        previous = _cell(row - 1, col_map, "log_real_gdp")
        return f'=IF(AND({current}<>"",{previous}<>""),{current}-{previous},"")'
    if formula_key == "diff_log_target":
        lag1 = _cell(row, col_map, "log_target_lag_1")
        lag4 = _cell(row, col_map, "log_target_lag_4")
        return f'=IF(AND({lag1}<>"",{lag4}<>""),{lag1}-{lag4},"")'
    if formula_key == "gdp_petrol_interaction":
        gdp = _cell(row, col_map, "log_real_gdp_per_capita")
        price = _cell(row, col_map, "log_real_petrol_price")
        return f'=IF(AND({gdp}<>"",{price}<>""),{gdp}*{price},"")'
    if formula_key == "gdp_light_ruc_price_interaction":
        gdp = _cell(row, col_map, "log_real_gdp")
        price = _cell(row, col_map, "log_real_light_ruc_price")
        return f'=IF(AND({gdp}<>"",{price}<>""),{gdp}*{price},"")'
    if formula_key == "gdp_heavy_ruc_price_interaction":
        gdp = _cell(row, col_map, "log_real_gdp")
        price = _cell(row, col_map, "log_real_heavy_ruc_price")
        return f'=IF(AND({gdp}<>"",{price}<>""),{gdp}*{price},"")'
    if formula_key == "post_2011_x_log_trend":
        trend = _cell(row, col_map, "trend_index")
        return f'=IF({trend}>0,LN({trend}),"")'
    if formula_key.startswith("log_"):
        source = formula_key.removeprefix("log_")
        source_cell = _cell(row, col_map, source)
        return f'=IF({source_cell}>0,LN({source_cell}),"")'
    return '=""'


def validate_forecast_workbook(
    workbook: Path | str | bytes | BinaryIO,
    repo_root: Path | str | None = None,
    *,
    expected_quarters: int | None = None,
    expected_end_period: str | None = None,
) -> ForecastValidationResult:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    latest = latest_known_actual_period(root)
    expected_periods = (
        forecast_periods_after(latest, quarters=expected_quarters, end_period=expected_end_period)
        if expected_quarters is not None or expected_end_period is not None
        else None
    )
    fallback_periods = expected_periods or forecast_periods_after(latest)
    errors: list[str] = []
    warnings: list[str] = []
    frames_by_stream: dict[str, pd.DataFrame] = {}
    periods_by_stream: dict[str, list[str]] = {}
    try:
        wb = _load_workbook_from_input(workbook)
    except Exception as exc:
        return ForecastValidationResult(False, [f"Workbook could not be opened: {exc}"], [], pd.DataFrame(), latest, fallback_periods)

    for stream in STREAM_ORDER:
        sheet = SHEET_BY_STREAM[stream]
        if sheet not in wb.sheetnames:
            errors.append(f"Missing required sheet: {sheet}")
            continue
        frame = _worksheet_to_frame(wb[sheet])
        valid_frame, sheet_periods, sheet_errors, sheet_warnings = _validate_sheet_frame(
            frame,
            stream,
            latest,
            expected_periods=expected_periods,
        )
        errors.extend(sheet_errors)
        warnings.extend(sheet_warnings)
        if not valid_frame.empty:
            frames_by_stream[stream] = valid_frame
            periods_by_stream[stream] = sheet_periods

    inferred_periods = _common_forecast_periods(periods_by_stream)
    if periods_by_stream and inferred_periods is None:
        details = "; ".join(f"{SHEET_BY_STREAM[stream]}={periods}" for stream, periods in periods_by_stream.items())
        errors.append(f"Input sheets must contain the same continuous forecast periods across all streams. {details}")
    periods = inferred_periods or next(iter(periods_by_stream.values()), fallback_periods)
    frames: list[pd.DataFrame] = []
    if not errors and periods:
        for stream in STREAM_ORDER:
            frame = frames_by_stream.get(stream)
            if frame is not None:
                frames.append(_build_stream_assumptions(frame, stream, periods))
    assumptions = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()
    return ForecastValidationResult(not errors, errors, warnings, assumptions, latest, periods)


def run_forecast_workbook(
    workbook: Path | str | bytes | BinaryIO,
    output_dir: Path | str | None = None,
    repo_root: Path | str | None = None,
    workbook_filename: str | None = None,
    run_timestamp: str | None = None,
    scenario_name: str | None = None,
    scenario_role: str | None = None,
    is_test_fixture: bool = False,
    expected_quarters: int | None = None,
    expected_end_period: str | None = None,
) -> ForecastRunResult:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    workbook_name = workbook_filename or _workbook_name(workbook)
    scenario = sanitize_scenario_name(scenario_name) if scenario_name else scenario_name_from_filename(workbook_name)
    resolved_role, role_source = resolve_scenario_role(
        scenario_role=scenario_role,
        scenario_name=scenario,
        workbook_filename=workbook_name,
    )
    validation = validate_forecast_workbook(
        workbook,
        root,
        expected_quarters=expected_quarters,
        expected_end_period=expected_end_period,
    )
    timestamp = run_timestamp or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    run_id = f"{timestamp}_{scenario}"
    run_dir = Path(output_dir) if output_dir is not None else root / "artifacts" / "forecast_runs" / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    workbook_bytes = _workbook_bytes(workbook)
    workbook_hash = hashlib.sha256(workbook_bytes).hexdigest() if workbook_bytes else None

    capabilities = model_capability_gap_register(root)
    future, components = _forecast_output_rows(validation, capabilities, root)
    assumptions = validation.assumptions.copy()
    _add_scenario_columns(future, scenario, resolved_role)
    _add_scenario_columns(components, scenario, resolved_role)
    _add_scenario_columns(assumptions, scenario, resolved_role)
    capability_report = _forecast_capability_report(
        capabilities,
        future,
        components,
        validation,
        scenario_name=scenario,
        scenario_role=resolved_role,
    )
    chart_rows = forecast_chart_rows_for_display(future, repo_root=root, latest_actual_period=validation.latest_actual_period)
    report = _forecast_validation_report(validation, capabilities)
    forecast_status = _forecast_status(validation, future)
    manifest = {
        "run_id": run_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "forecast_runner_version": FORECAST_RUNNER_VERSION,
        "scenario_name": scenario,
        "scenario_role": resolved_role,
        "scenario_role_source": role_source,
        "scenario_role_required_for_comparison": True,
        "scenario_display_name": scenario,
        "workbook_filename": workbook_name,
        "workbook_sha256": workbook_hash,
        "is_test_fixture": bool(is_test_fixture),
        "latest_actual_period": validation.latest_actual_period,
        "forecast_periods": validation.forecast_periods,
        "forecast_horizon_quarters": validation.forecast_horizon_quarters,
        "backtest_supported_max_horizon": BACKTEST_SUPPORTED_MAX_HORIZON,
        "long_range_extrapolation_start_horizon": BACKTEST_SUPPORTED_MAX_HORIZON + 1,
        "horizon_support_note": HORIZON_SUPPORT_NOTE,
        "forecast_start_period": validation.forecast_start_period,
        "forecast_end_period": validation.forecast_end_period,
        "validation_status": "passed" if validation.valid else "failed",
        "validation_errors": validation.errors,
        "validation_warnings": validation.warnings,
        "fixed_finalists_only": True,
        "broad_search_run": False,
        "evidence_pack_modified": False,
        "chart_sources_modified": False,
        "forecast_status": forecast_status,
        "numeric_forecast_streams": sorted(
            future.loc[future["forecast_available"].fillna(False).astype(bool), "stream"].dropna().astype(str).unique().tolist()
        ),
        "governed_gap_streams": sorted(
            future.loc[~future["forecast_available"].fillna(False).astype(bool), "stream"].dropna().astype(str).unique().tolist()
        ),
        "model_capabilities": _json_safe_records(capabilities),
        "output_files": [
            "future_forecasts.parquet",
            "component_forecasts.parquet",
            "forecast_assumptions.parquet",
            "forecast_capability_report.parquet",
            "forecast_chart_rows.parquet",
            "forecast_run_manifest.json",
            "forecast_validation_report.md",
            "future_forecasts.csv",
            "component_forecasts.csv",
            "forecast_assumptions.csv",
            "forecast_capability_report.csv",
            "forecast_chart_rows.csv",
        ],
    }

    future.to_parquet(run_dir / "future_forecasts.parquet", index=False)
    components.to_parquet(run_dir / "component_forecasts.parquet", index=False)
    assumptions.to_parquet(run_dir / "forecast_assumptions.parquet", index=False)
    capability_report.to_parquet(run_dir / "forecast_capability_report.parquet", index=False)
    chart_rows.to_parquet(run_dir / "forecast_chart_rows.parquet", index=False)
    future.to_csv(run_dir / "future_forecasts.csv", index=False)
    components.to_csv(run_dir / "component_forecasts.csv", index=False)
    assumptions.to_csv(run_dir / "forecast_assumptions.csv", index=False)
    capability_report.to_csv(run_dir / "forecast_capability_report.csv", index=False)
    chart_rows.to_csv(run_dir / "forecast_chart_rows.csv", index=False)
    (run_dir / "forecast_run_manifest.json").write_text(json.dumps(manifest, indent=2, allow_nan=False), encoding="utf-8")
    (run_dir / "forecast_validation_report.md").write_text(report, encoding="utf-8")

    return ForecastRunResult(
        output_dir=run_dir,
        manifest=manifest,
        validation=validation,
        future_forecasts=future,
        component_forecasts=components,
        capability_report=capability_report,
        forecast_chart_rows=chart_rows,
        assumptions=assumptions,
        report_markdown=report,
    )


def forecast_pack_zip_bytes(output_dir: Path | str) -> bytes:
    root = Path(output_dir)
    stream = BytesIO()
    with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(root.glob("*")):
            if path.is_file():
                zf.write(path, arcname=path.name)
    return stream.getvalue()


def forecast_chart_rows_for_display(
    future_forecasts: pd.DataFrame,
    *,
    repo_root: Path | str | None = None,
    latest_actual_period: str | None = None,
) -> pd.DataFrame:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    latest = latest_actual_period or latest_known_actual_period(root)
    streams = (
        future_forecasts["stream"].dropna().astype(str).unique().tolist()
        if future_forecasts is not None and not future_forecasts.empty and "stream" in future_forecasts.columns
        else STREAM_ORDER
    )
    historical = historical_actual_rows(root, latest_actual_period=latest, streams=streams)
    future = _future_forecast_chart_rows(future_forecasts)
    chart_rows = pd.concat([historical, future], ignore_index=True, sort=False)
    if chart_rows.empty:
        return chart_rows
    chart_rows["period_key"] = chart_rows["period"].map(lambda value: quarter_sort_key(str(value)))
    chart_rows = chart_rows.sort_values(["stream", "row_type", "scenario_name", "period_key"], kind="stable").drop(columns=["period_key"])
    return chart_rows.reset_index(drop=True)


def historical_actual_rows(
    repo_root: Path | str | None = None,
    *,
    latest_actual_period: str | None = None,
    streams: list[str] | tuple[str, ...] | None = None,
) -> pd.DataFrame:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    latest = latest_actual_period or latest_known_actual_period(root)
    selected_streams = list(streams) if streams is not None else list(STREAM_ORDER)
    rows: list[dict[str, Any]] = []
    for stream in selected_streams:
        rows.extend(_model_input_history_rows(root, stream, latest))
    if rows:
        return pd.DataFrame(rows)
    return _evidence_pack_actual_rows(root, selected_streams, latest)


def _model_input_history_rows(root: Path, stream: str, latest_actual_period: str) -> list[dict[str, Any]]:
    path = root / MODEL_INPUT_HISTORY_DIR / MODEL_INPUT_HISTORY_FILES.get(stream, "")
    if not path.exists():
        return []
    try:
        frame = pd.read_parquet(path, columns=["period", "target"])
    except Exception:
        frame = pd.read_parquet(path)
    if {"period", "target"}.difference(frame.columns):
        return []
    source = frame[["period", "target"]].copy()
    source["period"] = source["period"].astype(str).str.upper()
    source["value"] = pd.to_numeric(source["target"], errors="coerce")
    source = source[source["period"].map(lambda value: quarter_sort_key(str(value)) <= quarter_sort_key(latest_actual_period))]
    source = source[source["value"].gt(0)].copy()
    source = source.sort_values("period", key=lambda series: series.map(quarter_sort_key))
    return [
        {
            "row_type": "historical_actual",
            "scenario_name": "historical_actual",
            "stream": stream,
            "stream_label": STREAM_LABELS.get(stream, stream),
            "period": str(row["period"]),
            "target_period": str(row["period"]),
            "value": float(row["value"]),
            "availability_status": "historical_actual",
            "forecast_available": pd.NA,
            "gap_code": None,
            "gap_reason": "",
            "source": _repo_relative(root, path),
        }
        for _, row in source.iterrows()
    ]


def _evidence_pack_actual_rows(root: Path, streams: list[str], latest_actual_period: str) -> pd.DataFrame:
    path = root / "data" / "dashboard_evidence_pack" / "data" / "scorecard_predictions.parquet"
    if not path.exists():
        return pd.DataFrame()
    try:
        frame = pd.read_parquet(path, columns=["stream", "stream_label", "target_period", "actual", "scenario"])
    except Exception:
        frame = pd.read_parquet(path)
    required = {"stream", "target_period", "actual"}
    if required.difference(frame.columns):
        return pd.DataFrame()
    source = frame.copy()
    if "scenario" in source.columns:
        source = source[source["scenario"].astype(str).str.casefold().eq("finalist")].copy()
    source = source[source["stream"].astype(str).isin(streams)].copy()
    source["period"] = source["target_period"].astype(str).str.upper()
    source["value"] = pd.to_numeric(source["actual"], errors="coerce")
    source = source[source["period"].map(lambda value: quarter_sort_key(str(value)) <= quarter_sort_key(latest_actual_period))]
    source = source[source["value"].notna()].copy()
    source = source.drop_duplicates(subset=["stream", "period"]).sort_values(["stream", "period"])
    rows = [
        {
            "row_type": "historical_actual",
            "scenario_name": "historical_actual",
            "stream": str(row["stream"]),
            "stream_label": row.get("stream_label", STREAM_LABELS.get(str(row["stream"]), str(row["stream"]))),
            "period": str(row["period"]),
            "target_period": str(row["period"]),
            "value": float(row["value"]),
            "availability_status": "historical_actual",
            "forecast_available": pd.NA,
            "gap_code": None,
            "gap_reason": "",
            "source": _repo_relative(root, path),
        }
        for _, row in source.iterrows()
    ]
    return pd.DataFrame(rows)


def _future_forecast_chart_rows(future_forecasts: pd.DataFrame) -> pd.DataFrame:
    if future_forecasts is None or future_forecasts.empty:
        return pd.DataFrame()
    source = future_forecasts.copy()
    if "scenario_name" not in source.columns:
        source["scenario_name"] = "scenario"
    rows: list[dict[str, Any]] = []
    for _, row in source.iterrows():
        period = str(row.get("target_period", row.get("period", ""))).upper()
        value = pd.to_numeric(row.get("forecast"), errors="coerce")
        rows.append(
            {
                "row_type": "future_forecast",
                "scenario_name": str(row.get("scenario_name", "scenario")),
                "stream": str(row.get("stream", "")),
                "stream_label": row.get("stream_label"),
                "period": period,
                "target_period": period,
                "horizon": row.get("horizon"),
                "horizon_support_status": row.get("horizon_support_status"),
                "horizon_support_label": row.get("horizon_support_label"),
                "horizon_support_note": row.get("horizon_support_note"),
                "backtest_supported_max_horizon": row.get("backtest_supported_max_horizon"),
                "value": float(value) if pd.notna(value) else pd.NA,
                "availability_status": row.get("availability_status"),
                "forecast_available": row.get("forecast_available"),
                "gap_code": row.get("gap_code"),
                "gap_reason": row.get("gap_reason"),
                "capability_status": row.get("capability_status"),
                "scorer_version": row.get("scorer_version"),
                "source_artifact_hashes": row.get("source_artifact_hashes"),
                "parity_status": row.get("parity_status"),
                "max_parity_delta": row.get("max_parity_delta"),
                "stored_replay_max_delta": row.get("stored_replay_max_delta"),
                "failing_component": row.get("failing_component"),
                "source": "future_forecasts",
            }
        )
    return pd.DataFrame(rows)


def _combine_forecast_chart_rows(scenario_results: list[ForecastRunResult]) -> pd.DataFrame:
    if not scenario_results:
        return pd.DataFrame()
    rows = pd.concat([result.forecast_chart_rows for result in scenario_results], ignore_index=True, sort=False)
    if rows.empty:
        return rows
    historical = rows[rows["row_type"].astype(str).eq("historical_actual")].drop_duplicates(
        subset=["row_type", "stream", "period"],
        keep="first",
    )
    future = rows[~rows["row_type"].astype(str).eq("historical_actual")]
    combined = pd.concat([historical, future], ignore_index=True, sort=False)
    combined["period_key"] = combined["period"].map(lambda value: quarter_sort_key(str(value)))
    combined = combined.sort_values(["stream", "row_type", "scenario_name", "period_key"], kind="stable").drop(columns=["period_key"])
    return combined.reset_index(drop=True)


def _scenario_role_records(scenario_results: list[ForecastRunResult]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    errors: list[str] = []
    for index, result in enumerate(scenario_results):
        manifest = result.manifest
        scenario_name = sanitize_scenario_name(str(manifest.get("scenario_name", f"scenario_{index + 1}")))
        workbook_filename = str(manifest.get("workbook_filename") or "")
        explicit_role = manifest.get("scenario_role")
        role = normalize_scenario_role(str(explicit_role)) if explicit_role else None
        role_source = str(manifest.get("scenario_role_source") or "")
        if role is None:
            role, role_source = resolve_scenario_role(
                scenario_role=None,
                scenario_name=scenario_name,
                workbook_filename=workbook_filename,
            )
        if role is None:
            errors.append(
                f"{workbook_filename or scenario_name}: scenario role is ambiguous. "
                "Choose Basecase or Comparison explicitly; upload order is not used."
            )
        records.append(
            {
                "result_index": index,
                "scenario_name": scenario_name,
                "scenario_role": role,
                "scenario_role_source": role_source,
                "scenario_display_name": str(manifest.get("scenario_display_name") or scenario_name),
                "workbook_filename": workbook_filename,
                "workbook_sha256": manifest.get("workbook_sha256"),
                "run_id": manifest.get("run_id"),
                "output_dir": _repo_relative(repo_root_from_here(), result.output_dir),
                "is_test_fixture": bool(manifest.get("is_test_fixture", False)),
            }
        )
    base_count = sum(1 for record in records if record["scenario_role"] == SCENARIO_ROLE_BASECASE)
    comparison_count = sum(1 for record in records if record["scenario_role"] == SCENARIO_ROLE_COMPARISON)
    if base_count != 1:
        errors.append(f"Scenario comparison requires exactly one Basecase role; found {base_count}.")
    if comparison_count < 1:
        errors.append("Scenario comparison requires at least one Comparison role; found 0.")
    if errors:
        raise ValueError("Scenario role validation failed: " + " ".join(errors))
    return records


def _scenario_input_delta_audit(
    scenario_results: list[ForecastRunResult],
    role_records: list[dict[str, Any]] | None = None,
) -> pd.DataFrame:
    columns = [
        "base_scenario",
        "base_scenario_role",
        "base_original_filename",
        "base_sanitized_display_name",
        "base_workbook_sha256",
        "comparison_scenario",
        "comparison_scenario_role",
        "comparison_original_filename",
        "comparison_sanitized_display_name",
        "comparison_workbook_sha256",
        "comparison_is_test_fixture",
        "stream",
        "stream_label",
        "target_period",
        "input_column",
        "base_value",
        "comparison_value",
        "pct_delta",
        "expected_pct_delta",
        "matches_expected_delta",
        "all_required_inputs_plus_2pct",
        "decision_grade_status",
        "assumption_scope_note",
    ]
    if not scenario_results:
        return pd.DataFrame(columns=columns)
    frames = [
        result.assumptions
        for result in scenario_results
        if isinstance(getattr(result, "assumptions", None), pd.DataFrame) and not result.assumptions.empty
    ]
    if not frames:
        return pd.DataFrame(columns=columns)
    assumptions = pd.concat(frames, ignore_index=True, sort=False)
    if "scenario_name" not in assumptions.columns:
        return pd.DataFrame(columns=columns)
    if role_records is None:
        role_records = _scenario_role_records(scenario_results)
    role_by_scenario = {str(record["scenario_name"]): record for record in role_records}
    base_records = [record for record in role_records if record["scenario_role"] == SCENARIO_ROLE_BASECASE]
    comparison_records = [record for record in role_records if record["scenario_role"] == SCENARIO_ROLE_COMPARISON]
    if len(base_records) != 1 or not comparison_records:
        return pd.DataFrame(columns=columns)
    base_record = base_records[0]
    base_scenario = str(base_record["scenario_name"])
    base = assumptions[assumptions["scenario_name"].astype(str).eq(base_scenario)].copy()
    rows: list[dict[str, Any]] = []
    for comparison_record in comparison_records:
        comparison_scenario = str(comparison_record["scenario_name"])
        comparison = assumptions[assumptions["scenario_name"].astype(str).eq(comparison_scenario)].copy()
        expected_delta: float | None = SMOKE_FIXTURE_EXPECTED_DELTA if comparison_record["is_test_fixture"] else None
        for stream in STREAM_ORDER:
            user_columns = [column.name for column in STREAM_COLUMNS[stream] if column.role == "user" and column.required]
            stream_base = base[base["stream"].astype(str).eq(stream)].copy()
            stream_comparison = comparison[comparison["stream"].astype(str).eq(stream)].copy()
            if stream_base.empty or stream_comparison.empty:
                continue
            join_keys = ["stream", "period"]
            merged = stream_base[join_keys + user_columns].merge(
                stream_comparison[join_keys + user_columns],
                on=join_keys,
                how="inner",
                suffixes=("_base", "_comparison"),
            )
            for _, record in merged.iterrows():
                for input_column in user_columns:
                    base_value = pd.to_numeric(record.get(f"{input_column}_base"), errors="coerce")
                    comparison_value = pd.to_numeric(record.get(f"{input_column}_comparison"), errors="coerce")
                    pct_delta = (
                        float(comparison_value) / float(base_value) - 1.0
                        if pd.notna(base_value) and pd.notna(comparison_value) and float(base_value) != 0.0
                        else np.nan
                    )
                    rows.append(
                        {
                            "base_scenario": base_scenario,
                            "base_scenario_role": base_record["scenario_role"],
                            "base_original_filename": base_record["workbook_filename"],
                            "base_sanitized_display_name": base_record["scenario_display_name"],
                            "base_workbook_sha256": base_record["workbook_sha256"],
                            "comparison_scenario": comparison_scenario,
                            "comparison_scenario_role": comparison_record["scenario_role"],
                            "comparison_original_filename": comparison_record["workbook_filename"],
                            "comparison_sanitized_display_name": comparison_record["scenario_display_name"],
                            "comparison_workbook_sha256": comparison_record["workbook_sha256"],
                            "comparison_is_test_fixture": bool(comparison_record["is_test_fixture"]),
                            "stream": stream,
                            "stream_label": STREAM_LABELS[stream],
                            "target_period": str(record.get("period")),
                            "input_column": input_column,
                            "base_value": float(base_value) if pd.notna(base_value) else pd.NA,
                            "comparison_value": float(comparison_value) if pd.notna(comparison_value) else pd.NA,
                            "pct_delta": pct_delta,
                            "expected_pct_delta": expected_delta,
                            "matches_expected_delta": (
                                bool(np.isfinite(pct_delta) and abs(pct_delta - expected_delta) <= SMOKE_FIXTURE_DELTA_TOLERANCE)
                                if expected_delta is not None
                                else pd.NA
                            ),
                            "all_required_inputs_plus_2pct": pd.NA,
                            "decision_grade_status": "pending_delta_summary",
                            "assumption_scope_note": "",
                        }
                    )
    return _finalize_scenario_input_delta_audit(pd.DataFrame(rows, columns=columns))


def _finalize_scenario_input_delta_audit(audit: pd.DataFrame) -> pd.DataFrame:
    if audit.empty:
        return audit
    out = audit.copy()
    for comparison_scenario, group in out.groupby("comparison_scenario", dropna=False):
        idx = group.index
        deltas = pd.to_numeric(group["pct_delta"], errors="coerce")
        finite = deltas[np.isfinite(deltas)]
        is_fixture = bool(group["comparison_is_test_fixture"].fillna(False).astype(bool).all())
        matches = group["matches_expected_delta"]
        all_plus2 = bool(is_fixture and not matches.empty and matches.fillna(False).astype(bool).all())
        out.loc[idx, "all_required_inputs_plus_2pct"] = all_plus2
        if all_plus2:
            status = "not_decision_grade_smoke_fixture"
            note = HIGH_POPULATION_SMOKE_FIXTURE_NOTE
        elif is_fixture:
            status = "test_fixture_delta_mismatch_review_required"
            note = "Workbook is flagged as a test fixture, but not every required user input is +2% versus basecase."
        elif finite.empty:
            status = "scenario_delta_audit_only"
            note = "No finite base/comparison input deltas could be calculated for this scenario."
        else:
            status = "scenario_delta_audit_only"
            note = (
                f"Required user-input deltas range from {float(finite.min()):.3%} to {float(finite.max()):.3%} "
                "versus the resolved basecase; this is not labelled as an all-inputs +2% smoke fixture."
            )
        out.loc[idx, "decision_grade_status"] = status
        out.loc[idx, "assumption_scope_note"] = note
    return out


def _scenario_assumption_delta_summary(audit: pd.DataFrame) -> list[dict[str, Any]]:
    if audit is None or audit.empty:
        return []
    rows: list[dict[str, Any]] = []
    for comparison_scenario, group in audit.groupby("comparison_scenario", dropna=False):
        deltas = pd.to_numeric(group["pct_delta"], errors="coerce")
        finite = deltas[np.isfinite(deltas)]
        rows.append(
            {
                "base_scenario": first_non_null(group["base_scenario"]),
                "comparison_scenario": str(comparison_scenario),
                "comparison_is_test_fixture": bool(group["comparison_is_test_fixture"].fillna(False).astype(bool).all()),
                "all_required_inputs_plus_2pct": bool(group["all_required_inputs_plus_2pct"].fillna(False).astype(bool).all()),
                "decision_grade_status": first_non_null(group["decision_grade_status"]),
                "min_pct_delta": float(finite.min()) if not finite.empty else None,
                "max_pct_delta": float(finite.max()) if not finite.empty else None,
                "assumption_scope_note": first_non_null(group["assumption_scope_note"]),
                "row_count": int(len(group)),
            }
        )
    return rows


def write_forecast_scenario_comparison(
    scenario_results: list[ForecastRunResult],
    output_dir: Path | str | None = None,
    repo_root: Path | str | None = None,
    run_timestamp: str | None = None,
) -> ForecastScenarioComparisonResult:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    role_records = _scenario_role_records(scenario_results)
    timestamp = run_timestamp or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    comparison_id = f"{timestamp}_scenario_comparison"
    comparison_dir = Path(output_dir) if output_dir is not None else root / "artifacts" / "forecast_runs" / comparison_id
    comparison_dir.mkdir(parents=True, exist_ok=True)
    future = (
        pd.concat([result.future_forecasts for result in scenario_results], ignore_index=True, sort=False)
        if scenario_results
        else pd.DataFrame()
    )
    capability = (
        pd.concat([result.capability_report for result in scenario_results], ignore_index=True, sort=False)
        if scenario_results
        else pd.DataFrame()
    )
    chart_rows = _combine_forecast_chart_rows(scenario_results)
    scenario_input_delta_audit = _scenario_input_delta_audit(scenario_results, role_records)
    role_by_scenario = {record["scenario_name"]: record for record in role_records}
    base_scenario = next(record["scenario_name"] for record in role_records if record["scenario_role"] == SCENARIO_ROLE_BASECASE)
    comparison_scenarios = [
        record["scenario_name"] for record in role_records if record["scenario_role"] == SCENARIO_ROLE_COMPARISON
    ]
    scenario_delta_summary = _scenario_assumption_delta_summary(scenario_input_delta_audit)
    manifest = {
        "comparison_id": comparison_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "forecast_runner_version": FORECAST_RUNNER_VERSION,
        "scenario_count": len(scenario_results),
        "scenario_role_validation": {
            "status": "passed",
            "base_scenario": base_scenario,
            "comparison_scenarios": comparison_scenarios,
            "rule": "exactly one basecase role and at least one comparison role; upload order is never used",
        },
        "backtest_supported_max_horizon": BACKTEST_SUPPORTED_MAX_HORIZON,
        "long_range_extrapolation_start_horizon": BACKTEST_SUPPORTED_MAX_HORIZON + 1,
        "horizon_support_note": HORIZON_SUPPORT_NOTE,
        "scenario_assumption_delta_summary": scenario_delta_summary,
        "scenarios": [
            {
                "scenario_name": result.manifest.get("scenario_name"),
                "scenario_role": role_by_scenario.get(result.manifest.get("scenario_name"), {}).get("scenario_role"),
                "scenario_role_source": role_by_scenario.get(result.manifest.get("scenario_name"), {}).get("scenario_role_source"),
                "scenario_display_name": role_by_scenario.get(result.manifest.get("scenario_name"), {}).get("scenario_display_name"),
                "is_test_fixture": role_by_scenario.get(result.manifest.get("scenario_name"), {}).get("is_test_fixture"),
                "run_id": result.manifest.get("run_id"),
                "output_dir": _repo_relative(root, result.output_dir),
                "workbook_filename": result.manifest.get("workbook_filename"),
                "workbook_sha256": result.manifest.get("workbook_sha256"),
                "forecast_horizon_quarters": result.manifest.get("forecast_horizon_quarters"),
                "forecast_start_period": result.manifest.get("forecast_start_period"),
                "forecast_end_period": result.manifest.get("forecast_end_period"),
                "forecast_status": result.manifest.get("forecast_status"),
            }
            for result in scenario_results
        ],
        "fixed_finalists_only": True,
        "broad_search_run": False,
        "evidence_pack_modified": False,
        "chart_sources_modified": False,
        "output_files": [
            "forecast_scenario_comparison.parquet",
            "forecast_scenario_comparison.csv",
            "forecast_scenario_capability_report.parquet",
            "forecast_scenario_capability_report.csv",
            "forecast_scenario_chart_rows.parquet",
            "forecast_scenario_chart_rows.csv",
            "scenario_input_delta_audit.parquet",
            "scenario_input_delta_audit.csv",
            "forecast_scenario_comparison_manifest.json",
        ],
    }
    future.to_parquet(comparison_dir / "forecast_scenario_comparison.parquet", index=False)
    future.to_csv(comparison_dir / "forecast_scenario_comparison.csv", index=False)
    capability.to_parquet(comparison_dir / "forecast_scenario_capability_report.parquet", index=False)
    capability.to_csv(comparison_dir / "forecast_scenario_capability_report.csv", index=False)
    chart_rows.to_parquet(comparison_dir / "forecast_scenario_chart_rows.parquet", index=False)
    chart_rows.to_csv(comparison_dir / "forecast_scenario_chart_rows.csv", index=False)
    scenario_input_delta_audit.to_parquet(comparison_dir / "scenario_input_delta_audit.parquet", index=False)
    scenario_input_delta_audit.to_csv(comparison_dir / "scenario_input_delta_audit.csv", index=False)
    (comparison_dir / "forecast_scenario_comparison_manifest.json").write_text(
        json.dumps(manifest, indent=2, allow_nan=False),
        encoding="utf-8",
    )
    return ForecastScenarioComparisonResult(
        output_dir=comparison_dir,
        manifest=manifest,
        scenario_results=scenario_results,
        future_forecasts=future,
        capability_report=capability,
        forecast_chart_rows=chart_rows,
        scenario_input_delta_audit=scenario_input_delta_audit,
    )


def create_completed_sample_workbook(
    output_path: Path | str,
    repo_root: Path | str | None = None,
    *,
    quarters: int | None = None,
    end_period: str | None = None,
    value_multiplier: float = 1.0,
) -> Path:
    path = build_forecast_input_template(output_path, repo_root=repo_root, quarters=quarters, end_period=end_period)
    wb = load_workbook(path)
    sample_values = {
        "real_gdp_per_capita_nzd": 72000.0,
        "population": 5300000.0,
        "unemployment_rate": 4.2,
        "real_petrol_price_cents_per_litre": 265.0,
        "real_gdp_sa_nzd": 320000000000.0,
        "real_diesel_price_cents_per_litre": 232.0,
        "real_light_ruc_price_nzd_per_1000km": 84.0,
        "lagged_real_light_ruc_price_nzd_per_1000km": 82.0,
        "real_heavy_ruc_price_nzd_per_1000km": 188.0,
        "lead_real_heavy_ruc_price_nzd_per_1000km": 190.0,
        "target_lag_1": 1000.0,
        "target_lag_4": 980.0,
    }
    for stream in STREAM_ORDER:
        ws = wb[SHEET_BY_STREAM[stream]]
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        ws.protection.sheet = False
        for row in range(2, ws.max_row + 1):
            for name, value in sample_values.items():
                if name in headers:
                    multiplier = 1.0 + ((row - 2) * 0.004)
                    ws.cell(row=row, column=headers[name], value=value * multiplier * value_multiplier)
        ws.protection.sheet = True
    wb.save(path)
    return path


def model_capability_gap_register(repo_root: Path | str | None = None) -> pd.DataFrame:
    root = Path(repo_root) if repo_root is not None else repo_root_from_here()
    rows: list[dict[str, Any]] = []
    for stream in STREAM_ORDER:
        rows.append(_stream_forward_capability(root, stream))
    return pd.DataFrame(rows)


def _stream_forward_capability(root: Path, stream: str) -> dict[str, Any]:
    if stream == "PED":
        return evaluate_ped_forward_scorer(root).to_capability_record()
    if stream == "HEAVY_RUC":
        return evaluate_heavy_ruc_forward_scorer(root).to_capability_record()

    history = root / MODEL_INPUT_HISTORY_DIR / MODEL_INPUT_HISTORY_FILES[stream]
    manifest = root / MODEL_INPUT_HISTORY_DIR / "manifest.json"
    if stream == "LIGHT_RUC":
        registry = root / "data" / "dashboard_evidence_pack_reproducibility" / "light_ruc" / "model_registry.parquet"
        requirements = root / "requirements.txt"
        required = [history, manifest, registry, requirements]
        missing = missing_paths(root, required)
        hashes = artifact_hashes(root, required)
        sklearn_error = _sklearn_import_error()
        if not missing and sklearn_error is None:
            finalist = governed_finalist(stream)
            return ForwardScorerAudit(
                stream=stream,
                stream_label=STREAM_LABELS[stream],
                model=finalist,
                capability_status=NUMERIC_FORECAST_AVAILABLE,
                gap_code=None,
                gap_reason="",
                repo_artifact_basis=existing_basis(root, required) + "; requirements.txt::scikit-learn",
                scorer_version=LIGHT_RUC_FORWARD_SCORER_VERSION,
                parity_status="passed_repo_local_recipe",
                max_parity_delta=None,
                stored_replay_max_delta=None,
                source_artifact_hashes=hashes,
                required_components=("base_schiff_ols", "residual_gbr", finalist),
                forecast_capability_available=True,
            ).to_capability_record()
        reason_parts: list[str] = []
        if missing:
            reason_parts.append("missing repo-local artifacts: " + ", ".join(missing))
        if sklearn_error is not None:
            reason_parts.append(f"scikit-learn runtime unavailable: {sklearn_error}")
        finalist = governed_finalist(stream)
        return ForwardScorerAudit(
            stream=stream,
            stream_label=STREAM_LABELS[stream],
            model=finalist,
            capability_status=GOVERNED_GAP,
            gap_code=MISSING_CAPABILITY_BY_STREAM[stream],
            gap_reason=MISSING_CAPABILITY_NOTES[stream] + " " + "; ".join(reason_parts),
            repo_artifact_basis=_capability_basis(root, stream),
            scorer_version=LIGHT_RUC_FORWARD_SCORER_VERSION,
            parity_status="not_run_missing_runtime_or_artifacts",
            source_artifact_hashes=hashes,
            missing_artifacts=tuple(missing),
            required_components=("base_schiff_ols", "residual_gbr", finalist),
            forecast_capability_available=False,
        ).to_capability_record()
    return ForwardScorerAudit(
        stream=stream,
        stream_label=STREAM_LABELS[stream],
        model=governed_finalist(stream),
        capability_status=GOVERNED_GAP,
        gap_code=MISSING_CAPABILITY_BY_STREAM[stream],
        gap_reason=MISSING_CAPABILITY_NOTES[stream],
        repo_artifact_basis=_capability_basis(root, stream),
        scorer_version="not_implemented",
        parity_status="not_run",
        forecast_capability_available=False,
    ).to_capability_record()


def _sklearn_import_error() -> str | None:
    try:
        import sklearn  # noqa: F401
    except Exception as exc:
        return f"{type(exc).__name__}: {exc}"
    return None


def _capability_basis(root: Path, stream: str) -> str:
    paths = [
        root / "data" / "dashboard_evidence_pack" / "data" / "model_registry.parquet",
        root / "data" / "dashboard_evidence_pack" / "data" / "model_coefficients.parquet",
        root / MODEL_INPUT_HISTORY_DIR / MODEL_INPUT_HISTORY_FILES[stream],
        root / MODEL_INPUT_HISTORY_DIR / "manifest.json",
    ]
    if stream == "PED":
        paths.append(root / "data" / "dashboard_evidence_pack_reproducibility" / "ped_inner_hpo" / "model_coefficients.parquet")
    if stream == "LIGHT_RUC":
        paths.append(root / "data" / "dashboard_evidence_pack_reproducibility" / "light_ruc" / "model_coefficients.parquet")
    if stream == "HEAVY_RUC":
        paths.append(root / "data" / "dashboard_evidence_pack_reproducibility" / "heavy_ruc" / "model_coefficients.parquet")
    return "; ".join(_repo_relative(root, path) for path in paths if path.exists())


def _load_workbook_from_input(workbook: Path | str | bytes | BinaryIO) -> Any:
    if isinstance(workbook, bytes):
        return load_workbook(BytesIO(workbook), data_only=False)
    if hasattr(workbook, "read"):
        content = workbook.read()
        if hasattr(workbook, "seek"):
            workbook.seek(0)
        return load_workbook(BytesIO(content), data_only=False)
    return load_workbook(Path(workbook), data_only=False)


def _workbook_bytes(workbook: Path | str | bytes | BinaryIO) -> bytes:
    if isinstance(workbook, bytes):
        return workbook
    if hasattr(workbook, "read"):
        content = workbook.read()
        if hasattr(workbook, "seek"):
            workbook.seek(0)
        return content
    path = Path(workbook)
    if not path.exists():
        return b""
    try:
        return path.read_bytes()
    except OSError:
        return b""


def _workbook_name(workbook: Path | str | bytes | BinaryIO) -> str:
    if isinstance(workbook, (str, Path)):
        return Path(workbook).name
    return getattr(workbook, "name", "uploaded_forecast_workbook.xlsx")


def _worksheet_to_frame(ws: Any) -> pd.DataFrame:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        rows.append({header: value for header, value in zip(headers, row, strict=False) if header})
    return pd.DataFrame(rows)


def _validate_sheet_frame(
    frame: pd.DataFrame,
    stream: str,
    latest_actual_period: str,
    *,
    expected_periods: list[str] | None = None,
) -> tuple[pd.DataFrame, list[str], list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    sheet = SHEET_BY_STREAM[stream]
    expected_headers = [column.name for column in STREAM_COLUMNS[stream]]
    missing_headers = [name for name in expected_headers if name not in frame.columns]
    if missing_headers:
        errors.append(f"{sheet}: missing required headers: {', '.join(missing_headers)}")
        return pd.DataFrame(), [], errors, warnings
    if frame.empty:
        errors.append(f"{sheet}: no forecast rows were found")
        return pd.DataFrame(), [], errors, warnings
    required_user_columns = [col.name for col in STREAM_COLUMNS[stream] if col.role == "user" and col.required]
    valid_indices: list[int] = []
    blank_indices: set[int] = set()
    partial_periods: list[str] = []
    for idx, row in frame.iterrows():
        values = pd.to_numeric(row[required_user_columns], errors="coerce")
        filled = values.notna()
        positive = values.gt(0)
        period = str(row.get("period", "")).upper()
        if positive.all():
            valid_indices.append(idx)
        elif not filled.any():
            blank_indices.add(idx)
        else:
            partial_periods.append(period or f"row {idx + 2}")
    if partial_periods:
        errors.append(
            f"{sheet}: rows with partial or non-positive required user entries are not valid forecast rows: "
            + ", ".join(partial_periods[:8])
        )
    if not valid_indices:
        errors.append(f"{sheet}: no valid forecast rows were found; fill at least one continuous quarter.")
        return pd.DataFrame(), [], errors, warnings
    first_valid = min(valid_indices)
    if any(idx < first_valid for idx in blank_indices):
        errors.append(f"{sheet}: valid forecast rows must start at the first generated forecast quarter.")
    observed_periods = frame.loc[valid_indices, "period"].dropna().astype(str).str.upper().tolist()
    expected_continuous = future_quarters_after(latest_actual_period, len(observed_periods))
    if observed_periods != expected_continuous:
        errors.append(
            f"{sheet}: valid forecast rows must be continuous from {expected_continuous[0]} through {expected_continuous[-1]} with no gaps."
        )
    if expected_periods is not None and observed_periods != expected_periods:
        errors.append(
            f"{sheet}: workbook valid rows are {observed_periods[0]} through {observed_periods[-1]}, "
            f"but the requested horizon is {expected_periods[0]} through {expected_periods[-1]}."
        )
    formula_headers = [col.name for col in STREAM_COLUMNS[stream] if col.role == "formula"]
    if not formula_headers:
        warnings.append(f"{sheet}: no formula columns were found")
    return frame.loc[valid_indices].copy(), observed_periods, errors, warnings


def _common_forecast_periods(periods_by_stream: dict[str, list[str]]) -> list[str] | None:
    if not periods_by_stream:
        return None
    values = list(periods_by_stream.values())
    first = values[0]
    if all(periods == first for periods in values[1:]):
        return first
    return None


def _build_stream_assumptions(frame: pd.DataFrame, stream: str, periods: list[str]) -> pd.DataFrame:
    out = frame.copy()
    out = out.head(len(periods)).copy()
    out["stream"] = stream
    out["stream_label"] = STREAM_LABELS[stream]
    out["model"] = governed_finalist(stream)
    out["period"] = periods[: len(out)]
    out["year"] = out["period"].map(lambda value: parse_period(value)[0])
    out["quarter"] = out["period"].map(lambda value: parse_period(value)[1])
    out["horizon"] = range(1, len(out) + 1)
    q = pd.to_numeric(out["quarter"], errors="coerce")
    out["q2_dummy"] = q.eq(2).astype(int)
    out["q3_dummy"] = q.eq(3).astype(int)
    out["q4_dummy"] = q.eq(4).astype(int)
    out["post_2020_dummy"] = pd.to_numeric(out["year"], errors="coerce").ge(2021).astype(int)
    out["trend_index"] = out["period"].map(lambda value: quarter_sort_key(value) - quarter_sort_key("2000Q1") + 1)

    for source in [
        "real_gdp_sa_nzd",
        "real_gdp_per_capita_nzd",
        "population",
        "unemployment_rate",
        "real_petrol_price_cents_per_litre",
        "real_diesel_price_cents_per_litre",
        "real_light_ruc_price_nzd_per_1000km",
        "lagged_real_light_ruc_price_nzd_per_1000km",
        "real_heavy_ruc_price_nzd_per_1000km",
        "lead_real_heavy_ruc_price_nzd_per_1000km",
        "target_lag_1",
        "target_lag_4",
    ]:
        if source in out.columns:
            numeric = pd.to_numeric(out[source], errors="coerce")
            out[source] = numeric
            out[f"log_{source}"] = np.where(numeric > 0, np.log(numeric), np.nan)

    if "log_real_gdp_sa_nzd" in out.columns:
        out["diff_log_real_gdp"] = out["log_real_gdp_sa_nzd"].diff()
        out["log_real_gdp"] = out["log_real_gdp_sa_nzd"]
    if "log_real_gdp_per_capita_nzd" in out.columns:
        out["log_real_gdp_per_capita"] = out["log_real_gdp_per_capita_nzd"]
    if "log_real_petrol_price_cents_per_litre" in out.columns:
        out["log_real_petrol_price"] = out["log_real_petrol_price_cents_per_litre"]
    if "log_real_diesel_price_cents_per_litre" in out.columns:
        out["log_real_diesel_price"] = out["log_real_diesel_price_cents_per_litre"]
    if "log_real_light_ruc_price_nzd_per_1000km" in out.columns:
        out["log_real_light_ruc_price"] = out["log_real_light_ruc_price_nzd_per_1000km"]
    if "log_lagged_real_light_ruc_price_nzd_per_1000km" in out.columns:
        out["log_lagged_real_light_ruc_price"] = out["log_lagged_real_light_ruc_price_nzd_per_1000km"]
    if "log_real_heavy_ruc_price_nzd_per_1000km" in out.columns:
        out["log_real_heavy_ruc_price"] = out["log_real_heavy_ruc_price_nzd_per_1000km"]
    if "log_lead_real_heavy_ruc_price_nzd_per_1000km" in out.columns:
        out["log_lead_real_heavy_ruc_price"] = out["log_lead_real_heavy_ruc_price_nzd_per_1000km"]
    if {"log_target_lag_1", "log_target_lag_4"}.issubset(out.columns):
        out["diff_log_target_lag_1_lag_4"] = out["log_target_lag_1"] - out["log_target_lag_4"]
    if {"log_real_gdp_per_capita", "log_real_petrol_price"}.issubset(out.columns):
        out["gdp_petrol_interaction"] = out["log_real_gdp_per_capita"] * out["log_real_petrol_price"]
    if {"log_real_gdp", "log_real_light_ruc_price"}.issubset(out.columns):
        out["gdp_light_ruc_price_interaction"] = out["log_real_gdp"] * out["log_real_light_ruc_price"]
    if {"log_real_gdp", "log_real_heavy_ruc_price"}.issubset(out.columns):
        out["gdp_heavy_ruc_price_interaction"] = out["log_real_gdp"] * out["log_real_heavy_ruc_price"]
    out["post_2011_x_log_trend"] = np.where(out["trend_index"] > 0, np.log(out["trend_index"]), np.nan)
    out["transform_status"] = "built_from_template_inputs"
    return out


def _forecast_output_rows(
    validation: ForecastValidationResult,
    capabilities: pd.DataFrame,
    repo_root: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    light_future: pd.DataFrame | None = None
    light_components: pd.DataFrame | None = None
    scorer_errors: dict[str, str] = {}
    cap_lookup = capabilities.set_index("stream").to_dict(orient="index")
    if validation.valid and cap_lookup.get("LIGHT_RUC", {}).get("forecast_capability_available"):
        try:
            light_future, light_components = _light_ruc_forward_forecast(validation, repo_root)
        except Exception as exc:
            scorer_errors["LIGHT_RUC"] = f"{type(exc).__name__}: {exc}"

    vnext_outputs: dict[str, tuple[pd.DataFrame, pd.DataFrame]] = {}
    for stream in ("PED", "HEAVY_RUC"):
        if validation.valid and cap_lookup.get(stream, {}).get("forecast_capability_available"):
            try:
                from .vnext_forward_integration import vnext_forward_forecast

                vnext_outputs[stream] = vnext_forward_forecast(validation, repo_root, stream)
            except Exception as exc:
                scorer_errors[stream] = f"{type(exc).__name__}: {exc}"

    future_rows: list[pd.DataFrame] = []
    component_rows: list[pd.DataFrame] = []
    for stream in STREAM_ORDER:
        if stream == "LIGHT_RUC" and light_future is not None and light_components is not None:
            future_rows.append(light_future)
            component_rows.append(light_components)
            continue
        if stream in vnext_outputs:
            future_rows.append(vnext_outputs[stream][0])
            component_rows.append(vnext_outputs[stream][1])
            continue
        error_suffix = f" Scorer error: {scorer_errors[stream]}" if stream in scorer_errors else ""
        future_rows.append(_gap_future_forecast_rows(validation, capabilities, stream, error_suffix=error_suffix))
        component_rows.append(_gap_component_forecast_rows(validation, capabilities, repo_root, stream, error_suffix=error_suffix))
    future = pd.concat(future_rows, ignore_index=True, sort=False) if future_rows else pd.DataFrame()
    components = pd.concat(component_rows, ignore_index=True, sort=False) if component_rows else pd.DataFrame()
    future = _attach_capability_metadata(future, capabilities)
    components = _attach_capability_metadata(components, capabilities)
    future = _add_horizon_support_columns(future)
    components = _add_horizon_support_columns(components)
    return future, components


def _gap_future_forecast_rows(
    validation: ForecastValidationResult,
    capabilities: pd.DataFrame,
    stream: str,
    *,
    error_suffix: str = "",
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    cap_lookup = capabilities.set_index("stream").to_dict(orient="index")
    cap = cap_lookup[stream]
    metadata = _capability_metadata_fields(cap)
    for horizon, period in enumerate(validation.forecast_periods, start=1):
        rows.append(
            {
                **metadata,
                "stream": stream,
                "stream_label": STREAM_LABELS[stream],
                "model": governed_finalist(stream),
                "target_period": period,
                "horizon": horizon,
                "forecast": pd.NA,
                "prediction": pd.NA,
                "forecast_available": False,
                "availability_status": "validation_failed" if not validation.valid else "governed_gap",
                "gap_code": "input_validation_failed" if not validation.valid else cap["gap_code"],
                "gap_reason": "; ".join(validation.errors) if not validation.valid else str(cap["gap_reason"]) + error_suffix,
                "fixed_finalist_only": True,
                "broad_search_run": False,
            }
        )
    return pd.DataFrame(rows)


def _gap_component_forecast_rows(
    validation: ForecastValidationResult,
    capabilities: pd.DataFrame,
    repo_root: Path,
    stream: str,
    *,
    error_suffix: str = "",
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    cap_lookup = capabilities.set_index("stream").to_dict(orient="index")
    components = _component_registry(repo_root)
    cap = cap_lookup[stream]
    metadata = _capability_metadata_fields(cap)
    stream_components = components.get(stream, [])
    for horizon, period in enumerate(validation.forecast_periods, start=1):
        for component in stream_components:
            rows.append(
                {
                    **metadata,
                    "stream": stream,
                    "stream_label": STREAM_LABELS[stream],
                    "model": governed_finalist(stream),
                    "component_model": component.get("component_model"),
                    "component_role": component.get("component_role"),
                    "component_weight": component.get("component_weight"),
                    "target_period": period,
                    "horizon": horizon,
                    "component_forecast": pd.NA,
                    "component_log_value": pd.NA,
                    "weighted_component_forecast": pd.NA,
                    "forecast_available": False,
                    "availability_status": "validation_failed" if not validation.valid else "governed_gap",
                    "gap_code": "input_validation_failed" if not validation.valid else cap["gap_code"],
                    "gap_reason": "; ".join(validation.errors) if not validation.valid else str(cap["gap_reason"]) + error_suffix,
                    "fixed_finalist_only": True,
                    "broad_search_run": False,
                }
            )
    return pd.DataFrame(rows)


def _capability_metadata_fields(cap: dict[str, Any] | pd.Series) -> dict[str, Any]:
    return {column: cap.get(column, pd.NA) for column in CAPABILITY_METADATA_COLUMNS}


def _attach_capability_metadata(frame: pd.DataFrame, capabilities: pd.DataFrame) -> pd.DataFrame:
    if frame is None or frame.empty or "stream" not in frame.columns or capabilities is None or capabilities.empty:
        return pd.DataFrame() if frame is None else frame
    cap_lookup = capabilities.set_index("stream").to_dict(orient="index")
    out = frame.copy()
    for column in CAPABILITY_METADATA_COLUMNS:
        out[column] = out["stream"].map(lambda stream: cap_lookup.get(str(stream), {}).get(column, pd.NA))
    return out


def _light_ruc_forward_forecast(validation: ForecastValidationResult, repo_root: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    from sklearn.ensemble import GradientBoostingRegressor

    history_path = repo_root / MODEL_INPUT_HISTORY_DIR / MODEL_INPUT_HISTORY_FILES["LIGHT_RUC"]
    history = pd.read_parquet(history_path)
    future = validation.assumptions[validation.assumptions["stream"].astype(str).eq("LIGHT_RUC")].copy()
    if future.empty:
        raise ValueError("Light RUC future assumption rows are missing from the validated workbook.")

    feature_frame = _light_ruc_feature_frame(history, future, validation.latest_actual_period)
    train = feature_frame[feature_frame["sample_scope"].eq("history")].copy()
    train = train[train["period_key"].le(quarter_sort_key(validation.latest_actual_period))].copy()
    train["target"] = pd.to_numeric(train["target"], errors="coerce")
    required = ["target", *LIGHT_RUC_BASE_FEATURES, *LIGHT_RUC_RESIDUAL_FEATURES]
    train = train.replace([np.inf, -np.inf], np.nan).dropna(subset=required)
    train = train[train["target"].gt(0)].sort_values("period_key").tail(LIGHT_RUC_WINDOW).copy()
    if len(train) < LIGHT_RUC_WINDOW:
        raise ValueError(f"Light RUC scorer requires {LIGHT_RUC_WINDOW} usable training rows; found {len(train)}.")

    future_features = feature_frame[feature_frame["sample_scope"].eq("future")].copy()
    future_features = future_features.replace([np.inf, -np.inf], np.nan)
    missing_future = _missing_feature_columns(future_features, LIGHT_RUC_RESIDUAL_FEATURES)
    if missing_future:
        raise ValueError("Light RUC future rows are missing required residual features: " + ", ".join(missing_future))

    y = np.log(train["target"].to_numpy(dtype=float))
    base_x = train[LIGHT_RUC_BASE_FEATURES].to_numpy(dtype=float)
    beta = _ols_fit(base_x, y)
    train_base_log = _ols_predict(base_x, beta)
    residual_target = y - train_base_log
    residual_model = GradientBoostingRegressor(
        n_estimators=150,
        max_depth=1,
        learning_rate=0.05,
        subsample=0.85,
        random_state=42,
        loss="squared_error",
    )
    residual_model.fit(train[LIGHT_RUC_RESIDUAL_FEATURES].to_numpy(dtype=float), residual_target)

    future_base_log = _ols_predict(future_features[LIGHT_RUC_BASE_FEATURES].to_numpy(dtype=float), beta)
    future_residual_log = residual_model.predict(future_features[LIGHT_RUC_RESIDUAL_FEATURES].to_numpy(dtype=float))
    final_log = future_base_log + future_residual_log
    forecast = np.exp(final_log)

    future_rows: list[dict[str, Any]] = []
    component_rows: list[dict[str, Any]] = []
    train_start = str(train.iloc[0]["period"])
    train_end = str(train.iloc[-1]["period"])
    for idx, (_, row) in enumerate(future_features.sort_values("horizon").iterrows()):
        horizon = int(row["horizon"])
        period = str(row["period"])
        base_log = float(future_base_log[idx])
        residual_log = float(future_residual_log[idx])
        final_log_value = float(final_log[idx])
        final_prediction = float(forecast[idx])
        base_prediction = float(np.exp(base_log))
        residual_multiplier = float(np.exp(residual_log))
        common = {
            "stream": "LIGHT_RUC",
            "stream_label": STREAM_LABELS["LIGHT_RUC"],
            "model": governed_finalist("LIGHT_RUC"),
            "target_period": period,
            "horizon": horizon,
            "availability_status": "numeric_forecast_available",
            "gap_code": None,
            "gap_reason": "",
            "fixed_finalist_only": True,
            "broad_search_run": False,
            "source_recipe": "Schiff-style OLS base plus GradientBoostingRegressor residual correction",
            "training_window_start": train_start,
            "training_window_end": train_end,
            "training_window_rows": len(train),
            "score_basis": "forward_assumption_workbook",
        }
        future_rows.append(
            {
                **common,
                "forecast": final_prediction,
                "prediction": final_prediction,
                "forecast_available": True,
                "base_forecast": base_prediction,
                "residual_log_correction": residual_log,
                "residual_multiplier": residual_multiplier,
                "final_log_forecast": final_log_value,
            }
        )
        component_rows.extend(
            [
                {
                    **common,
                    "component_model": "base_schiff_ols",
                    "component_role": "OLS base level prediction",
                    "component_weight": pd.NA,
                    "component_forecast": base_prediction,
                    "component_log_value": base_log,
                    "weighted_component_forecast": pd.NA,
                    "forecast_available": True,
                },
                {
                    **common,
                    "component_model": "residual_gbr",
                    "component_role": "GBM residual log-correction multiplier",
                    "component_weight": pd.NA,
                    "component_forecast": residual_multiplier,
                    "component_log_value": residual_log,
                    "weighted_component_forecast": pd.NA,
                    "forecast_available": True,
                },
                {
                    **common,
                    "component_model": governed_finalist("LIGHT_RUC"),
                    "component_role": "final level prediction",
                    "component_weight": 1.0,
                    "component_forecast": final_prediction,
                    "component_log_value": final_log_value,
                    "weighted_component_forecast": final_prediction,
                    "forecast_available": True,
                },
            ]
        )
    return pd.DataFrame(future_rows), pd.DataFrame(component_rows)


def _light_ruc_feature_frame(history: pd.DataFrame, future: pd.DataFrame, latest_actual_period: str) -> pd.DataFrame:
    history_cols = [
        "period",
        "year",
        "quarter",
        "target",
        "real_gdp_sa_nzd",
        "real_diesel_price_cents_per_litre",
        "real_light_ruc_price_nzd_per_1000km",
        "lagged_real_light_ruc_price_nzd_per_1000km",
        "log_real_gdp",
        "log_real_diesel_price",
        "log_real_light_ruc_price",
        "log_lagged_real_light_ruc_price",
        "q2_dummy",
        "q3_dummy",
        "q4_dummy",
        "post_2020_dummy",
    ]
    hist = history[[column for column in history_cols if column in history.columns]].copy()
    hist["sample_scope"] = "history"
    hist["horizon"] = pd.NA
    fut = future.copy()
    fut["sample_scope"] = "future"
    fut["target"] = pd.NA
    combined = pd.concat([hist, fut], ignore_index=True, sort=False)
    combined["period"] = combined["period"].astype(str).str.upper()
    combined["period_key"] = combined["period"].map(quarter_sort_key)
    combined = combined.sort_values("period_key").reset_index(drop=True)
    first_history_key = combined.loc[combined["sample_scope"].eq("history"), "period_key"].min()
    combined["period_index"] = combined["period_key"] - int(first_history_key) + 1
    combined["time_trend"] = combined["period_index"]
    combined["log_trend"] = np.where(combined["time_trend"] > 0, np.log(combined["time_trend"]), np.nan)
    combined["year"] = combined["period"].map(lambda value: parse_period(value)[0])
    combined["quarter"] = combined["period"].map(lambda value: parse_period(value)[1])
    q = pd.to_numeric(combined["quarter"], errors="coerce")
    combined["q2_dummy"] = q.eq(2).astype(int)
    combined["q3_dummy"] = q.eq(3).astype(int)
    combined["q4_dummy"] = q.eq(4).astype(int)
    combined["post_2020_dummy"] = pd.to_numeric(combined["year"], errors="coerce").ge(2021).astype(int)
    _ensure_log_column(combined, "real_gdp_sa_nzd", "log_real_gdp")
    _ensure_log_column(combined, "real_diesel_price_cents_per_litre", "log_real_diesel_price")
    _ensure_log_column(combined, "real_light_ruc_price_nzd_per_1000km", "log_real_light_ruc_price")
    _ensure_log_column(combined, "lagged_real_light_ruc_price_nzd_per_1000km", "log_lagged_real_light_ruc_price")
    log_diesel = pd.to_numeric(combined["log_real_diesel_price"], errors="coerce")
    log_ruc = pd.to_numeric(combined["log_real_light_ruc_price"], errors="coerce")
    log_gdp = pd.to_numeric(combined["log_real_gdp"], errors="coerce")
    post = pd.to_numeric(combined["post_2020_dummy"], errors="coerce").fillna(0)
    combined["diesel_x_ruc_price"] = log_diesel * log_ruc
    combined["gdp_x_post2020"] = log_gdp * post
    combined["ruc_x_post2020"] = log_ruc * post
    combined["diesel_x_post2020"] = log_diesel * post
    combined["log_real_diesel_price_diff1"] = log_diesel.diff()
    combined["log_real_diesel_price_lag1"] = log_diesel.shift(1)
    combined["log_real_diesel_price_lag4"] = log_diesel.shift(4)
    combined["log_real_light_ruc_price_diff1"] = log_ruc.diff()
    combined["log_real_light_ruc_price_lag1"] = log_ruc.shift(1)
    combined["log_real_light_ruc_price_lag4"] = log_ruc.shift(4)
    combined["log_real_gdp_diff1"] = log_gdp.diff()
    combined["log_real_gdp_lag1"] = log_gdp.shift(1)
    combined["log_real_gdp_lag4"] = log_gdp.shift(4)
    combined.loc[combined["period_key"].gt(quarter_sort_key(latest_actual_period)) & combined["sample_scope"].eq("history"), "sample_scope"] = (
        "future_placeholder"
    )
    return combined


def _ensure_log_column(frame: pd.DataFrame, source: str, target: str) -> None:
    if target in frame.columns and pd.to_numeric(frame[target], errors="coerce").notna().any():
        frame[target] = pd.to_numeric(frame[target], errors="coerce")
        return
    numeric = pd.to_numeric(frame.get(source), errors="coerce")
    frame[target] = np.where(numeric > 0, np.log(numeric), np.nan)


def _missing_feature_columns(frame: pd.DataFrame, columns: list[str]) -> list[str]:
    missing: list[str] = []
    for column in columns:
        if column not in frame.columns or pd.to_numeric(frame[column], errors="coerce").isna().any():
            missing.append(column)
    return missing


def _ols_fit(features: np.ndarray, target: np.ndarray) -> np.ndarray:
    design = np.column_stack([np.ones(len(features)), features])
    return np.linalg.lstsq(design, target, rcond=None)[0]


def _ols_predict(features: np.ndarray, beta: np.ndarray) -> np.ndarray:
    design = np.column_stack([np.ones(len(features)), features])
    return design @ beta


def _component_registry(repo_root: Path) -> dict[str, list[dict[str, Any]]]:
    heavy_components: list[dict[str, Any]] = []
    heavy_path = repo_root / "data" / "dashboard_evidence_pack" / "data" / "ensemble_components.parquet"
    if heavy_path.exists():
        try:
            heavy = pd.read_parquet(heavy_path)
            heavy = heavy[heavy["stream"].astype(str).eq("HEAVY_RUC")]
            for _, row in heavy.iterrows():
                heavy_components.append(
                    {
                        "component_model": row.get("component_model"),
                        "component_role": "weighted_ensemble_component",
                        "component_weight": pd.to_numeric(row.get("weight"), errors="coerce"),
                    }
                )
        except Exception:
            heavy_components = []
    return {
        "PED": [
            {
                "component_model": "hpo::PED__HPOREFINE_solver_static_convex_top18",
                "component_role": "outer_hpo_static_solver_component",
                "component_weight": 1.0,
            },
            {
                "component_model": "PED inner HPO/static-solver members",
                "component_role": "inner_component_trace",
                "component_weight": pd.NA,
            },
        ],
        "LIGHT_RUC": [
            {"component_model": "base_schiff_ols", "component_role": "OLS base level prediction", "component_weight": pd.NA},
            {
                "component_model": "residual_gbm_correction",
                "component_role": "GBM residual correction",
                "component_weight": pd.NA,
            },
        ],
        "HEAVY_RUC": heavy_components
        or [
            {"component_model": "C1-C4 heavy RUC components", "component_role": "weighted_ensemble_component", "component_weight": pd.NA}
        ],
    }


def _any_forecast_available(capabilities: pd.DataFrame) -> bool:
    if capabilities.empty or "forecast_capability_available" not in capabilities.columns:
        return False
    return capabilities["forecast_capability_available"].fillna(False).astype(bool).any()


def _forecast_status(validation: ForecastValidationResult, future: pd.DataFrame) -> str:
    if not validation.valid:
        return "validation_failed"
    if future.empty or "forecast_available" not in future.columns:
        return "governed_gap"
    by_stream = future.groupby("stream")["forecast_available"].any()
    available_count = int(by_stream.fillna(False).astype(bool).sum())
    if available_count == 0:
        return "governed_gap"
    if available_count == len(STREAM_ORDER):
        return "numeric_forecast_available"
    return "mixed_numeric_and_governed_gap"


def _add_scenario_columns(frame: pd.DataFrame, scenario_name: str, scenario_role: str | None = None) -> None:
    if frame.empty:
        frame["scenario_name"] = pd.Series(dtype=str)
        frame["scenario_role"] = pd.Series(dtype=str)
        return
    if "scenario_name" in frame.columns:
        frame["scenario_name"] = scenario_name
    else:
        frame.insert(0, "scenario_name", scenario_name)
    if "scenario_role" in frame.columns:
        frame["scenario_role"] = scenario_role
    else:
        frame.insert(1, "scenario_role", scenario_role)


def _forecast_capability_report(
    capabilities: pd.DataFrame,
    future: pd.DataFrame,
    components: pd.DataFrame,
    validation: ForecastValidationResult,
    *,
    scenario_name: str,
    scenario_role: str | None = None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    future_group = future.groupby("stream", dropna=False) if not future.empty and "stream" in future.columns else {}
    component_group = components.groupby("stream", dropna=False) if not components.empty and "stream" in components.columns else {}
    for _, cap in capabilities.iterrows():
        stream = str(cap["stream"])
        stream_future = future_group.get_group(stream) if hasattr(future_group, "groups") and stream in future_group.groups else pd.DataFrame()
        stream_components = (
            component_group.get_group(stream) if hasattr(component_group, "groups") and stream in component_group.groups else pd.DataFrame()
        )
        forecast_available = bool(stream_future.get("forecast_available", pd.Series(dtype=bool)).fillna(False).astype(bool).any())
        numeric_count = int(pd.to_numeric(stream_future.get("forecast"), errors="coerce").notna().sum()) if not stream_future.empty else 0
        gap_reason = ""
        gap_code = None
        if not forecast_available and not stream_future.empty:
            gap_code = first_non_null(stream_future.get("gap_code"))
            gap_reason = str(first_non_null(stream_future.get("gap_reason")) or "")
        capability_status = NUMERIC_FORECAST_AVAILABLE if forecast_available else cap.get("capability_status")
        rows.append(
            {
                "scenario_name": scenario_name,
                "scenario_role": scenario_role,
                "stream": stream,
                "stream_label": cap.get("stream_label"),
                "model": cap.get("model"),
                "validation_status": "passed" if validation.valid else "failed",
                "forecast_available": forecast_available,
                "numeric_forecast_rows": numeric_count,
                "governed_gap_rows": int(len(stream_future) - numeric_count),
                "component_trace_rows": int(len(stream_components)),
                "capability_status": capability_status,
                "gap_code": gap_code if gap_code is not None else cap.get("gap_code"),
                "gap_reason": gap_reason if gap_reason else cap.get("gap_reason"),
                "repo_artifact_basis": cap.get("repo_artifact_basis"),
                "scorer_version": cap.get("scorer_version"),
                "source_artifact_hashes": cap.get("source_artifact_hashes"),
                "parity_status": cap.get("parity_status"),
                "max_parity_delta": cap.get("max_parity_delta"),
                "stored_replay_max_delta": cap.get("stored_replay_max_delta"),
                "required_components": cap.get("required_components"),
                "missing_artifacts": cap.get("missing_artifacts"),
                "failing_component": cap.get("failing_component"),
                "fixed_finalists_only": True,
                "broad_search_run": False,
            }
        )
    return pd.DataFrame(rows)


def _forecast_validation_report(validation: ForecastValidationResult, capabilities: pd.DataFrame) -> str:
    lines = [
        "# Forecast validation report",
        "",
        f"- Runner version: `{FORECAST_RUNNER_VERSION}`",
        f"- Latest known actual quarter: `{validation.latest_actual_period}`",
        f"- Forecast horizon: `{validation.forecast_periods[0]}` to `{validation.forecast_periods[-1]}`",
        f"- Horizon support: {HORIZON_SUPPORT_NOTE}",
        f"- Input validation: `{'passed' if validation.valid else 'failed'}`",
        "- Broad candidate search run: `false`",
        "- Existing dashboard evidence/KPI/chart-source files modified: `false`",
        "",
        "## Validation messages",
        "",
    ]
    if validation.errors:
        lines.extend(f"- ERROR: {message}" for message in validation.errors)
    if validation.warnings:
        lines.extend(f"- WARNING: {message}" for message in validation.warnings)
    if not validation.errors and not validation.warnings:
        lines.append("- OK: Workbook inputs passed validation.")
    lines.extend(["", "## Forecast capability", ""])
    for _, row in capabilities.iterrows():
        lines.extend(
            [
                f"### {row['stream_label']}",
                "",
                f"- Fixed finalist: `{row['model']}`",
                f"- Status: `{row['capability_status']}`",
                f"- Scorer version: `{row.get('scorer_version')}`",
                f"- Parity status: `{row.get('parity_status')}`",
                f"- Max parity delta: `{row.get('max_parity_delta')}`",
                f"- Failing component: `{row.get('failing_component')}`",
                f"- Gap code: `{row['gap_code']}`",
                f"- Reason: {row['gap_reason'] or 'Repo-local fixed-finalist scorer is available.'}",
                f"- Artifact basis: `{row['repo_artifact_basis']}`",
                "",
            ]
        )
    return "\n".join(lines).strip() + "\n"


def _repo_relative(root: Path, path: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.name


def first_non_null(series: pd.Series | None) -> Any:
    if series is None:
        return None
    values = series.dropna()
    return values.iloc[0] if not values.empty else None


def _json_safe_records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for raw in frame.to_dict(orient="records"):
        record: dict[str, Any] = {}
        for key, value in raw.items():
            if pd.isna(value):
                record[key] = None
            elif isinstance(value, np.generic):
                record[key] = value.item()
            else:
                record[key] = value
        records.append(record)
    return records
