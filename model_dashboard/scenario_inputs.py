from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import hashlib
import json
from pathlib import Path
import re
import shutil
from typing import Any, BinaryIO, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SCENARIO_INPUT_SCHEMA_VERSION = "nltf-scenario-input-materializer-v1"
SCENARIO_INPUT_DIRNAME = "scenario_inputs"
RAW_WORKBOOK_SIZE_LIMIT_BYTES = 50 * 1024 * 1024

SCENARIO_INPUT_CELLS_STEM = "scenario_input_cells"
SCENARIO_INPUT_LONG_STEM = "scenario_input_long"
SCENARIO_INPUT_WIDE_STEM = "scenario_input_wide"
SCENARIO_FEATURE_LINEAGE_STEM = "scenario_feature_lineage"
SCENARIO_INPUT_MANIFEST = "scenario_input_manifest.json"

TEMPLATE_SHEET_STREAMS = {
    "PED Inputs": "PED",
    "Light RUC Inputs": "LIGHT_RUC",
    "Heavy RUC Inputs": "HEAVY_RUC",
}

SYSTEM_VARIABLES = {"period", "year", "quarter", "horizon"}

LIGHT_RUC_MODEL_FEATURE_SOURCES = {
    "log_real_diesel_price": ("log_real_diesel_price",),
    "log_real_light_ruc_price": ("log_real_light_ruc_price",),
    "log_lagged_real_light_ruc_price": ("log_lagged_real_light_ruc_price",),
    "log_real_gdp": ("log_real_gdp",),
    "post_2020_dummy": ("post_2020_dummy",),
    "q2_dummy": ("q2_dummy",),
    "q3_dummy": ("q3_dummy",),
    "q4_dummy": ("q4_dummy",),
    "diesel_x_ruc_price": ("log_real_diesel_price", "log_real_light_ruc_price"),
    "gdp_x_post2020": ("log_real_gdp", "post_2020_dummy"),
    "ruc_x_post2020": ("log_real_light_ruc_price", "post_2020_dummy"),
    "diesel_x_post2020": ("log_real_diesel_price", "post_2020_dummy"),
    "time_trend": ("trend_index",),
    "log_trend": ("trend_index",),
    "log_real_diesel_price_diff1": ("log_real_diesel_price",),
    "log_real_diesel_price_lag1": ("log_real_diesel_price",),
    "log_real_diesel_price_lag4": ("log_real_diesel_price",),
    "log_real_light_ruc_price_diff1": ("log_real_light_ruc_price",),
    "log_real_light_ruc_price_lag1": ("log_real_light_ruc_price",),
    "log_real_light_ruc_price_lag4": ("log_real_light_ruc_price",),
    "log_real_gdp_diff1": ("log_real_gdp",),
    "log_real_gdp_lag1": ("log_real_gdp",),
    "log_real_gdp_lag4": ("log_real_gdp",),
}

VNEXT_TARGET_LAG_FEATURES = (
    "target__lag1",
    "target__lag2",
    "target__lag4",
    "target__diff1",
    "target__diff4",
    "target__roll4_mean",
    "target__roll8_mean",
)

VNEXT_TIME_FEATURES = (
    "time__trend",
    "time__trend_sq",
    "time__post2011",
    "time__post2011_trend",
    "time__post2020",
    "time__covid2020",
    "time__q2",
    "time__q3",
    "time__q4",
)

VNEXT_LEVEL_SOURCES = {
    "PED": {
        "gdp_pc": "real_gdp_per_capita_nzd",
        "petrol": "real_petrol_price_cents_per_litre",
        "unemp": "unemployment_rate",
        "population": "population",
    },
    "HEAVY_RUC": {
        "gdp": "real_gdp_sa_nzd",
        "diesel": "real_diesel_price_cents_per_litre",
        "heavy_price": "real_heavy_ruc_price_nzd_per_1000km",
        "light_price": "real_light_ruc_price_nzd_per_1000km",
        "unemp": "unemployment_rate",
    },
}

VNEXT_LOG_SOURCES = {
    "PED": {
        "gdp_pc": "log_real_gdp_per_capita",
        "petrol": "log_real_petrol_price",
        "unemp": "log_unemployment_rate",
    },
    "HEAVY_RUC": {
        "gdp": "log_real_gdp",
        "diesel": "log_real_diesel_price",
        "heavy_price": "log_real_heavy_ruc_price",
        "light_price": "log_real_light_ruc_price",
        "unemp": "log_unemployment_rate",
    },
}

VNEXT_POLICY_PRICE_BASES = {
    "PED": ("petrol",),
    "HEAVY_RUC": ("heavy_price", "diesel"),
}

UNIT_HINTS = {
    "population": "persons",
    "real_gdp_per_capita_nzd": "real NZD/person",
    "real_gdp_sa_nzd": "real NZD",
    "unemployment_rate": "percentage points",
    "real_petrol_price_cents_per_litre": "cents/litre",
    "real_diesel_price_cents_per_litre": "cents/litre",
    "ped_base_rate_cents_per_litre": "cents/litre",
    "real_light_ruc_price_nzd_per_1000km": "NZD/1000km",
    "lagged_real_light_ruc_price_nzd_per_1000km": "NZD/1000km",
    "light_ruc_nominal_rate_nzd_per_1000km": "NZD/1000km",
    "real_heavy_ruc_price_nzd_per_1000km": "NZD/1000km",
    "lead_real_heavy_ruc_price_nzd_per_1000km": "NZD/1000km",
    "heavy_ruc_nominal_rate_nzd_per_1000km": "NZD/1000km",
    "target_lag_1": "model target units",
    "target_lag_4": "model target units",
}


@dataclass(frozen=True)
class ScenarioWorkbookInput:
    workbook: Path | str | bytes | BinaryIO
    scenario_name: str
    scenario_role: str
    workbook_filename: str


def materialize_scenario_inputs(
    workbooks: Iterable[ScenarioWorkbookInput],
    output_dir: Path | str,
    *,
    copy_raw: bool = True,
    raw_size_limit_bytes: int = RAW_WORKBOOK_SIZE_LIMIT_BYTES,
    created_by: str = "scenario_input_materializer",
    repo_root: Path | str | None = None,
) -> dict[str, Any]:
    """Materialize scenario workbook contents into repo-local audit artifacts.

    The output is intentionally plain parquet/csv/json so Streamlit can read
    committed scenario variables without loading Excel. Raw workbook copies are
    stored only when they are below the governed size limit.
    """

    output = Path(output_dir)
    root = Path(repo_root).resolve() if repo_root is not None else Path.cwd().resolve()
    fallback_base = output.parent
    output.mkdir(parents=True, exist_ok=True)
    raw_dir = output / "raw"
    if copy_raw:
        raw_dir.mkdir(parents=True, exist_ok=True)

    all_cells: list[pd.DataFrame] = []
    all_long: list[pd.DataFrame] = []
    all_wide: list[pd.DataFrame] = []
    workbook_records: list[dict[str, Any]] = []
    sheet_inventory_records: list[dict[str, Any]] = []

    for item in workbooks:
        workbook_bytes = _workbook_bytes(item.workbook)
        workbook_sha = hashlib.sha256(workbook_bytes).hexdigest()
        workbook_size = len(workbook_bytes)
        raw_repo_path = ""
        raw_status = "hash_only"
        if copy_raw and workbook_bytes and workbook_size <= raw_size_limit_bytes:
            raw_name = f"{workbook_sha[:12]}_{_safe_filename(item.workbook_filename)}"
            raw_path = raw_dir / raw_name
            raw_path.write_bytes(workbook_bytes)
            raw_repo_path = _relative_path(raw_path, root, fallback_base=fallback_base)
            raw_status = "copied_repo_local_raw_workbook"
        elif workbook_size > raw_size_limit_bytes:
            raw_status = "too_large_hash_only"

        wb = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
        cells = _workbook_cells_frame(
            wb,
            scenario_name=item.scenario_name,
            scenario_role=item.scenario_role,
            workbook_filename=item.workbook_filename,
            workbook_sha256=workbook_sha,
        )
        long = _scenario_input_long_from_cells(cells)
        wide = _scenario_input_wide_from_long(long)
        sheet_inventory = _sheet_inventory_from_cells(cells)
        all_cells.append(cells)
        all_long.append(long)
        all_wide.append(wide)
        sheet_inventory_records.extend(sheet_inventory)
        workbook_records.append(
            {
                "scenario_name": item.scenario_name,
                "role": item.scenario_role,
                "workbook_filename": item.workbook_filename,
                "workbook_sha256": workbook_sha,
                "size_bytes": workbook_size,
                "raw_status": raw_status,
                "raw_repo_relative_path": raw_repo_path,
                "sheet_count": len(wb.sheetnames),
                "non_empty_cell_count": int(len(cells)),
                "long_row_count": int(len(long)),
                "wide_row_count": int(len(wide)),
                "sheet_inventory": sheet_inventory,
            }
        )

    cells_frame = pd.concat(all_cells, ignore_index=True, sort=False) if all_cells else _empty_cells_frame()
    long_frame = pd.concat(all_long, ignore_index=True, sort=False) if all_long else _empty_long_frame()
    wide_frame = pd.concat(all_wide, ignore_index=True, sort=False) if all_wide else _empty_wide_frame()
    lineage_frame = scenario_feature_lineage_from_wide(wide_frame)

    output_files = _write_artifact_pair(output, SCENARIO_INPUT_CELLS_STEM, cells_frame, root=root, fallback_base=fallback_base)
    output_files.update(_write_artifact_pair(output, SCENARIO_INPUT_LONG_STEM, long_frame, root=root, fallback_base=fallback_base))
    output_files.update(_write_artifact_pair(output, SCENARIO_INPUT_WIDE_STEM, wide_frame, root=root, fallback_base=fallback_base))
    output_files.update(_write_artifact_pair(output, SCENARIO_FEATURE_LINEAGE_STEM, lineage_frame, root=root, fallback_base=fallback_base))

    manifest = {
        "schema_version": SCENARIO_INPUT_SCHEMA_VERSION,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": created_by,
        "repo_relative_output_dir": _relative_path(output, root, fallback_base=fallback_base),
        "source_policy": "committed scenario input artifacts only; Streamlit must not load Excel at runtime",
        "raw_workbook_size_limit_bytes": int(raw_size_limit_bytes),
        "workbooks": workbook_records,
        "sheet_inventory": sheet_inventory_records,
        "output_files": output_files,
        "row_counts": {
            "scenario_input_cells": int(len(cells_frame)),
            "scenario_input_long": int(len(long_frame)),
            "scenario_input_wide": int(len(wide_frame)),
            "scenario_feature_lineage": int(len(lineage_frame)),
        },
    }
    manifest_path = output / SCENARIO_INPUT_MANIFEST
    manifest_path.write_text(json.dumps(manifest, indent=2, allow_nan=False) + "\n", encoding="utf-8")
    return manifest


def combine_scenario_input_dirs(
    input_dirs: Iterable[Path | str],
    output_dir: Path | str,
    *,
    created_by: str = "scenario_input_combiner",
    repo_root: Path | str | None = None,
) -> dict[str, Any]:
    output = Path(output_dir)
    root = Path(repo_root).resolve() if repo_root is not None else Path.cwd().resolve()
    fallback_base = output.parent
    output.mkdir(parents=True, exist_ok=True)
    raw_dir = output / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    manifests: list[dict[str, Any]] = []
    frame_map: dict[str, list[pd.DataFrame]] = {
        SCENARIO_INPUT_CELLS_STEM: [],
        SCENARIO_INPUT_LONG_STEM: [],
        SCENARIO_INPUT_WIDE_STEM: [],
        SCENARIO_FEATURE_LINEAGE_STEM: [],
    }
    copied_raw_paths: dict[str, Path] = {}
    for source in input_dirs:
        source_dir = Path(source)
        manifest_path = source_dir / SCENARIO_INPUT_MANIFEST
        if manifest_path.exists():
            manifests.append(json.loads(manifest_path.read_text(encoding="utf-8")))
        for stem in frame_map:
            path = source_dir / f"{stem}.parquet"
            if path.exists():
                frame_map[stem].append(pd.read_parquet(path))
        source_raw = source_dir / "raw"
        if source_raw.exists():
            for workbook in source_raw.glob("*.xlsx"):
                target = raw_dir / workbook.name
                if not target.exists():
                    shutil.copy2(workbook, target)
                copied_raw_paths[workbook.name] = target

    output_files: dict[str, Any] = {}
    row_counts: dict[str, int] = {}
    combined_frames: dict[str, pd.DataFrame] = {}
    for stem, frames in frame_map.items():
        frame = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()
        if stem == SCENARIO_FEATURE_LINEAGE_STEM and frame.empty and frame_map[SCENARIO_INPUT_WIDE_STEM]:
            frame = scenario_feature_lineage_from_wide(pd.concat(frame_map[SCENARIO_INPUT_WIDE_STEM], ignore_index=True, sort=False))
        combined_frames[stem] = frame
        output_files.update(_write_artifact_pair(output, stem, frame, root=root, fallback_base=fallback_base))
        row_counts[stem] = int(len(frame))

    workbooks: list[dict[str, Any]] = []
    for manifest in manifests:
        for record in manifest.get("workbooks", []):
            if isinstance(record, dict):
                workbook_record = dict(record)
                raw_name = Path(str(workbook_record.get("raw_repo_relative_path") or "")).name
                if raw_name and raw_name in copied_raw_paths:
                    workbook_record["raw_repo_relative_path"] = _relative_path(
                        copied_raw_paths[raw_name],
                        root,
                        fallback_base=fallback_base,
                    )
                workbooks.append(workbook_record)

    sheet_inventory = _sheet_inventory_from_cells(combined_frames.get(SCENARIO_INPUT_CELLS_STEM, pd.DataFrame()))

    manifest = {
        "schema_version": SCENARIO_INPUT_SCHEMA_VERSION,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": created_by,
        "repo_relative_output_dir": _relative_path(output, root, fallback_base=fallback_base),
        "source_policy": "combined committed scenario input artifacts; Streamlit must not load Excel at runtime",
        "source_manifests": [
            {
                "repo_relative_output_dir": item.get("repo_relative_output_dir", ""),
                "created_at": item.get("created_at", ""),
                "schema_version": item.get("schema_version", ""),
            }
            for item in manifests
        ],
        "workbooks": workbooks,
        "sheet_inventory": sheet_inventory,
        "output_files": output_files,
        "row_counts": row_counts,
    }
    (output / SCENARIO_INPUT_MANIFEST).write_text(json.dumps(manifest, indent=2, allow_nan=False) + "\n", encoding="utf-8")
    return manifest


def scenario_feature_lineage_from_wide(wide: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "stream",
        "model_id",
        "feature_name",
        "scenario_name",
        "period",
        "value",
        "source_artifact",
        "canonical_variable",
        "fallback_flag",
        "fallback_reason",
        "source_status",
        "lineage_role",
        "feature_source_variables",
        "feature_engineering_basis",
        "feature_lineage_status",
    ]
    if wide is None or wide.empty:
        return pd.DataFrame(columns=columns)
    metadata = {
        "scenario_name",
        "role",
        "workbook_filename",
        "workbook_sha256",
        "stream",
        "sheet",
        "period",
        "canonical_period",
        "horizon",
        "source_artifact",
    }
    rows: list[dict[str, Any]] = []
    for record in wide.to_dict(orient="records"):
        stream = str(record.get("stream") or "")
        scenario = str(record.get("scenario_name") or "")
        period = str(record.get("canonical_period") or record.get("period") or "")
        for feature_name, value in record.items():
            if feature_name in metadata or pd.isna(value):
                continue
            canonical = _canonical_variable(feature_name)
            rows.append(
                {
                    "stream": stream,
                    "model_id": _model_id_for_stream(stream),
                    "feature_name": feature_name,
                    "scenario_name": scenario,
                    "period": period,
                    "value": value,
                    "source_artifact": f"{SCENARIO_INPUT_DIRNAME}/{SCENARIO_INPUT_LONG_STEM}.parquet",
                    "canonical_variable": canonical,
                    "fallback_flag": False,
                    "fallback_reason": "",
                    "source_status": "committed_scenario_input",
                    "lineage_role": "source_variable",
                    "feature_source_variables": canonical,
                    "feature_engineering_basis": "workbook_cell_materialized_to_scenario_input_wide",
                    "feature_lineage_status": "committed_source_variable",
                }
            )
        rows.extend(_model_feature_lineage_rows(record, stream=stream, scenario=scenario, period=period))
    return pd.DataFrame(rows, columns=columns)


def _model_feature_lineage_rows(record: dict[str, Any], *, stream: str, scenario: str, period: str) -> list[dict[str, Any]]:
    model_id = _model_id_for_stream(stream)
    if not model_id:
        return []
    rows: list[dict[str, Any]] = []
    for feature_name, source_variables in _model_feature_source_map(stream).items():
        source_variables = tuple(source_variables)
        missing = [variable for variable in source_variables if variable not in record and variable != "canonical_period"]
        target_recursive = feature_name.startswith("target__")
        time_feature = feature_name.startswith("time__")
        history_context = _feature_requires_history_context(feature_name)
        fallback_flag = bool(target_recursive or missing)
        fallback_reason = ""
        if target_recursive:
            fallback_reason = (
                "Governed recursive target-lag feature generated from repo model_input_history and prior fixed-finalist "
                "forecasts; it is not a direct workbook cell."
            )
        elif missing:
            fallback_reason = "Missing committed scenario source variables: " + "; ".join(missing)
        source_status = "committed_scenario_input"
        source_artifact = f"{SCENARIO_INPUT_DIRNAME}/{SCENARIO_INPUT_LONG_STEM}.parquet"
        if time_feature:
            source_status = "derived_from_committed_scenario_period"
            source_artifact = f"{SCENARIO_INPUT_DIRNAME}/{SCENARIO_INPUT_WIDE_STEM}.parquet"
        elif history_context:
            source_status = "derived_from_committed_scenario_input_and_repo_history"
            source_artifact = (
                f"{SCENARIO_INPUT_DIRNAME}/{SCENARIO_INPUT_LONG_STEM}.parquet; "
                f"data/model_input_history/{_model_input_history_file(stream)}"
            )
        if target_recursive:
            source_status = "governed_recursive_target_feature"
            source_artifact = f"data/model_input_history/{_model_input_history_file(stream)}"
        value = _model_feature_source_value(record, source_variables)
        rows.append(
            {
                "stream": stream,
                "model_id": model_id,
                "feature_name": feature_name,
                "scenario_name": scenario,
                "period": period,
                "value": value,
                "source_artifact": source_artifact,
                "canonical_variable": "; ".join(source_variables),
                "fallback_flag": fallback_flag,
                "fallback_reason": fallback_reason,
                "source_status": source_status,
                "lineage_role": "model_feature",
                "feature_source_variables": "; ".join(source_variables),
                "feature_engineering_basis": _feature_engineering_basis(feature_name, stream),
                "feature_lineage_status": "governed_fallback" if fallback_flag else "mapped_to_committed_source",
            }
        )
    return rows


def _model_feature_source_map(stream: str) -> dict[str, tuple[str, ...]]:
    if stream == "LIGHT_RUC":
        return dict(LIGHT_RUC_MODEL_FEATURE_SOURCES)
    if stream in {"PED", "HEAVY_RUC"}:
        return _vnext_feature_source_map(stream)
    return {}


def _vnext_feature_source_map(stream: str) -> dict[str, tuple[str, ...]]:
    features: dict[str, tuple[str, ...]] = {}
    for short, variable in VNEXT_LEVEL_SOURCES.get(stream, {}).items():
        features[f"{short}__level"] = (variable,)
    for short, variable in VNEXT_LOG_SOURCES.get(stream, {}).items():
        features[f"{short}__log"] = (variable,)
        for lag in (1, 2, 4):
            features[f"{short}__log_lag{lag}"] = (variable,)
        for diff in (1, 4):
            features[f"{short}__log_diff{diff}"] = (variable,)
    for feature_name in VNEXT_TIME_FEATURES:
        features[feature_name] = ("canonical_period",)
    for short in VNEXT_POLICY_PRICE_BASES.get(stream, ()):
        variable = VNEXT_LOG_SOURCES.get(stream, {}).get(short, "")
        if not variable:
            continue
        for feature_name in (
            f"policy__{short}_log_change_1",
            f"policy__{short}_log_change_4",
            f"policy__{short}_jump_up_1",
            f"policy__{short}_cut_1",
            f"policy__{short}_abs_change_1",
        ):
            features[feature_name] = (variable,)
        for lag in (1, 2, 4):
            for source in (
                f"policy__{short}_jump_up_1",
                f"policy__{short}_cut_1",
                f"policy__{short}_abs_change_1",
            ):
                features[f"{source}_lag{lag}"] = (variable,)
    for feature_name in VNEXT_TARGET_LAG_FEATURES:
        features[feature_name] = ("target_lag_1", "target_lag_4")
    return features


def _feature_requires_history_context(feature_name: str) -> bool:
    return any(token in feature_name for token in ("_lag", "__log_lag", "_diff", "__log_diff", "_roll", "policy__"))


def _model_input_history_file(stream: str) -> str:
    if stream == "LIGHT_RUC":
        return "light_ruc_inputs.parquet"
    if stream == "HEAVY_RUC":
        return "heavy_ruc_inputs.parquet"
    if stream == "PED":
        return "ped_inputs.parquet"
    return ""


def _model_feature_source_value(record: dict[str, Any], source_variables: tuple[str, ...]) -> Any:
    values: list[str] = []
    for variable in source_variables:
        if variable == "canonical_period":
            value = record.get("canonical_period") or record.get("period")
        else:
            value = record.get(variable, "")
        if pd.notna(value) and str(value) != "":
            values.append(f"{variable}={value}")
    return "; ".join(values)


def _feature_engineering_basis(feature_name: str, stream: str) -> str:
    if stream == "LIGHT_RUC":
        if _feature_requires_history_context(feature_name):
            return "forecast_runner_light_ruc_feature_frame_with_repo_history_context"
        return "forecast_runner_light_ruc_direct_feature"
    if feature_name.startswith("time__"):
        return "vnext_engineer_features_period_deterministic"
    if feature_name.startswith("target__"):
        return "vnext_recursive_target_lag_policy"
    if feature_name.startswith("policy__") or "__log_lag" in feature_name or "__log_diff" in feature_name:
        return "vnext_engineer_features_with_repo_history_context"
    return "vnext_engineer_features_direct_source"


def _sheet_inventory_from_cells(cells: pd.DataFrame) -> list[dict[str, Any]]:
    if cells is None or cells.empty:
        return []
    required = {"scenario_name", "role", "workbook_filename", "workbook_sha256", "sheet", "stream", "cell", "row_number", "column_number"}
    if required.difference(cells.columns):
        return []
    hash_columns = [
        "cell",
        "row_number",
        "column_number",
        "row_label",
        "column_label",
        "period",
        "canonical_period",
        "variable_name",
        "canonical_variable",
        "value",
        "value_type",
        "source_status",
    ]
    records: list[dict[str, Any]] = []
    source = cells.copy()
    source["row_number"] = pd.to_numeric(source["row_number"], errors="coerce")
    source["column_number"] = pd.to_numeric(source["column_number"], errors="coerce")
    for keys, group in source.groupby(
        ["scenario_name", "role", "workbook_filename", "workbook_sha256", "sheet"],
        dropna=False,
        sort=True,
    ):
        scenario_name, role, workbook_filename, workbook_sha256, sheet = (str(value or "") for value in keys)
        ordered = group.sort_values(["row_number", "column_number"], kind="stable").copy()
        hash_frame = ordered.reindex(columns=hash_columns).copy()
        if "value" in hash_frame.columns:
            hash_frame["value"] = hash_frame["value"].map(_stringify_excel_value)
        for column in hash_frame.columns:
            hash_frame[column] = hash_frame[column].where(hash_frame[column].notna(), "").astype(str)
        sheet_payload = hash_frame.to_csv(index=False, lineterminator="\n").encode("utf-8")
        row_numbers = pd.to_numeric(ordered["row_number"], errors="coerce").dropna()
        column_numbers = pd.to_numeric(ordered["column_number"], errors="coerce").dropna()
        records.append(
            {
                "scenario_name": scenario_name,
                "role": role,
                "workbook_filename": workbook_filename,
                "workbook_sha256": workbook_sha256,
                "sheet": sheet,
                "stream": _first_text(ordered.get("stream", pd.Series(dtype=object))),
                "non_empty_cell_count": int(len(ordered)),
                "materialized_cell_count": int(len(ordered)),
                "non_empty_row_count": int(row_numbers.nunique()),
                "non_empty_column_count": int(column_numbers.nunique()),
                "first_cell": str(ordered["cell"].iloc[0]) if not ordered.empty else "",
                "last_cell": str(ordered["cell"].iloc[-1]) if not ordered.empty else "",
                "materialized_cells_sha256": hashlib.sha256(sheet_payload).hexdigest(),
                "source_status": "all_non_empty_cells_materialized",
            }
        )
    return records


def _first_text(values: Any) -> str:
    if values is None:
        return ""
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _workbook_bytes(workbook: Path | str | bytes | BinaryIO) -> bytes:
    if isinstance(workbook, bytes):
        return workbook
    if hasattr(workbook, "read"):
        content = workbook.read()
        if hasattr(workbook, "seek"):
            workbook.seek(0)
        return content
    return Path(workbook).read_bytes()


def _workbook_cells_frame(
    wb: Any,
    *,
    scenario_name: str,
    scenario_role: str,
    workbook_filename: str,
    workbook_sha256: str,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        stream = TEMPLATE_SHEET_STREAMS.get(sheet_name, "")
        header_by_col = {
            cell.column: str(cell.value).strip()
            for cell in ws[1]
            if cell.value is not None and str(cell.value).strip()
        }
        period_by_row = {}
        period_col = next((col for col, name in header_by_col.items() if _canonical_variable(name) == "period"), None)
        if period_col:
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=period_col).value
                if value is not None and str(value).strip():
                    period_by_row[row_idx] = str(value).strip().upper()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                variable = header_by_col.get(cell.column, "")
                canonical = _canonical_variable(variable) if variable else ""
                period = period_by_row.get(cell.row, "")
                rows.append(
                    {
                        "scenario_name": scenario_name,
                        "role": scenario_role,
                        "workbook_filename": workbook_filename,
                        "workbook_sha256": workbook_sha256,
                        "sheet": sheet_name,
                        "stream": stream,
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "row_number": cell.row,
                        "column_number": cell.column,
                        "row_label": period or str(cell.row),
                        "column_label": variable or get_column_letter(cell.column),
                        "period": period,
                        "canonical_period": _canonical_period(period),
                        "variable_name": variable,
                        "canonical_variable": canonical,
                        "value": cell.value,
                        "unit": _unit_for_variable(canonical),
                        "value_type": _value_type(cell.value),
                        "source_status": "materialized_from_workbook_cell",
                    }
                )
    return pd.DataFrame(rows, columns=_empty_cells_frame().columns)


def _scenario_input_long_from_cells(cells: pd.DataFrame) -> pd.DataFrame:
    if cells.empty:
        return _empty_long_frame()
    source = cells[cells["stream"].astype(str).ne("") & cells["variable_name"].astype(str).ne("")].copy()
    source = source[source["row_number"].gt(1)].copy()
    source["source_artifact"] = f"{SCENARIO_INPUT_DIRNAME}/{SCENARIO_INPUT_CELLS_STEM}.parquet"
    source["range"] = source["cell"]
    source["value"] = source["value"].map(_stringify_excel_value)
    cols = list(_empty_long_frame().columns)
    return source.reindex(columns=cols).reset_index(drop=True)


def _scenario_input_wide_from_long(long: pd.DataFrame) -> pd.DataFrame:
    if long.empty:
        return _empty_wide_frame()
    source = long.copy()
    index_cols = [
        "scenario_name",
        "role",
        "workbook_filename",
        "workbook_sha256",
        "stream",
        "sheet",
        "period",
        "canonical_period",
    ]
    rows = source[index_cols + ["canonical_variable", "value"]].copy()
    rows = rows[rows["canonical_variable"].astype(str).ne("")]
    rows = rows[~rows["canonical_variable"].astype(str).isin(set(index_cols))]
    rows = rows.drop_duplicates(index_cols + ["canonical_variable"], keep="first")
    wide = rows.pivot(index=index_cols, columns="canonical_variable", values="value").reset_index()
    wide.columns = [str(col) for col in wide.columns]
    if "horizon" in wide.columns:
        wide = wide.sort_values(["scenario_name", "stream", "horizon"], kind="stable")
    else:
        wide = wide.sort_values(["scenario_name", "stream", "canonical_period"], kind="stable")
    wide["source_artifact"] = f"{SCENARIO_INPUT_DIRNAME}/{SCENARIO_INPUT_LONG_STEM}.parquet"
    return wide.reset_index(drop=True)


def _write_artifact_pair(
    output_dir: Path,
    stem: str,
    frame: pd.DataFrame,
    *,
    root: Path,
    fallback_base: Path,
) -> dict[str, Any]:
    csv_path = output_dir / f"{stem}.csv"
    parquet_path = output_dir / f"{stem}.parquet"
    prepared = frame.copy()
    for column in prepared.columns:
        if prepared[column].dtype == object:
            prepared[column] = prepared[column].where(prepared[column].notna(), "").astype(str)
    prepared.to_csv(csv_path, index=False)
    prepared.to_parquet(parquet_path, index=False)
    return {
        f"{stem}.csv": _file_record(csv_path, root=root, fallback_base=fallback_base),
        f"{stem}.parquet": _file_record(parquet_path, root=root, fallback_base=fallback_base),
    }


def _file_record(path: Path, *, root: Path, fallback_base: Path) -> dict[str, Any]:
    return {
        "repo_relative_path": _relative_path(path, root, fallback_base=fallback_base),
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "size_bytes": path.stat().st_size,
    }


def _empty_cells_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "scenario_name",
            "role",
            "workbook_filename",
            "workbook_sha256",
            "sheet",
            "stream",
            "cell",
            "row_number",
            "column_number",
            "row_label",
            "column_label",
            "period",
            "canonical_period",
            "variable_name",
            "canonical_variable",
            "value",
            "unit",
            "value_type",
            "source_status",
        ]
    )


def _empty_long_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "scenario_name",
            "role",
            "workbook_filename",
            "workbook_sha256",
            "sheet",
            "stream",
            "cell",
            "range",
            "row_label",
            "column_label",
            "period",
            "canonical_period",
            "variable_name",
            "canonical_variable",
            "value",
            "unit",
            "value_type",
            "source_status",
            "source_artifact",
        ]
    )


def _empty_wide_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "scenario_name",
            "role",
            "workbook_filename",
            "workbook_sha256",
            "stream",
            "sheet",
            "period",
            "canonical_period",
            "source_artifact",
        ]
    )


def _canonical_variable(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _canonical_period(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text if re.fullmatch(r"\d{4}Q[1-4]", text) else ""


def _safe_filename(value: str) -> str:
    name = Path(str(value or "workbook.xlsx")).name
    safe = re.sub(r"[^A-Za-z0-9_. -]+", "_", name).strip(" .")
    return safe or "workbook.xlsx"


def _value_type(value: Any) -> str:
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "numeric"
    if hasattr(value, "isoformat"):
        return "date_or_datetime"
    return "text"


def _stringify_excel_value(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _unit_for_variable(canonical_variable: str) -> str:
    if canonical_variable in UNIT_HINTS:
        return UNIT_HINTS[canonical_variable]
    if canonical_variable in SYSTEM_VARIABLES:
        return "system"
    if canonical_variable.startswith("log_"):
        return "log transformed"
    if "interaction" in canonical_variable:
        return "derived interaction"
    if canonical_variable.endswith("_dummy"):
        return "0/1"
    return ""


def _model_id_for_stream(stream: str) -> str:
    if stream == "PED":
        return "PED__VNEXT_SOLVED_CONVEX_TOP2"
    if stream == "LIGHT_RUC":
        return "dynamic_RESID_GBR_n150_d1_lr0.05_w36"
    if stream == "HEAVY_RUC":
        return "HEAVY_RUC__VNEXT_SOLVED_CONVEX_TOP4"
    return ""


def _relative_path(path: Path, root: Path, *, fallback_base: Path | None = None) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(root.resolve()).as_posix()
    except ValueError:
        if fallback_base is not None:
            try:
                return resolved.relative_to(fallback_base.resolve()).as_posix()
            except ValueError:
                pass
        return path.name
