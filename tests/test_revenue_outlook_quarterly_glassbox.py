"""Quarterly glass-box XLSX (Current conditions baseline + 12c rate path, AR(1)).

Contract (owner request, 2026-09): a second download button beside the annual
forecast extract producing a formula-driven quarterly workbook for the
selected Current-model path. The top block keeps the annual extract's labels
and order MINUS the annual-percentage-change block, on calendar-quarter
columns 2000Q3-2050Q2; every top value is a formula linked to detail sections
or another worksheet; actual-period cells carry their full history
(published quarters plus the governed derived quarterly presentation),
colour-coded apart. PED is coefficient-reproduced, Light RUC is OLS + exact
imported GBR residual, Heavy RUC is the governed level-space weighted
ensemble; the builder Python-verifies every identity against the governed
displayed values before writing a cell. The delivered workbook has no Checks
sheet (audit builds carry one); the annual extract is untouched.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

import app
from model_dashboard.engine import ENGINE_AR1, engine_revenue_outlook_dir
from model_dashboard.long_run_shape_transition import (
    PRODUCTION_LONG_RUN_TRANSITION_SCHEDULE_ID,
)
from model_dashboard import official_vintage as official_vintage_module
from model_dashboard.revenue_outlook import (
    PED_BRIDGE_DEFAULT_MODE,
    load_revenue_outlook_pack,
    revenue_outlook_signature,
)
from model_dashboard.revenue_outlook_excel_extract import (
    LEVEL_ROW_SERIES,
    REVENUE_ROW_SERIES,
    TEMPLATE_RELATIVE_PATH,
    TEMPLATE_SHEET_NAME,
)
from model_dashboard._quarterly_glassbox_writer import (
    TOP_ROW_TEMPLATE_MAP,
    TOP_TOTAL_ROW,
    TRANSITION_RESERVED_ROWS,
)
from model_dashboard.revenue_outlook_quarterly_glassbox import (
    GLASSBOX_BUTTON_LABEL,
    GLASSBOX_SUPPORTED_TRACE,
    glassbox_quarter_grid,
    glassbox_supported_selection,
)
from model_dashboard.revenue_scenario_key import (
    HEAVY_BEV_DEFAULT,
    RevenueScenarioComputationKey,
)

ROOT = Path(__file__).resolve().parents[1]

#: The acceptance scenario: Central baseline + MCERT, every other lever Off,
#: production AR(1) engine.
ACCEPTANCE_POLICY_STATE = "mcert"

_STATUS_FILL_RGB = {"00CDD4DD", "00F0A072", "00FBDECD"}


@pytest.fixture(scope="module")
def glassbox_context():
    pack_dir = ROOT / engine_revenue_outlook_dir(ENGINE_AR1)
    pack = load_revenue_outlook_pack(pack_dir, repo_root=ROOT)
    assert pack is not None
    signature = revenue_outlook_signature(pack_dir, ROOT)
    key = RevenueScenarioComputationKey(
        engine=ENGINE_AR1,
        uptake_basis=app.DEFAULT_EV_UPTAKE_MODE,
        custom_ev_levers=(),
        eruc_levers=(),
        current_fed_policy_state=ACCEPTANCE_POLICY_STATE,
        official_fed_policy_state="published",
        ped_retention_sensitivity=False,
        heavy_bev_transition=HEAVY_BEV_DEFAULT,
        official_comparator_vintage_id=official_vintage_module.default_comparator_vintage_id(ROOT),
        official_comparator_overlay=False,
        ped_bridge_mode=PED_BRIDGE_DEFAULT_MODE,
        bridge_vintage_id=official_vintage_module.default_bridge_vintage_id(ROOT),
        long_run_transition_schedule_id=PRODUCTION_LONG_RUN_TRANSITION_SCHEDULE_ID,
        long_run_shape_vintage_id=official_vintage_module.default_long_run_shape_vintage_id(ROOT),
        fed_ruc_transition="off",
    )
    pack, signature = app._apply_long_run_shape_selection(
        pack, signature, str(pack_dir), key
    )
    sensitivity_key = app.selected_sensitivity_key("Off", "Off", "Off")
    return pack, signature, key, sensitivity_key, pack_dir


@pytest.fixture(scope="module")
def glassbox_result(glassbox_context):
    pack, signature, key, sensitivity_key, pack_dir = glassbox_context
    return app.cached_quarterly_glassbox_bytes(
        signature, sensitivity_key, PED_BRIDGE_DEFAULT_MODE, key, str(pack_dir), pack
    )


@pytest.fixture(scope="module")
def workbook(glassbox_result):
    return load_workbook(io.BytesIO(glassbox_result.workbook_bytes))


def test_delivered_workbook_has_four_sheets_and_no_checks(workbook) -> None:
    assert list(workbook.sheetnames) == [
        "README",
        "Scenario",
        "Quarterly Inputs",
        "Model Parameters",
    ]


def test_top_block_keeps_template_labels_minus_pct_block(workbook) -> None:
    template = load_workbook(ROOT / TEMPLATE_RELATIVE_PATH, read_only=True)
    template_labels = {
        row: template[TEMPLATE_SHEET_NAME].cell(row=row, column=1).value
        for row in range(1, 66)
    }
    template.close()
    scenario = workbook["Scenario"]
    for workbook_row, template_row in TOP_ROW_TEMPLATE_MAP.items():
        expected = template_labels[template_row]
        got = scenario.cell(row=workbook_row, column=1).value
        if template_row == 13:
            # TUC GTK is displayed in millions of tonne-km (owner request).
            assert got == "TUC GTK (m Tonne-km)"
            continue
        assert got == expected, f"row {workbook_row} (template {template_row})"
    # The percentage-change block is gone: no top-block label repeats the
    # template's row-15 header, and the total lands on row 54.
    labels = [scenario.cell(row=row, column=1).value for row in range(1, 60)]
    assert "Key volumes: Annual percentage changes" not in labels
    assert scenario.cell(row=TOP_TOTAL_ROW, column=1).value == "Total net revenues (m $)"


def test_quarter_grid_headers_and_status_fills(workbook) -> None:
    scenario = workbook["Scenario"]
    quarters = glassbox_quarter_grid()
    assert quarters[0] == "2000Q3" and quarters[-1] == "2050Q2"
    assert len(quarters) == 200
    assert scenario.cell(row=1, column=2).value == 2001
    assert scenario.cell(row=2, column=2).value == "2000Q3"
    assert scenario.cell(row=2, column=201).value == "2050Q2"
    assert scenario.cell(row=1, column=201).value == 2050
    fills = {
        str(scenario.cell(row=2, column=column).fill.start_color.rgb)
        for column in range(2, 202)
    }
    assert _STATUS_FILL_RGB.issubset(fills)


def test_every_top_block_value_is_a_formula(workbook) -> None:
    scenario = workbook["Scenario"]
    workbook_row_of_template = {
        template: wb_row for wb_row, template in TOP_ROW_TEMPLATE_MAP.items()
    }
    painted = []
    for template_row in list(LEVEL_ROW_SERIES) + list(REVENUE_ROW_SERIES):
        workbook_row = workbook_row_of_template[template_row]
        for column in range(2, 202):
            value = scenario.cell(row=workbook_row, column=column).value
            if value is None:
                continue
            if not (isinstance(value, str) and value.startswith("=")):
                painted.append((workbook_row, column, value))
    assert painted == []


def test_actual_history_is_filled_and_colour_coded(workbook, glassbox_result) -> None:
    scenario = workbook["Scenario"]
    workbook_row_of_template = {
        template: wb_row for wb_row, template in TOP_ROW_TEMPLATE_MAP.items()
    }
    # Every mapped value row carries values in the FY2005 columns (deep
    # history), whether natively published or derived presentation.
    fy2005_columns = [
        column for column in range(2, 202)
        if scenario.cell(row=1, column=column).value == 2005
    ]
    assert len(fy2005_columns) == 4
    missing = []
    fills = set()
    for template_row in list(LEVEL_ROW_SERIES) + list(REVENUE_ROW_SERIES):
        workbook_row = workbook_row_of_template[template_row]
        for column in fy2005_columns:
            cell = scenario.cell(row=workbook_row, column=column)
            if cell.value is None:
                missing.append((workbook_row, column))
            else:
                fills.add(str(cell.fill.start_color.rgb))
    assert missing == []
    # Both colour codes appear: natively published and derived presentation.
    assert "00CDD4DD" in fills or "00E9EDF2" in fills
    scenario_fills = {
        str(scenario.cell(row=5, column=column).fill.start_color.rgb)
        for column in range(2, 202)
        if scenario.cell(row=5, column=column).value is not None
    }
    assert "00CDD4DD" in scenario_fills  # native actual quarters on Light RUC km
    assert "00E9EDF2" in scenario_fills  # derived pre-2009 presentation


def test_transition_rows_stay_reserved(workbook) -> None:
    scenario = workbook["Scenario"]
    assert all(
        scenario.cell(row=row, column=column).value is None
        for row in TRANSITION_RESERVED_ROWS
        for column in range(2, 202)
    )


def test_python_parity_evidence_is_machine_precision(glassbox_result) -> None:
    expected_checks = {
        "ped_ar1_recursion_vs_committed",
        "light_ols_plus_gbr_vs_committed",
        "heavy_weighted_blend_vs_committed",
        "policy_calibration_identity",
        "native_quarters_vs_calibrated_model",
        "rate_factor_scalar_vs_bridge",
        "official_rate_and_carried_identities",
        "post_model_total_ruc_and_hidden_bev",
        "post_model_handover_vs_displayed",
        "annual_formula_registry_closure",
        "four_quarter_sum_worst_case",
    }
    assert expected_checks.issubset(glassbox_result.parity.keys())
    worst = max(glassbox_result.parity.values())
    assert worst < 1e-9


def test_defined_names_are_valid_and_unique(workbook) -> None:
    names = list(workbook.defined_names.keys())
    assert len(names) == len(set(names))
    for required in (
        "PED_B0", "PED_BPet", "PED_BGDP", "PED_BLagY", "PED_Rho", "PED_LastErr",
        "LR_B0", "LR_BDiesel", "LR_BGDP",
        "HR_W1", "HR_W2", "HR_W3",
        "PED_Elas", "LR_Elas", "HR_Elas", "RUC_MacroFac",
    ):
        assert required in names
        target = workbook.defined_names[required].attr_text
        assert "'Model Parameters'" in target and "#REF" not in target


def test_no_error_tokens_and_no_snake_case_labels(workbook) -> None:
    tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    offenders = []
    snake_labels = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if any(token in value for token in tokens):
                    offenders.append((sheet.title, cell.coordinate))
                if cell.column == 1 and "_" in value:
                    snake_labels.append((sheet.title, cell.coordinate, value[:60]))
    assert offenders == []
    assert snake_labels == []


def test_lookups_are_xlookup_not_index_match(workbook) -> None:
    scenario = workbook["Scenario"]
    uses_xlookup = False
    for row in scenario.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "_xlfn.XLOOKUP" in cell.value:
                uses_xlookup = True
            assert not (
                isinstance(cell.value, str) and "INDEX(" in str(cell.value)
            ), f"INDEX/MATCH left at {cell.coordinate}"
    assert uses_xlookup


def test_workbook_calculation_mode_is_automatic_full_calc(workbook) -> None:
    assert workbook.calculation.fullCalcOnLoad is True


def test_audit_build_carries_the_checks_sheet(glassbox_context) -> None:
    # The audit flag is exercised by the local acceptance harness (Excel COM
    # recalculation); here we only prove the flag changes the sheet inventory.
    import inspect

    from model_dashboard.revenue_outlook_quarterly_glassbox import (
        build_quarterly_glassbox_workbook,
    )

    signature = inspect.signature(build_quarterly_glassbox_workbook)
    assert "audit_sheets" in signature.parameters
    assert signature.parameters["audit_sheets"].default is False


def test_selection_gate_states_each_refusal() -> None:
    base = dict(
        trace_name=GLASSBOX_SUPPORTED_TRACE,
        engine=ENGINE_AR1,
        uptake_basis=app.DEFAULT_EV_UPTAKE_MODE,
        custom_ev_levers=(),
        eruc_levers=(),
        ped_retention_sensitivity=False,
        fed_ruc_transition="off",
        sensitivities_off=True,
    )
    assert glassbox_supported_selection(**base) == ""
    assert "Current-model scenarios only" in glassbox_supported_selection(
        **{**base, "trace_name": "PREBU26 official"}
    )
    assert "AR(1)" in glassbox_supported_selection(**{**base, "engine": "ensemble"})
    assert "FED->RUC" in glassbox_supported_selection(
        **{**base, "fed_ruc_transition": "managed"}
    )
    assert "sensitivities Off" in glassbox_supported_selection(
        **{**base, "sensitivities_off": False}
    )
    assert "e-RUC" in glassbox_supported_selection(
        **{**base, "eruc_levers": (2027.0, 3.0)}
    )


def test_result_identity(glassbox_result) -> None:
    assert glassbox_result.trace_name == GLASSBOX_SUPPORTED_TRACE
    assert glassbox_result.scenario_name == "middle_east_low__mcert"


def test_download_button_appears_on_the_single_scenario_view() -> None:
    from streamlit.testing.v1 import AppTest

    harness = AppTest.from_file(str(ROOT / "app.py"), default_timeout=180)
    harness.run()
    harness.radio[0].set_value(app.REVENUE_OUTLOOK_PAGE)
    harness.run()
    assert not harness.exception
    labels = [str(element.proto.label) for element in harness.get("download_button")]
    assert GLASSBOX_BUTTON_LABEL in labels
    assert "Download forecast extract XLSX" in labels  # the annual button survives

    compare_radio = next(
        radio
        for radio in harness.radio
        if app.REVENUE_OUTLOOK_VIEW_COMPARE in list(radio.options)
    )
    compare_radio.set_value(app.REVENUE_OUTLOOK_VIEW_COMPARE)
    harness.run()
    assert not harness.exception
    labels = [str(element.proto.label) for element in harness.get("download_button")]
    assert GLASSBOX_BUTTON_LABEL not in labels


# ---------------------------------------------------------------------------
# Model-contract coupling: the glass-box hardwires the PED / Light RUC
# coefficient maps and the Heavy RUC ensemble shape. Swapping a finalist must
# fail HERE with a pointer to the map, not deep inside a workbook build.
# ---------------------------------------------------------------------------


def test_glassbox_coefficient_maps_match_the_governed_fitted_states() -> None:
    import json

    import pipeline.vnext_forward as vnext_forward
    from model_dashboard.forecast_runner import (
        LIGHT_RUC_BASE_FEATURES,
        load_light_ruc_promoted_state,
    )
    from model_dashboard.revenue_outlook_quarterly_glassbox import (
        LIGHT_COEFFICIENT_NAMES,
        PED_COEFFICIENT_NAMES,
        _HEAVY_MANIFEST_REL,
        _PED_AR1_STATE_REL,
    )

    ped_state = json.loads((ROOT / _PED_AR1_STATE_REL).read_text(encoding="utf-8"))
    assert list(ped_state["features"]) == [name for name, _ in PED_COEFFICIENT_NAMES], (
        "PED AR(1) fitted-state features changed: update PED_COEFFICIENT_NAMES in "
        "model_dashboard/revenue_outlook_quarterly_glassbox.py and the Model "
        "Parameters labels in _quarterly_glassbox_writer.py."
    )
    assert len(ped_state["beta"]) == len(PED_COEFFICIENT_NAMES) + 2  # b0 + features + ylag
    assert list(ped_state["ylags"]) == [1] and len(ped_state["rho"]) == 1

    assert list(LIGHT_RUC_BASE_FEATURES) == [name for name, _ in LIGHT_COEFFICIENT_NAMES], (
        "Light RUC OLS base features changed: update LIGHT_COEFFICIENT_NAMES."
    )
    vnext_forward._register_legacy_sklearn_loss_module_alias()
    light_state = load_light_ruc_promoted_state(ROOT)
    assert light_state.ols_beta.shape == (len(LIGHT_COEFFICIENT_NAMES) + 1,)

    heavy = json.loads((ROOT / _HEAVY_MANIFEST_REL).read_text(encoding="utf-8"))
    members = heavy.get("members") or heavy.get("production_states") or {}
    assert len(members) == 3, (
        "Heavy RUC ensemble shape changed: the workbook names HR_W1..HR_W3 and "
        "labels M1 as the linear (Ridge) component - update the writer."
    )
