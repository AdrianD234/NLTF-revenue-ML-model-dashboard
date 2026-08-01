from __future__ import annotations

import json
import hashlib
from pathlib import Path
import shutil

import numpy as np
import pandas as pd
from openpyxl import load_workbook
import pytest

import model_dashboard.revenue_outlook as revenue_outlook_module
from model_dashboard.forecast_runner import (
    SHEET_BY_STREAM,
    create_completed_sample_workbook,
    run_forecast_workbook,
    write_forecast_scenario_comparison,
)
from model_dashboard.light_fleet_allocation import (
    CONVENTIONAL_ANCHOR_SERIES_ID,
    composition_shares,
    LAST_DECISION_GRADE_ANNUAL_FY,
    LAST_DECISION_GRADE_QUARTER,
)
from model_dashboard.series_inventory_contract import (
    FIRST_POST_MODEL_FY,
    LAST_POST_MODEL_FY,
)
from model_dashboard.revenue_outlook import (
    CANONICAL_JOIN_KEY_COLUMNS,
    CURRENT_REVENUE_OUTLOOK_DIR,
    FAN_SOURCE_CURRENT_BACKTEST,
    FAN_SOURCE_MBU26_ARCHIVED,
    FAN_SOURCE_SCENARIO_SPREAD,
    FUTURE_RATE_COLUMNS,
    PED_BRIDGE_DEFAULT_MODE,
    PED_EFFICIENCY_BASELINE_SCENARIO_ID,
    REVENUE_OUTLOOK_SCHEMA_VERSION,
    SENSITIVITY_SEED_WORKBOOK_SHA256,
    SOURCE_COMPARISON_OUTPUT_DIR_POLICY,
    apply_ped_bridge_mode_layer,
    apply_revenue_sensitivity_layer,
    apply_ped_efficiency_sensitivity,
    build_revenue_outlook_pack,
    load_revenue_outlook_pack,
    ped_efficiency_adjustment_frame,
    revenue_sensitivity_impact_audit_frame,
    promote_revenue_outlook_pack,
    validate_promotable_comparison,
)


ROOT = Path(__file__).resolve().parents[1]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _min_fy(frame: pd.DataFrame) -> int | None:
    """Mirror of _max_fy; frames disagree on whether the column is FY or june_year."""
    if frame is None or frame.empty:
        return None
    for column in ("FY", "june_year"):
        if column in frame.columns:
            years = pd.to_numeric(frame[column], errors="coerce")
            if years.notna().any():
                return int(years.min())
    for column in ("target_period", "annual_period", "period"):
        if column in frame.columns:
            years = frame[column].astype(str).str.extract(r"FY(\d{4})", expand=False)
            years = pd.to_numeric(years, errors="coerce")
            if years.notna().any():
                return int(years.min())
    return None


def _max_fy(frame: pd.DataFrame) -> int | None:
    if frame is None or frame.empty:
        return None
    for column in ("FY", "june_year"):
        if column in frame.columns:
            years = pd.to_numeric(frame[column], errors="coerce")
            if years.notna().any():
                return int(years.max())
    for column in ("target_period", "annual_period", "period"):
        if column in frame.columns:
            years = frame[column].astype(str).str.extract(r"FY(\d{4})", expand=False)
            years = pd.to_numeric(years, errors="coerce")
            if years.notna().any():
                return int(years.max())
    return None


def _comparison(tmp_path: Path, *, blank_rates: bool = False, fixture: bool = False):
    base = create_completed_sample_workbook(tmp_path / "NLTF_forecast_input_template_basecase.xlsx", repo_root=ROOT, quarters=4)
    comparison = create_completed_sample_workbook(
        tmp_path / "NLTF_forecast_input_template_high_population.xlsx",
        repo_root=ROOT,
        quarters=4,
        value_multiplier=1.02,
    )
    if blank_rates:
        _blank_nominal_rate_columns(base)
        _blank_nominal_rate_columns(comparison)
    results = [
        run_forecast_workbook(
            base,
            output_dir=tmp_path / "basecase_run",
            repo_root=ROOT,
            run_timestamp="revenue-test",
            scenario_name="basecase",
            scenario_role="basecase",
            is_test_fixture=fixture,
            expected_quarters=4,
        ),
        run_forecast_workbook(
            comparison,
            output_dir=tmp_path / "comparison_run",
            repo_root=ROOT,
            run_timestamp="revenue-test",
            scenario_name="high_population",
            scenario_role="comparison",
            is_test_fixture=fixture,
            expected_quarters=4,
        ),
    ]
    return write_forecast_scenario_comparison(
        results,
        output_dir=tmp_path / "scenario_comparison",
        repo_root=ROOT,
        run_timestamp="revenue-test",
    )


def _blank_nominal_rate_columns(path: Path) -> None:
    workbook = load_workbook(path)
    for stream, (rate_col, source_col, cpi_col) in FUTURE_RATE_COLUMNS.items():
        sheet_name = SHEET_BY_STREAM[stream]
        ws = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        ws.protection.sheet = False
        for column_name in (rate_col, source_col, cpi_col):
            if column_name not in headers:
                continue
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=headers[column_name]).value = None
        ws.protection.sheet = True
    workbook.save(path)


def test_revenue_outlook_pack_computes_ruc_formula_and_honest_ped_gap(tmp_path: Path) -> None:
    comparison = _comparison(tmp_path)
    pack = promote_revenue_outlook_pack(comparison, repo_root=ROOT, output_dir=tmp_path / "pack", promoted_by="pytest")

    assert pack.manifest["schema_version"] == REVENUE_OUTLOOK_SCHEMA_VERSION
    assert pack.manifest["pack_status"] == "explicitly_promoted_current_outlook"
    assert pack.manifest["source_policy"].startswith("explicit_promoted_pack_or_in_session_reviewed_result_only")
    assert pack.manifest["source_comparison"]["output_dir_policy"] == SOURCE_COMPARISON_OUTPUT_DIR_POLICY
    assert "output_dir" not in pack.manifest["source_comparison"]
    assert all("output_dir" not in scenario for scenario in pack.manifest["source_comparison"]["scenarios"])
    assert pack.manifest["join_key_contract"]["columns"] == CANONICAL_JOIN_KEY_COLUMNS
    assert (pack.output_dir / "manifest.json").exists()
    assert (pack.output_dir / "future_revenue_forecasts.parquet").exists()
    assert load_revenue_outlook_pack(pack.output_dir, repo_root=ROOT) is not None
    scenario_input_manifest = json.loads(
        (pack.output_dir / "scenario_inputs" / "scenario_input_manifest.json").read_text(encoding="utf-8")
    )
    assert str(tmp_path) not in json.dumps(scenario_input_manifest)
    assert "C:\\Users" not in json.dumps(scenario_input_manifest)
    for workbook in scenario_input_manifest["workbooks"]:
        raw_path = Path(workbook["raw_repo_relative_path"])
        assert not raw_path.is_absolute()
        assert (ROOT / raw_path).exists() or (pack.output_dir / raw_path).exists()
        assert "scenario_inputs/raw" in raw_path.as_posix()
    for frame in [pack.future_revenue_forecasts, pack.revenue_bridge_components, pack.revenue_chart_rows]:
        assert set(CANONICAL_JOIN_KEY_COLUMNS).issubset(frame.columns)
        assert frame["canonical_join_key"].astype(str).str.count(r"\|").eq(2).all()
        assert not frame["canonical_join_key"].astype(str).str.contains(r"\|\||^\||\|$").any()

    light = pack.future_revenue_forecasts[
        pack.future_revenue_forecasts["stream"].eq("LIGHT_RUC")
        & pack.future_revenue_forecasts["bridge_status"].eq("available")
    ].copy()
    assert not light.empty
    row = light.iloc[0]
    expected = float(row["activity_forecast"]) / 1000.0 * float(row["rate_value"])
    assert abs(float(row["revenue_forecast_nzd"]) - expected) <= 1e-6
    assert bool(row["forecast_available"])

    ped = pack.future_revenue_forecasts[pack.future_revenue_forecasts["stream"].eq("PED")].copy()
    assert not ped.empty
    assert set(ped["bridge_status"].dropna().unique()) == {"ped_bridge_source_history_missing"}
    assert ped["revenue_forecast_nzd"].isna().all()
    assert not ped["revenue_forecast_nzd"].fillna("").astype(str).isin({"0", "0.0"}).any()

    reconciliations = pack.revenue_bridge_components[
        pack.revenue_bridge_components["component_type"].eq("historical_revenue_reconciliation")
    ].copy()
    assert set(reconciliations["stream"]) == {"LIGHT_RUC", "HEAVY_RUC"}
    deltas = pd.to_numeric(reconciliations["reconciliation_max_abs_delta_nzd"], errors="coerce")
    assert deltas.notna().all()
    assert float(deltas.max()) <= 1e-4


def test_revenue_outlook_missing_rates_are_gaps_not_zeroes(tmp_path: Path) -> None:
    comparison = _comparison(tmp_path, blank_rates=True)
    pack = build_revenue_outlook_pack(comparison, repo_root=ROOT, output_dir=tmp_path / "pack")
    ruc = pack.future_revenue_forecasts[pack.future_revenue_forecasts["stream"].isin(["LIGHT_RUC", "HEAVY_RUC"])].copy()
    assert not ruc.empty
    numeric_activity = pd.to_numeric(ruc["activity_forecast"], errors="coerce").notna()
    assert ruc.loc[numeric_activity, "bridge_status"].isin(["nominal_rate_missing", "activity_forecast_gap"]).all()
    assert ruc["revenue_forecast_nzd"].isna().all()
    assert not ruc["revenue_forecast_nzd"].fillna("").astype(str).isin({"0", "0.0"}).any()


def test_revenue_outlook_blocks_fixture_publication(tmp_path: Path) -> None:
    comparison = _comparison(tmp_path, fixture=True)
    errors = validate_promotable_comparison(comparison)
    assert any("Test fixture scenario cannot be promoted" in message for message in errors)


def test_committed_current_revenue_outlook_pack_is_repo_local_and_hash_backed() -> None:
    pack_dir = ROOT / CURRENT_REVENUE_OUTLOOK_DIR
    manifest_path = pack_dir / "manifest.json"
    assert manifest_path.exists()
    manifest_text = manifest_path.read_text(encoding="utf-8")
    assert "C:\\Users" not in manifest_text
    assert "Downloads" not in manifest_text
    assert "test-output" not in manifest_text
    assert "revenue_outlook_promotion" not in manifest_text
    manifest = json.loads(manifest_text)
    assert manifest["schema_version"] == REVENUE_OUTLOOK_SCHEMA_VERSION
    assert manifest["source_comparison"]["output_dir_policy"] == SOURCE_COMPARISON_OUTPUT_DIR_POLICY
    assert "output_dir" not in manifest["source_comparison"]
    assert all("output_dir" not in scenario for scenario in manifest["source_comparison"]["scenarios"])
    assert manifest["join_key_contract"]["columns"] == CANONICAL_JOIN_KEY_COLUMNS
    assert "canonical stream, period and scenario keys" in manifest["join_key_contract"]["rule"]
    assert manifest["repo_relative_output_dir"] == "data/current_revenue_outlook"
    assert manifest["source_hashes"]["model_input_history"]
    assert all(item.get("workbook_sha256") for item in manifest["source_hashes"]["workbooks"])
    assert manifest["mbu26_annual_spine"]["status"] == "mbu26_annual_spine_vendored"
    assert manifest["mbu26_annual_spine"]["source_release"] == "MBU26"
    assert manifest["mbu26_annual_spine"]["workbook_sha256"] == "9aaff21f72c0a10cfa972a29d3c4f716495c79cbd72fc28e8008a65558454e12"
    assert manifest["mbu26_annual_spine"]["sheet"] == "MBU26"
    assert manifest["source_hashes"]["mbu26_annual_spine"]["repo_relative_path"] == "data/revenue_model_source_pack/mbu26_annual_spine"
    # Deliberate re-freeze (official-vintage framework): the runtime source
    # block now describes the bridge-assumption vintage (BEFU26). The MBU26
    # prior vintage keeps its own pinned block above.
    assert manifest["revenue_source_pack"]["status"] == "official_vintage_pack_vendored"
    assert manifest["revenue_source_pack"]["source_pack_version"] == "BEFU26"
    assert manifest["revenue_source_pack"]["raw_workbook_sha256"] == "7d6e5b19119ca8b5272ca2205c0735719033d82484ce674cfb595e6f45d085ff"
    assert manifest["revenue_source_pack"]["selections"]["release_round"] == "BEFU26"
    assert manifest["revenue_source_pack"]["selections"]["series"] == "Total NLTF revenue"
    assert manifest["revenue_source_pack"]["dashboard_default_selections"]["series"] == "Total NLTF revenue"
    assert manifest["revenue_source_pack"]["source_workbook_selections"]["sheet"] == "Baseline"
    assert "BEFU26 source pack" in manifest["revenue_source_pack"]["default_selection_policy"]
    assert manifest["official_vintages"]["official_comparator_vintage_id"] == "BEFU26"
    assert manifest["official_vintages"]["bridge_assumption_vintage_id"] == "BEFU26"
    assert set(manifest["official_vintages"]["available"]) == {"BEFU26", "MBU26"}
    assert manifest["revenue_line_reconciliation"]["repo_relative_path"] == "data/current_revenue_outlook/revenue_line_reconciliation.csv"
    assert manifest["revenue_stack_components"]["repo_relative_path"] == "data/current_revenue_outlook/revenue_stack_components.csv"
    assert "aggregates are overlays only" in manifest["revenue_stack_components"]["scope"]
    assert manifest["ev_phev_split_assumptions"]["repo_relative_path"] == "data/current_revenue_outlook/ev_phev_split_assumptions.csv"
    # The status now says what the table IS (an audit-only legacy split beside
    # the active conventional anchor) rather than claiming the retired
    # PED+Light migration superseded it and is therefore the live path.
    assert manifest["ev_phev_split_assumptions"]["allocation_status"] == "legacy_split_audit_only_conventional_anchor_active"
    assert (
        manifest["ev_phev_ped_light_drift_assumptions"]["repo_relative_path"]
        == "data/current_revenue_outlook/ev_phev_ped_light_drift_assumptions.csv"
    )
    # The drift table is retained for audit continuity only. The settled
    # decision-facing architecture is the conventional anchor plus exact-VFM
    # composition, so the manifest must describe it that way and label the
    # lambda material as retired rather than re-freezing the obsolete story.
    drift_block = manifest["ev_phev_ped_light_drift_assumptions"]
    assert drift_block["decision_facing"] is False
    assert "AUDIT ONLY, NON-DECISION-FACING" in drift_block["scope"]
    assert "conventional_anchor" in drift_block["composition_architecture"]
    retired = drift_block["retired_lambda_material"]
    assert retired["status"] == "retired_from_decision_facing_use"
    assert retired["audit_default_mode"] == "optimized"
    assert float(retired["audit_smoothness_penalty"]) > 0
    assert "conventional_anchor_exact_vfm" in manifest["equations"]["EV_PHEV_CLASS_COMPOSITION"] or (
        "CONVENTIONAL class" in manifest["equations"]["EV_PHEV_CLASS_COMPOSITION"]
    )
    assert manifest["ped_revenue_bridge_audit"]["repo_relative_path"] == "data/current_revenue_outlook/ped_revenue_bridge_audit.csv"
    assert "raw VKTpc x population" in manifest["ped_revenue_bridge_audit"]["scope"]
    assert manifest["ped_revenue_bridge_audit"]["default_bridge_mode"] == PED_BRIDGE_DEFAULT_MODE
    assert manifest["ped_revenue_bridge_audit"]["population_proxy_warning_rows"] >= 1
    assert manifest["ped_bridge_shape_fit_metrics"]["repo_relative_path"] == "data/current_revenue_outlook/ped_bridge_shape_fit_metrics.csv"
    assert "raw-vs-optimized" in manifest["ped_bridge_shape_fit_metrics"]["scope"]
    assert manifest["ped_bridge_mode_config"]["repo_relative_path"] == "data/current_revenue_outlook/ped_bridge_mode_config.csv"
    assert manifest["ped_bridge_mode_config"]["default_bridge_mode"] == PED_BRIDGE_DEFAULT_MODE
    assert manifest["ped_efficiency_scenarios"]["repo_relative_path"] == "data/current_revenue_outlook/ped_efficiency_scenarios.csv"
    assert manifest["ped_efficiency_scenarios"]["default_scenario_id"] == PED_EFFICIENCY_BASELINE_SCENARIO_ID
    assert manifest["ped_efficiency_scenarios"]["default_runtime_treatment"] == "0pct_no_change"
    assert manifest["sensitivity_seed_inputs"]["repo_relative_path"] == "data/current_revenue_outlook/sensitivity_seed_inputs.csv"
    assert manifest["sensitivity_seed_inputs"]["source_workbook_sha256"] == SENSITIVITY_SEED_WORKBOOK_SHA256
    assert "fleet transition" in manifest["sensitivity_seed_inputs"]["excluded_scope"].lower()
    assert "crude/oil shock" in manifest["sensitivity_seed_inputs"]["excluded_scope"]
    assert manifest["sensitivity_config"]["repo_relative_path"] == "data/current_revenue_outlook/sensitivity_config.csv"
    assert manifest["sensitivity_config"]["default_runtime_treatment"] == "all_off_no_change"
    assert manifest["sensitivity_impact_audit"]["repo_relative_path"] == "data/current_revenue_outlook/sensitivity_impact_audit.csv"
    assert manifest["scenario_role_contract"]["repo_relative_path"] == "data/current_revenue_outlook/scenario_role_contract.csv"
    assert "behavioural intensity metric" not in str(manifest["scenario_role_contract"].get("note", ""))
    assert manifest["scenario_inputs"]["status"] == "available"
    assert manifest["scenario_inputs"]["repo_relative_output_dir"] == "data/current_revenue_outlook/scenario_inputs"
    assert manifest["scenario_inputs"]["schema_version"] == "nltf-scenario-input-materializer-v1"
    assert manifest["scenario_inputs"]["row_counts"] == {
        "scenario_input_cells": 15472,
        "scenario_input_long": 15200,
        "scenario_input_wide": 600,
        "scenario_feature_lineage": 44800,
    }
    assert manifest["scenario_input_delta_audit"]["repo_relative_path"] == (
        "data/current_revenue_outlook/scenario_input_delta_audit.csv"
    )
    assert manifest["scenario_input_delta_audit"]["status"] == "available"
    assert manifest["scenario_input_delta_audit"]["source"] == "scenario_inputs/scenario_input_long.parquet"
    assert "workbook-cell base/comparison deltas" in manifest["scenario_input_delta_audit"]["scope"].lower()
    scenario_input_manifest_path = pack_dir / "scenario_inputs" / "scenario_input_manifest.json"
    assert scenario_input_manifest_path.exists()
    assert manifest["scenario_inputs"]["manifest_sha256"] == _sha256(scenario_input_manifest_path)
    scenario_input_manifest_text = scenario_input_manifest_path.read_text(encoding="utf-8")
    assert "C:\\Users" not in scenario_input_manifest_text
    assert "Downloads" not in scenario_input_manifest_text
    scenario_input_manifest = json.loads(scenario_input_manifest_text)
    # Canonical generation route is two-stage promote -> rebuild, which
    # materialises each reviewed workbook then combines. Both routes are proven
    # to produce identical scenario-input content by
    # tests/test_scenario_input_route_equivalence.py; only this descriptive
    # wording differs. See docs/PACK_PROVENANCE_FINDING.md.
    assert scenario_input_manifest["source_policy"] == "combined committed scenario input artifacts; Streamlit must not load Excel at runtime"
    assert scenario_input_manifest["row_counts"] == manifest["scenario_inputs"]["row_counts"]
    assert len(scenario_input_manifest["workbooks"]) == 2
    sheet_inventory = scenario_input_manifest["sheet_inventory"]
    assert len(sheet_inventory) == 10
    assert {row["sheet"] for row in sheet_inventory} == {
        "README",
        "PED Inputs",
        "Light RUC Inputs",
        "Heavy RUC Inputs",
        "Assumptions",
    }
    assert {row["source_status"] for row in sheet_inventory} == {"all_non_empty_cells_materialized"}
    assert sum(row["materialized_cell_count"] for row in sheet_inventory) == manifest["scenario_inputs"]["row_counts"][
        "scenario_input_cells"
    ]
    assert all(row["materialized_cell_count"] == row["non_empty_cell_count"] for row in sheet_inventory)
    assert all(len(row["materialized_cells_sha256"]) == 64 for row in sheet_inventory)
    assert {workbook["workbook_sha256"] for workbook in scenario_input_manifest["workbooks"]} == {
        "d0644d353ee5a073602186cf7ac5c16e707d5350e16fd037b73a65528067cc6a",
        "6213ce565cf1f4a058a3ea9f1af4d5476a8b0423a4d8747905c3cba128380ce1",
    }
    assert {workbook["raw_status"] for workbook in scenario_input_manifest["workbooks"]} == {
        "copied_repo_local_raw_workbook"
    }
    for workbook in scenario_input_manifest["workbooks"]:
        raw_path = ROOT / workbook["raw_repo_relative_path"]
        assert raw_path.exists()
        assert raw_path.stat().st_size == workbook["size_bytes"]
        assert raw_path.stat().st_size < 50 * 1024 * 1024
        assert _sha256(raw_path) == workbook["workbook_sha256"]
        assert len(workbook["sheet_inventory"]) == workbook["sheet_count"]
        assert sum(row["materialized_cell_count"] for row in workbook["sheet_inventory"]) == workbook[
            "non_empty_cell_count"
        ]
    for output_file, metadata in scenario_input_manifest["output_files"].items():
        assert metadata["sha256"] == _sha256(ROOT / metadata["repo_relative_path"]), output_file
        assert metadata["repo_relative_path"].startswith("data/current_revenue_outlook/scenario_inputs/")
    assert (
        manifest["scenario_feature_lineage"]["repo_relative_path"]
        == "data/current_revenue_outlook/scenario_feature_lineage.csv"
    )
    assert manifest["scenario_feature_lineage"]["source"] == "scenario_inputs/scenario_feature_lineage.parquet"
    assert (
        manifest["scenario_input_replay_mismatch_report"]["repo_relative_path"]
        == "data/current_revenue_outlook/scenario_input_replay_mismatch_report.csv"
    )
    assert manifest["scenario_input_replay_mismatch_report"]["status"] == "passed_no_mismatch"
    assert "raises" in manifest["scenario_input_replay_mismatch_report"]["fail_policy"]
    # Updated to the runtime contract: Light RUC is anchored on the CONVENTIONAL
    # forecast with exact VFM class shares. The previous assertions pinned the
    # retired optimized-migration description.
    assert manifest["target_semantics_audit"]["LIGHT_RUC"]["status"] == "conventional_anchor_with_exact_vfm_composition"
    light_ruc_note = manifest["data_vintage_manifest_notes"]["light_ruc_target_semantics"]
    assert "anchored on the CONVENTIONAL model forecast" in light_ruc_note
    assert "VFM202405" in light_ruc_note
    assert manifest["revenue_formula_residuals"]["repo_relative_path"] == "data/current_revenue_outlook/revenue_formula_residuals.csv"
    assert manifest["series_alias_audit"]["repo_relative_path"] == "data/current_revenue_outlook/series_alias_audit.csv"
    assert manifest["fan_availability"]["repo_relative_path"] == "data/current_revenue_outlook/fan_availability.csv"
    assert manifest["fan_band_rows"]["repo_relative_path"] == "data/current_revenue_outlook/fan_band_rows.csv"
    assert sorted(manifest["output_hashes"]) == [
        "ev_phev_ped_light_drift_assumptions.csv",
        "ev_phev_ped_light_drift_assumptions.parquet",
        "ev_phev_split_assumptions.csv",
        "ev_phev_split_assumptions.parquet",
        "fan_availability.csv",
        "fan_availability.parquet",
        "fan_band_rows.csv",
        "fan_band_rows.parquet",
        "future_revenue_forecasts.csv",
        "future_revenue_forecasts.parquet",
        "horizon_contract_audit.csv",
        "horizon_contract_audit.parquet",
        "light_ruc_horizon_availability.csv",
        "light_ruc_horizon_availability.parquet",
        "path_trace_status.csv",
        "path_trace_status.parquet",
        "ped_bridge_mode_config.csv",
        "ped_bridge_mode_config.parquet",
        "ped_bridge_shape_fit_metrics.csv",
        "ped_bridge_shape_fit_metrics.parquet",
        "ped_efficiency_scenarios.csv",
        "ped_efficiency_scenarios.parquet",
        "ped_revenue_bridge_audit.csv",
        "ped_revenue_bridge_audit.parquet",
        "quarterly_reconstitution_audit.csv",
        "quarterly_reconstitution_audit.parquet",
        "raw_quarterly_forecast_audit.csv",
        "raw_quarterly_forecast_audit.parquet",
        "revenue_bridge_components.csv",
        "revenue_bridge_components.parquet",
        "revenue_chart_rows.csv",
        "revenue_chart_rows.parquet",
        "revenue_formula_residuals.csv",
        "revenue_formula_residuals.parquet",
        "revenue_line_reconciliation.csv",
        "revenue_line_reconciliation.parquet",
        "revenue_stack_components.csv",
        "revenue_stack_components.parquet",
        "row_reconciliation.csv",
        "row_reconciliation.parquet",
        "runtime_cutoff_audit.csv",
        "runtime_cutoff_audit.parquet",
        "runtime_trace_audit.csv",
        "runtime_trace_audit.parquet",
        "scenario_feature_lineage.csv",
        "scenario_feature_lineage.parquet",
        "scenario_input_delta_audit.csv",
        "scenario_input_delta_audit.parquet",
        "scenario_input_replay_mismatch_report.csv",
        "scenario_input_replay_mismatch_report.parquet",
        "scenario_role_contract.csv",
        "scenario_role_contract.parquet",
        "sensitivity_config.csv",
        "sensitivity_config.parquet",
        "sensitivity_impact_audit.csv",
        "sensitivity_impact_audit.parquet",
        "sensitivity_seed_inputs.csv",
        "sensitivity_seed_inputs.parquet",
        "series_alias_audit.csv",
        "series_alias_audit.parquet",
        "series_trace_contract.csv",
        "series_trace_contract.parquet",
        "stream_vintage_status.csv",
        "stream_vintage_status.parquet",
        "trace_source_contract.csv",
        "trace_source_contract.parquet",
    ]
    for filename, metadata in manifest["output_hashes"].items():
        assert metadata["sha256"] == _sha256(pack_dir / filename)
    for filename in ["future_revenue_forecasts.parquet", "revenue_bridge_components.parquet", "revenue_chart_rows.parquet"]:
        frame = pd.read_parquet(pack_dir / filename)
        assert set(CANONICAL_JOIN_KEY_COLUMNS).issubset(frame.columns)
        assert frame["canonical_join_key"].astype(str).str.count(r"\|").eq(2).all()
    for path in pack_dir.iterdir():
        if path.is_file():
            assert path.stat().st_size < 50 * 1024 * 1024


def test_committed_current_revenue_outlook_runtime_contract() -> None:
    pack_dir = ROOT / CURRENT_REVENUE_OUTLOOK_DIR
    manifest = json.loads((pack_dir / "manifest.json").read_text(encoding="utf-8"))
    chart = pd.read_parquet(pack_dir / "revenue_chart_rows.parquet")
    bridge = pd.read_parquet(pack_dir / "revenue_bridge_components.parquet")
    future = pd.read_parquet(pack_dir / "future_revenue_forecasts.parquet")
    audit = pd.read_parquet(pack_dir / "runtime_trace_audit.parquet")
    line_reconciliation = pd.read_parquet(pack_dir / "revenue_line_reconciliation.parquet")
    stack_components = pd.read_parquet(pack_dir / "revenue_stack_components.parquet")
    ev_phev_split = pd.read_parquet(pack_dir / "ev_phev_split_assumptions.parquet")
    ev_phev_drift = pd.read_parquet(pack_dir / "ev_phev_ped_light_drift_assumptions.parquet")
    ped_bridge_audit = pd.read_parquet(pack_dir / "ped_revenue_bridge_audit.parquet")
    ped_bridge_shape_fit = pd.read_parquet(pack_dir / "ped_bridge_shape_fit_metrics.parquet")
    ped_bridge_mode_config = pd.read_parquet(pack_dir / "ped_bridge_mode_config.parquet")
    ped_efficiency_scenarios = pd.read_parquet(pack_dir / "ped_efficiency_scenarios.parquet")
    sensitivity_seed_inputs = pd.read_parquet(pack_dir / "sensitivity_seed_inputs.parquet")
    sensitivity_config = pd.read_parquet(pack_dir / "sensitivity_config.parquet")
    sensitivity_impact_audit = pd.read_parquet(pack_dir / "sensitivity_impact_audit.parquet")
    scenario_role_contract = pd.read_parquet(pack_dir / "scenario_role_contract.parquet")
    residuals = pd.read_parquet(pack_dir / "revenue_formula_residuals.parquet")
    alias_audit = pd.read_parquet(pack_dir / "series_alias_audit.parquet")
    runtime_cutoff_audit = pd.read_parquet(pack_dir / "runtime_cutoff_audit.parquet")
    fan_availability = pd.read_parquet(pack_dir / "fan_availability.parquet")
    fan_bands = pd.read_parquet(pack_dir / "fan_band_rows.parquet")
    scenario_feature_lineage = pd.read_parquet(pack_dir / "scenario_feature_lineage.parquet")
    scenario_input_delta = pd.read_parquet(pack_dir / "scenario_input_delta_audit.parquet")
    scenario_input_replay = pd.read_parquet(pack_dir / "scenario_input_replay_mismatch_report.parquet")
    scenario_input_wide = pd.read_parquet(pack_dir / "scenario_inputs" / "scenario_input_wide.parquet")

    assert manifest["runtime_pack_type"] == "official_vintage_actual_current_finalist_official_comparator"
    assert manifest["bridge_status_by_stream"] == {
        "PED": ["available"],
        "LIGHT_RUC": ["available"],
        "HEAVY_RUC": ["available"],
    }
    assert "workbook model" not in json.dumps(manifest).lower()
    assert "annual_model_paths" not in json.dumps(manifest).lower()
    assert "nominal_rate_missing" not in json.dumps(manifest)
    assert "ped_bridge_source_history_missing" not in json.dumps(manifest)
    assert "extrapolated_model_extension" not in json.dumps(manifest)
    assert "extrapolated from FY2046" not in json.dumps(manifest)
    runtime_cutoff_fy = int(manifest["period_rule"]["runtime_cutoff_fy"])
    # FY2050 was the pre-policy horizon. The current model now stops at the
    # governed decision-facing cutoff; the official comparator publishes over
    # its own source horizon, which the same manifest records separately.
    assert runtime_cutoff_fy == LAST_DECISION_GRADE_ANNUAL_FY
    period_rule = manifest["period_rule"]
    assert period_rule["current_light_ruc_quarterly_cutoff"] == LAST_DECISION_GRADE_QUARTER
    assert int(period_rule["current_light_ruc_annual_cutoff_fy"]) == LAST_DECISION_GRADE_ANNUAL_FY
    assert int(period_rule["official_comparator_cutoff_fy"]) > runtime_cutoff_fy
    assert str(period_rule["raw_audit_source_horizon"]).startswith("20")
    assert manifest["runtime_cutoff_audit"]["repo_relative_path"] == "data/current_revenue_outlook/runtime_cutoff_audit.csv"
    assert manifest["runtime_cutoff_audit"]["runtime_cutoff_fy"] == runtime_cutoff_fy
    assert not runtime_cutoff_audit.empty
    # The required-components audit row follows the bridge-assumption vintage
    # (BEFU26 since the official-vintage framework).
    assert set(runtime_cutoff_audit["audit_component"].astype(str)) == {
        "current_finalist_base",
        "current_finalist_comparison",
        "befu26_required_components_rates_splits",
        "runtime_cutoff",
    }
    assert pd.to_numeric(runtime_cutoff_audit["runtime_cutoff_fy"], errors="coerce").eq(runtime_cutoff_fy).all()
    # The invariant this protects is that the FY2051-FY2055 gradient extension
    # stays disabled. It previously asserted "no extrapolated model extension is
    # used", which stopped being true when the governed FY2031-FY2050 post-model
    # layer was added: the assertion was pinning stale wording, not the property.
    runtime_cutoff_note = manifest["data_vintage_manifest_notes"]["runtime_cutoff"].lower()
    assert "fy2051-fy2055 gradient extension remains disabled" in runtime_cutoff_note
    assert "econometric segment stops at fy2030" in runtime_cutoff_note
    assert f"FY{runtime_cutoff_fy}" in manifest["data_vintage_manifest_notes"]["official_horizon_note"]
    # These frames are MIXED: they carry current-model rows and official
    # comparator rows together. Asserting one scalar cutoff over the whole
    # frame is precisely what let the current-model H20 rule truncate the
    # published official comparator, so each scope is checked against its own
    # horizon.
    official_cutoff_fy = int(manifest["period_rule"]["official_comparator_cutoff_fy"])

    def _is_official(frame: pd.DataFrame) -> pd.Series:
        """Official-comparator rows, however this frame labels its scope.

        Fan bands carry no scenario_role and identify the comparator only by
        scenario_name, so a role-only test would silently classify official
        rows as current and re-impose the current cutoff on them.
        """
        mask = pd.Series(False, index=frame.index)
        if "scenario_role" in frame.columns:
            mask |= frame["scenario_role"].astype(str).eq("official_comparator")
        if "scenario_name" in frame.columns:
            mask |= frame["scenario_name"].astype(str).eq("mbu26_official")
        if "source_path" in frame.columns:
            mask |= frame["source_path"].astype(str).str.contains("MBU26", na=False)
        return mask

    def _current_only(frame: pd.DataFrame) -> pd.DataFrame:
        return frame[~_is_official(frame)].copy()

    def _official_only(frame: pd.DataFrame) -> pd.DataFrame:
        return frame[_is_official(frame)].copy()

    def _annual_only(frame: pd.DataFrame) -> pd.DataFrame:
        # Quarterly rows are governed by HORIZON, not by June year: 2030Q3/H19
        # and 2030Q4/H20 publish and both sit in June year 2031, whose ANNUAL
        # result is correctly withheld. Mixing the two grains under one June-year
        # ceiling is what entangled the quarterly and annual contracts.
        if "time_grain" in frame.columns:
            return frame[frame["time_grain"].astype(str).eq("june_year")].copy()
        return frame

    # The ECONOMETRIC segment still stops at the runtime cutoff. FY2031-FY2050
    # now publishes as the separately governed post-model extrapolation, so
    # the ceiling is asserted per segment rather than over the whole frame -
    # otherwise the long-run layer would look like a cutoff breach.
    for frame in (chart, line_reconciliation, stack_components, fan_bands, future, bridge):
        current_rows = _annual_only(_current_only(frame))
        if current_rows.empty:
            continue
        # Two frames carry a segment label under different names: chart/line
        # rows use forecast_segment, the fan bands use fan_segment. Both split
        # at the same FY2030 seam.
        if "forecast_segment" in current_rows.columns:
            segments = current_rows["forecast_segment"].fillna("").astype(str)
            long_run = segments.eq("post_model_extrapolation")
        elif "fan_segment" in current_rows.columns:
            segments = current_rows["fan_segment"].fillna("").astype(str)
            long_run = segments.eq("long_run_scenario_envelope")
        else:
            assert _max_fy(current_rows) == runtime_cutoff_fy
            continue
        within = current_rows[~long_run]
        beyond = current_rows[long_run]
        if not within.empty:
            assert _max_fy(within) == runtime_cutoff_fy
        if not beyond.empty:
            assert _max_fy(beyond) == LAST_POST_MODEL_FY
            assert _min_fy(beyond) >= FIRST_POST_MODEL_FY

    # The supported quarters survive and carry the next June year.
    current_quarterly = chart[
        chart["time_grain"].astype(str).eq("quarterly")
        & chart["row_type"].astype(str).eq("future_forecast")
    ]
    assert set(current_quarterly["period"].astype(str)) >= {"2030Q3", "2030Q4"}
    assert _max_fy(current_quarterly) == runtime_cutoff_fy + 1

    official_rows = _official_only(chart)
    assert not official_rows.empty
    assert _max_fy(official_rows) == official_cutoff_fy

    official_source = pd.read_parquet(ROOT / "data/revenue_model_source_pack/mbu26_annual_spine/mbu26_official_annual.parquet")
    assert (_max_fy(official_source) or 0) > runtime_cutoff_fy
    assert _max_fy(official_source) == official_cutoff_fy
    displayed = chart[
        chart["time_grain"].astype(str).eq("june_year")
        & chart["plot_allowed"].astype(str).str.lower().isin(["true", "1"])
    ].copy()
    # Displayed current rows now run to FY2050; the ECONOMETRIC portion still
    # stops at the runtime cutoff and the remainder is the labelled
    # post-model segment.
    displayed_current = _current_only(displayed)
    displayed_segments = displayed_current["forecast_segment"].fillna("").astype(str)
    assert _max_fy(displayed_current[~displayed_segments.eq("post_model_extrapolation")]) == runtime_cutoff_fy
    assert _max_fy(displayed_current[displayed_segments.eq("post_model_extrapolation")]) == LAST_POST_MODEL_FY
    current_line = line_reconciliation[line_reconciliation["source_path"].astype(str).str.startswith("Current finalist")].copy()
    line_segments = current_line["forecast_segment"].fillna("").astype(str)
    assert _max_fy(current_line[~line_segments.eq("post_model_extrapolation")]) == runtime_cutoff_fy
    assert _max_fy(current_line[line_segments.eq("post_model_extrapolation")]) == LAST_POST_MODEL_FY
    runtime_tables = [chart, line_reconciliation, stack_components, audit, fan_bands]
    for frame in runtime_tables:
        assert not frame.astype(str).stack().str.contains("extrapolated_model_extension", regex=False).any()
        assert not frame.astype(str).stack().str.contains("extrapolated from FY2046", regex=False).any()
    assert manifest["target_semantics_audit"]["HEAVY_RUC"]["status"] == "not_reclassified"
    assert manifest["scenario_role_contract"]["repo_relative_path"] == "data/current_revenue_outlook/scenario_role_contract.csv"
    assert not scenario_role_contract.empty
    required_contract_columns = {
        "scenario_name",
        "scenario_role",
        "differing_fields",
        "population_only_flag",
        "behavioural_driver_flag",
        "affected_series",
        "interpretation",
        "display_policy",
        "affects_ped_vktpc_directly",
        "affects_bridge_scaling",
        "stream_differing_fields",
        "ped_vktpc_direct_fields",
        "bridge_scaling_fields",
        "bridge_only_fields",
        "unknown_fields",
    }
    assert required_contract_columns.issubset(scenario_role_contract.columns)
    comparison_ped_contract = scenario_role_contract[
        scenario_role_contract["scenario_name"].astype(str).eq("current_comparison_1")
        & scenario_role_contract["affected_series"].astype(str).eq("ped_vkt_per_capita")
    ].iloc[0]
    assert not bool(comparison_ped_contract["population_only_flag"])
    assert bool(comparison_ped_contract["behavioural_driver_flag"])
    assert comparison_ped_contract["display_policy"] == "keep_trace_relabel_comparison_behavioural_path"
    assert "population__level" in str(comparison_ped_contract["ped_population_feature_fields"])
    assert bool(comparison_ped_contract["affects_ped_vktpc_directly"])
    assert bool(comparison_ped_contract["affects_bridge_scaling"])
    assert "gdp_petrol_interaction" in str(comparison_ped_contract["ped_vktpc_direct_fields"])
    assert "population" in str(comparison_ped_contract["ped_vktpc_direct_fields"])
    assert str(comparison_ped_contract["bridge_scaling_fields"]) == "population"
    assert str(comparison_ped_contract["bridge_only_fields"]) == ""
    assert str(comparison_ped_contract["unknown_fields"]) == ""
    assert pd.to_numeric(comparison_ped_contract["runtime_delta_min"], errors="coerce") < 0
    assert "behavioural intensity metric" not in str(comparison_ped_contract["notes"])
    comparison_revenue_contract = scenario_role_contract[
        scenario_role_contract["scenario_name"].astype(str).eq("current_comparison_1")
        & scenario_role_contract["affected_series"].astype(str).eq("gross_ped_revenue")
    ].iloc[0]
    assert comparison_revenue_contract["display_policy"] == "keep_comparison_trace_scale_or_bridge"
    assert comparison_revenue_contract["population_path_policy"] == "scenario_input_population_from_committed_workbook_artifacts"
    assert "population:population/scale" in str(comparison_revenue_contract["field_classification"])
    assert "real_gdp_sa_nzd:macro" in str(comparison_revenue_contract["field_classification"])
    assert "unemployment_rate:macro" in str(comparison_revenue_contract["field_classification"])
    assert "gdp_petrol_interaction:price/rate/policy" in str(comparison_revenue_contract["field_classification"])
    assert "target_lag_1:behavioural" in str(comparison_revenue_contract["field_classification"])
    assert bool(comparison_revenue_contract["affects_ped_vktpc_directly"])
    assert bool(comparison_revenue_contract["affects_bridge_scaling"])
    total_fed_contract = scenario_role_contract[
        scenario_role_contract["scenario_name"].astype(str).eq("current_comparison_1")
        & scenario_role_contract["affected_series"].astype(str).eq("total_fed_ruc_net_revenue")
    ].iloc[0]
    assert bool(total_fed_contract["affects_ped_vktpc_directly"])
    assert bool(total_fed_contract["affects_bridge_scaling"])
    total_ruc_contract = scenario_role_contract[
        scenario_role_contract["scenario_name"].astype(str).eq("current_comparison_1")
        & scenario_role_contract["affected_series"].astype(str).eq("total_ruc_net_revenue")
    ].iloc[0]
    assert not bool(total_ruc_contract["affects_ped_vktpc_directly"])
    assert not bool(total_ruc_contract["affects_bridge_scaling"])
    comparison_categories = {
        part.split(":", 1)[1].strip()
        for text in scenario_role_contract["field_classification"].dropna().astype(str)
        for part in text.split(";")
        if ":" in part
    }
    assert {"population/scale", "macro", "price/rate/policy", "behavioural"}.issubset(comparison_categories)
    assert "scenario_input_wide" in str(comparison_revenue_contract["source_basis"])
    assert not scenario_input_delta.empty
    required_delta_columns = {
        "base_workbook_sha256",
        "comparison_workbook_sha256",
        "base_cell",
        "comparison_cell",
        "canonical_period",
        "canonical_variable",
        "base_value",
        "comparison_value",
        "absolute_delta",
        "pct_delta",
        "field_classification",
        "affects_ped_vktpc_directly",
        "affects_bridge_scaling",
        "source_status",
    }
    assert required_delta_columns.issubset(scenario_input_delta.columns)
    assert set(scenario_input_delta["source_status"].dropna().astype(str)) == {"committed_scenario_input_delta"}
    assert scenario_input_delta["base_workbook_sha256"].astype(str).eq(
        "d0644d353ee5a073602186cf7ac5c16e707d5350e16fd037b73a65528067cc6a"
    ).all()
    assert scenario_input_delta["comparison_workbook_sha256"].astype(str).eq(
        "6213ce565cf1f4a058a3ea9f1af4d5476a8b0423a4d8747905c3cba128380ce1"
    ).all()
    assert scenario_input_delta["base_cell"].astype(str).str.len().gt(0).all()
    assert scenario_input_delta["comparison_cell"].astype(str).str.len().gt(0).all()
    assert {"PED", "LIGHT_RUC", "HEAVY_RUC"}.issubset(set(scenario_input_delta["stream"].astype(str)))
    assert {"population/scale", "macro", "price/rate/policy", "behavioural"}.issubset(
        set(scenario_input_delta["field_classification"].dropna().astype(str))
    )
    ped_population_delta = scenario_input_delta[
        scenario_input_delta["stream"].astype(str).eq("PED")
        & scenario_input_delta["canonical_variable"].astype(str).eq("population")
    ].copy()
    assert not ped_population_delta.empty
    assert ped_population_delta["field_classification"].astype(str).eq("population/scale").all()
    assert ped_population_delta["affects_ped_vktpc_directly"].astype(bool).all()
    assert ped_population_delta["affects_bridge_scaling"].astype(bool).all()
    assert not scenario_input_delta.astype(str).stack().str.contains(r"C:\\Users|Downloads|OneDrive", regex=True).any()
    required_source_split_columns = {
        "vktpc_source_file",
        "vktpc_source_cell",
        "vktpc_source_status",
        "population_source_file",
        "population_source_cell",
        "population_source_status",
    }
    assert required_source_split_columns.issubset(line_reconciliation.columns)
    assert required_source_split_columns.issubset(chart.columns)
    source_split_lines = line_reconciliation[
        line_reconciliation["scenario_name"].astype(str).eq("current_comparison_1")
        & line_reconciliation["series_id"].astype(str).isin(["ped_volume", "gross_ped_revenue"])
    ].copy()
    assert not source_split_lines.empty
    assert source_split_lines["vktpc_source_file"].astype(str).eq("forecast_scenario_comparison.parquet").all()
    assert source_split_lines["vktpc_source_status"].astype(str).eq("current_finalist_model").all()
    assert source_split_lines["vktpc_source_cell"].astype(str).str.len().gt(0).all()
    assert source_split_lines["population_source_file"].astype(str).eq("scenario_inputs/scenario_input_wide.parquet").all()
    assert source_split_lines["population_source_cell"].astype(str).str.contains("scenario_input_wide.parquet:PED:population").all()
    assert source_split_lines["population_source_status"].astype(str).str.startswith("scenario_input_population").all()
    source_split_chart = chart[
        chart["scenario_name"].astype(str).eq("current_comparison_1")
        & chart["series_id"].astype(str).eq("gross_ped_revenue")
    ].copy()
    assert not source_split_chart.empty
    assert source_split_chart["vktpc_source_file"].astype(str).eq("forecast_scenario_comparison.parquet").all()
    assert source_split_chart["population_source_file"].astype(str).eq("scenario_inputs/scenario_input_wide.parquet").all()
    assert not scenario_feature_lineage.empty
    assert set(scenario_feature_lineage["stream"].dropna().astype(str)) == {"PED", "LIGHT_RUC", "HEAVY_RUC"}
    required_lineage_columns = {
        "lineage_role",
        "feature_source_variables",
        "feature_engineering_basis",
        "feature_lineage_status",
    }
    assert required_lineage_columns.issubset(scenario_feature_lineage.columns)
    assert {"source_variable", "model_feature"}.issubset(
        set(scenario_feature_lineage["lineage_role"].dropna().astype(str))
    )
    source_variable_lineage = scenario_feature_lineage[
        scenario_feature_lineage["lineage_role"].astype(str).eq("source_variable")
    ].copy()
    model_feature_lineage = scenario_feature_lineage[
        scenario_feature_lineage["lineage_role"].astype(str).eq("model_feature")
    ].copy()
    assert set(source_variable_lineage["source_status"].dropna().astype(str)) == {"committed_scenario_input"}
    assert not source_variable_lineage["fallback_flag"].astype(bool).any()
    assert source_variable_lineage["source_artifact"].eq("scenario_inputs/scenario_input_long.parquet").all()
    assert not model_feature_lineage.empty
    assert set(model_feature_lineage["stream"].dropna().astype(str)) == {"PED", "LIGHT_RUC", "HEAVY_RUC"}
    required_model_features = {
        "LIGHT_RUC": {"diesel_x_ruc_price", "log_real_gdp_lag4"},
        "PED": {"gdp_pc__log", "policy__petrol_abs_change_1_lag4", "target__roll8_mean"},
        "HEAVY_RUC": {"heavy_price__log", "policy__diesel_abs_change_1_lag4", "target__roll8_mean"},
    }
    for stream, expected_features in required_model_features.items():
        features = set(
            model_feature_lineage.loc[model_feature_lineage["stream"].astype(str).eq(stream), "feature_name"].astype(str)
        )
        assert expected_features.issubset(features), stream
    target_lineage = model_feature_lineage[model_feature_lineage["feature_name"].astype(str).str.startswith("target__")]
    assert not target_lineage.empty
    assert target_lineage["fallback_flag"].astype(bool).all()
    assert target_lineage["fallback_reason"].astype(str).str.contains("recursive target-lag", regex=False).all()
    assert scenario_feature_lineage["canonical_variable"].astype(str).str.len().gt(0).all()
    assert set(scenario_input_wide["scenario_name"].dropna().astype(str)) == {
        "current_basecase",
        "current_comparison_1",
    }
    ped_population_inputs = scenario_input_wide[
        scenario_input_wide["stream"].astype(str).eq("PED")
        & scenario_input_wide["population"].fillna("").astype(str).ne("")
    ]
    assert not ped_population_inputs.empty
    assert set(ped_population_inputs["scenario_name"].dropna().astype(str)) == {
        "current_basecase",
        "current_comparison_1",
    }
    assert not scenario_input_replay.empty
    assert set(scenario_input_replay["mismatch_status"].dropna().astype(str)) == {
        "pass",
        "not_applicable",
    }
    assert "mismatch" not in set(scenario_input_replay["mismatch_status"].dropna().astype(str))
    matched_replay = scenario_input_replay[
        scenario_input_replay["scenario_input_status"].astype(str).eq("matched_committed_scenario_input")
    ].copy()
    assert not matched_replay.empty
    assert matched_replay["workbook_sha256"].astype(str).eq(matched_replay["manifest_workbook_sha256"].astype(str)).all()
    assert pd.to_numeric(matched_replay["required_feature_count"], errors="coerce").gt(0).all()
    assert pd.to_numeric(matched_replay["missing_required_feature_count"], errors="coerce").eq(0).all()
    required_replay_columns = {
        "replay_forecast_value",
        "promoted_forecast_value",
        "replay_abs_delta",
        "replay_tolerance",
        "replay_status",
    }
    assert required_replay_columns.issubset(scenario_input_replay.columns)
    assert matched_replay["replay_status"].astype(str).eq("pass").all()
    replay_values = pd.to_numeric(matched_replay["replay_forecast_value"], errors="coerce")
    promoted_values = pd.to_numeric(matched_replay["promoted_forecast_value"], errors="coerce")
    replay_deltas = pd.to_numeric(matched_replay["replay_abs_delta"], errors="coerce")
    replay_tolerances = pd.to_numeric(matched_replay["replay_tolerance"], errors="coerce")
    assert replay_values.notna().all()
    assert promoted_values.notna().all()
    assert replay_deltas.notna().all()
    assert replay_tolerances.notna().all()
    assert replay_tolerances.gt(0).all()
    assert replay_deltas.le(replay_tolerances).all()
    assert "governed_model_extension_not_replayed_from_workbook" not in set(
        scenario_input_replay["scenario_input_status"].dropna().astype(str)
    )
    assert not scenario_input_replay["annual_period"].dropna().astype(str).isin(
        {"FY2051", "FY2052", "FY2053", "FY2054", "FY2055"}
    ).any()
    assert not ev_phev_split.empty
    assert ev_phev_split["used_by_current_finalist"].astype(bool).any()
    current_split = ev_phev_split[ev_phev_split["used_by_current_finalist"].astype(bool)].copy()
    assert set(current_split["allocation_status"].dropna().astype(str)) == {
        "legacy_split_audit_only_conventional_anchor_active"
    }
    assert set(current_split["source_path"].dropna().astype(str)) == {
        "Current finalist Base case",
        "Current finalist High population/comparison",
    }
    assert pd.to_numeric(current_split["current_allocation_residual_km"], errors="coerce").abs().max() > 0
    assert not ev_phev_drift.empty
    assert set(ev_phev_drift["lambda_mode"].dropna().astype(str)) == {
        "optimized",
        "fixed_light_only",
        "fixed_ped_only",
        "mbu_ratio",
    }
    required_drift_columns = {
        "smoothed_target_PED_light_petrol_km",
        "smoothed_target_conventional_light_km",
        "smoothed_target_BEV_km",
        "smoothed_target_PHEV_km",
        "smoothed_target_EV_total_km",
        "current_migration_revenue_total",
        "old_light_only_migration_revenue_total",
        "migration_revenue_delta",
    }
    assert required_drift_columns.issubset(ev_phev_drift.columns)
    optimized_drift = ev_phev_drift[
        ev_phev_drift["lambda_mode"].astype(str).eq("optimized")
        & ev_phev_drift["source_path"].astype(str).str.startswith("Current finalist")
    ].copy()
    assert not optimized_drift.empty
    lambda_values = pd.to_numeric(optimized_drift["lambda_value"], errors="coerce")
    lambda_lower = pd.to_numeric(optimized_drift["lambda_lower_bound"], errors="coerce")
    lambda_upper = pd.to_numeric(optimized_drift["lambda_upper_bound"], errors="coerce")
    assert lambda_values.between(0, 1).all()
    assert (lambda_values >= lambda_lower).all()
    assert (lambda_values <= lambda_upper).all()
    assert (
        optimized_drift.sort_values(["source_path", "FY"])
        .groupby("source_path")["lambda_value"]
        .apply(lambda values: pd.to_numeric(values, errors="coerce").diff().diff().abs().max())
        .max()
        < 0.01
    )
    assert pd.to_numeric(optimized_drift["component_sum_residual_km"], errors="coerce").abs().max() <= 1e-6
    for component in [
        "current_PED_light_petrol_km",
        "current_conventional_light_km",
        "current_BEV_km",
        "current_PHEV_km",
    ]:
        assert pd.to_numeric(optimized_drift[component], errors="coerce").ge(0).all()
    for smoothed_col, current_col in {
        "smoothed_target_PED_light_petrol_km": "current_PED_light_petrol_km",
        "smoothed_target_conventional_light_km": "current_conventional_light_km",
        "smoothed_target_BEV_km": "current_BEV_km",
        "smoothed_target_PHEV_km": "current_PHEV_km",
    }.items():
        assert pd.to_numeric(optimized_drift[smoothed_col], errors="coerce").to_numpy() == pytest.approx(
            pd.to_numeric(optimized_drift[current_col], errors="coerce").to_numpy()
        )
    assert pd.to_numeric(optimized_drift["smoothed_target_EV_total_km"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(optimized_drift["smoothed_target_BEV_km"], errors="coerce")
            + pd.to_numeric(optimized_drift["smoothed_target_PHEV_km"], errors="coerce")
        ).to_numpy()
    )
    assert pd.to_numeric(optimized_drift["current_migration_revenue_total"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(optimized_drift["current_PED_revenue"], errors="coerce")
            + pd.to_numeric(optimized_drift["current_light_ruc_net_revenue"], errors="coerce")
            + pd.to_numeric(optimized_drift["current_light_bev_ruc_net_revenue"], errors="coerce")
            + pd.to_numeric(optimized_drift["current_phev_ruc_net_revenue"], errors="coerce")
        ).to_numpy()
    )
    assert pd.to_numeric(optimized_drift["migration_revenue_delta"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(optimized_drift["current_migration_revenue_total"], errors="coerce")
            - pd.to_numeric(optimized_drift["old_light_only_migration_revenue_total"], errors="coerce")
        ).to_numpy()
    )
    assert not ped_bridge_audit.empty
    assert set(ped_bridge_audit["source_path"].dropna().astype(str)) == {
        "Current finalist Base case",
        "Current finalist High population/comparison",
    }
    assert int(pd.to_numeric(ped_bridge_audit["FY"], errors="coerce").min()) == 2026
    assert int(pd.to_numeric(ped_bridge_audit["FY"], errors="coerce").max()) == runtime_cutoff_fy
    required_ped_bridge_columns = {
        "scenario",
        "ped_vktpc_model",
        "ped_vkt_per_capita",
        "scenario_population",
        "population_million",
        "population_source_status",
        "population_fallback_flag",
        "raw_light_petrol_vkt",
        "raw_light_petrol_vkt_million_km",
        "adjusted_light_petrol_vkt_million_km",
        "optimized_light_petrol_vkt",
        "optimized_light_petrol_vkt_million_km",
        "optimization_delta",
        "optimization_delta_million_km",
        "base_litres_per_100km",
        "ped_volume_raw",
        "ped_volume_raw_million_litres",
        "ped_volume_optimized",
        "ped_volume_optimized_million_litres",
        "ped_volume_million_litres",
        "ped_rate",
        "ped_rate_nzd_per_litre",
        "gross_ped_revenue_raw",
        "gross_ped_revenue_raw_million_nzd",
        "gross_ped_revenue_optimized",
        "gross_ped_revenue_optimized_million_nzd",
        "gross_ped_revenue_million_nzd",
        "total_nltf_raw",
        "total_nltf_raw_million_nzd",
        "total_nltf_optimized",
        "total_nltf_optimized_million_nzd",
        "total_nltf_net_revenue_million_nzd",
        "official_light_petrol_vkt_million_km",
        "official_ped_volume_million_litres",
        "official_gross_ped_revenue_million_nzd",
        "vktpc_source_cell",
        "population_source_cell",
        "migration_source_cells",
        "formula",
    }
    assert required_ped_bridge_columns.issubset(ped_bridge_audit.columns)
    assert pd.to_numeric(ped_bridge_audit["population_million"], errors="coerce").gt(0).all()
    assert pd.to_numeric(ped_bridge_audit["base_litres_per_100km"], errors="coerce").gt(0).all()
    assert pd.to_numeric(ped_bridge_audit["ped_volume_million_litres"], errors="coerce").gt(0).all()
    assert pd.to_numeric(ped_bridge_audit["gross_ped_revenue_million_nzd"], errors="coerce").gt(0).all()
    assert pd.to_numeric(ped_bridge_audit["ped_vktpc_model"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(ped_bridge_audit["ped_vkt_per_capita"], errors="coerce").to_numpy()
    )
    assert pd.to_numeric(ped_bridge_audit["raw_light_petrol_vkt"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(ped_bridge_audit["raw_light_petrol_vkt_million_km"], errors="coerce").to_numpy()
    )
    assert pd.to_numeric(ped_bridge_audit["ped_volume_raw"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(ped_bridge_audit["ped_volume_raw_million_litres"], errors="coerce").to_numpy()
    )
    assert pd.to_numeric(ped_bridge_audit["raw_light_petrol_vkt_million_km"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(ped_bridge_audit["ped_vktpc_model"], errors="coerce")
            * pd.to_numeric(ped_bridge_audit["scenario_population"], errors="coerce")
            / 1_000_000.0
        ).to_numpy()
    )
    assert pd.to_numeric(ped_bridge_audit["ped_volume_raw_million_litres"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(ped_bridge_audit["raw_light_petrol_vkt_million_km"], errors="coerce")
            * pd.to_numeric(ped_bridge_audit["base_litres_per_100km"], errors="coerce")
            / 100.0
        ).to_numpy()
    )
    assert pd.to_numeric(ped_bridge_audit["ped_volume_optimized_million_litres"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(ped_bridge_audit["optimized_light_petrol_vkt_million_km"], errors="coerce")
            * pd.to_numeric(ped_bridge_audit["base_litres_per_100km"], errors="coerce")
            / 100.0
        ).to_numpy()
    )
    assert ped_bridge_audit["population_fallback_flag"].fillna(False).astype(bool).any()
    assert ped_bridge_audit.loc[
        ped_bridge_audit["population_fallback_flag"].fillna(False).astype(bool), "population_source_cell"
    ].astype(str).str.contains("population_proxy", regex=False).all()
    assert not ped_bridge_shape_fit.empty
    assert {"raw", "optimized"}.issubset(set(ped_bridge_shape_fit["bridge_variant"].astype(str)))
    base_light_fit = ped_bridge_shape_fit[
        ped_bridge_shape_fit["source_path"].astype(str).eq("Current finalist Base case")
        & ped_bridge_shape_fit["official_comparator_series_id"].astype(str).eq("light_petrol_vkt")
    ]
    raw_mae = float(base_light_fit.loc[base_light_fit["bridge_variant"].astype(str).eq("raw"), "mean_abs_error"].iloc[0])
    opt_mae = float(base_light_fit.loc[base_light_fit["bridge_variant"].astype(str).eq("optimized"), "mean_abs_error"].iloc[0])
    # Both variants must be measured and reported. Which one sits closer to
    # MBU26 is deliberately NOT asserted: MBU26 proximity is descriptive, not a
    # target, and "optimized" named a fit calibrated under the retired lambda
    # architecture. The governed decision is that the raw AR(1) bridge is the
    # production default, which is asserted directly below.
    assert np.isfinite(raw_mae) and raw_mae > 0
    assert np.isfinite(opt_mae) and opt_mae > 0
    assert PED_BRIDGE_DEFAULT_MODE == "raw_model"
    assert set(ped_bridge_mode_config["bridge_mode"].astype(str)) == {
        "raw_model",
        "blend_25",
        "blend_50",
        "blend_75",
        "optimized_migration",
        PED_BRIDGE_DEFAULT_MODE,
    }
    assert ped_bridge_mode_config.loc[
        ped_bridge_mode_config["bridge_mode"].astype(str).eq(PED_BRIDGE_DEFAULT_MODE), "default_selected"
    ].astype(bool).all()
    assert not ped_efficiency_scenarios.empty
    assert set(ped_efficiency_scenarios["scenario_id"].astype(str)) == {
        PED_EFFICIENCY_BASELINE_SCENARIO_ID,
        "efficiency_0_5pct_pa",
        "efficiency_1_0pct_pa",
        "efficiency_1_5pct_pa",
        "efficiency_2_0pct_pa",
    }
    assert ped_efficiency_scenarios["start_fy"].astype(int).eq(2026).all()
    assert ped_efficiency_scenarios["end_fy"].astype(int).eq(runtime_cutoff_fy).all()
    assert ped_efficiency_scenarios["notes"].astype(str).str.contains("does not change VKTpc forecasts", regex=False).all()
    assert not sensitivity_seed_inputs.empty
    assert set(sensitivity_seed_inputs["family"].astype(str)) == {
        "fleet_efficiency",
        "pt_mode_shift",
        "freight_rail_shift",
        "demand_elasticity",
    }
    assert set(sensitivity_seed_inputs["scenario_level"].astype(str)) == {"Low", "Med", "High"}
    workbook_seed_rows = sensitivity_seed_inputs[
        ~sensitivity_seed_inputs["family"].astype(str).eq("freight_rail_shift")
    ]
    assert workbook_seed_rows["workbook_sha256"].astype(str).eq(SENSITIVITY_SEED_WORKBOOK_SHA256).all()
    assert workbook_seed_rows["sheet"].astype(str).eq("Inputs (TI)").all()
    freight_seed_rows = sensitivity_seed_inputs[sensitivity_seed_inputs["family"].astype(str).eq("freight_rail_shift")]
    assert len(freight_seed_rows) == 3
    assert freight_seed_rows["stream"].astype(str).eq("HEAVY_RUC").all()
    assert freight_seed_rows["workbook_sha256"].astype(str).eq("").all()
    assert freight_seed_rows["source_status"].astype(str).eq("analyst_inference_no_official_mot_scenario").all()
    assert freight_seed_rows["note"].astype(str).str.contains("analyst inferences", case=False, regex=False).all()
    assert {"C181", "D181", "E181", "C206", "D206", "E206", "C213", "D213", "E213", "C266", "D266", "E266"}.issubset(
        set(sensitivity_seed_inputs["cell"].astype(str))
    )
    assert not sensitivity_seed_inputs.astype(str).stack().str.contains("C:\\Users", regex=False).any()
    assert not sensitivity_seed_inputs.astype(str).stack().str.contains("Downloads", regex=False).any()
    assert not sensitivity_config.empty
    assert set(sensitivity_config["family"].astype(str)) == {
        "fleet_efficiency",
        "pt_mode_shift",
        "freight_rail_shift",
        "demand_elasticity",
    }
    for family in ["fleet_efficiency", "pt_mode_shift", "freight_rail_shift", "demand_elasticity"]:
        assert {"Off", "Low", "Med", "High", "Custom"}.issubset(
            set(sensitivity_config.loc[sensitivity_config["family"].astype(str).eq(family), "selection"].astype(str))
        )
    assert sensitivity_config.loc[sensitivity_config["selection"].astype(str).eq("Off"), "default_selected"].astype(bool).all()
    freight_config = sensitivity_config[sensitivity_config["family"].astype(str).eq("freight_rail_shift")]
    assert freight_config["stream"].astype(str).eq("HEAVY_RUC").all()
    assert freight_config["start_fy"].astype(int).eq(2030).all()
    assert freight_config["custom_allowed"].astype(bool).all()
    assert freight_config["source_workbook_sha256"].astype(str).eq("").all()
    assert freight_config["notes"].astype(str).str.contains("analyst inferences", case=False, regex=False).all()
    assert (
        freight_config.set_index("selection")["value"].astype(float).loc[["Low", "Med", "High"]].tolist()
        == [0.0025, 0.005, 0.010]
    )
    assert not sensitivity_config.astype(str).stack().str.contains("crude-to-pump", case=False, regex=False).any()
    assert not sensitivity_config.astype(str).stack().str.contains("fleet transition target", case=False, regex=False).any()
    assert not sensitivity_impact_audit.empty
    assert set(sensitivity_impact_audit["selected_fleet_efficiency"].astype(str)) == {"Off"}
    assert set(sensitivity_impact_audit["selected_pt_mode_shift"].astype(str)) == {"Off"}
    assert set(sensitivity_impact_audit["selected_freight_rail_shift"].astype(str)) == {"Off"}
    assert set(sensitivity_impact_audit["selected_demand_elasticity"].astype(str)) == {"Off"}
    assert pd.to_numeric(sensitivity_impact_audit["delta"], errors="coerce").abs().max() == pytest.approx(0.0, abs=0)
    evidence = (
        ev_phev_split[
            pd.to_numeric(ev_phev_split["FY"], errors="coerce").isin([2024, 2025])
            & ev_phev_split["model_input_full_year"].astype(bool)
        ]
        .drop_duplicates("FY")
        .set_index("FY")
    )
    assert set(evidence["target_semantics_status"]) == {"matches_conventional_light_not_total_universe"}
    assert evidence["target_matches_conventional_light"].astype(bool).all()
    assert not evidence["target_matches_total_light_universe"].astype(bool).any()
    assert pd.to_numeric(evidence["target_minus_conventional_light_km"], errors="coerce").abs().max() == pytest.approx(0.0, abs=1e-9)
    assert pd.to_numeric(evidence["target_minus_total_light_universe_km"], errors="coerce").lt(0).all()

    allowed_traces = {
        "Actual",
        "BEFU26 official",
        "MBU26 official",
        "Current finalist Base case",
        "Current finalist High population/comparison",
        "Current finalist comparison behavioural path",
    }
    displayed = chart[chart["time_grain"].astype(str).eq("june_year") & chart["plot_allowed"].astype(str).str.lower().isin(["true", "1"])]
    assert set(displayed["trace_name"].dropna().unique()) == allowed_traces
    assert displayed[
        displayed["row_type"].astype(str).eq("historical_actual")
        & pd.to_numeric(displayed["june_year"], errors="coerce").gt(2025)
    ].empty
    dashboard_series = {str(value) for value in displayed["series_id"].dropna().unique()}
    assert dashboard_series == {
        "gross_fed_revenue",
        "gross_ped_revenue",
        "heavy_ruc_net_km",
        "heavy_ruc_net_revenue",
        "light_bev_ruc_net_km",
        "light_bev_ruc_net_revenue",
        "light_ruc_net_km",
        "light_ruc_net_revenue",
        "net_fed_revenue",
        "net_mvr_revenue",
        "ped_volume",
        "ped_vkt_per_capita",
        "phev_ruc_net_km",
        "phev_ruc_net_revenue",
        "total_fed_ruc_net_revenue",
        "total_nltf_net_revenue",
        "total_ruc_net_revenue",
    }
    assert "current_light_ruc_conventional_modelled_km" not in dashboard_series
    for series_id, series_rows in displayed.groupby("series_id"):
        traces = set(series_rows["trace_name"].dropna().astype(str))
        expected_current_trace = (
            "Current finalist comparison behavioural path"
            if str(series_id) == "ped_vkt_per_capita"
            else "Current finalist High population/comparison"
        )
        expected_traces = {
            "Actual",
            "MBU26 official",
            "Current finalist Base case",
            expected_current_trace,
        }
        assert expected_traces.issubset(traces), series_id
        actual_rows = series_rows[series_rows["trace_name"].astype(str).eq("Actual")]
        official_rows = series_rows[series_rows["trace_name"].astype(str).eq("MBU26 official")]
        assert not actual_rows.empty, series_id
        assert not official_rows.empty, series_id
        assert not actual_rows["source_file"].fillna("").astype(str).str.contains("forecast_scenario|annual_model_paths|selected_dashboard", case=False).any(), series_id
        assert not official_rows["source_file"].fillna("").astype(str).str.contains("forecast_scenario|annual_model_paths|selected_dashboard", case=False).any(), series_id
        official_years = set(pd.to_numeric(official_rows["june_year"], errors="coerce").dropna().astype(int))
        assert {2026, 2027}.issubset(official_years), series_id
        current_rows = series_rows[
            series_rows["trace_name"].astype(str).isin(
                [
                    "Current finalist Base case",
                    expected_current_trace,
                ]
            )
            & pd.to_numeric(series_rows["june_year"], errors="coerce").ge(2026)
        ]
        assert {
            "Current finalist Base case",
            expected_current_trace,
        }.issubset(set(current_rows["trace_name"].dropna().astype(str))), series_id
        assert not current_rows["source_file"].fillna("").astype(str).str.contains("annual_model_paths|selected_dashboard", case=False).any(), series_id
    assert "light_petrol_vkt_per_capita" not in set(chart["series_id"].dropna().astype(str))
    ped_displayed = displayed[displayed["series_id"].astype(str).eq("ped_vkt_per_capita")].copy()
    ped_by_trace = {
        trace: set(pd.to_numeric(group["june_year"], errors="coerce").dropna().astype(int))
        for trace, group in ped_displayed.groupby("trace_name")
    }
    assert 2025 in ped_by_trace["Actual"]
    assert {2026, 2027}.issubset(ped_by_trace["MBU26 official"])
    assert {2026, 2027}.issubset(ped_by_trace["Current finalist Base case"])
    assert "Current finalist High population/comparison" not in ped_by_trace
    assert {2026, 2027}.issubset(ped_by_trace["Current finalist comparison behavioural path"])
    official_ped_fy2026 = ped_displayed[
        ped_displayed["trace_name"].astype(str).eq("MBU26 official")
        & pd.to_numeric(ped_displayed["june_year"], errors="coerce").eq(2026)
    ].iloc[0]
    assert float(official_ped_fy2026["value"]) > 5000
    assert official_ped_fy2026["source_cell"] == "AB17"

    runtime_text = pd.concat(
        [
            chart[["source_file", "source", "model_basis"]].astype(str).stack(),
            bridge[["source", "source_basis", "model_id"]].astype(str).stack(),
            future[["source", "model_id"]].astype(str).stack(),
            stack_components[["source_file", "source_basis", "model_id"]].astype(str).stack(),
            scenario_role_contract.astype(str).stack(),
            scenario_feature_lineage.astype(str).stack(),
            scenario_input_replay.astype(str).stack(),
        ],
        ignore_index=True,
    ).str.cat(sep="\n")
    assert "annual_model_paths.csv" not in runtime_text
    assert "selected_dashboard" not in runtime_text.lower()
    assert "scenario_workbook_population_path_not_committed" not in runtime_text
    assert "source_workbook_cell_delta_unavailable" not in runtime_text
    assert "schiff" not in runtime_text.lower()
    assert "Official comparator: selected MOT/BEFU" not in runtime_text
    assert "Official comparator: rolling BEFU 1Y" not in runtime_text

    current = chart[
        chart["time_grain"].astype(str).eq("june_year")
        & chart["trace_role"].astype(str).eq("in_house_current_finalist")
        & chart["fed_path"].astype(str).eq("Current planned path")
    ].copy()
    fy2026 = current[current["period"].astype(str).eq("FY2026")].set_index(["series_id", "scenario_name"])
    assert fy2026.loc[("gross_ped_revenue", "current_basecase"), "model_id"] == "PED__VNEXT_SOLVED_CONVEX_TOP2"
    assert fy2026.loc[("light_ruc_net_revenue", "current_basecase"), "model_id"] == "dynamic_RESID_GBR_n150_d1_lr0.05_w36"
    assert "dynamic_RESID_GBR_n150_d1_lr0.05_w36" in fy2026.loc[("light_bev_ruc_net_revenue", "current_basecase"), "model_id"]
    assert "PED__VNEXT_SOLVED_CONVEX_TOP2" in fy2026.loc[("light_bev_ruc_net_revenue", "current_basecase"), "model_id"]
    assert "dynamic_RESID_GBR_n150_d1_lr0.05_w36" in fy2026.loc[("phev_ruc_net_revenue", "current_basecase"), "model_id"]
    assert "PED__VNEXT_SOLVED_CONVEX_TOP2" in fy2026.loc[("phev_ruc_net_revenue", "current_basecase"), "model_id"]
    assert fy2026.loc[("heavy_ruc_net_revenue", "current_basecase"), "model_id"] == "HEAVY_RUC__VNEXT_SOLVED_CONVEX_TOP4"
    assert "PED__VNEXT_SOLVED_CONVEX_TOP2" in fy2026.loc[("total_nltf_net_revenue", "current_basecase"), "model_id"]
    assert fy2026.loc[("total_nltf_net_revenue", "current_basecase"), "data_scope"] == "current_nowcast"
    assert fy2026.loc[("total_nltf_net_revenue", "current_basecase"), "actual_quarters"] == "2025Q3; 2025Q4"
    assert fy2026.loc[("total_nltf_net_revenue", "current_basecase"), "forecast_quarters"] == "2026Q1; 2026Q2"
    # Conventional-anchor correction. PED is now the raw AR(1) bridge with no
    # lambda migration subtraction, so gross PED revenue rises; Light BEV and
    # PHEV are allocated from the Base pool rather than added on. Every value
    # here is a PACK-STAGE (S0) figure - final displayed values additionally
    # carry Treasury macro and the policy response, and are recorded in
    # artifacts/p0_light_fleet_fix/gold_path_audit.csv. Tolerance unchanged at
    # 1e-6; no tolerance has been widened.
    # 2026Q1 actuals refresh re-freeze: FY2026 Light/Heavy quarters mix the
    # accepted 2026Q1 Core actuals (3 actual + 1 forecast), moving every
    # Light-dependent leaf; PED remains the strict 2025Q4-cutoff replay so the
    # PED/FED leaves are byte-stable:
    #   light_bev_ruc_net_revenue 84.492358 -> 83.429855
    #   total_nltf_net_revenue    4635.051597 -> 4628.169746  (-6.882, -0.149%)
    # BEFU26 bridge-vintage re-freeze: the annual bridge assumptions (PED
    # rate/intensity, RUC effective rates, admin, refunds, MVR, TUC and fixed
    # lines) now come from the BEFU26 official vintage instead of MBU26.
    # Activity forecasts and the Light BEV/PHEV class rates are unchanged;
    # the movement is fully attributed in
    # artifacts/official_vintage_befu26/current_bridge_vintage_impact.csv:
    #   gross_ped_revenue      2143.356964 -> 2145.034189  (+1.677)
    #   total_nltf_net_revenue 4628.169746 -> 4599.805745  (-28.364, -0.613%)
    # Tolerance unchanged at 1e-6; no tolerance has been widened.
    assert float(fy2026.loc[("gross_ped_revenue", "current_basecase"), "value"]) == pytest.approx(2145.0341891615885, abs=1e-6)
    assert float(fy2026.loc[("gross_fed_revenue", "current_basecase"), "value"]) == pytest.approx(2187.079351801779, abs=1e-6)
    assert float(fy2026.loc[("net_fed_revenue", "current_basecase"), "value"]) == pytest.approx(2114.6925455991745, abs=1e-6)
    assert float(fy2026.loc[("light_bev_ruc_net_revenue", "current_basecase"), "value"]) == pytest.approx(83.42985462699842, abs=1e-6)
    assert float(fy2026.loc[("phev_ruc_net_revenue", "current_basecase"), "value"]) == pytest.approx(22.364680858017962, abs=1e-6)
    assert float(fy2026.loc[("total_ruc_net_revenue", "current_basecase"), "value"]) == pytest.approx(2073.8323474108515, abs=1e-6)
    assert float(fy2026.loc[("total_nltf_net_revenue", "current_basecase"), "value"]) == pytest.approx(4599.805745017675, abs=1e-6)
    assert float(fy2026.loc[("gross_ped_revenue", "current_comparison_1"), "value"]) == pytest.approx(2147.3065370616973, abs=1e-6)
    assert float(fy2026.loc[("total_nltf_net_revenue", "current_comparison_1"), "value"]) == pytest.approx(4603.093811944454, abs=1e-6)

    anchor = current[current["period"].astype(str).eq("FY2025")].set_index(["series_id", "scenario_name"])
    assert anchor.loc[("total_nltf_net_revenue", "current_basecase"), "data_scope"] == "actual_anchor"
    # The FY2025 anchor is sourced from the bridge-assumption vintage pack
    # (BEFU26 generic stems) since the official-vintage framework.
    assert anchor.loc[("total_nltf_net_revenue", "current_basecase"), "source_file"] == "annual_spine.csv"

    assert set(bridge["bridge_status"].dropna().astype(str).unique()) == {"available"}
    replacements = bridge[bridge["component_type"].astype(str).eq("replacement_line")]
    replacement_streams = {
        "gross_ped_revenue",
        "light_ruc_net_revenue",
        "light_bev_ruc_net_revenue",
        "phev_ruc_net_revenue",
        "heavy_ruc_net_revenue",
    }
    assert set(replacements["stream"].unique()) == replacement_streams
    replacement_counts = replacements.groupby(["period", "scenario_name", "fed_path"])["stream"].agg(lambda values: set(values))
    assert replacement_counts.map(lambda values: values == replacement_streams).all()

    assert not audit.empty
    assert {2024, 2025, 2026, 2027}.issubset(set(pd.to_numeric(audit["june_year"], errors="coerce").dropna().astype(int)))
    assert {
        "series_id",
        "trace_name",
        "trace_type",
        "trace_role",
        "trace_source",
        "source_file",
        "source_cell",
        "formula",
        "model_id",
        "replacement_only",
        "actual_quarters",
        "forecast_quarters",
        "anchor_flag",
        "nowcast_flag",
    }.issubset(audit.columns)

    assert set(line_reconciliation["source_path"].dropna().unique()) == {
        "BEFU26 official",
        "MBU26 official",
        "Current finalist Base case",
        "Current finalist High population/comparison",
    }
    required_lines = {
        "Light RUC net km",
        "Heavy RUC net km",
        "Light BEV RUC net km",
        "Heavy BEV RUC net km",
        "PHEV RUC net km",
        "PED volume",
        "Light petrol VKT",
        "PED VKT per capita",
        "TUC GTK",
        "Light RUC net revenue",
        "Heavy RUC net revenue",
        "Light BEV RUC net revenue",
        "Heavy BEV RUC net revenue",
        "PHEV RUC net revenue",
        "RUC refunds",
        "Gross RUC",
        "RUC admin",
        "RUC net admin",
        "RUC net admin/refunds",
        "Gross PED",
        "LPG",
        "CNG",
        "Gross FED",
        "FED refunds",
        "Net FED",
        "MR1",
        "MR2",
        "MR13",
        "Gross MVR",
        "MVR admin",
        "MVR net admin & COO",
        "MVR refunds",
        "MVR net admin/refunds/COO",
        "TUC net revenue",
        "Total gross revenues",
        "Total admin fees",
        "Total revenues net of admin fees",
        "Total refunds",
        "Total NLTF revenue",
    }
    for source_path in ["MBU26 official", "Current finalist Base case", "Current finalist High population/comparison"]:
        path_rows = line_reconciliation[line_reconciliation["source_path"].astype(str).eq(source_path)]
        assert required_lines.issubset(set(path_rows["line_label"].astype(str)))
    for source_path in ["Current finalist Base case", "Current finalist High population/comparison"]:
        path_rows = line_reconciliation[line_reconciliation["source_path"].astype(str).eq(source_path)]
        assert "Current finalist Light RUC conventional modelled km" in set(path_rows["line_label"].astype(str))
    assert "light_petrol_vkt_per_capita" not in set(line_reconciliation["series_id"].dropna().astype(str))

    required_stack_cols = {
        "composition_mode",
        "source_path",
        "FY",
        "section",
        "line_label",
        "series_id",
        "value",
        "signed_contribution",
        "stack_value",
        "stack_total_by_FY",
        "overlay_total_value",
        "overlay_series_id",
        "overlay_label",
        "stack_overlay_residual",
        "stack_overlay_status",
        "unit",
        "row_role",
        "stack_role",
        "formula_role",
        "raw_value",
        "source_file",
        "source_cell",
        "formula",
        "stack_value_clean",
        "clean_stack_value",
        "chart_visible",
        "legend_visible",
        "net_effect_group",
        "clean_stack_total_by_FY",
        "clean_overlay_total_value",
        "clean_overlay_residual",
        "clean_overlay_status",
        "replacement_flag",
        "model_id",
        "quarter_composition",
        "actual_quarters",
        "forecast_quarters",
        "residual_vs_official",
        "stack_balance_residual",
        "formula_residual_status",
    }
    assert required_stack_cols.issubset(stack_components.columns)
    assert list(stack_components["composition_mode"].dropna().astype(str).drop_duplicates()) == [
        "Gross-to-net bridge audit",
        "Gross contribution stack",
    ]
    assert set(stack_components["source_path"].dropna().unique()) == {
        "BEFU26 official",
        "MBU26 official",
        "Current finalist Base case",
        "Current finalist High population/comparison",
    }
    required_stack_lines = set(required_lines).difference({"MR13"})
    required_stack_lines.update({"MR13/COO", "RUC refunds gross add-back", "MR13/COO gross add-back"})
    for source_path in ["MBU26 official", "Current finalist Base case", "Current finalist High population/comparison"]:
        path_rows = stack_components[stack_components["source_path"].astype(str).eq(source_path)]
        assert required_stack_lines.issubset(set(path_rows["line_label"].astype(str)))
        assert "Total RUC+PED" in set(path_rows["line_label"].astype(str))
    assert "light_petrol_vkt_per_capita" not in set(stack_components["series_id"].dropna().astype(str))
    aggregate_series = {
        "gross_ruc_revenue",
        "ruc_revenue_net_admin",
        "total_ruc_net_revenue",
        "total_fed_ruc_net_revenue",
        "gross_fed_revenue",
        "net_fed_revenue",
        "gross_mvr_revenue",
        "mvr_revenue_net_admin_coo",
        "net_mvr_revenue",
        "total_gross_revenue",
        "total_admin_fees",
        "total_revenue_net_admin",
        "total_refunds",
        "total_nltf_net_revenue",
    }
    assert set(
        stack_components.loc[
            stack_components["stack_role"].astype(str).eq("aggregate_overlay"),
            "series_id",
        ].astype(str)
    ).issuperset(aggregate_series)
    assert stack_components[
        stack_components["series_id"].astype(str).isin(aggregate_series)
        & stack_components["stack_role"].astype(str).isin(["component_positive", "component_negative", "offset_not_stacked"])
    ].empty
    assert stack_components.loc[
        stack_components["stack_role"].astype(str).eq("aggregate_overlay"),
        "stack_value",
    ].isna().all()
    fy2026_mbu26 = stack_components[
        stack_components["source_path"].astype(str).eq("MBU26 official")
        & stack_components["composition_mode"].astype(str).eq("Gross contribution stack")
        & pd.to_numeric(stack_components["FY"], errors="coerce").eq(2026)
    ]
    assert fy2026_mbu26.set_index("series_id").loc["gross_fed_revenue", "stack_role"] == "aggregate_overlay"
    assert fy2026_mbu26.set_index("series_id").loc["gross_ped_revenue", "stack_role"] == "component_positive"
    assert fy2026_mbu26.set_index("series_id").loc["gross_lpg_revenue", "stack_role"] == "component_positive"
    assert fy2026_mbu26.set_index("series_id").loc["gross_cng_revenue", "stack_role"] == "component_positive"
    assert fy2026_mbu26.set_index("series_id").loc["fed_refunds", "stack_role"] == "audit_context"
    assert fy2026_mbu26.set_index("series_id").loc["ruc_refunds", "stack_role"] == "component_positive"
    assert fy2026_mbu26.set_index("series_id").loc["coo_revenue", "stack_role"] == "component_positive"
    assert "offset_not_stacked" not in set(stack_components["stack_role"].dropna().astype(str))
    negative_rows = stack_components[stack_components["stack_role"].astype(str).eq("component_negative")].copy()
    assert not negative_rows.empty
    assert pd.to_numeric(negative_rows["signed_contribution"], errors="coerce").to_numpy() == pytest.approx(
        -pd.to_numeric(negative_rows["value"], errors="coerce").to_numpy()
    )
    bridge_offsets = stack_components[
        stack_components["composition_mode"].astype(str).eq("Gross-to-net bridge audit")
        & stack_components["series_id"].astype(str).isin(["coo_revenue", "ruc_refunds"])
    ]
    assert set(bridge_offsets["stack_role"].dropna().astype(str)) == {"component_negative"}
    assert pd.to_numeric(bridge_offsets["stack_value"], errors="coerce").to_numpy() == pytest.approx(
        -pd.to_numeric(bridge_offsets["value"], errors="coerce").to_numpy()
    )
    assert pd.to_numeric(bridge_offsets["clean_stack_value"], errors="coerce").to_numpy() == pytest.approx(0.0)
    assert not bridge_offsets["chart_visible"].fillna(True).astype(bool).any()
    assert not bridge_offsets["legend_visible"].fillna(True).astype(bool).any()
    bridge_addbacks = stack_components[
        stack_components["composition_mode"].astype(str).eq("Gross-to-net bridge audit")
        & stack_components["series_id"].astype(str).isin(["coo_gross_mvr_addback", "ruc_refunds_gross_addback"])
    ]
    assert set(bridge_addbacks["stack_role"].dropna().astype(str)) == {"component_positive"}
    assert set(bridge_addbacks["formula_role"].dropna().astype(str)) == {"gross_addback"}
    assert pd.to_numeric(bridge_addbacks["stack_value"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(bridge_addbacks["value"], errors="coerce").to_numpy()
    )
    assert pd.to_numeric(bridge_addbacks["clean_stack_value"], errors="coerce").to_numpy() == pytest.approx(0.0)
    assert not bridge_addbacks["chart_visible"].fillna(True).astype(bool).any()
    assert not bridge_addbacks["legend_visible"].fillna(True).astype(bool).any()
    assert set(
        stack_components.loc[
            stack_components["composition_mode"].astype(str).eq("Gross-to-net bridge audit")
            & stack_components["series_id"].astype(str).isin(["ruc_refunds", "ruc_refunds_gross_addback"]),
            "net_effect_group",
        ].dropna().astype(str)
    ) == {"ruc_refunds_internal_zero_net_pair"}
    assert set(
        stack_components.loc[
            stack_components["composition_mode"].astype(str).eq("Gross-to-net bridge audit")
            & stack_components["series_id"].astype(str).isin(["coo_revenue", "coo_gross_mvr_addback"]),
            "net_effect_group",
        ].dropna().astype(str)
    ) == {"mvr_mr13_coo_internal_zero_net_pair"}
    stack_residuals = stack_components[["source_path", "composition_mode", "FY", "stack_overlay_residual", "stack_overlay_status"]].drop_duplicates()
    assert set(stack_residuals["stack_overlay_status"].dropna().astype(str)) == {"balanced"}
    assert pd.to_numeric(stack_residuals["stack_overlay_residual"], errors="coerce").abs().max() <= 1.0
    component_sums = (
        stack_components[stack_components["stack_role"].isin(["component_positive", "component_negative"])]
        .groupby(["source_path", "composition_mode", "FY"])["stack_value"]
        .sum()
    )
    bridge_totals = stack_components[
        stack_components["composition_mode"].eq("Gross-to-net bridge audit")
        & stack_components["series_id"].eq("total_nltf_net_revenue")
    ].set_index(["source_path", "composition_mode", "FY"])["value"]
    gross_totals = stack_components[
        stack_components["composition_mode"].eq("Gross contribution stack")
        & stack_components["series_id"].eq("total_gross_revenue")
    ].set_index(["source_path", "composition_mode", "FY"])["value"]
    target_totals = pd.concat([bridge_totals, gross_totals])
    diff = pd.to_numeric(component_sums, errors="coerce") - pd.to_numeric(target_totals, errors="coerce")
    assert diff.abs().max() <= 1.0
    clean_component_sums = (
        stack_components[
            stack_components["stack_role"].isin(["component_positive", "component_negative"])
            & stack_components["chart_visible"].fillna(False).astype(bool)
        ]
        .groupby(["source_path", "composition_mode", "FY"])["clean_stack_value"]
        .sum()
    )
    clean_diff = pd.to_numeric(clean_component_sums, errors="coerce") - pd.to_numeric(target_totals, errors="coerce")
    assert clean_diff.abs().max() <= 1.0
    clean_status = stack_components[["source_path", "composition_mode", "FY", "clean_overlay_residual", "clean_overlay_status"]].drop_duplicates()
    assert set(clean_status["clean_overlay_status"].dropna().astype(str)) == {"balanced"}
    assert pd.to_numeric(clean_status["clean_overlay_residual"], errors="coerce").abs().max() <= 1.0
    overlay_targets = stack_components[["composition_mode", "overlay_series_id", "overlay_label"]].drop_duplicates()
    assert set(overlay_targets[overlay_targets["composition_mode"].eq("Gross-to-net bridge audit")]["overlay_series_id"]) == {"total_nltf_net_revenue"}
    assert set(overlay_targets[overlay_targets["composition_mode"].eq("Gross contribution stack")]["overlay_series_id"]) == {"total_gross_revenue"}
    current_stack = stack_components[stack_components["source_path"].astype(str).str.startswith("Current finalist")]
    current_pre_forecast = current_stack[
        pd.to_numeric(current_stack["FY"], errors="coerce").le(2025)
        & current_stack["row_role"].astype(str).isin(["leaf", "deduction", "replacement_line"])
    ]
    # Actual history is copied from the bridge-assumption vintage pack
    # (BEFU26 generic stems) since the official-vintage framework.
    assert set(current_pre_forecast["source_file"].dropna().astype(str)) == {"annual_spine.csv"}
    assert set(current_pre_forecast["source_basis"].dropna().astype(str)) == {"BEFU26 actual anchor"}
    assert not current_pre_forecast["source_file"].astype(str).str.contains("forecast_scenario", case=False).any()
    assert current_pre_forecast["forecast_quarters"].fillna("").astype(str).str.strip().eq("").all()
    current_fy2026_replacements = current_stack[
        pd.to_numeric(current_stack["FY"], errors="coerce").eq(2026)
        & current_stack["replacement_flag"].astype(str).str.lower().isin(["true", "1"])
    ]
    # Per-stream seam since the 2026Q1 actuals refresh: PED replacement lines
    # keep the 2-actual/2-forecast FY2026 mix (strict 2025Q4 accepted cutoff),
    # while Light/Heavy-derived lines carry the accepted 2026Q1 actual.
    assert set(current_fy2026_replacements["actual_quarters"].dropna().astype(str)) == {
        "2025Q3; 2025Q4",
        "2025Q3; 2025Q4; 2026Q1",
    }
    assert set(current_fy2026_replacements["forecast_quarters"].dropna().astype(str)) == {
        "2026Q1; 2026Q2",
        "2026Q2",
    }
    def _mix_values(series_id: str, column: str) -> set[str]:
        rows = current_fy2026_replacements[
            current_fy2026_replacements["series_id"].astype(str).eq(series_id)
        ]
        return set(rows[column].dropna().astype(str))

    assert _mix_values("gross_ped_revenue", "actual_quarters") == {"2025Q3; 2025Q4"}
    assert _mix_values("light_ruc_net_revenue", "actual_quarters") == {"2025Q3; 2025Q4; 2026Q1"}
    assert _mix_values("heavy_ruc_net_revenue", "forecast_quarters") == {"2026Q2"}
    current_replacements = set(
        current_stack.loc[
            current_stack["replacement_flag"].astype(str).str.lower().isin(["true", "1"]),
            "series_id",
        ].astype(str)
    )
    assert current_replacements == replacement_streams
    stack_text = stack_components.astype(str).to_csv(index=False)
    assert "C:\\Users" not in stack_text
    assert "Downloads" not in stack_text
    assert "OneDrive" not in stack_text
    assert ".xlsx" not in stack_text

    assert {
        "source_label",
        "source_series_id",
        "runtime_series_id",
        "dashboard_label",
        "unit",
        "source_row",
        "source_cell",
        "alias_reason",
        "status",
    }.issubset(alias_audit.columns)
    ped_alias = alias_audit[alias_audit["source_series_id"].astype(str).eq("light_petrol_vkt_per_capita")].iloc[0]
    assert ped_alias["runtime_series_id"] == "ped_vkt_per_capita"
    assert ped_alias["dashboard_label"] == "PED VKT per capita"
    assert ped_alias["status"] == "canonical_mapping"

    required_fan_availability_cols = {
        "series_id",
        "series_label",
        "fan_source",
        "available",
        "reason",
        "source_file",
        "model_id",
        "horizon_scope",
        "interpretation",
    }
    required_fan_band_cols = {
        "series_id",
        "fan_source",
        "scenario_name",
        "FY",
        "period",
        "central",
        "lower50",
        "upper50",
        "lower80",
        "upper80",
        "unit",
        "method",
        "source_file",
        "model_id",
    }
    assert required_fan_availability_cols.issubset(fan_availability.columns)
    assert required_fan_band_cols.issubset(fan_bands.columns)
    assert set(fan_availability["series_id"].dropna().astype(str)) == dashboard_series
    for series_id in ["ped_vkt_per_capita", "light_ruc_net_km", "heavy_ruc_net_km"]:
        rows = fan_availability[fan_availability["series_id"].astype(str).eq(series_id)]
        current_row = rows[rows["fan_source"].astype(str).eq(FAN_SOURCE_CURRENT_BACKTEST)].iloc[0]
        assert str(current_row["available"]).lower() in {"true", "1"}
        assert "annual_predictions.parquet" in current_row["source_file"]
        band_rows = fan_bands[
            fan_bands["series_id"].astype(str).eq(series_id)
            & fan_bands["fan_source"].astype(str).eq(FAN_SOURCE_CURRENT_BACKTEST)
        ]
        assert not band_rows.empty
        assert band_rows["method"].astype(str).eq("empirical_current_finalist_annual_backtest_error").all()
    for series_id in ["gross_ped_revenue", "light_ruc_net_revenue", "heavy_ruc_net_revenue"]:
        current_row = fan_availability[
            fan_availability["series_id"].astype(str).eq(series_id)
            & fan_availability["fan_source"].astype(str).eq(FAN_SOURCE_CURRENT_BACKTEST)
        ].iloc[0]
        assert str(current_row["available"]).lower() in {"true", "1"}
        assert "deterministic" in current_row["interpretation"].lower()
        assert "excludes" in current_row["interpretation"].lower()
    total_nltf_current = fan_availability[
        fan_availability["series_id"].astype(str).eq("total_nltf_net_revenue")
        & fan_availability["fan_source"].astype(str).eq(FAN_SOURCE_CURRENT_BACKTEST)
    ].iloc[0]
    assert str(total_nltf_current["available"]).lower() in {"false", "0"}
    assert "not been propagated" in total_nltf_current["reason"]
    assert fan_bands[
        fan_bands["series_id"].astype(str).eq("total_nltf_net_revenue")
        & fan_bands["fan_source"].astype(str).eq(FAN_SOURCE_CURRENT_BACKTEST)
    ].empty
    assert not fan_bands[fan_bands["fan_source"].astype(str).eq(FAN_SOURCE_SCENARIO_SPREAD)].empty
    assert fan_bands.loc[fan_bands["fan_source"].astype(str).eq(FAN_SOURCE_SCENARIO_SPREAD), "method"].astype(str).eq(
        "scenario_spread_not_probabilistic"
    ).all()
    assert not fan_bands.loc[fan_bands["fan_source"].astype(str).eq(FAN_SOURCE_SCENARIO_SPREAD), "method"].astype(str).str.contains(
        "probability|confidence", case=False
    ).any()
    assert not fan_bands[fan_bands["fan_source"].astype(str).eq(FAN_SOURCE_MBU26_ARCHIVED)].empty
    ped_mbu26 = fan_availability[
        fan_availability["series_id"].astype(str).eq("ped_vkt_per_capita")
        & fan_availability["fan_source"].astype(str).eq(FAN_SOURCE_MBU26_ARCHIVED)
    ].iloc[0]
    assert str(ped_mbu26["available"]).lower() in {"false", "0"}
    assert "not PED VKT per capita" in ped_mbu26["reason"]

    base_lines = line_reconciliation[
        line_reconciliation["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(line_reconciliation["FY"], errors="coerce").eq(2026)
    ].set_index("series_id")
    value = lambda series_id: float(base_lines.loc[series_id, "value"])
    drift_base_2026 = optimized_drift[
        optimized_drift["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(optimized_drift["FY"], errors="coerce").eq(2026)
    ].iloc[0]
    # The light classes and PED are no longer derived from the lambda drift
    # table. That table is retired audit evidence, so cross-checking the
    # decision-facing pack against it asserted the architecture this PR
    # removes. The pack must instead satisfy the conventional-anchor contract.
    conventional = value("light_ruc_net_km")
    assert value("current_light_ruc_conventional_modelled_km") == pytest.approx(conventional, abs=1e-9)
    shares, vfm_scenario = composition_shares(2026, repo_root=ROOT, uptake_basis="MoT VFM base")
    assert vfm_scenario == "Base_EV"
    pool = conventional / shares["conventional"]
    assert value("light_bev_ruc_net_km") == pytest.approx(pool * shares["bev"], abs=1e-6)
    assert value("phev_ruc_net_km") == pytest.approx(pool * shares["phev"], abs=1e-6)
    assert conventional + value("light_bev_ruc_net_km") + value("phev_ruc_net_km") == pytest.approx(
        pool, abs=1e-6
    )
    # Lambda independence is asserted directly in
    # tests/test_canonical_base_composition.py rather than inferred from an
    # incidental inequality against a retired audit table.
    assert value("gross_fed_revenue") == pytest.approx(value("gross_ped_revenue") + value("gross_lpg_revenue") + value("gross_cng_revenue"), abs=1e-9)
    assert value("net_fed_revenue") == pytest.approx(value("gross_fed_revenue") - value("fed_refunds"), abs=1e-9)
    assert value("gross_ruc_revenue") == pytest.approx(
        value("light_ruc_net_revenue")
        + value("heavy_ruc_net_revenue")
        + value("light_bev_ruc_net_revenue")
        + value("heavy_bev_ruc_net_revenue")
        + value("phev_ruc_net_revenue")
        + value("ruc_refunds"),
        abs=1e-9,
    )
    assert value("total_ruc_net_revenue") == pytest.approx(value("gross_ruc_revenue") - value("ruc_admin_revenue") - value("ruc_refunds"), abs=1e-9)
    assert value("total_nltf_net_revenue") == pytest.approx(value("total_revenue_net_admin") - value("total_refunds"), abs=1e-9)
    assert value("gross_ped_revenue") > 2000
    assert value("total_nltf_net_revenue") > 4500
    for series_id in ["gross_ped_revenue", "gross_fed_revenue", "net_fed_revenue", "total_ruc_net_revenue", "total_nltf_net_revenue"]:
        assert float(fy2026.loc[(series_id, "current_basecase"), "value"]) == pytest.approx(value(series_id), abs=1e-9)

    base_lines_2029 = line_reconciliation[
        line_reconciliation["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(line_reconciliation["FY"], errors="coerce").eq(2029)
    ].set_index("series_id")
    drift_base_2029 = optimized_drift[
        optimized_drift["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(optimized_drift["FY"], errors="coerce").eq(2029)
    ].iloc[0]
    fixed_light_base_2029 = ev_phev_drift[
        ev_phev_drift["source_path"].astype(str).eq("Current finalist Base case")
        & ev_phev_drift["lambda_mode"].astype(str).eq("fixed_light_only")
        & pd.to_numeric(ev_phev_drift["FY"], errors="coerce").eq(2029)
    ].iloc[0]
    assert float(drift_base_2029["weighted_sse"]) < float(fixed_light_base_2029["weighted_sse"])
    old_2029_migration_bundle = (
        float(drift_base_2029["old_light_only_PED_revenue"])
        + float(drift_base_2029["old_light_only_light_ruc_net_revenue"])
        + float(drift_base_2029["old_light_only_light_bev_ruc_net_revenue"])
        + float(drift_base_2029["old_light_only_phev_ruc_net_revenue"])
    )
    current_2029_migration_bundle = (
        float(drift_base_2029["current_PED_revenue"])
        + float(drift_base_2029["current_light_ruc_net_revenue"])
        + float(drift_base_2029["current_light_bev_ruc_net_revenue"])
        + float(drift_base_2029["current_phev_ruc_net_revenue"])
    )
    current_2029_total_nltf = float(base_lines_2029.loc["total_nltf_net_revenue", "value"])
    light_only_2029_total_nltf = current_2029_total_nltf - current_2029_migration_bundle + old_2029_migration_bundle
    assert current_2029_migration_bundle != pytest.approx(old_2029_migration_bundle, abs=1e-6)
    assert current_2029_total_nltf != pytest.approx(light_only_2029_total_nltf, abs=1e-6)
    assert current_2029_total_nltf - light_only_2029_total_nltf == pytest.approx(
        current_2029_migration_bundle - old_2029_migration_bundle,
        abs=1e-6,
    )

    base_cutoff_line = line_reconciliation[
        line_reconciliation["source_path"].astype(str).eq("Current finalist Base case")
        & line_reconciliation["series_id"].astype(str).eq("current_light_ruc_conventional_modelled_km")
        & pd.to_numeric(line_reconciliation["FY"], errors="coerce").between(2046, runtime_cutoff_fy, inclusive="both")
    ].copy()
    base_cutoff_line["FY_numeric"] = pd.to_numeric(base_cutoff_line["FY"], errors="coerce").astype(int)
    assert set(base_cutoff_line["FY_numeric"]) == set(range(2046, runtime_cutoff_fy + 1))
    assert base_cutoff_line["value_status"].astype(str).ne("extrapolated_model_extension").all()
    # The econometric segment still ends at the runtime cutoff; the labelled
    # post-model segment carries FY2031-FY2050.
    current_line_rows = line_reconciliation[
        line_reconciliation["source_path"].astype(str).str.startswith("Current finalist")
    ]
    econometric_line = current_line_rows[
        ~current_line_rows["forecast_segment"].fillna("").astype(str).eq("post_model_extrapolation")
    ]
    assert pd.to_numeric(econometric_line["FY"], errors="coerce").max() == runtime_cutoff_fy
    post_model_line = current_line_rows[
        current_line_rows["forecast_segment"].fillna("").astype(str).eq("post_model_extrapolation")
    ]
    assert pd.to_numeric(post_model_line["FY"], errors="coerce").max() == LAST_POST_MODEL_FY

    current_residuals = residuals[
        residuals["source_path"].astype(str).str.startswith("Current finalist")
        & residuals["output_series_id"].isin(["gross_fed_revenue", "net_fed_revenue", "total_ruc_net_revenue", "total_nltf_net_revenue"])
    ]
    assert set(current_residuals["status"].dropna().unique()) == {"reconciled"}


# Everything the PED bridge is allowed to move, and everything it must not.
# The conventional anchor is the whole point of the correction: a PED-side
# overlay may never reach Light RUC, its class split, or Heavy RUC.
_PED_CHAIN_SERIES = {
    "light_petrol_vkt",
    "ped_volume",
    "ped_vkt_per_capita",
    "gross_ped_revenue",
    "gross_fed_revenue",
    "net_fed_revenue",
    "total_gross_revenue",
    "total_revenue_net_admin",
    "total_fed_ruc_net_revenue",
    "total_nltf_net_revenue",
}
_LIGHT_RUC_FAMILY = {
    CONVENTIONAL_ANCHOR_SERIES_ID,
    "light_ruc_net_km",
    "light_ruc_net_revenue",
    "light_bev_ruc_net_km",
    "light_bev_ruc_net_revenue",
    "phev_ruc_net_km",
    "phev_ruc_net_revenue",
    "heavy_ruc_net_km",
    "heavy_ruc_net_revenue",
    "total_ruc_net_revenue",
}


def _assert_movement_confined_to_ped_chain(
    adjusted: pd.DataFrame,
    original: pd.DataFrame,
    value_column: str,
) -> None:
    """A PED-side overlay may move the PED chain and nothing else."""
    moved = (
        pd.to_numeric(adjusted[value_column], errors="coerce")
        .sub(pd.to_numeric(original[value_column], errors="coerce"))
        .abs()
        .gt(1e-9)
        .fillna(False)
        .to_numpy()
    )
    if not moved.any():
        return
    # line_reconciliation labels rows as series_id; bridge_components and
    # future_revenue_forecasts carry the same series names under "stream".
    label_column = "series_id" if "series_id" in adjusted.columns else "stream"
    moved_labels = set(adjusted.loc[moved, label_column].astype(str))
    assert moved_labels <= _PED_CHAIN_SERIES, (
        f"PED overlay moved {sorted(moved_labels - _PED_CHAIN_SERIES)}"
    )
    assert not moved_labels & _LIGHT_RUC_FAMILY, (
        "the conventional anchor must survive a PED-side overlay untouched"
    )


def test_ped_bridge_modes_materialize_raw_optimized_and_reconcile() -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None
    audit = pack.ped_revenue_bridge_audit
    assert not audit.empty
    required = {
        "ped_vktpc_model",
        "raw_light_petrol_vkt",
        "raw_light_petrol_vkt_million_km",
        "optimized_light_petrol_vkt",
        "optimized_light_petrol_vkt_million_km",
        "optimization_delta",
        "optimization_delta_million_km",
        "ped_volume_raw",
        "ped_volume_raw_million_litres",
        "ped_volume_optimized",
        "ped_volume_optimized_million_litres",
        "gross_ped_revenue_raw",
        "gross_ped_revenue_raw_million_nzd",
        "gross_ped_revenue_optimized",
        "gross_ped_revenue_optimized_million_nzd",
    }
    assert required.issubset(audit.columns)
    current_base = audit[audit["source_path"].astype(str).eq("Current finalist Base case")].copy()
    assert pd.to_numeric(current_base["optimization_delta_million_km"], errors="coerce").abs().max() > 100
    assert pd.to_numeric(current_base["raw_light_petrol_vkt_million_km"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(current_base["ped_vktpc_model"], errors="coerce")
            * pd.to_numeric(current_base["scenario_population"], errors="coerce")
            / 1_000_000.0
        ).to_numpy()
    )
    assert pd.to_numeric(current_base["ped_volume_raw_million_litres"], errors="coerce").to_numpy() == pytest.approx(
        (
            pd.to_numeric(current_base["raw_light_petrol_vkt_million_km"], errors="coerce")
            * pd.to_numeric(current_base["base_litres_per_100km"], errors="coerce")
            / 100.0
        ).to_numpy()
    )
    assert audit["population_fallback_flag"].fillna(False).astype(bool).any()
    assert not audit.astype(str).stack().str.contains("C:\\Users", regex=False).any()
    assert not audit.astype(str).stack().str.contains("Downloads", regex=False).any()

    default = apply_ped_bridge_mode_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        bridge_mode=PED_BRIDGE_DEFAULT_MODE,
    )
    explicit_raw = apply_ped_bridge_mode_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        bridge_mode="raw_model",
    )
    for key, value_column in [
        ("chart_rows", "value"),
        ("line_reconciliation", "value"),
        ("revenue_bridge_components", "component_value"),
        ("future_revenue_forecasts", "revenue_forecast_nzd"),
    ]:
        assert pd.to_numeric(default[key][value_column], errors="coerce").to_numpy() == pytest.approx(
            pd.to_numeric(explicit_raw[key][value_column], errors="coerce").to_numpy(),
            abs=0,
        )
    assert pd.to_numeric(
        default["ped_revenue_bridge_audit"]["ped_volume_million_litres"], errors="coerce"
    ).to_numpy() == pytest.approx(
        pd.to_numeric(default["ped_revenue_bridge_audit"]["ped_volume_raw_million_litres"], errors="coerce").to_numpy(),
        abs=1e-9,
    )
    default_petrol = default["chart_rows"][
        default["chart_rows"]["series_id"].astype(str).eq("light_petrol_vkt")
        & default["chart_rows"]["time_grain"].astype(str).eq("june_year")
        & default["chart_rows"]["trace_role"].astype(str).eq(
            "in_house_current_finalist"
        )
    ].copy()
    assert not default_petrol.empty
    assert not default_petrol.duplicated(
        ["trace_name", "scenario_name", "fed_path", "june_year"]
    ).any()
    raw_petrol = default["ped_revenue_bridge_audit"][
        default["ped_revenue_bridge_audit"]["source_path"]
        .astype(str)
        .str.startswith("Current finalist")
    ][
        [
            "source_path",
            "scenario_name",
            "fed_path",
            "FY",
            "raw_light_petrol_vkt_million_km",
        ]
    ].copy()
    aligned = default_petrol.merge(
        raw_petrol,
        left_on=["trace_name", "scenario_name", "fed_path", "june_year"],
        right_on=["source_path", "scenario_name", "fed_path", "FY"],
        how="left",
        validate="one_to_one",
    )
    assert aligned["raw_light_petrol_vkt_million_km"].notna().all()
    assert pd.to_numeric(aligned["value"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(
            aligned["raw_light_petrol_vkt_million_km"], errors="coerce"
        ).to_numpy(),
        abs=1e-9,
    )

    optimized = apply_ped_bridge_mode_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        bridge_mode="optimized_migration",
    )
    # The optimized bridge changes light-petrol activity, so everything
    # downstream of PED volume moves with it. Blanket equality outside
    # light_petrol_vkt encoded the retired lambda transfer, under which the
    # bridge moved a migration total shared with Light RUC. PED and Light RUC
    # are now independent, so the meaningful assertion is which family moves -
    # not that nothing else does.
    for key, value_column, original in [
        ("chart_rows", "value", pack.revenue_chart_rows),
        ("line_reconciliation", "value", pack.revenue_line_reconciliation),
        ("revenue_bridge_components", "component_value", pack.revenue_bridge_components),
        ("future_revenue_forecasts", "revenue_forecast_nzd", pack.future_revenue_forecasts),
    ]:
        _assert_movement_confined_to_ped_chain(optimized[key], original, value_column)
    optimized_petrol = optimized["chart_rows"][
        optimized["chart_rows"]["series_id"].astype(str).eq("light_petrol_vkt")
    ]
    assert not optimized_petrol.empty

    for mode in ["raw_model", "blend_25", "blend_50", "blend_75", "optimized_migration", PED_BRIDGE_DEFAULT_MODE]:
        result = apply_ped_bridge_mode_layer(
            chart_rows=pack.revenue_chart_rows,
            line_reconciliation=pack.revenue_line_reconciliation,
            bridge_components=pack.revenue_bridge_components,
            future_revenue_forecasts=pack.future_revenue_forecasts,
            ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
            bridge_mode=mode,
        )
        mode_audit = result["ped_revenue_bridge_audit"]
        assert pd.to_numeric(mode_audit["ped_volume_million_litres"], errors="coerce").to_numpy() == pytest.approx(
            (
                pd.to_numeric(mode_audit["adjusted_light_petrol_vkt_million_km"], errors="coerce")
                * pd.to_numeric(mode_audit["base_litres_per_100km"], errors="coerce")
                / 100.0
            ).to_numpy()
        )
        current_residuals = result["revenue_formula_residuals"][
            result["revenue_formula_residuals"]["source_path"].astype(str).str.startswith("Current finalist")
        ]
        assert set(current_residuals["status"].dropna().astype(str)) == {"reconciled"}
        official_original = pack.revenue_chart_rows[
            pack.revenue_chart_rows["trace_role"].astype(str).eq("official_external_comparator")
        ].copy()
        official_adjusted = result["chart_rows"][
            result["chart_rows"]["trace_role"].astype(str).eq("official_external_comparator")
        ].copy()
        assert pd.to_numeric(official_adjusted["value"], errors="coerce").to_numpy() == pytest.approx(
            pd.to_numeric(official_original["value"], errors="coerce").to_numpy(),
            abs=0,
        )
        current_formula_rows = result["chart_rows"][
            result["chart_rows"]["time_grain"].astype(str).eq("june_year")
            & result["chart_rows"]["trace_role"].astype(str).eq("in_house_current_finalist")
            & result["chart_rows"]["series_id"].astype(str).isin(
                [
                    "net_fed_revenue",
                    "total_ruc_net_revenue",
                    "total_fed_ruc_net_revenue",
                ]
            )
        ].pivot_table(
            index=["trace_name", "scenario_name", "fed_path", "june_year"],
            columns="series_id",
            values="value",
            aggfunc="first",
        )
        formula_residual = (
            pd.to_numeric(current_formula_rows["total_fed_ruc_net_revenue"], errors="coerce")
            - pd.to_numeric(current_formula_rows["net_fed_revenue"], errors="coerce")
            - pd.to_numeric(current_formula_rows["total_ruc_net_revenue"], errors="coerce")
        )
        assert formula_residual.abs().max() == pytest.approx(0.0, abs=1e-9)


def test_ped_efficiency_sensitivity_noops_baseline_and_reconciles_rollups() -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    baseline = apply_ped_efficiency_sensitivity(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        ped_efficiency_scenarios=pack.ped_efficiency_scenarios,
        scenario_id=PED_EFFICIENCY_BASELINE_SCENARIO_ID,
    )
    # Same contract as the bridge-mode case: a PED-side overlay may move the
    # PED chain only, and the baseline scenario should move nothing at all.
    for key, value_column, original in [
        ("chart_rows", "value", pack.revenue_chart_rows),
        ("line_reconciliation", "value", pack.revenue_line_reconciliation),
        ("revenue_bridge_components", "component_value", pack.revenue_bridge_components),
        ("future_revenue_forecasts", "revenue_forecast_nzd", pack.future_revenue_forecasts),
    ]:
        _assert_movement_confined_to_ped_chain(baseline[key], original, value_column)

    sensitivity = apply_ped_efficiency_sensitivity(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        ped_efficiency_scenarios=pack.ped_efficiency_scenarios,
        scenario_id="efficiency_1_0pct_pa",
    )
    adjustment = sensitivity["ped_efficiency_adjustment"]
    assert not adjustment.empty
    assert adjustment["efficiency_label"].astype(str).eq("1.0% p.a.").all()
    assert pd.to_numeric(adjustment["adjusted_litres_per_100km"], errors="coerce").lt(
        pd.to_numeric(adjustment["base_litres_per_100km"], errors="coerce")
    ).all()
    assert pd.to_numeric(adjustment["adjusted_ped_volume_million_litres"], errors="coerce").lt(
        pd.to_numeric(adjustment["baseline_ped_volume_million_litres"], errors="coerce")
    ).all()
    assert pd.to_numeric(adjustment["adjusted_gross_ped_revenue_million_nzd"], errors="coerce").lt(
        pd.to_numeric(adjustment["baseline_gross_ped_revenue_million_nzd"], errors="coerce")
    ).all()
    assert pd.to_numeric(adjustment["total_nltf_net_revenue_delta_million_nzd"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(adjustment["gross_ped_revenue_delta_million_nzd"], errors="coerce").to_numpy()
    )
    current_residuals = sensitivity["revenue_formula_residuals"][
        sensitivity["revenue_formula_residuals"]["source_path"].astype(str).str.startswith("Current finalist")
    ]
    assert set(current_residuals["status"].dropna().astype(str)) == {"reconciled"}

    adjusted_chart = sensitivity["chart_rows"]
    official_original = pack.revenue_chart_rows[
        pack.revenue_chart_rows["trace_role"].astype(str).eq("official_external_comparator")
    ].copy()
    official_adjusted = adjusted_chart[
        adjusted_chart["trace_role"].astype(str).eq("official_external_comparator")
    ].copy()
    assert pd.to_numeric(official_adjusted["value"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(official_original["value"], errors="coerce").to_numpy(),
        abs=0,
    )
    unchanged_ev_phev = pack.revenue_chart_rows[
        pack.revenue_chart_rows["trace_role"].astype(str).eq("in_house_current_finalist")
        & pack.revenue_chart_rows["series_id"].astype(str).isin(
            ["light_bev_ruc_net_km", "phev_ruc_net_km", "light_bev_ruc_net_revenue", "phev_ruc_net_revenue"]
        )
    ].copy()
    adjusted_ev_phev = adjusted_chart.loc[unchanged_ev_phev.index]
    assert pd.to_numeric(adjusted_ev_phev["value"], errors="coerce").to_numpy() == pytest.approx(
        pd.to_numeric(unchanged_ev_phev["value"], errors="coerce").to_numpy(),
        abs=0,
    )


def test_revenue_sensitivity_layer_off_preserves_runtime_values() -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    baseline = apply_revenue_sensitivity_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        sensitivity_config=pack.sensitivity_config,
    )
    # Same contract as the bridge-mode case: a PED-side overlay may move the
    # PED chain only, and the baseline scenario should move nothing at all.
    for key, value_column, original in [
        ("chart_rows", "value", pack.revenue_chart_rows),
        ("line_reconciliation", "value", pack.revenue_line_reconciliation),
        ("revenue_bridge_components", "component_value", pack.revenue_bridge_components),
        ("future_revenue_forecasts", "revenue_forecast_nzd", pack.future_revenue_forecasts),
    ]:
        _assert_movement_confined_to_ped_chain(baseline[key], original, value_column)
    assert pd.to_numeric(baseline["sensitivity_impact_audit"]["delta"], errors="coerce").abs().max() == pytest.approx(0.0, abs=0)


def test_revenue_sensitivity_layer_off_reuses_formula_residuals(monkeypatch) -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    calls = 0
    original = revenue_outlook_module.revenue_formula_residual_frame

    def counted_formula_residuals(line_reconciliation: pd.DataFrame) -> pd.DataFrame:
        nonlocal calls
        calls += 1
        return original(line_reconciliation)

    monkeypatch.setattr(revenue_outlook_module, "revenue_formula_residual_frame", counted_formula_residuals)
    baseline = revenue_outlook_module.apply_revenue_sensitivity_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        sensitivity_config=pack.sensitivity_config,
    )

    assert calls == 1
    assert len(baseline["revenue_formula_residuals"]) == len(pack.revenue_formula_residuals)
    assert not baseline["revenue_stack_components"].empty


def test_revenue_sensitivity_efficiency_lowers_ped_revenue_holding_vkt_fixed() -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    sensitivity = apply_revenue_sensitivity_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        sensitivity_config=pack.sensitivity_config,
        fleet_efficiency="Med",
    )
    audit = sensitivity["sensitivity_impact_audit"]
    rows = audit[
        audit["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(audit["FY"], errors="coerce").eq(2029)
    ].set_index("series_id")
    assert rows.loc["light_petrol_vkt", "adjusted"] == pytest.approx(rows.loc["light_petrol_vkt", "baseline"], abs=0)
    assert rows.loc["ped_vkt_per_capita", "adjusted"] == pytest.approx(rows.loc["ped_vkt_per_capita", "baseline"], abs=0)
    assert rows.loc["ped_volume", "adjusted"] < rows.loc["ped_volume", "baseline"]
    assert rows.loc["gross_ped_revenue", "adjusted"] < rows.loc["gross_ped_revenue", "baseline"]
    assert rows.loc["light_ruc_net_km", "adjusted"] == pytest.approx(rows.loc["light_ruc_net_km", "baseline"], abs=0)
    current_residuals = sensitivity["revenue_formula_residuals"][
        sensitivity["revenue_formula_residuals"]["source_path"].astype(str).str.startswith("Current finalist")
    ]
    assert set(current_residuals["status"].dropna().astype(str)) == {"reconciled"}


def test_revenue_sensitivity_pt_shift_preserves_ev_phev_shares() -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    sensitivity = apply_revenue_sensitivity_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        sensitivity_config=pack.sensitivity_config,
        pt_mode_shift="Med",
    )
    audit = sensitivity["sensitivity_impact_audit"]
    rows = audit[
        audit["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(audit["FY"], errors="coerce").eq(LAST_DECISION_GRADE_ANNUAL_FY)
    ].set_index("series_id")
    expected_factor = (1 - 0.005) ** (LAST_DECISION_GRADE_ANNUAL_FY - 2030 + 1)
    for series_id in ["light_petrol_vkt", "light_ruc_net_km", "light_bev_ruc_net_km", "phev_ruc_net_km"]:
        assert rows.loc[series_id, "adjusted"] == pytest.approx(rows.loc[series_id, "baseline"] * expected_factor)
    baseline_total = rows.loc["light_ruc_net_km", "baseline"] + rows.loc["light_bev_ruc_net_km", "baseline"] + rows.loc["phev_ruc_net_km", "baseline"]
    adjusted_total = rows.loc["light_ruc_net_km", "adjusted"] + rows.loc["light_bev_ruc_net_km", "adjusted"] + rows.loc["phev_ruc_net_km", "adjusted"]
    for series_id in ["light_ruc_net_km", "light_bev_ruc_net_km", "phev_ruc_net_km"]:
        assert rows.loc[series_id, "adjusted"] / adjusted_total == pytest.approx(rows.loc[series_id, "baseline"] / baseline_total)
    assert rows.loc["heavy_ruc_net_km", "adjusted"] == pytest.approx(rows.loc["heavy_ruc_net_km", "baseline"], abs=0)


def test_revenue_sensitivity_freight_rail_shift_scales_heavy_ruc_only() -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    sensitivity = apply_revenue_sensitivity_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        sensitivity_config=pack.sensitivity_config,
        freight_rail_shift="Med",
    )
    audit = sensitivity["sensitivity_impact_audit"]
    assert set(audit["selected_freight_rail_shift"].astype(str)) == {"Med"}
    rows = audit[
        audit["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(audit["FY"], errors="coerce").eq(LAST_DECISION_GRADE_ANNUAL_FY)
    ].set_index("series_id")
    expected_factor = (1 - 0.005) ** (LAST_DECISION_GRADE_ANNUAL_FY - 2030 + 1)
    for series_id in ["heavy_ruc_net_km", "heavy_ruc_net_revenue"]:
        assert rows.loc[series_id, "adjusted"] == pytest.approx(rows.loc[series_id, "baseline"] * expected_factor)
    for series_id in ["light_petrol_vkt", "light_ruc_net_km", "light_bev_ruc_net_km", "phev_ruc_net_km", "ped_volume", "gross_ped_revenue"]:
        assert rows.loc[series_id, "adjusted"] == pytest.approx(rows.loc[series_id, "baseline"], abs=0)
    heavy_delta = rows.loc["heavy_ruc_net_revenue", "adjusted"] - rows.loc["heavy_ruc_net_revenue", "baseline"]
    assert heavy_delta < 0
    for rollup_id in ["total_ruc_net_revenue", "total_nltf_net_revenue"]:
        assert rows.loc[rollup_id, "adjusted"] - rows.loc[rollup_id, "baseline"] == pytest.approx(heavy_delta)
    pre_start = audit[
        audit["source_path"].astype(str).eq("Current finalist Base case")
        & pd.to_numeric(audit["FY"], errors="coerce").eq(2029)
    ].set_index("series_id")
    assert pre_start.loc["heavy_ruc_net_km", "adjusted"] == pytest.approx(pre_start.loc["heavy_ruc_net_km", "baseline"], abs=0)
    current_residuals = sensitivity["revenue_formula_residuals"][
        sensitivity["revenue_formula_residuals"]["source_path"].astype(str).str.startswith("Current finalist")
    ]
    assert set(current_residuals["status"].dropna().astype(str)) == {"reconciled"}


def test_revenue_sensitivity_demand_elasticity_responds_to_cost_ratio() -> None:
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    lower_cost = apply_revenue_sensitivity_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        sensitivity_config=pack.sensitivity_config,
        demand_elasticity="Med",
        cost_per_km_ratio=0.9,
    )["sensitivity_impact_audit"]
    higher_cost = apply_revenue_sensitivity_layer(
        chart_rows=pack.revenue_chart_rows,
        line_reconciliation=pack.revenue_line_reconciliation,
        bridge_components=pack.revenue_bridge_components,
        future_revenue_forecasts=pack.future_revenue_forecasts,
        ped_revenue_bridge_audit=pack.ped_revenue_bridge_audit,
        sensitivity_config=pack.sensitivity_config,
        demand_elasticity="Med",
        cost_per_km_ratio=1.1,
    )["sensitivity_impact_audit"]
    for audit, relation in [(lower_cost, "gt"), (higher_cost, "lt")]:
        rows = audit[
            audit["source_path"].astype(str).eq("Current finalist Base case")
            & pd.to_numeric(audit["FY"], errors="coerce").eq(2029)
        ].set_index("series_id")
        for series_id in ["light_petrol_vkt", "light_ruc_net_km", "heavy_ruc_net_km"]:
            if relation == "gt":
                assert rows.loc[series_id, "adjusted"] > rows.loc[series_id, "baseline"]
            else:
                assert rows.loc[series_id, "adjusted"] < rows.loc[series_id, "baseline"]


def test_current_revenue_outlook_runtime_artifact_hashes_are_frozen() -> None:
    pack_dir = ROOT / CURRENT_REVENUE_OUTLOOK_DIR
    expected_hashes = {
        'conflict_fuel_price_scenarios.csv': 'ad379997aa4044cdabf7d948787c926e06a434447ff076640cfab317eda53c73',
        'conflict_gdp_calibration.csv': '032606a84ed1e7716197ced405d3026b39c8c00e0a56882174b45c13eb8fb655',
        'ev_phev_ped_light_drift_assumptions.csv': '07e1cfb9c6279a3b9a84fcfe772096b4213694880c496ce1eedac7231c8a8984',
        'ev_phev_ped_light_drift_assumptions.parquet': '015ec183ddcfb8fe4adc17bb3b60f8ef6c357f8c9f51672fa5924399ca5cb9e4',
        'ev_phev_split_assumptions.csv': '01cdc3b0491b03e9ac8513058ddd87797797765d0ab65f818144d7b7df56ffa0',
        'ev_phev_split_assumptions.parquet': '3e76a9493aa6d5b03bd3f336ec9ccd8156f571af1da47da9424f5eb8c176a963',
        'fan_availability.csv': '9c89fa4263e8b07b1e3c8675b5f8f4d23d3668dece4fb89d1460559b2fbf83e5',
        'fan_availability.parquet': '2bab0334fd9a7350e40840b4697ae71d259263f781e2332bc8103690b98a164d',
        'fan_band_rows.csv': '24e47a51b316e125a4111b5499f65731c808fc31e0d605ef80bee7cb66c9ad36',
        'fan_band_rows.parquet': '8b2aac6445c37ff9d51af18260093c9c8d9a6e89b6abf60f31652ca9283df53b',
        'future_revenue_forecasts.csv': '93ce13e7b48a860ef1d528fe673ff06b579752d7f8d9378bc0f72b042f76d618',
        'future_revenue_forecasts.parquet': '403c0adcbc5ede8566358b7622aea6e181527cbda7b6dbda7bf042d0b59f3576',
        'horizon_contract_audit.csv': 'b6e117b5656c479275986187c56b3eafbe5ef2362313c7d6f9a418359ea667f7',
        'horizon_contract_audit.parquet': '015e1a42d88eaafe2fa966800a5b499f2b90b8331a6a062b81e76d080d1c5882',
        'light_ruc_horizon_availability.csv': 'ca87c6ac93033496a928ae07159824dac642917ba104c030161dab5ff46b063f',
        'light_ruc_horizon_availability.parquet': '39abda810443b1d91e04b26399a96c0c30e2fe70a30cbde7440790c3145f5f48',
        'manifest.json': '072598d1557c7fbce56645817cdff03e3baccc48453bf0d639394a3368f7b04d',
        'manifest.md': 'f021e551cf021f8be307e7ab0088e8857e40eced50375bd6d3e55e2093eca628',
        'path_trace_status.csv': '9d9eabcd54cbba3b468b9e7153218a954e43c9efec031b2089f741c4fff20ea5',
        'path_trace_status.parquet': '56c8b08e1d5efffa930e703c52b9e9de848e21d94c5ae9a760ca47083a4a1df1',
        'ped_bridge_mode_config.csv': '60583741fcd8484df3e4f166a82e49a06fdeb0fd353756fcfdf48a1f9786efc4',
        'ped_bridge_mode_config.parquet': '7b512297e05dfea806ccb86c984b816f3d2b7da7012be5ecef4f531dcae77c39',
        'ped_bridge_shape_fit_metrics.csv': '3a435ddaa5ecfecda0d068a453acd617e01c7b8fe2c313973d5694d125171fb8',
        'ped_bridge_shape_fit_metrics.parquet': '2e2bf09c973c9b6a4a157f34310c56c00d8b997e53910b3b056bf5c985201040',
        'ped_efficiency_scenarios.csv': '340a91407a2d1c7565ef3b5339bb14554d55b7dd4c4c48b66b39847af268a37f',
        'ped_efficiency_scenarios.parquet': '529f9bad8f06391846215906986166528384b0e7a9dfe50dc681d998437dbd35',
        'ped_revenue_bridge_audit.csv': '3ac01f1743632d104f1c341eeb6af5bbcae42d34965b9c6d0127522a8be99858',
        'ped_revenue_bridge_audit.parquet': '9d7bfd2ce74f299b1bc282833098f15d69cb03a6dc6e60f97f38f84ee0430730',
        'quarterly_reconstitution_audit.csv': '0960e1d20d0a8cea8a6f9042d2a704ee9d10995d70770322350a45e43bdfb304',
        'quarterly_reconstitution_audit.parquet': '02330b0320ace399dd885815c2f5d27e177f3d3b56106f6c51d2ac8682971a5f',
        'raw_quarterly_forecast_audit.csv': '2d84801c1c39073b5f340a06ed9fec644ce2faf8abb30a288a4314eaee4ec222',
        'raw_quarterly_forecast_audit.parquet': '5062e4c77f4d7e926d9a9494b51132f476623710adab1ba19f35e2b68a98f300',
        'revenue_bridge_components.csv': '1a9c4f2a37329da2ab56070210c2db89c449bab2faa62362fe3a0a591c0daf9f',
        'revenue_bridge_components.parquet': '0a0a74c594ddc41c14b7089f7cce8c801cd9d733a526f859700f614009b30cf4',
        'revenue_chart_rows.csv': 'eb5176d5ec78d52a603f405d5730b7e6212565b0d98598853b6b226936791749',
        'revenue_chart_rows.parquet': 'da2655cd252780495b61a582471b4938e1cf2873874655d97094909e27b7c778',
        'revenue_formula_residuals.csv': '6bebd38530a9fc92141064a02d4fe611dfd22c12142e90203eb979fe9b62955c',
        'revenue_formula_residuals.parquet': '78044187267396940e7504a602662b48eb7c25094a1abe6930087302e9100835',
        'revenue_line_reconciliation.csv': 'd1d57befba13fa755cec1a680c942332b0d1cefd189e5758ec42a74a18aec4d4',
        'revenue_line_reconciliation.parquet': 'f8ddacdfcd5bc50b1982fe5bcabf6242af00257e6291c0c1b2ca4b182dec9022',
        'revenue_stack_components.csv': '6b90610da5393a305d26a1db61a3667cb7d8f98e25f9ff77ad81ca994bccc91a',
        'revenue_stack_components.parquet': 'e49c70d51f54d2e58c48e5045d5253d8579a619c9fe9bea4cf4c136dd02f7de8',
        'row_reconciliation.csv': 'c3b08c6a325833c85b042335acb58cd858e3b7377adb92fc3570507934258c9b',
        'row_reconciliation.parquet': '848f119deaddd1bac51883f85ff6f3b870ea583e3bcabef688d43376a3ba10ae',
        'runtime_cutoff_audit.csv': 'ddf70dcd8ce5d2965e56bcb248a96c3efee10367e2b31e5005c2107d96aa18ed',
        'runtime_cutoff_audit.parquet': '80a9881662173029f23d0d423f002e257abb272f8f7f980e00bbdfc3ca6584ca',
        'runtime_trace_audit.csv': '21a1cd8b59ba2f0edb9c06d105565f625b5644b3c42628468a5687faf50de756',
        'runtime_trace_audit.parquet': '4932e7c0fb20d441d2218fd44e0ba839d81e0783609beef3d62bf0ca7bad5917',
        'scenario_feature_lineage.csv': 'b123c97090bd282009225a0ac2cfc36226d20017412f820dfeec6af34411b30d',
        'scenario_feature_lineage.parquet': 'dbc4ac3f2693f6de0915aebd21bc85e2f794f878eaefef0eb9031de468fb00e9',
        'scenario_input_delta_audit.csv': '546a2e00c02b247368196a106499d1a1473619f5770d70794ab76428e1919cf8',
        'scenario_input_delta_audit.parquet': '578f75991a26df530ba6715c8ff1a0b842800b3c76b29914b06be1643f486c5a',
        'scenario_input_replay_mismatch_report.csv': 'd490cb117bd39913cd604a9285b311fe28e47be6dd9c10e049d03ba070d22012',
        'scenario_input_replay_mismatch_report.parquet': 'ff274edfb7d0a3fd999e22b0aed54a125b751edfa3dfb6830c688c95c7b4ef4a',
        'scenario_role_contract.csv': '08b0f3e2b11ddddcfbafe9698b44af6076996cf99727350e9c2c7d39bc4dd74d',
        'scenario_role_contract.parquet': '3ceb32bd026e46352c72bb026b9f408f584dbbb2403631746395374cadd20b09',
        'sensitivity_config.csv': '3c29b44c845f67ee27d71baa3420aa7af1a077e09904675ca6475a63fef129b1',
        'sensitivity_config.parquet': '48e4fe218d6e136c8253c5435f0d80960ccd2698b5132800ba7ba16f57f7b9f4',
        'sensitivity_impact_audit.csv': '768c0da3f7c2c11c510849f07b7b2838bf2a1721df474c08e6f5b63a056492e2',
        'sensitivity_impact_audit.parquet': 'c4bf77b8aceecb7f58ee04a706f61be7bd0cb4c78bed61a9f3ade0c82c5e7f89',
        'sensitivity_seed_inputs.csv': '5181058396fbdb3896ade20b3ea335955fc186af85cccd365c7d86531852f3a6',
        'sensitivity_seed_inputs.parquet': '23a8ba664690235b92fac25912b9c8ab0eefe990f94a58258b901fafa306b310',
        'series_alias_audit.csv': '225bd2e4407bb97aa40b515a0ad35f6501e0b15779aa3efa5ee0884290efbcf7',
        'series_alias_audit.parquet': '09f87b5713889de262fcd8c257d3f829266e10ec10a45b520ac081b19ac96eca',
        'series_trace_contract.csv': 'c717badc9b8056755483a6c494820b62cd44bb734a99f8e0b12fe75188bf4db6',
        'series_trace_contract.parquet': '866482606699335a37eccac418f667a77ad93492b05e6f0e43d9aea4651c4e35',
        'stream_vintage_status.csv': 'c606a559f4f96634cace973eda9b78f4c6975f08b45842a2186a818f1a845b05',
        'stream_vintage_status.parquet': 'c9d5e63a9418cb87c505f98d55a78e4a9d813c04d7dbc26c625ddaee4fb2a73c',
        'trace_source_contract.csv': '36137d8e5cac28530159151e04ac3094fedd49ffb43b5dff3aa91e00172fd95c',
        'trace_source_contract.parquet': '85d2a78906c8756e145112e1ab2f0ef8e2e26801a390f0db83cb30629b66c971',
        'treasury_befu26_macro_path.csv': '1dd2fcbbc122ad8dd502e44ff1833f93ff26169dcd7dc47026da787adaf8aa1e',
    }
    assert {path.name: _sha256(path) for path in sorted(pack_dir.iterdir()) if path.is_file()} == expected_hashes


def test_revenue_outlook_loader_rejects_hash_mismatched_promoted_pack(tmp_path: Path) -> None:
    pack_copy = tmp_path / "current_revenue_outlook"
    shutil.copytree(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, pack_copy)
    chart_csv = pack_copy / "revenue_chart_rows.csv"
    chart_csv.write_text(chart_csv.read_text(encoding="utf-8") + "\n# tampered\n", encoding="utf-8")

    with pytest.raises(ValueError, match="revenue_chart_rows.csv hash mismatch"):
        load_revenue_outlook_pack(pack_copy, repo_root=ROOT)


def test_revenue_outlook_loader_accepts_git_normalized_csv_line_endings(tmp_path: Path) -> None:
    pack_copy = tmp_path / "current_revenue_outlook"
    shutil.copytree(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, pack_copy)
    manifest = json.loads((pack_copy / "manifest.json").read_text(encoding="utf-8"))
    for filename in manifest["output_hashes"]:
        if str(filename).endswith(".csv"):
            path = pack_copy / filename
            path.write_bytes(path.read_bytes().replace(b"\r\n", b"\n"))

    pack = load_revenue_outlook_pack(pack_copy, repo_root=ROOT)

    assert pack is not None
    assert not pack.revenue_chart_rows.empty


def test_revenue_outlook_loader_rejects_missing_required_runtime_file(tmp_path: Path) -> None:
    pack_copy = tmp_path / "current_revenue_outlook"
    shutil.copytree(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, pack_copy)
    (pack_copy / "future_revenue_forecasts.parquet").unlink()

    with pytest.raises(ValueError, match="future_revenue_forecasts.parquet is missing"):
        load_revenue_outlook_pack(pack_copy, repo_root=ROOT)


def test_revenue_outlook_loader_uses_committed_csv_fallback_when_parquet_engine_unavailable(monkeypatch) -> None:
    def fail_read_parquet(*args, **kwargs):
        raise ImportError("simulated missing parquet engine")

    monkeypatch.setattr(pd, "read_parquet", fail_read_parquet)
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)

    assert pack is not None
    assert not pack.revenue_chart_rows.empty
    assert not pack.revenue_bridge_components.empty
    assert not pack.future_revenue_forecasts.empty
    assert not pack.revenue_chart_rows.astype(str).stack().str.contains(r"C:\\Users|Downloads|OneDrive", regex=True).any()


def test_revenue_outlook_loader_uses_committed_runtime_pack_without_excel_or_local_paths(monkeypatch) -> None:
    def fail_read_excel(*args, **kwargs):
        raise AssertionError("Revenue Outlook runtime loader must not read Excel workbooks")

    monkeypatch.setattr(pd, "read_excel", fail_read_excel)
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None

    forbidden = r"C:\\Users|C:/Users|Downloads|OneDrive"
    for name, value in vars(pack).items():
        if isinstance(value, pd.DataFrame) and not value.empty:
            assert not value.astype(str).stack().str.contains(forbidden, regex=True).any(), name


def test_ped_bridge_audit_application_preserves_fallback_key_semantics() -> None:
    audit = pd.DataFrame(
        [
            {
                "FY": 2029,
                "source_path": "Current finalist Base case",
                "scenario_name": "current_basecase",
                "fed_path": "Current planned path",
                "series_id": "gross_ped_revenue",
                "adjusted": 15.0,
                "selected_ped_bridge_label": "Raw model bridge",
                "bridge_alpha": 0.0,
                "delta": 5.0,
                "gap_reason": "",
            }
        ]
    )

    def apply(frame: pd.DataFrame, *, source: bool = True, fed: bool = True, mask: bool = False) -> pd.DataFrame:
        return revenue_outlook_module._apply_ped_bridge_mode_audit_to_frame(
            frame,
            audit,
            value_column="value",
            fy_column="period",
            series_column="series_id",
            source_path_column="source_path" if source else None,
            scenario_column="scenario_name",
            fed_path_column="fed_path" if fed else None,
            current_mask_column="trace_role" if mask else None,
        )

    base = {
        "period": "FY2029",
        "series_id": "gross_ped_revenue",
        "scenario_name": "current_basecase",
        "value": 10.0,
    }
    exact = apply(pd.DataFrame([{**base, "source_path": "Current finalist Base case", "fed_path": "Current planned path"}]))
    no_source = apply(pd.DataFrame([{**base, "fed_path": "Current planned path"}]), source=False)
    no_fed = apply(pd.DataFrame([{**base, "source_path": "Current finalist Base case"}]), fed=False)
    no_source_no_fed = apply(pd.DataFrame([base]), source=False, fed=False)
    masked = apply(
        pd.DataFrame(
            [{**base, "source_path": "Current finalist Base case", "fed_path": "Current planned path", "trace_role": "official_comparator"}]
        ),
        mask=True,
    )

    for frame in [exact, no_source, no_fed, no_source_no_fed]:
        assert frame.loc[0, "value"] == pytest.approx(15.0)
        assert frame.loc[0, "ped_bridge_value_delta"] == pytest.approx(5.0)
        assert "Raw model bridge" in str(frame.loc[0, "ped_bridge_mode_label"])
    assert masked.loc[0, "value"] == pytest.approx(10.0)
