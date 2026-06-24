from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path
import subprocess
import sys

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from model_dashboard.forecast_runner import (
    BACKTEST_SUPPORTED_MAX_HORIZON,
    DEFAULT_FORECAST_HORIZON_QUARTERS,
    HIGH_POPULATION_SMOKE_FIXTURE_NOTE,
    HORIZON_SUPPORT_NOTE,
    SHEET_BY_STREAM,
    STREAM_COLUMNS,
    STREAM_ORDER,
    TEMPLATE_FILENAME,
    build_forecast_input_template,
    create_completed_sample_workbook,
    forecast_template_filename,
    future_quarters_after,
    latest_known_actual_period,
    model_capability_gap_register,
    resolve_scenario_role,
    run_forecast_workbook,
    scenario_name_from_filename,
    validate_forecast_workbook,
    write_forecast_scenario_comparison,
)


ROOT = Path(__file__).resolve().parents[1]
C4 = "HEAVY_RUC__dynamic_no_leads__GBR_learning_rate0_08_max_depth1_n_estimators150__ylag__w40"
CURRENT_PED_FINALIST = "PED__VNEXT_SOLVED_CONVEX_TOP2"
CURRENT_HEAVY_FINALIST = "HEAVY_RUC__VNEXT_SOLVED_CONVEX_TOP4"
CURRENT_LIGHT_FINALIST = "dynamic_RESID_GBR_n150_d1_lr0.05_w36"


def _expected_periods(horizon: int) -> list[str]:
    return future_quarters_after(latest_known_actual_period(ROOT), horizon)


def _assert_template_contract(path: Path, horizon: int) -> None:
    wb = load_workbook(path, data_only=False)
    assert wb.sheetnames == ["README", "PED Inputs", "Light RUC Inputs", "Heavy RUC Inputs"]
    expected_periods = _expected_periods(horizon)
    for stream in STREAM_ORDER:
        ws = wb[SHEET_BY_STREAM[stream]]
        headers = [cell.value for cell in ws[1]]
        expected_headers = [column.name for column in STREAM_COLUMNS[stream]]
        assert headers == expected_headers
        assert ws.max_row == horizon + 1
        assert [ws.cell(row=row, column=1).value for row in range(2, horizon + 2)] == expected_periods
        assert all(cell.comment for cell in ws[1] if cell.value)
        assert ws.protection.sheet is True
        formula_headers = [column.name for column in STREAM_COLUMNS[stream] if column.role == "formula"]
        assert {"q2_dummy", "q3_dummy", "q4_dummy", "post_2020_dummy", "trend_index"}.issubset(formula_headers)
        assert any(header.startswith("log_") for header in formula_headers)
        assert any(header.startswith("diff_") for header in formula_headers)
        assert any("interaction" in header for header in formula_headers)
        for header in formula_headers:
            value = ws.cell(row=2, column=headers.index(header) + 1).value
            assert isinstance(value, str) and value.startswith("="), header
            last_value = ws.cell(row=horizon + 1, column=headers.index(header) + 1).value
            assert isinstance(last_value, str) and last_value.startswith("="), header


def _sorted_forecast_values(frame: pd.DataFrame) -> pd.DataFrame:
    columns = ["stream", "model", "target_period", "horizon", "forecast", "forecast_available"]
    out = frame[columns].copy()
    out["forecast"] = pd.to_numeric(out["forecast"], errors="coerce")
    return out.sort_values(["stream", "target_period"], kind="stable").reset_index(drop=True)


def _tracked_data_hashes() -> dict[str, str]:
    result = subprocess.run(
        ["git", "ls-files", "data/dashboard_evidence_pack", "artifacts/chart_sources"],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    hashes: dict[str, str] = {}
    for rel_path in result.stdout.splitlines():
        path = ROOT / rel_path
        if path.is_file():
            hashes[rel_path] = hashlib.sha256(path.read_bytes()).hexdigest()
    return hashes


def _forecast_values_by_role(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    if "scenario_role" not in out.columns or out["scenario_role"].isna().all():
        scenario = out.get("scenario_name", pd.Series("", index=out.index)).astype(str)
        out["scenario_role"] = scenario.map({"basecase": "basecase", "high_population": "comparison"}).fillna(scenario)
    out["forecast"] = pd.to_numeric(out["forecast"], errors="coerce")
    columns = ["scenario_role", "stream", "model", "target_period", "horizon", "forecast", "forecast_available"]
    return out[columns].sort_values(["scenario_role", "stream", "target_period"], kind="stable").reset_index(drop=True)


def _assert_source_derived_continuity(result) -> None:
    for stream in STREAM_ORDER:
        historical = result.forecast_chart_rows[
            result.forecast_chart_rows["row_type"].astype(str).eq("historical_actual")
            & result.forecast_chart_rows["stream"].astype(str).eq(stream)
        ].copy()
        historical["period_key"] = historical["period"].astype(str).map(lambda value: int(value[:4]) * 4 + int(value[-1]))
        historical = historical.sort_values("period_key", kind="stable")
        history_values = pd.to_numeric(historical["value"], errors="coerce").dropna()
        recent_qoq = history_values.pct_change().dropna().abs().tail(20)
        assert not recent_qoq.empty, stream
        tolerance = max(float(recent_qoq.quantile(0.95)) * 3.0, float(recent_qoq.max()) * 2.0)

        future = result.future_forecasts[result.future_forecasts["stream"].eq(stream)].sort_values("target_period")
        future_values = pd.to_numeric(future["forecast"], errors="coerce")
        actual_to_h1 = abs(float(future_values.iloc[0]) / float(history_values.iloc[-1]) - 1.0)
        h1_h12_qoq = future_values.head(BACKTEST_SUPPORTED_MAX_HORIZON).pct_change().dropna().abs()

        assert actual_to_h1 <= tolerance, f"{stream} latest actual-to-H1 {actual_to_h1:.6f} exceeds source tolerance {tolerance:.6f}"
        assert float(h1_h12_qoq.max()) <= tolerance, (
            f"{stream} H1-H{BACKTEST_SUPPORTED_MAX_HORIZON} QoQ {float(h1_h12_qoq.max()):.6f} "
            f"exceeds source tolerance {tolerance:.6f}"
        )


def _set_required_user_values(path: Path, stream: str, excel_row: int, value: float) -> None:
    wb = load_workbook(path)
    ws = wb[SHEET_BY_STREAM[stream]]
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    ws.protection.sheet = False
    for column in STREAM_COLUMNS[stream]:
        if column.role == "user" and column.required:
            ws.cell(row=excel_row, column=headers[column.name]).value = value
    ws.protection.sheet = True
    wb.save(path)


def _multiply_target_lag_columns(path: Path, streams: tuple[str, ...], multiplier: float) -> None:
    wb = load_workbook(path)
    for stream in streams:
        ws = wb[SHEET_BY_STREAM[stream]]
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        ws.protection.sheet = False
        for column_name in ("target_lag_1", "target_lag_4"):
            col = headers[column_name]
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    ws.cell(row=row, column=col).value = float(value) * multiplier
        ws.protection.sheet = True
    wb.save(path)


def test_forecast_template_workbook_contract_for_variable_horizons(tmp_path: Path) -> None:
    one_quarter = tmp_path / forecast_template_filename(quarters=1)
    default_template = tmp_path / forecast_template_filename(quarters=DEFAULT_FORECAST_HORIZON_QUARTERS)
    long_template = tmp_path / forecast_template_filename(end_period="2050Q4")
    build_forecast_input_template(one_quarter, repo_root=ROOT, quarters=1)
    build_forecast_input_template(default_template, repo_root=ROOT)
    build_forecast_input_template(long_template, repo_root=ROOT, end_period="2050Q4")

    _assert_template_contract(one_quarter, 1)
    _assert_template_contract(default_template, DEFAULT_FORECAST_HORIZON_QUARTERS)
    _assert_template_contract(long_template, 100)
    assert forecast_template_filename(quarters=20) == "NLTF_forecast_input_template_20q.xlsx"
    assert forecast_template_filename(end_period="2050q4") == "NLTF_forecast_input_template_to_2050Q4.xlsx"
    for invalid_horizon in [0, -1]:
        try:
            forecast_template_filename(quarters=invalid_horizon)
        except ValueError as exc:
            assert "at least 1 quarter" in str(exc)
        else:
            raise AssertionError(f"Expected invalid horizon {invalid_horizon} to fail.")


def test_committed_forecast_template_exists_and_is_small() -> None:
    path = ROOT / "templates" / TEMPLATE_FILENAME
    assert path.exists()
    assert path.stat().st_size < 1_000_000
    _assert_template_contract(path, DEFAULT_FORECAST_HORIZON_QUARTERS)
    validation = validate_forecast_workbook(path, repo_root=ROOT)
    assert not validation.valid
    assert validation.latest_actual_period == "2025Q4"
    assert validation.forecast_periods == _expected_periods(DEFAULT_FORECAST_HORIZON_QUARTERS)
    assert any("no valid forecast rows" in message for message in validation.errors)


def test_runner_handles_missing_inputs_cleanly(tmp_path: Path) -> None:
    template = tmp_path / TEMPLATE_FILENAME
    build_forecast_input_template(template, repo_root=ROOT)
    result = run_forecast_workbook(template, output_dir=tmp_path / "blank_run", repo_root=ROOT, run_timestamp="blank-run")
    assert result.manifest["validation_status"] == "failed"
    assert result.manifest["forecast_status"] == "validation_failed"
    assert result.manifest["broad_search_run"] is False
    assert result.future_forecasts["forecast_available"].eq(False).all()
    assert result.future_forecasts["forecast"].isna().all()
    assert not result.future_forecasts["forecast"].fillna("").astype(str).isin({"0", "0.0"}).any()
    for name in [
        "future_forecasts.parquet",
        "component_forecasts.parquet",
        "forecast_assumptions.parquet",
        "forecast_capability_report.parquet",
        "forecast_chart_rows.parquet",
        "forecast_run_manifest.json",
        "forecast_validation_report.md",
        "forecast_capability_report.csv",
        "forecast_chart_rows.csv",
    ]:
        assert (result.output_dir / name).exists(), name
    assert result.forecast_chart_rows["row_type"].isin(["historical_actual", "future_forecast"]).all()
    assert result.forecast_chart_rows[result.forecast_chart_rows["row_type"].eq("historical_actual")]["value"].notna().any()


def test_model_input_history_pack_is_committed_and_hash_backed() -> None:
    history_dir = ROOT / "data" / "model_input_history"
    manifest_path = history_dir / "manifest.json"
    assert manifest_path.exists()
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["source_basename"] == "Master Copy revenue modelling workbook.xlsx"
    assert manifest["source_sha256"]
    assert manifest["workbook_full_path_public"] is False
    paths = {
        "PED": history_dir / "ped_inputs.parquet",
        "LIGHT_RUC": history_dir / "light_ruc_inputs.parquet",
        "HEAVY_RUC": history_dir / "heavy_ruc_inputs.parquet",
    }
    for stream, path in paths.items():
        assert path.exists(), stream
        assert path.stat().st_size < 1_000_000
        frame = pd.read_parquet(path)
        assert len(frame) >= 90
        assert {"period", "target", "year", "quarter"}.issubset(frame.columns)
        assert pd.to_numeric(frame["target"], errors="coerce").gt(0).any()


def test_model_capability_register_is_parity_gated_and_hash_backed(monkeypatch) -> None:
    # Pin to the legacy governance path; vNext capability is covered separately.
    import model_dashboard.vnext_forward_integration as vfi

    monkeypatch.setattr(vfi, "evaluate_vnext_forward_scorer", lambda root, stream: None)
    capabilities = model_capability_gap_register(ROOT).set_index("stream")
    assert capabilities.loc["LIGHT_RUC", "capability_status"] == "numeric_forecast_available"
    assert capabilities.loc["LIGHT_RUC", "forecast_capability_available"] == True
    assert capabilities.loc["LIGHT_RUC", "parity_status"] == "passed_repo_local_recipe"

    assert capabilities.loc["HEAVY_RUC", "capability_status"] == "parity_failed"
    assert capabilities.loc["HEAVY_RUC", "forecast_capability_available"] == False
    assert capabilities.loc["HEAVY_RUC", "gap_code"] == "heavy_ruc_component_forward_scorers_missing"
    assert capabilities.loc["HEAVY_RUC", "parity_status"] == "failed_canonical_history_component_replay"
    assert 0 <= capabilities.loc["HEAVY_RUC", "stored_replay_max_delta"] <= 1e-6
    assert capabilities.loc["HEAVY_RUC", "max_parity_delta"] > 1
    assert capabilities.loc["HEAVY_RUC", "failing_component"] == C4
    assert "Source-script Stage 1 workbook history was recovered" in capabilities.loc["HEAVY_RUC", "gap_reason"]
    assert "target-lagged GBM components C3/C4 still exceed parity tolerance" in capabilities.loc["HEAVY_RUC", "gap_reason"]

    assert capabilities.loc["PED", "capability_status"] == "parity_failed"
    assert capabilities.loc["PED", "forecast_capability_available"] == False
    assert capabilities.loc["PED", "gap_code"] == "ped_inner_hpo_static_solver_forward_scorer_missing"
    assert capabilities.loc["PED", "parity_status"] == "failed_inner_hpo_replay_delta"
    assert capabilities.loc["PED", "max_parity_delta"] > 1
    assert "feature_level_refit_not_attempted" in capabilities.loc["PED", "gap_reason"]

    for stream in STREAM_ORDER:
        assert capabilities.loc[stream, "scorer_version"]
        assert capabilities.loc[stream, "source_artifact_hashes"]


def test_completed_basecase_and_high_population_validate_to_2050q4(tmp_path: Path) -> None:
    basecase = create_completed_sample_workbook(
        tmp_path / "NLTF_forecast_input_template_basecase.xlsx",
        repo_root=ROOT,
        end_period="2050Q4",
    )
    high_population = create_completed_sample_workbook(
        tmp_path / "NLTF_forecast_input_template_high_population.xlsx",
        repo_root=ROOT,
        end_period="2050Q4",
        value_multiplier=1.02,
    )
    for workbook in [basecase, high_population]:
        validation = validate_forecast_workbook(workbook, repo_root=ROOT, expected_end_period="2050Q4")
        assert validation.valid
        assert validation.forecast_start_period == "2026Q1"
        assert validation.forecast_end_period == "2050Q4"
        assert len(validation.forecast_periods) == 100
        assert len(validation.assumptions) == 100 * len(STREAM_ORDER)


def test_completed_workbook_builds_transforms_and_numeric_light_outputs(tmp_path: Path, monkeypatch) -> None:
    # Pin to the legacy governance path; vNext capability is covered separately.
    import model_dashboard.vnext_forward_integration as vfi

    monkeypatch.setattr(vfi, "evaluate_vnext_forward_scorer", lambda root, stream: None)
    workbook = create_completed_sample_workbook(tmp_path / "completed.xlsx", repo_root=ROOT)
    result = run_forecast_workbook(
        workbook,
        output_dir=tmp_path / "sample_run",
        repo_root=ROOT,
        run_timestamp="sample-run",
        scenario_name="basecase",
    )
    assert result.manifest["validation_status"] == "passed"
    assert result.manifest["forecast_status"] == "mixed_numeric_and_governed_gap"
    assert result.manifest["scenario_name"] == "basecase"
    assert result.manifest["workbook_filename"] == "completed.xlsx"
    assert result.manifest["workbook_sha256"]
    assert result.manifest["forecast_horizon_quarters"] == DEFAULT_FORECAST_HORIZON_QUARTERS
    assert result.manifest["forecast_start_period"] == "2026Q1"
    assert result.manifest["forecast_end_period"] == "2030Q4"
    assert result.manifest["numeric_forecast_streams"] == ["LIGHT_RUC"]
    assert result.manifest["governed_gap_streams"] == ["HEAVY_RUC", "PED"]
    assert result.manifest["fixed_finalists_only"] is True
    assert result.manifest["broad_search_run"] is False
    assert result.manifest["evidence_pack_modified"] is False
    assert result.manifest["chart_sources_modified"] is False
    assert len(result.future_forecasts) == DEFAULT_FORECAST_HORIZON_QUARTERS * len(STREAM_ORDER)
    assert "row_type" not in result.future_forecasts.columns
    assert result.future_forecasts["scenario_name"].eq("basecase").all()
    assert result.capability_report["scenario_name"].eq("basecase").all()
    assert result.assumptions["scenario_name"].eq("basecase").all()
    metadata_columns = {
        "scorer_version",
        "source_artifact_hashes",
        "parity_status",
        "max_parity_delta",
        "stored_replay_max_delta",
        "capability_status",
        "failing_component",
    }
    assert metadata_columns.issubset(result.future_forecasts.columns)
    assert metadata_columns.issubset(result.component_forecasts.columns)
    assert metadata_columns.issubset(result.capability_report.columns)
    assert metadata_columns.issubset(result.forecast_chart_rows.columns)
    light = result.future_forecasts[result.future_forecasts["stream"].eq("LIGHT_RUC")]
    gaps = result.future_forecasts[~result.future_forecasts["stream"].eq("LIGHT_RUC")]
    assert len(light) == DEFAULT_FORECAST_HORIZON_QUARTERS
    assert light["forecast_available"].eq(True).all()
    assert pd.to_numeric(light["forecast"], errors="coerce").notna().all()
    assert (pd.to_numeric(light["forecast"], errors="coerce") > 0).all()
    assert light["availability_status"].eq("numeric_forecast_available").all()
    assert light["capability_status"].eq("numeric_forecast_available").all()
    assert light["parity_status"].eq("passed_repo_local_recipe").all()
    assert gaps["forecast_available"].eq(False).all()
    assert gaps["forecast"].isna().all()
    assert gaps["availability_status"].eq("governed_gap").all()
    assert set(gaps["gap_code"]) == {
        "ped_inner_hpo_static_solver_forward_scorer_missing",
        "heavy_ruc_component_forward_scorers_missing",
    }
    gap_status = gaps.drop_duplicates("stream").set_index("stream")["capability_status"].to_dict()
    assert gap_status == {"PED": "parity_failed", "HEAVY_RUC": "parity_failed"}
    gap_parity = gaps.drop_duplicates("stream").set_index("stream")["parity_status"].to_dict()
    assert gap_parity["PED"] == "failed_inner_hpo_replay_delta"
    assert gap_parity["HEAVY_RUC"] == "failed_canonical_history_component_replay"
    gap_failing_component = gaps.drop_duplicates("stream").set_index("stream")["failing_component"].to_dict()
    assert gap_failing_component["HEAVY_RUC"] == C4
    heavy_reason = " ".join(gaps[gaps["stream"].eq("HEAVY_RUC")]["gap_reason"].dropna().astype(str).unique())
    ped_reason = " ".join(gaps[gaps["stream"].eq("PED")]["gap_reason"].dropna().astype(str).unique())
    assert "C1 ElasticNet dynamic no-leads ylag w64" in heavy_reason
    assert "C4 GBM dynamic no-leads ylag w40" in heavy_reason
    assert "inner HPO/static-solver forward scorer" in ped_reason
    light_components = result.component_forecasts[result.component_forecasts["stream"].eq("LIGHT_RUC")]
    assert {"base_schiff_ols", "residual_gbr", "dynamic_RESID_GBR_n150_d1_lr0.05_w36"}.issubset(
        set(light_components["component_model"].astype(str))
    )
    assert pd.to_numeric(light_components["component_forecast"], errors="coerce").notna().all()
    assert (result.future_forecasts["forecast_available"] == result.future_forecasts["stream"].eq("LIGHT_RUC")).all()
    capability = result.capability_report.set_index("stream")
    assert capability.loc["LIGHT_RUC", "forecast_available"] == True
    assert capability.loc["LIGHT_RUC", "numeric_forecast_rows"] == DEFAULT_FORECAST_HORIZON_QUARTERS
    assert capability.loc["LIGHT_RUC", "capability_status"] == "numeric_forecast_available"
    assert capability.loc["PED", "capability_status"] == "parity_failed"
    assert capability.loc["HEAVY_RUC", "capability_status"] == "parity_failed"
    assert capability.loc["PED", "max_parity_delta"] > 1
    assert capability.loc["HEAVY_RUC", "max_parity_delta"] > 1
    assert 0 <= capability.loc["HEAVY_RUC", "stored_replay_max_delta"] <= 1e-6
    assumptions = result.assumptions
    for column in [
        "q2_dummy",
        "q3_dummy",
        "q4_dummy",
        "post_2020_dummy",
        "trend_index",
        "log_real_gdp",
        "log_real_gdp_per_capita",
        "diff_log_real_gdp",
        "gdp_light_ruc_price_interaction",
        "gdp_heavy_ruc_price_interaction",
    ]:
        assert column in assumptions.columns
    assert assumptions["transform_status"].eq("built_from_template_inputs").all()
    manifest = json.loads((result.output_dir / "forecast_run_manifest.json").read_text(encoding="utf-8"))
    assert manifest["broad_search_run"] is False
    assert manifest["output_files"]
    assert all(metadata_columns.issubset(record.keys()) for record in manifest["model_capabilities"])
    capabilities_by_stream = {record["stream"]: record for record in manifest["model_capabilities"]}
    assert capabilities_by_stream["LIGHT_RUC"]["capability_status"] == "numeric_forecast_available"
    assert capabilities_by_stream["PED"]["capability_status"] == "parity_failed"
    assert capabilities_by_stream["HEAVY_RUC"]["capability_status"] == "parity_failed"
    assert capabilities_by_stream["PED"]["source_artifact_hashes"]
    assert capabilities_by_stream["HEAVY_RUC"]["source_artifact_hashes"]
    assert capabilities_by_stream["HEAVY_RUC"]["failing_component"] == C4
    assert "forecast_chart_rows.parquet" in manifest["output_files"]
    assert (result.output_dir / "forecast_capability_report.csv").exists()
    assert (result.output_dir / "forecast_chart_rows.parquet").exists()
    assert (result.output_dir / "forecast_chart_rows.csv").exists()
    chart_rows = result.forecast_chart_rows
    assert {"row_type", "scenario_name", "stream", "period", "value", "availability_status"}.issubset(chart_rows.columns)
    assert {"historical_actual", "future_forecast"} == set(chart_rows["row_type"].dropna().astype(str).unique())
    historical = chart_rows[chart_rows["row_type"].eq("historical_actual")]
    assert set(historical["stream"]) == set(STREAM_ORDER)
    assert historical["scenario_name"].eq("historical_actual").all()
    assert historical["period"].map(lambda value: int(str(value)[:4])).max() == 2025
    future_chart = chart_rows[chart_rows["row_type"].eq("future_forecast")]
    assert len(future_chart) == len(result.future_forecasts)
    light_chart = future_chart[future_chart["stream"].eq("LIGHT_RUC")].sort_values("period")
    pd.testing.assert_series_equal(
        light_chart["value"].reset_index(drop=True).astype(float),
        light.sort_values("target_period")["forecast"].reset_index(drop=True).astype(float),
        check_names=False,
    )
    gap_chart = future_chart[~future_chart["stream"].eq("LIGHT_RUC")]
    assert gap_chart["value"].isna().all()
    assert gap_chart.drop_duplicates("stream").set_index("stream")["capability_status"].to_dict() == {
        "PED": "parity_failed",
        "HEAVY_RUC": "parity_failed",
    }


def _clear_required_user_values(path: Path, stream: str, excel_row: int) -> None:
    wb = load_workbook(path)
    ws = wb[SHEET_BY_STREAM[stream]]
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    ws.protection.sheet = False
    for column in STREAM_COLUMNS[stream]:
        if column.role == "user" and column.required:
            ws.cell(row=excel_row, column=headers[column.name]).value = None
    ws.protection.sheet = True
    wb.save(path)


def test_validation_accepts_variable_horizons_and_rejects_gaps_or_unequal_periods(tmp_path: Path) -> None:
    one_quarter = create_completed_sample_workbook(tmp_path / "one_quarter.xlsx", repo_root=ROOT, quarters=1)
    validation = validate_forecast_workbook(one_quarter, repo_root=ROOT)
    assert validation.valid
    assert validation.forecast_periods == ["2026Q1"]
    assert len(validation.assumptions) == len(STREAM_ORDER)

    three_quarters = create_completed_sample_workbook(tmp_path / "three_quarters.xlsx", repo_root=ROOT, quarters=3)
    validation = validate_forecast_workbook(three_quarters, repo_root=ROOT, expected_quarters=3)
    assert validation.valid
    assert validation.forecast_periods == ["2026Q1", "2026Q2", "2026Q3"]
    assert len(validation.assumptions) == 3 * len(STREAM_ORDER)

    gapped = create_completed_sample_workbook(tmp_path / "gapped.xlsx", repo_root=ROOT, quarters=3)
    for stream in STREAM_ORDER:
        _clear_required_user_values(gapped, stream, excel_row=3)
    gap_validation = validate_forecast_workbook(gapped, repo_root=ROOT)
    assert not gap_validation.valid
    assert any("continuous from 2026Q1 through 2026Q2 with no gaps" in message for message in gap_validation.errors)

    unequal = create_completed_sample_workbook(tmp_path / "unequal.xlsx", repo_root=ROOT, quarters=3)
    _clear_required_user_values(unequal, "HEAVY_RUC", excel_row=4)
    unequal_validation = validate_forecast_workbook(unequal, repo_root=ROOT)
    assert not unequal_validation.valid
    assert any("same continuous forecast periods across all streams" in message for message in unequal_validation.errors)


def test_validation_rejects_nonpositive_required_inputs(tmp_path: Path) -> None:
    workbook = create_completed_sample_workbook(tmp_path / "nonpositive.xlsx", repo_root=ROOT, quarters=2)
    for stream in STREAM_ORDER:
        _set_required_user_values(workbook, stream, excel_row=2, value=0.0)
    validation = validate_forecast_workbook(workbook, repo_root=ROOT, expected_quarters=2)
    assert not validation.valid
    assert any("partial or non-positive required user entries" in message for message in validation.errors)


def test_filename_suffix_creates_expected_scenario_name() -> None:
    assert scenario_name_from_filename("NLTF_forecast_input_template_basecase.xlsx") == "basecase"
    assert scenario_name_from_filename("NLTF_forecast_input_template_high_fuel.xlsx") == "high_fuel"
    assert scenario_name_from_filename("forecast_input_downside.xlsx") == "downside"
    assert scenario_name_from_filename("completed_upside.xlsx") == "upside"
    assert scenario_name_from_filename("completed.xlsx") == "completed"


def test_scenario_role_inference_handles_horizon_copy_filename_variants() -> None:
    cases = [
        ("NLTF forecast input template to 2050Q4 Basecase (2) Copy.xlsx", "basecase"),
        ("to_2050q4_basecase_2_copy.xlsx", "basecase"),
        ("NLTF_forecast_input_template_to_2050Q4_high population Copy.xlsx", "comparison"),
        ("to_2050q4_high_population_2_copy.xlsx", "comparison"),
    ]
    for filename, expected_role in cases:
        scenario = scenario_name_from_filename(filename)
        role, source = resolve_scenario_role(scenario_name=scenario, workbook_filename=filename)
        assert role == expected_role
        assert source == "inferred_from_name"
    role, source = resolve_scenario_role(scenario_name="to_2050q4_2_copy", workbook_filename="to_2050q4_2_copy.xlsx")
    assert role is None
    assert source == "ambiguous"


def test_cli_runs_variable_horizon_with_scenario_name(tmp_path: Path) -> None:
    workbook = create_completed_sample_workbook(
        tmp_path / "NLTF_forecast_input_template_high_fuel.xlsx",
        repo_root=ROOT,
        quarters=1,
    )
    output_dir = tmp_path / "cli_run"
    result = subprocess.run(
        [
            sys.executable,
            str(ROOT / "scripts" / "run_forecast_pack.py"),
            str(workbook),
            "--scenario-name",
            "high_fuel",
            "--quarters",
            "1",
            "--output-dir",
            str(output_dir),
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )
    assert result.returncode == 0, result.stderr
    assert "Scenario: high_fuel" in result.stdout
    assert "Scenario role: ambiguous" in result.stdout
    manifest = json.loads((output_dir / "forecast_run_manifest.json").read_text(encoding="utf-8"))
    assert manifest["scenario_name"] == "high_fuel"
    assert manifest["scenario_role"] is None
    assert manifest["forecast_horizon_quarters"] == 1
    future = pd.read_parquet(output_dir / "future_forecasts.parquet")
    assert len(future) == len(STREAM_ORDER)
    light = future[future["stream"].eq("LIGHT_RUC")]
    assert len(light) == 1
    assert light["forecast_available"].eq(True).all()


def test_scenario_comparison_artifacts(tmp_path: Path) -> None:
    base = create_completed_sample_workbook(
        tmp_path / "NLTF forecast input template to 2050Q4 Basecase (2) Copy.xlsx",
        repo_root=ROOT,
        quarters=1,
    )
    high_population = create_completed_sample_workbook(
        tmp_path / "NLTF_forecast_input_template_to_2050Q4_high population Copy.xlsx",
        repo_root=ROOT,
        quarters=2,
        value_multiplier=1.02,
    )
    results = [
        run_forecast_workbook(
            base,
            output_dir=tmp_path / "basecase_run",
            repo_root=ROOT,
            run_timestamp="comparison-smoke",
        ),
        run_forecast_workbook(
            high_population,
            output_dir=tmp_path / "high_population_run",
            repo_root=ROOT,
            run_timestamp="comparison-smoke",
        ),
    ]
    comparison = write_forecast_scenario_comparison(
        results,
        output_dir=tmp_path / "comparison",
        repo_root=ROOT,
        run_timestamp="comparison-smoke",
    )
    assert comparison.manifest["scenario_count"] == 2
    assert comparison.manifest["scenario_role_validation"]["base_scenario"] == results[0].manifest["scenario_name"]
    assert {row["scenario_role"] for row in comparison.manifest["scenarios"]} == {"basecase", "comparison"}
    assert len(comparison.future_forecasts) == sum(len(result.future_forecasts) for result in results)
    assert set(comparison.future_forecasts["scenario_name"]) == {result.manifest["scenario_name"] for result in results}
    assert set(comparison.forecast_chart_rows["scenario_name"]) >= {"historical_actual", *[result.manifest["scenario_name"] for result in results]}
    historical = comparison.forecast_chart_rows[comparison.forecast_chart_rows["row_type"].eq("historical_actual")]
    assert historical.duplicated(subset=["stream", "period"]).sum() == 0
    for name in [
        "forecast_scenario_comparison.parquet",
        "forecast_scenario_comparison.csv",
        "forecast_scenario_capability_report.parquet",
        "forecast_scenario_capability_report.csv",
        "forecast_scenario_chart_rows.parquet",
        "forecast_scenario_chart_rows.csv",
        "scenario_input_delta_audit.parquet",
        "scenario_input_delta_audit.csv",
        "forecast_scenario_comparison_manifest.json",
    ]:
        assert (comparison.output_dir / name).exists(), name
    assert comparison.manifest["horizon_support_note"] == HORIZON_SUPPORT_NOTE
    assert "high_population_fixture_note" not in comparison.manifest
    assert not comparison.scenario_input_delta_audit.empty
    audit = comparison.scenario_input_delta_audit
    assert audit["base_scenario_role"].eq("basecase").all()
    assert audit["comparison_scenario_role"].eq("comparison").all()
    assert audit["base_original_filename"].str.contains("Basecase", regex=False).all()
    assert audit["comparison_original_filename"].str.contains("high population", regex=False).all()
    assert audit["base_workbook_sha256"].str.fullmatch(r"[0-9a-f]{64}").all()
    assert audit["comparison_workbook_sha256"].str.fullmatch(r"[0-9a-f]{64}").all()
    assert audit["decision_grade_status"].eq("scenario_delta_audit_only").all()
    assert not audit["all_required_inputs_plus_2pct"].fillna(False).astype(bool).any()
    assert audit["assumption_scope_note"].str.contains("not labelled as an all-inputs +2% smoke fixture", regex=False).all()
    assert pd.to_numeric(audit["pct_delta"], errors="coerce").dropna().between(0.019999999, 0.020000001).all()


def test_scenario_comparison_refuses_ambiguous_roles(tmp_path: Path) -> None:
    first = create_completed_sample_workbook(tmp_path / "to_2050q4_first_copy.xlsx", repo_root=ROOT, quarters=1)
    second = create_completed_sample_workbook(tmp_path / "to_2050q4_second_copy.xlsx", repo_root=ROOT, quarters=1)
    results = [
        run_forecast_workbook(first, output_dir=tmp_path / "first", repo_root=ROOT, run_timestamp="ambiguous"),
        run_forecast_workbook(second, output_dir=tmp_path / "second", repo_root=ROOT, run_timestamp="ambiguous"),
    ]
    try:
        write_forecast_scenario_comparison(results, output_dir=tmp_path / "comparison", repo_root=ROOT)
    except ValueError as exc:
        message = str(exc)
        assert "scenario role is ambiguous" in message
        assert "upload order is not used" in message
    else:
        raise AssertionError("Ambiguous scenario roles should fail comparison packaging.")


def test_all_inputs_plus_two_smoke_note_requires_fixture_flag(tmp_path: Path) -> None:
    base = create_completed_sample_workbook(tmp_path / "basecase.xlsx", repo_root=ROOT, quarters=1)
    fixture = create_completed_sample_workbook(tmp_path / "high_population.xlsx", repo_root=ROOT, quarters=1, value_multiplier=1.02)
    results = [
        run_forecast_workbook(
            base,
            output_dir=tmp_path / "basecase_run",
            repo_root=ROOT,
            run_timestamp="fixture",
            scenario_name="basecase",
            scenario_role="basecase",
        ),
        run_forecast_workbook(
            fixture,
            output_dir=tmp_path / "fixture_run",
            repo_root=ROOT,
            run_timestamp="fixture",
            scenario_name="high_population",
            scenario_role="comparison",
            is_test_fixture=True,
        ),
    ]
    comparison = write_forecast_scenario_comparison(results, output_dir=tmp_path / "comparison", repo_root=ROOT)
    audit = comparison.scenario_input_delta_audit
    assert audit["decision_grade_status"].eq("not_decision_grade_smoke_fixture").all()
    assert audit["all_required_inputs_plus_2pct"].astype(bool).all()
    assert audit["assumption_scope_note"].eq(HIGH_POPULATION_SMOKE_FIXTURE_NOTE).all()
    summary = comparison.manifest["scenario_assumption_delta_summary"][0]
    assert summary["all_required_inputs_plus_2pct"] is True
    assert summary["comparison_is_test_fixture"] is True


def test_forecast_runner_does_not_touch_evidence_pack(tmp_path: Path) -> None:
    evidence_dir = ROOT / "data" / "dashboard_evidence_pack"
    before = {
        path.relative_to(evidence_dir).as_posix(): path.read_bytes()
        for path in evidence_dir.rglob("*")
        if path.is_file()
    }
    workbook = create_completed_sample_workbook(tmp_path / "completed.xlsx", repo_root=ROOT)
    run_forecast_workbook(workbook, output_dir=tmp_path / "sample_run", repo_root=ROOT, run_timestamp="evidence-isolation")
    after = {
        path.relative_to(evidence_dir).as_posix(): path.read_bytes()
        for path in evidence_dir.rglob("*")
        if path.is_file()
    }
    assert after == before


def test_forecast_runner_does_not_touch_chart_sources(tmp_path: Path) -> None:
    chart_dir = ROOT / "artifacts" / "chart_sources"
    before = {
        path.relative_to(chart_dir).as_posix(): path.read_bytes()
        for path in chart_dir.glob("*.csv")
        if path.is_file()
    }
    workbook = create_completed_sample_workbook(tmp_path / "completed.xlsx", repo_root=ROOT)
    run_forecast_workbook(workbook, output_dir=tmp_path / "sample_run", repo_root=ROOT, run_timestamp="chart-isolation")
    after = {
        path.relative_to(chart_dir).as_posix(): path.read_bytes()
        for path in chart_dir.glob("*.csv")
        if path.is_file()
    }
    assert after == before


def test_forecast_run_artifacts_are_repo_ignored() -> None:
    ignore_text = (ROOT / ".gitignore").read_text(encoding="utf-8")
    assert "artifacts/forecast_runs/" in ignore_text
    assert "artifacts/forecast_uploads/" in ignore_text
    assert "!templates/NLTF_forecast_input_template_12q.xlsx" in ignore_text
    assert "!templates/NLTF_forecast_input_template_20q.xlsx" in ignore_text
    result = subprocess.run(["git", "ls-files"], cwd=ROOT, text=True, capture_output=True, check=True)
    tracked = set(result.stdout.splitlines())
    assert "templates/NLTF_forecast_input_template_20q.xlsx" in tracked
    assert not any(path.startswith("artifacts/forecast_runs/") for path in tracked)
    assert not any(path.startswith("artifacts/forecast_uploads/") for path in tracked)
    tracked_workbooks = {path for path in tracked if path.lower().endswith((".xlsx", ".xls"))}
    assert tracked_workbooks <= {
        "templates/NLTF_forecast_input_template_12q.xlsx",
        "templates/NLTF_forecast_input_template_20q.xlsx",
    }


def test_completed_workbook_numeric_all_streams_with_vnext(tmp_path) -> None:
    """With the vNext packs present and parity passing, PED and Heavy RUC must
    produce numeric fixed-finalist forecasts alongside Light RUC."""
    import model_dashboard.vnext_forward_integration as vfi

    if not (vfi.vnext_pack_present(ROOT, "PED") and vfi.vnext_pack_present(ROOT, "HEAVY_RUC")):
        import pytest

        pytest.skip("vNext packs not present")
    ped_audit = vfi.evaluate_vnext_forward_scorer(ROOT, "PED")
    heavy_audit = vfi.evaluate_vnext_forward_scorer(ROOT, "HEAVY_RUC")
    if not (ped_audit.forecast_capability_available and heavy_audit.forecast_capability_available):
        import pytest

        pytest.skip("vNext parity gates not passing in this environment")
    workbook = create_completed_sample_workbook(tmp_path / "completed.xlsx", repo_root=ROOT)
    result = run_forecast_workbook(
        workbook,
        output_dir=tmp_path / "vnext_run",
        repo_root=ROOT,
        run_timestamp="vnext-run",
        scenario_name="vnext_check",
    )
    future = result.future_forecasts
    for stream in STREAM_ORDER:
        rows = future[future["stream"].eq(stream)]
        assert rows["forecast_available"].astype(bool).all(), stream
        assert pd.to_numeric(rows["forecast"], errors="coerce").gt(0).all(), stream
    assert result.manifest["forecast_status"] == "numeric_forecast_available"
    assert sorted(result.manifest["numeric_forecast_streams"]) == sorted(STREAM_ORDER)
    assert result.manifest["governed_gap_streams"] == []
    capability = result.capability_report.set_index("stream")
    assert capability.loc["PED", "capability_status"] == "numeric_forecast_available"
    assert capability.loc["HEAVY_RUC", "capability_status"] == "numeric_forecast_available"
    assert str(capability.loc["PED", "model"]).startswith("PED__VNEXT")
    assert str(capability.loc["HEAVY_RUC", "model"]).startswith("HEAVY_RUC__VNEXT")


def test_100q_forecast_builder_current_finalists_all_streams_numeric_and_deterministic(tmp_path: Path) -> None:
    capabilities = model_capability_gap_register(ROOT).set_index("stream")
    assert capabilities["forecast_capability_available"].astype(bool).all()
    assert capabilities.loc["PED", "parity_status"] == "passed"
    assert capabilities.loc["HEAVY_RUC", "parity_status"] == "passed"
    assert float(capabilities.loc["PED", "stored_replay_max_delta"]) <= 1e-6
    assert float(capabilities.loc["HEAVY_RUC", "stored_replay_max_delta"]) <= 1e-6

    workbooks = {
        "basecase": create_completed_sample_workbook(
            tmp_path / "NLTF_forecast_input_template_basecase.xlsx",
            repo_root=ROOT,
            end_period="2050Q4",
        ),
        "high_population": create_completed_sample_workbook(
            tmp_path / "NLTF_forecast_input_template_high_population.xlsx",
            repo_root=ROOT,
            end_period="2050Q4",
            value_multiplier=1.02,
        ),
    }
    first_results = []
    for scenario, workbook in workbooks.items():
        first = run_forecast_workbook(
            workbook,
            output_dir=tmp_path / f"{scenario}_run_1",
            repo_root=ROOT,
            run_timestamp="determinism-1",
            scenario_name=scenario,
            scenario_role="basecase" if scenario == "basecase" else "comparison",
            expected_end_period="2050Q4",
        )
        second = run_forecast_workbook(
            workbook,
            output_dir=tmp_path / f"{scenario}_run_2",
            repo_root=ROOT,
            run_timestamp="determinism-2",
            scenario_name=scenario,
            scenario_role="basecase" if scenario == "basecase" else "comparison",
            expected_end_period="2050Q4",
        )
        first_results.append(first)
        pd.testing.assert_frame_equal(
            _sorted_forecast_values(first.future_forecasts),
            _sorted_forecast_values(second.future_forecasts),
            check_exact=True,
        )
        assert first.manifest["validation_status"] == "passed"
        assert first.manifest["scenario_role"] == ("basecase" if scenario == "basecase" else "comparison")
        assert first.manifest["forecast_status"] == "numeric_forecast_available"
        assert first.manifest["forecast_horizon_quarters"] == 100
        assert first.manifest["forecast_start_period"] == "2026Q1"
        assert first.manifest["forecast_end_period"] == "2050Q4"
        assert first.manifest["numeric_forecast_streams"] == sorted(STREAM_ORDER)
        assert first.manifest["governed_gap_streams"] == []
        assert first.manifest["fixed_finalists_only"] is True
        assert first.manifest["broad_search_run"] is False
        assert first.manifest["horizon_support_note"] == HORIZON_SUPPORT_NOTE
        assert len(first.future_forecasts) == 300
        model_by_stream = first.future_forecasts.drop_duplicates("stream").set_index("stream")["model"].to_dict()
        assert model_by_stream == {
            "PED": CURRENT_PED_FINALIST,
            "LIGHT_RUC": CURRENT_LIGHT_FINALIST,
            "HEAVY_RUC": CURRENT_HEAVY_FINALIST,
        }
        for stream in STREAM_ORDER:
            rows = first.future_forecasts[first.future_forecasts["stream"].eq(stream)].sort_values("target_period")
            values = pd.to_numeric(rows["forecast"], errors="coerce")
            assert len(rows) == 100
            assert rows["forecast_available"].astype(bool).all()
            assert values.notna().all() and np.isfinite(values).all()
            assert values.gt(0).all()
            pct_change = values.pct_change().dropna().abs()
            assert pct_change.lt(2.0).all(), stream
            assert set(rows["horizon_support_status"]) == {
                "backtest_supported_h1_12",
                "long_range_extrapolation_h13_plus",
            }

        vnext_components = first.component_forecasts[
            first.component_forecasts["stream"].isin(["PED", "HEAVY_RUC"])
            & ~first.component_forecasts["component_label"].astype(str).eq("FINAL")
        ].copy()
        reconciled = (
            vnext_components.groupby(["stream", "target_period"], dropna=False)["weighted_component_forecast"]
            .sum()
            .reset_index(name="component_weighted_sum")
            .merge(
                first.future_forecasts[["stream", "target_period", "forecast"]],
                on=["stream", "target_period"],
                how="inner",
            )
        )
        assert not reconciled.empty
        delta = (
            pd.to_numeric(reconciled["component_weighted_sum"], errors="coerce")
            - pd.to_numeric(reconciled["forecast"], errors="coerce")
        ).abs()
        assert float(delta.max()) <= 1e-6

    comparison = write_forecast_scenario_comparison(
        first_results,
        output_dir=tmp_path / "comparison",
        repo_root=ROOT,
        run_timestamp="determinism-comparison",
    )
    assert len(comparison.future_forecasts) == 600
    assert comparison.manifest["fixed_finalists_only"] is True
    assert comparison.manifest["broad_search_run"] is False
    assert comparison.manifest["horizon_support_note"] == HORIZON_SUPPORT_NOTE
    assert comparison.manifest["scenario_role_validation"]["base_scenario"] == "basecase"
    assert comparison.manifest["scenario_role_validation"]["comparison_scenarios"] == ["high_population"]
    assert "high_population_fixture_note" not in comparison.manifest
    assert "scenario_input_delta_audit.csv" in comparison.manifest["output_files"]
    for scenario in ["basecase", "high_population"]:
        rows = comparison.future_forecasts[comparison.future_forecasts["scenario_name"].eq(scenario)]
        assert len(rows) == 300
        assert set(rows["stream"]) == set(STREAM_ORDER)
        for stream in STREAM_ORDER:
            stream_rows = rows[rows["stream"].eq(stream)].sort_values("target_period")
            assert stream_rows["target_period"].tolist() == _expected_periods(100)
    audit = comparison.scenario_input_delta_audit
    assert len(audit) == 100 * sum(
        len([column for column in STREAM_COLUMNS[stream] if column.role == "user" and column.required])
        for stream in STREAM_ORDER
    )
    assert audit["base_scenario"].eq("basecase").all()
    assert audit["comparison_scenario"].eq("high_population").all()
    assert audit["base_scenario_role"].eq("basecase").all()
    assert audit["comparison_scenario_role"].eq("comparison").all()
    assert audit["decision_grade_status"].eq("scenario_delta_audit_only").all()
    assert not audit["all_required_inputs_plus_2pct"].fillna(False).astype(bool).any()
    assert set(audit["input_column"]).issuperset(
        {"unemployment_rate", "real_petrol_price_cents_per_litre", "real_diesel_price_cents_per_litre", "target_lag_1", "target_lag_4"}
    )
    assert audit["assumption_scope_note"].str.contains("not labelled as an all-inputs +2% smoke fixture", regex=False).all()
    summary = comparison.manifest["scenario_assumption_delta_summary"][0]
    assert summary["comparison_scenario"] == "high_population"
    assert summary["comparison_is_test_fixture"] is False
    assert summary["all_required_inputs_plus_2pct"] is False


def test_exact_2050q4_workbooks_preserve_forecasts_and_source_continuity(tmp_path: Path) -> None:
    base_path = os.environ.get("NLTF_EXACT_BASECASE_WORKBOOK")
    comparison_path = os.environ.get("NLTF_EXACT_COMPARISON_WORKBOOK")
    if not base_path or not comparison_path:
        pytest.skip("Set NLTF_EXACT_BASECASE_WORKBOOK and NLTF_EXACT_COMPARISON_WORKBOOK to run exact workbook regression.")
    base_workbook = Path(base_path)
    comparison_workbook = Path(comparison_path)
    if not base_workbook.exists() or not comparison_workbook.exists():
        pytest.skip("Exact workbook paths are not available in this environment.")

    before_hashes = _tracked_data_hashes()
    results = [
        run_forecast_workbook(
            base_workbook,
            output_dir=tmp_path / "exact_basecase_run",
            repo_root=ROOT,
            workbook_filename=base_workbook.name,
            run_timestamp="exact-env",
            expected_end_period="2050Q4",
        ),
        run_forecast_workbook(
            comparison_workbook,
            output_dir=tmp_path / "exact_comparison_run",
            repo_root=ROOT,
            workbook_filename=comparison_workbook.name,
            run_timestamp="exact-env",
            expected_end_period="2050Q4",
        ),
    ]

    for result in results:
        assert result.manifest["forecast_status"] == "numeric_forecast_available"
        assert result.manifest["forecast_horizon_quarters"] == 100
        assert result.manifest["forecast_start_period"] == "2026Q1"
        assert result.manifest["forecast_end_period"] == "2050Q4"
        assert len(result.future_forecasts) == 300
        assert set(result.future_forecasts["model"]) == {CURRENT_PED_FINALIST, CURRENT_LIGHT_FINALIST, CURRENT_HEAVY_FINALIST}
        _assert_source_derived_continuity(result)

    comparison = write_forecast_scenario_comparison(
        results,
        output_dir=tmp_path / "exact_comparison",
        repo_root=ROOT,
        run_timestamp="exact-env-comparison",
    )
    assert len(comparison.future_forecasts) == 600
    assert comparison.manifest["scenario_role_validation"]["base_scenario"] == results[0].manifest["scenario_name"]
    assert comparison.manifest["scenario_role_validation"]["comparison_scenarios"] == [results[1].manifest["scenario_name"]]
    assert "high_population_fixture_note" not in comparison.manifest
    assert comparison.manifest["evidence_pack_modified"] is False
    assert comparison.manifest["chart_sources_modified"] is False

    audit = comparison.scenario_input_delta_audit
    assert audit["base_scenario_role"].eq("basecase").all()
    assert audit["comparison_scenario_role"].eq("comparison").all()
    assert audit["base_original_filename"].eq(base_workbook.name).all()
    assert audit["comparison_original_filename"].eq(comparison_workbook.name).all()
    assert audit["base_workbook_sha256"].eq(hashlib.sha256(base_workbook.read_bytes()).hexdigest()).all()
    assert audit["comparison_workbook_sha256"].eq(hashlib.sha256(comparison_workbook.read_bytes()).hexdigest()).all()
    assert audit["decision_grade_status"].eq("scenario_delta_audit_only").all()
    assert not audit["all_required_inputs_plus_2pct"].fillna(False).astype(bool).any()
    sample = audit[pd.to_numeric(audit["pct_delta"], errors="coerce").notna()].iloc[0]
    expected_delta = float(sample["comparison_value"]) / float(sample["base_value"]) - 1.0
    assert abs(float(sample["pct_delta"]) - expected_delta) <= 1e-12

    baseline_dir = os.environ.get("NLTF_EXACT_BASELINE_COMPARISON_DIR")
    if baseline_dir:
        baseline_path = Path(baseline_dir) / "forecast_scenario_comparison.parquet"
        assert baseline_path.exists(), f"Requested exact baseline comparison does not exist: {baseline_path}"
        baseline = pd.read_parquet(baseline_path)
        pd.testing.assert_frame_equal(
            _forecast_values_by_role(comparison.future_forecasts),
            _forecast_values_by_role(baseline),
            check_exact=True,
        )

    assert _tracked_data_hashes() == before_hashes


def test_vnext_y_lag_members_ignore_future_workbook_target_lag_leakage(tmp_path: Path) -> None:
    base = create_completed_sample_workbook(
        tmp_path / "base_target_lags.xlsx",
        repo_root=ROOT,
        end_period="2050Q4",
    )
    mutated = create_completed_sample_workbook(
        tmp_path / "mutated_target_lags.xlsx",
        repo_root=ROOT,
        end_period="2050Q4",
    )
    _multiply_target_lag_columns(mutated, ("PED", "HEAVY_RUC"), multiplier=10_000.0)
    base_result = run_forecast_workbook(
        base,
        output_dir=tmp_path / "base_run",
        repo_root=ROOT,
        run_timestamp="lag-base",
        scenario_name="base_lags",
        expected_end_period="2050Q4",
    )
    mutated_result = run_forecast_workbook(
        mutated,
        output_dir=tmp_path / "mutated_run",
        repo_root=ROOT,
        run_timestamp="lag-mutated",
        scenario_name="mutated_lags",
        expected_end_period="2050Q4",
    )
    for stream in ["PED", "HEAVY_RUC"]:
        base_rows = _sorted_forecast_values(base_result.future_forecasts[base_result.future_forecasts["stream"].eq(stream)])
        mutated_rows = _sorted_forecast_values(mutated_result.future_forecasts[mutated_result.future_forecasts["stream"].eq(stream)])
        pd.testing.assert_frame_equal(base_rows, mutated_rows, check_exact=True)
