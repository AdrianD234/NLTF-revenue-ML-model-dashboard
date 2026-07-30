"""Blocking gates for the governed quarterly-actuals refresh (2026Q1 vintage).

Covers the workbook snapshot contract, the canonical-history invariants, the
identity checks, the PED/heavy governance rules, and the plug-and-play
contract proving a later workbook with the same three main sheet schemas
refreshes the next quarter without code edits.
"""
from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "references" / "NLTF_model_input_sheet_actuals_to_2026Q1_complete1.xlsx"
WORKBOOK_SHA256 = "be951103cfa0fc4415583044397eea11982c2d83b0679bd97984cb0b0cf305a3"
HISTORY_DIR = ROOT / "data" / "model_input_history"
ARTIFACTS = ROOT / "artifacts" / "actuals_refresh_2026q1"

sys.path.insert(0, str(ROOT))

from scripts import refresh_model_actuals as rma  # noqa: E402

STREAM_FILES = {
    "ped": "ped_inputs.parquet",
    "light_ruc": "light_ruc_inputs.parquet",
    "heavy_ruc": "heavy_ruc_inputs.parquet",
}

SENTINELS = {
    "light_ruc": {
        "target": 3_196_014_020.0,
        "light_ruc_revenue_nzd": 248_999_340.25,
        "light_ruc_price_nominal_nzd_per_1000km": 77.90933916178503,
        "diesel_price_nominal_cents_per_litre": 214.6,
    },
    "heavy_ruc": {
        "target": 1_039_372_457.0,
        "heavy_ruc_revenue_nzd": 334_926_318.92,
        "heavy_ruc_price_nominal_nzd_per_1000km": 322.2389785916754,
        "lead_real_heavy_ruc_price_nzd_per_1000km": 318.7803374122621,
    },
}
PED_BRIDGE_TARGET = 1355.8193183068997
PED_BRIDGE_VKT = 7_268_954_111.238782


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _history(stream: str) -> pd.DataFrame:
    return pd.read_parquet(HISTORY_DIR / STREAM_FILES[stream])


# --------------------------------------------------------------- workbook
class TestWorkbookSnapshot:
    def test_sha256_matches_the_governed_snapshot(self) -> None:
        assert WORKBOOK.exists(), "immutable source workbook snapshot is missing"
        assert _sha256(WORKBOOK) == WORKBOOK_SHA256

    def test_all_seven_sheets_present_with_expected_ranges(self) -> None:
        import openpyxl

        wb = openpyxl.load_workbook(WORKBOOK)
        assert wb.sheetnames == [
            "PED Inputs",
            "Light RUC Inputs",
            "Heavy RUC Inputs",
            "Q1 2026 Filled Rows",
            "Q1 2026 Source Map",
            "Q1 2026 Reconciliation",
            "Agent Build Spec",
        ]
        dims = {name: wb[name].dimensions for name in wb.sheetnames}
        assert dims["PED Inputs"] == "A1:AK98"
        assert dims["Light RUC Inputs"] == "A1:AW98"
        assert dims["Heavy RUC Inputs"] == "A1:AL98"

    def test_97_continuous_quarters_per_stream(self) -> None:
        xl = pd.ExcelFile(WORKBOOK)
        for sheet in ("PED Inputs", "Light RUC Inputs", "Heavy RUC Inputs"):
            frame = xl.parse(sheet)
            periods = frame["period"].astype(str).tolist()
            assert len(periods) == 97
            assert periods[0] == "2002Q1" and periods[-1] == "2026Q1"
            keys = [rma.quarter_sort_key(p) for p in periods]
            assert keys == list(range(keys[0], keys[0] + 97)), "quarter sequence must be continuous"
            assert len(set(periods)) == 97, "no duplicate stream/period"

    def test_sentinel_q1_values(self) -> None:
        xl = pd.ExcelFile(WORKBOOK)
        for stream, sheet in (("light_ruc", "Light RUC Inputs"), ("heavy_ruc", "Heavy RUC Inputs")):
            row = xl.parse(sheet).query("period == '2026Q1'").iloc[0]
            for column, expected in SENTINELS[stream].items():
                assert float(row[column]) == pytest.approx(expected, rel=1e-12), (stream, column)
        ped = xl.parse("PED Inputs").query("period == '2026Q1'").iloc[0]
        assert float(ped["target"]) == pytest.approx(PED_BRIDGE_TARGET, rel=1e-12)
        assert float(ped["light_petrol_vkt_total_km"]) == pytest.approx(PED_BRIDGE_VKT, rel=1e-12)
        assert float(ped["population"]) == 5_361_300
        assert float(ped["unemployment_percent"]) == pytest.approx(5.3)

    def test_source_map_covers_every_populated_model_column(self) -> None:
        lineage = pd.read_csv(ARTIFACTS / "source_lineage.csv")
        cells = lineage[lineage.columns[1]].dropna().astype(str)
        mapped = {c.split("!")[0].strip() for c in cells if "!" in c}
        assert {"PED Inputs", "Light RUC Inputs", "Heavy RUC Inputs"}.issubset(mapped)
        # 37 + 49 (minus 6 shared identifier repeats...) - assert full per-sheet counts.
        counts = cells[cells.str.contains("!")].str.split("!").str[0].value_counts()
        assert counts["PED Inputs"] == 39  # 37 columns + data_status + notes rows
        assert counts["Light RUC Inputs"] == 49
        assert counts["Heavy RUC Inputs"] == 38


# --------------------------------------------------------------- history
class TestCanonicalHistory:
    def test_2026q1_appended_once_and_continuous(self) -> None:
        for stream in STREAM_FILES:
            hist = _history(stream)
            periods = hist["period"].astype(str).tolist()
            assert periods.count("2026Q1") == 1
            keys = [rma.quarter_sort_key(p) for p in periods]
            assert keys == list(range(keys[0], keys[0] + len(keys)))

    def test_accepted_targets_and_ped_placeholder(self) -> None:
        light = _history("light_ruc").query("period == '2026Q1'").iloc[0]
        heavy = _history("heavy_ruc").query("period == '2026Q1'").iloc[0]
        ped = _history("ped").query("period == '2026Q1'").iloc[0]
        assert float(light["target"]) == SENTINELS["light_ruc"]["target"]
        assert float(heavy["target"]) == SENTINELS["heavy_ruc"]["target"]
        assert float(ped["target"]) == 0.0, "provisional PED bridge must never enter the fitting target column"
        assert float(ped["light_petrol_vkt_total_km"]) == 0.0
        assert str(ped["data_status"]).lower().startswith("provisional")

    def test_lags_roll_from_2025q4_and_2025q1(self) -> None:
        for stream in STREAM_FILES:
            hist = _history(stream).set_index("period")
            q1 = hist.loc["2026Q1"]
            assert float(q1["target_lag_1"]) == pytest.approx(float(hist.loc["2025Q4", "target"]), rel=1e-12)
            assert float(q1["target_lag_4"]) == pytest.approx(float(hist.loc["2025Q1", "target"]), rel=1e-12)

    def test_positive_log_fields_finite_for_accepted_rows(self) -> None:
        for stream in ("light_ruc", "heavy_ruc"):
            q1 = _history(stream).query("period == '2026Q1'").iloc[0]
            for column in ("log_target", "log_real_gdp", "log_trend"):
                assert np.isfinite(float(q1[column])), (stream, column)

    def test_provisional_sidecar_matches_the_workbook_bridge(self) -> None:
        sidecar = json.loads((HISTORY_DIR / "ped_provisional_bridge.json").read_text(encoding="utf-8"))
        entry = sidecar["entries"][-1]
        assert entry["period"] == "2026Q1"
        assert entry["provisional_target_km_per_person"] == pytest.approx(PED_BRIDGE_TARGET, rel=1e-12)
        assert entry["target_method"] == "mbu26_residual_core_ped_share"
        assert entry["eligible_for_refit"] is False
        assert entry["display_as_observed_actual"] is False
        assert sidecar["governance"]["never_fit_on_provisional"] is True


# --------------------------------------------------------------- identities
class TestIdentities:
    def test_rate_identities_at_q1(self) -> None:
        light = _history("light_ruc").query("period == '2026Q1'").iloc[0]
        heavy = _history("heavy_ruc").query("period == '2026Q1'").iloc[0]
        assert float(light["light_ruc_price_nominal_nzd_per_1000km"]) == pytest.approx(
            float(light["light_ruc_revenue_nzd"]) / float(light["target"]) * 1000.0, rel=1e-9
        )
        assert float(heavy["heavy_ruc_price_nominal_nzd_per_1000km"]) == pytest.approx(
            float(heavy["heavy_ruc_revenue_nzd"]) / float(heavy["target"]) * 1000.0, rel=1e-9
        )

    def test_gdp_identity_at_q1(self) -> None:
        for stream in ("light_ruc", "heavy_ruc"):
            q1 = _history(stream).query("period == '2026Q1'").iloc[0]
            assert float(q1["real_gdp_sa_nzd"]) == pytest.approx(
                float(q1["nominal_gdp_sa_nzd"]) * float(q1["gdp_rebasing_factor"]), rel=1e-12
            )

    def test_ped_bridge_identity(self) -> None:
        sidecar = json.loads((HISTORY_DIR / "ped_provisional_bridge.json").read_text(encoding="utf-8"))
        entry = sidecar["entries"][-1]
        assert entry["provisional_target_km_per_person"] == pytest.approx(
            entry["provisional_light_petrol_vkt_total_km"] / entry["population"], rel=1e-12
        )

    def test_cross_stream_shared_inputs_agree_at_q1(self) -> None:
        light = _history("light_ruc").query("period == '2026Q1'").iloc[0]
        heavy = _history("heavy_ruc").query("period == '2026Q1'").iloc[0]
        ped = _history("ped").query("period == '2026Q1'").iloc[0]
        assert float(heavy["real_diesel_price_cents_per_litre"]) == pytest.approx(
            float(light["real_diesel_price_cents_per_litre"]), rel=1e-12
        )
        assert float(heavy["real_light_ruc_price_nzd_per_1000km"]) == pytest.approx(
            float(light["real_light_ruc_price_nzd_per_1000km"]), rel=1e-12
        )
        assert float(heavy["unemployment_rate"]) == pytest.approx(float(ped["unemployment_rate"]), rel=1e-12)

    def test_ped_annual_residual_bridge_reconciles(self) -> None:
        recon = pd.read_csv(ARTIFACTS / "q1_reconciliation.csv")
        label_col, value_col = recon.columns[0], recon.columns[2]
        lookup = {
            str(row[label_col]).strip(): row[value_col]
            for _, row in recon.iterrows()
            if pd.notna(row[label_col])
        }
        annual = float(lookup["MBU26 FY2026 Light petrol VKT"])
        q3 = float(lookup["Less actual 2025Q3 VKT"])
        q4 = float(lookup["Less actual 2025Q4 VKT"])
        q1 = float(lookup["Q1 petrol-only VKT"])
        q2 = float(lookup["Q2 petrol-only VKT"])
        assert q3 + q4 + q1 + q2 == pytest.approx(annual, rel=1e-12)


# --------------------------------------------------------------- governance
class TestGovernance:
    def test_ped_provisional_never_enters_estimation(self) -> None:
        state = json.loads(
            (
                ROOT
                / "data"
                / "dashboard_evidence_pack_reproducibility"
                / "ped_ar1"
                / "ar1_fitted_state.json"
            ).read_text(encoding="utf-8")
        )
        assert state["latest_actual"] == "2025Q4"
        assert state["train_window_end"] == "2025Q4"

    def test_per_stream_latest_accepted_periods(self) -> None:
        from model_dashboard.forecast_runner import stream_latest_accepted_periods

        latest = stream_latest_accepted_periods(ROOT)
        assert latest == {"PED": "2025Q4", "LIGHT_RUC": "2026Q1", "HEAVY_RUC": "2026Q1"}

    def test_pack_vintage_status_exposes_per_stream_seam(self) -> None:
        for pack in (
            ROOT / "data" / "current_revenue_outlook",
            ROOT / "data" / "engine_ar1" / "current_revenue_outlook",
        ):
            frame = pd.read_csv(pack / "stream_vintage_status.csv")
            row = frame.set_index("stream")
            assert row.loc["PED", "latest_accepted_exact_actual"] == "2025Q4"
            assert row.loc["LIGHT_RUC", "latest_accepted_exact_actual"] == "2026Q1"
            assert row.loc["HEAVY_RUC", "latest_accepted_exact_actual"] == "2026Q1"
            assert row.loc["PED", "first_forecast_quarter"] == "2026Q1"
            assert row.loc["LIGHT_RUC", "first_forecast_quarter"] == "2026Q2"
            assert "not an observed actual" in str(row.loc["PED", "provisional_seed"])

    def test_ped_q1_not_displayed_as_observed_actual(self) -> None:
        for pack in (
            ROOT / "data" / "current_revenue_outlook",
            ROOT / "data" / "engine_ar1" / "current_revenue_outlook",
        ):
            rows = pd.read_csv(pack / "revenue_chart_rows.csv", low_memory=False)
            ped_hist = rows[
                rows["row_type"].astype(str).eq("historical_actual")
                & rows["series_id"].astype(str).eq("ped_vkt_per_capita")
                & rows["time_grain"].astype(str).eq("quarterly")
            ]
            assert "2026Q1" not in set(ped_hist["period"].astype(str))
            light_hist = rows[
                rows["row_type"].astype(str).eq("historical_actual")
                & rows["series_id"].astype(str).eq("light_ruc_net_km")
                & rows["time_grain"].astype(str).eq("quarterly")
            ]
            assert "2026Q1" in set(light_hist["period"].astype(str))

    def test_no_superseded_forecast_rows_remain(self) -> None:
        for pack in (
            ROOT / "data" / "current_revenue_outlook",
            ROOT / "data" / "engine_ar1" / "current_revenue_outlook",
        ):
            rows = pd.read_csv(pack / "revenue_chart_rows.csv", low_memory=False)
            forecast = rows[
                rows["row_type"].astype(str).eq("future_forecast")
                & rows["time_grain"].astype(str).eq("quarterly")
            ]
            for stream in ("LIGHT_RUC", "HEAVY_RUC"):
                periods = set(forecast[forecast["stream"].astype(str).eq(stream)]["period"].astype(str))
                assert "2026Q1" not in periods, (pack, stream)
            assert "2026Q1" in set(
                forecast[forecast["stream"].astype(str).eq("PED")]["period"].astype(str)
            )

    def test_heavy_lead_labelled_retrospective(self) -> None:
        audit = pd.read_csv(ARTIFACTS / "heavy_lead_vintage_audit.csv")
        q1 = audit[
            audit["period"].astype(str).eq("2026Q1")
            & audit["column"].eq("lead_real_heavy_ruc_price_nzd_per_1000km")
        ].iloc[0]
        assert q1["vintage_status"] == "retrospective_history"
        q4 = audit[
            audit["period"].astype(str).eq("2025Q4")
            & audit["column"].eq("lead_real_heavy_ruc_price_nzd_per_1000km")
        ].iloc[0]
        assert bool(q4["applied_to_canonical_history"]) is False
        # And the canonical 2025Q4 lead is untouched (placeholder retained).
        heavy = _history("heavy_ruc").set_index("period")
        assert float(heavy.loc["2025Q4", "lead_real_heavy_ruc_price_nzd_per_1000km"]) == 0.0

    def test_promoted_heavy_state_uses_no_lead_features(self) -> None:
        diag = pd.read_csv(ARTIFACTS / "heavy_lead_replay_diagnostic.csv")
        assert int(diag["lead_feature_count"].iloc[0]) == 0


# --------------------------------------------------------------- plug-and-play
def _write_values_workbook(path: Path, frames: dict[str, pd.DataFrame]) -> None:
    """Values-only workbook with the three canonical main sheets."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_names = {"ped": "PED Inputs", "light_ruc": "Light RUC Inputs", "heavy_ruc": "Heavy RUC Inputs"}
    for stream, frame in frames.items():
        ws = wb.create_sheet(sheet_names[stream])
        ws.append([str(c) for c in frame.columns])
        for _, row in frame.iterrows():
            values = []
            for value in row.tolist():
                if isinstance(value, (np.integer,)):
                    value = int(value)
                elif isinstance(value, (np.floating,)):
                    value = float(value)
                elif pd.isna(value):
                    value = None
                values.append(value)
            ws.append(values)
    wb.save(path)


def _next_quarter_frames(history_dir: Path) -> dict[str, pd.DataFrame]:
    """History frames plus a derived 2026Q2 row per stream (via the central
    derivation path; this proves append plumbing, while Q1 parity against the
    Excel-computed workbook proves derivation correctness)."""
    frames: dict[str, pd.DataFrame] = {
        stream: pd.read_parquet(history_dir / filename) for stream, filename in STREAM_FILES.items()
    }
    period = "2026Q2"
    raw = {
        "ped": {
            "nominal_gdp_sa_nzd": 115_600_000_000.0,
            "population": 5_372_000.0,
            "gdp_deflator_sa": 1572.0,
            "petrol_price_nominal_cents_per_litre": 271.2,
            "cpi_rebasing_factor": 1339.0 / 1359.0,
            "light_petrol_vkt_total_km": 7_900_000_000.0,
            "unemployment_percent": 5.2,
        },
        "light_ruc": {
            "target": 3_250_000_000.0,
            "nominal_gdp_sa_nzd": 115_600_000_000.0,
            "gdp_deflator_sa": 1572.0,
            "diesel_price_nominal_cents_per_litre": 216.0,
            "cpi_rebasing_factor": 1339.0 / 1359.0,
            "light_ruc_revenue_nzd": 253_500_000.0,
        },
        "heavy_ruc": {
            "target": 1_050_000_000.0,
            "nominal_gdp_sa_nzd": 115_600_000_000.0,
            "gdp_deflator_sa": 1572.0,
            "heavy_ruc_revenue_nzd": 338_000_000.0,
        },
    }
    notes = {"data_status": "Historical official/source data", "notes": "synthetic plug-and-play test row"}
    pending: dict[str, list] = {s: [] for s in STREAM_FILES}
    cross: dict[str, dict] = {}
    out: dict[str, pd.DataFrame] = {}
    for stream in ("ped", "light_ruc", "heavy_ruc"):
        base_deflator = rma.gdp_rebase_base_deflator(frames[stream])
        row = rma.derive_stream_row(
            stream, period, raw[stream], dict(notes), frames[stream], pending, base_deflator, cross
        )
        row = {k: v for k, v in row.items() if k in set(frames[stream].columns)}
        cross[stream] = row
        pending[stream].append(row)
        appended, _how = rma.replace_or_append(frames[stream], row)
        out[stream] = appended
    return out


@pytest.fixture()
def temp_refresh_env(tmp_path: Path) -> dict[str, Path]:
    history_dir = tmp_path / "model_input_history"
    history_dir.mkdir()
    for filename in STREAM_FILES.values():
        shutil.copy2(HISTORY_DIR / filename, history_dir / filename)
    shutil.copy2(HISTORY_DIR / "manifest.json", history_dir / "manifest.json")
    return {"history_dir": history_dir, "out_dir": tmp_path / "out", "tmp": tmp_path}


def _run_refresh(workbook: Path, env: dict[str, Path], *extra: str) -> int:
    argv = [
        "--workbook", str(workbook),
        "--output-dir", str(env["out_dir"]),
        "--history-dir", str(env["history_dir"]),
        *extra,
    ]
    return rma.main(argv)


class TestPlugAndPlayContract:
    def test_next_quarter_append_works_without_code_changes(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        workbook = env["tmp"] / "next_quarter.xlsx"
        _write_values_workbook(workbook, frames)
        before_sha = _sha256(workbook)
        assert _run_refresh(workbook, env) == 0
        hist = pd.read_parquet(env["history_dir"] / "light_ruc_inputs.parquet")
        assert "2026Q2" in set(hist["period"].astype(str))
        assert float(hist.set_index("period").loc["2026Q2", "target"]) == 3_250_000_000.0
        # PED provisional row remains excluded from fitting after the append.
        ped = pd.read_parquet(env["history_dir"] / "ped_inputs.parquet").set_index("period")
        assert float(ped.loc["2026Q1", "target"]) == 0.0
        # The source workbook is immutable.
        assert _sha256(workbook) == before_sha

    def test_duplicate_period_is_idempotent_only_when_identical(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        workbook = env["tmp"] / "next_quarter.xlsx"
        _write_values_workbook(workbook, frames)
        assert _run_refresh(workbook, env) == 0
        # Identical rerun: governed no-op.
        assert _run_refresh(workbook, env) == 0

    def test_conflicting_replacement_fails_closed(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        workbook = env["tmp"] / "next_quarter.xlsx"
        _write_values_workbook(workbook, frames)
        assert _run_refresh(workbook, env) == 0
        # Same period, different accepted value: must fail.
        conflicted = {s: f.copy() for s, f in frames.items()}
        idx = conflicted["light_ruc"].index[conflicted["light_ruc"]["period"].astype(str) == "2026Q2"][0]
        conflicted["light_ruc"].at[idx, "target"] = 3_333_333_333.0
        workbook2 = env["tmp"] / "conflict.xlsx"
        _write_values_workbook(workbook2, conflicted)
        assert _run_refresh(workbook2, env) == 1

    def test_missing_intermediate_quarter_fails(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        skipped = {}
        for stream, frame in frames.items():
            frame = frame.copy()
            idx = frame.index[frame["period"].astype(str) == "2026Q2"][0]
            frame.at[idx, "period"] = "2026Q3"
            skipped[stream] = frame
        workbook = env["tmp"] / "gap.xlsx"
        _write_values_workbook(workbook, skipped)
        assert _run_refresh(workbook, env) == 1

    def test_changed_header_fails(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        renamed = {s: f.copy() for s, f in frames.items()}
        renamed["light_ruc"] = renamed["light_ruc"].rename(columns={"target": "target_km_renamed"})
        workbook = env["tmp"] / "renamed.xlsx"
        _write_values_workbook(workbook, renamed)
        assert _run_refresh(workbook, env) == 1

    def test_unknown_unit_scale_fails(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        wrong = {s: f.copy() for s, f in frames.items()}
        idx = wrong["light_ruc"].index[wrong["light_ruc"]["period"].astype(str) == "2026Q2"][0]
        # Million-km instead of km: three orders of magnitude off.
        wrong["light_ruc"].at[idx, "target"] = 3_250.0
        workbook = env["tmp"] / "wrong_unit.xlsx"
        _write_values_workbook(workbook, wrong)
        assert _run_refresh(workbook, env) == 1

    def test_ped_accepted_mode_requires_governance_approval(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        workbook = env["tmp"] / "next_quarter.xlsx"
        _write_values_workbook(workbook, frames)
        assert _run_refresh(workbook, env, "--ped-mode", "accepted") == 1
        assert (
            _run_refresh(
                workbook, env, "--ped-mode", "accepted", "--governance-approval", "test-owner: decision ref"
            )
            == 0
        )
        ped = pd.read_parquet(env["history_dir"] / "ped_inputs.parquet").set_index("period")
        # Under an explicit governed acceptance the PED target may enter history.
        assert float(ped.loc["2026Q2", "target"]) > 0.0

    def test_provisional_ped_mode_never_writes_target(self, temp_refresh_env) -> None:
        env = temp_refresh_env
        frames = _next_quarter_frames(env["history_dir"])
        workbook = env["tmp"] / "next_quarter.xlsx"
        _write_values_workbook(workbook, frames)
        assert _run_refresh(workbook, env, "--ped-mode", "provisional_replay_only") == 0
        ped = pd.read_parquet(env["history_dir"] / "ped_inputs.parquet").set_index("period")
        assert float(ped.loc["2026Q2", "target"]) == 0.0
        sidecar = json.loads((env["history_dir"] / "ped_provisional_bridge.json").read_text(encoding="utf-8"))
        assert sidecar["entries"][-1]["period"] == "2026Q2"
