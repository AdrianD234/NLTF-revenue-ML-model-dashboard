"""Plug-and-play contract for future official vintages.

A future BEFU/PREFU/MBU workbook with the governed schema must require only a
registry entry and one materialisation command — no code changes. These tests
drive synthetic workbooks through the real materializer inside a temporary
repo root and prove both the happy path and every fail-closed path.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Callable

import pandas as pd
import pytest

from model_dashboard import official_vintage as ov

START_FY = 2001
ACTUAL_END_FY = 2025
ST_END_FY = 2030
END_FY = 2055

_ACTIVITY_LABELS = [
    item["source_label"]
    for item in ov.CANONICAL_SERIES_DEFINITIONS
    if item["section"] == "Key volumes"
]
_SECTION_LABELS = {
    section: [
        item["source_label"]
        for item in ov.CANONICAL_SERIES_DEFINITIONS
        if item["section"] == section
    ]
    for section in ("RUC", "FED", "MVR", "TUC", "Totals")
}


def _series_values(fy: int) -> dict[str, float]:
    i = float(fy - START_FY)
    v: dict[str, float] = {
        "light_ruc_net_km": 1000.0 + i,
        "heavy_ruc_net_km": 500.0 + i,
        "light_bev_ruc_net_km": 10.0 + 0.5 * i,
        "heavy_bev_ruc_net_km": 5.0,
        "phev_ruc_net_km": 3.0,
        "ped_volume": 3000.0 - i,
        "light_petrol_vkt": 20000.0 - 2.0 * i,
        "ped_vkt_per_capita": 4000.0 - i,
        "tuc_gtk": 1.0e9,
        "light_ruc_net_revenue": 100.0 + i,
        "heavy_ruc_net_revenue": 200.0 + i,
        "light_bev_ruc_net_revenue": 1.0 + 0.1 * i,
        "heavy_bev_ruc_net_revenue": 2.0,
        "phev_ruc_net_revenue": 3.0,
        "ruc_refunds": 5.0,
        "ruc_admin_revenue": 4.0,
        "gross_ped_revenue": 500.0 + i,
        "gross_lpg_revenue": 10.0,
        "gross_cng_revenue": 1.0,
        "fed_refunds": 6.0,
        "mr1_revenue": 50.0 + 0.2 * i,
        "mr2_revenue": 10.0,
        "coo_revenue": 2.0,
        "mvr_admin_revenue": 7.0,
        "mvr_refunds": 1.0,
        "tuc_net_revenue": 8.0,
    }
    v["gross_ruc_revenue"] = (
        v["light_ruc_net_revenue"]
        + v["heavy_ruc_net_revenue"]
        + v["light_bev_ruc_net_revenue"]
        + v["heavy_bev_ruc_net_revenue"]
        + v["phev_ruc_net_revenue"]
        + v["ruc_refunds"]
    )
    v["ruc_revenue_net_admin"] = v["gross_ruc_revenue"] - v["ruc_admin_revenue"]
    v["total_ruc_net_revenue"] = v["ruc_revenue_net_admin"] - v["ruc_refunds"]
    v["gross_fed_revenue"] = v["gross_ped_revenue"] + v["gross_lpg_revenue"] + v["gross_cng_revenue"]
    v["net_fed_revenue"] = v["gross_fed_revenue"] - v["fed_refunds"]
    v["gross_mvr_revenue"] = v["mr1_revenue"] + v["mr2_revenue"] + v["coo_revenue"]
    v["mvr_revenue_net_admin_coo"] = v["mr1_revenue"] + v["mr2_revenue"] - v["mvr_admin_revenue"]
    v["net_mvr_revenue"] = v["mvr_revenue_net_admin_coo"] - v["mvr_refunds"]
    v["total_gross_revenue"] = (
        v["gross_ruc_revenue"] + v["gross_fed_revenue"] + v["gross_mvr_revenue"] + v["tuc_net_revenue"]
    )
    v["total_admin_fees"] = v["ruc_admin_revenue"] + v["mvr_admin_revenue"] + v["coo_revenue"]
    v["total_revenue_net_admin"] = v["total_gross_revenue"] - v["total_admin_fees"]
    v["total_refunds"] = v["ruc_refunds"] + v["fed_refunds"] + v["mvr_refunds"]
    v["total_nltf_net_revenue"] = v["total_revenue_net_admin"] - v["total_refunds"]
    return v


def write_vintage_workbook(
    path: Path,
    *,
    sheet: str = "Baseline",
    end_fy: int = END_FY,
    mutate: Callable[[Any], None] | None = None,
) -> None:
    import openpyxl

    label_by_id = {
        str(item["series_id"]): str(item["source_label"])
        for item in ov.CANONICAL_SERIES_DEFINITIONS
    }
    id_by_label = {label: series_id for series_id, label in label_by_id.items()}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    years = list(range(START_FY, end_fy + 1))
    ws.cell(1, 1, "YE June")
    ws.cell(2, 1, "Period")
    for offset, fy in enumerate(years):
        column = 2 + offset
        ws.cell(1, column, fy)
        status = (
            "ACTUAL"
            if fy <= ACTUAL_END_FY
            else "ST_FORECAST"
            if fy <= ST_END_FY
            else "LT_FORECAST"
        )
        ws.cell(2, column, status)

    def write_block(anchor_row: int, anchor_label: str, labels: list[str], *, values: bool) -> None:
        ws.cell(anchor_row, 1, anchor_label)
        for index, label in enumerate(labels):
            row = anchor_row + 1 + index
            ws.cell(row, 1, label)
            for offset, fy in enumerate(years):
                if values:
                    ws.cell(row, 2 + offset, _series_values(fy)[id_by_label[label]])
                else:
                    ws.cell(row, 2 + offset, 0.0 if fy > START_FY else None)

    write_block(4, "Key volumes: Level", _ACTIVITY_LABELS, values=True)
    write_block(15, "Key volumes: Annual percentage changes", _ACTIVITY_LABELS, values=False)
    ws.cell(26, 1, "Revenues: Level ($m ex GST)")
    write_block(27, "Road User Charges", _SECTION_LABELS["RUC"], values=True)
    write_block(39, "Fuel Excise Duties", _SECTION_LABELS["FED"], values=True)
    write_block(47, "Motor Vehicle Register", _SECTION_LABELS["MVR"], values=True)
    write_block(57, "Track User Charges", _SECTION_LABELS["TUC"], values=True)
    write_block(60, "TOTALS", _SECTION_LABELS["Totals"], values=True)
    if mutate is not None:
        mutate(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


BASE_ENTRY = {
    "vintage_id": "BASE0",
    "display_name": "BASE0 official",
    "release_round": "BASE0",
    "release_date": None,
    "source_workbook": "references/base0.xlsx",
    "workbook_sha256": "0" * 64,
    "source_sheet": "Baseline",
    "schema_version": ov.OFFICIAL_VINTAGE_PACK_SCHEMA_VERSION,
    "actual_start_fy": START_FY,
    "actual_end_fy": ACTUAL_END_FY,
    "short_forecast_start_fy": ACTUAL_END_FY + 1,
    "short_forecast_end_fy": ST_END_FY,
    "long_forecast_start_fy": ST_END_FY + 1,
    "long_forecast_end_fy": END_FY,
    "source_horizon_fy": END_FY,
    "source_pack_path": "data/revenue_model_source_pack/official_vintages/base0",
    "pack_format": "official_vintage_v1",
    "file_stems": dict(ov.GENERIC_FILE_STEMS),
    "layout": {"year_header_row": 1, "period_status_row": 2, "label_column": 1},
    "available": True,
    "is_latest": True,
    "is_default_comparator": True,
    "is_default_bridge_vintage": True,
    "status": "active_official_vintage",
}


@pytest.fixture()
def temp_root(tmp_path: Path) -> Path:
    root = tmp_path / "repo"
    registry = {
        "schema_version": ov.OFFICIAL_VINTAGE_REGISTRY_SCHEMA_VERSION,
        "vintages": [json.loads(json.dumps(BASE_ENTRY))],
    }
    target = root / ov.OFFICIAL_VINTAGE_REGISTRY_PATH
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(registry, indent=2) + "\n", encoding="utf-8")
    return root


def _materialize(root: Path, vintage_id: str, workbook: Path, **kwargs: Any) -> dict[str, Any]:
    return ov.materialize_official_vintage(
        vintage_id,
        workbook_path=workbook,
        repo_root=root,
        sheet=kwargs.pop("sheet", "Baseline"),
        **kwargs,
    )


class TestPlugAndPlay:
    def test_prefu26_style_vintage_ingests_without_code_changes(self, temp_root: Path):
        workbook = temp_root / "references" / "PREFU26 revenue forecast.xlsx"
        write_vintage_workbook(workbook)
        manifest = _materialize(
            temp_root,
            "PREFU26",
            workbook,
            display_name="PREFU26 official",
            set_latest=True,
            set_default_comparator=True,
            set_default_bridge_vintage=True,
        )
        assert manifest["vintage_id"] == "PREFU26"
        pack = ov.load_official_vintage("PREFU26", repo_root=temp_root)
        assert pack is not None
        assert len(pack.annual_spine) == 39 * (END_FY - START_FY + 1)
        assert set(pack.formula_audit["status"]) == {"reconciled"}
        assert ov.default_comparator_vintage_id(temp_root) == "PREFU26"
        assert ov.default_bridge_vintage_id(temp_root) == "PREFU26"
        assert ov.latest_official_vintage_id(temp_root) == "PREFU26"
        entry = ov.official_vintage_entry("BASE0", repo_root=temp_root)
        assert entry["is_latest"] is False

    def test_later_vintage_with_different_sheet_name(self, temp_root: Path):
        workbook = temp_root / "references" / "MBU27 revenue forecast.xlsx"
        write_vintage_workbook(workbook, sheet="MBU27 Baseline")
        manifest = _materialize(temp_root, "MBU27", workbook, sheet="MBU27 Baseline")
        assert manifest["workbook"]["sheet"] == "MBU27 Baseline"
        pack = ov.load_official_vintage("MBU27", repo_root=temp_root)
        assert pack.registry_entry["source_sheet"] == "MBU27 Baseline"

    def test_later_source_horizon_is_inferred_and_registered(self, temp_root: Path):
        workbook = temp_root / "references" / "BEFU27 revenue forecast.xlsx"
        write_vintage_workbook(workbook, end_fy=2060)
        _materialize(temp_root, "BEFU27", workbook)
        entry = ov.official_vintage_entry("BEFU27", repo_root=temp_root)
        assert entry["source_horizon_fy"] == 2060
        assert entry["long_forecast_end_fy"] == 2060
        pack = ov.load_official_vintage("BEFU27", repo_root=temp_root)
        assert pack.source_horizon_fy == 2060

    def test_identical_reingestion_is_idempotent(self, temp_root: Path):
        workbook = temp_root / "references" / "PREFU26 revenue forecast.xlsx"
        write_vintage_workbook(workbook)
        _materialize(temp_root, "PREFU26", workbook)
        pack_dir = temp_root / "data" / "revenue_model_source_pack" / "official_vintages" / "prefu26"
        first = (pack_dir / "manifest.json").read_text(encoding="utf-8")
        _materialize(temp_root, "PREFU26", workbook)
        second = (pack_dir / "manifest.json").read_text(encoding="utf-8")
        assert first == second

    def test_conflicting_reingestion_fails_closed(self, temp_root: Path):
        workbook = temp_root / "references" / "PREFU26 revenue forecast.xlsx"
        write_vintage_workbook(workbook)
        _materialize(temp_root, "PREFU26", workbook)
        pack_dir = temp_root / "data" / "revenue_model_source_pack" / "official_vintages" / "prefu26"
        before = (pack_dir / "manifest.json").read_text(encoding="utf-8")

        def bump_total(ws: Any) -> None:
            ws.cell(65, 30, float(ws.cell(65, 30).value) + 1.0)

        write_vintage_workbook(workbook, mutate=bump_total)
        with pytest.raises(ov.OfficialVintageError, match="does not match the registered"):
            _materialize(temp_root, "PREFU26", workbook)
        assert (pack_dir / "manifest.json").read_text(encoding="utf-8") == before

    def test_missing_required_row_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V1.xlsx"
        write_vintage_workbook(
            workbook, mutate=lambda ws: setattr(ws.cell(58, 1), "value", None)
        )
        with pytest.raises(ov.OfficialVintageError, match="not found in section"):
            _materialize(temp_root, "V1", workbook)

    def test_renamed_label_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V2.xlsx"
        write_vintage_workbook(
            workbook, mutate=lambda ws: ws.cell(28, 1, "Light RUC receipts (m $)")
        )
        with pytest.raises(ov.OfficialVintageError, match="not found in section"):
            _materialize(temp_root, "V2", workbook)

    def test_changed_unit_in_label_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V3.xlsx"
        write_vintage_workbook(
            workbook, mutate=lambda ws: ws.cell(5, 1, "Light RUC net km (bn km)")
        )
        with pytest.raises(ov.OfficialVintageError, match="not found in section"):
            _materialize(temp_root, "V3", workbook)

    def test_missing_total_row_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V4.xlsx"
        write_vintage_workbook(
            workbook, mutate=lambda ws: setattr(ws.cell(65, 1), "value", None)
        )
        with pytest.raises(ov.OfficialVintageError, match="not found in section"):
            _materialize(temp_root, "V4", workbook)

    def test_duplicate_label_in_section_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V5.xlsx"
        write_vintage_workbook(
            workbook, mutate=lambda ws: ws.cell(59, 1, "TUC net revenue (m $)")
        )
        with pytest.raises(ov.OfficialVintageError, match="duplicated at rows"):
            _materialize(temp_root, "V5", workbook)

    def test_duplicate_year_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V6.xlsx"
        write_vintage_workbook(workbook, mutate=lambda ws: ws.cell(1, 3, 2001))
        with pytest.raises(ov.OfficialVintageError, match="duplicate year columns"):
            _materialize(temp_root, "V6", workbook)

    def test_unknown_period_status_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V7.xlsx"
        write_vintage_workbook(workbook, mutate=lambda ws: ws.cell(2, 30, "PROJECTION"))
        with pytest.raises(ov.OfficialVintageError, match="unknown Period values"):
            _materialize(temp_root, "V7", workbook)

    def test_missing_required_value_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V8.xlsx"
        write_vintage_workbook(
            workbook, mutate=lambda ws: setattr(ws.cell(65, 30), "value", None)
        )
        with pytest.raises(ov.OfficialVintageError, match="missing required value"):
            _materialize(temp_root, "V8", workbook)

    def test_non_numeric_value_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V9.xlsx"
        write_vintage_workbook(workbook, mutate=lambda ws: ws.cell(61, 30, "n/a"))
        with pytest.raises(ov.OfficialVintageError, match="non-numeric value"):
            _materialize(temp_root, "V9", workbook)

    def test_formula_without_cached_value_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V10.xlsx"
        write_vintage_workbook(workbook, mutate=lambda ws: ws.cell(64, 30, "=B64+C64"))
        with pytest.raises(ov.OfficialVintageError, match="no cached value|non-numeric value"):
            _materialize(temp_root, "V10", workbook)

    def test_sentinel_mismatch_fails(self, temp_root: Path):
        workbook = temp_root / "references" / "V11.xlsx"
        write_vintage_workbook(workbook)
        _materialize(temp_root, "V11", workbook)
        registry = ov.load_official_vintage_registry(temp_root)
        for entry in registry["vintages"]:
            if entry["vintage_id"] == "V11":
                entry["sentinels"] = [
                    {"series_id": "total_nltf_net_revenue", "fy": 2030, "value": -1.0}
                ]
        target = ov.registry_path(temp_root)
        target.write_text(json.dumps(registry, indent=2) + "\n", encoding="utf-8")
        with pytest.raises(ov.OfficialVintageError, match="sentinel validation failed"):
            _materialize(temp_root, "V11", workbook)

    def test_percentage_change_block_is_extractable(self, temp_root: Path):
        import openpyxl

        workbook = temp_root / "references" / "V12.xlsx"
        write_vintage_workbook(workbook)
        _materialize(temp_root, "V12", workbook)
        entry = ov.official_vintage_entry("V12", repo_root=temp_root)
        wb = openpyxl.load_workbook(workbook, data_only=True)
        try:
            changes = ov.extract_annual_percentage_changes(wb["Baseline"], entry)
        finally:
            wb.close()
        assert set(changes["series_id"]) == {
            item["series_id"]
            for item in ov.CANONICAL_SERIES_DEFINITIONS
            if item["section"] == "Key volumes"
        }
        assert len(changes) == 9 * (END_FY - START_FY + 1)
        assert pd.to_numeric(
            changes.loc[changes["FY"] > START_FY, "pct_change"], errors="coerce"
        ).notna().all()
