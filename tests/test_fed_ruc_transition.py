"""Focused acceptance tests for the fleetwide FED->RUC transition overlay.

The transition replaces all fuel excise duty with RUC from 1 Jan 2028 under
two governed leakage cases (managed / stress). These tests prove the
consultant's acceptance criteria on the REAL governed inputs: the
materialised policy-runtime chart rows are exactly what the overlay consumes
on the fast path, so the module-level tests here exercise the same frames
production reads, and the app-level tests drive the full overlay-chain path
(key -> cached_scenario_overlay_rows -> view -> extract).
"""
from __future__ import annotations

import io
import json
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

import app
from model_dashboard.fed_ruc_transition import (
    FED_RUC_TRANSITION_MANAGED,
    FED_RUC_TRANSITION_OFF,
    FED_RUC_TRANSITION_PARAMS,
    FED_RUC_TRANSITION_STATES,
    FED_RUC_TRANSITION_STRESS,
    FedRucTransitionError,
    LIGHT_PETROL_RUC_SERIES,
    TRANSITION_COLLECTION_COST_SERIES,
    TRANSITION_LEAKAGE_SERIES,
    TRANSITION_ONEOFF_COST_SERIES,
    TRANSITION_START_FY,
    apply_fed_ruc_transition_to_chart_rows,
    apply_fed_ruc_transition_to_quarterly_rows,
    fed_ruc_transition_marker_present,
    fiscal_year_quarters,
    leakage_rate,
    normalise_fed_ruc_transition_state,
    transition_age,
)
from model_dashboard.rate_paths import ped_quarterly_rate_schedules
from model_dashboard.revenue_outlook import (
    CURRENT_REVENUE_OUTLOOK_DIR,
    PED_BRIDGE_DEFAULT_MODE,
    load_revenue_outlook_pack,
    revenue_outlook_signature,
)

ROOT = Path(__file__).resolve().parents[1]
FED_PATH = "Current planned path"
CATALOGUE_STATE_DIR = (
    ROOT / "data" / "revenue_outlook_policy_runtime" / "ensemble" / "frames"
)
CURRENT_SCENARIOS = (
    "current_basecase",
    "current_comparison_1",
    "middle_east_low",
    "middle_east_medium",
    "middle_east_high",
)
TOTAL_SERIES = "total_nltf_net_revenue"


def _catalogue_rows(current_state: str = "delayed_6m") -> pd.DataFrame:
    path = (
        CATALOGUE_STATE_DIR
        / f"ensemble__cur-{current_state}__off-published"
        / "chart_rows.parquet"
    )
    if not path.is_file():
        pytest.skip(f"materialised state missing: {path}")
    return pd.read_parquet(path)


@pytest.fixture(scope="module")
def schedule() -> pd.DataFrame:
    return ped_quarterly_rate_schedules(ROOT)


@pytest.fixture(scope="module")
def baseline_rows() -> pd.DataFrame:
    return _catalogue_rows()


def _apply(rows: pd.DataFrame, schedule: pd.DataFrame, state: str, policy="delayed_6m"):
    return apply_fed_ruc_transition_to_chart_rows(
        rows, schedule, state, policy_state=policy
    )


@pytest.fixture(scope="module")
def managed(baseline_rows, schedule):
    return _apply(baseline_rows, schedule, FED_RUC_TRANSITION_MANAGED)


@pytest.fixture(scope="module")
def stress(baseline_rows, schedule):
    return _apply(baseline_rows, schedule, FED_RUC_TRANSITION_STRESS)


def _annual(rows: pd.DataFrame, scenario: str) -> pd.DataFrame:
    scoped = rows[
        rows["time_grain"].astype(str).eq("june_year")
        & rows["scenario_name"].astype(str).eq(scenario)
    ]
    return scoped.pivot_table(
        index="june_year", columns="series_id", values="value", aggfunc="first"
    )


def _value(frame: pd.DataFrame, fy: int, series: str) -> float | None:
    if series not in frame.columns or fy not in frame.index:
        return None
    value = frame.at[fy, series]
    return None if pd.isna(value) else float(value)


# ------------------------------------------------------------- vocabulary


def test_states_and_normalisation_fail_closed() -> None:
    assert FED_RUC_TRANSITION_STATES == ("off", "managed", "stress")
    assert normalise_fed_ruc_transition_state("") == FED_RUC_TRANSITION_OFF
    assert normalise_fed_ruc_transition_state(None) == FED_RUC_TRANSITION_OFF
    assert normalise_fed_ruc_transition_state("Managed") == FED_RUC_TRANSITION_MANAGED
    with pytest.raises(FedRucTransitionError):
        normalise_fed_ruc_transition_state("fleetwide")


def test_leakage_schedules_match_the_documented_parameters() -> None:
    managed_params = FED_RUC_TRANSITION_PARAMS[FED_RUC_TRANSITION_MANAGED]
    stress_params = FED_RUC_TRANSITION_PARAMS[FED_RUC_TRANSITION_STRESS]
    # Transition year: recoverable*(1-rho) + unrecoverable.
    assert leakage_rate(1, managed_params) == pytest.approx(0.059)
    assert leakage_rate(1, stress_params) == pytest.approx(0.065)
    # Year 2: 20% known at the case's recovery rate plus 3% unknown.
    assert leakage_rate(2, managed_params) == pytest.approx(0.20 * 0.30 + 0.03)
    assert leakage_rate(2, stress_params) == pytest.approx(0.20 * 0.50 + 0.03)
    # Terminal (year 8+): long-run known at recovery plus terminal unknown.
    assert leakage_rate(8, managed_params) == pytest.approx(0.09 * 0.30 + 0.05)
    assert leakage_rate(25, managed_params) == pytest.approx(0.09 * 0.30 + 0.05)
    assert leakage_rate(8, stress_params) == pytest.approx(0.12 * 0.50 + 0.11)
    # Stress leaks strictly more than managed at every age.
    for age in range(1, 30):
        assert leakage_rate(age, stress_params) > leakage_rate(age, managed_params)
    # Costs are the documented pairs.
    assert managed_params.one_off_cost_m == 34.0
    assert managed_params.ongoing_cost_m_per_year == 16.0
    assert stress_params.one_off_cost_m == 111.0
    assert stress_params.ongoing_cost_m_per_year == 31.0


def test_transition_ages_step_at_calendar_years() -> None:
    assert transition_age("2027Q4") == 0
    assert transition_age("2028Q1") == 1
    assert transition_age("2028Q4") == 1
    assert transition_age("2029Q1") == 2
    assert fiscal_year_quarters(2028) == ("2027Q3", "2027Q4", "2028Q1", "2028Q2")


# ------------------------------------------------------ acceptance 1 and 2


def test_off_state_returns_the_rows_unchanged(baseline_rows, schedule) -> None:
    rows, audit = _apply(baseline_rows, schedule, FED_RUC_TRANSITION_OFF)
    assert rows is baseline_rows
    assert audit.empty
    assert not fed_ruc_transition_marker_present(baseline_rows)


@pytest.mark.parametrize("case", ["managed", "stress"])
def test_nothing_changes_before_fy2028(request, baseline_rows, case) -> None:
    rows, _ = request.getfixturevalue(case)
    fy = pd.to_numeric(baseline_rows["june_year"], errors="coerce")
    early = baseline_rows[
        baseline_rows["time_grain"].astype(str).eq("june_year") & fy.lt(TRANSITION_START_FY)
    ]
    fy_new = pd.to_numeric(rows["june_year"], errors="coerce")
    early_new = rows[
        rows["time_grain"].astype(str).eq("june_year") & fy_new.lt(TRANSITION_START_FY)
    ]
    merged = early.merge(
        early_new,
        on=["series_id", "scenario_name", "period"],
        suffixes=("_old", "_new"),
    )
    assert len(merged) == len(early)
    old = pd.to_numeric(merged["value_old"], errors="coerce")
    new = pd.to_numeric(merged["value_new"], errors="coerce")
    both = old.notna() & new.notna()
    assert np.array_equal(old[both].to_numpy(), new[both].to_numpy())
    assert (old.isna() == new.isna()).all()


def test_quarterly_activity_rows_are_untouched(baseline_rows, managed) -> None:
    rows, _ = managed
    old_q = baseline_rows[baseline_rows["time_grain"].astype(str).eq("quarterly")]
    new_q = rows[rows["time_grain"].astype(str).eq("quarterly")]
    merged = old_q.merge(
        new_q, on=["series_id", "scenario_name", "period"], suffixes=("_old", "_new")
    )
    assert len(merged) == len(old_q)
    old = pd.to_numeric(merged["value_old"], errors="coerce")
    new = pd.to_numeric(merged["value_new"], errors="coerce")
    both = old.notna() & new.notna()
    assert np.array_equal(old[both].to_numpy(), new[both].to_numpy())


# ------------------------------------------------------------ acceptance 3


@pytest.mark.parametrize("case", ["managed", "stress"])
def test_fy2028_has_two_pre_and_two_post_calendar_quarters(request, case) -> None:
    _, audit = request.getfixturevalue(case)
    quarters = audit[
        audit["scenario_name"].eq("current_basecase")
        & audit["june_year"].eq(TRANSITION_START_FY)
        & ~audit["quarter"].astype(str).eq("FY total")
    ]
    assert sorted(quarters["quarter"]) == ["2028Q1", "2028Q2"]
    assert set(quarters["transition_age"]) == {1}
    annual = audit[
        audit["scenario_name"].eq("current_basecase")
        & audit["june_year"].eq(TRANSITION_START_FY)
        & audit["quarter"].astype(str).eq("FY total")
    ]
    share = float(annual["pre_transition_fed_share"].iloc[0])
    assert 0.0 < share < 1.0


# ------------------------------------------------------------ acceptance 4


@pytest.mark.parametrize("case", ["managed", "stress"])
def test_net_fed_and_refunds_are_zero_from_2028q1(request, baseline_rows, case) -> None:
    rows, audit = request.getfixturevalue(case)
    for scenario in CURRENT_SCENARIOS:
        before = _annual(baseline_rows, scenario)
        after = _annual(rows, scenario)
        share = float(
            audit[
                audit["scenario_name"].eq(scenario)
                & audit["june_year"].eq(TRANSITION_START_FY)
                & audit["quarter"].astype(str).eq("FY total")
            ]["pre_transition_fed_share"].iloc[0]
        )
        for fy in after.index:
            fy = int(fy)
            if fy < TRANSITION_START_FY:
                continue
            for series in ("gross_ped_revenue", "gross_fed_revenue", "net_fed_revenue"):
                old = _value(before, fy, series)
                new = _value(after, fy, series)
                if old is None:
                    continue
                expected = old * share if fy == TRANSITION_START_FY else 0.0
                assert new == pytest.approx(expected, abs=1e-9), (scenario, fy, series)
            # Implied refunds (gross - net) retire with the duty itself.
            gross_new = _value(after, fy, "gross_fed_revenue")
            net_new = _value(after, fy, "net_fed_revenue")
            if gross_new is not None and net_new is not None and fy > TRANSITION_START_FY:
                assert gross_new - net_new == pytest.approx(0.0, abs=1e-9)


# ------------------------------------------------------------ acceptance 5


@pytest.mark.parametrize("case", ["managed", "stress"])
def test_petrol_ruc_is_k_times_r_over_1000_before_leakage(
    request, baseline_rows, case
) -> None:
    rows, audit = request.getfixturevalue(case)
    quarters = audit[~audit["quarter"].astype(str).eq("FY total")]
    gross = quarters["gross_petrol_ruc_m"].astype(float)
    identity = (
        quarters["petrol_vkt_million_km"].astype(float)
        * quarters["light_ruc_rate_per_1000km"].astype(float)
        / 1000.0
    )
    assert np.allclose(gross.to_numpy(), identity.to_numpy(), rtol=0, atol=1e-9)
    collected = quarters["collected_petrol_ruc_m"].astype(float)
    leak = quarters["leakage_m"].astype(float)
    assert np.allclose((gross - leak).to_numpy(), collected.to_numpy(), atol=1e-9)
    # The quarterly VKT split sums back to the governed annual series, and
    # the annual collected total is exactly the chart's new series.
    before = _annual(baseline_rows, "current_basecase")
    after = _annual(rows, "current_basecase")
    base_quarters = quarters[quarters["scenario_name"].eq("current_basecase")]
    for fy, group in base_quarters.groupby("june_year"):
        fy = int(fy)
        annual_vkt = _value(before, fy, "light_petrol_vkt")
        expected_quarters = 2 if fy == TRANSITION_START_FY else 4
        assert len(group) == expected_quarters
        assert group["petrol_vkt_million_km"].astype(float).sum() == pytest.approx(
            annual_vkt / 4.0 * expected_quarters
        )
        assert _value(after, fy, LIGHT_PETROL_RUC_SERIES) == pytest.approx(
            group["collected_petrol_ruc_m"].astype(float).sum()
        )
        assert _value(after, fy, TRANSITION_LEAKAGE_SERIES) == pytest.approx(
            group["leakage_m"].astype(float).sum()
        )


# ------------------------------------------------------ acceptance 6 and 7


def test_managed_collects_more_than_stress_in_every_affected_quarter(
    managed, stress
) -> None:
    managed_q = managed[1][~managed[1]["quarter"].astype(str).eq("FY total")]
    stress_q = stress[1][~stress[1]["quarter"].astype(str).eq("FY total")]
    merged = managed_q.merge(
        stress_q, on=["scenario_name", "june_year", "quarter"], suffixes=("_m", "_s")
    )
    assert len(merged) == len(managed_q)
    assert (
        merged["collected_petrol_ruc_m_m"].astype(float)
        > merged["collected_petrol_ruc_m_s"].astype(float)
    ).all()


def test_cumulative_shortfalls_never_reverse_sign(
    baseline_rows, managed, stress
) -> None:
    base = _annual(baseline_rows, "current_basecase")[TOTAL_SERIES]
    managed_total = _annual(managed[0], "current_basecase")[TOTAL_SERIES]
    stress_total = _annual(stress[0], "current_basecase")[TOTAL_SERIES]
    affected = [fy for fy in base.index if int(fy) >= TRANSITION_START_FY]
    managed_gap = (managed_total.loc[affected] - stress_total.loc[affected]).astype(float)
    # Managed never collects less than stress in any year...
    assert (managed_gap > 0).all()
    # ...so the cumulative managed-vs-stress gap is monotone: it never
    # reverses sign or direction over the horizon.
    cumulative = managed_gap.cumsum()
    assert (cumulative.diff().dropna() > 0).all()
    assert (cumulative > 0).all()


# ------------------------------------------------------ acceptance 8 and 9


@pytest.mark.parametrize("case", ["managed", "stress"])
def test_leakage_applies_only_to_the_newly_transitioned_base(
    request, baseline_rows, case
) -> None:
    rows, audit = request.getfixturevalue(case)
    before = _annual(baseline_rows, "current_basecase")
    after = _annual(rows, "current_basecase")
    share = float(
        audit[
            audit["scenario_name"].eq("current_basecase")
            & audit["june_year"].eq(TRANSITION_START_FY)
            & audit["quarter"].astype(str).eq("FY total")
        ]["pre_transition_fed_share"].iloc[0]
    )
    for fy in after.index:
        fy = int(fy)
        if fy < TRANSITION_START_FY:
            continue
        # Existing RUC classes: conventional light, light BEV and heavy are
        # untouched by the transition (no leakage, no re-pricing).
        for series in (
            "light_ruc_net_revenue",
            "light_bev_ruc_net_revenue",
            "heavy_ruc_net_revenue",
            "light_ruc_net_km",
            "phev_ruc_net_km",
        ):
            old = _value(before, fy, series)
            new = _value(after, fy, series)
            if old is not None:
                assert new == pytest.approx(old, abs=1e-12), (fy, series)
        # PHEVs move to the FULL light rate for post-transition quarters -
        # explicit re-pricing, never leakage: the uplift is exactly the
        # implied full-light/PHEV rate ratio on the post-transition share.
        phev_old = _value(before, fy, "phev_ruc_net_revenue")
        phev_new = _value(after, fy, "phev_ruc_net_revenue")
        light_rate = _value(before, fy, "light_ruc_net_revenue") / _value(
            before, fy, "light_ruc_net_km"
        )
        phev_rate = phev_old / _value(before, fy, "phev_ruc_net_km")
        post_share = (1.0 - share) if fy == TRANSITION_START_FY else 1.0
        expected = phev_old * (1.0 + post_share * (light_rate / phev_rate - 1.0))
        assert phev_new == pytest.approx(expected, rel=1e-9), fy
        assert phev_new > phev_old  # the discount compensated petrol FED
        # No FED/RUC double charge: once petrol pays RUC, net FED is gone.
        if fy > TRANSITION_START_FY:
            assert _value(after, fy, "net_fed_revenue") == pytest.approx(0.0, abs=1e-9)
            assert _value(after, fy, LIGHT_PETROL_RUC_SERIES) > 0.0


# ----------------------------------------------------------- acceptance 10


@pytest.mark.parametrize(
    "policy_state", ["published", "delayed_6m", "delayed_12m", "delayed_36m", "off"]
)
def test_composes_with_representative_timing_states(schedule, policy_state) -> None:
    rows = _catalogue_rows(policy_state)
    adjusted, audit = _apply(rows, schedule, FED_RUC_TRANSITION_MANAGED, policy=policy_state)
    before = _annual(rows, "current_basecase")
    after = _annual(adjusted, "current_basecase")
    assert _value(after, 2029, "net_fed_revenue") == pytest.approx(0.0, abs=1e-9)
    assert _value(after, 2029, LIGHT_PETROL_RUC_SERIES) > 0.0
    annual = audit[
        audit["scenario_name"].eq("current_basecase")
        & audit["june_year"].eq(TRANSITION_START_FY)
        & audit["quarter"].astype(str).eq("FY total")
    ]
    share = float(annual["pre_transition_fed_share"].iloc[0])
    assert 0.0 < share < 1.0
    # The transition expresses the SELECTED state's staircase through RUC:
    # the petrol pool pays the rate already implied by that state's rows.
    implied = (
        _value(before, 2030, "light_ruc_net_revenue")
        / _value(before, 2030, "light_ruc_net_km")
        * 1000.0
    )
    quarters = audit[
        audit["scenario_name"].eq("current_basecase")
        & audit["june_year"].eq(2030)
        & ~audit["quarter"].astype(str).eq("FY total")
    ]
    assert quarters["light_ruc_rate_per_1000km"].astype(float).mean() == pytest.approx(
        implied, rel=1e-9
    )
    # Identity closes under every timing state.
    delta_total = _value(after, 2035, TOTAL_SERIES) - _value(before, 2035, TOTAL_SERIES)
    components = (
        (_value(after, 2035, "net_fed_revenue") - _value(before, 2035, "net_fed_revenue"))
        + _value(after, 2035, LIGHT_PETROL_RUC_SERIES)
        + (
            _value(after, 2035, "phev_ruc_net_revenue")
            - _value(before, 2035, "phev_ruc_net_revenue")
        )
        - _value(after, 2035, TRANSITION_COLLECTION_COST_SERIES)
    )
    assert delta_total == pytest.approx(components, abs=1e-6)


# ----------------------------------------------------------- acceptance 11


@pytest.mark.parametrize("case", ["managed", "stress"])
def test_revenue_identities_close_for_every_scenario_and_year(
    request, baseline_rows, case
) -> None:
    rows, _ = request.getfixturevalue(case)
    for scenario in CURRENT_SCENARIOS:
        before = _annual(baseline_rows, scenario)
        after = _annual(rows, scenario)
        for fy in after.index:
            fy = int(fy)
            if fy < TRANSITION_START_FY:
                continue
            total_old = _value(before, fy, TOTAL_SERIES)
            total_new = _value(after, fy, TOTAL_SERIES)
            if total_old is None or total_new is None:
                continue
            oneoff = _value(after, fy, TRANSITION_ONEOFF_COST_SERIES) or 0.0
            delta = (
                (_value(after, fy, "net_fed_revenue") - _value(before, fy, "net_fed_revenue"))
                + (_value(after, fy, LIGHT_PETROL_RUC_SERIES) or 0.0)
                + (
                    _value(after, fy, "phev_ruc_net_revenue")
                    - _value(before, fy, "phev_ruc_net_revenue")
                )
                - (_value(after, fy, TRANSITION_COLLECTION_COST_SERIES) or 0.0)
                - oneoff
            )
            assert total_new - total_old == pytest.approx(delta, abs=1e-6), (scenario, fy)
            # Total RUC gains exactly the collected petrol RUC + PHEV uplift.
            ruc_delta = _value(after, fy, "total_ruc_net_revenue") - _value(
                before, fy, "total_ruc_net_revenue"
            )
            assert ruc_delta == pytest.approx(
                (_value(after, fy, LIGHT_PETROL_RUC_SERIES) or 0.0)
                + (
                    _value(after, fy, "phev_ruc_net_revenue")
                    - _value(before, fy, "phev_ruc_net_revenue")
                ),
                abs=1e-6,
            ), (scenario, fy)
    # One-off cost is recorded once, in the transition year only.
    after_base = _annual(rows, "current_basecase")
    assert _value(after_base, TRANSITION_START_FY, TRANSITION_ONEOFF_COST_SERIES) == (
        FED_RUC_TRANSITION_PARAMS[
            FED_RUC_TRANSITION_MANAGED if case == "managed" else FED_RUC_TRANSITION_STRESS
        ].one_off_cost_m
    )
    assert _value(after_base, TRANSITION_START_FY + 1, TRANSITION_ONEOFF_COST_SERIES) is None


# ----------------------------------------------------------- acceptance 13


@pytest.mark.parametrize("case", ["managed", "stress"])
def test_official_comparators_and_actuals_are_untouched(
    request, baseline_rows, case
) -> None:
    rows, _ = request.getfixturevalue(case)
    role = baseline_rows["scenario_name"].astype(str)
    officials = baseline_rows[
        baseline_rows["scenario_role"].astype(str).isin(["official_comparator", "actual"])
    ]
    new_officials = rows[
        rows["scenario_role"].astype(str).isin(["official_comparator", "actual"])
    ]
    merged = officials.merge(
        new_officials,
        on=["series_id", "scenario_name", "time_grain", "period"],
        suffixes=("_old", "_new"),
    )
    assert len(merged) == len(officials)
    old = pd.to_numeric(merged["value_old"], errors="coerce")
    new = pd.to_numeric(merged["value_new"], errors="coerce")
    both = old.notna() & new.notna()
    assert np.array_equal(old[both].to_numpy(), new[both].to_numpy())
    del role


def test_every_current_scenario_carries_the_transition(managed) -> None:
    rows, audit = managed
    audited = set(audit["scenario_name"].astype(str))
    for scenario in CURRENT_SCENARIOS:
        assert scenario in audited
        after = _annual(rows, scenario)
        assert _value(after, 2035, LIGHT_PETROL_RUC_SERIES) > 0.0


# --------------------------------------------------------- quarterly grain


def test_quarterly_display_zeroes_fed_from_2028q1_and_preserves_the_fy_sum() -> None:
    quarters = pd.DataFrame(
        {
            "series_id": ["net_fed_revenue"] * 8,
            "scenario_name": ["current_basecase"] * 8,
            "scenario_role": ["basecase"] * 8,
            "time_grain": ["quarterly"] * 8,
            "period": [
                "2027Q3", "2027Q4", "2028Q1", "2028Q2",
                "2028Q3", "2028Q4", "2029Q1", "2029Q2",
            ],
            "june_year": [2028] * 4 + [2029] * 4,
            "value": [300.0, 310.0, 320.0, 330.0, 0.0, 0.0, 0.0, 0.0],
            "value_status": ["x"] * 8,
            "data_scope": ["x"] * 8,
        }
    )
    adjusted = apply_fed_ruc_transition_to_quarterly_rows(quarters)
    by_period = adjusted.set_index("period")["value"]
    assert by_period["2028Q1"] == 0.0 and by_period["2028Q2"] == 0.0
    # FY2028's surviving half-year keeps the fiscal-year sum exactly.
    assert by_period[["2027Q3", "2027Q4"]].sum() == pytest.approx(1260.0)
    # Relative within-year shape of the surviving quarters is preserved.
    assert by_period["2027Q4"] / by_period["2027Q3"] == pytest.approx(310.0 / 300.0)


# ------------------------------------------------- app-level real-path tests


@pytest.fixture(scope="module")
def app_context():
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None
    signature = revenue_outlook_signature(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, ROOT)
    manifest = json.loads(
        (ROOT / "data" / "revenue_outlook_policy_runtime" / "ensemble" / "manifest.json")
        .read_text(encoding="utf-8")
    )
    pins = dict(manifest["pinned_key_fields"])
    pins.pop("engine", None)
    key_off = app.RevenueScenarioComputationKey(
        engine="ensemble",
        current_fed_policy_state=app.FED_POLICY_DELAYED_6M,
        official_fed_policy_state=app.FED_POLICY_PUBLISHED,
        **{k: (tuple(v) if isinstance(v, list) else v) for k, v in pins.items()},
    )
    sensitivity = app.selected_sensitivity_key("Off", "Off", "Off", freight_rail_shift="Off")
    return pack, signature, sensitivity, key_off


def _overlay_rows(app_context, key):
    pack, signature, sensitivity, _ = app_context
    rows, *_ = app.cached_scenario_overlay_rows(
        signature, sensitivity, PED_BRIDGE_DEFAULT_MODE, key, pack
    )
    return rows


def test_selector_moves_plotted_values_through_the_real_path(app_context) -> None:
    """The wiring test the selector-wiring lesson demands: key -> overlay
    rows -> a VALUE moves at a forecast FY while earlier years, officials and
    actuals stay bit-identical."""
    _, _, _, key_off = app_context
    rows_off = _overlay_rows(app_context, key_off)
    rows_on = _overlay_rows(
        app_context, key_off.replace(fed_ruc_transition=FED_RUC_TRANSITION_MANAGED)
    )

    def total(rows, fy):
        annual = _annual(rows, "current_basecase")
        return _value(annual, fy, TOTAL_SERIES)

    assert total(rows_on, 2035) != total(rows_off, 2035)
    assert total(rows_on, 2027) == total(rows_off, 2027)
    officials_off = rows_off[
        rows_off["scenario_role"].astype(str).isin(["official_comparator", "actual"])
    ]
    officials_on = rows_on[
        rows_on["scenario_role"].astype(str).isin(["official_comparator", "actual"])
    ]
    merged = officials_off.merge(
        officials_on,
        on=["series_id", "scenario_name", "time_grain", "period"],
        suffixes=("_off", "_on"),
    )
    old = pd.to_numeric(merged["value_off"], errors="coerce")
    new = pd.to_numeric(merged["value_on"], errors="coerce")
    both = old.notna() & new.notna()
    assert len(merged) == len(officials_off)
    assert np.array_equal(old[both].to_numpy(), new[both].to_numpy())


def test_view_quarterly_net_fed_is_zero_from_2028q1(app_context) -> None:
    pack, signature, sensitivity, key_off = app_context
    key_on = key_off.replace(fed_ruc_transition=FED_RUC_TRANSITION_MANAGED)
    view = app.cached_revenue_outlook_view(
        signature,
        "Net FED revenue",
        "quarterly",
        FED_PATH,
        ("Actual", "Current finalist Base case"),
        sensitivity,
        PED_BRIDGE_DEFAULT_MODE,
        key_on,
        pack,
    )
    quarters = view["filtered_rows"]
    quarters = quarters[
        quarters["trace_name"].astype(str).eq("Current finalist Base case")
        & quarters["time_grain"].astype(str).eq("quarterly")
    ]
    values = pd.to_numeric(quarters["value"], errors="coerce")
    period = quarters["period"].astype(str)
    from model_dashboard.fed_policy_states import quarter_serial

    serial = period.map(quarter_serial)
    post = serial >= quarter_serial("2028Q1")
    assert post.any()
    assert values[post].abs().max() == pytest.approx(0.0, abs=1e-9)
    # FY2028's surviving quarters still sum to the annual value on screen.
    fy2028 = pd.to_numeric(quarters["june_year"], errors="coerce").eq(2028)
    annual_view = app.cached_revenue_outlook_view(
        signature,
        "Net FED revenue",
        "june_year",
        FED_PATH,
        ("Actual", "Current finalist Base case"),
        sensitivity,
        PED_BRIDGE_DEFAULT_MODE,
        key_on,
        pack,
    )
    annual_rows = annual_view["filtered_rows"]
    annual_rows = annual_rows[
        annual_rows["trace_name"].astype(str).eq("Current finalist Base case")
        & annual_rows["time_grain"].astype(str).eq("june_year")
        & pd.to_numeric(annual_rows["june_year"], errors="coerce").eq(2028)
    ]
    assert values[fy2028].sum() == pytest.approx(
        float(pd.to_numeric(annual_rows["value"], errors="coerce").iloc[0]), rel=1e-9
    )


def test_extract_matches_the_displayed_view(app_context) -> None:
    from openpyxl import load_workbook

    pack, signature, sensitivity, key_off = app_context
    key_on = key_off.replace(fed_ruc_transition=FED_RUC_TRANSITION_STRESS)
    traces = (
        "Actual",
        "Current finalist Base case",
        "Persistent downside",
    )
    result = app.cached_revenue_outlook_extract_bytes(
        signature, sensitivity, PED_BRIDGE_DEFAULT_MODE, key_on, traces, pack
    )
    assert result.skipped_traces == []
    workbook = load_workbook(io.BytesIO(result.workbook_bytes))
    sheet = workbook["Current Base"]
    rows_on = _overlay_rows(app_context, key_on)
    annual = _annual(rows_on, "current_basecase")

    def column_for(fy: int) -> int:
        return 2 + (fy - 2001)

    for fy in (2029, 2035, 2050):
        assert sheet.cell(row=66, column=column_for(fy)).value == pytest.approx(
            _value(annual, fy, LIGHT_PETROL_RUC_SERIES), rel=1e-9
        )
        assert sheet.cell(row=67, column=column_for(fy)).value == pytest.approx(
            _value(annual, fy, TRANSITION_LEAKAGE_SERIES), rel=1e-9
        )
        assert sheet.cell(row=45, column=column_for(fy)).value == pytest.approx(0.0, abs=1e-9)
    assert sheet.cell(row=66, column=1).value == "Light petrol RUC (m $)"
    # The displayed total (row 65) is the transition-adjusted NLTF total.
    assert sheet.cell(row=65, column=column_for(2035)).value == pytest.approx(
        _value(annual, 2035, TOTAL_SERIES), rel=1e-9
    )
    # Transition state is named in the workbook metadata.
    assert "FED -> RUC transition" in str(workbook.properties.description)


def test_persistent_downside_composes_with_the_transition(app_context) -> None:
    pack, signature, sensitivity, key_off = app_context
    key_on = key_off.replace(fed_ruc_transition=FED_RUC_TRANSITION_MANAGED)
    view = app.cached_revenue_outlook_view(
        signature,
        "Total NLTF revenue",
        "june_year",
        FED_PATH,
        (
            "Actual",
            "Current finalist Base case",
            "Persistent downside",
        ),
        sensitivity,
        PED_BRIDGE_DEFAULT_MODE,
        key_on,
        pack,
    )
    chart_rows = view["chart_rows"]
    downside = _annual(chart_rows, "persistent_downside")
    base = _annual(chart_rows, "current_basecase")
    assert _value(downside, 2035, LIGHT_PETROL_RUC_SERIES) is not None
    assert _value(downside, 2035, LIGHT_PETROL_RUC_SERIES) < _value(
        base, 2035, LIGHT_PETROL_RUC_SERIES
    )
    assert _value(downside, 2035, TOTAL_SERIES) < _value(base, 2035, TOTAL_SERIES)
    assert _value(downside, 2035, "net_fed_revenue") == pytest.approx(0.0, abs=1e-9)


def test_transition_audit_is_derived_from_the_displayed_rows(app_context) -> None:
    pack, signature, sensitivity, key_off = app_context
    key_on = key_off.replace(fed_ruc_transition=FED_RUC_TRANSITION_MANAGED)
    audit = app.cached_fed_ruc_transition_audit(
        signature, sensitivity, PED_BRIDGE_DEFAULT_MODE, key_on, pack
    )
    assert not audit.empty
    rows_on = _overlay_rows(app_context, key_on)
    annual = _annual(rows_on, "current_basecase")
    fy_total = audit[
        audit["scenario_name"].eq("current_basecase")
        & audit["june_year"].eq(2035)
        & audit["quarter"].astype(str).eq("FY total")
    ]
    assert float(fy_total["collected_petrol_ruc_m"].iloc[0]) == pytest.approx(
        _value(annual, 2035, LIGHT_PETROL_RUC_SERIES), rel=1e-12
    )
    # The non-road FED audit reports against the documents' reference value.
    assert float(fy_total["nonroad_fed_reference_m"].iloc[0]) == 150.0
    off_audit = app.cached_fed_ruc_transition_audit(
        signature, sensitivity, PED_BRIDGE_DEFAULT_MODE, key_off, pack
    )
    assert off_audit.empty


# ---------------------------------------------------------- A/B and UI keys


def test_comparison_b_keys_carry_the_transition() -> None:
    page_sens = app.selected_sensitivity_key("Off", "Off", "Off", freight_rail_shift="Off")
    page_key = app.RevenueScenarioComputationKey(
        engine="ensemble",
        uptake_basis=app.DEFAULT_EV_UPTAKE_MODE,
        current_fed_policy_state=app.FED_POLICY_DELAYED_6M,
        official_fed_policy_state=app.FED_POLICY_PUBLISHED,
    )
    _, key_b = app._comparison_scenario_b_keys(
        page_sens,
        page_key,
        fleet="Off",
        pt_shift="Off",
        freight="Off",
        eruc_values=(),
        fed_policy_state=app.FED_POLICY_DELAYED_6M,
        fed_ruc_transition=FED_RUC_TRANSITION_STRESS,
    )
    assert key_b.fed_ruc_transition == FED_RUC_TRANSITION_STRESS
    # Every other field is the page's own.
    assert key_b.engine == page_key.engine
    assert key_b.official_fed_policy_state == page_key.official_fed_policy_state


def test_official_comparator_side_locks_the_transition_off() -> None:
    page_key = app.RevenueScenarioComputationKey(
        engine="ensemble",
        uptake_basis=app.DEFAULT_EV_UPTAKE_MODE,
        current_fed_policy_state=app.FED_POLICY_DELAYED_6M,
        official_fed_policy_state=app.FED_POLICY_PUBLISHED,
        fed_ruc_transition=FED_RUC_TRANSITION_MANAGED,
    )
    _, key_official, _ = app._comparison_official_scenario_keys(
        page_key,
        official_policy_state=app.FED_POLICY_PUBLISHED,
        selected_vid="PREBU26",
    )
    assert key_official.fed_ruc_transition == FED_RUC_TRANSITION_OFF


def test_bands_are_withheld_while_a_transition_is_selected() -> None:
    key = app.RevenueScenarioComputationKey(
        engine="ensemble", fed_ruc_transition=FED_RUC_TRANSITION_MANAGED
    )
    assert app._uncertainty_bands_withheld_for_fed_ruc_transition(key)
    assert not app._uncertainty_bands_withheld_for_fed_ruc_transition(
        key.replace(fed_ruc_transition=FED_RUC_TRANSITION_OFF)
    )


def test_transition_field_changes_only_its_own_resolution() -> None:
    key = app.RevenueScenarioComputationKey(
        engine="ensemble",
        uptake_basis=app.DEFAULT_EV_UPTAKE_MODE,
        current_fed_policy_state=app.FED_POLICY_DELAYED_6M,
        official_fed_policy_state=app.FED_POLICY_PUBLISHED,
    )
    changed = key.replace(fed_ruc_transition=FED_RUC_TRANSITION_STRESS)
    assert app._fed_ruc_transition_state(key) == FED_RUC_TRANSITION_OFF
    assert app._fed_ruc_transition_state(changed) == FED_RUC_TRANSITION_STRESS
    assert app._fed_policy_state_scope(key) == app._fed_policy_state_scope(changed)
    assert app._official_vintage_scope(key) == app._official_vintage_scope(changed)
    assert app._resolve_uptake_basis(key) == app._resolve_uptake_basis(changed)
    assert key.digest() != changed.digest()
