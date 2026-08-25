"""Template-matched XLSX forecast extract (BEFU26 layout, FY2001-FY2050).

Contract (owner request, 2026-08): one worksheet per selected scenario path,
rows 1-65 exactly as the reference workbook (row 65 = Total net revenues),
years B:AY ending FY2050, all page levers (Fleet efficiency incl. Custom, PT
mode shift, uptake, 12c policy) reflected in the exported numbers, blanks -
never zeros - where a line item is not governed for a path.
"""

from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import app
from model_dashboard.revenue_outlook import (
    CURRENT_REVENUE_OUTLOOK_DIR,
    PED_BRIDGE_DEFAULT_MODE,
    load_revenue_outlook_pack,
    revenue_outlook_signature,
)
from model_dashboard.revenue_outlook_excel_extract import (
    EXTRACT_LAST_ROW,
    LAST_DATA_COLUMN,
    ROW_SERIES,
    TEMPLATE_RELATIVE_PATH,
    RevenueOutlookExtractError,
    build_revenue_outlook_extract,
    extract_sheet_name,
)

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT / TEMPLATE_RELATIVE_PATH

DEFAULT_TRACES = (
    "Actual",
    "PREBU26 official",
    "Current finalist Base case",
    "Current finalist High population/comparison",
)


@pytest.fixture(scope="module")
def extract_context():
    pack = load_revenue_outlook_pack(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, repo_root=ROOT)
    assert pack is not None
    signature = revenue_outlook_signature(ROOT / CURRENT_REVENUE_OUTLOOK_DIR, ROOT)
    return pack, signature


def _default_uptake_key():
    return (app.DEFAULT_EV_UPTAKE_MODE, (), (), 0, 0)


def _extract(extract_context, traces, sensitivity_key=None):
    pack, signature = extract_context
    key = sensitivity_key or app.selected_sensitivity_key("Off", "Off", "Off")
    return app.cached_revenue_outlook_extract_bytes(
        signature, key, PED_BRIDGE_DEFAULT_MODE, _default_uptake_key(), tuple(traces), pack
    )


def _workbook(result):
    return load_workbook(io.BytesIO(result.workbook_bytes))


def _row_values(worksheet, row):
    return {
        int(worksheet.cell(row=1, column=column).value): worksheet.cell(row=row, column=column).value
        for column in range(2, LAST_DATA_COLUMN + 1)
    }


@pytest.fixture(scope="module")
def default_result(extract_context):
    return _extract(extract_context, DEFAULT_TRACES)


@pytest.fixture(scope="module")
def custom_fleet_result(extract_context):
    key = app.selected_sensitivity_key(
        "Custom", "Off", "Off", custom_fleet_efficiency_pct=10.0
    )
    return _extract(extract_context, DEFAULT_TRACES, key)


@pytest.fixture(scope="module")
def pt_result(extract_context):
    key = app.selected_sensitivity_key("Off", "Med", "Off")
    return _extract(extract_context, DEFAULT_TRACES, key)


# ---------------------------------------------------------------------------
# worksheet inventory and layout
# ---------------------------------------------------------------------------


def test_single_selected_path_creates_one_worksheet(extract_context) -> None:
    result = _extract(extract_context, ("Actual", "Current finalist Base case"))
    assert result.sheet_names == ["Current Base"]
    assert result.exported_traces == ["Current finalist Base case"]


def test_three_selected_paths_create_three_worksheets(default_result) -> None:
    assert default_result.sheet_names == [
        "PREBU26 Official",
        "Current Base",
        "High Population",
    ]
    assert not default_result.skipped_traces


def test_conflict_trace_exports_from_the_aligned_frames(extract_context) -> None:
    result = _extract(
        extract_context,
        ("Current finalist Base case", "Temporary fuel shock (Treasury Medium)"),
    )
    assert result.sheet_names == ["Current Base", "Temporary Shock"]


def test_every_sheet_is_a1_ay65_with_no_fy2051_columns(default_result) -> None:
    workbook = _workbook(default_result)
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        assert worksheet.max_row == EXTRACT_LAST_ROW, name
        assert worksheet.max_column == LAST_DATA_COLUMN, name
        assert int(worksheet.cell(row=1, column=LAST_DATA_COLUMN).value) == 2050, name
        assert worksheet.cell(row=1, column=LAST_DATA_COLUMN + 1).value is None, name


def test_row_labels_match_the_reference_template_exactly(default_result) -> None:
    template = load_workbook(TEMPLATE_PATH)["Baseline"]
    expected = [template.cell(row=row, column=1).value for row in range(1, EXTRACT_LAST_ROW + 1)]
    workbook = _workbook(default_result)
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        got = [worksheet.cell(row=row, column=1).value for row in range(1, EXTRACT_LAST_ROW + 1)]
        assert got == expected, name
    assert expected[EXTRACT_LAST_ROW - 1] == "Total net revenues (m $)"


def test_period_row_classifies_actual_st_and_lt_segments(default_result) -> None:
    workbook = _workbook(default_result)
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        for column in range(2, LAST_DATA_COLUMN + 1):
            year = int(worksheet.cell(row=1, column=column).value)
            period = str(worksheet.cell(row=2, column=column).value)
            if year <= 2025:
                assert period == "ACTUAL", (name, year)
            elif year <= 2030:
                assert period == "ST_FORECAST", (name, year)
            else:
                assert period == "LT_FORECAST", (name, year)


def test_no_formula_error_tokens_and_workbook_reopens(default_result) -> None:
    workbook = _workbook(default_result)  # re-open is itself the parse check
    errors = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}
    for name in workbook.sheetnames:
        for row in workbook[name].iter_rows(values_only=True):
            for value in row:
                assert not (isinstance(value, str) and value.strip() in errors), name
                # No formulas at all: the reference extract is hardcoded values.
                assert not (isinstance(value, str) and value.startswith("=")), name


def test_sheet_names_are_valid_unique_and_readable() -> None:
    taken: set[str] = set()
    first = extract_sheet_name("Current finalist Base case", taken)
    duplicate = extract_sheet_name("Current finalist Base case", taken)
    awkward = extract_sheet_name("A[very]:long/trace*name?" + "x" * 40, taken)
    for name in (first, duplicate, awkward):
        assert 0 < len(name) <= 31
        assert not set(name) & set(':\\/?*[]')
    assert first == "Current Base"
    assert duplicate != first
    assert len({first, duplicate, awkward}) == 3


def test_export_refuses_when_no_trace_resolves(extract_context) -> None:
    pack, _ = extract_context
    with pytest.raises(RevenueOutlookExtractError):
        build_revenue_outlook_extract(
            selected_traces=("Actual",),
            line_reconciliation=pack.revenue_line_reconciliation,
            stack_components=pack.revenue_stack_components,
            pack_stack_components=pack.revenue_stack_components,
            template_path=TEMPLATE_PATH,
        )


# ---------------------------------------------------------------------------
# values: canonical view identity and lever pass-through
# ---------------------------------------------------------------------------


def test_total_net_revenues_match_the_canonical_aligned_view(
    extract_context, default_result
) -> None:
    pack, signature = extract_context
    line, _, _, _ = app.cached_aligned_scenario_detail_frames(
        signature,
        app.selected_sensitivity_key("Off", "Off", "Off"),
        PED_BRIDGE_DEFAULT_MODE,
        _default_uptake_key(),
        pack,
    )
    line = line.copy()
    line["fy"] = pd.to_numeric(line["FY"], errors="coerce")
    workbook = _workbook(default_result)
    for sheet, source_path in [
        ("Current Base", "Current finalist Base case"),
        ("PREBU26 Official", "PREBU26 official"),
    ]:
        totals = _row_values(workbook[sheet], EXTRACT_LAST_ROW)
        for fy in (2026, 2030, 2050):
            expected = line[
                line["source_path"].astype(str).eq(source_path)
                & line["series_id"].astype(str).eq("total_nltf_net_revenue")
                & line["fy"].eq(fy)
            ]["value"]
            assert not expected.empty, (sheet, fy)
            assert float(totals[fy]) == pytest.approx(float(expected.iloc[0]), rel=1e-12), (
                sheet,
                fy,
            )


def test_custom_fleet_efficiency_changes_exported_ped_values(
    default_result, custom_fleet_result
) -> None:
    baseline = _row_values(_workbook(default_result)["Current Base"], 10)
    adjusted = _row_values(_workbook(custom_fleet_result)["Current Base"], 10)
    for fy, expected_factor in ((2026, 0.9), (2030, 0.9**5), (2031, 0.9**6), (2050, 0.9**25)):
        assert float(adjusted[fy]) == pytest.approx(
            float(baseline[fy]) * expected_factor, rel=1e-9
        ), fy
    revenue_baseline = _row_values(_workbook(default_result)["Current Base"], 40)
    revenue_adjusted = _row_values(_workbook(custom_fleet_result)["Current Base"], 40)
    assert float(revenue_adjusted[2050]) < float(revenue_baseline[2050])


def test_pt_mode_shift_changes_exported_light_activity(default_result, pt_result) -> None:
    for row in (5, 11):  # Light RUC net km, Light petrol VKT
        baseline = _row_values(_workbook(default_result)["Current Base"], row)
        adjusted = _row_values(_workbook(pt_result)["Current Base"], row)
        for fy, exponent in ((2026, 1), (2030, 5), (2050, 25)):
            assert float(adjusted[fy]) == pytest.approx(
                float(baseline[fy]) * (1 - 0.005) ** exponent, rel=1e-9
            ), (row, fy)
    heavy_baseline = _row_values(_workbook(default_result)["Current Base"], 6)
    heavy_adjusted = _row_values(_workbook(pt_result)["Current Base"], 6)
    for fy in (2026, 2050):
        assert float(heavy_adjusted[fy]) == float(heavy_baseline[fy]), fy


def test_official_sheet_is_not_altered_by_current_only_levers(
    default_result, custom_fleet_result
) -> None:
    baseline_wb = _workbook(default_result)["PREBU26 Official"]
    adjusted_wb = _workbook(custom_fleet_result)["PREBU26 Official"]
    for row in ROW_SERIES:
        assert _row_values(baseline_wb, row) == _row_values(adjusted_wb, row), row


def test_actual_year_history_is_scenario_invariant(default_result) -> None:
    workbook = _workbook(default_result)
    base = workbook["Current Base"]
    high = workbook["High Population"]
    for row in (5, 10, 65):
        base_values = _row_values(base, row)
        high_values = _row_values(high, row)
        for fy in range(2001, 2025):
            assert base_values[fy] == high_values[fy], (row, fy)


def test_percentage_rows_are_year_on_year_changes_of_the_levels(default_result) -> None:
    worksheet = _workbook(default_result)["Current Base"]
    levels = _row_values(worksheet, 5)
    changes = _row_values(worksheet, 16)
    assert changes[2001] is None
    for fy in (2010, 2030, 2050):
        assert float(changes[fy]) == pytest.approx(
            float(levels[fy]) / float(levels[fy - 1]) - 1.0, rel=1e-12
        ), fy


def test_blank_line_items_stay_blank_not_zero(default_result) -> None:
    for sheet, labels in default_result.blank_rows.items():
        assert labels, sheet  # only genuinely blank rows are reported
    workbook = _workbook(default_result)
    # A cell with no governed value must be None, never 0.0-filled: check the
    # per-capita history the template itself leaves blank in FY2001.
    assert workbook["Current Base"].cell(row=12, column=2).value in (None,)


# ---------------------------------------------------------------------------
# page placement
# ---------------------------------------------------------------------------


def test_behavioural_path_is_declared_non_exportable_not_silently_skipped(
    extract_context,
) -> None:
    """A selected trace with no sheet must be DECLARED, never silent.

    The behavioural path is a single-series relabel of the comparison PED
    intensity trace, so it has no scenario sheet by stated policy: it lands
    in ``declared_non_exportable`` with a reason, ``skipped_traces`` stays
    empty (that bucket is always a defect), and the download still builds.
    """
    result = _extract(
        extract_context,
        ("Current finalist Base case", "Current finalist comparison behavioural path"),
    )
    assert result.sheet_names == ["Current Base"]
    assert result.exported_traces == ["Current finalist Base case"]
    assert result.skipped_traces == []
    assert result.declared_non_exportable == [
        "Current finalist comparison behavioural path"
    ]
    # The workbook bytes are a real XLSX: the download continues to work.
    workbook = _workbook(result)
    assert workbook.sheetnames == ["Current Base"]


def test_declared_registry_pins_the_runtime_trace_names() -> None:
    """The literals in NON_EXPORTABLE_TRACE_REASONS cannot drift.

    The extract module deliberately avoids importing the heavy runtime
    module, so the behavioural-path literal is pinned here against the
    runtime constant instead.
    """
    from model_dashboard.revenue_outlook import PED_COMPARISON_BEHAVIOURAL_TRACE_NAME
    from model_dashboard.revenue_outlook_excel_extract import (
        NON_EXPORTABLE_TRACE_REASONS,
    )

    assert PED_COMPARISON_BEHAVIOURAL_TRACE_NAME in NON_EXPORTABLE_TRACE_REASONS
    assert "Actual" in NON_EXPORTABLE_TRACE_REASONS
    for reason in NON_EXPORTABLE_TRACE_REASONS.values():
        assert reason.strip()


# ---------------------------------------------------------------------------
# extract coverage gate: every selectable trace exports or is declared
# ---------------------------------------------------------------------------


def _full_picker_inventory(pack) -> list[str]:
    """The Show-on-chart trace inventory, built the way the page builds it."""
    options = list(app._revenue_outlook_trace_options(pack.revenue_chart_rows))
    if (
        app._prebu_defer_workbook_signature() is not None
        and app.PREBU_DEFER_TRACE_NAME not in options
    ):
        options.append(app.PREBU_DEFER_TRACE_NAME)
    return options


def test_every_selectable_trace_is_exported_or_declared(extract_context) -> None:
    """The gate this file exists for: no silent third bucket, ever.

    Every trace the Show-on-chart picker can offer must either produce a
    worksheet or appear in the declared non-exportable registry with a
    stated reason. A newly added trace that does neither fails CI here
    instead of silently vanishing from the reader's download.
    """
    pack, _ = extract_context
    options = _full_picker_inventory(pack)
    assert len(options) >= 8, options
    result = _extract(extract_context, tuple(options))
    assert result.skipped_traces == [], (
        "selected traces silently missing from the extract: "
        f"{result.skipped_traces}"
    )
    accounted = set(result.exported_traces) | set(result.declared_non_exportable)
    assert accounted == set(options)
    assert len(result.sheet_names) == len(result.exported_traces)


def test_persistent_downside_sheet_matches_the_displayed_view(extract_context) -> None:
    """The exported downside totals are the on-screen downside totals."""
    pack, signature = extract_context
    traces = (
        "Actual",
        "Current finalist Base case",
        "Middle East conflict: High",
        "Persistent downside",
    )
    result = _extract(extract_context, traces)
    assert "Persistent Downside" in result.sheet_names
    workbook = _workbook(result)
    downside_totals = _row_values(workbook["Persistent Downside"], EXTRACT_LAST_ROW)
    base_totals = _row_values(workbook["Current Base"], EXTRACT_LAST_ROW)

    view = app.cached_revenue_outlook_view(
        signature,
        "Total NLTF revenue",
        "june_year",
        "Current planned path",
        traces,
        app.selected_sensitivity_key("Off", "Off", "Off"),
        PED_BRIDGE_DEFAULT_MODE,
        _default_uptake_key(),
        pack,
    )
    chart = view["chart_rows"]
    displayed = chart[
        chart["scenario_name"].astype(str).eq("persistent_downside")
        & chart["series_id"].astype(str).eq("total_nltf_net_revenue")
        & chart["time_grain"].astype(str).eq("june_year")
    ].set_index("june_year")["value"]
    for fy in (2035, 2042, 2050):
        assert downside_totals[fy] == pytest.approx(float(displayed.loc[fy]), abs=1e-9)
        # And the downside stays below the central sheet - the wedge survives
        # the round-trip into the workbook.
        assert downside_totals[fy] < base_totals[fy]


def test_high_population_sheet_matches_the_retethered_view(extract_context) -> None:
    """The exported comparison values are the re-tethered on-screen values.

    Before the extract read the final view layer, the High population sheet
    exported pre-retether values while the chart showed re-tethered ones.
    """
    pack, signature = extract_context
    traces = ("Actual", "Current finalist High population/comparison")
    result = _extract(extract_context, traces)
    workbook = _workbook(result)
    exported = _row_values(workbook["High Population"], EXTRACT_LAST_ROW)

    view = app.cached_revenue_outlook_view(
        signature,
        "Total NLTF revenue",
        "june_year",
        "Current planned path",
        traces,
        app.selected_sensitivity_key("Off", "Off", "Off"),
        PED_BRIDGE_DEFAULT_MODE,
        _default_uptake_key(),
        pack,
    )
    chart = view["chart_rows"]
    displayed = chart[
        chart["scenario_name"].astype(str).eq("current_comparison_1")
        & chart["series_id"].astype(str).eq("total_nltf_net_revenue")
        & chart["time_grain"].astype(str).eq("june_year")
    ].set_index("june_year")["value"]
    for fy in (2027, 2030, 2040):
        assert exported[fy] == pytest.approx(float(displayed.loc[fy]), abs=1e-9)


def test_prebu_defer_reference_workbook_exports_its_own_series(extract_context) -> None:
    """The display-only reference workbook gets a sheet from its own values."""
    workbook_signature = app._prebu_defer_workbook_signature()
    if workbook_signature is None:
        pytest.skip("references/PREBU defer.xlsx is not deployed in this checkout")
    result = _extract(
        extract_context,
        ("Actual", "Current finalist Base case", app.PREBU_DEFER_TRACE_NAME),
    )
    assert "PREBU26 Deferral Ref" in result.sheet_names
    workbook = _workbook(result)
    exported = _row_values(workbook["PREBU26 Deferral Ref"], EXTRACT_LAST_ROW)
    frames = app.cached_prebu_defer_workbook_frames(workbook_signature)
    source = frames["series"].get("total_nltf_net_revenue")
    assert source is not None and not source.empty
    for fy in (2030, 2040, 2050):
        if fy in source.index:
            assert exported[fy] == pytest.approx(float(source.loc[fy]), abs=1e-9)


def test_download_button_belongs_to_the_single_scenario_view() -> None:
    """The extract downloads from Single scenario; compare mode does not offer it."""
    from streamlit.testing.v1 import AppTest

    at = AppTest.from_file(str(ROOT / "app.py"), default_timeout=120)
    at.run()
    at.radio[0].set_value(app.REVENUE_OUTLOOK_PAGE)
    at.run()
    assert not at.exception
    labels = [str(element.proto.label) for element in at.get("download_button")]
    assert "Download forecast extract XLSX" in labels

    compare_radio = next(
        radio for radio in at.radio if app.REVENUE_OUTLOOK_VIEW_COMPARE in list(radio.options)
    )
    compare_radio.set_value(app.REVENUE_OUTLOOK_VIEW_COMPARE)
    at.run()
    assert not at.exception
    labels = [str(element.proto.label) for element in at.get("download_button")]
    assert "Download forecast extract XLSX" not in labels
