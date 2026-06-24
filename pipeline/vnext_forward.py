"""Fixed-finalist forward scorer for vNext models.

Scores future assumption rows (from a completed forecast input workbook)
with the saved production estimators. No model search is run. Numeric
forecasts are emitted only when the stream's parity audit status is
``passed`` and the runtime production-state gate passes; otherwise governed
gap rows (missing values) are emitted.
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

from . import GENERATION, PIPELINE_VERSION
from .vnext_core import (
    BASE_SERIES,
    LEVEL_SERIES,
    MAX_HORIZON,
    PARITY_TOLERANCE_ABS,
    STREAM_LABELS,
    engineer_features,
    load_stream_data,
    parse_period,
    period_str,
    target_lag_row,
)

VNEXT_SCORER_VERSION = "vnext-forward-scorer-v1"


class FittedStateHashMismatch(RuntimeError):
    """Raised when a fitted production state differs from its manifest hash."""

SHEET_BY_STREAM = {
    "PED": "PED Inputs",
    "LIGHT_RUC": "Light RUC Inputs",
    "HEAVY_RUC": "Heavy RUC Inputs",
}

# Template user-entry columns required per stream (positive numeric).
REQUIRED_USER_COLUMNS = {
    "HEAVY_RUC": [
        "real_gdp_sa_nzd",
        "real_diesel_price_cents_per_litre",
        "unemployment_rate",
        "real_light_ruc_price_nzd_per_1000km",
        "real_heavy_ruc_price_nzd_per_1000km",
    ],
    "PED": [
        "real_gdp_per_capita_nzd",
        "population",
        "unemployment_rate",
        "real_petrol_price_cents_per_litre",
    ],
}

# canonical history column -> derived log column used by engineer_features
LOG_DERIVATIONS = {
    "HEAVY_RUC": {
        "log_real_gdp": "real_gdp_sa_nzd",
        "log_real_diesel_price": "real_diesel_price_cents_per_litre",
        "log_real_heavy_ruc_price": "real_heavy_ruc_price_nzd_per_1000km",
        "log_real_light_ruc_price": "real_light_ruc_price_nzd_per_1000km",
        "log_unemployment_rate": "unemployment_rate",
    },
    "PED": {
        "log_real_gdp_per_capita": "real_gdp_per_capita_nzd",
        "log_real_petrol_price": "real_petrol_price_cents_per_litre",
        "log_unemployment_rate": "unemployment_rate",
    },
}


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def state_dir(stream: str) -> Path:
    return repo_root() / "data" / "dashboard_evidence_pack_reproducibility" / f"{stream.lower()}_{GENERATION}"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def fitted_state_hash_errors(sdir: Path, manifest: Dict[str, Any]) -> List[str]:
    errors: List[str] = []
    for label, entry in manifest.get("production_states", {}).items():
        state_file = str(entry.get("file", "")).strip()
        expected = str(entry.get("sha256", "")).strip().lower()
        if not state_file:
            errors.append(f"{label}: manifest production state has no file")
            continue
        if not expected:
            errors.append(f"{label}: manifest production state has no sha256")
            continue
        path = sdir / state_file
        if not path.exists():
            errors.append(f"{label}: fitted state is missing at {state_file}")
            continue
        actual = sha256_file(path)
        if actual.lower() != expected:
            errors.append(f"{label}: fitted state SHA256 mismatch for {state_file}; expected {expected}, got {actual}")
    return errors


@dataclass
class VNextScorer:
    stream: str
    manifest: Dict[str, Any]
    parity: Dict[str, Any]
    bundles: Dict[str, Any]
    runtime_state_gate_delta: float

    @property
    def finalist(self) -> str:
        return self.manifest["finalist_model"]

    @property
    def numeric_enabled(self) -> bool:
        return (self.parity.get("parity_status") == "passed"
                and self.runtime_state_gate_delta <= PARITY_TOLERANCE_ABS)


def load_scorer(stream: str) -> Optional[VNextScorer]:
    import joblib

    sdir = state_dir(stream)
    mpath = sdir / "fitted_model_manifest.json"
    ppath = sdir / "forward_scorer_parity_audit.json"
    if not mpath.exists() or not ppath.exists():
        return None
    manifest = json.loads(mpath.read_text(encoding="utf-8"))
    parity = json.loads(ppath.read_text(encoding="utf-8"))
    hash_errors = fitted_state_hash_errors(sdir, manifest)
    if hash_errors:
        raise FittedStateHashMismatch("; ".join(hash_errors))
    bundles = {}
    for label, entry in manifest["production_states"].items():
        path = sdir / entry["file"]
        if not path.exists():
            return None
        bundles[label] = joblib.load(path)
    gate = _runtime_production_state_gate(stream, sdir, manifest, bundles)
    return VNextScorer(stream, manifest, parity, bundles, gate)


def _runtime_production_state_gate(stream: str, sdir: Path, manifest: Dict[str, Any],
                                   bundles: Dict[str, Any]) -> float:
    """Replay archived production training-fit predictions from the saved
    state and matrices. Run before every forecast; guards against silent
    environment or artifact drift."""
    tf_path = sdir / "training_fit_predictions.parquet"
    mat_path = sdir / "training_feature_matrices.parquet"
    if not tf_path.exists() or not mat_path.exists():
        return float("inf")
    tf = pd.read_parquet(tf_path)
    mats = pd.read_parquet(mat_path)
    max_delta = 0.0
    members = {label: entry["component_model"] if isinstance(entry, dict) and "component_model" in entry else None
               for label, entry in manifest["production_states"].items()}
    member_models = {m["component_label"]: m["component_model"] for m in manifest["members"]}
    for label, bundle in bundles.items():
        g = mats[(mats["component_label"] == label) & (mats["origin"] == "production")]
        component_model = member_models.get(label)
        archived = tf[tf["component_model"] == component_model].drop_duplicates("training_period")
        if g.empty or archived.empty:
            return float("inf")
        X = g[bundle["feature_cols"]].astype(float).fillna(0.0)
        model = bundle["model"]
        if isinstance(model, dict) and model.get("kind") == "residual":
            Xb = g[bundle["base_cols"]].astype(float).fillna(0.0)
            fit_log = model["base"].predict(Xb.to_numpy(float)) + model["resid"].predict(X.to_numpy(float))
        else:
            fit_log = model.predict(X.to_numpy(float))
        fit_level = np.exp(np.asarray(fit_log, dtype=float))
        archived_map = dict(zip(archived["training_period"], archived["training_fit_pred"]))
        for tp, fl in zip(g["training_period"], fit_level):
            if tp in archived_map:
                max_delta = max(max_delta, abs(float(fl) - float(archived_map[tp])))
    return max_delta


# ---------------------------------------------------------------------------
# Workbook parsing (pipeline CLI path; the dashboard runner has its own
# validated parser and passes an assumptions frame directly).
# ---------------------------------------------------------------------------

def parse_workbook_assumptions(workbook_path: Path, stream: str,
                               latest_actual: pd.Period) -> pd.DataFrame:
    df = pd.read_excel(workbook_path, sheet_name=SHEET_BY_STREAM[stream])
    df.columns = [str(c).strip() for c in df.columns]
    required = REQUIRED_USER_COLUMNS[stream]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{SHEET_BY_STREAM[stream]}: missing required columns {missing}")
    valid_rows = []
    for _, row in df.iterrows():
        vals = pd.to_numeric(row[required], errors="coerce")
        if vals.notna().all() and (vals > 0).all():
            valid_rows.append(row)
        elif vals.notna().any():
            break
        else:
            break
    if not valid_rows:
        raise ValueError(f"{SHEET_BY_STREAM[stream]}: no valid forecast rows found")
    out = pd.DataFrame(valid_rows).reset_index(drop=True)
    periods = [parse_period(p) for p in out["period"]]
    expected = [latest_actual + i + 1 for i in range(len(periods))]
    if periods != expected:
        raise ValueError(
            f"{SHEET_BY_STREAM[stream]}: forecast rows must be continuous from "
            f"{period_str(expected[0])}; found {[period_str(p) for p in periods[:4]]}...")
    out["__period__"] = periods
    return out.set_index("__period__")


def build_future_canonical_frame(assumptions: pd.DataFrame, stream: str) -> pd.DataFrame:
    """Map validated workbook assumption rows to the canonical history schema
    columns needed by the engineered feature registry."""
    cols_needed = set(LEVEL_SERIES[stream].values()) | set(BASE_SERIES[stream].values())
    out = pd.DataFrame(index=assumptions.index)
    for col in cols_needed:
        if col in assumptions.columns:
            out[col] = pd.to_numeric(assumptions[col], errors="coerce").astype(float)
    for log_col, src in LOG_DERIVATIONS[stream].items():
        if log_col in cols_needed:
            s = pd.to_numeric(assumptions[src], errors="coerce").astype(float)
            out[log_col] = np.where(s > 0, np.log(s.where(s > 0)), np.nan)
    missing = [c for c in cols_needed if c not in out.columns]
    if missing:
        raise ValueError(f"Cannot derive canonical columns for {stream}: {missing}")
    return out


def forward_forecast(stream: str, assumptions: pd.DataFrame,
                     scorer: Optional[VNextScorer] = None) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    """Score future assumption rows with the fixed vNext finalist.

    Returns (future_rows, component_rows, capability_record). When the
    parity/state gates fail, forecast values are missing (governed gap).
    """
    scorer = scorer or load_scorer(stream)
    sd = load_stream_data(repo_root(), stream)
    periods = list(assumptions.index)
    capability = _capability_record(stream, scorer)

    if scorer is None or not scorer.numeric_enabled:
        gap_future, gap_comp = _gap_rows(stream, periods, capability)
        return gap_future, gap_comp, capability

    future_canonical = build_future_canonical_frame(assumptions, stream)
    needed = sorted(set(LEVEL_SERIES[stream].values()) | set(BASE_SERIES[stream].values()))
    hist_canonical = sd.history[needed].copy()
    for c in needed:
        hist_canonical[c] = pd.to_numeric(hist_canonical[c], errors="coerce").astype(float)
    combined = pd.concat([hist_canonical, future_canonical[needed]], axis=0)
    combined = combined[~combined.index.duplicated(keep="last")].sort_index()
    exog_full = engineer_features(combined, stream)

    members = scorer.manifest["members"]
    weights = {m["component_label"]: float(m["component_weight"]) for m in members}
    base_y_hist = {p: float(sd.y_log.loc[p]) for p in sd.y_log.index
                   if pd.notna(sd.y_log.loc[p]) and p <= sd.latest_actual}

    comp_levels: Dict[str, List[float]] = {}
    comp_logs: Dict[str, List[float]] = {}
    for m in members:
        label = m["component_label"]
        bundle = scorer.bundles[label]
        include_ylags = bool(bundle["spec"]["include_target_lags"])
        y_hist = dict(base_y_hist)
        levels, logs = [], []
        for p in periods:
            row: Dict[str, float] = exog_full.loc[p].to_dict() if p in exog_full.index else {}
            if include_ylags:
                row.update(target_lag_row(p, y_hist))
            x, feature_gap = _feature_contract_frame(row, bundle["feature_cols"], bundle, stream, label, p)
            if feature_gap is not None:
                gap_capability = _feature_contract_gap_capability(capability, [feature_gap])
                gap_future, _ = _gap_rows(stream, periods, gap_capability)
                return (
                    gap_future,
                    _feature_contract_gap_component_rows(stream, periods, gap_capability, [feature_gap]),
                    gap_capability,
                )
            model = bundle["model"]
            if isinstance(model, dict) and model.get("kind") == "residual":
                xb, base_gap = _feature_contract_frame(row, bundle["base_cols"], bundle, stream, label, p)
                if base_gap is not None:
                    gap_capability = _feature_contract_gap_capability(capability, [base_gap])
                    gap_future, _ = _gap_rows(stream, periods, gap_capability)
                    return (
                        gap_future,
                        _feature_contract_gap_component_rows(stream, periods, gap_capability, [base_gap]),
                        gap_capability,
                    )
                pred_log = float(model["base"].predict(xb.to_numpy(float))[0]
                                 + model["resid"].predict(x.to_numpy(float))[0])
            else:
                pred_log = float(model.predict(x.to_numpy(float))[0])
            if include_ylags and np.isfinite(pred_log):
                y_hist[p] = pred_log
            logs.append(pred_log)
            levels.append(float(np.exp(pred_log)))
        comp_levels[label] = levels
        comp_logs[label] = logs

    future_rows, component_rows = [], []
    for i, p in enumerate(periods):
        final_level = float(sum(weights[lbl] * comp_levels[lbl][i] for lbl in comp_levels))
        common = {
            "stream": stream, "stream_label": STREAM_LABELS[stream],
            "model": scorer.finalist, "target_period": period_str(p),
            "horizon": i + 1,
            "availability_status": "numeric_forecast_available",
            "gap_code": None, "gap_reason": "",
            "fixed_finalist_only": True, "broad_search_run": False,
            "score_basis": "forward_assumption_workbook",
            **_scorer_metadata(capability),
        }
        future_rows.append({**common, "forecast": final_level, "prediction": final_level,
                            "forecast_available": True})
        for m in members:
            lbl = m["component_label"]
            component_rows.append({**common,
                                   "component_model": m["component_model"],
                                   "component_label": lbl,
                                   "component_role": "weighted level component",
                                   "component_weight": weights[lbl],
                                   "component_forecast": comp_levels[lbl][i],
                                   "component_log_value": comp_logs[lbl][i],
                                   "weighted_component_forecast": weights[lbl] * comp_levels[lbl][i],
                                   "forecast_available": True})
        component_rows.append({**common, "component_model": scorer.finalist,
                               "component_label": "FINAL",
                               "component_role": "final weighted level prediction",
                               "component_weight": 1.0,
                               "component_forecast": final_level,
                               "component_log_value": float(np.log(final_level)) if final_level > 0 else np.nan,
                               "weighted_component_forecast": final_level,
                               "forecast_available": True})
    return pd.DataFrame(future_rows), pd.DataFrame(component_rows), capability


def _zero_imputed_feature_set(bundle: Dict[str, Any]) -> set[str]:
    contract = bundle.get("feature_contract", {}) if isinstance(bundle, dict) else {}
    values = []
    for key in ("zero_imputed_features", "zero_fill_features"):
        raw = bundle.get(key) if isinstance(bundle, dict) else None
        if raw is None and isinstance(contract, dict):
            raw = contract.get(key)
        if isinstance(raw, str):
            values.extend([item.strip() for item in raw.split(",")])
        elif isinstance(raw, (list, tuple, set)):
            values.extend([str(item).strip() for item in raw])
    return {value for value in values if value}


def _feature_contract_frame(
    row: Dict[str, Any],
    columns: Sequence[str],
    bundle: Dict[str, Any],
    stream: str,
    member: str,
    period: pd.Period,
) -> Tuple[pd.DataFrame, Dict[str, Any] | None]:
    frame = pd.DataFrame([{column: row.get(column, np.nan) for column in columns}]).astype(float)
    allowed_zero = _zero_imputed_feature_set(bundle)
    missing = [
        column
        for column in columns
        if column in frame.columns and not np.isfinite(pd.to_numeric(frame.at[0, column], errors="coerce"))
    ]
    zero_fill = [column for column in missing if column in allowed_zero]
    if zero_fill:
        frame.loc[0, zero_fill] = 0.0
    review_required = [column for column in missing if column not in allowed_zero]
    if review_required:
        return frame, {
            "stream": stream,
            "member": member,
            "period": period_str(period),
            "missing_or_nonfinite_features": review_required,
        }
    return frame, None


def _feature_contract_gap_capability(capability: Dict[str, Any], gaps: list[Dict[str, Any]]) -> Dict[str, Any]:
    out = dict(capability)
    stream = str(out.get("stream", "stream")).lower()
    out["capability_status"] = "feature_contract_review_required"
    out["gap_code"] = f"{stream}_vnext_feature_contract_gap"
    out["gap_reason"] = "Missing/nonfinite future features require governed review before scoring: " + json.dumps(gaps, sort_keys=True)
    out["forecast_capability_available"] = False
    out["feature_contract_gaps"] = json.dumps(gaps, sort_keys=True)
    return out


def _feature_contract_gap_component_rows(
    stream: str,
    periods: Sequence[pd.Period],
    capability: Dict[str, Any],
    gaps: list[Dict[str, Any]],
) -> pd.DataFrame:
    _, component_rows = _gap_rows(stream, periods, capability)
    if component_rows.empty:
        return component_rows
    component_rows["component_role"] = "feature contract governed gap"
    component_rows["feature_contract_gaps"] = json.dumps(gaps, sort_keys=True)
    return component_rows


def _scorer_metadata(capability: Dict[str, Any]) -> Dict[str, Any]:
    return {k: capability.get(k) for k in
            ("scorer_version", "source_artifact_hashes", "parity_status",
             "max_parity_delta", "stored_replay_max_delta", "failing_component",
             "capability_status")}


def _capability_record(stream: str, scorer: Optional[VNextScorer]) -> Dict[str, Any]:
    if scorer is None:
        return {"stream": stream, "stream_label": STREAM_LABELS[stream],
                "model": None, "capability_status": "insufficient_artifacts",
                "gap_code": f"{stream.lower()}_vnext_state_missing",
                "gap_reason": "vNext fitted state or manifests are not present; run pipeline finalize.",
                "scorer_version": VNEXT_SCORER_VERSION, "parity_status": "not_run",
                "max_parity_delta": None, "stored_replay_max_delta": None,
                "failing_component": None, "source_artifact_hashes": None,
                "forecast_capability_available": False}
    hashes = {label: entry["sha256"] for label, entry in scorer.manifest["production_states"].items()}
    status = "numeric_forecast_available" if scorer.numeric_enabled else "parity_failed"
    return {"stream": stream, "stream_label": STREAM_LABELS[stream],
            "model": scorer.finalist, "capability_status": status,
            "gap_code": None if scorer.numeric_enabled else f"{stream.lower()}_vnext_parity_failed",
            "gap_reason": "" if scorer.numeric_enabled else
            "vNext parity or runtime state gate failed; numeric forecasts withheld.",
            "scorer_version": VNEXT_SCORER_VERSION,
            "parity_status": scorer.parity.get("parity_status"),
            "max_parity_delta": max(scorer.parity.get("state_replay_max_abs_delta", np.nan),
                                    scorer.parity.get("recipe_replay_max_abs_delta", np.nan)),
            "stored_replay_max_delta": scorer.runtime_state_gate_delta,
            "failing_component": None,
            "source_artifact_hashes": json.dumps(hashes, sort_keys=True),
            "forecast_capability_available": scorer.numeric_enabled}


def _gap_rows(stream: str, periods: Sequence[pd.Period],
              capability: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    future_rows, component_rows = [], []
    for i, p in enumerate(periods):
        common = {"stream": stream, "stream_label": STREAM_LABELS[stream],
                  "model": capability.get("model"),
                  "target_period": period_str(p), "horizon": i + 1,
                  "availability_status": capability["capability_status"],
                  "gap_code": capability.get("gap_code"),
                  "gap_reason": capability.get("gap_reason"),
                  "fixed_finalist_only": True, "broad_search_run": False,
                  "score_basis": "forward_assumption_workbook",
                  **_scorer_metadata(capability)}
        future_rows.append({**common, "forecast": pd.NA, "prediction": pd.NA,
                            "forecast_available": False})
        component_rows.append({**common, "component_model": capability.get("model"),
                               "component_label": "FINAL",
                               "component_role": "governed gap",
                               "component_weight": pd.NA, "component_forecast": pd.NA,
                               "component_log_value": pd.NA,
                               "weighted_component_forecast": pd.NA,
                               "forecast_available": False})
    return pd.DataFrame(future_rows), pd.DataFrame(component_rows)


# ---------------------------------------------------------------------------
# Flat-forward baseline workbook (governed demo / sensitivity baseline)
# ---------------------------------------------------------------------------

EXTRA_USER_COLUMNS = {
    "HEAVY_RUC": ["lagged_real_light_ruc_price_nzd_per_1000km",
                  "lead_real_heavy_ruc_price_nzd_per_1000km", "target_lag_1", "target_lag_4"],
    "PED": ["target_lag_1", "target_lag_4"],
}
LIGHT_HISTORY_USER_COLUMNS = [
    "real_gdp_sa_nzd", "real_diesel_price_cents_per_litre",
    "real_light_ruc_price_nzd_per_1000km", "lagged_real_light_ruc_price_nzd_per_1000km",
    "target_lag_1", "target_lag_4",
]


def create_flat_forward_workbook(output_path: Path | str, quarters: int = 12) -> Path:
    """Build a completed assumption workbook whose user entries hold every
    driver flat at its last positive actual value from the canonical history.

    This is the governed demo/sensitivity baseline: unit-consistent with the
    canonical input basis by construction. It is NOT an economic forecast.
    """
    from model_dashboard.forecast_runner import build_forecast_input_template
    from openpyxl import load_workbook

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    build_forecast_input_template(output_path, quarters=quarters)

    last_values: Dict[str, Dict[str, float]] = {}
    for stream, hist_file in (("PED", "ped_inputs.parquet"),
                              ("LIGHT_RUC", "light_ruc_inputs.parquet"),
                              ("HEAVY_RUC", "heavy_ruc_inputs.parquet")):
        hist = pd.read_parquet(repo_root() / "data" / "model_input_history" / hist_file)
        vals: Dict[str, float] = {}
        cols = (REQUIRED_USER_COLUMNS.get(stream, []) + EXTRA_USER_COLUMNS.get(stream, [])
                if stream != "LIGHT_RUC" else list(LIGHT_HISTORY_USER_COLUMNS))
        target = pd.to_numeric(hist["target"], errors="coerce")
        last_target = float(target[target > 0].iloc[-1])
        for col in cols:
            if col in ("target_lag_1", "target_lag_4"):
                vals[col] = last_target
                continue
            if col not in hist.columns:
                continue
            s = pd.to_numeric(hist[col], errors="coerce")
            s = s[s > 0]
            if not s.empty:
                vals[col] = float(s.iloc[-1])
        last_values[stream] = vals

    wb = load_workbook(output_path)
    for stream, sheet in SHEET_BY_STREAM.items():
        ws = wb[sheet]
        headers = {str(c.value).strip(): i + 1 for i, c in enumerate(ws[1]) if c.value is not None}
        n_rows = min(quarters, ws.max_row - 1)
        for col_name, value in last_values[stream].items():
            if col_name not in headers:
                continue
            col = headers[col_name]
            for row in range(2, 2 + n_rows):
                ws.cell(row=row, column=col).value = float(value)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Pipeline CLI stage
# ---------------------------------------------------------------------------

def stage_forecast(streams: Sequence[str], workbooks: Optional[Sequence[str]]) -> None:
    from .vnext_run import sha256_file, vnext_dir

    if not workbooks:
        default = repo_root() / "artifacts" / "vnext" / "flat_forward_baseline_12q.xlsx"
        if not default.exists():
            create_flat_forward_workbook(default, quarters=12)
            print(f"[forecast] built flat-forward baseline workbook: {default}")
        workbooks = [str(default)]

    for wb in workbooks:
        wb_path = Path(wb)
        scenario = wb_path.stem.replace("NLTF_forecast_input_template_", "") or "scenario"
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        run_dir = repo_root() / "artifacts" / "vnext" / "forecast_runs" / f"{ts}_{scenario}"
        run_dir.mkdir(parents=True, exist_ok=True)
        all_future, all_comp, capabilities = [], [], []
        for stream in streams:
            scorer = load_scorer(stream)
            sd = load_stream_data(repo_root(), stream)
            assumptions = parse_workbook_assumptions(wb_path, stream, sd.latest_actual)
            future, comp, capability = forward_forecast(stream, assumptions, scorer)
            future["scenario_name"] = scenario
            comp["scenario_name"] = scenario
            all_future.append(future)
            all_comp.append(comp)
            capabilities.append(capability)
            n_num = int(future["forecast_available"].sum()) if "forecast_available" in future else 0
            print(f"[forecast] {stream} / {scenario}: {n_num}/{len(future)} numeric rows "
                  f"({capability['capability_status']})")
        fut = pd.concat(all_future, ignore_index=True)
        cmp_ = pd.concat(all_comp, ignore_index=True)
        cap = pd.DataFrame(capabilities)
        fut.to_parquet(run_dir / "future_forecasts.parquet", index=False)
        cmp_.to_parquet(run_dir / "component_forecasts.parquet", index=False)
        cap.to_parquet(run_dir / "forecast_capability_report.parquet", index=False)
        manifest = {
            "created_at": utc_iso(),
            "pipeline_version": PIPELINE_VERSION,
            "scorer_version": VNEXT_SCORER_VERSION,
            "scenario_name": scenario,
            "workbook_filename": wb_path.name,
            "workbook_sha256": sha256_file(wb_path),
            "fixed_finalists_only": True,
            "broad_search_run": False,
            "evidence_pack_modified": False,
            "streams": {c["stream"]: {"model": c["model"], "capability_status": c["capability_status"],
                                      "parity_status": c["parity_status"]}
                        for c in capabilities},
        }
        (run_dir / "forecast_run_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        print(f"[forecast] wrote {run_dir}")


def utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()
